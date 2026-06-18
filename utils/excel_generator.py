import os
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.chart import BarChart, Reference

logger = logging.getLogger("excel_generator")

class ExcelSheet:
    def __init__(self, worksheet, title, subtitle=None):
        self.ws = worksheet
        self.title = title
        self.subtitle = subtitle
        self.columns = []
        self.headers_set = False
        self.start_row = 5  # Reserve rows 1-4 for logo and titles
        self._write_title_block()
        
    def _write_title_block(self):
        # Maroon header block
        self.ws.merge_cells('C1:H2')
        title_cell = self.ws['C1']
        title_cell.value = self.title
        title_cell.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Load University logo in cell A1 if exists
        logo_path = None
        possible_paths = [
            os.path.join(os.getcwd(), 'static', 'images', 'dypatil_logo.png'),
            os.path.join(os.getcwd(), 'static', 'img', 'logo.png')
        ]
        for path in possible_paths:
            if os.path.exists(path):
                logo_path = path
                break
                
        if logo_path:
            try:
                img = OpenpyxlImage(logo_path)
                img.width = 110
                img.height = 40
                self.ws.add_image(img, 'A1')
            except Exception as e:
                logger.warning(f"Failed to embed logo in Excel: {e}")
                
        # Set subtitle if present
        if self.subtitle:
            self.ws.merge_cells('A3:H3')
            sub_cell = self.ws['A3']
            sub_cell.value = self.subtitle
            sub_cell.font = Font(name='Arial', size=10, italic=True, color='555555')
            sub_cell.alignment = Alignment(horizontal='left', vertical='center')
            
    def set_headers(self, columns: list[dict]):
        """
        Sets sheet headers. Columns list structure:
        [ { 'key': 'roll', 'label': 'Roll No', 'width': 10, 'format': 'text', 'align': 'center' }, ... ]
        """
        self.columns = columns
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='medium', color='800000')
        )
        
        for col_idx, col in enumerate(columns, start=1):
            cell = self.ws.cell(row=self.start_row, column=col_idx)
            cell.value = col.get('label', '')
            cell.font = Font(name='Arial', size=10, bold=True, color='333333')
            cell.fill = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Explicit width override if defined
            if 'width' in col:
                col_letter = get_column_letter(col_idx)
                self.ws.column_dimensions[col_letter].width = col['width']
                
        self.ws.row_dimensions[self.start_row].height = 28
        self.headers_set = True
        
    def add_rows(self, data: list[dict]):
        if not self.headers_set:
            raise ValueError("Headers must be set before adding data rows.")
            
        current_row = self.ws.max_row + 1
        
        alt_fill = PatternFill(start_color='FDFDFD', end_color='FDFDFD', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='EFEFEF'),
            right=Side(style='thin', color='EFEFEF'),
            top=Side(style='thin', color='EFEFEF'),
            bottom=Side(style='thin', color='EFEFEF')
        )
        
        for row_offset, row_data in enumerate(data):
            row_num = current_row + row_offset
            is_alt = (row_offset % 2 == 1)
            row_fill = alt_fill if is_alt else white_fill
            
            for col_idx, col in enumerate(self.columns, start=1):
                cell = self.ws.cell(row=row_num, column=col_idx)
                val = row_data.get(col['key'], None)
                
                # Default empty values to empty string
                if val is None:
                    val = ""
                    
                # Format cell according to configuration
                fmt = col.get('format', 'text')
                if fmt == 'percentage':
                    try:
                        val_float = float(val)
                        # If value is passed as 75.0 instead of 0.75, normalize it
                        if val_float > 1.0:
                            cell.value = val_float / 100.0
                        else:
                            cell.value = val_float
                        cell.number_format = '0.00%'
                    except (ValueError, TypeError):
                        cell.value = val
                elif fmt == 'integer':
                    try:
                        cell.value = int(val)
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        cell.value = val
                elif fmt == 'float':
                    try:
                        cell.value = float(val)
                        cell.number_format = '#,##0.00'
                    except (ValueError, TypeError):
                        cell.value = val
                else:
                    cell.value = str(val)
                    
                cell.font = Font(name='Arial', size=10, color='333333')
                cell.fill = row_fill
                cell.border = thin_border
                
                align = col.get('align', 'left')
                cell.alignment = Alignment(horizontal=align, vertical='center')
                
            self.ws.row_dimensions[row_num].height = 20
            
        self._auto_fit_columns()
        
    def add_summary_row(self, totals: dict):
        row_num = self.ws.max_row + 1
        summary_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        summary_font = Font(name='Arial', size=10, bold=True, color='1F4E79')
        
        thick_border = Border(
            top=Side(style='thin', color='888888'),
            bottom=Side(style='double', color='000000')
        )
        
        for col_idx, col in enumerate(self.columns, start=1):
            cell = self.ws.cell(row=row_num, column=col_idx)
            val = totals.get(col['key'], "")
            cell.value = val
            cell.font = summary_font
            cell.fill = summary_fill
            cell.border = thick_border
            
            fmt = col.get('format')
            if fmt == 'percentage':
                cell.number_format = '0.00%'
            elif fmt == 'integer':
                cell.number_format = '#,##0'
            elif fmt == 'float':
                cell.number_format = '#,##0.00'
                
            align = col.get('align', 'center')
            cell.alignment = Alignment(horizontal=align, vertical='center')
            
        self.ws.row_dimensions[row_num].height = 22
        
    def apply_conditional_format(self, col_key, rules: list[dict]):
        """
        Applies a list of rules: [ { 'min': 0, 'max': 0.75, 'fill_color': 'FFC7CE' } ]
        """
        col_idx = None
        for idx, col in enumerate(self.columns, start=1):
            if col['key'] == col_key:
                col_idx = idx
                break
                
        if col_idx is None:
            return
            
        for r in range(self.start_row + 1, self.ws.max_row + 1):
            cell = self.ws.cell(row=r, column=col_idx)
            val = cell.value
            if val is not None:
                try:
                    val_float = float(val)
                    for rule in rules:
                        min_val = rule.get('min')
                        max_val = rule.get('max')
                        fill_color = rule.get('fill_color')
                        
                        match = True
                        if min_val is not None and val_float < min_val:
                            match = False
                        if max_val is not None and val_float > max_val:
                            match = False
                            
                        if match and fill_color:
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                except (ValueError, TypeError):
                    pass
                    
    def freeze(self, row=5, col=2):
        col_letter = get_column_letter(col + 1)
        self.ws.freeze_panes = f"{col_letter}{row + 1}"
        
    def add_chart(self, chart_type, data_range, title):
        """
        data_range should be: { 'data_col': 5, 'cats_col': 2, 'min_row': 5, 'max_row': 20 }
        """
        if chart_type == 'bar':
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = title
            chart.y_axis.title = "Value"
            chart.x_axis.title = "Item"
            chart.width = 18
            chart.height = 10
            
            data = Reference(self.ws, min_col=data_range['data_col'], min_row=data_range['min_row'], max_row=data_range['max_row'])
            cats = Reference(self.ws, min_col=data_range['cats_col'], min_row=data_range['min_row'] + 1, max_row=data_range['max_row'])
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Place chart to the right of the table
            self.ws.add_chart(chart, f"J{self.start_row}")
            
    def _auto_fit_columns(self):
        for col in self.ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Auto width check skipping the merged title blocks
            for cell in col:
                if cell.row >= self.start_row and cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            # Enforce minimum size of 10
            self.ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

class ExcelReport:
    def __init__(self, title, subtitle=None):
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        self.title = title
        self.subtitle = subtitle
        self.sheets = {}
        
    def add_sheet(self, name) -> ExcelSheet:
        ws = self.wb.create_sheet(title=name)
        sheet = ExcelSheet(ws, self.title, self.subtitle)
        self.sheets[name] = sheet
        return sheet
        
    def save(self, output_path) -> int:
        self.wb.save(output_path)
        if os.path.exists(output_path):
            return os.path.getsize(output_path)
        return 0
