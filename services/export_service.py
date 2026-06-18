import io
from openpyxl import Workbook
from openpyxl.styles import Font

class ExportService:
    @staticmethod
    def to_excel(data, headers, title="Export"):
        wb = Workbook()
        ws = wb.active
        ws.title = title
        
        # Headers
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            
        # Data
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num).value = cell_value
                
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
