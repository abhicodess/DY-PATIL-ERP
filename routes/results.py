"""
routes/results.py — Complete Results Dashboard Module
DY Patil University ERP
"""

from flask import (Blueprint, render_template, request, redirect,
                   session, send_file, jsonify, flash)
from functools import wraps
import psycopg2
import psycopg2.extras, os, io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border,
                              Side, GradientFill)
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Config ──────────────────────────────────────────────────
from config import SEMESTERS, DEPARTMENTS, DIVISIONS, YEARS
SECTIONS = DIVISIONS

# Grade thresholds (DY Patil standard)
GRADE_TABLE = [
    (75, "O",  "Outstanding",  "#059669"),
    (70, "A+", "Excellent",    "#2563eb"),
    (60, "A",  "Very Good",    "#7c3aed"),
    (55, "B+", "Good",         "#d97706"),
    (50, "B",  "Above Average","#ea580c"),
    (45, "C",  "Average",      "#dc2626"),
    (0,  "F",  "Fail",         "#b91c1c"),
]

results_bp = Blueprint("results_bp", __name__)

# ── DB helpers (standalone, same DB as main app) ────────────
from utils.pg_wrapper import get_db, qry as _qry, qone as _qone, exe as _exe


from blueprints.auth.decorators import login_required
admin_required = login_required("admin")

def is_hod_or_admin():
    role = session.get("role")
    if role == "admin":
        return True
    if role == "faculty":
        if session.get("designation") == "HOD":
            return True
        from utils.pg_wrapper import qone
        fac_id = session.get("faculty_id")
        if fac_id:
            fac = qone("SELECT designation FROM faculty WHERE id = %s", (fac_id,))
            if fac and fac.get("designation") == "HOD":
                return True
    return False

def hod_or_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("role"):
            flash("Please log in to access this page.", "warning")
            return redirect(url_for('auth.login', next=request.url))
        if not is_hod_or_admin():
            from flask import abort
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ═══════════════════════════════════════════════════════════
#  CORE CALCULATION ENGINE
# ═══════════════════════════════════════════════════════════

def calc_grade(percentage):
    """Return (grade_letter, label, hex_color) for a percentage."""
    for threshold, letter, label, color in GRADE_TABLE:
        if percentage >= threshold:
            return letter, label, color
    return "F", "Fail", "#b91c1c"

def calc_result_status(subject_rows):
    """
    PASS  → all subjects ≥ 40% of their total
    ATKT  → 1–2 subjects below 40%
    FAIL  → 3+ subjects below 40%
    Returns (status, below_count, list_of_failed_subjects)
    """
    below = []
    for r in subject_rows:
        marks = r["marks"] if r["marks"] is not None else 0
        total = r["total"] if r["total"] and r["total"] > 0 else 100
        if (marks / total * 100) < 40:
            below.append(r["subject"])
    n = len(below)
    if n == 0:   status = "PASS"
    elif n <= 2: status = "ATKT"
    else:        status = "FAIL"
    return status, n, below

def assign_section(rank, total_students):
    """Assign section A/B/C/D based on rank (1-indexed)."""
    if total_students == 0:
        return "A"
    q = max(1, total_students // 4)
    if   rank <= q:     return "A"
    elif rank <= q * 2: return "B"
    elif rank <= q * 3: return "C"
    else:               return "D"

def build_student_records(dept=None, semester=None, year=None,
                           section_filter=None, status_filter=None,
                           search=None, published_only=False):
    """
    Core engine: fetches all results rows, groups per student,
    computes section / cumulative / percentage / grade / status.
    Returns list of student dicts sorted by roll.
    """
    sql = "SELECT * FROM results WHERE 1=1"
    params = []
    if dept:      sql += " AND department=%s";  params.append(dept)
    if semester:  sql += " AND semester=%s";     params.append(semester)
    if year:      sql += " AND year=%s";         params.append(year)
    if published_only:
        sql += " AND published=1"
    sql += " ORDER BY roll, student_name, subject"
    rows = _qry(sql, params)

    # Group by (roll, student_name)
    student_map = {}  # key = (roll, name) → {subjects: [], meta: {}}
    for r in rows:
        key = (str(r["roll"] or ""), str(r["student_name"] or ""))
        if key not in student_map:
            student_map[key] = {
                "roll":       r["roll"] or "",
                "name":       r["student_name"] or "",
                "dept":       r["department"] or "",
                "year":       r["year"] or "",
                "semester":   r["semester"] or "",
                "subjects":   [],
            }
        student_map[key]["subjects"].append(dict(r))

    # Build ordered list (sorted by roll for section assignment)
    students = sorted(student_map.values(), key=lambda x: x["roll"])
    total_count = len(students)

    result_list = []
    for rank, s in enumerate(students, 1):
        subj_rows = s["subjects"]

        # Cumulative
        cum_obtained = sum(
            (r["marks"] if r["marks"] is not None else 0)
            for r in subj_rows
        )
        cum_max = sum(
            (r["total"] if r["total"] and r["total"] > 0 else 100)
            for r in subj_rows
        )
        pct = round(cum_obtained / cum_max * 100, 2) if cum_max else 0

        grade_letter, grade_label, grade_color = calc_grade(pct)
        status, fail_count, failed_subjs = calc_result_status(subj_rows)
        section = assign_section(rank, total_count)

        # Check if any subject is AB (absent: marks=0 and special flag)
        has_absent = any(
            r.get("exam_type","").upper() == "ABSENT" or
            str(r.get("marks","")).upper() == "AB"
            for r in subj_rows
        )

        record = {
            "rank":          rank,
            "roll":          s["roll"],
            "name":          s["name"],
            "dept":          s["dept"],
            "year":          s["year"],
            "semester":      s["semester"],
            "section":       section,
            "subjects":      subj_rows,
            "cum_obtained":  round(cum_obtained, 1),
            "cum_max":       round(cum_max, 1),
            "percentage":    pct,
            "grade":         grade_letter,
            "grade_label":   grade_label,
            "grade_color":   grade_color,
            "status":        status,
            "fail_count":    fail_count,
            "failed_subjs":  failed_subjs,
            "has_absent":    has_absent,
        }
        result_list.append(record)

    # ── Apply post-computed filters ──
    if section_filter:
        result_list = [r for r in result_list if r["section"] == section_filter]
    if status_filter:
        result_list = [r for r in result_list if r["status"] == status_filter]
    if search:
        q = search.lower()
        result_list = [
            r for r in result_list
            if q in r["name"].lower() or q in r["roll"].lower()
        ]

    return result_list

def get_subject_list(dept=None, semester=None):
    """Return distinct subject names for the given filter."""
    sql = "SELECT DISTINCT subject FROM results WHERE 1=1"
    p = []
    if dept:     sql += " AND department=%s"; p.append(dept)
    if semester: sql += " AND semester=%s";   p.append(semester)
    sql += " ORDER BY subject"
    return [r["subject"] for r in _qry(sql, p)]

# ═══════════════════════════════════════════════════════════
#  ROUTE 1 — MAIN DASHBOARD
# ═══════════════════════════════════════════════════════════

@results_bp.route("/results_dashboard")
@admin_required
def results_dashboard():
    dept     = request.args.get("dept","").strip()
    semester = request.args.get("semester","").strip()
    year     = request.args.get("year","").strip()
    section  = request.args.get("section","").strip()
    status   = request.args.get("status","").strip()
    search   = request.args.get("q","").strip()
    page     = max(1, int(request.args.get("page","1") or 1))
    per_page = 50

    students = build_student_records(
        dept=dept, semester=semester, year=year,
        section_filter=section or None,
        status_filter=status or None,
        search=search or None,
    )

    # ── Summary KPIs ──
    total_students = len(students)
    pass_count  = sum(1 for s in students if s["status"] == "PASS")
    atkt_count  = sum(1 for s in students if s["status"] == "ATKT")
    fail_count  = sum(1 for s in students if s["status"] == "FAIL")
    absent_count= sum(1 for s in students if s["has_absent"])

    avg_pct = round(
        sum(s["percentage"] for s in students) / total_students, 1
    ) if total_students else 0

    topper = max(students, key=lambda s: s["percentage"]) if students else None

    # ── Grade distribution ──
    grade_dist = {"O":0,"A+":0,"A":0,"B+":0,"B":0,"C":0,"F":0}
    for s in students:
        grade_dist[s["grade"]] = grade_dist.get(s["grade"], 0) + 1

    # ── Section-wise summary ──
    sec_summary = {}
    for sec in SECTIONS:
        sec_students = [s for s in students if s["section"] == sec]
        if sec_students:
            sec_summary[sec] = {
                "count": len(sec_students),
                "avg":   round(sum(s["percentage"] for s in sec_students) / len(sec_students), 1),
                "pass":  sum(1 for s in sec_students if s["status"] == "PASS"),
            }

    # ── Pagination ──
    total_pages = max(1, (total_students + per_page - 1) // per_page)
    page = min(page, total_pages)
    paginated = students[(page-1)*per_page : page*per_page]

    # Subject list for table headers
    subjects = get_subject_list(dept or None, semester or None)

    return render_template("results/results_dashboard.html",
        students=paginated,
        all_students=students,
        subjects=subjects,
        # Filters
        dept=dept, semester=semester, year=year,
        section=section, status=status, search=search,
        # KPIs
        total_students=total_students,
        pass_count=pass_count, atkt_count=atkt_count, fail_count=fail_count,
        absent_count=absent_count, avg_pct=avg_pct, topper=topper,
        grade_dist=grade_dist,
        sec_summary=sec_summary,
        # Pagination
        page=page, total_pages=total_pages, per_page=per_page,
        # Config
        DEPARTMENTS=DEPARTMENTS, SEMESTERS=SEMESTERS, SECTIONS=SECTIONS,
        GRADE_TABLE=GRADE_TABLE,
    )

@results_bp.route("/results_import", methods=["POST"])
@admin_required
def results_import():
    from openpyxl import load_workbook
    from flask import flash, redirect, current_app
    import re
    
    f = request.files.get("file")
    if not f:
        flash("No file selected", "error")
        return redirect("/results_dashboard")
        
    try:
        wb = load_workbook(f, data_only=True)
        current_app.logger.info(f"Uploaded file: {f.filename}, Sheet names: {wb.sheetnames}")
        sheets_to_process = []
        for name in wb.sheetnames:
            ws = wb[name]
            has_student_data = False
            for r in range(1, min(ws.max_row + 1, 11)):
                row_vals = [str(ws.cell(r, c).value or "").lower().strip() for c in range(1, min(ws.max_column + 1, 15))]
                if any("student name" in v or "prn" in v or "roll" in v for v in row_vals):
                    has_student_data = True
                    break
            if has_student_data:
                sheets_to_process.append(ws)
                
        current_app.logger.info(f"Sheets containing student data to process: {[s.title for s in sheets_to_process]}")
        if not sheets_to_process:
            flash("No sheets with valid student name or PRN columns found in the Excel file.", "error")
            return redirect("/results_dashboard")
            
        added = 0
        for ws in sheets_to_process:
            current_app.logger.info(f"--- Inspecting Sheet: {ws.title} ---")
            for r in range(1, 5):
                row_vals = [f"Col {c}: {ws.cell(r, c).value}" for c in range(1, min(ws.max_column + 1, 16)) if ws.cell(r, c).value is not None]
                current_app.logger.info(f"Row {r} non-empty cells: {row_vals}")
                
            # 1. Detect header row
            hdr_row = 1
            for r in range(1, min(ws.max_row + 1, 11)):
                row_vals = [str(ws.cell(r, c).value or "").lower().strip() for c in range(1, min(ws.max_column + 1, 15))]
                if any("student name" in v or "prn" in v or "roll" in v for v in row_vals):
                    hdr_row = r
                    break
                    
            hdrs = [str(ws.cell(hdr_row, c).value or "").lower().strip() for c in range(1, ws.max_column + 1)]
            current_app.logger.info(f"Sheet '{ws.title}': Detected hdr_row={hdr_row}")
            current_app.logger.info(f"Sheet '{ws.title}': First 15 headers={hdrs[:15]}")
            
            def get_col_index(keywords):
                for kw in keywords:
                    for idx, h in enumerate(hdrs):
                        if kw in h:
                            return idx + 1
                return None
                
            col_name = get_col_index(["student name", "name", "student"])
            col_roll = get_col_index(["roll no", "roll number", "roll", "prn"])
            col_dept = get_col_index(["department", "dept", "branch"])
            col_year = get_col_index(["year", "class year"])
            col_sem  = get_col_index(["semester", "sem"])
            
            # Detect flat format columns
            col_sub  = get_col_index(["subject", "sub", "course"])
            col_marks = get_col_index(["marks", "obtained", "score"])
            col_total = get_col_index(["total", "max", "out of"])
            col_exam = get_col_index(["exam type", "exam", "type"])
            
            current_app.logger.info(f"Sheet '{ws.title}': col_name={col_name}, col_roll={col_roll}, col_sub={col_sub}, col_marks={col_marks}")
            if not col_name:
                current_app.logger.warning(f"Sheet '{ws.title}': Skipped because name column was not found.")
                continue
                
            # Global sheet level metadata resolution
            sheet_dept = "CS"
            title_lower = ws.title.lower()
            if "comp" in title_lower or "cs" in title_lower:
                sheet_dept = "CS"
            elif "aids" in title_lower:
                sheet_dept = "AIDS"
            elif "aiml" in title_lower:
                sheet_dept = "AIML"
            elif "it" in title_lower:
                sheet_dept = "IT"
                
            sheet_sem = None
            for r in range(1, min(ws.max_row + 1, 6)):
                for c in range(1, min(ws.max_column + 1, 16)):
                    val = str(ws.cell(r, c).value or "").strip().lower()
                    if not val:
                        continue
                    if "sem:" in val or "semester" in val:
                        for s in SEMESTERS:
                            if s.lower() in val:
                                sheet_sem = s
                                break
                    elif "sem" in val:
                        for s in SEMESTERS:
                            if f"sem {s.lower()}" in val or f"sem-{s.lower()}" in val or f"sem: {s.lower()}" in val:
                                sheet_sem = s
                                break
                    if "computer" in val or "comp" in val or "cs" in val:
                        sheet_dept = "CS"
                    elif "aids" in val or "artificial" in val:
                        sheet_dept = "AIDS"
                    elif "aiml" in val:
                        sheet_dept = "AIML"
                    elif "information technology" in val or "it" in val:
                        sheet_dept = "IT"
            
            if not sheet_sem:
                filename_lower = f.filename.lower()
                for s in SEMESTERS:
                    if f"sem_{s.lower()}" in filename_lower or f"sem {s.lower()}" in filename_lower or f"sem-{s.lower()}" in filename_lower or f"sem{s.lower()}" in filename_lower:
                        sheet_sem = s
                        break
            
            if not sheet_sem:
                sheet_sem = "I"
                
            def sem_to_year(sem_val):
                if sem_val in ("I", "II"): return "I"
                if sem_val in ("III", "IV"): return "II"
                if sem_val in ("V", "VI"): return "III"
                if sem_val in ("VII", "VIII"): return "IV"
                return "I"
                
            sheet_year = sem_to_year(sheet_sem)
            is_flat = col_sub is not None and col_marks is not None
            current_app.logger.info(f"Sheet '{ws.title}': is_flat={is_flat}, sheet_dept={sheet_dept}, sheet_sem={sheet_sem}")
            
            if is_flat:
                # FLAT FORMAT (one row per student per subject)
                for r_idx in range(hdr_row + 1, ws.max_row + 1):
                    name = str(ws.cell(r_idx, col_name).value or "").strip()
                    if not name or name.lower() in ("name", "student name", "student"):
                        continue
                    if any(name.lower().startswith(x) for x in ("total", "average", "class", "topper", "pass", "fail", "atkt")):
                        continue
                        
                    roll = str(ws.cell(r_idx, col_roll).value or "").strip() if col_roll else ""
                    dept = str(ws.cell(r_idx, col_dept).value or "").strip() if col_dept else ""
                    year = str(ws.cell(r_idx, col_year).value or "").strip() if col_year else ""
                    sem  = str(ws.cell(r_idx, col_sem).value or "").strip() if col_sem else ""
                    sub  = str(ws.cell(r_idx, col_sub).value or "").strip() if col_sub else ""
                    
                    marks_val = ws.cell(r_idx, col_marks).value
                    total_val = ws.cell(r_idx, col_total).value if col_total else None
                    exam_val = str(ws.cell(r_idx, col_exam).value or "Semester Exam").strip() if col_exam else "Semester Exam"
                    
                    if str(marks_val or "").strip().upper() == "AB":
                        marks = 0.0
                        exam_val = "ABSENT"
                    else:
                        try:
                            marks = float(marks_val) if marks_val is not None else 0.0
                        except ValueError:
                            continue
                            
                    total = 60.0
                    if total_val is not None:
                        try:
                            total = float(total_val)
                        except ValueError:
                            pass
                            
                    if roll and (not dept or not year or not sem):
                        student = _qone("SELECT department, year FROM students WHERE roll=%s", (roll,))
                        if student:
                            if not dept: dept = student["department"]
                            if not year: year = student["year"]
                    if not dept: dept = sheet_dept
                    if not sem: sem = sheet_sem
                    if not year: year = sem_to_year(sem)
                    
                    pct = (marks / total * 100) if total > 0 else 0
                    grade, _, _ = calc_grade(pct)
                    result = "Pass" if pct >= 40 else "Fail"
                    
                    existing = _qone(
                        "SELECT id FROM results WHERE student_name=%s AND roll=%s AND subject=%s AND exam_type=%s",
                        (name, roll, sub, exam_val)
                    )
                    if existing:
                        _exe(
                            "UPDATE results SET department=%s, year=%s, semester=%s, marks=%s, total=%s, grade=%s, result=%s, published=1 WHERE id=%s",
                            (dept, year, sem, marks, total, grade, result, existing["id"])
                        )
                    else:
                        _exe(
                            "INSERT INTO results (student_name, roll, department, year, semester, subject, marks, total, exam_type, grade, result, published) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)",
                            (name, roll, dept, year, sem, sub, marks, total, exam_val, grade, result)
                        )
                    added += 1
            else:
                # CUMULATIVE GRID FORMAT (each student is one row, subjects are columns)
                def get_subject_name(c_idx):
                    min_col = max(col_name, col_roll) + 1
                    # Try Row hdr_row - 1
                    if hdr_row > 1:
                        for temp_c in range(c_idx, min_col - 1, -1):
                            val = ws.cell(hdr_row - 1, temp_c).value
                            if val and str(val).strip():
                                return str(val).strip()
                    # Try Row hdr_row - 2
                    if hdr_row > 2:
                        for temp_c in range(c_idx, min_col - 1, -1):
                            val = ws.cell(hdr_row - 2, temp_c).value
                            if val and str(val).strip():
                                return str(val).strip()
                    return None
                    
                def parse_max_marks(h_val):
                    m = re.search(r"\d+", h_val)
                    if m:
                        return float(m.group())
                    return 60.0
                    
                subject_cols = []
                for idx in range(1, ws.max_column + 1):
                    if idx in (col_name, col_roll):
                        continue
                    h_val = str(ws.cell(hdr_row, idx).value or "").strip().lower()
                    if "total" in h_val:
                        subj_name = get_subject_name(idx)
                        if subj_name:
                            subject_cols.append({
                                "col_idx": idx,
                                "subject": subj_name,
                                "total": parse_max_marks(h_val)
                            })
                            
                current_app.logger.info(f"Sheet '{ws.title}': Detected subject columns={subject_cols}")
                if not subject_cols:
                    current_app.logger.warning(f"Sheet '{ws.title}': Skipped because no subject columns with 'TOTAL' in headers were found.")
                    continue
                    
                for r_idx in range(hdr_row + 1, ws.max_row + 1):
                    name = str(ws.cell(r_idx, col_name).value or "").strip()
                    if not name or name.lower() in ("name", "student name", "student"):
                        continue
                    if any(name.lower().startswith(x) for x in ("total", "average", "class", "topper", "pass", "fail", "atkt")):
                        continue
                        
                    roll = str(ws.cell(r_idx, col_roll).value or "").strip() if col_roll else ""
                    dept = str(ws.cell(r_idx, col_dept).value or "").strip() if col_dept else ""
                    year = str(ws.cell(r_idx, col_year).value or "").strip() if col_year else ""
                    sem  = str(ws.cell(r_idx, col_sem).value or "").strip() if col_sem else ""
                    
                    if roll and (not dept or not year or not sem):
                        student = _qone("SELECT department, year FROM students WHERE roll=%s", (roll,))
                        if student:
                            if not dept: dept = student["department"]
                            if not year: year = student["year"]
                    if not dept: dept = sheet_dept
                    if not sem: sem = sheet_sem
                    if not year: year = sem_to_year(sem)
                    
                    current_app.logger.info(f"Sheet '{ws.title}': Row {r_idx} student name='{name}', roll='{roll}', dept='{dept}', sem='{sem}'")
                    for sc in subject_cols:
                        sub = sc["subject"]
                        total = sc["total"]
                        marks_val = ws.cell(r_idx, sc["col_idx"]).value
                        
                        exam_val = "Semester Exam"
                        if str(marks_val or "").strip().upper() == "AB":
                            marks = 0.0
                            exam_val = "ABSENT"
                        else:
                            try:
                                marks = float(marks_val) if marks_val is not None else 0.0
                            except ValueError:
                                continue
                                
                        pct = (marks / total * 100) if total > 0 else 0
                        grade, _, _ = calc_grade(pct)
                        result = "Pass" if pct >= 40 else "Fail"
                        
                        existing = _qone(
                            "SELECT id FROM results WHERE student_name=%s AND roll=%s AND subject=%s AND exam_type=%s",
                            (name, roll, sub, exam_val)
                        )
                        if existing:
                            _exe(
                                "UPDATE results SET department=%s, year=%s, semester=%s, marks=%s, total=%s, grade=%s, result=%s, published=1 WHERE id=%s",
                                (dept, year, sem, marks, total, grade, result, existing["id"])
                            )
                        else:
                            _exe(
                                "INSERT INTO results (student_name, roll, department, year, semester, subject, marks, total, exam_type, grade, result, published) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)",
                                (name, roll, dept, year, sem, sub, marks, total, exam_val, grade, result)
                            )
                        added += 1
                        
        flash(f"Successfully imported/updated {added} student results.", "success")
        
    except Exception as e:
        import traceback
        current_app.logger.error(f"Excel import error: {traceback.format_exc()}")
        flash(f"Error parsing file: {str(e)}", "error")
        
    return redirect("/results_dashboard")

# ═══════════════════════════════════════════════════════════
#  ROUTE 2 — ANALYTICS PAGE
# ═══════════════════════════════════════════════════════════

@results_bp.route("/results_analytics")
@admin_required
def results_analytics():
    dept     = request.args.get("dept","").strip()
    semester = request.args.get("semester","").strip()

    students = build_student_records(dept=dept or None, semester=semester or None)
    subjects = get_subject_list(dept or None, semester or None)

    # ── Subject-wise analytics ──
    subj_analytics = []
    for subj in subjects:
        subj_marks = []
        below40 = 0
        for s in students:
            for r in s["subjects"]:
                if r["subject"] == subj:
                    m = r["marks"] if r["marks"] is not None else 0
                    t = r["total"] if r["total"] and r["total"] > 0 else 100
                    pct_val = m / t * 100
                    subj_marks.append(pct_val)
                    if pct_val < 40:
                        below40 += 1
        if subj_marks:
            subj_analytics.append({
                "subject":  subj,
                "avg":      round(sum(subj_marks) / len(subj_marks), 1),
                "highest":  round(max(subj_marks), 1),
                "lowest":   round(min(subj_marks), 1),
                "below40":  below40,
                "count":    len(subj_marks),
                "pass_rate":round((len(subj_marks)-below40)/len(subj_marks)*100,1),
            })

    # ── Top 10 students ──
    top10 = sorted(students, key=lambda s: s["percentage"], reverse=True)[:10]

    # ── Section-wise breakdown ──
    sec_breakdown = {}
    for sec in SECTIONS:
        sec_s = [s for s in students if s["section"] == sec]
        if not sec_s:
            continue
        sec_breakdown[sec] = {
            "count":       len(sec_s),
            "avg":         round(sum(s["percentage"] for s in sec_s)/len(sec_s),1),
            "pass":        sum(1 for s in sec_s if s["status"]=="PASS"),
            "atkt":        sum(1 for s in sec_s if s["status"]=="ATKT"),
            "fail":        sum(1 for s in sec_s if s["status"]=="FAIL"),
            "topper":      max(sec_s, key=lambda s: s["percentage"]),
        }

    # ── Failed students per subject ──
    failed_per_subj = {subj: 0 for subj in subjects}
    for s in students:
        for subj in s["failed_subjs"]:
            if subj in failed_per_subj:
                failed_per_subj[subj] += 1

    # ── Grade distribution ──
    grade_dist = {"O":0,"A+":0,"A":0,"B+":0,"B":0,"C":0,"F":0}
    for s in students:
        grade_dist[s["grade"]] = grade_dist.get(s["grade"],0) + 1

    total_students = len(students)
    avg_pct = round(sum(s["percentage"] for s in students)/total_students,1) if total_students else 0
    fail_count = sum(1 for s in students if s["status"] == "FAIL")

    return render_template("results/results_analytics.html",
        dept=dept, semester=semester,
        total_students=total_students, avg_pct=avg_pct,
        fail_count=fail_count,
        subj_analytics=subj_analytics, top10=top10,
        sec_breakdown=sec_breakdown, failed_per_subj=failed_per_subj,
        grade_dist=grade_dist,
        DEPARTMENTS=DEPARTMENTS, SEMESTERS=SEMESTERS,
        GRADE_TABLE=GRADE_TABLE,
    )

# ═══════════════════════════════════════════════════════════
#  ROUTE 3 — INDIVIDUAL REPORT CARD (printable PDF)
# ═══════════════════════════════════════════════════════════

@results_bp.route("/results_reportcard/<roll>")
@admin_required
def results_reportcard(roll):
    dept     = request.args.get("dept","").strip()
    semester = request.args.get("semester","").strip()

    all_students = build_student_records(dept=dept or None, semester=semester or None)
    student = next((s for s in all_students if s["roll"] == roll), None)
    if not student:
        return redirect("/results_dashboard?error=student_not_found")

    # Class rank
    ranked = sorted(all_students, key=lambda s: s["percentage"], reverse=True)
    rank = next((i+1 for i,s in enumerate(ranked) if s["roll"]==roll), "—")

    return render_template("results/results_reportcard.html",
        student=student, rank=rank,
        total_students=len(all_students),
        dept=dept, semester=semester,
        GRADE_TABLE=GRADE_TABLE,
        now=datetime.now().strftime("%d %b %Y"),
    )

# ═══════════════════════════════════════════════════════════
#  ROUTE 4 — EXPORT EXCEL (full formatted workbook)
# ═══════════════════════════════════════════════════════════

@results_bp.route("/results_export_excel")
@admin_required
def results_export_excel():
    dept     = request.args.get("dept","").strip()
    semester = request.args.get("semester","").strip()
    section  = request.args.get("section","").strip()
    status   = request.args.get("status","").strip()

    students = build_student_records(
        dept=dept or None, semester=semester or None,
        section_filter=section or None,
        status_filter=status or None,
    )
    subjects = get_subject_list(dept or None, semester or None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # ── Styles ──
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
    SUBJ_FILL = PatternFill("solid", fgColor="2563EB")
    ALT_FILL  = PatternFill("solid", fgColor="F8FAFC")
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN_SIDE = Side(style="thin", color="CBD5E1")
    BORDER    = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

    STATUS_COLORS = {"PASS":"C6EFCE","ATKT":"FFEB9C","FAIL":"FFC7CE"}
    GRADE_COLORS  = {
        "O":"A7F3D0","A+":"BFDBFE","A":"DDD6FE",
        "B+":"FDE68A","B":"FED7AA","C":"FECACA","F":"FCA5A5"
    }

    # ── Title row ──
    title = f"Results — {dept or 'All Depts'} | SEM {semester or 'All'}"
    ws.merge_cells(f"A1:{get_column_letter(8 + len(subjects))}1")
    t_cell = ws["A1"]
    t_cell.value = title
    t_cell.font  = Font(bold=True, size=14, color="FFFFFF")
    t_cell.fill  = PatternFill("solid", fgColor="0F172A")
    t_cell.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # ── Header row ──
    base_headers = ["#", "Roll No.", "Student Name", "Section",
                    "Dept", "Semester"]
    all_headers  = base_headers + subjects + ["Total Obtained", "Total Max",
                                               "Percentage", "Grade", "Status"]
    for c, h in enumerate(all_headers, 1):
        cell = ws.cell(2, c, h)
        cell.font      = HDR_FONT
        cell.fill      = SUBJ_FILL if c > len(base_headers) and c <= len(base_headers)+len(subjects) else HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
    ws.row_dimensions[2].height = 36

    # ── Data rows ──
    for ri, s in enumerate(students, 3):
        row_fill = ALT_FILL if ri % 2 == 0 else None
        subj_marks_map = {r["subject"]: r["marks"] for r in s["subjects"]}

        base_vals = [ri-2, s["roll"], s["name"], s["section"],
                     s["dept"], s["semester"]]
        subj_vals = [subj_marks_map.get(subj, "—") for subj in subjects]
        extra_vals = [s["cum_obtained"], s["cum_max"],
                      f"{s['percentage']}%", s["grade"], s["status"]]

        all_vals = base_vals + subj_vals + extra_vals
        for ci, val in enumerate(all_vals, 1):
            cell = ws.cell(ri, ci, val)
            cell.alignment = CENTER
            cell.border    = BORDER
            if row_fill:
                cell.fill = row_fill

            col_count = len(base_headers) + len(subjects)
            if ci == col_count + 3:  # Percentage
                cell.font = Font(bold=True)
            if ci == col_count + 4:  # Grade
                gc = GRADE_COLORS.get(s["grade"])
                if gc:
                    cell.fill = PatternFill("solid", fgColor=gc)
                    cell.font = Font(bold=True)
            if ci == col_count + 5:  # Status
                sc = STATUS_COLORS.get(s["status"])
                if sc:
                    cell.fill = PatternFill("solid", fgColor=sc)
                    cell.font = Font(bold=True)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    for i in range(len(base_headers)+1, len(base_headers)+len(subjects)+6):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # ── Freeze panes ──
    ws.freeze_panes = "D3"

    # ════ Sheet 2: Subject-wise Analytics ════
    ws2 = wb.create_sheet("Analytics")
    ws2["A1"] = "Subject-wise Analytics"
    ws2["A1"].font = Font(bold=True, size=13)

    ah = ["Subject","Avg %","Highest %","Lowest %","Pass Rate %","Below 40 Count"]
    for c, h in enumerate(ah, 1):
        cell = ws2.cell(2, c, h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CENTER; cell.border = BORDER

    subjects_analytics = []
    for subj in subjects:
        vals = []
        below40 = 0
        for s in students:
            for r in s["subjects"]:
                if r["subject"] == subj:
                    m = r["marks"] if r["marks"] is not None else 0
                    t = r["total"] if r["total"] and r["total"] > 0 else 100
                    p = m / t * 100
                    vals.append(p)
                    if p < 40: below40 += 1
        if vals:
            avg_v = round(sum(vals)/len(vals),1)
            pr    = round((len(vals)-below40)/len(vals)*100,1)
            subjects_analytics.append([subj, avg_v, round(max(vals),1),
                                        round(min(vals),1), pr, below40])

    for ri, row in enumerate(subjects_analytics, 3):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(ri, ci, val)
            cell.alignment = CENTER; cell.border = BORDER
            if ci == 2:  # Avg
                fill_col = "C6EFCE" if row[1] >= 60 else ("FFEB9C" if row[1] >= 40 else "FFC7CE")
                cell.fill = PatternFill("solid", fgColor=fill_col)

    for i, w in enumerate([40,12,12,12,14,14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ════ Sheet 3: Top 10 ════
    ws3 = wb.create_sheet("Top 10")
    ws3["A1"] = "Top 10 Performers"
    ws3["A1"].font = Font(bold=True, size=13)
    top_h = ["Rank","Roll","Name","Section","%","Grade","Status"]
    for c,h in enumerate(top_h,1):
        cell = ws3.cell(2,c,h)
        cell.font=HDR_FONT; cell.fill=HDR_FILL
        cell.alignment=CENTER; cell.border=BORDER
    top10 = sorted(students, key=lambda s: s["percentage"], reverse=True)[:10]
    for ri, s in enumerate(top10, 3):
        for ci,v in enumerate([ri-2,s["roll"],s["name"],s["section"],
                                f"{s['percentage']}%",s["grade"],s["status"]],1):
            cell = ws3.cell(ri,ci,v)
            cell.alignment=CENTER; cell.border=BORDER
            if ri == 3:
                cell.fill = PatternFill("solid",fgColor="FDE68A")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"results_{dept or 'all'}_{semester or 'all'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════════════════════
#  ROUTE 5 — API: Chart data (JSON)
# ═══════════════════════════════════════════════════════════

@results_bp.route("/results_chart_data")
@admin_required
def results_chart_data():
    dept     = request.args.get("dept","").strip()
    semester = request.args.get("semester","").strip()

    students = build_student_records(dept=dept or None, semester=semester or None)
    subjects = get_subject_list(dept or None, semester or None)

    # Grade distribution
    grade_dist = {"O":0,"A+":0,"A":0,"B+":0,"B":0,"C":0,"F":0}
    for s in students:
        grade_dist[s["grade"]] = grade_dist.get(s["grade"],0) + 1

    # Subject avg
    subj_avgs = {}
    for subj in subjects:
        vals = [
            (r["marks"] or 0) / (r["total"] or 100) * 100
            for s in students for r in s["subjects"] if r["subject"]==subj
        ]
        subj_avgs[subj] = round(sum(vals)/len(vals),1) if vals else 0

    # Section performance
    sec_perf = {}
    for sec in SECTIONS:
        ss = [s["percentage"] for s in students if s["section"]==sec]
        sec_perf[sec] = round(sum(ss)/len(ss),1) if ss else 0

    return jsonify({
        "grade_dist":  grade_dist,
        "subj_avgs":   subj_avgs,
        "sec_perf":    sec_perf,
        "status_dist": {
            "PASS":  sum(1 for s in students if s["status"]=="PASS"),
            "ATKT":  sum(1 for s in students if s["status"]=="ATKT"),
            "FAIL":  sum(1 for s in students if s["status"]=="FAIL"),
        },
        "total": len(students),
    })

# ═══════════════════════════════════════════════════════════
#  LEGACY RESULTS MANAGEMENT MODULE
# ═══════════════════════════════════════════════════════════

def pct(obtained, total):
    return (obtained / total * 100) if total else 0.0

def grade(obtained, total):
    p = pct(obtained, total)
    for threshold, letter, _, _ in GRADE_TABLE:
        if p >= threshold:
            return letter
    return "F"

def _do_import(f):
    from openpyxl import load_workbook
    from flask import flash, current_app, session
    import re
    from services.results_service import get_components_for_subject, parse_marks_value
    
    wb = load_workbook(f, data_only=True)
    current_app.logger.info(f"Uploaded file: {f.filename}, Sheet names: {wb.sheetnames}")
    sheets_to_process = []
    for name in wb.sheetnames:
        ws = wb[name]
        has_student_data = False
        for r in range(1, min(ws.max_row + 1, 11)):
            row_vals = [str(ws.cell(r, c).value or "").lower().strip() for c in range(1, min(ws.max_column + 1, 15))]
            if any("student name" in v or "prn" in v or "roll" in v for v in row_vals):
                has_student_data = True
                break
        if has_student_data:
            sheets_to_process.append(ws)
            
    current_app.logger.info(f"Sheets containing student data to process: {[s.title for s in sheets_to_process]}")
    if not sheets_to_process:
        flash("No sheets with valid student name or PRN columns found in the Excel file.", "error")
        return False
        
    session["import_warnings"] = []
    added = 0
    for ws in sheets_to_process:
        current_app.logger.info(f"--- Inspecting Sheet: {ws.title} ---")
        for r in range(1, 5):
            row_vals = [f"Col {c}: {ws.cell(r, c).value}" for c in range(1, min(ws.max_column + 1, 16)) if ws.cell(r, c).value is not None]
            current_app.logger.info(f"Row {r} non-empty cells: {row_vals}")
            
        hdr_row = 1
        for r in range(1, min(ws.max_row + 1, 11)):
            row_vals = [str(ws.cell(r, c).value or "").lower().strip() for c in range(1, min(ws.max_column + 1, 15))]
            if any("student name" in v or "prn" in v or "roll" in v for v in row_vals):
                hdr_row = r
                break
                
        hdrs = [str(ws.cell(hdr_row, c).value or "").lower().strip() for c in range(1, ws.max_column + 1)]
        current_app.logger.info(f"Sheet '{ws.title}': Detected hdr_row={hdr_row}")
        
        def get_col_index(keywords):
            for kw in keywords:
                for idx, h in enumerate(hdrs):
                    if kw in h:
                        return idx + 1
            return None
            
        col_name = get_col_index(["student name", "name", "student"])
        col_roll = get_col_index(["roll no", "roll number", "roll", "prn"])
        col_dept = get_col_index(["department", "dept", "branch"])
        col_year = get_col_index(["year", "class year"])
        col_sem  = get_col_index(["semester", "sem"])
        
        col_sub  = get_col_index(["subject", "sub", "course"])
        col_marks = get_col_index(["marks", "obtained", "score"])
        col_total = get_col_index(["total", "max", "out of"])
        col_exam = get_col_index(["exam type", "exam", "type"])
        
        if not col_name:
            current_app.logger.warning(f"Sheet '{ws.title}': Skipped because name column was not found.")
            continue
            
        sheet_dept = "CS"
        title_lower = ws.title.lower()
        if "comp" in title_lower or "cs" in title_lower:
            sheet_dept = "CS"
        elif "aids" in title_lower:
            sheet_dept = "AIDS"
        elif "aiml" in title_lower:
            sheet_dept = "AIML"
        elif "it" in title_lower:
            sheet_dept = "IT"
            
        sheet_sem = None
        for r in range(1, min(ws.max_row + 1, 6)):
            for c in range(1, min(ws.max_column + 1, 16)):
                val = str(ws.cell(r, c).value or "").strip().lower()
                if not val:
                    continue
                if "sem:" in val or "semester" in val:
                    for s in SEMESTERS:
                        if s.lower() in val:
                            sheet_sem = s
                            break
                elif "sem" in val:
                    for s in SEMESTERS:
                        if f"sem {s.lower()}" in val or f"sem-{s.lower()}" in val or f"sem: {s.lower()}" in val:
                            sheet_sem = s
                            break
                if "computer" in val or "comp" in val or "cs" in val:
                    sheet_dept = "CS"
                elif "aids" in val or "artificial" in val:
                    sheet_dept = "AIDS"
                elif "aiml" in val:
                    sheet_dept = "AIML"
                elif "information technology" in val or "it" in val:
                    sheet_dept = "IT"
        
        if not sheet_sem:
            filename_lower = f.filename.lower()
            for s in SEMESTERS:
                if f"sem_{s.lower()}" in filename_lower or f"sem {s.lower()}" in filename_lower or f"sem-{s.lower()}" in filename_lower or f"sem{s.lower()}" in filename_lower:
                    sheet_sem = s
                    break
        
        if not sheet_sem:
            sheet_sem = "I"
            
        def sem_to_year(sem_val):
            if sem_val in ("I", "II"): return "I"
            if sem_val in ("III", "IV"): return "II"
            if sem_val in ("V", "VI"): return "III"
            if sem_val in ("VII", "VIII"): return "IV"
            return "I"
            
        sheet_year = sem_to_year(sheet_sem)
        is_flat = col_sub is not None and col_marks is not None
        
        if is_flat:
            for r_idx in range(hdr_row + 1, ws.max_row + 1):
                name = str(ws.cell(r_idx, col_name).value or "").strip()
                if not name or name.lower() in ("name", "student name", "student"):
                    continue
                if any(name.lower().startswith(x) for x in ("total", "average", "class", "topper", "pass", "fail", "atkt")):
                    continue
                    
                roll = str(ws.cell(r_idx, col_roll).value or "").strip() if col_roll else ""
                dept = str(ws.cell(r_idx, col_dept).value or "").strip() if col_dept else ""
                year = str(ws.cell(r_idx, col_year).value or "").strip() if col_year else ""
                sem  = str(ws.cell(r_idx, col_sem).value or "").strip() if col_sem else ""
                sub  = str(ws.cell(r_idx, col_sub).value or "").strip() if col_sub else ""
                
                marks_val = ws.cell(r_idx, col_marks).value
                total_val = ws.cell(r_idx, col_total).value if col_total else None
                exam_val = str(ws.cell(r_idx, col_exam).value or "Semester Exam").strip() if col_exam else "Semester Exam"
                
                # Stage 3: unified AB handling via parse_marks_value
                marks, is_absent = parse_marks_value(marks_val)
                if is_absent:
                    exam_val = "ABSENT"
                    marks = 0.0
                elif marks is None:
                    continue
                        
                if roll and (not dept or not year or not sem):
                    student = _qone("SELECT department, year FROM students WHERE roll=%s", (roll,))
                    if student:
                        if not dept: dept = student["department"]
                        if not year: year = student["year"]
                if not dept: dept = sheet_dept
                if not sem: sem = sheet_sem
                if not year: year = sem_to_year(sem)
                
                # Stage 3: check configured max_total from subject_mark_components
                sub_info = get_components_for_subject(sub, dept, sem)
                db_max_total = sub_info["max_total"]

                total_val_float = None
                if total_val is not None:
                    try:
                        total_val_float = float(total_val)
                    except ValueError:
                        pass

                # Mismatch = warning row; skip insert, require admin confirmation
                if total_val_float is not None and abs(total_val_float - db_max_total) > 0.01:
                    warnings = session.get("import_warnings", [])
                    warnings.append(
                        f"MISMATCH|Row {r_idx} ({ws.title}): '{sub}' sheet total "
                        f"({total_val_float}) != configured total ({db_max_total}). "
                        f"Row skipped - confirm manually to import."
                    )
                    session["import_warnings"] = warnings
                    continue

                total = db_max_total
                
                pct_val = (marks / total * 100) if total > 0 else 0
                grade_letter, _, _ = calc_grade(pct_val)
                result = "Pass" if pct_val >= 40 else "Fail"
                
                existing = _qone(
                    "SELECT id FROM results WHERE student_name=%s AND roll=%s AND subject=%s AND exam_type=%s",
                    (name, roll, sub, exam_val)
                )
                if existing:
                    _exe(
                        "UPDATE results SET department=%s, year=%s, semester=%s, marks=%s, total=%s, grade=%s, result=%s, published=1 WHERE id=%s",
                        (dept, year, sem, marks, total, grade_letter, result, existing["id"])
                    )
                else:
                    _exe(
                        "INSERT INTO results (student_name, roll, department, year, semester, subject, marks, total, exam_type, grade, result, published) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)",
                        (name, roll, dept, year, sem, sub, marks, total, exam_val, grade_letter, result)
                    )
                added += 1
        else:
            def get_subject_name(c_idx):
                min_col = max(col_name, col_roll) + 1
                if hdr_row > 1:
                    for temp_c in range(c_idx, min_col - 1, -1):
                        val = ws.cell(hdr_row - 1, temp_c).value
                        if val and str(val).strip():
                            return str(val).strip()
                if hdr_row > 2:
                    for temp_c in range(c_idx, min_col - 1, -1):
                        val = ws.cell(hdr_row - 2, temp_c).value
                        if val and str(val).strip():
                            return str(val).strip()
                return None
                
            def parse_max_marks(h_val):
                m = re.search(r"\d+", h_val)
                if m: return float(m.group())
                return 60.0
                
            subject_cols = []
            for idx in range(1, ws.max_column + 1):
                if idx in (col_name, col_roll):
                    continue
                h_val = str(ws.cell(hdr_row, idx).value or "").strip().lower()
                if "total" in h_val:
                    subj_name = get_subject_name(idx)
                    if subj_name:
                        subject_cols.append({
                            "col_idx": idx,
                            "subject": subj_name,
                            "total": parse_max_marks(h_val)
                        })
                        
            if not subject_cols:
                continue
                
            for r_idx in range(hdr_row + 1, ws.max_row + 1):
                name = str(ws.cell(r_idx, col_name).value or "").strip()
                if not name or name.lower() in ("name", "student name", "student"):
                    continue
                if any(name.lower().startswith(x) for x in ("total", "average", "class", "topper", "pass", "fail", "atkt")):
                    continue
                    
                roll = str(ws.cell(r_idx, col_roll).value or "").strip() if col_roll else ""
                dept = str(ws.cell(r_idx, col_dept).value or "").strip() if col_dept else ""
                year = str(ws.cell(r_idx, col_year).value or "").strip() if col_year else ""
                sem  = str(ws.cell(r_idx, col_sem).value or "").strip() if col_sem else ""
                
                if roll and (not dept or not year or not sem):
                    student = _qone("SELECT department, year FROM students WHERE roll=%s", (roll,))
                    if student:
                        if not dept: dept = student["department"]
                        if not year: year = student["year"]
                if not dept: dept = sheet_dept
                if not sem: sem = sheet_sem
                if not year: year = sem_to_year(sem)
                
                for sc in subject_cols:
                    sub = sc["subject"]
                    sheet_total = sc["total"]
                    marks_val = ws.cell(r_idx, sc["col_idx"]).value
                    
                    exam_val = "Semester Exam"
                    # Stage 3: unified AB handling via parse_marks_value
                    marks, is_absent = parse_marks_value(marks_val)
                    if is_absent:
                        marks = 0.0
                        exam_val = "ABSENT"
                    elif marks is None:
                        continue

                    # Stage 3: check configured max from subject_mark_components
                    sub_info = get_components_for_subject(sub, dept, sem)
                    db_max_total = sub_info["max_total"]

                    if abs(sheet_total - db_max_total) > 0.01:
                        warnings = session.get("import_warnings", [])
                        warnings.append(
                            f"MISMATCH|Row {r_idx} ({ws.title}): '{sub}' sheet total "
                            f"({sheet_total}) != configured total ({db_max_total}). "
                            f"Row skipped - confirm manually to import."
                        )
                        session["import_warnings"] = warnings
                        continue

                    total = db_max_total
                            
                    pct_val = (marks / total * 100) if total > 0 else 0
                    grade_letter, _, _ = calc_grade(pct_val)
                    result = "Pass" if pct_val >= 40 else "Fail"
                    
                    existing = _qone(
                        "SELECT id FROM results WHERE student_name=%s AND roll=%s AND subject=%s AND exam_type=%s",
                        (name, roll, sub, exam_val)
                    )
                    if existing:
                        _exe(
                            "UPDATE results SET department=%s, year=%s, semester=%s, marks=%s, total=%s, grade=%s, result=%s, published=1 WHERE id=%s",
                            (dept, year, sem, marks, total, grade_letter, result, existing["id"])
                        )
                    else:
                        _exe(
                            "INSERT INTO results (student_name, roll, department, year, semester, subject, marks, total, exam_type, grade, result, published) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)",
                            (name, roll, dept, year, sem, sub, marks, total, exam_val, grade_letter, result)
                        )
                    added += 1
                    
    flash(f"Successfully imported/updated {added} student results.", "success")
    import_warns = session.get("import_warnings", [])
    if import_warns:
        flash(f"Import finished with {len(import_warns)} warnings/mismatches. Please check.", "warning")
    return True
@results_bp.route("/admin_results")
@admin_required
def admin_results():
    dept     = request.args.get("dept","").strip()
    semester = request.args.get("semester","").strip()
    year     = request.args.get("year","").strip()
    q        = request.args.get("q","").strip()
    published= request.args.get("published","").strip()
    exam_type = request.args.get("exam_type","").strip()

    sql = """SELECT r.*, f.name as faculty_name
             FROM results r LEFT JOIN faculty f ON r.faculty_id=f.id WHERE 1=1"""
    params = []
    if dept:      sql += " AND r.department=%s";       params.append(dept)
    if semester:  sql += " AND r.semester=%s";          params.append(semester)
    if year:      sql += " AND r.year=%s";              params.append(year)
    if q:         sql += " AND (r.student_name ILIKE %s OR r.roll ILIKE %s)"; params += [f"%{q}%",f"%{q}%"]
    if published != "": sql += " AND r.published=%s";  params.append(int(published))
    if exam_type: sql += " AND r.exam_type=%s";         params.append(exam_type)
    sql += " ORDER BY r.department, r.student_name, r.semester"
    results_list = _qry(sql, params)

    # Summary counts
    total_r    = _qone("SELECT COUNT(*) as c FROM results")["c"] or 0
    published_r= _qone("SELECT COUNT(*) as c FROM results WHERE published=1")["c"] or 0
    pass_r     = _qone("SELECT COUNT(*) as c FROM results WHERE result='Pass' AND published=1")["c"] or 0
    fail_r     = _qone("SELECT COUNT(*) as c FROM results WHERE result='Fail' AND published=1")["c"] or 0

    exam_types = [r["exam_type"] for r in _qry("SELECT DISTINCT exam_type FROM results WHERE exam_type IS NOT NULL AND exam_type != ''")]
    if not exam_types:
        exam_types = ["Semester Exam", "Internal Assessment", "Re-evaluation"]

    return render_template("admin/admin_results.html",
        results=results_list, dept=dept, semester=semester, year=year,
        q=q, published=published, exam_type=exam_type, exam_types=exam_types,
        total_r=total_r, published_r=published_r, pass_r=pass_r, fail_r=fail_r,
        DEPARTMENTS=DEPARTMENTS, SEMESTERS=SEMESTERS, YEARS=YEARS)

@results_bp.route("/admin_save_result", methods=["POST"])
@admin_required
def admin_save_result():
    student_name = request.form.get("student_name","").strip()
    roll_row = _qone("SELECT roll,department,year FROM students WHERE name=%s", (student_name,))
    roll = roll_row["roll"] if roll_row else request.form.get("roll","")
    dept = roll_row["department"] if roll_row else request.form.get("department","")
    yr   = roll_row["year"] if roll_row else request.form.get("year","")
    subject_val = request.form.get("subject","").strip()
    sem_val = request.form.get("semester","I").strip()

    # Stage 3: unified component lookup
    from services.results_service import get_components_for_subject, write_audit_log
    sub_info  = get_components_for_subject(subject_val, dept, sem_val)
    max_total = sub_info["max_total"]
    comp_caps = {c["component_name"].lower(): c["max_marks"] for c in sub_info["components"]}

    def _cap(form_key, comp_name, fallback_max):
        val = float(request.form.get(form_key, 0) or 0)
        limit = comp_caps.get(comp_name.lower(), fallback_max)
        return min(val, limit)

    assignment_marks    = _cap("assignment_marks",   "Assignment",        5.0)
    attendance_marks    = _cap("attendance_marks",   "Attendance",        5.0)
    teacher_assessment  = _cap("teacher_assessment", "Teacher Assessment",10.0)
    ut_marks            = _cap("ut_marks",           "Unit Test",         20.0)
    mse_marks           = _cap("mse_marks",          "Mid-Sem Exam",      20.0)
    tw_marks            = _cap("tw_marks",           "Term Work",          0.0)
    pr_or_marks         = _cap("pr_or_marks",        "Practical/Oral",     0.0)

    marks_val = assignment_marks + attendance_marks + teacher_assessment + ut_marks + mse_marks + tw_marks + pr_or_marks
    if marks_val == 0:
        marks_val = float(request.form.get("marks", 0) or 0)

    marks_val = min(marks_val, max_total)
    pct_val   = pct(marks_val, max_total)
    g, _, _   = calc_grade(pct_val)
    result_val= "Pass" if pct_val >= 40 else "Fail"
    status_val = request.form.get("status", "draft").strip()

    _exe("""INSERT INTO results(student_name,roll,department,year,semester,subject,
                               marks,total,exam_type,grade,result,published,status,
                               assignment_marks, attendance_marks, ut_marks, mse_marks,
                               teaching_assessment, tw_marks, pr_or_marks)
           VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (student_name, roll, dept, yr, sem_val, subject_val,
         marks_val, max_total, request.form.get("exam_type","Semester Exam"),
         g, result_val, 0, status_val, assignment_marks, attendance_marks, ut_marks, mse_marks,
         teacher_assessment, tw_marks, pr_or_marks))

    # Audit log
    new_row = _qone("SELECT id FROM results WHERE student_name=%s AND subject=%s AND semester=%s ORDER BY id DESC LIMIT 1",
                    (student_name, subject_val, sem_val))
    if new_row:
        write_audit_log(new_row["id"], "created", session.get("user_id"))

    return redirect("/admin_results?success=1")

@results_bp.route("/admin_edit_result", methods=["POST"])
@admin_required
def admin_edit_result():
    rid = request.form.get("result_id","")
    existing = _qone("SELECT department, semester, subject FROM results WHERE id=%s", (rid,))
    dept = existing["department"] if existing else ""
    subject_val = request.form.get("subject", existing["subject"] if existing else "").strip()
    sem_val = request.form.get("semester", existing["semester"] if existing else "").strip()

    # Stage 3: unified component lookup
    from services.results_service import get_components_for_subject, write_audit_log
    sub_info  = get_components_for_subject(subject_val, dept, sem_val)
    total_val = sub_info["max_total"]
    comp_caps = {c["component_name"].lower(): c["max_marks"] for c in sub_info["components"]}

    def _cap(form_key, comp_name, fallback_max):
        val = float(request.form.get(form_key, 0) or 0)
        limit = comp_caps.get(comp_name.lower(), fallback_max)
        return min(val, limit)

    assignment_marks    = _cap("assignment_marks",   "Assignment",        5.0)
    attendance_marks    = _cap("attendance_marks",   "Attendance",        5.0)
    teacher_assessment  = _cap("teacher_assessment", "Teacher Assessment",10.0)
    ut_marks            = _cap("ut_marks",           "Unit Test",         20.0)
    mse_marks           = _cap("mse_marks",          "Mid-Sem Exam",      20.0)
    tw_marks            = _cap("tw_marks",           "Term Work",          0.0)
    pr_or_marks         = _cap("pr_or_marks",        "Practical/Oral",     0.0)

    marks_val = assignment_marks + attendance_marks + teacher_assessment + ut_marks + mse_marks + tw_marks + pr_or_marks
    if marks_val == 0:
        marks_val = float(request.form.get("marks", 0) or 0)

    marks_val = min(marks_val, total_val)
    pct_val = pct(marks_val, total_val)
    g, _, _ = calc_grade(pct_val)
    result_val = "Pass" if pct_val >= 40 else "Fail"
    status_val = request.form.get("status", "draft").strip()

    _exe("""UPDATE results SET semester=%s,subject=%s,marks=%s,total=%s,
           exam_type=%s,grade=%s,result=%s, assignment_marks=%s, attendance_marks=%s, ut_marks=%s, mse_marks=%s,
           teaching_assessment=%s, tw_marks=%s, pr_or_marks=%s, status=%s WHERE id=%s""",
        (sem_val, subject_val,
         marks_val, total_val, request.form.get("exam_type",""),
         g, result_val, assignment_marks, attendance_marks, ut_marks, mse_marks,
         teacher_assessment, tw_marks, pr_or_marks, status_val, rid))

    write_audit_log(rid, "edited", session.get("user_id"))
    return redirect("/admin_results?updated=1")

@results_bp.route("/admin_delete_result", methods=["POST"])
@admin_required
def admin_delete_result():
    rid = request.form.get("result_id", "")
    from services.results_service import write_audit_log
    write_audit_log(rid, "deleted", session.get("user_id"))
    _exe("DELETE FROM results WHERE id=%s", (rid,))
    return redirect("/admin_results?deleted=1")

@results_bp.route("/admin_submit_result", methods=["POST"])
@admin_required
def admin_submit_result():
    """Move result from draft -> submitted."""
    rid = request.form.get("result_id", "")
    existing = _qone("SELECT status FROM results WHERE id=%s", (rid,))
    if not existing or existing["status"] not in ("draft", None, ""):
        flash("Only draft results can be submitted.", "warning")
        return redirect("/admin_results")
    _exe("UPDATE results SET status='submitted' WHERE id=%s", (rid,))
    from services.results_service import write_audit_log
    write_audit_log(rid, "submitted", session.get("user_id"))
    return redirect("/admin_results?submitted=1")


@results_bp.route("/admin_verify_result", methods=["POST"])
@admin_required
def admin_verify_result():
    """Move result from submitted -> verified (HOD verification)."""
    rid = request.form.get("result_id", "")
    existing = _qone("SELECT status FROM results WHERE id=%s", (rid,))
    if not existing or existing["status"] not in ("submitted",):
        flash("Only submitted results can be verified.", "warning")
        return redirect("/admin_results")
    _exe("UPDATE results SET status='verified', approved_by=%s, approved_at=NOW() WHERE id=%s",
         (session.get("username"), rid))
    from services.results_service import write_audit_log
    write_audit_log(rid, "verified", session.get("user_id"))
    return redirect("/admin_results?verified=1")


@results_bp.route("/admin_approve_result", methods=["POST"])
@admin_required
def admin_approve_result():
    """Move result from verified -> approved (Principal/Admin approval). Only approved can be published."""
    rid = request.form.get("result_id", "")
    reason = request.form.get("reason", "").strip()
    existing = _qone("SELECT status FROM results WHERE id=%s", (rid,))
    if not existing or existing["status"] not in ("verified",):
        flash("Only verified results can be approved.", "warning")
        return redirect("/admin_results")
    _exe("UPDATE results SET status='approved', approved_by=%s, approved_at=NOW() WHERE id=%s",
         (session.get("username"), rid))
    from services.results_service import write_audit_log
    write_audit_log(rid, "approved", session.get("user_id"), reason or None)
    return redirect("/admin_results?approved=1")


@results_bp.route("/admin_bulk_submit", methods=["POST"])
@admin_required
def admin_bulk_submit():
    """Bulk move draft results for a semester+dept to submitted."""
    semester = request.form.get("semester", "").strip()
    dept     = request.form.get("dept", "").strip()
    if not semester:
        flash("Semester is required.", "warning")
        return redirect("/admin_results")
    sql = "UPDATE results SET status='submitted' WHERE semester=%s AND (status='draft' OR status IS NULL OR status='')"
    params = [semester]
    if dept:
        sql += " AND department=%s"
        params.append(dept)
    cur = _exe(sql, params)
    flash(f"{cur.rowcount} draft results submitted for review.", "success")
    return redirect("/admin_results")


@results_bp.route("/admin_bulk_verify", methods=["POST"])
@admin_required
def admin_bulk_verify():
    """Bulk verify all submitted results for a semester+dept."""
    semester = request.form.get("semester", "").strip()
    dept     = request.form.get("dept", "").strip()
    if not semester:
        flash("Semester is required.", "warning")
        return redirect("/admin_results")
    sql = "UPDATE results SET status='verified', approved_by=%s, approved_at=NOW() WHERE semester=%s AND status='submitted'"
    params = [session.get("username"), semester]
    if dept:
        sql += " AND department=%s"
        params.append(dept)
    cur = _exe(sql, params)
    flash(f"{cur.rowcount} results verified.", "success")
    return redirect("/admin_results")


@results_bp.route("/admin_bulk_approve", methods=["POST"])
@admin_required
def admin_bulk_approve():
    """Bulk approve all verified results for a semester+dept."""
    semester = request.form.get("semester", "").strip()
    dept     = request.form.get("dept", "").strip()
    if not semester:
        flash("Semester is required.", "warning")
        return redirect("/admin_results")
    sql = "UPDATE results SET status='approved', approved_by=%s, approved_at=NOW() WHERE semester=%s AND status='verified'"
    params = [session.get("username"), semester]
    if dept:
        sql += " AND department=%s"
        params.append(dept)
    cur = _exe(sql, params)
    flash(f"{cur.rowcount} results approved and ready to publish.", "success")
    return redirect("/admin_results")


@results_bp.route("/admin_publish_validate")
@admin_required
def admin_publish_validate():
    """
    Stage 7: Pre-publish validation.
    Returns JSON with a list of issues that would block or warn on publish.
    Called via JS from the publish modal before the admin submits the form.
    """
    semester = request.args.get("semester", "").strip()
    dept     = request.args.get("dept", "").strip()

    if not semester:
        return jsonify({"ok": False, "errors": ["Semester is required."], "warnings": []})

    base_sql = "SELECT * FROM results WHERE semester=%s AND status='approved'"
    params = [semester]
    if dept:
        base_sql += " AND department=%s"
        params.append(dept)

    rows = _qry(base_sql, params) or []
    errors   = []
    warnings = []

    if not rows:
        errors.append(f"No approved results found for {semester}" + (f" / {dept}" if dept else "") + ". Run bulk approve first.")
        return jsonify({"ok": False, "errors": errors, "warnings": warnings})

    zero_marks = [r for r in rows if (r.get("marks") or 0) == 0 and (r.get("exam_type") or "") != "ABSENT"]
    if zero_marks:
        warnings.append(f"{len(zero_marks)} result(s) have 0 marks (not marked absent). Verify before publishing.")

    missing_grade = [r for r in rows if not r.get("grade")]
    if missing_grade:
        errors.append(f"{len(missing_grade)} result(s) are missing a grade. Recalculate before publishing.")

    missing_total = [r for r in rows if (r.get("total") or 0) == 0]
    if missing_total:
        errors.append(f"{len(missing_total)} result(s) have total=0. Fix component data before publishing.")

    no_student_id = [r for r in rows if not r.get("student_id")]
    if no_student_id:
        warnings.append(f"{len(no_student_id)} result(s) have no linked student_id - notification will be skipped for those.")

    return jsonify({
        "ok": len(errors) == 0,
        "ready": len(rows),
        "errors": errors,
        "warnings": warnings
    })


@results_bp.route("/admin_publish_results", methods=["POST"])
@admin_required
def admin_publish_results():
    semester = request.form.get("semester","").strip()
    dept     = request.form.get("dept","").strip()
    from services.results_service import write_audit_log
    if semester:
        sql_skipped = "SELECT COUNT(*) as c FROM results WHERE semester=%s AND status != 'approved'"
        params = [semester]
        if dept:
            sql_skipped += " AND department=%s"
            params.append(dept)
        skipped_count = _qone(sql_skipped, params)["c"] or 0

        sql_pub = "UPDATE results SET published=1, status='published' WHERE semester=%s AND status='approved'"
        params_pub = [semester]
        if dept:
            sql_pub += " AND department=%s"
            params_pub.append(dept)
        cur = _exe(sql_pub, params_pub)
        published_count = cur.rowcount

        # Audit each published row
        published_ids = _qry("SELECT id FROM results WHERE semester=%s AND status='published'"
                             + (" AND department=%s" if dept else ""),
                             ([semester, dept] if dept else [semester]))
        for row in (published_ids or []):
            write_audit_log(row["id"], "published", session.get("user_id"))

        flash(f"Published {published_count} results. {skipped_count} skipped (not yet approved).", "success")

        # Stage 6: Scoped notifications - only affected students, not a broadcast
        if published_count > 0:
            try:
                from services.notification_service import NotificationService
                # Fetch unique student_ids whose results were just published
                notif_sql = (
                    "SELECT DISTINCT student_id FROM results WHERE semester=%s AND status='published'"
                    + (" AND department=%s" if dept else "")
                )
                notif_rows = _qry(notif_sql, ([semester, dept] if dept else [semester])) or []
                for row in notif_rows:
                    if row.get("student_id"):
                        NotificationService.send_notification(
                            row["student_id"], "student",
                            f"Your {semester} semester results have been published. Log in to view your marks."
                        )
            except Exception:
                pass  # Notification failures must never block the publish flow

    return redirect("/admin_results")


@results_bp.route("/admin_unpublish_results", methods=["POST"])
@admin_required
def admin_unpublish_results():
    semester = request.form.get("semester","").strip()
    dept     = request.form.get("dept","").strip()
    reason   = request.form.get("reason", "").strip()
    from services.results_service import write_audit_log
    if semester:
        # Fetch IDs before update for audit
        id_sql = "SELECT id FROM results WHERE semester=%s AND published=1"
        id_params = [semester]
        if dept:
            id_sql += " AND department=%s"
            id_params.append(dept)
        to_unpublish = _qry(id_sql, id_params) or []

        sql = "UPDATE results SET published=0, status='approved' WHERE semester=%s AND published=1"
        params = [semester]
        if dept:
            sql += " AND department=%s"
            params.append(dept)
        _exe(sql, params)

        for row in to_unpublish:
            write_audit_log(row["id"], "unpublished", session.get("user_id"), reason or None)

    return redirect("/admin_results?unpublished_ok=1")

@results_bp.route("/admin_import_results", methods=["POST"])
@admin_required
def admin_import_results():
    f = request.files.get("file")
    if not f:
        flash("No file selected", "error")
        return redirect("/admin_results")
    try:
        _do_import(f)
    except Exception as e:
        import traceback
        from flask import current_app
        current_app.logger.error(f"Excel import error: {traceback.format_exc()}")
        flash(f"Error parsing file: {str(e)}", "error")
    return redirect("/admin_results")

@results_bp.route("/export_results_excel")
@admin_required
def export_results_excel():
    return results_export_excel()

@results_bp.route("/admin_copy_marks", methods=["POST"])
@admin_required
def admin_copy_marks():
    roll = request.form.get("roll", "").strip()
    src_sem = request.form.get("src_sem", "").strip()
    src_sub = request.form.get("src_subject", "").strip()
    src_exam = request.form.get("src_exam_type", "").strip()
    
    dest_sem = request.form.get("dest_sem", "").strip()
    dest_sub = request.form.get("dest_subject", "").strip()
    dest_exam = request.form.get("dest_exam_type", "").strip()
    
    if not (roll and src_sem and src_sub and src_exam and dest_sem and dest_sub and dest_exam):
        flash("All fields are required to copy marks.", "error")
        return redirect("/admin_results")
        
    # Find source marks record
    src_result = _qone("""
        SELECT * FROM results 
        WHERE roll = %s AND semester = %s AND subject = %s AND exam_type = %s
    """, (roll, src_sem, src_sub, src_exam))
    
    if not src_result:
        flash(f"No source marks found for Roll {roll}, Subject '{src_sub}', {src_exam} (Sem {src_sem}).", "error")
        return redirect("/admin_results")
        
    # Upgrade: look up from subject_mark_components
    from services.results_service import get_components_for_subject
    sub_info = get_components_for_subject(dest_sub, src_result["department"], dest_sem)
    dest_total = sub_info["max_total"]
    
    # Calculate grade and result based on destination subject's max_total
    dest_marks = min(src_result["marks"] or 0.0, dest_total)
    pct_val = (dest_marks / dest_total * 100.0) if dest_total > 0 else 0.0
    dest_grade, _, _ = calc_grade(pct_val)
    dest_res = "Pass" if pct_val >= 40 else "Fail"
    
    # Check if a destination result record already exists
    dest_result = _qone("""
        SELECT id FROM results 
        WHERE roll = %s AND semester = %s AND subject = %s AND exam_type = %s
    """, (roll, dest_sem, dest_sub, dest_exam))
    
    if dest_result:
        _exe("""
            UPDATE results 
            SET marks = %s, total = %s, grade = %s, result = %s,
                assignment_marks = %s, attendance_marks = %s, ut_marks = %s, mse_marks = %s,
                teaching_assessment = %s, tw_marks = %s, pr_or_marks = %s, status = 'draft'
            WHERE id = %s
        """, (dest_marks, dest_total, dest_grade, dest_res,
              src_result["assignment_marks"], src_result["attendance_marks"], src_result["ut_marks"], src_result["mse_marks"],
              src_result["teaching_assessment"], src_result["tw_marks"], src_result["pr_or_marks"], dest_result["id"]))
    else:
        _exe("""
            INSERT INTO results (student_name, roll, department, year, semester, subject,
                                marks, total, exam_type, grade, result, published, status,
                                assignment_marks, attendance_marks, ut_marks, mse_marks,
                                teaching_assessment, tw_marks, pr_or_marks)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, 'draft', %s, %s, %s, %s, %s, %s, %s)
        """, (src_result["student_name"], roll, src_result["department"], src_result["year"], dest_sem, dest_sub,
              dest_marks, dest_total, dest_exam, dest_grade, dest_res,
              src_result["assignment_marks"], src_result["attendance_marks"], src_result["ut_marks"], src_result["mse_marks"],
              src_result["teaching_assessment"], src_result["tw_marks"], src_result["pr_or_marks"]))
              
    flash(f"Successfully copied marks for student {src_result['student_name']} (Roll {roll}) to Sem {dest_sem} - {dest_sub} ({dest_exam}).", "success")
    return redirect("/admin_results")


@results_bp.route("/admin_compute_ranks")
@admin_required
def admin_compute_ranks():
    dept = request.args.get("dept", "").strip()
    semester = request.args.get("semester", "").strip()
    subject = request.args.get("subject", "").strip()
    
    sql = "SELECT * FROM results WHERE 1=1"
    params = []
    if dept:
        sql += " AND department = %s"
        params.append(dept)
    if semester:
        sql += " AND semester = %s"
        params.append(semester)
    if subject:
        sql += " AND subject = %s"
        params.append(subject)
        
    rows = _qry(sql, params)
    if not rows:
        return jsonify({"success": True, "ranks": []})
        
    # Group by student (roll, name)
    student_map = {}
    for r in rows:
        key = (r["roll"], r["student_name"])
        if key not in student_map:
            student_map[key] = {
                "roll": r["roll"] or "N/A",
                "name": r["student_name"],
                "obtained": 0.0,
                "total": 0.0,
            }
        student_map[key]["obtained"] += r["marks"] or 0.0
        student_map[key]["total"] += r["total"] or 0.0
        
    # Compute percentage
    student_list = []
    for key, s in student_map.items():
        pct_val = (s["obtained"] / s["total"] * 100.0) if s["total"] > 0 else 0.0
        student_list.append({
            "roll": s["roll"],
            "name": s["name"],
            "obtained": round(s["obtained"], 2),
            "total": round(s["total"], 2),
            "percentage": round(pct_val, 2)
        })
        
    # Sort by percentage descending
    student_list.sort(key=lambda x: x["percentage"], reverse=True)
    
    # Assign ranks with tie handling
    ranked_list = []
    current_rank = 0
    prev_pct = -1
    for i, s in enumerate(student_list):
        if s["percentage"] != prev_pct:
            current_rank = i + 1
            prev_pct = s["percentage"]
        ranked_list.append({
            "rank": current_rank,
            **s
        })

    # Write computed ranks back to DB (rank_in_subject if filtered by subject, else rank_in_class)
    rank_col = "rank_in_subject" if subject else "rank_in_class"
    for item in ranked_list:
        _exe(
            f"UPDATE results SET {rank_col}=%s WHERE roll=%s AND semester=%s"
            + (" AND subject=%s" if subject else ""),
            ([item["rank"], item["roll"], semester] + ([subject] if subject else [])) if semester
            else [item["rank"], item["roll"]] + ([semester, subject] if subject else [])
        )

    return jsonify({"success": True, "ranks": ranked_list})


@results_bp.route("/admin_bulk_delete_results", methods=["POST"])
@admin_required
def admin_bulk_delete_results():
    dept = request.form.get("dept", "").strip()
    semester = request.form.get("semester", "").strip()
    year = request.form.get("year", "").strip()
    q = request.form.get("q", "").strip()
    published = request.form.get("published", "").strip()
    exam_type = request.form.get("exam_type", "").strip()
    
    sql = "DELETE FROM results WHERE 1=1"
    params = []
    if dept:
        sql += " AND department=%s"
        params.append(dept)
    if semester:
        sql += " AND semester=%s"
        params.append(semester)
    if year:
        sql += " AND year=%s"
        params.append(year)
    if q:
        sql += " AND (student_name ILIKE %s OR roll ILIKE %s)"
        params += [f"%{q}%", f"%{q}%"]
    if published != "":
        sql += " AND published=%s"
        params.append(int(published))
    if exam_type:
        sql += " AND exam_type=%s"
        params.append(exam_type)
        
    # Fetch IDs for audit before deletion
    id_sql = "SELECT id FROM results WHERE 1=1"
    id_params = []
    if dept: id_sql += " AND department=%s"; id_params.append(dept)
    if semester: id_sql += " AND semester=%s"; id_params.append(semester)
    if year: id_sql += " AND year=%s"; id_params.append(year)
    if q:
        id_sql += " AND (student_name ILIKE %s OR roll ILIKE %s)"
        id_params += [f"%{q}%", f"%{q}%"]
    if published != "": id_sql += " AND published=%s"; id_params.append(int(published))
    if exam_type: id_sql += " AND exam_type=%s"; id_params.append(exam_type)
    to_delete = _qry(id_sql, id_params) or []

    cur = _exe(sql, params)
    deleted_count = cur.rowcount

    from services.results_service import write_audit_log
    for row in to_delete:
        write_audit_log(row["id"], "deleted", session.get("user_id"), "bulk delete")

    flash(f"Successfully deleted {deleted_count} results matching the current filter selection.", "success")
    return redirect("/admin_results")


# ═══════════════════════════════════════════════════════════
#  HOD — SHARE RESULTS VIA SMS
# ═══════════════════════════════════════════════════════════

@results_bp.route("/share_results_sms", methods=["POST"])
@hod_or_admin_required
def share_results_sms():
    """
    Sends SMS to parents of all students who FAIL or have ATKT
    in the selected semester/department filter.
    """
    dept     = request.form.get("dept", "").strip()
    semester = request.form.get("semester", "").strip()
    status_f = request.form.get("status", "FAIL").strip()   # FAIL / ATKT / both

    # We need build_student_records helper
    from routes.results import build_student_records
    students = build_student_records(
        dept=dept, semester=semester,
        status_filter=status_f if status_f in ("FAIL", "ATKT") else None,
    )

    if not students:
        return jsonify({"success": False, "error": "No students match the filter."}), 400

    sent, failed = 0, 0
    details = []

    for s in students:
        if s["status"] not in ("FAIL", "ATKT"):
            continue

        # Fetch parent contact from students table
        row = _qone(
            "SELECT parent_contact FROM students WHERE roll = %s LIMIT 1",
            (s["roll"],)
        )
        parent_phone = (row or {}).get("parent_contact", "")

        if not parent_phone:
            failed += 1
            details.append({"roll": s["roll"], "name": s["name"], "status": "no_phone"})
            continue

        context = {
            "student_name": s["name"],
            "roll":         s["roll"],
            "semester":     s["semester"],
            "status":       s["status"],
            "percentage":   s["percentage"],
        }

        try:
            from services.sms_service import SMSService
            result = SMSService.send_immediate(parent_phone, "result_alert", context)
            if result.get("success"):
                sent += 1
                details.append({"roll": s["roll"], "name": s["name"], "status": "sent"})
            else:
                failed += 1
                details.append({"roll": s["roll"], "name": s["name"],
                                 "status": "failed", "error": result.get("error", "")})
        except Exception as e:
            failed += 1
            details.append({"roll": s["roll"], "name": s["name"],
                             "status": "error", "error": str(e)})

    return jsonify({
        "success": True,
        "sent":    sent,
        "failed":  failed,
        "total":   sent + failed,
        "details": details
    })

