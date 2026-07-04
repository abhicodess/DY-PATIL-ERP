import os
import io
import re
import math
import logging
import tempfile
from datetime import datetime, date, timedelta
from flask import Blueprint, request, redirect, session, jsonify, render_template, send_file, url_for, flash
from werkzeug.security import generate_password_hash
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

try:
    import pandas as pd
except ImportError:
    pd = None

import psycopg2
import psycopg2.extras

from utils.pg_wrapper import qry, qone, exe, get_db
from blueprints.auth.decorators import login_required
from config import DEPARTMENTS, DESIGNATIONS, YEARS, DIVISIONS
from routes.upload_attendance import is_attendance_upload_allowed, process_attendance_upload
from utils.helpers import hash_password, pct
from models.attendance import ensure_attendance_upload_tables
from models.attendance_model import ensure_attendance_engine_schema
from services.attendance_service import init_attendance_engine
from routes.attendance import handle_import as handle_attendance_import

logger = logging.getLogger("imports")
imports_bp = Blueprint("imports", __name__)

# Re-define helper or alias names used in monolith if any
_re = re

DEFAULT_STUDENT_PASSWORD = os.environ.get("DEFAULT_STUDENT_PASSWORD")
DEFAULT_FACULTY_PASSWORD = os.environ.get("DEFAULT_FACULTY_PASSWORD")

SLOT_COLORS = {
    "Theory":    "#2563EB",
    "Lab":       "#7C3AED",
    "Elective":  "#059669",
    "Minor":     "#D97706",
    "Practical": "#0891B2",
    "Other":     "#6B7280",
}

def assign_color(subject, slot_type):
    if slot_type and slot_type in SLOT_COLORS:
        return SLOT_COLORS[slot_type]
    s = (subject or "").lower()
    if "lab" in s:       return SLOT_COLORS["Lab"]
    if "elective" in s:  return SLOT_COLORS["Elective"]
    if "minor" in s:     return SLOT_COLORS["Minor"]
    return SLOT_COLORS["Theory"]

def _class_meta_from_attendance_text(text, filename=""):
    raw = text or ""
    program_m = re.search(r"Program[:\s]*(.+)", raw, re.I)
    dept_m = re.search(r"Department\s+of\s+(.+)", raw, re.I)
    source = "\n".join(
        part for part in (
            program_m.group(1) if program_m else "",
            dept_m.group(1) if dept_m else "",
            filename or "",
        ) if part
    ) or f"{raw}\n{filename or ''}"
    dept = ""
    div = ""
    year = ""
    div_m = re.search(r"Div[.\s]*([A-Z])", source, re.I)
    if div_m:
        div = div_m.group(1).upper()
    if re.search(r"AIML", source, re.I):
        dept = "AIML"
    elif re.search(r"AI\s*&?\s*DS|AIDS", source, re.I):
        dept = "AIDS"
    elif re.search(r"Comp|Computer|CE", source, re.I):
        dept = "CS"
    elif re.search(r"\bIT\b", source, re.I):
        dept = "IT"
    if re.search(r"\bS\.?\s*Y\.?\b|\bSY\b|S\.Y\.B", source, re.I):
        year = "II"
    elif re.search(r"\bF\.?\s*Y\.?\b|\bFE\b", source, re.I):
        year = "I"
    elif re.search(r"\bT\.?\s*Y\.?\b|\bTE\b", source, re.I):
        year = "III"
    elif re.search(r"\bB\.?\s*E\.?\b|\bBE\b", source, re.I):
        year = "IV"
    return dept, div, year

def grade(marks, total):
    p = pct(marks, total)
    if p >= 75: return "O"
    if p >= 65: return "A"
    if p >= 55: return "B"
    if p >= 45: return "C"
    if p >= 35: return "D"
    return "F"

def normalise_date(raw):
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%m-%Y","%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None

def normalise_status(raw):
    s = str(raw).strip().upper()
    if s in ("P","PRESENT","1","YES"): return "Present"
    if s in ("A","ABSENT","0","NO"):   return "Absent"
    if s in ("LATE"): return "Late"
    if s in ("EXCUSED", "E"): return "Excused"
    if s in ("L","LEAVE","ML","CL", "MEDICAL"): return "Leave"
    return None

def normalize_time(t):
    if not t: return t
    t = str(t).strip()
    parts = re.split(r'\s*-\s*', t)
    result = []
    for part in parts:
        m = re.match(r'^(\d{1,2}):(\d{2})$', part.strip())
        if m:
            result.append(f"{int(m.group(1)):02d}:{m.group(2)}")
        else:
            result.append(part.strip())
    return '-'.join(result)

def resolve_student_id(name, roll=None):
    """Resolve students.id from name; optional roll disambiguates duplicate names."""
    if not name or not str(name).strip():
        return None
    name = str(name).strip()
    r = str(roll).strip() if roll is not None and str(roll).strip() else None
    if r:
        row = qone("SELECT id FROM students WHERE name=%s AND roll=%s", (name, r))
        if row:
            return row["id"]
    row = qone("SELECT id FROM students WHERE name=%s ORDER BY id LIMIT 1", (name,))
    return row["id"] if row else None

def shortage_needed(attended, total):
    if total <= 0:
        return 0, 0
    needed = math.ceil(0.75 * total - attended)
    shortage = max(0, needed)
    can_miss = 0
    if shortage == 0:
        can_miss = max(0, int((attended - 0.75 * total) / 0.75))
    return shortage, can_miss

def today_str():
    return date.today().strftime("%Y-%m-%d")

@imports_bp.route("/import_students_excel", methods=["POST"])
@login_required("admin")
def import_students_excel():
    f = request.files.get("file")
    if not f: return redirect("/students/")
    wb = load_workbook(f, data_only=True)
    if _is_attendance_student_workbook(wb):
        added, updated, skipped = _import_attendance_workbook_students(wb, f.filename)
        return redirect(f"/students/?imported={added}&updated={updated}&skipped={skipped}&format=attendance_xlsx")
    ws = wb.active
    added = skipped = 0
    # Find header row
    hdr_row = 1
    for i,row in enumerate(ws.iter_rows(max_row=10),1):
        vals = [str(c.value or "").lower() for c in row]
        if any("name" in v for v in vals): hdr_row = i; break
    # Map columns
    headers = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def col(keywords):
        for k in keywords:
            for i,h in enumerate(headers):
                if k in h: return i+1
        return None
    cn = col(["name"]); cr = col(["roll"]); cd = col(["dept","department"])
    cy = col(["year"]); ce = col(["email"]); cc = col(["contact","phone","mobile"])
    if not cn: return redirect("/students/")
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        name = str(row[cn-1] or "").strip()
        roll = str(row[cr-1] if cr else "").strip()
        dept = str(row[cd-1] if cd else "").strip()
        year = str(row[cy-1] if cy else "").strip()
        email= str(row[ce-1] if ce else "").strip()
        contact_raw = row[cc-1] if cc and cc <= len(row) else None
        contact = str(int(contact_raw)) if isinstance(contact_raw, float) else str(contact_raw or "").strip()
        if not name or not roll: continue
        prn = roll # Sync PRN with Roll
        try:
            exe("INSERT INTO students(name,roll,prn,department,year,email,contact_number,password) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)",
                (name, roll, prn, dept, year, email, contact, hash_password(_default_student_password())))
            added += 1
        except psycopg2.IntegrityError:
            skipped += 1
    return redirect(f"/students/?imported={added}&skipped={skipped}")



# ── FACULTY ───────────────────────────────────────────────

@imports_bp.route("/import_faculty_excel", methods=["POST"])
@login_required("admin")
def import_faculty_excel():
    f = request.files.get("file")
    if not f: return redirect("/faculty/")
    wb = load_workbook(f, data_only=True); ws = wb.active
    added = skipped = 0
    hdr_row = 1
    for i,row in enumerate(ws.iter_rows(max_row=10),1):
        vals = [str(c.value or "").lower() for c in row]
        if any("name" in v for v in vals): hdr_row = i; break
    headers = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def col(kws):
        for k in kws:
            for i,h in enumerate(headers):
                if k in h: return i+1
        return None
    cn=col(["name"]); cd=col(["dept","department"]); ce=col(["email"]); cp=col(["phone","contact"])
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        if not row or not row[0]: continue
        name  = str(row[cn-1] if cn else "").strip()
        dept  = str(row[cd-1] if cd else "").strip()
        email = str(row[ce-1] if ce else "").strip()
        phone = str(row[cp-1] if cp else "").strip()
        if not name or not email: continue
        try:
            exe("INSERT INTO faculty(name,department,email,phone,password) VALUES(%s,%s,%s,%s,%s)",
                (name, dept, email, phone, hash_password(_default_faculty_password())))
            added += 1
        except psycopg2.IntegrityError:
            skipped += 1
    return redirect(f"/faculty/?imported={added}&skipped={skipped}")


# ── SUBJECTS ──────────────────────────────────────────────

@imports_bp.route("/import_attendance_excel", methods=["POST"])
def import_attendance_excel_v2():
    if not is_attendance_upload_allowed(session):
        return "Unauthorized", 403

    f = request.files.get("file")
    subject = request.form.get("subject", "").strip()
    result = process_attendance_upload(f, session)
    if result["ok"]:
        saved = result["result"]
        return redirect(
            f"/attendance_dashboard?saved={saved['saved']}&skipped={saved['skipped']}"
            f"&students={saved['students']}&format=advanced_upload&batch_id={saved['batch_id']}"
        )

    message = result["error"]
    fallback_errors = (
        "Could not detect attendance table header",
        "Could not detect the total lectures row",
        "No valid subject columns were detected",
    )
    if f and any(token in message for token in fallback_errors):
        try:
            f.seek(0)
        except Exception:
            pass
        added, err = _parse_attendance_excel(f, subject)
        if err == "select_subject":
            return redirect("/attendance?error=select_subject")
        if added:
            return redirect(f"/view_attendance?saved={added}")

    return redirect(f"/attendance?error={str(message)[:160].replace('\r', '').replace('\n', ' ')}")



@imports_bp.route("/import_final_attendance_report", methods=["POST"])
def import_final_attendance_report():
    if not is_attendance_upload_allowed(session):
        return "Unauthorized", 403
    f = request.files.get("file")
    result = process_attendance_upload(f, session)
    if result["ok"]:
        saved = result["result"]
        return redirect(
            f"/attendance_dashboard?saved={saved['saved']}&skipped={saved['skipped']}"
            f"&students={saved['students']}&format=advanced_upload&batch_id={saved['batch_id']}"
        )
    return redirect(f"/attendance?error={str(result['error'])[:160].replace('\r', '').replace('\n', ' ')}")



@imports_bp.route("/import_timetable_v2", methods=["POST"])
@login_required("admin")
def import_timetable_v2():
    f = request.files.get("file")
    if not f: return redirect("/timetable/?error=nofile")
    added = _parse_timetable_excel(f)
    return redirect(f"/timetable/?imported={added}")


@imports_bp.route("/faculty_import_attendance", methods=["POST"])
@login_required("faculty")
def faculty_import_attendance():
    f = request.files.get("file")
    if not f:
        return redirect("/faculty/attendance_portal?tab=import&error=no_file")
    return handle_attendance_import(f, session, request.form.get("subject", "").strip())



@imports_bp.route("/faculty_import_attendance_v2", methods=["POST"])
@login_required("faculty")
def faculty_import_attendance_v2():
    f = request.files.get("file")
    if not f:
        return redirect("/faculty/attendance_portal?tab=import&error=no_file")
    return handle_attendance_import(f, session, request.form.get("subject", "").strip())


# ── NEW: Import Faculty from Faculty_Details_.xlsx ─────────

@imports_bp.route("/import_faculty_details", methods=["POST"])
@login_required("admin")
def import_faculty_details():
    f = request.files.get("file")
    if not f: return redirect("/faculty/")
    added, skipped = _parse_faculty_excel(f)
    return redirect(f"/faculty/?imported={added}&skipped={skipped}")


# ── FIXED: Timetable import (uses exact DY Patil parser) ───

@imports_bp.route("/import_students_v2", methods=["POST"])
@login_required("admin")
def import_students_v2():
    f = request.files.get("file")
    if not f: return redirect("/students/")
    wb = load_workbook(f, data_only=True)
    if _is_attendance_student_workbook(wb):
        added, updated, skipped = _import_attendance_workbook_students(wb, f.filename)
        return redirect(f"/students/?imported={added}&updated={updated}&skipped={skipped}&format=attendance_xlsx")
    ws = wb.active
    added = skipped = 0
    hdr_row = 1
    for i in range(1, min(ws.max_row+1, 10)):
        vals = [str(ws.cell(i,c).value or "").lower() for c in range(1,8)]
        if any("name" in v for v in vals): hdr_row=i; break
    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None
    cn=gcol(["name"]); cr=gcol(["roll"]); cd=gcol(["dept","department"])
    cy=gcol(["year"]); ce=gcol(["email"])
    if not cn: return redirect("/students/?error=bad_format")
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        name  = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        roll  = str(row[cr-1] or "").strip() if cr and cr <= len(row) else ""
        dept  = str(row[cd-1] or "").strip() if cd and cd <= len(row) else ""
        year  = str(row[cy-1] or "").strip() if cy and cy <= len(row) else ""
        email = str(row[ce-1] or "").strip() if ce and ce <= len(row) else ""
        if not name or not roll: continue
        try:
            exe("INSERT INTO students(name,roll,department,year,email,password) VALUES(%s,%s,%s,%s,%s,%s)",
                (name, roll, dept, year, email, hash_password(_default_student_password())))
            added += 1
        except Exception: skipped += 1
    return redirect(f"/students/?imported={added}&skipped={skipped}")



# ── NEW: Import Subjects from Excel ────────────────────────

@imports_bp.route("/import_subjects_v2", methods=["POST"])
@login_required("admin")
def import_subjects_v2():
    f = request.files.get("file")
    if not f: return redirect("/admin/subjects")
    wb = load_workbook(f, data_only=True)
    if _is_attendance_student_workbook(wb):
        added, skipped = _import_attendance_workbook_subjects(wb, f.filename)
        return redirect(f"/admin/subjects?imported={added}&skipped={skipped}&format=attendance_xlsx")
    tt_subjects = _read_subjects_from_wb(wb)
    if tt_subjects:
        added = skipped = 0
        for info in tt_subjects.values():
            name = info["name"]
            if _subject_exists(name, info["code"], info["dept"], info["sem"]):
                skipped += 1
                continue
            try:
                exe(
                    "INSERT INTO subjects(name,department,subject_code,teacher,semester) VALUES(%s,%s,%s,%s,%s)",
                    (name, info["dept"], info["code"], info["teacher"], info["sem"]),
                )
                added += 1
            except Exception:
                skipped += 1
        return redirect(f"/admin/subjects?imported={added}&skipped={skipped}&format=timetable_xlsx")
    ws = wb.active
    added = skipped = 0
    hdr_row = 1
    for i in range(1, min(ws.max_row+1, 10)):
        vals = [str(ws.cell(i,c).value or "").lower() for c in range(1,8)]
        if any("name" in v or "subject" in v for v in vals): hdr_row=i; break
    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None
    cn=gcol(["name","subject"]); cd=gcol(["dept","department"])
    cc=gcol(["code"]); ct=gcol(["teacher"]); cs=gcol(["sem"])
    if not cn: return redirect("/admin/subjects?error=bad_format")
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        name = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        if not name: continue
        dept = str(row[cd-1] or "").strip() if cd and cd <= len(row) else ""
        code = str(row[cc-1] or "").strip() if cc and cc <= len(row) else ""
        tchr = str(row[ct-1] or "").strip() if ct and ct <= len(row) else ""
        sem  = str(row[cs-1] or "I").strip() if cs and cs <= len(row) else "I"
        try:
            exe("INSERT INTO subjects(name,department,subject_code,teacher,semester) VALUES(%s,%s,%s,%s,%s)",
                (name,dept,code,tchr,sem))
            added += 1
        except Exception: skipped += 1
    return redirect(f"/admin/subjects?imported={added}&skipped={skipped}")


# ── NEW: Import Marks from Excel ───────────────────────────

@imports_bp.route("/import_students_smart", methods=["POST"])
@login_required("admin")
def import_students_smart():
    """
    Handles BE-AIDS_students-2025-26_list_.xlsx format:
    Row 1: empty
    Row 2: PRN | Name As Per Marksheet | Mobile No | Email | ...
    Row 3+: data
    Auto-detects dept from filename (BE-AIDS → AIDS, BE-CS → CS etc)
    Auto-detects year from BE/TE/SE/FE prefix
    """
    files = [f for f in request.files.getlist("file") if f and f.filename]
    f = files[0] if files else None
    if not f:
        return redirect("/students/?error=no_file")

    if any((x.filename or "").lower().endswith(".pdf") for x in files):
        added = skipped = updated = 0
        for pdf_file in files:
            if not (pdf_file.filename or "").lower().endswith(".pdf"):
                continue
            rows = _parse_students_from_attendance_pdf(pdf_file)
            for item in rows:
                result = _upsert_pdf_student(item)
                if result == "added":
                    added += 1
                elif result == "updated":
                    updated += 1
                else:
                    skipped += 1
        return redirect(f"/students/?imported={added}&updated={updated}&skipped={skipped}&format=pdf")

    fname   = f.filename.upper()
    dept    = ""
    year    = ""
    # Auto-detect department from filename
    for d in ["AIDS","AIML","CS","IT","CIVIL","MECH","ENTC"]:
        if d in fname: dept = d; break
    # Auto-detect year from filename
    year_map = {"BE":"IV","TE":"III","SE":"II","FE":"I"}
    for k,v in year_map.items():
        if k in fname: year = v; break

    # Allow override from form
    dept = request.form.get("department","").strip() or dept
    year = request.form.get("year","").strip()       or year

    try:
        wb = load_workbook(f, data_only=True)
    except Exception as e:
        return redirect("/students/?error=bad_excel")

    if _is_attendance_student_workbook(wb):
        added, updated, skipped = _import_attendance_workbook_students(wb, f.filename)
        return redirect(f"/students/?imported={added}&updated={updated}&skipped={skipped}&format=attendance_xlsx")

    ws = wb.active
    added = skipped = 0

    # Find header row
    hdr_row = 1
    for ri in range(1, min(ws.max_row+1, 6)):
        vals = [str(ws.cell(ri,c).value or "").lower() for c in range(1, min(ws.max_column+1,6))]
        if any("name" in v or "prn" in v for v in vals):
            hdr_row = ri; break

    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip()
            for c in range(1, ws.max_column+1)]

    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None

    # BE-AIDS format: PRN=roll, Name=name, Email=email
    cn   = gcol(["name as per","name"])
    cr   = gcol(["prn","roll"])
    ce   = gcol(["email"])
    cc   = gcol(["mobile","contact","phone"])
    cdept= gcol(["dept","department","branch"])
    cyear= gcol(["year","class","sem"])

    if not cn:
        return redirect("/students/?error=bad_format")

    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        if not row: continue
        name = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        roll = str(row[cr-1] or "").strip() if cr and cr <= len(row) else ""
        eml  = str(row[ce-1] or "").strip() if ce and ce <= len(row) else ""
        contact_raw = row[cc-1] if cc and cc <= len(row) else None
        contact = str(int(contact_raw)) if isinstance(contact_raw, float) else str(contact_raw or "").strip()

        # Skip header-like rows or empty
        if not name or not roll: continue
        if name.lower() in ("name","name as per marksheet","student name",""):
            continue
        if roll.lower() in ("prn","roll no","roll number","sr.no","sr no"):
            continue

        # Get dept/year from file columns or fall back to auto-detected
        row_dept = str(row[cdept-1] or "").strip() if cdept and cdept <= len(row) else dept
        row_year = str(row[cyear-1] or "").strip() if cyear and cyear <= len(row) else year

        try:
            exe("INSERT INTO students(name,roll,department,year,email,contact_number,password) VALUES(%s,%s,%s,%s,%s,%s,%s)",
                (name.strip(), roll.strip(), row_dept or dept, row_year or year, eml, contact, hash_password(_default_student_password())))
            added += 1
        except Exception:
            skipped += 1

    return redirect(f"/students/?imported={added}&skipped={skipped}")



@imports_bp.route("/import_faculty_smart", methods=["POST"])
@login_required("admin")
def import_faculty_smart():
    """
    Handles Faculty_Details_.xlsx:
    Row 1-2: junk
    Row 3: Sr. No. | Name of Faculty | Department | Contact No | Email Id | Signature
    Row 4+: data with empty rows between
    """
    f = request.files.get("file")
    if not f:
        return redirect("/faculty/?error=no_file")

    try:
        wb = load_workbook(f, data_only=True)
    except Exception as e:
        return redirect("/faculty/?error=bad_excel")
    ws = wb.active
    added = skipped = 0

    # Find header row — look for "Name of Faculty" or "Email"
    hdr_row = 3
    for ri in range(1, min(ws.max_row+1, 8)):
        vals = [str(ws.cell(ri,c).value or "").lower() for c in range(1, 8)]
        if any("name of faculty" in v or ("name" in v and "faculty" in v) for v in vals):
            hdr_row = ri; break
        if any("email" in v for v in vals) and any("name" in v for v in vals):
            hdr_row = ri; break

    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip()
            for c in range(1, ws.max_column+1)]

    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None

    cn  = gcol(["name of faculty","name"])
    cd  = gcol(["department","dept"])
    cp  = gcol(["contact","phone","mobile"])
    ce  = gcol(["email"])

    if not cn:
        return redirect("/faculty/?error=bad_format")

    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        if not row: continue
        name  = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        dept  = str(row[cd-1] or "").strip() if cd and cd <= len(row) else ""
        email = str(row[ce-1] or "").strip() if ce and ce <= len(row) else ""

        # Skip empty rows, header rows, placeholder rows
        if not name: continue
        if name.lower() in ("name of faculty","name","faculty name","sr. no.",""):
            continue
        if name.replace(".","").replace(" ","").isdigit():
            continue

        # Phone: might be stored as float like 8830488239.0
        phone_raw = row[cp-1] if cp and cp <= len(row) else None
        if isinstance(phone_raw, float):
            phone = str(int(phone_raw))
        else:
            phone = str(phone_raw or "").strip()

        try:
            exe("INSERT INTO faculty(name,department,email,phone,password) VALUES(%s,%s,%s,%s,%s)",
                (name, dept.strip(), email, phone, hash_password(_default_faculty_password())))
            added += 1
        except Exception:
            skipped += 1

    return redirect(f"/faculty/?imported={added}&skipped={skipped}")



# ════════════════════════════════════════════════════════════
#  QUICK WIN 1 — PASSWORD RESET (OTP, no email needed)
# ════════════════════════════════════════════════════════════
_reset_otps = {}   # { (role, identifier): (otp, expires_ts) }


@imports_bp.route("/admin_import_results", methods=["POST"])
@login_required("admin")
def admin_import_results():
    f = request.files.get("file")
    if not f: return redirect("/admin_results?error=no_file")
    
    try:
        if pd is None:
            raise ImportError("pandas is not installed")
        df = pd.read_excel(f, header=None)
    except Exception as e:
        return redirect("/admin_results?error=invalid_file_format")
        
    added = 0
    skipped = 0
    
    # 1. Find the header row (contains 'student name', 'prn', or 'roll')
    hdr_row_idx = -1
    for idx, row in df.iterrows():
        row_str = ' '.join([str(x).lower() for x in row.values])
        if 'student name' in row_str or 'prn ' in row_str or 'name' in row_str:
            hdr_row_idx = idx
            break
            
    if hdr_row_idx == -1: return redirect("/admin_results?error=no_student_column_found")
    
    # 2. Extract Columns
    columns = df.iloc[hdr_row_idx].fillna('').astype(str).str.lower().str.strip()
    
    # Subject row is assumed to be the row immediately above the sub-headers
    if hdr_row_idx > 0:
        subjects_row_code = df.iloc[hdr_row_idx - 1].ffill()
        if hdr_row_idx >= 2:
            subjects_row_name = df.iloc[hdr_row_idx - 2].ffill()
            combined = []
            for name_val, code_val in zip(subjects_row_name, subjects_row_code):
                ns = str(name_val).strip().replace('nan','')
                cs = str(code_val).strip().replace('nan','')
                if ns and len(ns) > 2 and ns.lower() not in ['nan','none','blank']:
                    combined.append(ns)
                else:
                    combined.append(cs)
            subjects_row = pd.Series(combined)
        else:
            subjects_row = subjects_row_code
    else:
        subjects_row = pd.Series(['']*len(columns))
    
    # Default semester mapping if passed from form
    sem_val = request.form.get("semester", "I")
    
    # 3. Parse Data
    for idx in range(hdr_row_idx + 1, len(df)):
        row = df.iloc[idx]
        name = ""
        roll = ""
        for c_idx, col_name in enumerate(columns):
            if 'name' in col_name: name = str(row[c_idx]).strip()
            elif 'prn' in col_name or 'roll' in col_name: roll = str(row[c_idx]).strip()
            
        if not name or str(name).lower() in ('nan', 'none', ''): continue
        if str(name).lower().startswith('sr.'): continue # Skip header re-print
        
        # Resolve Student
        roll_row = qone("SELECT roll,department,year FROM students WHERE name=%s OR roll=%s", (name, roll))
        dept_v = roll_row["department"] if roll_row else ""
        yr_v   = roll_row["year"] if roll_row else ""
        
        # Group marks by mapped Subject
        subject_marks = {}
        for c_idx, col_name in enumerate(columns):
            sub_name = str(subjects_row[c_idx]).strip()
            
            # Skip non-subject header fields (like the ones above Sr.No, Name, Roll)
            if not sub_name or str(sub_name).lower() in ['nan','none','sr.no.','student name','prn number','name','roll','dept']:
                continue
                
            val = row[c_idx]
            try:
                val = float(val) if pd.notna(val) else 0
            except:
                val = 0
                
            if sub_name not in subject_marks:
                subject_marks[sub_name] = {'assign':0, 'attend':0, 'ta':0, 'ut':0, 'mse':0, 'tw':0, 'pr_or':0, 'total':0}
                
            if 'assign' in col_name: subject_marks[sub_name]['assign'] = val
            elif 'attend' in col_name: subject_marks[sub_name]['attend'] = val
            elif 'teach' in col_name or 'assess' in col_name: subject_marks[sub_name]['ta'] = val
            elif 'ut' in col_name or 'unit' in col_name: subject_marks[sub_name]['ut'] = val
            elif 'mse' in col_name or 'mid' in col_name: subject_marks[sub_name]['mse'] = val
            elif 'tw' in col_name or 'term' in col_name: subject_marks[sub_name]['tw'] = val
            elif 'pr' in col_name or 'or' in col_name: subject_marks[sub_name]['pr_or'] = val
            elif 'total' in col_name: subject_marks[sub_name]['total'] = val

        # Save Subject Iteratively
        for subj, marks in subject_marks.items():
            marks['assign'] = min(marks['assign'], 5.0)
            marks['attend'] = min(marks['attend'], 5.0)
            marks['ta']     = min(marks['ta'], 10.0)
            marks['ut']     = min(marks['ut'], 20.0)
            marks['mse']    = min(marks['mse'], 20.0)
            
            total_sum = marks['assign'] + marks['attend'] + marks['ta'] + marks['ut'] + marks['mse'] + marks['tw'] + marks['pr_or']
            final_total_obtained = total_sum if total_sum > 0 else marks['total']
            
            if final_total_obtained <= 0 and marks['total'] <= 0:
                continue # Skip if completely empty marks
                
            final_total_obtained = min(final_total_obtained, 60.0)
            max_marks = 60.0
            
            # Evaluate Grade
            g   = grade(final_total_obtained, max_marks)
            res = "Pass" if pct(final_total_obtained, max_marks) >= 35 else "Fail"
            
            try:
                exe("""INSERT INTO results(student_name,roll,department,year,semester,subject,
                                           marks,total,exam_type,grade,result,published,
                                           assignment_marks, attendance_marks, ut_marks, mse_marks,
                                           teaching_assessment, tw_marks, pr_or_marks)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,%s,%s,%s,%s,%s,%s,%s)""",
                    (name, roll, dept_v, yr_v, sem_val, subj, 
                     final_total_obtained, max_marks, "Semester Exam", g, res,
                     marks['assign'], marks['attend'], marks['ut'], marks['mse'],
                     marks['ta'], marks['tw'], marks['pr_or']))
                added += 1
            except Exception as e:
                skipped += 1
                
    return redirect(f"/admin_results?imported={added}&skipped={skipped}")



@imports_bp.route("/admin_import_sem7", methods=["POST"])
@login_required("admin")
def admin_import_sem7():
    """
    Import DY Patil SEM VII Master Sheet (AIDS / IT / COMP tabs).
    """
    f = request.files.get("file")
    if not f:
        return redirect("/admin_results?error=no_file")

    try:
        wb = load_workbook(f, data_only=True)
    except Exception:
        return redirect("/admin_results?error=invalid_file")

    MAX_MARKS = {"AIDS": 510, "IT": 510, "COMP": 485}
    GRADE_TABLE = [(75,"O"),(70,"A+"),(60,"A"),(55,"B+"),(50,"B"),(45,"C"),(0,"F")]

    def sem7_grade(obtained, total):
        if not total:
            return "F"
        p = obtained / total * 100
        for threshold, g in GRADE_TABLE:
            if p >= threshold:
                return g
        return "F"

    added = skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        dept = sheet_name.strip().upper()  # AIDS / IT / COMP

        subj_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        hdr_row  = [str(ws.cell(3, c).value or "").strip() for c in range(1, ws.max_column + 1)]

        total_cols = []  
        last_subj  = "General"
        for i, sv in enumerate(subj_row):
            if sv and str(sv).strip():
                last_subj = str(sv).strip()
            hv = hdr_row[i].upper() if i < len(hdr_row) else ""
            if "TOTAL" in hv and i > 2:
                m = re.search(r"\((\d+)\)", hv)
                max_m = int(m.group(1)) if m else 100
                total_cols.append((i + 1, last_subj, max_m))

        if not total_cols:
            continue

        for row_idx in range(4, ws.max_row + 1):
            sr_cell = ws.cell(row_idx, 1).value
            if sr_cell is None:
                continue
            try:
                int(float(str(sr_cell)))
            except (ValueError, TypeError):
                continue

            name = str(ws.cell(row_idx, 2).value or "").strip()
            prn  = str(ws.cell(row_idx, 3).value or "").strip()
            if not name or name.lower() in ("student name", "name", "nan"):
                continue

            db_student = qone("SELECT roll, department, year FROM students WHERE name=%s OR roll=%s", (name, prn))
            db_dept = db_student["department"] if db_student else dept
            db_year = db_student["year"]       if db_student else "IV"
            db_roll = db_student["roll"]       if db_student else prn

            for col_1, subj_name, max_m in total_cols:
                raw_val = ws.cell(row_idx, col_1).value
                if raw_val is None:
                    continue
                try:
                    marks_val = float(raw_val)
                except (ValueError, TypeError):
                    continue

                g   = sem7_grade(marks_val, max_m)
                res = "Pass" if (marks_val / max_m * 100 >= 35) else "Fail"

                existing = qone("SELECT id FROM results WHERE student_name=%s AND subject=%s AND semester=%s", (name, subj_name, "VII"))
                if existing:
                    skipped += 1
                    continue

                try:
                    exe("""INSERT INTO results
                           (student_name, roll, department, year, semester,
                            subject, marks, total, exam_type, grade, result, published,
                            assignment_marks, attendance_marks, ut_marks, mse_marks,
                            teaching_assessment, tw_marks, pr_or_marks)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,0,0,0,0,0,0,0)""",
                        (name, db_roll, db_dept, db_year, "VII",
                         subj_name, marks_val, float(max_m),
                         "Semester Exam", g, res))
                    added += 1
                except Exception:
                    skipped += 1

    return redirect(f"/admin_results?imported={added}&skipped={skipped}")


# ── Faculty: Enter + view results for their subjects ────────

@imports_bp.route("/import_subjects_from_tt", methods=["POST"])
@login_required("admin")
def import_subjects_from_tt():
    f = request.files.get("file")
    if not f: return redirect("/admin/subjects?error=no_file")
    wb = load_workbook(f, data_only=True)
    if _is_attendance_student_workbook(wb):
        added, skipped = _import_attendance_workbook_subjects(wb, f.filename)
        return redirect(f"/admin/subjects?imported={added}&skipped={skipped}&format=attendance_xlsx")
    all_subjects = _read_subjects_from_wb(wb)
    added = skipped = 0
    for info in all_subjects.values():
        name = info["name"]
        if _subject_exists(name, info["code"], info["dept"], info["sem"]):
            skipped += 1; continue
        exe("INSERT INTO subjects(name,subject_code,teacher,department,semester) VALUES(%s,%s,%s,%s,%s)",
            (name, info["code"], info["teacher"], info["dept"], info["sem"]))
        added += 1
    return redirect(f"/admin/subjects?imported={added}&skipped={skipped}")



@imports_bp.route("/import_final_pdf", methods=["POST"])
@login_required("admin")
def import_final_pdf():
    """Import DY Patil Final Attendance Report PDF (multi-page, all branches/divs)."""

    f = request.files.get("file")
    if not f:
        return redirect("/attendance_dashboard?error=nofile")

    added = skipped = students_found = students_created = 0

    def _clean(s):
        """Normalize cell text: collapse newlines/spaces."""
        return " ".join(str(s or "").replace("\n", " ").split()).strip()

    def _safe_int(v):
        v = str(v or "").strip()
        try:
            return int(float(v)) if v and v.replace(".","").isdigit() else 0
        except Exception:
            return 0

    def _dept_div_from_program(text):
        """Extract department and division from program line like
        'S. Y. B. Tech Comp. Engg. (Div. A)' or 'S. Y. B. Tech AIML (Div. A)'"""
        dept, div = "", ""
        m = _re.search(r'Div[.\s]*([A-Z])', text, _re.I)
        if m:
            div = m.group(1).upper()
        if _re.search(r'AIML', text, _re.I):
            dept = "AIML"
        elif _re.search(r'AI.*DS|AIDS', text, _re.I):
            dept = "AIDS"
        elif _re.search(r'Comp|CE', text, _re.I):
            dept = "CS"
        elif _re.search(r'\bIT\b', text, _re.I):
            dept = "IT"
        return dept, div

    def _ensure_student(name, roll, dept, div):
        """Return student id; create student record if not found."""
        nonlocal students_created
        if not name:
            return None
        # Try by roll first, then name
        row = None
        if roll:
            row = qone("SELECT id FROM students WHERE roll=%s", (roll,))
        if not row:
            row = qone("SELECT id FROM students WHERE name=%s ORDER BY id LIMIT 1", (name,))
        if row:
            return row["id"]
        # Auto-create
        year = "II"  # S.Y. = Second Year
        try:
            new_id = exe(
                "INSERT INTO students(name,roll,department,year,division,password) VALUES(%s,%s,%s,%s,%s,%s)",
                (name, roll or "", dept or "CS", year, div or "", hash_password(_default_student_password()))
            )
            students_created += 1
            return new_id
        except Exception:
            return None

    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name

        with pdfplumber.open(tmp_path) as pdf:
            # Grab full text for program/dept detection
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)

        # Detect dept/div from full_text
        prog_match = _re.search(r'Program[:\s]*(S\.?\s*Y\.?.*)', full_text, _re.I)
        prog_line  = prog_match.group(1) if prog_match else full_text[:200]
        global_dept, global_div = _dept_div_from_program(prog_line)

        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 5:
                        continue

                    # ── Step 1: Locate header row (has "NAME OF STUDENT") ──
                    hdr_idx = None
                    for ri, row in enumerate(table[:6]):
                        row_str = " ".join(_clean(c).upper() for c in row)
                        if "NAME" in row_str and ("ROLL" in row_str or "SR." in row_str or "SR\nNO" in row_str):
                            hdr_idx = ri
                            break
                    if hdr_idx is None:
                        continue

                    # Clean header: collapse multi-line subject names
                    headers = [_clean(c) for c in table[hdr_idx]]
                    name_col = next((i for i,h in enumerate(headers) if "NAME" in h.upper()), None)
                    roll_col = next((i for i,h in enumerate(headers) if "ROLL" in h.upper()), None)
                    if name_col is None:
                        continue

                    # ── Step 2: Build subject column list ──
                    # Subject cols = all columns after roll/name that are not Total/%
                    skip_until = max(name_col, roll_col or 0)
                    subj_cols = []
                    for ci in range(skip_until + 1, len(headers)):
                        h = headers[ci]
                        hu = h.upper().replace(" ", "")
                        if hu in ("", "TOTAL", "%OFATTENDANCE", "PERCENTAGEOFATTENDANCE", "%"):
                            continue  # skip silently (don't break — % can be last col)
                        if h:
                            subj_cols.append((ci, h))

                    if not subj_cols:
                        continue

                    # ── Step 3: Find "TOTAL NO. OF LECTURES CONDUCTED" row ──
                    total_lec_row = None
                    for ri in range(hdr_idx + 1, min(hdr_idx + 7, len(table))):
                        name_cell = _clean(table[ri][name_col] if name_col < len(table[ri]) else "")
                        nc_up = name_cell.upper()
                        if "LECTURE" in nc_up and ("TOTAL" in nc_up or "NO." in nc_up or "CONDUCTED" in nc_up):
                            total_lec_row = ri
                            break

                    total_lecs = {}
                    if total_lec_row is not None:
                        for ci, subj in subj_cols:
                            try:
                                total_lecs[subj] = _safe_int(table[total_lec_row][ci])
                            except Exception:
                                total_lecs[subj] = 0

                    # ── Step 4: Parse student rows ──
                    data_start = (total_lec_row + 1) if total_lec_row is not None else (hdr_idx + 4)
                    for row in table[data_start:]:
                        if not row:
                            continue
                        name_v = _clean(row[name_col]) if name_col < len(row) else ""
                        roll_v = _clean(row[roll_col]) if roll_col and roll_col < len(row) else ""

                        # Skip empty, header-repeat, or footer rows
                        if not name_v:
                            continue
                        nu = name_v.upper()
                        if nu in ("NAME OF STUDENT", "NAME", "FACULTY NAME", "SIGNATURE"):
                            continue
                        if any(nu.startswith(k) for k in ("PROF.", "DR.", "HOD", "FACULTY", "S.Y.", "ACADEMIC")):
                            continue
                        # Skip rows that are all zeros or look like summary rows
                        row_vals = [_clean(row[ci]) if ci < len(row) else "" for ci, _ in subj_cols]
                        if not any(v for v in row_vals):
                            continue

                        students_found += 1
                        s_id = _ensure_student(name_v, roll_v, global_dept, global_div)
                        if not s_id:
                            skipped += 1
                            continue

                        for ci, subj in subj_cols:
                            att_val  = _clean(row[ci]) if ci < len(row) else "0"
                            att_count = _safe_int(att_val)
                            total_v  = total_lecs.get(subj, 0) or att_count

                            try:
                                conn = get_db()
                                conn.execute("""
                                    INSERT INTO attendance_summary
                                        (student_id, student_name, subject, attended, total, division, semester)
                                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                                    ON CONFLICT(student_id, subject) DO UPDATE SET
                                        attended  = excluded.attended,
                                        total     = excluded.total,
                                        division  = excluded.division,
                                        semester  = excluded.semester
                                """, (s_id, name_v, subj, att_count, total_v, global_div, "IV"))
                                conn.commit()
                                conn.close()
                                added += 1
                            except Exception:
                                skipped += 1

        os.unlink(tmp_path)

    except Exception as e:
        return redirect(f"/attendance_dashboard?error=parse_failed&msg={str(e)[:100]}")

    return redirect(
        f"/attendance_dashboard?saved={added}&skipped={skipped}"
        f"&students={students_found}&new_students={students_created}&format=pdf"
    )


# ════════════════════════════════════════════════════════════
#  SMART QR ATTENDANCE API
# ════════════════════════════════════════════════════════════
import uuid


def _parse_timetable_excel(file_obj, simulate=False):
    """
    Parse DY Patil TY-TT Excel. Reads subject lookup table (rows 15+)
    to expand abbreviations: ML(ST) → Machine Learning / Prof.Shakil Tamboli
    """

    _DAYS = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]
    _SKIP = {"short break","lunch break","lunch","break",
             "short            break","lunch         break",""}

    def _lookups(ws):
        sl = {}; fl = {}
        for ri in range(1, ws.max_row+1):
            if str(ws.cell(ri,1).value or "").strip().lower() == "sr.no":
                for ri2 in range(ri+1, ws.max_row+1):
                    if not ws.cell(ri2,1).value: continue
                    sc = str(ws.cell(ri2,3).value or "").strip()
                    fc = str(ws.cell(ri2,5).value or "").strip()
                    # subject abbrev: "Machine Learning(ML)" → ML→Machine Learning
                    m = re.search(r'\(([A-Za-z0-9 ]+(?:\s+Lab)?)\)\s*$', sc)
                    if m:
                        sl[m.group(1).strip()] = sc[:sc.rfind("(")].strip()
                    # faculty abbrev: "Prof.X Yadav(XY)" → XY→Prof.X Yadav
                    for fp in re.split(r'\s*/\s*', fc):
                        fp = fp.strip()
                        fm = re.search(r'\(([A-Z]{1,4})\)\s*$', fp)
                        if fm:
                            fl[fm.group(1)] = fp[:fp.rfind("(")].strip()
                break
        return sl, fl

    def _slot(raw, sl, fl):
        results = []
        for entry in _re.split(r',\s*\n\s*', raw.strip()):
            entry = entry.strip().rstrip(',').strip()
            if not entry: continue
            # room
            room = ""
            rm = _re.search(r'-\s*((?:Lab|Room)[-\s]\w+)', entry)
            if rm: room = rm.group(1).strip()
            # teacher from last matching paren
            teacher = ""
            parens  = _re.findall(r'\(([^)]+)\)', entry)
            for p in reversed(parens):
                p = p.strip()
                if p in fl:                                  teacher = fl[p]; break
                if _re.match(r'^[A-Z]{1,4}$', p):           teacher = p;    break
                if _re.match(r'^[A-Z][a-z]+$',p) and len(p)>3: teacher = p; break
            # clean subject text
            clean = entry
            clean = _re.sub(r'\([A-Z]\d+\)', '', clean)   # batch (A1)
            for p in parens:
                if _re.match(r'^[A-Z]{1,4}$', p.strip()):
                    clean = clean.replace(f'({p})','',1)
            clean = _re.sub(r'-\s*(?:Lab|Room)[-\s\w, ]+', '', clean)
            clean = _re.sub(r'\s+-\s*$','',clean).strip().rstrip(',').strip()
            abbrev  = clean.split('(')[0].strip()
            subject = sl.get(abbrev, abbrev) if abbrev else clean
            if not subject: continue
            stype = ("Lab" if "lab" in subject.lower() else
                     "Elective" if "elective" in subject.lower() else
                     "Minor" if "minor" in subject.lower() else "Theory")
            results.append((subject, teacher, room, stype))
        return results

    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception:
        return 0
    added = 0
    
    # Pre-cache maps to achieve 10,000x insert speedup
    fac_map = {r['name']: r['id'] for r in qry("SELECT id, name FROM faculty")}
    div_map = {r['division']: r for r in qry("SELECT DISTINCT division, department, year FROM students WHERE division IS NOT NULL")}
    sub_map = {r['name']: r['id'] for r in qry("SELECT id, name FROM subjects")}
    inserts = []

    for ws in wb.worksheets:
        sl, fl = _lookups(ws)
        division = ws.title.strip()
        h6 = str(ws.cell(6,1).value or "").strip()
        dept = ("CS" if any(x in h6 for x in ["CSE","Computer"]) else
                "IT" if any(x in h6 for x in ["IT","Information"]) else
                "AIDS" if "Data" in h6 else "AIML" if "Artificial" in h6 else "")
        sem  = "VI" if any(x in h6 for x in ["T.E","TE"]) else ""

        # find time-header row
        time_row = None
        for ri in range(1, 15):
            vals = [str(ws.cell(ri,c).value or "").strip() for c in range(1, ws.max_column+1)]
            if sum(1 for v in vals if _re.match(r'\d+:\d+\s*[-]\s*\d+:\d+', v)) >= 2:
                time_row = ri; break
        if not time_row: continue

        # col → time slot
        time_cols = {}
        for ci in range(2, ws.max_column+1):
            v = str(ws.cell(time_row, ci).value or "").strip()
            if _re.match(r'\d+:\d+\s*[-]\s*\d+:\d+', v):
                time_cols[ci] = v

        # parse day rows
        for ri in range(time_row+1, time_row+8):
            day_v   = str(ws.cell(ri, 2).value or "").strip()
            matched = next((d for d in _DAYS if d.lower()==day_v.lower()), None)
            if not matched: continue
            for ci, ts in time_cols.items():
                raw  = str(ws.cell(ri, ci).value or "").strip()
                norm = " ".join(raw.lower().split())
                if not raw or norm in _SKIP: continue
                if "break" in norm and len(norm) < 25: continue
                for subject, teacher, room, slot_type in _slot(raw, sl, fl):
                    color = assign_color(subject, slot_type)
                    time_s = normalize_time(ts)
                    start_time, end_time = None, None
                    m = _re.match(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", time_s)
                    if m:
                        h1, m1, h2, m2 = map(int, m.groups())
                        if h1 < 7: h1 += 12
                        if h2 < 7: h2 += 12
                        start_time = f"{h1:02d}:{m1:02d}:00"
                        end_time = f"{h2:02d}:{m2:02d}:00"

                    # In-memory mapping to save 3,000+ db queries
                    faculty_id = fac_map.get(teacher, 1)
                    branch = div_map.get(division, {}).get("department", "Unknown")
                    year = div_map.get(division, {}).get("year", "Unknown")
                    subject_id = sub_map.get(subject, None)
                    
                    if not simulate:
                        inserts.append((matched, time_s, start_time, end_time, subject_id, subject, teacher, room, division, sem, slot_type, color, faculty_id, branch, year))
                    added += 1
    if not simulate and inserts:
        # Bulk Insert
        conn = get_db()
        try:
            cur = conn.cursor()
            psycopg2.extras.execute_values(
                cur,
                "INSERT INTO timetable (day,time,start_time,end_time,subject_id,subject,teacher,room,division,semester,slot_type,color,faculty_id,branch,year) VALUES %s",
                inserts
            )
            # if we grabbed raw cursor, make sure it commits via connection
            if hasattr(conn, 'conn'): conn.conn.commit()
            else: conn.commit()
        except Exception as e:
            if hasattr(conn, 'conn'): conn.conn.rollback()
            else: conn.rollback()
            raise e
        finally:
            conn.close()
            
    return added








def _parse_faculty_excel(file_obj):
    """
    Supports Faculty_Details_.xlsx:
    Row 3 = headers (Sr.No | Name of Faculty | Department | Contact No | Email Id | Signature)
    Row 4+ = data, skip empty rows (col1 is None)
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    added = skipped = 0

    # Find header row (has "Name of Faculty" or "name")
    hdr_row = 3
    for i in range(1, min(ws.max_row + 1, 10)):
        vals = [str(ws.cell(i, c).value or "").lower() for c in range(1, 8)]
        if any("name" in v and "faculty" in v for v in vals) or any("email" in v for v in vals):
            hdr_row = i; break

    hdrs = [str(ws.cell(hdr_row, c).value or "").lower().strip()
            for c in range(1, ws.max_column + 1)]
    def gcol(kws):
        for k in kws:
            for i, h in enumerate(hdrs):
                if k in h: return i + 1
        return None

    cn = gcol(["name"])
    ce = gcol(["email"])
    cd = gcol(["dept","department"])
    cp = gcol(["contact","phone"])

    if not cn:
        return 0, 0

    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row: continue
        name  = str(row[cn-1] or "").strip() if cn and cn <= len(row) else ""
        email = str(row[ce-1] or "").strip() if ce and ce <= len(row) else ""
        dept  = str(row[cd-1] or "").strip() if cd and cd <= len(row) else ""
        phone_raw = row[cp-1] if cp and cp <= len(row) else None
        phone = str(int(phone_raw)) if isinstance(phone_raw, float) else str(phone_raw or "").strip()

        if not name or name.lower() in ("name of faculty", "sr. no.", ""):
            continue

        try:
            exe("INSERT INTO faculty(name,department,email,phone,password) VALUES(%s,%s,%s,%s,%s)",
                (name, dept, email, phone, hash_password(_default_faculty_password())))
            added += 1
        except Exception:
            skipped += 1

    return added, skipped


# ── Admin attendance import (uses shared parser) ────

def _parse_attendance_excel(file_obj, subject):
    """
    Supports:
    1. DY Patil Ai_Attendance_sheet.xlsx
       Row2 col5+ = datetime  |  Row4+ col4 = name, col5+ = 0/1
    2. Flat format: Name | Subject | Date | Status
    Returns (added_count, error_msg)
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    added = 0

    # Detect DY Patil: row 2 has datetime objects
    date_map = {}
    for c in range(1, min(ws.max_column + 1, 60)):
        v = ws.cell(2, c).value
        if isinstance(v, (datetime, date)):
            date_map[c] = v.strftime("%Y-%m-%d")

    if len(date_map) >= 2:
        if not subject:
            return 0, "select_subject"
        for ri in range(4, ws.max_row + 1):
            name = str(ws.cell(ri, 4).value or "").strip()
            if not name or name.isdigit() or "name" in name.lower():
                continue
            for col_idx, att_date in date_map.items():
                raw = ws.cell(ri, col_idx).value
                v   = str(raw).strip() if raw is not None else ""
                if   v in ("1","P","p","Present"):       status = "Present"
                elif v in ("0","A","a","Absent"):         status = "Absent"
                elif v in ("L","l","Leave","ML","CL"):    status = "Leave"
                else: continue
                sid = resolve_student_id(name)
                exe(
                    "INSERT INTO attendance(student_id,student_name,subject,date,status) VALUES(%s,%s,%s,%s,%s)",
                    (sid, name, subject, att_date, status),
                )
                added += 1
    else:
        hdr_row = 1
        for i in range(1, min(ws.max_row + 1, 10)):
            vals = [str(ws.cell(i, c).value or "").lower() for c in range(1, 8)]
            if any("name" in v or "student" in v for v in vals):
                hdr_row = i; break
        hdrs = [str(ws.cell(hdr_row, c).value or "").lower().strip()
                for c in range(1, ws.max_column + 1)]
        def gcol(kws):
            for k in kws:
                for i, h in enumerate(hdrs):
                    if k in h: return i + 1
            return None
        cn = gcol(["name","student"]); cs = gcol(["subject"])
        cd = gcol(["date"]);           cst = gcol(["status"])
        if cn:
            for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                name = str(row[cn-1] or "").strip()
                if not name or name.isdigit(): continue
                subj     = str(row[cs-1] if cs else "").strip() or subject
                d_raw    = row[cd-1]  if cd  else None
                st_raw   = row[cst-1] if cst else "Present"
                att_date = normalise_date(d_raw) if d_raw else today_str()
                status   = normalise_status(st_raw) or "Present"
                if name and subj and att_date:
                    sid = resolve_student_id(name)
                    exe(
                        "INSERT INTO attendance(student_id,student_name,subject,date,status) VALUES(?,?,?,?,?)",
                        (sid, name, subj, att_date, status),
                    )
                    added += 1
    return added, None



def _parse_dypatil_final_report(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    def clean(value):
        return " ".join(str(value or "").replace("\n", " ").split()).strip()

    def to_int(value):
        try:
            if value is None or value == "":
                return 0
            return int(float(str(value).strip()))
        except (TypeError, ValueError):
            return 0

    meta_text = "\n".join(
        clean(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, 10) + 1)
        for c in range(1, min(ws.max_column, 8) + 1)
    )
    dept, div, year = _class_meta_from_attendance_text(meta_text, getattr(file_obj, "filename", ""))
    semester = _semester_from_attendance_text(meta_text, "II")

    header_row = None
    headers = []
    for r in range(1, min(ws.max_row, 25) + 1):
        values = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        row_text = " ".join(v.upper() for v in values)
        if "ROLL" in row_text and "NAME" in row_text:
            header_row = r
            headers = values
            break
    if not header_row:
        return 0, 0, 0

    roll_col = next((i + 1 for i, h in enumerate(headers) if "ROLL" in h.upper()), None)
    name_col = next((i + 1 for i, h in enumerate(headers) if "NAME" in h.upper()), None)
    if not roll_col or not name_col:
        return 0, 0, 0

    total_row = None
    for r in range(header_row + 1, min(ws.max_row, header_row + 8) + 1):
        row_text = " ".join(clean(ws.cell(r, c).value).upper() for c in range(1, ws.max_column + 1))
        if "LECTURES" in row_text and ("TOTAL" in row_text or "CONDUCTED" in row_text):
            total_row = r
            break
    if not total_row:
        return 0, 0, 0

    subject_cols = []
    for col, raw_name in enumerate(headers, 1):
        if col <= name_col:
            continue
        name = clean(raw_name)
        upper = name.upper()
        if not name:
            continue
        if upper == "TOTAL" or "ATTEND" in upper or upper == "%":
            break
        total = to_int(ws.cell(total_row, col).value)
        if total <= 0:
            continue
        subject_cols.append((col, name, total))

    added = skipped = students_processed = 0

    def get_or_create_student(name, roll):
        row = qone("SELECT id FROM students WHERE roll=%s", (roll,)) if roll else None
        if not row:
            row = qone(
                "SELECT id FROM students WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) AND department=%s AND division=%s ORDER BY id LIMIT 1",
                (name, dept or "", div or ""),
            )
        if row:
            return row["id"]
        try:
            exe(
                "INSERT INTO students(name,roll,department,year,division,password) VALUES(%s,%s,%s,%s,%s,%s)",
                (name, roll or "", dept or "CS", year or "II", div or "", hash_password(_default_student_password())),
            )
            row = qone("SELECT id FROM students WHERE roll=%s", (roll,)) if roll else None
            if not row:
                row = qone("SELECT id FROM students WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) ORDER BY id DESC LIMIT 1", (name,))
            return row["id"] if row else None
        except Exception:
            return None

    for r in range(total_row + 1, ws.max_row + 1):
        roll = clean(ws.cell(r, roll_col).value).upper()
        name = clean(ws.cell(r, name_col).value)
        row_text = " ".join(clean(ws.cell(r, c).value).upper() for c in range(1, min(ws.max_column, 6) + 1))
        if "FACULTY" in row_text or "SIGNATURE" in row_text:
            break
        if not name or not roll or not re.match(r"^[A-Z]?\d{1,3}[A-Z]?$", roll, re.I):
            continue

        sid = get_or_create_student(name, roll)
        if not sid:
            skipped += len(subject_cols)
            continue
        students_processed += 1

        combined = {}
        for col, subject, total in subject_cols:
            item = combined.setdefault(subject, {"attended": 0, "total": 0})
            item["attended"] += to_int(ws.cell(r, col).value)
            item["total"] += total

        for subject, counts in combined.items():
            try:
                exe(
                    """
                    INSERT INTO attendance_summary
                        (student_id, student_name, subject, attended, total, division, semester)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT(student_id, subject) DO UPDATE SET
                        student_name = excluded.student_name,
                        attended = excluded.attended,
                        total = excluded.total,
                        division = excluded.division,
                        semester = excluded.semester
                    """,
                    (sid, name, subject, counts["attended"], counts["total"], div or "", semester or "II"),
                )
                added += 1
            except Exception:
                skipped += 1

    return added, skipped, students_processed



def _is_dypatil_final_report(file_obj):
    """Peek at the workbook to detect DY Patil Final Attendance Report format."""
    try:
        wb = load_workbook(file_obj, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
        wb.close()
        file_obj.seek(0)
        # Signature: row 7 col C contains 'NAME OF STUDENT' (case-insensitive)
        if len(rows) >= 7:
            r7 = rows[6]  # 0-indexed
            header_text = " ".join(str(v or "").upper() for v in r7)
            if "NAME OF STUDENT" in header_text:
                return True
            # Also check row 4 for 'FINAL ATTENDANCE REPORT'
            r4 = rows[3]
            r4_text = " ".join(str(v or "").upper() for v in r4)
            if "FINAL ATTENDANCE REPORT" in r4_text:
                return True
        return False
    except Exception:
        file_obj.seek(0)
        return False



def _parse_students_from_attendance_pdf(file_obj):
    def clean(value):
        return " ".join(str(value or "").replace("\n", " ").split()).strip()

    rows = []
    seen = set()
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            file_obj.seek(0)
            file_obj.save(tmp.name)
            tmp_path = tmp.name

        with pdfplumber.open(tmp_path) as pdf:
            full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            dept, div, year = _class_meta_from_attendance_text(full_text, getattr(file_obj, "filename", ""))
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    if not table:
                        continue
                    header_idx = None
                    for idx, row in enumerate(table[:8]):
                        row_text = " ".join(clean(cell).upper() for cell in row)
                        if "NAME" in row_text and "ROLL" in row_text:
                            header_idx = idx
                            break
                    if header_idx is None:
                        continue
                    headers = [clean(cell).upper() for cell in table[header_idx]]
                    name_col = next((i for i, h in enumerate(headers) if "NAME" in h), None)
                    roll_col = next((i for i, h in enumerate(headers) if "ROLL" in h), None)
                    if name_col is None or roll_col is None:
                        continue
                    for row in table[header_idx + 1:]:
                        if not row:
                            continue
                        name = clean(row[name_col] if name_col < len(row) else "")
                        roll = clean(row[roll_col] if roll_col < len(row) else "")
                        if not name or not roll:
                            continue
                        upper_name = name.upper()
                        if (
                            "LECTURES CONDUCTED" in upper_name
                            or upper_name in ("NAME", "NAME OF STUDENT")
                            or upper_name.startswith(("PROF.", "DR.", "FACULTY", "HOD"))
                        ):
                            continue
                        if not re.match(r"^[A-Z]?\d{1,3}[A-Z]?$", roll, re.I):
                            continue
                        key = (dept, div, roll.upper(), name.upper())
                        if key in seen:
                            continue
                        seen.add(key)
                        rows.append({
                            "name": name,
                            "roll": roll.upper(),
                            "department": dept or request.form.get("department", "").strip(),
                            "division": div or request.form.get("division", "").strip(),
                            "year": year or request.form.get("year", "").strip() or "II",
                        })
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    return rows



def _parse_students_from_attendance_workbook(wb, filename=""):
    def clean(value):
        return " ".join(str(value or "").replace("\n", " ").split()).strip()

    ws = wb.active
    header_text = "\n".join(
        clean(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, 10) + 1)
        for c in range(1, min(ws.max_column, 8) + 1)
    )
    dept, div, year = _class_meta_from_attendance_text(header_text, filename)

    header_row = None
    headers = []
    for r in range(1, min(ws.max_row, 20) + 1):
        row_headers = [clean(ws.cell(r, c).value).upper() for c in range(1, ws.max_column + 1)]
        row_text = " ".join(row_headers)
        if "ROLL" in row_text and "NAME" in row_text:
            header_row = r
            headers = row_headers
            break
    if not header_row:
        return []

    name_col = next((i + 1 for i, h in enumerate(headers) if "NAME" in h), None)
    roll_col = next((i + 1 for i, h in enumerate(headers) if "ROLL" in h), None)
    if not name_col or not roll_col:
        return []

    rows = []
    seen = set()
    for r in range(header_row + 1, ws.max_row + 1):
        name = clean(ws.cell(r, name_col).value)
        roll = clean(ws.cell(r, roll_col).value).upper()
        if not name or not roll:
            continue
        upper_name = name.upper()
        if "LECTURES CONDUCTED" in upper_name or upper_name in ("NAME", "NAME OF STUDENT"):
            continue
        if not re.match(r"^[A-Z]?\d{1,3}[A-Z]?$", roll, re.I):
            continue
        key = (dept, div, roll, upper_name)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            "name": name,
            "roll": roll,
            "department": dept or request.form.get("department", "").strip(),
            "division": div or request.form.get("division", "").strip(),
            "year": year or request.form.get("year", "").strip() or "II",
        })
    return rows



def _is_attendance_student_workbook(wb):
    ws = wb.active
    text = "\n".join(
        str(ws.cell(r, c).value or "").upper()
        for r in range(1, min(ws.max_row, 12) + 1)
        for c in range(1, min(ws.max_column, 8) + 1)
    )
    return "FINAL ATTENDANCE REPORT" in text and "ROLL" in text and "NAME OF STUDENT" in text



def _semester_from_attendance_text(text, default="II"):
    value = text or ""
    m = re.search(r"Semester\s*[-:]?\s*([IVX]+|\d+)", value, re.I)
    if not m:
        return default
    raw = m.group(1).upper()
    number_map = {"1": "I", "2": "II", "3": "III", "4": "IV", "5": "V", "6": "VI", "7": "VII", "8": "VIII"}
    return number_map.get(raw, raw)



def _parse_subjects_from_attendance_workbook(wb, filename=""):
    def clean(value):
        return " ".join(str(value or "").replace("\n", " ").split()).strip()

    ws = wb.active
    meta_text = "\n".join(
        clean(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, 10) + 1)
        for c in range(1, min(ws.max_column, 8) + 1)
    )
    dept, _, _ = _class_meta_from_attendance_text(meta_text, filename)
    semester = _semester_from_attendance_text(meta_text)

    header_row = None
    headers = []
    for r in range(1, min(ws.max_row, 20) + 1):
        row_headers = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        row_text = " ".join(h.upper() for h in row_headers)
        if "ROLL" in row_text and "NAME" in row_text:
            header_row = r
            headers = row_headers
            break
    if not header_row:
        return []

    code_row = header_row + 1
    type_row = header_row + 2
    subjects = []
    seen = set()
    for col in range(1, ws.max_column + 1):
        name = clean(headers[col - 1])
        name = re.sub(r"\s*-\s*", "-", name)
        code = re.sub(r"\s+", "", clean(ws.cell(code_row, col).value))
        lecture_type = clean(ws.cell(type_row, col).value)
        upper_name = name.upper()
        if not name or upper_name in ("SR. NO.", "ROLL NO.", "NAME OF STUDENT", "TOTAL"):
            continue
        if "ATTEND" in upper_name or upper_name == "%":
            continue
        if not code or not re.search(r"[A-Z]{2,}.*\d", code, re.I):
            continue
        key = (name.lower(), code.upper(), dept, semester)
        if key in seen:
            continue
        seen.add(key)
        subjects.append({
            "name": name,
            "code": code,
            "department": dept,
            "semester": semester,
            "lecture_type": lecture_type,
        })
    return subjects



def _import_attendance_workbook_subjects(wb, filename=""):
    added = skipped = 0
    for subject in _parse_subjects_from_attendance_workbook(wb, filename):
        existing = qone(
            "SELECT id FROM subjects WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) AND COALESCE(subject_code,'')=%s AND department=%s AND semester=%s",
            (subject["name"], subject["code"], subject["department"], subject["semester"]),
        )
        if existing:
            skipped += 1
            continue
        try:
            exe(
                "INSERT INTO subjects(name,department,subject_code,teacher,semester) VALUES(%s,%s,%s,%s,%s)",
                (subject["name"], subject["department"], subject["code"], "", subject["semester"]),
            )
            added += 1
        except Exception:
            skipped += 1
    return added, skipped



def _unique_pdf_roll(raw_roll, dept, div, existing_id=None):
    raw_roll = (raw_roll or "").strip().upper()
    existing = qone("SELECT id,name,department,division FROM students WHERE roll=%s", (raw_roll,))
    if not existing or (existing_id and existing["id"] == existing_id):
        return raw_roll
    if (existing["department"] or "") == (dept or "") and (existing["division"] or "") == (div or ""):
        return raw_roll
    prefixed = "-".join(x for x in (dept, div, raw_roll) if x)
    if not prefixed:
        prefixed = raw_roll
    suffix = 2
    candidate = prefixed
    while qone("SELECT id FROM students WHERE roll=%s", (candidate,)):
        candidate = f"{prefixed}-{suffix}"
        suffix += 1
    return candidate



def _upsert_pdf_student(item):
    name = (item.get("name") or "").strip()
    raw_roll = (item.get("roll") or "").strip().upper()
    dept = (item.get("department") or "").strip()
    div = (item.get("division") or "").strip()
    year = (item.get("year") or "II").strip()
    if not name or not raw_roll:
        return "skipped"

    existing = qone(
        "SELECT id,roll FROM students WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) AND department=%s AND division=%s",
        (name, dept, div),
    )
    if existing:
        exe("UPDATE students SET year=%s WHERE id=%s", (year, existing["id"]))
        return "updated"

    roll = _unique_pdf_roll(raw_roll, dept, div)
    try:
        exe(
            "INSERT INTO students(name,roll,department,year,division,password) VALUES(%s,%s,%s,%s,%s,%s)",
            (name, roll, dept, year, div, hash_password(_default_student_password())),
        )
        return "added"
    except Exception:
        return "skipped"



def _import_attendance_workbook_students(wb, filename=""):
    added = skipped = updated = 0
    rows = _parse_students_from_attendance_workbook(wb, filename)
    for item in rows:
        result = _upsert_pdf_student(item)
        if result == "added":
            added += 1
        elif result == "updated":
            updated += 1
        else:
            skipped += 1
    return added, updated, skipped



def _cell_text(value):
    return " ".join(str(value or "").replace("\n", " ").split()).strip()



def _norm_subject_code(value):
    return re.sub(r"\s+", "", _cell_text(value)).upper()



def _clean_subject_name(value):
    name = _cell_text(value)
    name = re.sub(r"\s*-\s*", " - ", name)
    name = re.sub(r"\s+", " ", name).strip(" -")
    return name



def _clean_faculty_name(value):
    text = _cell_text(value)
    if not text:
        return ""
    parts = [p.strip() for p in re.split(r"\s*/\s*|\s*,\s*", text) if p.strip()]
    for part in parts or [text]:
        cleaned = re.sub(r"\s*\([A-Z0-9]{1,8}\)\s*$", "", part).strip()
        if cleaned:
            return cleaned
    return text



def _dept_from_timetable_text(text):
    upper = re.sub(r"[^A-Z0-9]+", " ", (text or "").upper())
    if "AIDS" in upper or ("AI" in upper and "DS" in upper):
        return "AIDS"
    if "AIML" in upper or ("AI" in upper and "ML" in upper):
        return "AIML"
    if "INFORMATION TECHNOLOGY" in upper or re.search(r"\bIT\b", upper):
        return "IT"
    if "COMPUTER ENGINEERING" in upper or "CSE" in upper or re.search(r"\bCE\b", upper):
        return "CS"
    return ""



def _sem_from_timetable_text(text):
    upper = (text or "").upper()
    m = re.search(r"SEM(?:ESTER)?\s*[:\-]?\s*(I{1,3}|IV|V|VI{0,3}|[1-8])\b", upper)
    if m:
        val = m.group(1)
        return {"1": "I", "2": "II", "3": "III", "4": "IV", "5": "V", "6": "VI", "7": "VII", "8": "VIII"}.get(val, val)
    if re.search(r"\b(FY|F\.Y|FIRST YEAR)\b", upper):
        return "II"
    if re.search(r"\b(SY|S\.Y|SE|S\.E|SECOND YEAR)\b", upper):
        return "IV"
    if re.search(r"\b(TY|T\.Y|TE|T\.E|THIRD YEAR)\b", upper):
        return "VI"
    return ""



def _subject_exists(name, code, dept, sem):
    if code:
        existing = qone(
            "SELECT id FROM subjects WHERE COALESCE(subject_code,'')=%s "
            "AND COALESCE(department,'')=%s AND COALESCE(semester,'')=%s",
            (code or "", dept or "", sem or ""),
        )
        if existing:
            return existing
    return qone(
        "SELECT id FROM subjects WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) "
        "AND COALESCE(subject_code,'')=%s AND COALESCE(department,'')=%s "
        "AND COALESCE(semester,'')=%s",
        (name, code or "", dept or "", sem or ""),
    )



def _read_subjects_from_wb(wb):
    all_subjects = {}
    for ws in wb.worksheets:
        sheet_text = " ".join(
            _cell_text(ws.cell(r, c).value)
            for r in range(1, min(ws.max_row, 12) + 1)
            for c in range(1, min(ws.max_column, 8) + 1)
        )
        meta_text = f"{ws.title} {sheet_text}"
        dept = _dept_from_timetable_text(ws.title) or _dept_from_timetable_text(meta_text)
        sem = _sem_from_timetable_text(meta_text)

        for ri in range(1, ws.max_row + 1):
            row_vals = [_cell_text(ws.cell(ri, c).value) for c in range(1, ws.max_column + 1)]
            row_lowers = [v.lower().replace(" ", "").replace(".", "") for v in row_vals]
            if not any(v in ("srno", "sr", "serialno") for v in row_lowers):
                continue
            if not any("subjectcode" in v or v == "code" for v in row_lowers):
                continue
            if not any("subjectname" in v or v == "subject" for v in row_lowers):
                continue

            def find_col(*needles):
                for idx, value in enumerate(row_lowers, 1):
                    if any(needle in value for needle in needles):
                        return idx
                return None

            code_col = find_col("subjectcode", "code") or 2
            name_col = find_col("subjectname") or find_col("name") or 3
            faculty_col = find_col("facultyname", "faculty", "teacher") or 4
            sr_col = find_col("srno", "serial") or 1
            blank_run = 0

            for ri2 in range(ri + 1, ws.max_row + 1):
                sr_v = _cell_text(ws.cell(ri2, sr_col).value)
                code_v = _norm_subject_code(ws.cell(ri2, code_col).value)
                name_v = _clean_subject_name(ws.cell(ri2, name_col).value)
                fac_v = _clean_faculty_name(ws.cell(ri2, faculty_col).value)
                row_text = " ".join(_cell_text(ws.cell(ri2, c).value) for c in range(1, min(ws.max_column, 8) + 1))
                row_upper = row_text.upper()

                if not any([sr_v, code_v, name_v, fac_v]):
                    blank_run += 1
                    if blank_run >= 3:
                        break
                    continue
                blank_run = 0

                if any(stop in row_upper for stop in ("HOD", "TIME TABLE", "TIME-TABLE", "COORDINATOR")):
                    break
                if not name_v or "subject" in name_v.lower():
                    continue
                if not re.search(r"\d", sr_v) and not code_v:
                    continue
                if name_v.lower() in ("lunch", "break", "recess"):
                    continue

                if code_v:
                    key = (code_v, dept, sem)
                else:
                    key = (name_v.lower(), dept, sem)
                if any(
                    existing["name"].lower() == name_v.lower()
                    and existing["dept"] == dept
                    and existing["sem"] == sem
                    for existing in all_subjects.values()
                ):
                    continue
                if key not in all_subjects:
                    all_subjects[key] = {"name": name_v, "code": code_v, "teacher": fac_v, "dept": dept, "sem": sem}
            break
    return all_subjects



def _default_student_password():
    if DEFAULT_STUDENT_PASSWORD:
        return DEFAULT_STUDENT_PASSWORD
    raise RuntimeError("DEFAULT_STUDENT_PASSWORD must be set when creating students without explicit password")



def _default_faculty_password():
    if DEFAULT_FACULTY_PASSWORD:
        return DEFAULT_FACULTY_PASSWORD
    raise RuntimeError("DEFAULT_FACULTY_PASSWORD must be set when creating faculty without explicit password")



def _ensure_summary_table():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance_summary (
            id          SERIAL PRIMARY KEY,
            student_id  INTEGER,
            student_name TEXT,
            subject     TEXT NOT NULL,
            attended    INTEGER NOT NULL DEFAULT 0,
            total       INTEGER NOT NULL DEFAULT 0,
            division    TEXT DEFAULT '',
            semester    TEXT DEFAULT '',
            UNIQUE(student_id, subject)
        )
    """)
    try:
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_summary_student_id ON attendance_summary(student_id)"
        )
    except Exception:
        pass
    conn.commit(); conn.close()


try:
    _ensure_summary_table()
    ensure_attendance_upload_tables()
    ensure_attendance_engine_schema()
    init_attendance_engine()
except RuntimeError:
    pass


# ─── Unified helpers ──────────────────────────────────────


def _cumulative_result_dict(present, total, source):
    """Build the same dict shape as get_cumulative from aggregated counts."""
    absent = max(0, total - present)
    pct_val = round(present / total * 100, 1) if total else 0
    if pct_val >= 75:
        status = "Good"
    elif pct_val >= 50:
        status = "Average"
    else:
        status = "Low"
    shortage, can_miss = shortage_needed(present, total)
    return {
        "present":    present,
        "attended":   present,
        "absent":     absent,
        "total":      total,
        "percentage": pct_val,
        "status":     status,
        "source":     source,
        "shortage":   shortage,
        "can_miss":   can_miss,
    }


# ── FACULTY TIMETABLE IMPORT ──────────────────────────────────────

def _parse_faculty_timetable_excel(file_obj, faculty_id, faculty_name):
    _DAYS = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]
    _SKIP = {"short break","lunch break","lunch","break",
             "short            break","lunch         break",""}

    def _lookups(ws):
        sl = {}; fl = {}
        for ri in range(1, ws.max_row+1):
            if str(ws.cell(ri,1).value or "").strip().lower() == "sr.no":
                for ri2 in range(ri+1, ws.max_row+1):
                    if not ws.cell(ri2,1).value: continue
                    sc = str(ws.cell(ri2,3).value or "").strip()
                    fc = str(ws.cell(ri2,5).value or "").strip()
                    m = re.search(r'\(([A-Za-z0-9 ]+(?:\s+Lab)?)\)\s*$', sc)
                    if m:
                        sl[m.group(1).strip()] = sc[:sc.rfind("(")].strip()
                    for fp in re.split(r'\s*/\s*', fc):
                        fp = fp.strip()
                        fm = re.search(r'\(([A-Z]{1,4})\)\s*$', fp)
                        if fm:
                            fl[fm.group(1)] = fp[:fp.rfind("(")].strip()
                break
        return sl, fl

    def _slot(raw, sl, fl):
        results = []
        for entry in _re.split(r',\s*\n\s*', raw.strip()):
            entry = entry.strip().rstrip(',').strip()
            if not entry: continue
            room = ""
            rm = _re.search(r'-\s*((?:Lab|Room)[-\s]\w+)', entry)
            if rm: room = rm.group(1).strip()
            teacher = ""
            parens  = _re.findall(r'\(([^)]+)\)', entry)
            for p in reversed(parens):
                p = p.strip()
                if p in fl:                                  teacher = fl[p]; break
                if _re.match(r'^[A-Z]{1,4}$', p):           teacher = p;    break
                if _re.match(r'^[A-Z][a-z]+$',p) and len(p)>3: teacher = p; break
            clean = entry
            clean = _re.sub(r'\([A-Z]\d+\)', '', clean)
            for p in parens:
                if _re.match(r'^[A-Z]{1,4}$', p.strip()):
                    clean = clean.replace(f'({p})','',1)
            clean = _re.sub(r'-\s*(?:Lab|Room)[-\s\w, ]+', '', clean)
            clean = _re.sub(r'\s+-\s*$','',clean).strip().rstrip(',').strip()
            abbrev  = clean.split('(')[0].strip()
            subject = sl.get(abbrev, abbrev) if abbrev else clean
            if not subject: continue
            stype = ("Lab" if "lab" in subject.lower() else
                     "Elective" if "elective" in subject.lower() else
                     "Minor" if "minor" in subject.lower() else "Theory")
            results.append((subject, teacher, room, stype))
        return results

    wb = load_workbook(file_obj, data_only=True)
    added = 0
    inserts = []

    def is_teacher_match(t_name, f_name):
        if not t_name or not f_name:
            return False
        t_clean = t_name.lower().replace("prof.", "").replace("dr.", "").replace("mr.", "").replace("mrs.", "").strip()
        f_clean = f_name.lower().replace("prof.", "").replace("dr.", "").replace("mr.", "").replace("mrs.", "").strip()
        if t_clean in f_clean or f_clean in t_clean:
            return True
        t_words = [w for w in t_clean.split() if len(w) > 2]
        f_words = [w for w in f_clean.split() if len(w) > 2]
        for tw in t_words:
            if tw in f_words:
                return True
        return False

    for ws in wb.worksheets:
        sl, fl = _lookups(ws)
        division = ws.title.strip()
        h6 = str(ws.cell(6,1).value or "").strip()
        sem  = "VI" if any(x in h6 for x in ["T.E","TE"]) else ""

        time_row = None
        for ri in range(1, 15):
            vals = [str(ws.cell(ri,c).value or "").strip() for c in range(1, ws.max_column+1)]
            if sum(1 for v in vals if _re.match(r'\d+:\d+\s*[-]\s*\d+:\d+', v)) >= 2:
                time_row = ri; break
        if not time_row: continue

        time_cols = {}
        for ci in range(2, ws.max_column+1):
            v = str(ws.cell(time_row, ci).value or "").strip()
            if _re.match(r'\d+:\d+\s*[-]\s*\d+:\d+', v):
                time_cols[ci] = v

        for ri in range(time_row+1, time_row+8):
            day_v   = str(ws.cell(ri, 2).value or "").strip()
            matched = next((d for d in _DAYS if d.lower()==day_v.lower()), None)
            if not matched: continue
            for ci, ts in time_cols.items():
                raw  = str(ws.cell(ri, ci).value or "").strip()
                norm = " ".join(raw.lower().split())
                if not raw or norm in _SKIP: continue
                if "break" in norm and len(norm) < 25: continue
                for subject, teacher, room, slot_type in _slot(raw, sl, fl):
                    if is_teacher_match(teacher, faculty_name):
                        time_s = normalize_time(ts)
                        inserts.append((faculty_id, faculty_name, matched, time_s, subject, division, room, slot_type, sem, ''))
                        added += 1

    if inserts:
        conn = get_db()
        try:
            cur = conn.cursor()
            for row in inserts:
                cur.execute("""
                    SELECT id FROM faculty_timetable 
                    WHERE faculty_id = %s AND day = %s AND time_slot = %s AND status != 'rejected'
                """, (row[0], row[2], row[3]))
                if cur.fetchone():
                    continue
                cur.execute("""
                    INSERT INTO faculty_timetable (faculty_id, faculty_name, day, time_slot, subject, division, room, slot_type, semester, academic_year, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'draft')
                """, row)
            if hasattr(conn, 'conn'): conn.conn.commit()
            else: conn.commit()
        except Exception as e:
            if hasattr(conn, 'conn'): conn.conn.rollback()
            else: conn.rollback()
            raise e
        finally:
            conn.close()
            
    return added


@imports_bp.route("/import_faculty_timetable", methods=["POST"])
@login_required("faculty")
def import_faculty_timetable():
    f = request.files.get("file")
    if not f:
        flash("No file selected", "error")
        return redirect("/faculty_timetable")
    
    faculty_id = session.get('faculty_id')
    faculty_name = session.get('name', '')
    if not faculty_id:
        flash("Unauthorized", "error")
        return redirect("/faculty_timetable")
        
    try:
        added = _parse_faculty_timetable_excel(f, faculty_id, faculty_name)
        if added > 0:
            flash(f"Successfully imported {added} slots as drafts!", "success")
        else:
            flash("No matching timetable slots found for your name in the Excel file.", "warning")
    except Exception as e:
        logger.error(f"Error importing faculty timetable: {e}")
        flash(f"Error importing timetable: {str(e)}", "error")
        
    return redirect("/faculty_timetable")


@imports_bp.route("/admin/marks/import-excel", methods=["POST"])
@login_required("admin")
def admin_import_marks_excel():
    f = request.files.get("file")
    if not f:
        return jsonify({"errors": ["No file uploaded"]}), 400
        
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        
        if len(rows) < 4:
            return jsonify({"errors": ["Invalid file structure, too few rows"]}), 400
            
        row2 = rows[1] # subject codes
        row3 = rows[2] # headers
        
        subjects = []
        c = 3 # Col D is 3
        while c < len(row2):
            code = row2[c]
            if code:
                code_str = str(code).strip()
                if code_str:
                    subjects.append({
                        "code": code_str,
                        "col_start": c
                    })
                    c += 6
                    continue
            c += 1
            
        if not subjects:
            return jsonify({"errors": ["No subject codes found in row 2"]}), 400
            
        errors = []
        imported_count = 0
        
        for row_idx, row in enumerate(rows[3:], start=4):
            if not row or len(row) < 3:
                continue
            prn = row[2]
            if prn is None:
                continue
            prn_str = str(prn).strip()
            if not prn_str or prn_str.lower() in ('none', 'null', ''):
                continue
                
            student = qone("SELECT id, name, roll, department, division FROM students WHERE prn = %s", (prn_str,))
            if not student:
                errors.append(f"Row {row_idx}: Student with PRN '{prn_str}' not found")
                continue
                
            for subj in subjects:
                subj_code = subj["code"]
                start_col = subj["col_start"]
                
                if len(row) < start_col + 5:
                    errors.append(f"Row {row_idx}: Missing columns for subject {subj_code}")
                    continue
                    
                try:
                    assignment = float(row[start_col + 0] or 0.0)
                    attendance = float(row[start_col + 1] or 0.0)
                    teaching   = float(row[start_col + 2] or 0.0)
                    ut         = float(row[start_col + 3] or 0.0)
                    mse        = float(row[start_col + 4] or 0.0)
                except (ValueError, TypeError) as e:
                    errors.append(f"Row {row_idx}, subject {subj_code}: Invalid numeric marks: {e}")
                    continue
                    
                sub_master = qone("SELECT subject_name, semester FROM subjects_master WHERE subject_code = %s", (subj_code,))
                if sub_master:
                    subject_name = sub_master["subject_name"]
                    semester_val = sub_master["semester"] or "SEM IV"
                else:
                    sub_tbl = qone("SELECT name, semester FROM subjects WHERE subject_code = %s", (subj_code,))
                    subject_name = sub_tbl["name"] if sub_tbl else subj_code
                    semester_val = sub_tbl["semester"] if (sub_tbl and sub_tbl["semester"]) else "SEM IV"
                    
                from services.results_service import calculate_result
                total_val, grade_val, result_val, passed = calculate_result(assignment, attendance, teaching, ut, mse)
                
                existing = qone("SELECT id FROM marks WHERE student_id = %s AND (subject_code = %s OR subject = %s) AND semester = %s",
                                (student["id"], subj_code, subject_name, semester_val))
                
                fid = session.get("faculty_id") or session.get("user_id") or 1
                
                if existing:
                    exe("""
                        UPDATE marks
                        SET assignment_marks = %s,
                            attendance_marks = %s,
                            teaching_assessment = %s,
                            ut_marks = %s,
                            mse_marks = %s,
                            marks = %s,
                            total = %s,
                            grade = %s,
                            result = %s,
                            prn_number = %s,
                            subject_code = %s
                        WHERE id = %s
                    """, (assignment, attendance, teaching, ut, mse, total_val, 60.0, grade_val, result_val, prn_str, subj_code, existing["id"]))
                else:
                    exe("""
                        INSERT INTO marks (faculty_id, student_id, student_name, roll, subject, department,
                                           marks, total, exam_type, date, assignment_marks, attendance_marks,
                                           teaching_assessment, ut_marks, mse_marks, remarks, prn_number, subject_code, semester)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (fid, student["id"], student["name"], student["roll"], subject_name, student["department"],
                          total_val, 60.0, 'Semester Exam', datetime.now().strftime('%Y-%m-%d'),
                          assignment, attendance, teaching, ut, mse, '', prn_str, subj_code, semester_val))
                          
            imported_count += 1
            
        return jsonify({"imported": imported_count, "errors": errors}), 200
        
    except Exception as e:
        logger.error(f"Error importing marks: {e}", exc_info=True)
        return jsonify({"errors": [f"File import failed: {str(e)}"]}), 500



