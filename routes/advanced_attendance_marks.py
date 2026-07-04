import io
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from flask import Blueprint, render_template, request, session, redirect, jsonify, send_file, url_for

from utils.pg_wrapper import qry, qone, exe
from blueprints.auth.decorators import login_required
from utils.helpers import pct, grade
from config import Config
from routes.student_extra import (
    att_match_student_sql,
    marks_match_student_sql,
    att_match_student_params,
    marks_match_student_params
)

advanced_att_marks_bp = Blueprint("advanced_att_marks", __name__)

DEPARTMENTS = Config.DEPARTMENTS
YEARS = Config.YEARS

def get_student():
    sid = session.get("student_id")
    if not sid: return None
    row = qone("SELECT * FROM students WHERE id=%s", (sid,))
    return dict(row) if row else None

# ── Shared helper: build per-subject attendance+marks dict for a student ──────
def _student_subject_matrix(student_id, student_name):
    ms  = att_match_student_sql()
    mms = marks_match_student_sql()
    sp  = att_match_student_params(student_id, student_name)
    msp = marks_match_student_params(student_id, student_name)

    # ── All attendance rows ───────────────────────────────────────────
    att_rows = qry(
        f"SELECT subject, status FROM attendance WHERE {ms}",  # nosec B608 - ms is a safe formatted match string
        sp,
    )
    # ── All marks rows ────────────────────────────────────────────────
    marks_rows = qry(
        f"SELECT subject, exam_type, marks, total FROM marks WHERE {mms} ORDER BY subject, exam_type",  # nosec B608 - mms is a safe formatted match string
        msp,
    )

    # Build attendance map
    att_map = {}
    for r in att_rows:
        s = (r["subject"] or "").strip()
        if not s:
            continue
        att_map.setdefault(s, {"total": 0, "present": 0, "absent": 0, "late": 0})
        att_map[s]["total"] += 1
        if r["status"] == "Present":
            att_map[s]["present"] += 1
        elif r["status"] == "Absent":
            att_map[s]["absent"] += 1
        else:
            att_map[s]["late"] += 1

    # Build marks map
    marks_map = {}
    for r in marks_rows:
        s = (r["subject"] or "").strip()
        if not s:
            continue
        marks_map.setdefault(s, []).append({
            "exam_type": r["exam_type"] or "—",
            "marks":     round(float(r["marks"] or 0), 1),
            "total":     round(float(r["total"]  or 100), 1),
            "pct":       pct(r["marks"], r["total"]),
            "grade":     grade(r["marks"], r["total"]),
        })

    # Merge all known subjects
    all_subjects = sorted(set(list(att_map.keys()) + list(marks_map.keys())))

    result = []
    for subj in all_subjects:
        a = att_map.get(subj, {})
        m_list = marks_map.get(subj, [])

        att_total   = a.get("total", 0)
        att_present = a.get("present", 0)
        att_absent  = a.get("absent", 0)
        att_late    = a.get("late", 0)
        att_p       = pct(att_present, att_total) if att_total else 0

        if att_p >= 75:
            att_status = "Good"
        elif att_p >= 50:
            att_status = "Average"
        else:
            att_status = "Low" if att_total > 0 else "No Data"

        avg_marks_pct = None
        best_grade    = "N/A"
        if m_list:
            valid = [x["pct"] for x in m_list if x["total"] > 0]
            avg_marks_pct = round(sum(valid) / len(valid), 1) if valid else None
            grade_order   = ["A+", "A", "B+", "B", "C", "F"]
            grades_found  = [x["grade"] for x in m_list]
            best_grade    = next((g for g in grade_order if g in grades_found), "N/A")

        # Composite: 40% attendance weight + 60% marks weight (both out of 100)
        att_weight   = att_p * 0.40
        marks_weight = (avg_marks_pct or 0) * 0.60
        combined     = round(att_weight + marks_weight, 1)

        result.append({
            "subject":       subj,
            "att_total":     att_total,
            "att_present":   att_present,
            "att_absent":    att_absent,
            "att_late":      att_late,
            "att_pct":       att_p,
            "att_status":    att_status,
            "marks_list":    m_list,
            "avg_marks_pct": avg_marks_pct,
            "best_grade":    best_grade,
            "combined_score": combined,
        })

    return result

# ── Shared helper: overall summary for one student ────────────────────────────
def _student_overall_summary(student_id, student_name):
    ms  = att_match_student_sql()
    mms = marks_match_student_sql()
    sp  = att_match_student_params(student_id, student_name)
    msp = marks_match_student_params(student_id, student_name)

    att_rows   = qry(f"SELECT status FROM attendance WHERE {ms}", sp)  # nosec B608 - ms is safe
    marks_rows = qry(f"SELECT marks, total FROM marks WHERE {mms}", msp)  # nosec B608 - mms is safe

    total_att   = len(att_rows)
    present_att = sum(1 for r in att_rows if r["status"] == "Present")
    absent_att  = sum(1 for r in att_rows if r["status"] == "Absent")
    att_pct     = pct(present_att, total_att)

    total_obtained = sum(float(r["marks"] or 0) for r in marks_rows)
    total_max      = sum(float(r["total"]  or 100) for r in marks_rows)
    marks_pct      = round(total_obtained / total_max * 100, 1) if total_max else 0

    return {
        "att_total":       total_att,
        "att_present":     present_att,
        "att_absent":      absent_att,
        "att_pct":         att_pct,
        "marks_obtained":  round(total_obtained, 1),
        "marks_total":     round(total_max, 1),
        "marks_pct":       marks_pct,
        "overall_grade":   grade(total_obtained, total_max) if total_max else "N/A",
        "at_risk":         att_pct < 75 or marks_pct < 40,
    }

# ════════════════════════════════════════════════════════════
#  ADMIN — Advanced Attendance + Marks Dashboard
# ════════════════════════════════════════════════════════════

@advanced_att_marks_bp.route("/admin_att_marks_dashboard")
@login_required("admin")
def admin_att_marks_dashboard():
    dept     = request.args.get("dept",     "").strip()
    division = request.args.get("division", "").strip()
    year     = request.args.get("year",     "").strip()
    semester = request.args.get("semester", "").strip()
    q        = request.args.get("q",        "").strip()
    sort_by  = request.args.get("sort",     "att_pct")   # att_pct | marks_pct | combined | name

    sql    = "SELECT id,name,roll,department,year,division FROM students WHERE 1=1"
    params = []
    if dept:
        sql += " AND department=%s"
        params.append(dept)
    if division:
        sql += " AND division=%s"
        params.append(division)
    if year:
        sql += " AND year=%s"
        params.append(year)
    if q:
        sql += " AND (name LIKE %s OR roll LIKE %s)"
        params += [f"%{q}%", f"%{q}%"]
    sql += " ORDER BY name"
    students_all = qry(sql, params)

    rows = []
    for s in students_all:
        summary = _student_overall_summary(s["id"], s["name"])

        # Optionally filter by semester on marks
        if semester:
            m_rows = qry(
                f"SELECT marks, total FROM marks WHERE {marks_match_student_sql()} AND semester=%s",  # nosec B608 - safe match SQL
                list(marks_match_student_params(s["id"], s["name"])) + [semester],
            )
            if m_rows:
                ob = sum(float(r["marks"] or 0) for r in m_rows)
                mx = sum(float(r["total"]  or 100) for r in m_rows)
                summary["marks_pct"] = round(ob / mx * 100, 1) if mx else 0
                summary["marks_obtained"] = round(ob, 1)
                summary["marks_total"]    = round(mx, 1)

        combined = round(summary["att_pct"] * 0.40 + summary["marks_pct"] * 0.60, 1)
        summary["at_risk"] = summary["att_pct"] < 75 or summary["marks_pct"] < 40

        rows.append({
            "id":         s["id"],
            "name":       s["name"],
            "roll":       s["roll"] or "",
            "dept":       s["department"] or "",
            "year":       s["year"] or "",
            "division":   s["division"] or "",
            "att_pct":    summary["att_pct"],
            "marks_pct":  summary["marks_pct"],
            "combined":   combined,
            "at_risk":    summary["at_risk"],
            "att_total":  summary["att_total"],
            "att_present":summary["att_present"],
            "marks_obtained": summary["marks_obtained"],
            "marks_total":    summary["marks_total"],
            "grade":      summary["overall_grade"],
        })

    # Sorting
    if sort_by == "marks_pct":
        rows.sort(key=lambda x: x["marks_pct"])
    elif sort_by == "combined":
        rows.sort(key=lambda x: x["combined"])
    elif sort_by == "name":
        rows.sort(key=lambda x: x["name"])
    else:
        rows.sort(key=lambda x: x["att_pct"])

    total_students = len(rows)
    at_risk_count  = sum(1 for r in rows if r["at_risk"])
    avg_att_pct    = round(sum(r["att_pct"]   for r in rows) / total_students, 1) if total_students else 0
    avg_marks_pct  = round(sum(r["marks_pct"] for r in rows) / total_students, 1) if total_students else 0
    avg_combined   = round(sum(r["combined"]  for r in rows) / total_students, 1) if total_students else 0

    # Subject-level cross-student averages
    dept_clause = "AND s.department=%s" if dept else ""
    query_subj_att = f"""
        SELECT a.subject,
               COUNT(*) as total,
               SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) as present
        FROM attendance a
        JOIN students s ON a.student_id = s.id
        WHERE 1=1
        {dept_clause}
        GROUP BY a.subject ORDER BY a.subject
    """  # nosec B608 - dept_clause is composed of safe hardcoded literals and parameters
    att_subj_rows = qry(query_subj_att, [dept] if dept else [])

    subj_att_avg = [
        {
            "subject": r["subject"],
            "att_pct": pct(r["present"], r["total"]),
            "total":   r["total"],
        }
        for r in att_subj_rows
        if r["total"] > 0
    ]

    divs_list = [r["division"] for r in qry("SELECT DISTINCT division FROM students WHERE division IS NOT NULL AND division!='' ORDER BY division")]
    sems_list = [r["semester"] for r in qry("SELECT DISTINCT semester FROM marks WHERE semester IS NOT NULL AND semester!='' ORDER BY semester")]

    return render_template("attendance/adv_att_marks_dashboard.html",
        rows=rows,
        total_students=total_students,
        at_risk_count=at_risk_count,
        avg_att_pct=avg_att_pct,
        avg_marks_pct=avg_marks_pct,
        avg_combined=avg_combined,
        subj_att_avg=subj_att_avg,
        f_dept=dept, f_division=division, f_year=year,
        f_semester=semester, f_q=q, f_sort=sort_by,
        DEPARTMENTS=DEPARTMENTS, YEARS=YEARS,
        divs=divs_list, sems=sems_list,
    )

# ── Admin: Deep-dive for one student ──────────────────────────────────────────

@advanced_att_marks_bp.route("/admin_student_att_marks/<int:student_id>")
@login_required("admin")
def admin_student_att_marks(student_id):
    s = qone("SELECT * FROM students WHERE id=%s", (student_id,))
    if not s:
        return redirect(url_for('advanced_att_marks.admin_att_marks_dashboard', error='not_found'))
    s = dict(s)

    matrix  = _student_subject_matrix(s["id"], s["name"])
    summary = _student_overall_summary(s["id"], s["name"])

    # Attendance trend: last 30 days
    ms  = att_match_student_sql()
    sp  = att_match_student_params(s["id"], s["name"])
    trend_rows = qry(
        f"SELECT date, subject, status FROM attendance WHERE {ms} ORDER BY date DESC LIMIT 60",  # nosec B608 - ms is safe
        sp,
    )

    # Recent marks (all exams)
    mms = marks_match_student_sql()
    msp = marks_match_student_params(s["id"], s["name"])
    recent_marks = qry(
        f"SELECT * FROM marks WHERE {mms} ORDER BY date DESC LIMIT 20",  # nosec B608 - mms is safe
        msp,
    )
    recent_marks = [dict(r) for r in recent_marks]
    for m in recent_marks:
        m["pct"]   = pct(m["marks"], m["total"])
        m["grade"] = grade(m["marks"], m["total"])

    danger_subjects = [
        s_row for s_row in matrix
        if s_row["att_pct"] < 75 or (s_row["avg_marks_pct"] is not None and s_row["avg_marks_pct"] < 40)
    ]

    return render_template("attendance/adv_student_att_marks.html",
        student=s,
        matrix=matrix,
        summary=summary,
        trend_rows=[dict(r) for r in trend_rows],
        recent_marks=recent_marks,
        danger_subjects=danger_subjects,
    )

# ── Admin: Export combined Excel ──────────────────────────────────────────────

@advanced_att_marks_bp.route("/export_att_marks_excel")
@login_required("admin")
def export_att_marks_excel():
    dept     = request.args.get("dept",     "").strip()
    division = request.args.get("division", "").strip()
    year     = request.args.get("year",     "").strip()

    sql    = "SELECT id,name,roll,department,year,division FROM students WHERE 1=1"
    params = []
    if dept:
        sql += " AND department=%s"
        params.append(dept)
    if division:
        sql += " AND division=%s"
        params.append(division)
    if year:
        sql += " AND year=%s"
        params.append(year)
    sql += " ORDER BY name"
    students_all = qry(sql, params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance+Marks"
    hdrs = ["#", "Name", "Roll", "Dept", "Year", "Div",
            "Total Att", "Present", "Absent", "Att %",
            "Marks Obtained", "Marks Total", "Marks %", "Grade", "Combined Score", "Status"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")

    STATUS_COLORS = {"At Risk": "FFC7CE", "Good": "C6EFCE", "Average": "FFEB9C"}
    row_num = 2
    for idx, s in enumerate(students_all, 1):
        summary  = _student_overall_summary(s["id"], s["name"])
        combined = round(summary["att_pct"] * 0.40 + summary["marks_pct"] * 0.60, 1)
        status   = "At Risk" if summary["at_risk"] else ("Good" if summary["att_pct"] >= 75 else "Average")
        vals = [
            idx, s["name"], s["roll"] or "", s["department"] or "",
            s["year"] or "", s["division"] or "",
            summary["att_total"], summary["att_present"], summary["att_absent"],
            f"{summary['att_pct']}%",
            summary["marks_obtained"], summary["marks_total"],
            f"{summary['marks_pct']}%",
            summary["overall_grade"],
            combined,
            status,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row_num, c, v)
            if c == 16:   # Status column
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
        row_num += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or "")) + 4, 13
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"att_marks_combined{'_'+dept if dept else ''}{'_div'+division if division else ''}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ════════════════════════════════════════════════════════════
#  FACULTY — Attendance + Marks for their students/subjects
# ════════════════════════════════════════════════════════════

@advanced_att_marks_bp.route("/faculty_att_marks")
@login_required("faculty")
def faculty_att_marks():
    fid      = session["faculty_id"]
    fname    = session["name"]

    # Subjects this faculty member is linked to
    my_subjects = [r["name"] for r in qry(
        "SELECT name FROM subjects WHERE teacher LIKE %s ORDER BY name",
        (f"%{fname.split()[-1]}%",)   # match by last name fragment
    )]
    # Also collect any subjects they entered marks for
    mark_subjects = [r["subject"] for r in qry(
        "SELECT DISTINCT subject FROM marks WHERE faculty_id=%s ORDER BY subject", (fid,)
    )]
    all_subjects = sorted(set(my_subjects + mark_subjects))

    filter_subj = request.args.get("subject", "").strip()
    if filter_subj:
        all_subjects = [s for s in all_subjects if s == filter_subj]

    subject_data = []  # one block per subject
    for subj in all_subjects:
        student_ids_att = qry(
            "SELECT DISTINCT student_id FROM attendance WHERE subject=%s AND student_id IS NOT NULL",
            (subj,),
        )
        student_ids_marks = qry(
            "SELECT DISTINCT student_id FROM marks WHERE subject=%s AND faculty_id=%s AND student_id IS NOT NULL",
            (subj, fid),
        )
        all_ids = set(r["student_id"] for r in student_ids_att) | set(r["student_id"] for r in student_ids_marks)

        if not all_ids:
            continue

        students_in_subj = []
        for sid in sorted(all_ids):
            s = qone("SELECT id,name,roll,department FROM students WHERE id=%s", (sid,))
            if not s:
                continue
            s = dict(s)

            # Attendance for this subject
            att_rows = qry(
                "SELECT status FROM attendance WHERE student_id=%s AND subject=%s",
                (sid, subj),
            )
            att_total   = len(att_rows)
            att_present = sum(1 for r in att_rows if r["status"] == "Present")
            att_pct_val = pct(att_present, att_total)

            # Marks for this subject (from this faculty)
            m_rows = qry(
                "SELECT exam_type, marks, total FROM marks WHERE student_id=%s AND subject=%s AND faculty_id=%s ORDER BY exam_type",
                (sid, subj, fid),
            )
            m_list = [{
                "exam_type": r["exam_type"],
                "marks":     round(float(r["marks"] or 0), 1),
                "total":     round(float(r["total"]  or 100), 1),
                "pct":       pct(r["marks"], r["total"]),
                "grade":     grade(r["marks"], r["total"]),
            } for r in m_rows]

            avg_m = None
            if m_list:
                valid = [x["pct"] for x in m_list if x["total"] > 0]
                avg_m = round(sum(valid) / len(valid), 1) if valid else None

            combined = round(att_pct_val * 0.40 + (avg_m or 0) * 0.60, 1)
            at_risk  = att_pct_val < 75 or (avg_m is not None and avg_m < 40)

            students_in_subj.append({
                "id":         s["id"],
                "name":       s["name"],
                "roll":       s["roll"] or "",
                "dept":       s["department"] or "",
                "att_total":  att_total,
                "att_present":att_present,
                "att_pct":    att_pct_val,
                "marks":      m_list,
                "avg_marks_pct": avg_m,
                "combined":   combined,
                "at_risk":    at_risk,
            })

        students_in_subj.sort(key=lambda x: x["att_pct"])

        pcts = [s["att_pct"] for s in students_in_subj]
        avg_att = round(sum(pcts) / len(pcts), 1) if pcts else 0
        low_count = sum(1 for p in pcts if p < 75)

        subject_data.append({
            "subject":        subj,
            "students":       students_in_subj,
            "student_count":  len(students_in_subj),
            "avg_att_pct":    avg_att,
            "low_count":      low_count,
        })

    return render_template("faculty/faculty_att_marks.html",
        subject_data=subject_data,
        all_subjects=all_subjects,
        f_subject=filter_subj,
        faculty_name=fname,
    )

# ════════════════════════════════════════════════════════════
#  STUDENT — Own combined Attendance + Marks view
# ════════════════════════════════════════════════════════════

@advanced_att_marks_bp.route("/student_att_marks")
@login_required("student")
def student_att_marks():
    student = get_student()
    if not student:
        return redirect("/logout")

    matrix  = _student_subject_matrix(student["id"], student["name"])
    summary = _student_overall_summary(student["id"], student["name"])

    shortage = []
    for row in matrix:
        if row["att_pct"] < 75 and row["att_total"] > 0:
            t = row["att_total"]
            p = row["att_present"]
            needed = max(0, int((0.75 * t - p) / 0.25) + 1)
            shortage.append({
                "subject": row["subject"],
                "att_pct": row["att_pct"],
                "needed":  needed,
            })

    combined_score = round(
        summary["att_pct"] * 0.40 + summary["marks_pct"] * 0.60, 1
    )

    return render_template("student/student_att_marks.html",
        student=student,
        matrix=matrix,
        summary=summary,
        shortage=shortage,
        combined_score=combined_score,
    )

# ════════════════════════════════════════════════════════════
#  API — JSON endpoints for AJAX widgets
# ════════════════════════════════════════════════════════════

@advanced_att_marks_bp.route("/api/student_att_marks_summary/<int:student_id>")
@login_required("admin")
def api_student_att_marks_summary(student_id):
    s = qone("SELECT id,name,roll,department FROM students WHERE id=%s", (student_id,))
    if not s:
        return jsonify({"error": "Student not found"}), 404
    s = dict(s)
    matrix  = _student_subject_matrix(s["id"], s["name"])
    summary = _student_overall_summary(s["id"], s["name"])

    return jsonify({
        "student":  s,
        "summary":  summary,
        "subjects": matrix,
        "combined": round(summary["att_pct"] * 0.40 + summary["marks_pct"] * 0.60, 1),
    })

@advanced_att_marks_bp.route("/api/att_marks_dept_stats")
@login_required("admin")
def api_att_marks_dept_stats():
    stats = []
    for d in DEPARTMENTS:
        students = qry("SELECT id,name FROM students WHERE department=%s", (d,))
        att_pcts   = []
        marks_pcts = []
        for s in students:
            summary = _student_overall_summary(s["id"], s["name"])
            if summary["att_total"] > 0:
                att_pcts.append(summary["att_pct"])
            if summary["marks_total"] > 0:
                marks_pcts.append(summary["marks_pct"])
        stats.append({
            "dept":           d,
            "avg_att_pct":    round(sum(att_pcts)   / len(att_pcts),   1) if att_pcts   else 0,
            "avg_marks_pct":  round(sum(marks_pcts) / len(marks_pcts), 1) if marks_pcts else 0,
            "student_count":  len(students),
        })
    return jsonify(stats)
