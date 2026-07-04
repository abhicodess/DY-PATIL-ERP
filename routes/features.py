"""
routes/features.py — Extended features (Audit, Notifications, Backup)
"""
from flask import Blueprint, render_template, request, redirect, session, jsonify, send_file, current_app
import psycopg2
import os, datetime
from utils.pg_wrapper import exe, qry, qone

features_bp = Blueprint("features", __name__)
DATABASE = "college.db" # match legacy if needed

# ----------------- 1. AUDIT LOGS -----------------
def _ensure_audit_logs_table():
    exe("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id SERIAL PRIMARY KEY,
            action TEXT,
            details TEXT,
            role TEXT,
            user_id INTEGER,
            ip_addr TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)


def _ensure_notifications_table():
    exe("""
        CREATE TABLE IF NOT EXISTS notifications (
            id SERIAL PRIMARY KEY,
            title TEXT,
            message TEXT,
            role_target TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    try:
        exe("ALTER TABLE notifications ADD COLUMN IF NOT EXISTS attachment_path TEXT")
        exe("ALTER TABLE notifications ADD COLUMN IF NOT EXISTS attachment_name TEXT")
    except Exception:
        pass


def log_audit(action, details):
    role = session.get("role", "guest")
    user_id = session.get(f"{role}_id", 0)
    ip_addr = request.remote_addr
    try:
        _ensure_audit_logs_table()
        exe("INSERT INTO audit_logs (action, details, role, user_id, ip_addr) VALUES (%s, %s, %s, %s, %s)", 
            (action, details, role, user_id, ip_addr))
    except Exception as e:
        current_app.logger.error(f"Audit log failed: {e}")

@features_bp.route("/admin_audit_logs")
def admin_audit_logs():
    if session.get("role") != "admin": return redirect("/")
    _ensure_audit_logs_table()
    logs = qry("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 200")
    return render_template("common/audit_logs.html", logs=logs)


# ----------------- 2. NOTIFICATIONS SYSTEM -----------------
@features_bp.route("/api/notifications")
def api_notifications():
    role = session.get("role") or "guest"
    _ensure_notifications_table()
    rows = qry("SELECT * FROM notifications WHERE role_target=%s OR role_target='all' ORDER BY id DESC LIMIT 10", (role,))
    return jsonify([dict(r) for r in rows])

@features_bp.route("/admin_notifications", methods=["GET", "POST"])
def admin_notifications():
    if session.get("role") != "admin": return redirect("/")
    
    if request.method == "POST":
        import uuid
        title = request.form.get("title")
        message = request.form.get("message")
        role_target = request.form.get("role_target", "all")
        
        attachment = request.files.get("attachment")
        attachment_path = None
        attachment_name = None
        if attachment and attachment.filename:
            filename = attachment.filename
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            if ext in ["xlsx", "xls", "pdf", "jpg", "jpeg", "png"]:
                new_filename = f"{uuid.uuid4().hex}.{ext}"
                upload_dir = os.path.join(current_app.root_path, "static", "uploads")
                os.makedirs(upload_dir, exist_ok=True)
                attachment.save(os.path.join(upload_dir, new_filename))
                attachment_path = f"/static/uploads/{new_filename}"
                attachment_name = filename
                
        _ensure_notifications_table()
        exe("INSERT INTO notifications (title, message, role_target, attachment_path, attachment_name) VALUES (%s, %s, %s, %s, %s)", 
            (title, message, role_target, attachment_path, attachment_name))
        log_audit("Publish Notification", f"Title: {title}")
        return redirect("/admin_notifications")
    
    _ensure_notifications_table()
    logs = qry("SELECT * FROM notifications ORDER BY id DESC LIMIT 100")
    return render_template("admin/admin_notifications.html", notifications=logs)

@features_bp.route("/admin/import_notifications_excel", methods=["POST"])
def import_notifications_excel():
    if session.get("role") != "admin": return redirect("/")
    
    f = request.files.get("file")
    if not f or not f.filename:
        return redirect("/admin_notifications?error=no_file")
        
    try:
        from openpyxl import load_workbook
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        
        # Find header row — look for "title" or "message"
        hdr_row = 1
        for ri in range(1, min(ws.max_row+1, 10)):
            vals = [str(ws.cell(ri,c).value or "").lower() for c in range(1, min(ws.max_column+1, 10))]
            if any("title" in v or "message" in v for v in vals):
                hdr_row = ri
                break
                
        headers = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1, ws.max_column+1)]
        
        def gcol(kws):
            for k in kws:
                for i, h in enumerate(headers):
                    if k in h: return i+1
            return None
            
        col_title = gcol(["title", "subject"])
        col_message = gcol(["message", "content", "body"])
        col_audience = gcol(["audience", "role", "target", "role_target"])
        
        if not col_title or not col_message:
            return redirect("/admin_notifications?error=bad_format")
            
        _ensure_notifications_table()
        added = 0
        skipped = 0
        
        for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
            if not row or len(row) < max(col_title or 1, col_message or 1): continue
            title = str(row[col_title-1] or "").strip()
            message = str(row[col_message-1] or "").strip()
            audience = str(row[col_audience-1] or "all").strip().lower() if col_audience and col_audience <= len(row) else "all"
            
            if not title or not message:
                skipped += 1
                continue
                
            if audience not in ["all", "student", "faculty", "admin"]:
                audience = "all"
                
            exe("INSERT INTO notifications (title, message, role_target) VALUES (%s, %s, %s)", 
                (title, message, audience))
            added += 1
            
        log_audit("Import Notifications Excel", f"Imported: {added}, Skipped: {skipped}")
        return redirect(f"/admin_notifications?imported={added}&skipped={skipped}")
    except Exception as e:
        current_app.logger.error(f"Bulk notifications import failed: {e}")
        return redirect("/admin_notifications?error=import_failed")

@features_bp.route("/delete_notification/<int:nid>", methods=["POST"])
def delete_notification(nid):
    if session.get("role") != "admin": return jsonify({"error": "Unauthorized"}), 403
    exe("DELETE FROM notifications WHERE id=%s", (nid,))
    return redirect("/admin_notifications")


# ----------------- 3. BACKUP SYSTEM -----------------
@features_bp.route("/admin_backup")
def admin_backup():
    if session.get("role") != "admin": return redirect("/")
    # Keep legacy SQLite backup path if the file still exists locally.
    if not os.path.exists(DATABASE):
        return "Backup file not available (running on PostgreSQL)."
    log_audit("Database Backup", "Downloaded college.db")
    return send_file(DATABASE, as_attachment=True, download_name=f"backup_college_erp_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.db")


# ----------------- 4. API ENDPOINTS -----------------
@features_bp.route("/api/v1/stats")
def api_v1_stats():
    students_count = qone("SELECT count(*) as c FROM students")["c"]
    faculty_count = qone("SELECT count(*) as c FROM faculty")["c"]
    return jsonify({
        "status": "success",
        "timestamp": datetime.datetime.now().isoformat(),
        "metrics": {
            "total_students": students_count,
            "total_faculty": faculty_count
        }
    })

@features_bp.route("/api/v1/departments")
def api_v1_departments():
    depts = [dict(r) for r in qry("SELECT DISTINCT department FROM students WHERE department IS NOT NULL")]
    return jsonify({"status": "success", "data": depts})


# RESULT PUBLISH CONTROL IS ALREADY IN app.py 
# SMART DASHBOARD WILL BE IMPLEMENTED BY OVERRIDING index.html or admin.html and adding these links.
