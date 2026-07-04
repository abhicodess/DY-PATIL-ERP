import os
import io
import math
import logging
from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, redirect, session, jsonify, send_file, url_for, flash
from werkzeug.security import generate_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from utils.pg_wrapper import qry, qone, exe
from blueprints.auth.decorators import login_required
from config import DEPARTMENTS, DESIGNATIONS, YEARS, DIVISIONS

logger = logging.getLogger("admin_extra")
admin_extra_bp = Blueprint("admin_extra", __name__)

# Authentication decorator
def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Administrator access required.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

# Utility math functions
def pct(a, b):
    return round(a / b * 100) if b > 0 else 0

def grade(marks, total):
    p = pct(marks, total)
    if p >= 75: return "O"
    if p >= 65: return "A"
    if p >= 55: return "B"
    if p >= 45: return "C"
    if p >= 35: return "D"
    return "F"

def shortage_needed(attended, total):
    if total <= 0:
        return 0, 0
    needed = math.ceil(0.75 * total - attended)
    shortage = max(0, needed)
    can_miss = 0
    if shortage == 0:
        can_miss = max(0, int((attended - 0.75 * total) / 0.75))
    return shortage, can_miss

# 1. VIEW ATTENDANCE
@admin_extra_bp.route("/view_attendance")
@login_required(["admin", "faculty"])
def view_attendance():
    from routes.attendance import handle_view_records
    data = handle_view_records(request.args, session)
    
    # DISTINCT subjects
    subjects = [r["subject"] for r in qry("SELECT DISTINCT subject FROM attendance ORDER BY subject")]
    # DISTINCT divisions
    divisions = [r["division"] for r in qry("SELECT DISTINCT division FROM students WHERE division != '' ORDER BY division")]
    
    return render_template(
        "attendance/view_attendance.html",
        records=data.get("records", []),
        current_page=data.get("current_page", 1),
        total_pages=data.get("total_pages", 1),
        total_count=data.get("total_count", 0),
        stats=data.get("stats", {}),
        analytics=data.get("analytics", {}),
        subjects=subjects,
        divs=divisions
    )

# 2. ATTENDANCE DASHBOARD
@admin_extra_bp.route("/attendance_dashboard")
@login_required("admin")
def attendance_dashboard():
    # Initialize all template variables with safe defaults to prevent 500 errors
    all_students = []; defaulters = []; critical = []
    total_students = 0; avg_pct = 0; good_count = 0; avg_count = 0; low_count = 0
    at_risk_count = 0; critical_count = 0; safe_count = 0
    divisions = []; dept_stats = []; div_stats = []
    weekly_trend = []; status_dist = {}
    dept_labels = []; dept_pcts = []; div_labels = []; div_pcts = []
    trend_labels = []; trend_pcts = []; status_labels = []; status_counts = []
    top5_names = []; top5_pcts = []; topper = None

    timeframe = request.args.get("timeframe", "cumulative").strip().lower()

    # Fetch Students & Cumulative Data
    students_raw = qry("SELECT id, name, roll, division, department FROM students ORDER BY name")
    
    # Bulk cumulative calculation helper
    summary_map = {}
    if timeframe == "cumulative":
        try:
            if students_raw:
                ids = [int(s["id"]) for s in students_raw]
                ph = ",".join(["%s"] * len(ids))
                params = tuple(ids)
                for r in qry(f"SELECT student_id, COALESCE(SUM(attended),0) AS sm_att, COALESCE(SUM(total),0) AS sm_tot FROM attendance_summary WHERE student_id IN ({ph}) GROUP BY student_id", params):  # nosec B608 - safe IN placeholder string
                    summary_map[int(r["student_id"])] = (int(r["sm_att"]), int(r["sm_tot"]))
        except Exception as e:
            logger.warning(f"Error fetching attendance summary: {e}")

    daily_map = {}
    try:
        if students_raw:
            ids = [int(s["id"]) for s in students_raw]
            ph = ",".join(["%s"] * len(ids))
            
            # Base query
            sql_daily = f"SELECT student_id, COUNT(*) AS d_tot, SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) AS d_att FROM attendance WHERE student_id IN ({ph})"
            params_daily = list(ids)
            
            if timeframe == "daily":
                sql_daily += " AND date = CURRENT_DATE"
            elif timeframe == "weekly":
                sql_daily += " AND date >= CURRENT_DATE - INTERVAL '6 days'"
            elif timeframe == "monthly":
                sql_daily += " AND date >= CURRENT_DATE - INTERVAL '29 days'"
                
            sql_daily += " GROUP BY student_id"
            
            for r in qry(sql_daily, tuple(params_daily)):
                daily_map[int(r["student_id"])] = (int(r["d_att"]) if r["d_att"] is not None else 0, int(r["d_tot"]))
    except Exception as e:
        logger.warning(f"Error fetching daily attendance: {e}")

    for s in students_raw:
        sid = int(s["id"])
        att, tot = 0, 0
        if timeframe == "cumulative" and sid in summary_map and summary_map[sid][1] > 0:
            att, tot = summary_map[sid]
        elif sid in daily_map and daily_map[sid][1] > 0:
            att, tot = daily_map[sid]
        
        if tot == 0: continue
        
        pct_val = round(att / tot * 100, 1)
        status_val = "Good" if pct_val >= 75 else ("Average" if pct_val >= 50 else "Low")
        shortage_v, can_miss_v = shortage_needed(att, tot)
        
        all_students.append({
            "id": s["id"], "name": s["name"], "roll": s["roll"] or "",
            "division": s["division"] or "", "dept": s["department"] or "",
            "pct": pct_val, "status": status_val,
            "attended": att, "total": tot,
            "shortage": shortage_v, "can_miss": can_miss_v
        })

    if all_students:
        all_students.sort(key=lambda x: -x["pct"])
        for i, s in enumerate(all_students): s["rank"] = i + 1

        total_students = len(all_students)
        avg_pct = round(sum(s["pct"] for s in all_students) / total_students, 1)
        
        defaulters = [s for s in all_students if s["pct"] < 75]
        critical = [s for s in all_students if s["pct"] < 40]
        at_risk_count = len(defaulters)
        critical_count = len(critical)
        safe_count = total_students - at_risk_count
        
        good_count = sum(1 for s in all_students if s["pct"] >= 75)
        avg_count = sum(1 for s in all_students if 50 <= s["pct"] < 75)
        low_count = sum(1 for s in all_students if s["pct"] < 50)
        
        divisions = sorted(list(set(s["division"] for s in all_students if s["division"])))

        # Stats by Dept
        d_map = {}
        for s in all_students:
            d = s["dept"] or "Unknown"
            d_map.setdefault(d, {"cnt": 0, "sum": 0.0, "g": 0, "l": 0})
            d_map[d]["cnt"] += 1; d_map[d]["sum"] += s["pct"]
            if s["pct"] >= 75: d_map[d]["g"] += 1
            elif s["pct"] < 50: d_map[d]["l"] += 1
        dept_stats = sorted([{"dept": k, "count": v["cnt"], "avg_pct": round(v["sum"]/v["cnt"], 1), "good": v["g"], "low": v["l"]} for k, v in d_map.items()], key=lambda x: x["dept"])

        # Stats by Div
        v_map = {}
        for s in all_students:
            dv = s["division"] or "N/A"
            v_map.setdefault(dv, {"cnt": 0, "sum": 0.0})
            v_map[dv]["cnt"] += 1; v_map[dv]["sum"] += s["pct"]
        div_stats = sorted([{"division": k, "count": v["cnt"], "avg_pct": round(v["sum"]/v["cnt"], 1)} for k, v in v_map.items()], key=lambda x: x["division"])

        dept_labels = [d["dept"] for d in dept_stats]
        dept_pcts = [d["avg_pct"] for d in dept_stats]
        div_labels = [d["division"] for d in div_stats]
        div_pcts = [d["avg_pct"] for d in div_stats]
        
        top5 = all_students[:5]
        top5_names = [s["name"] for s in top5]
        top5_pcts = [s["pct"] for s in top5]
        topper = all_students[0]

    # Weekly Trend
    week_dates = [(date.today() - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6, -1, -1)]
    weekly_trend = []
    for d in week_dates:
        row = qone("SELECT COUNT(*) as tot, SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as p FROM attendance WHERE date::text = %s", (d,))
        tot = row["tot"] if row else 0
        p = row["p"] if row and row["p"] is not None else 0
        weekly_trend.append({"date": d[5:], "pct": pct(p, tot)})
    
    trend_labels = [t["date"] for t in weekly_trend]
    trend_pcts = [t["pct"] for t in weekly_trend]

    # Status Distribution
    for st in ["Present", "Absent", "Late", "Medical", "Leave"]:
        row = qone("SELECT COUNT(*) as c FROM attendance WHERE status=%s", (st,))
        status_dist[st] = row["c"] if row else 0
    status_labels = list(status_dist.keys())
    status_counts = [status_dist[k] for k in status_labels]

    return render_template("attendance/attendance_dashboard.html",
        all_students=all_students, defaulters=defaulters, critical=critical,
        total_students=total_students, avg_pct=avg_pct, good_count=good_count,
        avg_count=avg_count, low_count=low_count, at_risk_count=at_risk_count,
        critical_count=critical_count, safe_count=safe_count, divisions=divisions,
        dept_stats=dept_stats, div_stats=div_stats, weekly_trend=weekly_trend,
        status_dist=status_dist, dept_labels=dept_labels, dept_pcts=dept_pcts,
        div_labels=div_labels, div_pcts=div_pcts, trend_labels=trend_labels,
        trend_pcts=trend_pcts, status_labels=status_labels, status_counts=status_counts,
        top5_names=top5_names, top5_pcts=top5_pcts, topper=topper, timeframe=timeframe
    )

# 3. ADMIN RESULTS (Moved to routes/results.py)

# 4. ANALYTICS
@admin_extra_bp.route("/analytics")
@login_required("admin")
def analytics():
    # Monthly attendance trend (last 6 months)
    monthly = []
    today_d = date.today()
    for i in range(5, -1, -1):
        month_offset = (today_d.month - 1 - i) % 12
        year_offset = today_d.year + (today_d.month - 1 - i) // 12
        mn = month_offset + 1
        yn = year_offset
        label = f"{yn}-{mn:02d}"
        
        row = qone("SELECT COUNT(*) as tot, SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as p FROM attendance WHERE date::text LIKE %s", (f"{label}%",))
        tot = row["tot"] if row else 0
        p = row["p"] if row and row["p"] is not None else 0
        monthly.append({"label": label, "total": tot, "present": p, "pct": pct(p, tot)})

    # Bottom students by attendance (min 3 classes)
    students_all = qry("SELECT id, name, roll, department FROM students")
    student_att = []
    for s in students_all:
        rows = qry("SELECT status FROM attendance WHERE student_id = %s", (s["id"],))
        if len(rows) < 3: continue
        pre = sum(1 for r in rows if r["status"] == "Present")
        student_att.append({
            "name": s["name"], "roll": s["roll"], "dept": s["department"],
            "pct": pct(pre, len(rows)), "total": len(rows)
        })
    student_att.sort(key=lambda x: -x["pct"])
    bottom5 = sorted(student_att, key=lambda x: x["pct"])[:5]

    # Subject-wise average marks
    subj_rows = qry("SELECT subject, ROUND(AVG(marks*100.0/total)::numeric, 1) as avg FROM marks GROUP BY subject ORDER BY avg DESC LIMIT 10")
    subj_marks = [dict(r) for r in subj_rows]

    # Exam type breakdown
    exam_rows = qry("SELECT exam_type, ROUND(AVG(marks*100.0/total)::numeric, 1) as avg FROM marks GROUP BY exam_type")

    # Dept attendance
    dept_att = []
    for d in DEPARTMENTS:
        row = qone("SELECT COUNT(*) as tot, SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) as p FROM attendance a JOIN students s ON a.student_id=s.id WHERE s.department=%s", (d,))
        tot = row["tot"] if row else 0
        p = row["p"] if row and row["p"] is not None else 0
        dept_att.append({"dept": d, "pct": pct(p, tot), "total": tot})

    # Grade distribution
    grade_dist = {"A+": 0, "A": 0, "B+": 0, "B": 0, "C": 0, "F": 0}
    all_marks = qry("SELECT marks, total FROM marks WHERE total > 0")
    for m in all_marks:
        g = grade(m["marks"], m["total"])
        if g in grade_dist:
            grade_dist[g] += 1
    pass_count = sum(1 for m in all_marks if pct(m["marks"], m["total"]) >= 35)
    fail_count = len(all_marks) - pass_count

    # Dept-wise avg marks
    dept_marks = []
    for d in DEPARTMENTS:
        row = qone("SELECT ROUND(AVG(marks*100.0/total)::numeric, 1) as avg FROM marks WHERE department=%s AND total>0", (d,))
        dept_marks.append({"dept": d, "avg": float(row["avg"]) if row and row["avg"] is not None else 0.0})

    # Attendance status breakdown
    att_status = {}
    for s in ["Present", "Absent", "Leave", "Late", "Medical"]:
        row = qone("SELECT COUNT(*) as c FROM attendance WHERE status=%s", (s,))
        att_status[s] = row["c"] if row else 0

    return render_template("common/analytics.html",
        monthly=monthly,
        month_labels=[m["label"] for m in monthly],
        month_pcts=[m["pct"] for m in monthly],
        top5=student_att[:5], bottom5=bottom5,
        subj_marks=subj_marks,
        sm_labels=[r["subject"][:25] for r in subj_marks],
        sm_data=[r["avg"] for r in subj_marks],
        exam_breakdown=[dict(r) for r in exam_rows],
        exam_labels=[r["exam_type"] for r in exam_rows],
        exam_data=[r["avg"] for r in exam_rows],
        dept_att=dept_att,
        grade_dist=grade_dist,
        pass_count=pass_count, fail_count=fail_count,
        dept_marks=dept_marks,
        att_status=att_status,
        total_students=qone("SELECT COUNT(*) as c FROM students")["c"],
        total_att=qone("SELECT COUNT(*) as c FROM attendance")["c"],
        total_marks=qone("SELECT COUNT(*) as c FROM marks")["c"]
    )


# 6. CALENDAR
@admin_extra_bp.route("/calendar")
def calendar_view():
    role = session.get("role")
    if not role: return redirect("/login")
    month = request.args.get("month", "")
    if not month:
        month = date.today().strftime("%Y-%m")
    events = qry("SELECT * FROM events WHERE event_date LIKE %s ORDER BY event_date", (f"{month}%",))
    upcoming = qry("SELECT * FROM events WHERE event_date >= %s ORDER BY event_date LIMIT 10", (date.today().strftime("%Y-%m-%d"),))
    return render_template("common/calendar.html", events=events, upcoming=upcoming, month=month, role=role)

@admin_extra_bp.route("/save_event", methods=["POST"])
@login_required("admin")
def save_event():
    title = request.form.get("title", "").strip()
    event_date = request.form.get("event_date", "").strip()
    event_type = request.form.get("event_type", "Event").strip()
    description = request.form.get("description", "").strip()
    
    if title and event_date:
        exe("INSERT INTO events(title, event_date, event_type, description) VALUES(%s, %s, %s, %s)",
            (title, event_date, event_type, description))
        flash("Event saved successfully!", "success")
    else:
        flash("Title and Date are required.", "error")
    return redirect("/calendar")

# 7. ADMIN PROFILE
@admin_extra_bp.route("/admin_profile")
@login_required("admin")
def admin_profile():
    return render_template("admin/admin_profile.html")

# 8. EXPORT STUDENTS EXCEL
@admin_extra_bp.route("/export_students_excel")
@login_required("admin")
def export_students_excel():
    rows = qry("SELECT name, prn, roll, department, year, division, contact_number, email FROM students ORDER BY department, year, division, name")
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    hdr = ["Name", "PRN Number", "Department", "Year", "Division", "Contact Number", "Email"]
    for c, h in enumerate(hdr, 1):
        ws.cell(1, c, h).font = Font(bold=True)
    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row["name"])
        ws.cell(r, 2, row["prn"] or row["roll"])
        ws.cell(r, 3, row["department"])
        ws.cell(r, 4, row["year"])
        ws.cell(r, 5, row["division"] or "")
        ws.cell(r, 6, row["contact_number"] or "")
        ws.cell(r, 7, row["email"])
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="students.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# 9. ADD STUDENT / ADD FACULTY & SAVING
@admin_extra_bp.route("/add_student")
@login_required("admin")
def add_student():
    return render_template("admin/add_student.html", DEPARTMENTS=DEPARTMENTS, YEARS=YEARS)

@admin_extra_bp.route("/save_student", methods=["POST"])
@login_required("admin")
def save_student():
    name = request.form.get("name", "").strip()
    roll = request.form.get("roll", "").strip()
    dept = request.form.get("department", "").strip()
    year = request.form.get("year", "").strip()
    email = request.form.get("email", "").strip()
    division = request.form.get("division", "").strip()
    gender = request.form.get("gender", "").strip()
    dob = request.form.get("dob", "").strip() or None
    contact = request.form.get("contact_number", "").strip()
    parent = request.form.get("parent_contact", "").strip()
    address = request.form.get("address", "").strip()
    adm = request.form.get("admission_year", "").strip()
    
    if adm:
        try: adm = int(adm)
        except: adm = None
    else:
        adm = None
        
    pw_raw = request.form.get("password", "").strip() or "student123"
    pw = generate_password_hash(pw_raw)
    prn = request.form.get("prn", "").strip()
    roll = roll or prn

    if not name or not prn:
        return render_template("admin/add_student.html", DEPARTMENTS=DEPARTMENTS, YEARS=YEARS, error="Name and PRN Number are required.")
    
    try:
        exe("INSERT INTO students(name, roll, prn, department, year, email, password, division, gender, dob, contact_number, parent_contact, address, admission_year) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (name, roll, prn, dept, year, email, pw, division, gender, dob, contact, parent, address, adm))
        flash("Student saved successfully!", "success")
        return redirect("/students/?success=1")
    except Exception as e:
        logger.error(f"Student save failed: {e}")
        return render_template("admin/add_student.html", DEPARTMENTS=DEPARTMENTS, YEARS=YEARS, error=f"Could not save student. PRN may already exist.")

@admin_extra_bp.route("/add_faculty")
@login_required("admin")
def add_faculty():
    return render_template("admin/add_faculty.html", DEPARTMENTS=DEPARTMENTS, DESIGNATIONS=DESIGNATIONS)

@admin_extra_bp.route("/save_faculty", methods=["POST"])
@login_required("admin")
def save_faculty():
    name = request.form.get("name", "").strip()
    dept = request.form.get("department", "").strip()
    desig = request.form.get("designation", "").strip()
    email = request.form.get("email", "").strip()
    phone = request.form.get("phone", "").strip()
    qual = request.form.get("qualification", "").strip()
    jdate = request.form.get("joining_date", "").strip() or None
    pw_raw = request.form.get("password", "").strip() or "faculty123"
    pw = generate_password_hash(pw_raw)

    if not name or not email:
        return render_template("admin/add_faculty.html", DEPARTMENTS=DEPARTMENTS, DESIGNATIONS=DESIGNATIONS, error="Name and Email are required.")
    
    # Generate employee_id
    employee_id = f"FAC-{email.split('@')[0].upper()}"
    
    try:
        exe("INSERT INTO faculty(name, department, designation, email, phone, qualification, joining_date, password, employee_id) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (name, dept, desig, email, phone, qual, jdate, pw, employee_id))
        flash("Faculty saved successfully!", "success")
        return redirect("/faculty/?success=1")
    except Exception as e:
        logger.error(f"Faculty save failed: {e}")
        return render_template("admin/add_faculty.html", DEPARTMENTS=DEPARTMENTS, DESIGNATIONS=DESIGNATIONS, error=f"Could not save faculty. Email may already exist.")

# 10. DELETE STUDENT & WIPE SUMMARY
@admin_extra_bp.route("/delete_student", methods=["POST"])
@login_required("admin")
def delete_student():
    student_id = request.form.get("student_id", "").strip()
    if student_id:
        exe("DELETE FROM students WHERE id=%s", (student_id,))
        from utils.cache import erp_cache
        erp_cache.invalidate_pattern("student_list:*")
    return redirect("/students/")


# ── STUDENT MANAGEMENT ACTIONS ──────────────────────────────
@admin_extra_bp.route("/edit_student", methods=["POST"])
@login_required("admin")
def edit_student_route():
    sid = request.form.get("student_id")
    adm = request.form.get("admission_year", "").strip()
    if adm:
        try: adm = int(adm)
        except: adm = None
    else:
        adm = None

    prn = request.form.get("prn", "").strip()
    roll = request.form.get("roll", "").strip() or prn

    exe(
        "UPDATE students SET name=%s,roll=%s,prn=%s,department=%s,year=%s,email=%s,division=%s,gender=%s,dob=%s,contact_number=%s,parent_contact=%s,address=%s,admission_year=%s WHERE id=%s",
        (request.form.get("name", ""), roll, prn, request.form.get("department", ""),
         request.form.get("year", ""), request.form.get("email", ""), 
         request.form.get("division", ""), request.form.get("gender", ""),
         request.form.get("dob", "") or None, request.form.get("contact_number", ""),
         request.form.get("parent_contact", ""), request.form.get("address", ""), adm, sid)
    )
    from utils.cache import erp_cache
    erp_cache.invalidate_pattern("student_list:*")
    return redirect("/students/")

@admin_extra_bp.route("/bulk_delete_students", methods=["POST"])
@login_required("admin")
def bulk_delete_students_route():
    ids = request.form.getlist("ids[]")
    for i in ids:
        exe("DELETE FROM students WHERE id=%s", (i,))
    from utils.cache import erp_cache
    erp_cache.invalidate_pattern("student_list:*")
    return redirect(f"/students/?deleted={len(ids)}")

@admin_extra_bp.route("/fix_students", methods=["POST"])
@login_required("admin")
def fix_students_route():
    bad = qry("""SELECT id FROM students WHERE
        name IS NULL OR name='' OR name='Student Name' OR
        CAST(name AS TEXT) ~ '^[0-9]' OR
        roll IS NULL OR roll='' OR roll='Roll No.' OR roll='Roll'
    """)
    for r in bad:
        exe("DELETE FROM students WHERE id=%s", (r["id"],))
    from utils.cache import erp_cache
    erp_cache.invalidate_pattern("student_list:*")
    return redirect(f"/students/?cleaned={len(bad)}")


# ── FACULTY MANAGEMENT ACTIONS ──────────────────────────────
@admin_extra_bp.route("/edit_faculty", methods=["POST"])
@login_required("admin")
def edit_faculty_route():
    fid = request.form.get("faculty_id", "")
    try:
        exe(
            "UPDATE faculty SET name=%s,department=%s,designation=%s,email=%s,phone=%s,qualification=%s WHERE id=%s",
            (request.form.get("name", ""), request.form.get("department", ""), request.form.get("designation", ""),
             request.form.get("email", ""), request.form.get("phone", ""), request.form.get("qualification", ""), fid)
        )
    except Exception:
        pass
    return redirect("/faculty/")

@admin_extra_bp.route("/delete_faculty", methods=["POST"])
@login_required("admin")
def delete_faculty_route():
    fid = request.form.get("faculty_id", "").strip()
    if fid:
        exe("DELETE FROM faculty WHERE id=%s", (fid,))
    return redirect("/faculty/")

@admin_extra_bp.route("/bulk_delete_faculty", methods=["POST"])
@login_required("admin")
def bulk_delete_faculty_route():
    ids = request.form.getlist("ids[]")
    for i in ids:
        exe("DELETE FROM faculty WHERE id=%s", (i,))
    return redirect(f"/faculty/?deleted={len(ids)}")

@admin_extra_bp.route("/fix_faculty", methods=["POST"])
@login_required("admin")
def fix_faculty_route():
    bad = qry("""SELECT id FROM faculty WHERE
        name IS NULL OR name='' OR name='Name of Faculty' OR
        name='Faculty Name' OR name='Name' OR
        email IS NULL OR email='' OR email='Email Id' OR email='Email'
    """)
    for r in bad:
        exe("DELETE FROM faculty WHERE id=%s", (r["id"],))
    return redirect(f"/faculty/?cleaned={len(bad)}")


# ── STUDENT & FACULTY DATA EXPORTS ───────────────────────────
@admin_extra_bp.route("/api/students/export")
@login_required("admin")
def api_export_students():
    import io, csv
    rows = qry("SELECT * FROM students ORDER BY id DESC")
    si = io.StringIO()
    cw = csv.writer(si)
    if rows:
        cw.writerow(list(rows[0].keys()))
        for r in rows:
            cw.writerow(list(r.values()))
            
    from flask import make_response
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=students_export.csv"
    output.headers["Content-type"] = "text/csv"
    return output

@admin_extra_bp.route("/export_faculty_excel")
@login_required("admin")
def export_faculty_excel():
    rows = qry("SELECT name,department,designation,email,phone,qualification,joining_date FROM faculty ORDER BY name")
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty"
    hdrs = ["Name", "Department", "Designation", "Email", "Phone", "Qualification", "Joining Date"]
    for c, h in enumerate(hdrs, 1):
        ws.cell(1, c, h).font = Font(bold=True)
    for r, row in enumerate(rows, 2):
        for c, k in enumerate(["name", "department", "designation", "email", "phone", "qualification", "joining_date"], 1):
            ws.cell(r, c, row[k])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="faculty.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@admin_extra_bp.route("/clear_all_attendance_summary", methods=["POST"])
@login_required("admin")
def clear_all_attendance_summary():
    phrase = (request.form.get("confirm_phrase") or "").strip()
    if phrase != "DELETE ALL PDF SUMMARY":
        return redirect("/attendance_dashboard?error=bad_summary_confirm")
    exe("DELETE FROM attendance_summary")
    return redirect("/attendance_dashboard?summary_cleared=1")

def safe_int(val, default=0):
    try:
        return int(val)
    except (ValueError, TypeError):
        return default

# Defaulters Report Route
@admin_extra_bp.route("/admin/reports/defaulters")
@login_required("admin")
def reports_defaulters():
    threshold = safe_int(request.args.get("threshold", "75"))
    dept = request.args.get("department", "").strip()
    div = request.args.get("division", "").strip()
    subj = request.args.get("subject", "").strip()

    where_clauses = ["percentage < %s"]
    params = [threshold]

    if dept:
        where_clauses.append("department = %s")
        params.append(dept)
    if div:
        where_clauses.append("division = %s")
        params.append(div)
    if subj:
        where_clauses.append("subject = %s")
        params.append(subj)

    where_str = " AND ".join(where_clauses)

    query_str = f"""
        WITH subject_attendance AS (
            SELECT 
                student_id, student_name, subject, attended, total, division, department,
                CASE WHEN total > 0 THEN ROUND(attended * 100.0 / total, 1) ELSE 0.0 END as percentage
            FROM attendance_summary
            UNION ALL
            SELECT 
                a.student_id, s.name as student_name, a.subject, 
                COUNT(*) FILTER (WHERE a.status IN ('Present', 'Late')) as attended, 
                COUNT(*) as total, s.division, s.department,
                CASE WHEN COUNT(*) > 0 THEN ROUND(COUNT(*) FILTER (WHERE a.status IN ('Present', 'Late')) * 100.0 / COUNT(*), 1) ELSE 0.0 END as percentage
            FROM attendance a 
            JOIN students s ON a.student_id = s.id 
            WHERE a.student_id NOT IN (SELECT DISTINCT student_id FROM attendance_summary)
            GROUP BY a.student_id, s.name, a.subject, s.division, s.department
        )
        SELECT * FROM subject_attendance
        WHERE {where_str}
        ORDER BY department, division, student_name, subject
    """  # nosec B608 - where_str is composed of safe hardcoded literals and parameters
    
    records = qry(query_str, params)
    subjects = [r["subject"] for r in qry("SELECT DISTINCT subject FROM (SELECT subject FROM attendance_summary UNION SELECT subject FROM attendance) s WHERE subject IS NOT NULL AND subject != '' ORDER BY subject")]
    
    return render_template(
        "attendance/defaulters_report.html",
        records=records,
        subjects=subjects,
        departments=DEPARTMENTS,
        divisions=DIVISIONS,
        threshold=threshold,
        selected_dept=dept,
        selected_div=div,
        selected_subj=subj
    )

# Defaulters Excel Export Route
@admin_extra_bp.route("/admin/reports/defaulters/export")
@login_required("admin")
def reports_defaulters_export():
    threshold = safe_int(request.args.get("threshold", "75"))
    dept = request.args.get("department", "").strip()
    div = request.args.get("division", "").strip()
    subj = request.args.get("subject", "").strip()

    where_clauses = ["percentage < %s"]
    params = [threshold]

    if dept:
        where_clauses.append("department = %s")
        params.append(dept)
    if div:
        where_clauses.append("division = %s")
        params.append(div)
    if subj:
        where_clauses.append("subject = %s")
        params.append(subj)

    where_str = " AND ".join(where_clauses)

    query_str = f"""
        WITH subject_attendance AS (
            SELECT 
                student_id, student_name, subject, attended, total, division, department,
                CASE WHEN total > 0 THEN ROUND(attended * 100.0 / total, 1) ELSE 0.0 END as percentage
            FROM attendance_summary
            UNION ALL
            SELECT 
                a.student_id, s.name as student_name, a.subject, 
                COUNT(*) FILTER (WHERE a.status IN ('Present', 'Late')) as attended, 
                COUNT(*) as total, s.division, s.department,
                CASE WHEN COUNT(*) > 0 THEN ROUND(COUNT(*) FILTER (WHERE a.status IN ('Present', 'Late')) * 100.0 / COUNT(*), 1) ELSE 0.0 END as percentage
            FROM attendance a 
            JOIN students s ON a.student_id = s.id 
            WHERE a.student_id NOT IN (SELECT DISTINCT student_id FROM attendance_summary)
            GROUP BY a.student_id, s.name, a.subject, s.division, s.department
        )
        SELECT * FROM subject_attendance
        WHERE {where_str}
        ORDER BY department, division, student_name, subject
    """  # nosec B608 - where_str is composed of safe hardcoded literals and parameters
    
    records = qry(query_str, params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Subject Defaulters"

    hdr = ["Student Name", "Department", "Division", "Subject", "Attended", "Total", "Percentage"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(records, 2):
        ws.cell(r, 1, row["student_name"])
        ws.cell(r, 2, row["department"])
        ws.cell(r, 3, row["division"])
        ws.cell(r, 4, row["subject"])
        ws.cell(r, 5, row["attended"])
        ws.cell(r, 6, row["total"])
        pct_cell = ws.cell(r, 7, f"{row['percentage']}%")
        pct_cell.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"Subject_Defaulters_Report_{datetime.today().strftime('%Y-%m-%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Division Attendance Report Route
@admin_extra_bp.route("/admin/reports/division_attendance")
@login_required("admin")
def reports_division_attendance():
    today_dt = date.today()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    
    if not start_date_str:
        start_date = today_dt - timedelta(days=30)
        start_date_str = start_date.strftime("%Y-%m-%d")
    if not end_date_str:
        end_date_str = today_dt.strftime("%Y-%m-%d")

    dept = request.args.get("department", "").strip()
    div = request.args.get("division", "").strip()

    where_clauses = ["a.date >= %s", "a.date <= %s"]
    params = [start_date_str, end_date_str]

    if dept:
        where_clauses.append("s.department = %s")
        params.append(dept)
    if div:
        where_clauses.append("s.division = %s")
        params.append(div)

    where_str = " AND ".join(where_clauses)

    query_str = f"""
        SELECT 
            s.division,
            s.department,
            COUNT(*) FILTER (WHERE a.status IN ('Present', 'Late')) as attended,
            COUNT(*) as total,
            CASE WHEN COUNT(*) > 0 THEN ROUND(COUNT(*) FILTER (WHERE a.status IN ('Present', 'Late')) * 100.0 / COUNT(*), 1) ELSE 0.0 END as percentage
        FROM attendance a
        JOIN students s ON a.student_id = s.id
        WHERE {where_str}
        GROUP BY s.division, s.department
        ORDER BY s.department, s.division
    """  # nosec B608 - where_str is composed of safe hardcoded literals and parameters

    records = qry(query_str, params)

    return render_template(
        "attendance/division_attendance_report.html",
        records=records,
        departments=DEPARTMENTS,
        divisions=DIVISIONS,
        start_date=start_date_str,
        end_date=end_date_str,
        selected_dept=dept,
        selected_div=div
    )

# Division Attendance Excel Export Route
@admin_extra_bp.route("/admin/reports/division_attendance/export")
@login_required("admin")
def reports_division_attendance_export():
    today_dt = date.today()
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()
    
    if not start_date_str:
        start_date = today_dt - timedelta(days=30)
        start_date_str = start_date.strftime("%Y-%m-%d")
    if not end_date_str:
        end_date_str = today_dt.strftime("%Y-%m-%d")

    dept = request.args.get("department", "").strip()
    div = request.args.get("division", "").strip()

    where_clauses = ["a.date >= %s", "a.date <= %s"]
    params = [start_date_str, end_date_str]

    if dept:
        where_clauses.append("s.department = %s")
        params.append(dept)
    if div:
        where_clauses.append("s.division = %s")
        params.append(div)

    where_str = " AND ".join(where_clauses)

    query_str = f"""
        SELECT 
            s.division,
            s.department,
            COUNT(*) FILTER (WHERE a.status IN ('Present', 'Late')) as attended,
            COUNT(*) as total,
            CASE WHEN COUNT(*) > 0 THEN ROUND(COUNT(*) FILTER (WHERE a.status IN ('Present', 'Late')) * 100.0 / COUNT(*), 1) ELSE 0.0 END as percentage
        FROM attendance a
        JOIN students s ON a.student_id = s.id
        WHERE {where_str}
        GROUP BY s.division, s.department
        ORDER BY s.department, s.division
    """  # nosec B608 - where_str is composed of safe hardcoded literals and parameters

    records = qry(query_str, params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Division Attendance"

    hdr = ["Department", "Division", "Attended Lectures", "Total Lectures", "Average Attendance %"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(records, 2):
        ws.cell(r, 1, row["department"])
        ws.cell(r, 2, row["division"])
        ws.cell(r, 3, row["attended"])
        ws.cell(r, 4, row["total"])
        pct_cell = ws.cell(r, 5, f"{row['percentage']}%")
        pct_cell.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"Division_Attendance_Report_{start_date_str}_to_{end_date_str}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ── ATTENDANCE EXTENSION ROUTES ──────────────────────────────
@admin_extra_bp.route("/attendance", methods=["GET", "POST"])
@login_required(["admin", "faculty"])
def attendance_entry_page():
    if request.method == "POST":
        from routes.attendance import handle_single_mark
        return handle_single_mark(request.form, session)
        
    from services.attendance_service import attendance_page_context
    from utils.helpers import get_today_str
    ctx = attendance_page_context(
        role=session.get("role", "admin"),
        actor_id=session.get("faculty_id") or session.get("student_id"),
        actor_name=session.get("name", "")
    )
    return render_template(
        "attendance/attendance.html",
        students=ctx.get("students", []),
        subjects=ctx.get("subjects", []),
        divs=ctx.get("divs", []),
        today=ctx.get("today", get_today_str()),
        DEPARTMENTS=DEPARTMENTS,
    )

@admin_extra_bp.route("/save_attendance", methods=["POST"])
@login_required(["admin", "faculty"])
def save_single_attendance():
    from routes.attendance import handle_single_mark
    return handle_single_mark(request.form, session)

@admin_extra_bp.route("/attendance_bulk", methods=["POST"])
@admin_extra_bp.route("/save_bulk_attendance", methods=["POST"])
@login_required(["admin", "faculty"])
def save_bulk_attendance():
    from routes.attendance import handle_bulk_mark
    return handle_bulk_mark(request.form, session)

@admin_extra_bp.route("/download_attendance_template")
@login_required(["admin", "faculty"])
def download_attendance_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Template"
    headers = ["student_name", "subject", "date", "status", "remark"]
    sample = ["John Doe", "DBMS", date.today().strftime("%Y-%m-%d"), "Present", ""]
    for c, header in enumerate(headers, 1):
        cell = ws.cell(1, c, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        ws.cell(2, c, sample[c - 1])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")) + 4, 16)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="attendance_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_extra_bp.route("/api/admin/attendance_trend")
@login_required("admin")
def api_attendance_trend():
    view = request.args.get("view", "daily")
    dept = request.args.get("dept")
    division = request.args.get("division")
    
    # Query all attendance records (with optional filters)
    sql = """
        SELECT a.date, a.status 
        FROM attendance a
        INNER JOIN students s ON a.student_id = s.id
        WHERE 1=1
    """
    params = []
    if dept:
        sql += " AND s.department = %s"
        params.append(dept)
    if division:
        sql += " AND s.division = %s"
        params.append(division)
    
    sql += " ORDER BY a.date ASC"
    
    rows = qry(sql, params)
    
    # Aggregate in Python
    # We group by date
    daily_stats = {}
    for r in rows:
        dt_str = r["date"]
        if isinstance(dt_str, (date, datetime)):
            dt_str = dt_str.strftime("%Y-%m-%d")
        else:
            dt_str = str(dt_str)
        status = r["status"]
        if dt_str not in daily_stats:
            daily_stats[dt_str] = {"present": 0, "total": 0}
        daily_stats[dt_str]["total"] += 1
        if status in ("Present", "Late"):
            daily_stats[dt_str]["present"] += 1
            
    # Sort dates chronologically
    sorted_dates = sorted(daily_stats.keys())
    
    labels = []
    data = []
    
    if view == "weekly":
        # Group by week (Monday start)
        weeks = {}
        for d_str in sorted_dates:
            try:
                dt = datetime.strptime(d_str, "%Y-%m-%d")
                # Find the Monday of that week
                monday = dt - timedelta(days=dt.weekday())
                week_key = monday.strftime("%Y-%m-%d")
                if week_key not in weeks:
                    weeks[week_key] = {"present": 0, "total": 0}
                weeks[week_key]["present"] += daily_stats[d_str]["present"]
                weeks[week_key]["total"] += daily_stats[d_str]["total"]
            except Exception:
                continue
        # Sort weeks chronologically
        for wk in sorted(weeks.keys()):
            try:
                dt = datetime.strptime(wk, "%Y-%m-%d")
                labels.append(dt.strftime("%d %b"))
                pct_val = round(weeks[wk]["present"] / weeks[wk]["total"] * 100, 1) if weeks[wk]["total"] > 0 else 0
                data.append(pct_val)
            except Exception:
                continue
    elif view == "monthly":
        # Group by month
        months = {}
        for d_str in sorted_dates:
            try:
                dt = datetime.strptime(d_str, "%Y-%m-%d")
                month_key = dt.strftime("%Y-%m")
                if month_key not in months:
                    months[month_key] = {"present": 0, "total": 0}
                months[month_key]["present"] += daily_stats[d_str]["present"]
                months[month_key]["total"] += daily_stats[d_str]["total"]
            except Exception:
                continue
        # Sort months chronologically
        for mn in sorted(months.keys()):
            try:
                dt = datetime.strptime(mn + "-01", "%Y-%m-%d")
                labels.append(dt.strftime("%b %Y"))
                pct_val = round(months[mn]["present"] / months[mn]["total"] * 100, 1) if months[mn]["total"] > 0 else 0
                data.append(pct_val)
            except Exception:
                continue
    else: # daily
        # Show last 30 active days
        recent_dates = sorted_dates[-30:]
        for d_str in recent_dates:
            try:
                dt = datetime.strptime(d_str, "%Y-%m-%d")
                labels.append(dt.strftime("%d %b"))
                pct_val = round(daily_stats[d_str]["present"] / daily_stats[d_str]["total"] * 100, 1) if daily_stats[d_str]["total"] > 0 else 0
                data.append(pct_val)
            except Exception:
                continue
                
    return jsonify({
        "status": "success",
        "labels": labels,
        "data": data
    })


@admin_extra_bp.route("/api/admin/attendance_calendar")
@login_required("admin")
def api_attendance_calendar():
    try:
        year = int(request.args.get("year", datetime.now().year))
        month = int(request.args.get("month", datetime.now().month))
    except ValueError:
        return jsonify({"status": "error", "message": "Invalid year/month parameters"}), 400
        
    dept = request.args.get("dept")
    division = request.args.get("division")
    
    month_str = f"{year}-{month:02d}-%"
    
    # Query attendance grouped by date in that month
    sql = """
        SELECT a.date, a.status 
        FROM attendance a
        INNER JOIN students s ON a.student_id = s.id
        WHERE a.date::text LIKE %s
    """
    params = [month_str]
    if dept:
        sql += " AND s.department = %s"
        params.append(dept)
    if division:
        sql += " AND s.division = %s"
        params.append(division)
        
    rows = qry(sql, params)
    
    # Aggregate by date
    daily_stats = {}
    for r in rows:
        d_str = r["date"]
        if isinstance(d_str, (date, datetime)):
            d_str = d_str.strftime("%Y-%m-%d")
        else:
            d_str = str(d_str)
        status = r["status"]
        if d_str not in daily_stats:
            daily_stats[d_str] = {"present": 0, "absent": 0, "total": 0}
        daily_stats[d_str]["total"] += 1
        if status in ("Present", "Late"):
            daily_stats[d_str]["present"] += 1
        else:
            daily_stats[d_str]["absent"] += 1
            
    # Calculate percentage for each day
    calendar_data = {}
    for d_str, stats in daily_stats.items():
        pct_val = round(stats["present"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
        calendar_data[d_str] = {
            "pct": pct_val,
            "present": stats["present"],
            "absent": stats["absent"],
            "total": stats["total"]
        }
        
    return jsonify({
        "status": "success",
        "data": calendar_data
    })


@admin_extra_bp.route("/api/admin/attendance_date_details")
@login_required("admin")
def api_attendance_date_details():
    target_date = request.args.get("date")
    if not target_date:
        return jsonify({"status": "error", "message": "Date is required"}), 400
        
    # Get overall stats
    summary = qone("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN status IN ('Present', 'Late') THEN 1 ELSE 0 END) as present,
               SUM(CASE WHEN status = 'Absent' THEN 1 ELSE 0 END) as absent
        FROM attendance
        WHERE date::text = %s
    """, (target_date,))
    
    if not summary or summary["total"] == 0:
        return jsonify({"status": "success", "sessions": [], "stats": None})
        
    # Group by subject + division
    sessions = qry("""
        SELECT subject, division, faculty,
               COUNT(*) as total_students,
               SUM(CASE WHEN status IN ('Present', 'Late') THEN 1 ELSE 0 END) as present_students
        FROM attendance
        WHERE date::text = %s
        GROUP BY subject, division, faculty
        ORDER BY subject
    """, (target_date,))
    
    session_list = []
    for s in sessions:
        present_count = s["present_students"] or 0
        total_count = s["total_students"] or 0
        pct_val = round(present_count / total_count * 100, 1) if total_count > 0 else 0
        session_list.append({
            "subject": s["subject"],
            "division": s["division"] or "-",
            "faculty": s["faculty"] or "N/A",
            "present": present_count,
            "absent": total_count - present_count,
            "pct": pct_val
        })
        
    return jsonify({
        "status": "success",
        "stats": {
            "date": target_date,
            "total": summary["total"],
            "present": summary["present"] or 0,
            "absent": summary["absent"] or 0,
            "pct": round((summary["present"] or 0) / summary["total"] * 100, 1) if summary["total"] > 0 else 0
        },
        "sessions": session_list
    })


@admin_extra_bp.route("/admin/marks/export-excel", methods=["GET"])
@login_required("admin")
def admin_export_marks_excel():
    division = request.args.get("division", "").strip()
    semester = request.args.get("semester", "SEM IV").strip()
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    # Remove default Sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    tabs = ["Comp", "IT", "AIDS", "AIML"]
    tab_depts = {
        "Comp": ["CS", "Computer", "Comp"],
        "IT": ["IT", "Information Technology"],
        "AIDS": ["AIDS"],
        "AIML": ["AIML"]
    }
    
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True)
    
    for tab_name in tabs:
        ws = wb.create_sheet(title=tab_name)
        
        # Write Row 1 blank
        ws.append([])
        
        # Get subjects for this department and semester
        dept_names = tab_depts.get(tab_name, [tab_name])
        placeholders = ",".join(["%s"] * len(dept_names))
        
        query_sub = f"SELECT * FROM subjects_master WHERE semester = %s AND department IN ({placeholders}) ORDER BY subject_code"  # nosec B608
        params_sub = [semester] + dept_names
        subjects = qry(query_sub, params_sub)
        
        # Merged cells for subject codes in Row 2
        for idx, sub in enumerate(subjects):
            start_col = 4 + idx * 6
            end_col = start_col + 5
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=start_col, value=sub["subject_code"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            
        # Format Row 2 and Row 3 header cells
        num_cols = 3 + len(subjects) * 6
        for c in range(1, num_cols + 1):
            cell2 = ws.cell(row=2, column=c)
            if c >= 4:
                cell2.fill = header_fill
                cell2.font = header_font
            cell3 = ws.cell(row=3, column=c)
            cell3.fill = header_fill
            cell3.font = header_font
            cell3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # Write Row 3 headers
        ws.cell(row=3, column=1, value="Sr.No")
        ws.cell(row=3, column=2, value="Student Name")
        ws.cell(row=3, column=3, value="PRN")
        
        for idx, sub in enumerate(subjects):
            start_col = 4 + idx * 6
            ws.cell(row=3, column=start_col + 0, value="Assignment(05)")
            ws.cell(row=3, column=start_col + 1, value="Attendance(05)")
            ws.cell(row=3, column=start_col + 2, value="Teaching(10)")
            ws.cell(row=3, column=start_col + 3, value="UT(20)")
            ws.cell(row=3, column=start_col + 4, value="MSE(20)")
            ws.cell(row=3, column=start_col + 5, value="TOTAL(60)")
            
        # Get students for this department
        if division:
            query_stu = f"SELECT id, name, roll, prn FROM students WHERE department IN ({placeholders}) AND division = %s ORDER BY roll"  # nosec B608
            params_stu = dept_names + [division]
        else:
            query_stu = f"SELECT id, name, roll, prn FROM students WHERE department IN ({placeholders}) ORDER BY roll"  # nosec B608
            params_stu = dept_names
            
        students = qry(query_stu, params_stu)
        
        # Write Row 4+ student data
        for row_idx, student in enumerate(students, start=4):
            sr_no = row_idx - 3
            ws.cell(row=row_idx, column=1, value=sr_no)
            ws.cell(row=row_idx, column=2, value=student["name"])
            ws.cell(row=row_idx, column=3, value=student["prn"] or "")
            
            is_fail = False
            is_low = False
            has_any_marks = False
            
            for idx, sub in enumerate(subjects):
                subj_code = sub["subject_code"]
                subj_name = sub["subject_name"]
                
                m_row = qone("SELECT * FROM marks WHERE student_id = %s AND (subject_code = %s OR subject = %s) AND semester = %s",
                             (student["id"], subj_code, subj_name, semester))
                             
                start_col = 4 + idx * 6
                if m_row:
                    has_any_marks = True
                    ws.cell(row=row_idx, column=start_col + 0, value=m_row["assignment_marks"])
                    ws.cell(row=row_idx, column=start_col + 1, value=m_row["attendance_marks"])
                    ws.cell(row=row_idx, column=start_col + 2, value=m_row["teaching_assessment"])
                    ws.cell(row=row_idx, column=start_col + 3, value=m_row["ut_marks"])
                    ws.cell(row=row_idx, column=start_col + 4, value=m_row["mse_marks"])
                    total_obtained = m_row["marks"]
                    ws.cell(row=row_idx, column=start_col + 5, value=total_obtained)
                    
                    if total_obtained is not None:
                        if total_obtained < 24:
                            is_fail = True
                        elif 24 <= total_obtained <= 29:
                            is_low = True
                            
            if has_any_marks:
                if is_fail:
                    row_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    row_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                elif is_low:
                    row_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    row_font = Font(name="Arial", size=10, bold=False, color="000000")
                else:
                    row_fill = None
                    row_font = Font(name="Arial", size=10, bold=False)
            else:
                row_fill = None
                row_font = Font(name="Arial", size=10, bold=False)
                
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if row_fill:
                    cell.fill = row_fill
                if row_font:
                    cell.font = row_font
                    
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        for col_idx in range(4, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Master_SY_Sem_IV_{division or 'ALL'}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



