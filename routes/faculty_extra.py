from flask import Blueprint, request, redirect, session, render_template, send_file, jsonify, url_for
from datetime import datetime, date
import re
import io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils.pg_wrapper import qry, qone, exe
from blueprints.auth.decorators import login_required
from utils.helpers import pct, safe_int
from config import Config

faculty_extra_bp = Blueprint("faculty_extra", __name__)

DEPARTMENTS = Config.DEPARTMENTS
DIVISIONS = Config.DIVISIONS
SEMESTERS = Config.SEMESTERS
DAYS = Config.DAYS
DAY_ORD = Config.DAY_ORD

# ── Helpers ──────────────────────────────────────────────────

def today_str():
    return date.today().strftime("%Y-%m-%d")

def normalise_date(raw):
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None

def normalize_time(t):
    if not t:
        return t
    t = str(t).strip()
    parts = re.split(r'\s*-\s*', t)
    result = []
    for part in parts:
        m = re.match(r'^(\d{1,2}):(\d{2})$', part.strip())
        if m:
            result.append(f"{int(m.group(1)):02d}:{m.group(2)}")
        else:
            result.append(part.strip())
    return '-'.join(result)

def grade(marks, total):
    p = pct(marks, total)
    if p >= 75: return "O"
    if p >= 65: return "A"
    if p >= 55: return "B"
    if p >= 45: return "C"
    if p >= 35: return "D"
    return "F"

def resolve_student_id(name, roll=None):
    if not name or not str(name).strip():
        return None
    name = str(name).strip()
    r = str(roll).strip() if roll is not None and str(roll).strip() else None
    if r:
        row = qone("SELECT id FROM students WHERE name=%s AND roll=%s", (name, r))
        if row:
            return row["id"]
    row = qone("SELECT id FROM students WHERE name=%s ORDER BY id LIMIT 1", (name,))
    return row["id"] if row else None

def marks_match_student_params(sid, name):
    return sid, name

def marks_match_student_sql(prefix=""):
    p = f"{prefix}." if prefix else ""
    return f"({p}student_id = %s OR ({p}student_id IS NULL AND {p}student_name = %s))"

def validate_password_change(stored_hash, current_pw, new_pw, confirm_pw):
    from werkzeug.security import check_password_hash
    from utils.password_policy import validate_password
    if not check_password_hash(stored_hash, current_pw):
        return "invalid_current"
    if not new_pw:
        return "empty_new"
    if new_pw != confirm_pw:
        return "mismatch"
    is_valid, err_msg = validate_password(new_pw)
    if not is_valid:
        return err_msg
    return None

def hash_password(plain):
    from werkzeug.security import generate_password_hash
    return generate_password_hash(plain)

def _cumulative_marks_data_aggregated(dept="", student_filter="", faculty_id=None):
    data = []

    def _add(name, roll, d_dept, obtained, total_m, exams):
        if not total_m:
            return
        pv = round(obtained / total_m * 100, 1) if total_m else 0
        data.append({
            "name": name,
            "roll": roll or "",
            "dept": d_dept or "",
            "obtained": obtained,
            "total": total_m,
            "pct": pv,
            "grade": grade(obtained, total_m),
            "exams": exams,
        })

    wh = ["m.student_id IS NOT NULL"]
    params = []
    if faculty_id is not None:
        wh.append("m.faculty_id=%s")
        params.append(faculty_id)
    if dept:
        wh.append("m.department=%s")
        params.append(dept)
    if student_filter:
        wh.append("m.student_name LIKE %s")
        params.append(f"%{student_filter}%")
    ws = " AND ".join(wh)
    
    rows_std = qry(f"""
        SELECT MAX(m.student_name) AS student_name, MAX(m.roll) AS roll,
        MAX(m.department) AS department, SUM(m.marks) AS obtained,
        SUM(m.total) AS total_m, COUNT(*) AS exams
        FROM marks m WHERE {ws} GROUP BY m.student_id HAVING SUM(m.total) > 0
    """, params)  # nosec B608 - ws is composed of safe hardcoded literals and parameters
    for r in rows_std:
        _add(r["student_name"], r["roll"], r["department"], r["obtained"], r["total_m"], r["exams"])

    wh2 = ["m.student_id IS NULL"]
    params2 = []
    if faculty_id is not None:
        wh2.append("m.faculty_id=%s")
        params2.append(faculty_id)
    if dept:
        wh2.append("m.department=%s")
        params2.append(dept)
    if student_filter:
        wh2.append("m.student_name LIKE %s")
        params2.append(f"%{student_filter}%")
    ws2 = " AND ".join(wh2)
    
    rows_legacy = qry(f"""
        SELECT MAX(m.student_name) AS student_name, MAX(m.roll) AS roll,
        MAX(m.department) AS department, SUM(m.marks) AS obtained,
        SUM(m.total) AS total_m, COUNT(*) AS exams
        FROM marks m WHERE {ws2}
        GROUP BY m.student_name, m.roll, m.department HAVING SUM(m.total) > 0
    """, params2)  # nosec B608 - ws2 is composed of safe hardcoded literals and parameters
    for r in rows_legacy:
        _add(r["student_name"], r["roll"], r["department"], r["obtained"], r["total_m"], r["exams"])

    data.sort(key=lambda x: -x["pct"])
    return data

# ── Routes ───────────────────────────────────────────────────

@faculty_extra_bp.route("/faculty_marks")
@login_required("faculty")
def faculty_marks():
    fid = session["faculty_id"]
    students_list = qry("SELECT name,roll FROM students ORDER BY name")
    my_subjects   = qry("SELECT * FROM subjects WHERE teacher LIKE %s ORDER BY name", (f"%{session['name']}%",))
    marks         = qry("SELECT * FROM marks WHERE faculty_id=%s ORDER BY id DESC", (fid,))
    return render_template("faculty/faculty_marks.html",
                           students=students_list, my_subjects=my_subjects,
                           marks=marks, today=today_str())

@faculty_extra_bp.route("/faculty_save_marks", methods=["POST"])
@login_required("faculty")
def faculty_save_marks():
    fid = session["faculty_id"]
    if request.is_json:
        data = request.json
    else:
        data = request.form

    # COMPONENT 3: Validation — max marks per component
    COMPONENT_MAXES = {
        'assignment_marks': 5,
        'attendance_marks': 5,
        'teaching_assessment': 10,
        'ut_marks': 20,
        'mse_marks': 20,
    }
    errors = {}
    for field, max_val in COMPONENT_MAXES.items():
        val = float(data.get(field, 0) or 0)
        if val < 0:
            errors[field] = "Cannot be negative"
        if val > max_val:
            errors[field] = f"Max is {max_val}, got {val}"
    if errors:
        return jsonify({"errors": errors}), 400

    student_name = data.get("student_name", "").strip()
    roll_row = qone("SELECT id, roll, department, division, prn FROM students WHERE name=%s OR roll=%s", (student_name, student_name))
    if not roll_row:
        # fallback lookup by student_id or roll
        student_id = data.get("student_id")
        if student_id:
            roll_row = qone("SELECT id, roll, department, division, prn FROM students WHERE id=%s", (student_id,))

    stu_id = roll_row["id"] if roll_row else None
    roll   = roll_row["roll"] if roll_row else data.get("roll", "")
    dept   = roll_row["department"] if roll_row else data.get("department", "")
    prn_number = roll_row["prn"] if roll_row else data.get("prn_number", "")
    if not prn_number:
        prn_number = ""

    subject_name = data.get("subject", "").strip()
    sub_row = qone("SELECT subject_code FROM subjects_master WHERE subject_name=%s OR subject_code=%s", (subject_name, subject_name))
    subject_code = sub_row["subject_code"] if sub_row else ""

    exam_type = data.get("exam_type", "Semester Exam")
    assignment_m = float(data.get("assignment_marks", 0) or 0)
    attendance_m = float(data.get("attendance_marks", 0) or 0)
    teaching_m   = float(data.get("teaching_assessment", 0) or 0)
    ut_m         = float(data.get("ut_marks", 0) or 0)
    mse_m        = float(data.get("mse_marks", 0) or 0)
    remarks      = data.get("remarks", "").strip()
    date_val     = data.get("date", today_str())
    semester_val = data.get("semester", "SEM IV")

    # Insert marks row
    exe("""INSERT INTO marks(faculty_id, student_id, student_name, roll, subject, department,
                             marks, total, exam_type, date,
                             assignment_marks, attendance_marks, teaching_assessment,
                             ut_marks, mse_marks, remarks, prn_number, subject_code, semester)
           VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
        (fid, stu_id, student_name, roll,
         subject_name, dept,
         0.0, 60.0, # Will be updated below
         exam_type, date_val,
         assignment_m, attendance_m, teaching_m, ut_m, mse_m,
         remarks, prn_number, subject_code, semester_val))

    # COMPONENT 2: Auto-calculate total in marks save
    total = (
        float(assignment_m or 0) +
        float(attendance_m or 0) +
        float(teaching_m or 0) +
        float(ut_m or 0) +
        float(mse_m or 0)
    )
    exe("""UPDATE marks SET marks=%s, total=%s
           WHERE student_id=%s AND subject=%s AND semester=%s""",
        (total, 60.0, stu_id, subject_name, semester_val))

    # COMPONENT 4: Pass/Fail + Grade auto-calculation
    from services.results_service import calculate_result
    total_val, grade_val, result_val, passed = calculate_result(
        assignment_m, attendance_m, teaching_m, ut_m, mse_m
    )
    exe("""UPDATE marks SET grade=%s, result=%s
           WHERE student_id=%s AND subject=%s AND semester=%s""",
        (grade_val, result_val, stu_id, subject_name, semester_val))

    try:
        from services.admin_notification_service import admin_notifier
        admin_notifier.notify_admin(
            event_type    = 'marks_submitted',
            faculty_id    = fid,
            faculty_name  = session.get('name', 'Unknown'),
            subject       = subject_name,
            division      = roll_row["division"] if (roll_row and roll_row["division"]) else "Unknown",
            exam_type     = exam_type,
            student_count = 1,
        )
    except Exception:
        pass

    if request.is_json:
        return jsonify({"success": True, "total": total_val, "grade": grade_val, "result": result_val}), 200

    return redirect("/faculty_marks?success=1")

@faculty_extra_bp.route("/faculty_delete_marks", methods=["POST"])
@login_required("faculty")
def faculty_delete_marks():
    exe("DELETE FROM marks WHERE id=%s AND faculty_id=%s",
        (request.form.get("marks_id",""), session["faculty_id"]))
    return redirect("/faculty_marks")

@faculty_extra_bp.route("/faculty/api/students_by_subject")
@login_required("faculty")
def api_students_by_subject():
    subj_name = request.args.get("subject", "").strip()
    s_info = qone("SELECT division, department FROM subjects WHERE name=%s", (subj_name,))
    if not s_info:
        return jsonify([])
    
    div = s_info['division']
    dept = s_info['department']
    
    students = qry("""SELECT id, name, roll FROM students 
                      WHERE (division=%s OR %s='') AND (department=%s OR %s='')
                      ORDER BY name""", (div, div, dept, dept))
    return jsonify([dict(s) for s in students])


@faculty_extra_bp.route("/faculty_notices")
@login_required("faculty")
def faculty_notices():
    fid = session["faculty_id"]
    notices = qry("SELECT * FROM faculty_notices WHERE faculty_id=%s ORDER BY id DESC", (fid,))
    return render_template("faculty/faculty_notices.html", notices=notices)

@faculty_extra_bp.route("/faculty_save_notice", methods=["POST"])
@login_required("faculty")
def faculty_save_notice():
    exe("INSERT INTO faculty_notices(faculty_id,title,message) VALUES(%s,%s,%s)",
        (session["faculty_id"], request.form.get("title",""), request.form.get("message","")))
    return redirect("/faculty_notices?success=1")

@faculty_extra_bp.route("/faculty_delete_notice", methods=["POST"])
@login_required("faculty")
def faculty_delete_notice():
    exe("DELETE FROM faculty_notices WHERE id=%s AND faculty_id=%s",
        (request.form.get("notice_id",""), session["faculty_id"]))
    return redirect("/faculty_notices")

@faculty_extra_bp.route("/faculty_notes")
@login_required("faculty")
def faculty_notes():
    fid = session["faculty_id"]
    f_subject = request.args.get("subject","").strip()
    my_subjects = qry("SELECT name FROM subjects WHERE teacher LIKE %s ORDER BY name", (f"%{session['name']}%",))
    sql = "SELECT fn.*, f.name as faculty_name FROM faculty_notes fn JOIN faculty f ON fn.faculty_id=f.id WHERE fn.faculty_id=%s"
    params = [fid]
    if f_subject: 
        sql += " AND fn.subject=%s"
        params.append(f_subject)
    sql += " ORDER BY fn.id DESC"
    notes = qry(sql, params)
    notes_list = []
    for n in notes:
        d = dict(n)
        if isinstance(d.get("created_at"), (datetime, date)):
            d["created_at"] = d["created_at"].strftime("%Y-%m-%d")
        elif d.get("created_at"):
            d["created_at"] = str(d["created_at"])[:10]
        notes_list.append(d)
    return render_template("faculty/faculty_notes.html", notes=notes_list, my_subjects=my_subjects, f_subject=f_subject)

@faculty_extra_bp.route("/faculty_save_note", methods=["POST"])
@login_required("faculty")
def faculty_save_note():
    import os
    import uuid
    from flask import current_app
    
    attachment = request.files.get("attachment")
    attachment_path = None
    if attachment and attachment.filename:
        filename = attachment.filename
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext in ["xlsx", "xls", "pdf", "jpg", "jpeg", "png"]:
            new_filename = f"{uuid.uuid4().hex}.{ext}"
            upload_dir = os.path.join(current_app.root_path, "static", "uploads")
            os.makedirs(upload_dir, exist_ok=True)
            attachment.save(os.path.join(upload_dir, new_filename))
            attachment_path = f"/static/uploads/{new_filename}"
            
    exe("INSERT INTO faculty_notes(faculty_id,subject,title,content,note_type,attachment_path) VALUES(%s,%s,%s,%s,%s,%s)",
        (session["faculty_id"], request.form.get("subject",""), request.form.get("title",""),
         request.form.get("content",""), request.form.get("note_type","Lecture Note"), attachment_path))
    return redirect("/faculty_notes?success=1")

@faculty_extra_bp.route("/faculty_edit_note", methods=["POST"])
@login_required("faculty")
def faculty_edit_note():
    exe("UPDATE faculty_notes SET title=%s,content=%s,note_type=%s WHERE id=%s AND faculty_id=%s",
        (request.form.get("title",""), request.form.get("content",""), request.form.get("note_type",""),
         request.form.get("note_id",""), session["faculty_id"]))
    return redirect("/faculty_notes")

@faculty_extra_bp.route("/faculty_delete_note", methods=["POST"])
@login_required("faculty")
def faculty_delete_note():
    exe("DELETE FROM faculty_notes WHERE id=%s AND faculty_id=%s",
        (request.form.get("note_id",""), session["faculty_id"]))
    return redirect("/faculty_notes")

@faculty_extra_bp.route("/faculty_timetable")
@login_required("faculty")
def faculty_timetable():
    view    = request.args.get("view","grid")
    f_day   = request.args.get("day","").strip()
    f_type  = request.args.get("slot_type","").strip()
    f_branch = request.args.get("branch","").strip()
    f_year = request.args.get("year","").strip()
    f_div = request.args.get("division","").strip()
    f_subj = request.args.get("subject_id","").strip()

    fac_id  = session.get("faculty_id")

    faculty_name = session.get("name", "")
    master_rows = [dict(e) for e in qry(f"""
        SELECT * FROM timetable 
        WHERE faculty_id = %s 
        OR (teacher IS NOT NULL AND TRIM(REPLACE(teacher, ' ', '')) ILIKE TRIM(REPLACE(%s, ' ', '')))
        ORDER BY {DAY_ORD}, start_time
    """, (fac_id, faculty_name))]
    for m in master_rows:
        m['status'] = 'approved'
        m['teacher'] = faculty_name or m.get('teacher') or 'You'
        m['time_slot'] = m.get('time') or ''
        m['time'] = m.get('time') or ''

    self_rows = [dict(e) for e in qry("SELECT * FROM faculty_timetable WHERE faculty_id = %s AND status != 'approved'", (fac_id,))]
    for s in self_rows:
        s['time'] = s.get('time_slot') or ''
        s['teacher'] = session.get('name', 'You')

    all_entries = master_rows + self_rows
    for e in all_entries: 
        e["time"] = normalize_time(e.get("time",""))

    entries = all_entries
    if f_day:  entries = [e for e in entries if e["day"]==f_day]
    if f_type: entries = [e for e in entries if (e.get("slot_type") or "Theory")==f_type]
    if f_branch: entries = [e for e in entries if e.get("branch")==f_branch]
    if f_year: entries = [e for e in entries if e.get("year")==f_year]
    if f_div: entries = [e for e in entries if e.get("division")==f_div]
    if f_subj: entries = [e for e in entries if str(e.get("subject_id"))==str(f_subj)]

    seen=set(); raw=[]
    for e in all_entries:
        t=e["time"]
        if t and t not in seen: 
            seen.add(t)
            raw.append(t)

    def _sk(ts):
        m=re.match(r"(\d+):(\d+)",ts)
        if not m: return 999
        h=int(m.group(1)); mn=int(m.group(2))
        if h<7: h+=12
        return h*60+mn

    time_slots = sorted(raw, key=_sk)
    grid = {d:{t:[] for t in time_slots} for d in DAYS}
    for e in all_entries:
        if e["day"] in grid and e["time"] in grid[e["day"]]:
            grid[e["day"]][e["time"]].append(e)

    total   = len(all_entries)
    theory  = sum(1 for e in all_entries if (e.get("slot_type") or "Theory")=="Theory")
    lab     = sum(1 for e in all_entries if e.get("slot_type")=="Lab")
    days_active = len(set(e["day"] for e in all_entries))

    # Self-service stats queries
    total_slots_row = qone("SELECT COUNT(*) FROM faculty_timetable WHERE faculty_id=%s AND status='approved'", (fac_id,))
    total_slots = total_slots_row[0] if total_slots_row else 0

    theory_count_row = qone("SELECT COUNT(*) FROM faculty_timetable WHERE faculty_id=%s AND status='approved' AND slot_type='Theory'", (fac_id,))
    theory_count = theory_count_row[0] if theory_count_row else 0

    lab_count_row = qone("SELECT COUNT(*) FROM faculty_timetable WHERE faculty_id=%s AND status='approved' AND slot_type='Lab'", (fac_id,))
    lab_count = lab_count_row[0] if lab_count_row else 0

    active_days_row = qone("SELECT COUNT(DISTINCT day) FROM faculty_timetable WHERE faculty_id=%s AND status='approved'", (fac_id,))
    active_days_val = active_days_row[0] if active_days_row else 0

    time_slots_count_row = qone("SELECT COUNT(DISTINCT time_slot) FROM faculty_timetable WHERE faculty_id=%s AND status='approved'", (fac_id,))
    time_slots_count = time_slots_count_row[0] if time_slots_count_row else 0

    pending_count_row = qone("SELECT COUNT(*) FROM faculty_timetable WHERE faculty_id=%s AND status='pending'", (fac_id,))
    pending_count = pending_count_row[0] if pending_count_row else 0

    draft_count_row = qone("SELECT COUNT(*) FROM faculty_timetable WHERE faculty_id=%s AND status IN ('draft', 'rejected')", (fac_id,))
    draft_count = draft_count_row[0] if draft_count_row else 0

    return render_template("faculty/faculty_timetable.html",
        entries=entries, all_entries=all_entries,
        grid=grid, time_slots=time_slots,
        DAYS=DAYS, view=view, f_day=f_day, f_type=f_type,
        total=total, theory=theory, lab=lab,
        days_active=days_active,
        today_name=date.today().strftime("%A"),
        total_slots=total_slots, theory_count=theory_count,
        lab_count=lab_count, active_days=active_days_val,
        time_slots_count=time_slots_count, pending_count=pending_count,
        draft_count=draft_count,
        DEPARTMENTS=DEPARTMENTS
    )

@faculty_extra_bp.route("/faculty_profile")
@login_required("faculty")
def faculty_profile():
    profile = qone("SELECT * FROM faculty WHERE id=%s", (session["faculty_id"],))
    return render_template("faculty/faculty_profile.html", profile=dict(profile) if profile else {})

@faculty_extra_bp.route("/faculty_update_profile", methods=["POST"])
@login_required("faculty")
def faculty_update_profile():
    fid = session["faculty_id"]
    name = request.form.get("name","").strip()
    phone= request.form.get("phone","").strip()
    qual = request.form.get("qualification","").strip()
    current_pw = request.form.get("current_password","").strip()
    new_pw = request.form.get("new_password","").strip() or request.form.get("password","").strip()
    confirm_pw = request.form.get("confirm_password","").strip() or new_pw
    
    if name:
        exe("UPDATE faculty SET name=%s,phone=%s,qualification=%s WHERE id=%s", (name,phone,qual,fid))
        session["name"] = name
        
    if new_pw:
        row = qone("SELECT password FROM faculty WHERE id=%s", (fid,))
        err = validate_password_change(row["password"] if row else "", current_pw, new_pw, confirm_pw)
        if err:
            return redirect(f"/faculty_profile?error={err}")
        exe("UPDATE faculty SET password=%s WHERE id=%s", (hash_password(new_pw), fid))
        
    return redirect("/faculty_profile?success=1")

@faculty_extra_bp.route("/faculty_cumulative")
@login_required("faculty")
def faculty_cumulative():
    fid     = session["faculty_id"]
    student = request.args.get("student","").strip()

    data = _cumulative_marks_data_aggregated(dept="", student_filter=student, faculty_id=fid)

    return render_template("cumulative/cumulative_marks.html",
        data=data, dept="", student=student,
        DEPARTMENTS=DEPARTMENTS,
        total_students=len(data),
        role="faculty")

@faculty_extra_bp.route("/import_marks_v2", methods=["POST"])
@login_required("faculty")
def import_marks_v2():
    f = request.files.get("file")
    if not f: return redirect("/faculty_marks")
    fid = session["faculty_id"]
    wb = load_workbook(f, data_only=True); ws = wb.active
    added = 0
    hdr_row = 1
    for i in range(1, min(ws.max_row+1, 10)):
        vals = [str(ws.cell(i,c).value or "").lower() for c in range(1,8)]
        if any("name" in v or "student" in v or "mark" in v for v in vals): hdr_row=i; break
    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None
    cn=gcol(["name","student"]); cr=gcol(["roll"])
    cs=gcol(["subject"]); cm=gcol(["marks","obtained"])
    ct=gcol(["total"]); ce=gcol(["exam","type"]); cd=gcol(["date"])
    if not cn or not cm: return redirect("/faculty_marks?error=bad_format")
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        name = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        if not name: continue
        roll    = str(row[cr-1] or "").strip() if cr and cr <= len(row) else ""
        subj    = str(row[cs-1] or "").strip() if cs and cs <= len(row) else ""
        marks_v = row[cm-1] if cm <= len(row) else 0
        total_v = row[ct-1] if ct and ct <= len(row) else 100
        exam_t  = str(row[ce-1] or "Unit Test 1").strip() if ce and ce <= len(row) else "Unit Test 1"
        d_raw   = row[cd-1] if cd and cd <= len(row) else None
        att_date= normalise_date(d_raw) if d_raw else today_str()
        try:
            marks_f = float(marks_v or 0)
            total_f = float(total_v or 100)
        except: continue
        stu_id = resolve_student_id(name, roll)
        sr = qone("SELECT roll,department FROM students WHERE id=%s", (stu_id,)) if stu_id else None
        if sr:
            roll = sr["roll"] or roll
            dept = sr["department"]
        else:
            dept = ""

        exe(
            "INSERT INTO marks(faculty_id,student_id,student_name,roll,subject,department,marks,total,exam_type,date,remarks) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (fid, stu_id, name, roll, subj, dept, marks_f, total_f, exam_t, att_date, ""),
        )
        added += 1
    return redirect(f"/faculty_marks?imported={added}")

@faculty_extra_bp.route("/export_marks_excel")
@login_required("faculty")
def export_marks_excel():
    fid     = session["faculty_id"]
    subject = request.args.get("subject","").strip()
    student = request.args.get("student","").strip()

    sql = "SELECT * FROM marks WHERE faculty_id=%s"
    params = [fid]
    if subject: sql += " AND subject=%s"; params.append(subject)
    if student:
        sid = resolve_student_id(student)
        sql += f" AND ({marks_match_student_sql()})"
        params.extend(marks_match_student_params(sid, student))
    sql += " ORDER BY student_name, subject, exam_type"
    rows = qry(sql, params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Marks"

    hdrs = ["Student Name","Roll","Subject","Dept","Exam Type","Marks","Total","Percentage","Grade","Date"]
    for c,h in enumerate(hdrs,1):
        cell = ws.cell(1,c,h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.font = Font(bold=True, color="FFFFFF")

    for r,row in enumerate(rows,2):
        p = round(row["marks"]/row["total"]*100,1) if row["total"] else 0
        g = grade(row["marks"], row["total"])
        vals = [row["student_name"],row["roll"],row["subject"],row["department"],
                row["exam_type"],row["marks"],row["total"],f"{p}%",g,row["date"]]
        for c,v in enumerate(vals,1):
            ws.cell(r,c,v)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or ""))+4, 14)

    fname = f"marks{'_'+subject if subject else ''}{'_'+student if student else ''}.xlsx"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@faculty_extra_bp.route("/faculty_results")
@login_required("faculty")
def faculty_results():
    fid = session["faculty_id"]
    my_subjects = qry("SELECT name FROM subjects WHERE teacher LIKE %s ORDER BY name",
                      (f"%{session['name']}%",))
    results = qry("SELECT * FROM results WHERE faculty_id=%s ORDER BY id DESC", (fid,))
    students_list = qry("SELECT name,roll FROM students ORDER BY name")
    return render_template("faculty/faculty_results.html",
        results=results, my_subjects=my_subjects,
        students=students_list, today=today_str(), SEMESTERS=SEMESTERS)


@faculty_extra_bp.route("/faculty/results/add", methods=["GET"])
@login_required("faculty")
def faculty_add_result_page():
    my_subjects = qry("SELECT name FROM subjects WHERE teacher LIKE %s ORDER BY name",
                      (f"%{session['name']}%",))
    students_list = qry("SELECT name,roll FROM students ORDER BY name")
    return render_template("faculty/add_result.html",
        my_subjects=my_subjects, students=students_list,
        today=today_str(), SEMESTERS=SEMESTERS)


@faculty_extra_bp.route("/faculty/results/edit/<int:result_id>", methods=["GET"])
@login_required("faculty")
def faculty_edit_result_page(result_id):
    fid = session["faculty_id"]
    res_row = qone("SELECT * FROM results WHERE id=%s AND faculty_id=%s", (result_id, fid))
    if not res_row:
        from flask import flash
        flash("Result record not found.", "error")
        return redirect("/faculty_results")
    my_subjects = qry("SELECT name FROM subjects WHERE teacher LIKE %s ORDER BY name",
                      (f"%{session['name']}%",))
    students_list = qry("SELECT name,roll FROM students ORDER BY name")
    return render_template("faculty/edit_result.html",
        result=res_row, my_subjects=my_subjects, students=students_list,
        SEMESTERS=SEMESTERS)

@faculty_extra_bp.route("/faculty_save_result", methods=["POST"])
@login_required("faculty")
def faculty_save_result():
    fid = session["faculty_id"]
    student_name = request.form.get("student_name","").strip()
    roll_row = qone("SELECT roll,department,year FROM students WHERE name=%s", (student_name,))
    roll = roll_row["roll"] if roll_row else ""
    dept = roll_row["department"] if roll_row else ""
    yr   = roll_row["year"] if roll_row else ""

    assignment_m   = min(float(request.form.get("assignment_marks",   0) or 0), 5.0)
    attendance_m   = min(float(request.form.get("attendance_marks",   0) or 0), 5.0)
    teaching_m     = min(float(request.form.get("teaching_assessment",0) or 0), 10.0)
    ut_m           = min(float(request.form.get("ut_marks",           0) or 0), 20.0)
    mse_m          = min(float(request.form.get("mse_marks",          0) or 0), 20.0)

    marks_val = assignment_m + attendance_m + teaching_m + ut_m + mse_m
    if marks_val == 0:
        marks_val = min(float(request.form.get("marks", 0) or 0), 60.0)

    marks_val = min(marks_val, 60.0)
    total_val = 60.0
    g = request.form.get("grade", "").strip() or grade(marks_val, total_val)
    res = "Pass" if pct(marks_val, total_val) >= 35 else "Fail"

    exe("""INSERT INTO results(student_name,roll,department,year,semester,subject,
                               marks,total,exam_type,grade,result,faculty_id,published,
                               assignment_marks,attendance_marks,teaching_assessment,
                               ut_marks,mse_marks)
           VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,%s,%s,%s,%s,%s)""",
        (student_name, roll, dept, yr,
         request.form.get("semester","IV"),
         request.form.get("subject",""),
         marks_val, total_val,
         request.form.get("exam_type","Semester Exam"),
         g, res, fid,
         assignment_m, attendance_m, teaching_m, ut_m, mse_m))
    return redirect("/faculty_results?success=1")

@faculty_extra_bp.route("/faculty_edit_result", methods=["POST"])
@login_required("faculty")
def faculty_edit_result():
    rid = request.form.get("result_id", "")
    fid = session["faculty_id"]
    assignment_m = min(float(request.form.get("assignment_marks", 0) or 0), 5.0)
    attendance_m = min(float(request.form.get("attendance_marks", 0) or 0), 5.0)
    teaching_m = min(float(request.form.get("teacher_assessment", 0) or 0), 10.0)
    ut_m = min(float(request.form.get("ut_marks", 0) or 0), 20.0)
    mse_m = min(float(request.form.get("mse_marks", 0) or 0), 20.0)
    tw_m = float(request.form.get("tw_marks", 0) or 0)
    pr_or_m = float(request.form.get("pr_or_marks", 0) or 0)

    marks_val = assignment_m + attendance_m + teaching_m + ut_m + mse_m + tw_m + pr_or_m
    if marks_val == 0:
        marks_val = min(float(request.form.get("marks", 0) or 0), 60.0)

    marks_val = min(marks_val, 60.0)
    total_val = 60.0
    g = request.form.get("grade", "").strip() or grade(marks_val, total_val)
    res = "Pass" if pct(marks_val, total_val) >= 35 else "Fail"
    exe("""UPDATE results SET semester=%s,subject=%s,marks=%s,total=%s,
           exam_type=%s,grade=%s,result=%s, assignment_marks=%s, attendance_marks=%s,
           teaching_assessment=%s, ut_marks=%s, mse_marks=%s, tw_marks=%s, pr_or_marks=%s
           WHERE id=%s AND faculty_id=%s""",
        (request.form.get("semester", ""), request.form.get("subject", ""),
         marks_val, total_val, request.form.get("exam_type", ""),
         g, res, assignment_m, attendance_m, teaching_m, ut_m, mse_m, tw_m, pr_or_m,
         rid, fid))
    return redirect("/faculty_results?updated=1")

@faculty_extra_bp.route("/faculty_delete_result", methods=["POST"])
@login_required("faculty")
def faculty_delete_result():
    exe("DELETE FROM results WHERE id=%s AND faculty_id=%s",
        (request.form.get("result_id",""), session["faculty_id"]))
    return redirect("/faculty_results")

@faculty_extra_bp.route("/cumulative_marks")
@login_required("admin")
def cumulative_marks():
    dept    = request.args.get("dept","").strip()
    student = request.args.get("student","").strip()

    data = _cumulative_marks_data_aggregated(dept=dept, student_filter=student, faculty_id=None)

    return render_template("cumulative/cumulative_marks.html",
        data=data, dept=dept, student=student,
        DEPARTMENTS=DEPARTMENTS,
        total_students=len(data),
        role="admin")

@faculty_extra_bp.route("/export_cumulative_excel")
@login_required("admin")
def export_cumulative_excel():
    dept    = request.args.get("dept","").strip()
    student = request.args.get("student","").strip()

    data = _cumulative_marks_data_aggregated(dept=dept, student_filter=student, faculty_id=None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cumulative Marks"

    hdrs = ["Student Name", "Roll", "Dept", "Obtained", "Total", "Percentage", "Grade", "Exams"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.font = Font(bold=True, color="FFFFFF")

    for r, row in enumerate(data, 2):
        vals = [row["name"], row["roll"], row["dept"], row["obtained"], row["total"], f"{row['pct']}%", row["grade"], row["exams"]]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")) + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="cumulative_marks.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Assessment Module Routes ───────────────────────────────────

@faculty_extra_bp.route("/faculty/assessments")
@login_required("faculty")
def faculty_assessments():
    fid = session.get("faculty_id")
    # Fetch assigned subjects/divisions/departments
    my_assignments = qry("SELECT DISTINCT subject_name, division, department FROM faculty_subject_assignments WHERE faculty_id = %s", (fid,))
    if not my_assignments:
        my_assignments = qry("SELECT DISTINCT name as subject_name, division, department FROM subjects WHERE teacher LIKE %s", (f"%{session['name']}%",))

    subjects_list = sorted(list(set(a["subject_name"] for a in my_assignments if a["subject_name"])))
    divisions_list = sorted(list(set(a["division"] for a in my_assignments if a["division"])))
    departments_list = sorted(list(set(a["department"] for a in my_assignments if a["department"])))

    selected_subject = request.args.get("subject", "").strip()
    selected_div = request.args.get("division", "").strip()
    selected_dept = request.args.get("department", "").strip()

    students = []
    assessments_map = {}
    if selected_subject:
        sql = "SELECT * FROM students WHERE 1=1"
        params = []
        if selected_div:
            sql += " AND division = %s"
            params.append(selected_div)
        if selected_dept:
            sql += " AND department = %s"
            params.append(selected_dept)
        sql += " ORDER BY name"
        students = qry(sql, params)

        raw_assessments = qry("SELECT * FROM assessments WHERE subject = %s", (selected_subject,))
        assessments_map = {r["student_id"]: dict(r) for r in raw_assessments}

    return render_template(
        "faculty/faculty_assessments.html",
        my_assignments=my_assignments,
        subjects_list=subjects_list,
        divisions_list=divisions_list,
        departments_list=departments_list,
        students=students,
        assessments_map=assessments_map,
        selected_subject=selected_subject,
        selected_div=selected_div,
        selected_dept=selected_dept
    )

@faculty_extra_bp.route("/faculty/assessments/edit/<int:student_id>/<path:subject>")
@login_required("faculty")
def edit_assessment(student_id, subject):
    student = qone("SELECT * FROM students WHERE id = %s", (student_id,))
    if not student:
        return redirect("/faculty/assessments")

    assessment = qone("SELECT * FROM assessments WHERE student_id = %s AND subject = %s", (student_id, subject))
    if assessment:
        assessment = dict(assessment)
    else:
        assessment = {}

    return render_template(
        "faculty/edit_assessment.html",
        student=dict(student),
        subject=subject,
        assessment=assessment
    )

@faculty_extra_bp.route("/faculty/assessments/save", methods=["POST"])
@login_required("faculty")
def save_assessment():
    student_id = int(request.form.get("student_id"))
    subject = request.form.get("subject", "").strip()
    selected_div = request.form.get("selected_div", "").strip()
    selected_dept = request.form.get("selected_dept", "").strip()

    assignment_1 = request.form.get("assignment_1", "").strip()
    assignment_2 = request.form.get("assignment_2", "").strip()
    assignment_3 = request.form.get("assignment_3", "").strip()
    assignment_4 = request.form.get("assignment_4", "").strip()
    assignment_5 = request.form.get("assignment_5", "").strip()

    paper_q1 = request.form.get("paper_q1", "").strip()
    paper_q2 = request.form.get("paper_q2", "").strip()
    paper_q3 = request.form.get("paper_q3", "").strip()
    paper_q4 = request.form.get("paper_q4", "").strip()
    patent_publication = request.form.get("patent_publication", "").strip()
    copyright = request.form.get("copyright", "").strip()

    project_review_1 = request.form.get("project_review_1", "").strip()
    project_review_2 = request.form.get("project_review_2", "").strip()
    implementation_documentation = request.form.get("implementation_documentation", "").strip()
    remark = request.form.get("remark", "").strip()

    existing = qone("SELECT id FROM assessments WHERE student_id = %s AND subject = %s", (student_id, subject))
    if existing:
        exe("""
            UPDATE assessments SET 
                assignment_1 = %s, assignment_2 = %s, assignment_3 = %s, assignment_4 = %s, assignment_5 = %s,
                paper_q1 = %s, paper_q2 = %s, paper_q3 = %s, paper_q4 = %s,
                patent_publication = %s, copyright = %s,
                project_review_1 = %s, project_review_2 = %s,
                implementation_documentation = %s, remark = %s,
                faculty_id = %s, updated_at = %s
            WHERE id = %s
        """, (
            assignment_1, assignment_2, assignment_3, assignment_4, assignment_5,
            paper_q1, paper_q2, paper_q3, paper_q4,
            patent_publication, copyright,
            project_review_1, project_review_2,
            implementation_documentation, remark,
            session["faculty_id"], datetime.utcnow(), existing["id"]
        ))
    else:
        exe("""
            INSERT INTO assessments (
                student_id, subject, faculty_id,
                assignment_1, assignment_2, assignment_3, assignment_4, assignment_5,
                paper_q1, paper_q2, paper_q3, paper_q4,
                patent_publication, copyright,
                project_review_1, project_review_2,
                implementation_documentation, remark
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            student_id, subject, session["faculty_id"],
            assignment_1, assignment_2, assignment_3, assignment_4, assignment_5,
            paper_q1, paper_q2, paper_q3, paper_q4,
            patent_publication, copyright,
            project_review_1, project_review_2,
            implementation_documentation, remark
        ))

    from flask import flash
    flash("Assessment details updated successfully.", "success")
    return redirect(url_for("faculty_extra.faculty_assessments", subject=subject, division=selected_div, department=selected_dept))


@faculty_extra_bp.route("/faculty/marks/publish-ut", methods=["POST"])
@login_required("faculty")
def publish_ut():
    data = request.json if request.is_json else request.form
    subject_code = data.get("subject_code")
    division = data.get("division")
    semester = data.get("semester")
    
    if not (subject_code and division and semester):
        return jsonify({"errors": "Missing required fields"}), 400
        
    cur = exe("""
        UPDATE marks
        SET ut_published = TRUE
        WHERE ut_marks IS NOT NULL
          AND (subject_code = %s OR subject = %s)
          AND semester = %s
          AND student_id IN (
              SELECT id FROM students WHERE division = %s
          )
    """, (subject_code, subject_code, semester, division))
    
    return jsonify({"updated": cur.rowcount}), 200


@faculty_extra_bp.route("/faculty/marks/publish-mse", methods=["POST"])
@login_required("faculty")
def publish_mse():
    data = request.json if request.is_json else request.form
    subject_code = data.get("subject_code")
    division = data.get("division")
    semester = data.get("semester")
    
    if not (subject_code and division and semester):
        return jsonify({"errors": "Missing required fields"}), 400
        
    cur = exe("""
        UPDATE marks
        SET mse_published = TRUE
        WHERE mse_marks IS NOT NULL
          AND (subject_code = %s OR subject = %s)
          AND semester = %s
          AND student_id IN (
              SELECT id FROM students WHERE division = %s
          )
    """, (subject_code, subject_code, semester, division))
    
    return jsonify({"updated": cur.rowcount}), 200


@faculty_extra_bp.route("/admin/marks/publish-result", methods=["POST"])
@login_required("admin")
def publish_result():
    data = request.json if request.is_json else request.form
    subject_code = data.get("subject_code")
    division = data.get("division")
    semester = data.get("semester")
    
    if not (subject_code and division and semester):
        return jsonify({"errors": "Missing required fields"}), 400
        
    cur = exe("""
        UPDATE marks
        SET result_published = TRUE
        WHERE (subject_code = %s OR subject = %s)
          AND semester = %s
          AND student_id IN (
              SELECT id FROM students WHERE division = %s
          )
    """, (subject_code, subject_code, semester, division))
    
    return jsonify({"updated": cur.rowcount}), 200


@faculty_extra_bp.route("/faculty_import_results", methods=["POST"])
@login_required("faculty")
def faculty_import_results():
    f = request.files.get("file")
    if not f:
        from flask import flash
        flash("No file selected", "error")
        return redirect("/faculty_results")
    fid = session["faculty_id"]
    try:
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        added = 0
        hdr_row = 1
        for i in range(1, min(ws.max_row+1, 10)):
            vals = [str(ws.cell(i,c).value or "").lower() for c in range(1, min(ws.max_column+1, 15))]
            if any("name" in v or "student" in v or "marks" in v or "subject" in v for v in vals):
                hdr_row = i
                break
        hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1, ws.max_column+1)]
        def gcol(kws):
            for k in kws:
                for idx, h in enumerate(hdrs):
                    if k in h: return idx + 1
            return None
        cn = gcol(["name", "student"])
        cr = gcol(["roll"])
        cs = gcol(["subject"])
        csem = gcol(["semester", "sem"])
        c_exam = gcol(["exam type", "exam"])
        c_assign = gcol(["assignment"])
        c_attend = gcol(["attendance"])
        c_ta = gcol(["teacher assessment", "ta", "teaching assessment", "assess"])
        c_ut = gcol(["ut", "unit test"])
        c_mse = gcol(["mse", "mid-sem", "mid sem"])
        c_tw = gcol(["term work", "tw"])
        c_pr = gcol(["practical", "pr", "or"])
        if not cn or not cs:
            from flask import flash
            flash("Invalid Excel format. Name and Subject columns are required.", "error")
            return redirect("/faculty_results")
        for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
            if not row or len(row) < cn: continue
            name = str(row[cn-1] or "").strip()
            if not name or name.lower() in ("name", "student name", "student"): continue
            roll = str(row[cr-1] or "").strip() if cr and cr <= len(row) else ""
            subj = str(row[cs-1] or "").strip() if cs and cs <= len(row) else ""
            sem = str(row[csem-1] or "IV").strip() if csem and csem <= len(row) else "IV"
            exam_t = str(row[c_exam-1] or "Semester Exam").strip() if c_exam and c_exam <= len(row) else "Semester Exam"
            def get_val(col_idx):
                if col_idx and col_idx <= len(row) and row[col_idx-1] is not None:
                    try: return float(row[col_idx-1])
                    except: return 0.0
                return 0.0
            assign = get_val(c_assign)
            attend = get_val(c_attend)
            ta = get_val(c_ta)
            ut = get_val(c_ut)
            mse = get_val(c_mse)
            tw = get_val(c_tw)
            pr = get_val(c_pr)
            stu_id = resolve_student_id(name, roll)
            sr = qone("SELECT roll, department, year FROM students WHERE id=%s", (stu_id,)) if stu_id else None
            dept = sr["department"] if sr else ""
            yr = sr["year"] if sr else ""
            roll = sr["roll"] if sr else roll
            marks_val = min(assign + attend + ta + ut + mse + tw + pr, 60.0)
            total_val = 60.0
            g_letter = grade(marks_val, total_val)
            res_val = "Pass" if pct(marks_val, total_val) >= 35 else "Fail"
            existing = qone("SELECT id FROM results WHERE student_name=%s AND roll=%s AND subject=%s AND exam_type=%s AND semester=%s",
                            (name, roll, subj, exam_t, sem))
            if existing:
                exe("""UPDATE results SET marks=%s, total=%s, grade=%s, result=%s,
                                         assignment_marks=%s, attendance_marks=%s, teaching_assessment=%s,
                                         ut_marks=%s, mse_marks=%s, tw_marks=%s, pr_or_marks=%s
                       WHERE id=%s""",
                    (marks_val, total_val, g_letter, res_val, assign, attend, ta, ut, mse, tw, pr, existing["id"]))
            else:
                exe("""INSERT INTO results(student_name, roll, department, year, semester, subject,
                                           marks, total, exam_type, grade, result, faculty_id, published,
                                           assignment_marks, attendance_marks, teaching_assessment,
                                           ut_marks, mse_marks, tw_marks, pr_or_marks)
                       VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, %s, %s, %s, %s, %s, %s, %s)""",
                    (name, roll, dept, yr, sem, subj, marks_val, total_val, exam_t, g_letter, res_val, fid,
                     assign, attend, ta, ut, mse, tw, pr))
            added += 1
        from flask import flash
        flash(f"Successfully imported {added} results.", "success")
    except Exception as e:
        import traceback
        from flask import current_app, flash
        current_app.logger.error(f"Faculty results import error: {traceback.format_exc()}")
        flash(f"Error importing Excel: {str(e)}", "error")
    return redirect("/faculty_results")


@faculty_extra_bp.route("/faculty_export_results")
@login_required("faculty")
def faculty_export_results():
    fid = session["faculty_id"]
    subject = request.args.get("subject","").strip()
    semester = request.args.get("semester","").strip()
    
    sql = "SELECT * FROM results WHERE faculty_id=%s"
    params = [fid]
    if subject:
        sql += " AND subject=%s"
        params.append(subject)
    if semester:
        sql += " AND semester=%s"
        params.append(semester)
        
    sql += " ORDER BY student_name, subject, exam_type"
    rows = qry(sql, params)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    hdrs = ["Student Name", "Roll", "Subject", "Semester", "Exam Type",
            "Assignment", "Attendance", "Teacher Assessment", "UT", "MSE", "TW", "PR OR",
            "Aggregate", "Total", "Percentage", "Grade", "Result"]
            
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")
        
    for r, row in enumerate(rows, 2):
        obtained = float(row["marks"] or 0.0)
        tot = float(row["total"] or 60.0)
        p = round(obtained / tot * 100, 1) if tot else 0
        
        vals = [
            row["student_name"], row["roll"], row["subject"], row["semester"], row["exam_type"],
            row["assignment_marks"], row["attendance_marks"], row["teaching_assessment"],
            row["ut_marks"], row["mse_marks"], row["tw_marks"], row["pr_or_marks"],
            obtained, tot, f"{p}%", row["grade"], row["result"]
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)
            
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")) + 4, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    fname = f"results{'_'+subject if subject else ''}{'_'+semester if semester else ''}.xlsx"
    return send_file(output, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



