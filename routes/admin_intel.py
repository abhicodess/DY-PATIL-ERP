from flask import Blueprint, render_template, request, session, redirect, url_for, flash
from utils.pg_wrapper import qry, qone
from functools import wraps
import logging

admin_intel_bp = Blueprint('admin_intel', __name__)

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Administrator access required.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

@admin_intel_bp.route("/admin/intelligence")
@admin_required
def admin_attendance_intelligence():
    """
    Main Analytics Dashboard - Attendance Intelligence
    """
    from services.attendance_ai import AttendanceAI
    from config import Config
    DEPARTMENTS, DIVISIONS, YEARS = Config.DEPARTMENTS, Config.DIVISIONS, Config.YEARS
    
    dept = request.args.get("dept", "")
    division = request.args.get("division", "")
    year = request.args.get("year", "")
    
    # Fetch Real AI Insights
    insights = {
        "risk_profiles": AttendanceAI.get_risk_profiles(dept=dept, year=year, division=division),
        "subject_anomalies": AttendanceAI.get_department_comparison()[:5],
        "attendance_trend": AttendanceAI.get_weekly_trend(dept=dept, division=division),
        "summary": AttendanceAI.get_insights_summary()
    }
    
    return render_template("admin/attendance_intelligence.html", 
                           dept=dept, 
                           division=division,
                           year=year,
                           insights=insights,
                           DEPARTMENTS=DEPARTMENTS,
                           DIVISIONS=DIVISIONS,
                           YEARS=YEARS,
                           is_ai_enabled=True)

@admin_intel_bp.route("/admin/session-detail/<int:sid>")
@admin_required
def admin_session_detail(sid):
    """
    Deep Audit for a specific session.
    """
    session_info = qone("""
        SELECT 
            MIN(a.id) as id,
            a.date as lecture_date,
            a.subject,
            a.time_slot as start_time,
            a.division,
            a.semester as year,
            f.name as faculty_name
        FROM attendance a
        LEFT JOIN faculty f ON f.id::TEXT = a.faculty OR f.name = a.faculty
        WHERE a.id = %s
        GROUP BY a.date, a.subject, a.time_slot, a.division, a.semester, f.name
    """, (sid,))
    
    if not session_info:
        flash("Audit session not found", "error")
        return redirect(url_for("admin_intel.admin_faculty_logs"))
        
    records = qry("""
        SELECT a2.id as att_id, a2.status, s.name as student_name, s.roll, s.id as student_id
        FROM attendance a2
        JOIN students s ON a2.student_id = s.id
        WHERE a2.date = (SELECT date FROM attendance WHERE id = %s)
          AND a2.subject = (SELECT subject FROM attendance WHERE id = %s)
          AND a2.time_slot = (SELECT time_slot FROM attendance WHERE id = %s)
        ORDER BY s.roll
    """, (sid, sid, sid))
    
    summary = {
        "present": sum(1 for r in records if str(r["status"]).lower() == "present"),
        "absent": sum(1 for r in records if str(r["status"]).lower() == "absent"),
        "total": len(records)
    }
    
    return render_template("admin/attendance_session_detail.html", 
                           session=session_info, records=records, summary=summary)

@admin_intel_bp.route("/admin/faculty-logs")
@admin_required
def admin_faculty_logs():
    """Review logs for all faculty sessions."""
    from config import Config
    DEPARTMENTS, DIVISIONS, YEARS = Config.DEPARTMENTS, Config.DIVISIONS, Config.YEARS
    try:
        f_id = request.args.get("faculty_id")
        dept = request.args.get("dept")
        
        filter_sql = ""
        params = []
        if f_id: filter_sql += " AND f.id=%s"; params.append(f_id)
        if dept: filter_sql += " AND f.department=%s"; params.append(dept)
        
        params.append(100) # For LIMIT
        
        sql = """
            SELECT 
                MIN(a.id) as id, 
                f.name as faculty_name, 
                f.department as faculty_dept,
                f.department as branch,
                a.subject,
                a.division,
                a.date as lecture_date,
                a.time_slot,
                a.time_slot as start_time,
                a.semester as year,
                'Submitted' as status,
                'Theory' as method,
                NULL as created_at,
                COUNT(a.id) as total_students,
                SUM(CASE WHEN a.status = 'Present' THEN 1 ELSE 0 END) as present,
                SUM(CASE WHEN a.status = 'Absent' THEN 1 ELSE 0 END) as absent
            FROM attendance a
            LEFT JOIN faculty f ON f.id::TEXT = a.faculty OR f.name = a.faculty
            WHERE 1=1
            {filters}
            GROUP BY f.id, f.name, f.department, a.subject, a.division, a.date, a.time_slot, a.semester
            ORDER BY a.date DESC LIMIT %s
        """.format(filters=filter_sql)  # nosec B608 - filter_sql is built from safe hardcoded literals and parameters
        
        sessions_data = qry(sql, params)
        faculty_list = qry("SELECT id, name FROM faculty ORDER BY name")
        
        return render_template("admin/faculty_sessions.html", 
                               sessions=sessions_data, faculty_list=faculty_list,
                               DEPARTMENTS=DEPARTMENTS, DIVISIONS=DIVISIONS, YEARS=YEARS)
    except Exception as e:
        logging.error("Admin Faculty Logs Error: %s", e, exc_info=True)
        return render_template("errors/500.html"), 500

@admin_intel_bp.route("/shortage_report")
@admin_required
def shortage_report():
    from config import Config
    DEPARTMENTS, DIVISIONS, YEARS = Config.DEPARTMENTS, Config.DIVISIONS, Config.YEARS
    threshold_pct = int(request.args.get("threshold", "75"))
    dept = request.args.get("department", "").strip()
    div = request.args.get("division", "").strip()
    t = threshold_pct / 100.0
    
    sql = f"""
        WITH daily_stats AS (
            SELECT student_id,
                   COUNT(id) as d_total,
                   SUM(CASE WHEN status ILIKE 'Present' THEN 1 ELSE 0 END) as d_attended
            FROM attendance
            GROUP BY student_id
        ),
        summary_stats AS (
            SELECT student_id,
                   SUM(attended) as s_attended,
                   SUM(total) as s_total
            FROM attendance_summary
            GROUP BY student_id
        ),
        student_stats AS (
            SELECT s.id, s.name, s.roll, s.department as dept, s.division,
                   CASE 
                       WHEN COALESCE(sum_s.s_total, 0) > 0 THEN COALESCE(sum_s.s_attended, 0)
                       ELSE COALESCE(day_s.d_attended, 0)
                   END as attended,
                   CASE 
                       WHEN COALESCE(sum_s.s_total, 0) > 0 THEN COALESCE(sum_s.s_total, 0)
                       ELSE COALESCE(day_s.d_total, 0)
                   END as total
            FROM students s
            LEFT JOIN daily_stats day_s ON s.id = day_s.student_id
            LEFT JOIN summary_stats sum_s ON s.id = sum_s.student_id
            WHERE (%s='' OR s.department=%s) AND (%s='' OR s.division=%s)
        )
        SELECT *,
               ROUND(CASE WHEN total > 0 THEN (attended * 100.0 / total) ELSE 0 END, 1) as pct,
               CEIL(GREATEST(0, ({t} * total - attended) / (1 - {t} + 0.00001))) as shortage,
               FLOOR(GREATEST(0, (attended - {t} * total) / ({t} + 0.00001))) as can_miss
        FROM student_stats
        WHERE (CASE WHEN total > 0 THEN (attended * 100.0 / total) ELSE 0 END) < %s
        ORDER BY pct ASC, roll ASC
    """  # nosec B608 - t is a sanitized float constructed from an int cast
    defaulters = qry(sql, (dept, dept, div, div, threshold_pct))
    
    return render_template("attendance/shortage_report.html", 
                           defaulters=defaulters, 
                           threshold=threshold_pct, 
                           department=dept, 
                           division=div,
                           DEPARTMENTS=DEPARTMENTS, DIVISIONS=DIVISIONS, YEARS=YEARS)

@admin_intel_bp.route("/export_shortage_excel")
@admin_required
def export_shortage_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    
    threshold_pct = int(request.args.get("threshold", "75"))
    dept = request.args.get("department", "").strip()
    div = request.args.get("division", "").strip()
    
    sql = """
        WITH daily_stats AS (
            SELECT student_id,
                   COUNT(id) as d_total,
                   SUM(CASE WHEN status ILIKE 'Present' THEN 1 ELSE 0 END) as d_attended
            FROM attendance
            GROUP BY student_id
        ),
        summary_stats AS (
            SELECT student_id,
                   SUM(attended) as s_attended,
                   SUM(total) as s_total
            FROM attendance_summary
            GROUP BY student_id
        ),
        student_stats AS (
            SELECT s.id, s.name, s.roll, s.department as dept, s.division, s.year,
                   CASE 
                       WHEN COALESCE(sum_s.s_total, 0) > 0 THEN COALESCE(sum_s.s_attended, 0)
                       ELSE COALESCE(day_s.d_attended, 0)
                   END as attended,
                   CASE 
                       WHEN COALESCE(sum_s.s_total, 0) > 0 THEN COALESCE(sum_s.s_total, 0)
                       ELSE COALESCE(day_s.d_total, 0)
                   END as total
            FROM students s
            LEFT JOIN daily_stats day_s ON s.id = day_s.student_id
            LEFT JOIN summary_stats sum_s ON s.id = sum_s.student_id
            WHERE (%s='' OR s.department=%s) AND (%s='' OR s.division=%s)
        )
        SELECT *,
               ROUND(CASE WHEN total > 0 THEN (attended * 100.0 / total) ELSE 0 END, 1) as pct
        FROM student_stats
        WHERE (CASE WHEN total > 0 THEN (attended * 100.0 / total) ELSE 0 END) < %s
        ORDER BY pct ASC, roll ASC
    """
    rows = qry(sql, (dept, dept, div, div, threshold_pct))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Shortage Report"
    
    headers = ["Student", "Roll", "Department", "Division", "Year", "Total", "Present", "Absent", "Attendance %", "Threshold"]
    for c, header in enumerate(headers, 1):
        cell = ws.cell(1, c, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        
    out_row = 2
    for row in rows:
        total = int(row["total"] or 0)
        present = int(row["attended"] or 0)
        percentage = float(row["pct"] or 0.0)
        values = [
            row["name"], row["roll"], row["dept"], row["division"], row["year"],
            total, present, total - present, f"{percentage}%", f"{threshold_pct}%",
        ]
        for c, value in enumerate(values, 1):
            ws.cell(out_row, c, value)
        out_row += 1
        
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")) + 4, 14)
        
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="shortage_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@admin_intel_bp.route("/trigger_shortage_alert", methods=["POST"])
@admin_required
def trigger_shortage_alert():
    from services.parent_notification_service import ParentNotificationService
    
    threshold_pct = int(request.form.get("threshold", "75"))
    dept = request.form.get("department", "").strip()
    div = request.form.get("division", "").strip()
    
    sql = """
        WITH daily_stats AS (
            SELECT student_id,
                   COUNT(id) as d_total,
                   SUM(CASE WHEN status ILIKE 'Present' THEN 1 ELSE 0 END) as d_attended
            FROM attendance
            GROUP BY student_id
        ),
        summary_stats AS (
            SELECT student_id,
                   SUM(attended) as s_attended,
                   SUM(total) as s_total
            FROM attendance_summary
            GROUP BY student_id
        ),
        student_stats AS (
            SELECT s.id, s.name, s.department as dept, s.division,
                   CASE 
                       WHEN COALESCE(sum_s.s_total, 0) > 0 THEN COALESCE(sum_s.s_attended, 0)
                       ELSE COALESCE(day_s.d_attended, 0)
                   END as attended,
                   CASE 
                       WHEN COALESCE(sum_s.s_total, 0) > 0 THEN COALESCE(sum_s.s_total, 0)
                       ELSE COALESCE(day_s.d_total, 0)
                   END as total
            FROM students s
            LEFT JOIN daily_stats day_s ON s.id = day_s.student_id
            LEFT JOIN summary_stats sum_s ON s.id = sum_s.student_id
            WHERE (%s='' OR s.department=%s) AND (%s='' OR s.division=%s)
        )
        SELECT *,
               ROUND(CASE WHEN total > 0 THEN (attended * 100.0 / total) ELSE 0 END, 1) as pct
        FROM student_stats
        WHERE (CASE WHEN total > 0 THEN (attended * 100.0 / total) ELSE 0 END) < %s
    """
    defaulters = qry(sql, (dept, dept, div, div, threshold_pct))
    
    sent_count = 0
    for s in defaulters:
        pct_val = float(s["pct"] or 0.0)
        results = ParentNotificationService.notify_student_parents(
            student_id=s['id'],
            category='attendance',
            template_slug='defaulter_alert',
            context={'percentage': f"{pct_val}%"}
        )
        if any(r.get("success") for r in results):
            sent_count += 1
            
    return redirect(url_for("admin_intel.shortage_report", 
                             threshold=threshold_pct, 
                             department=dept, 
                             division=div, 
                             alert_triggered=sent_count))

@admin_intel_bp.route("/attendance_roll_call")
@admin_required
def attendance_roll_call():
    from config import Config
    DEPARTMENTS, DIVISIONS, YEARS = Config.DEPARTMENTS, Config.DIVISIONS, Config.YEARS
    from datetime import date, timedelta
    
    view = request.args.get("view", "daily")
    dept = request.args.get("dept", "").strip()
    year = request.args.get("year", "").strip()
    div  = request.args.get("division", "").strip()
    
    today = date.today()
    if view == "weekly":
        start_date = today - timedelta(days=7)
    elif view == "monthly":
        start_date = today - timedelta(days=30)
    elif view == "cumulative":
        start_date = date(2000, 1, 1)
    else:
        start_date = today

    # Standard Student Roll Call
    sql = """
        SELECT s.id, s.name, s.roll, s.division, s.department, s.year,
               (SELECT COALESCE(COUNT(a2.id), 0) FROM attendance a2 
                WHERE a2.student_id = s.id 
                AND (a2.date >= CAST(%s AS TEXT) OR CAST(%s AS TEXT) = '2000-01-01')) as total_sessions,
               (SELECT COALESCE(SUM(CASE WHEN a2.status ILIKE 'Present' THEN 1 ELSE 0 END), 0) FROM attendance a2
                WHERE a2.student_id = s.id 
                AND (a2.date >= CAST(%s AS TEXT) OR CAST(%s AS TEXT) = '2000-01-01')) as present_count
        FROM students s
        WHERE 1=1
    """
    params = [start_date, start_date, start_date, start_date]
    if dept: sql += " AND s.department=%s"; params.append(dept)
    if year: sql += " AND s.year=%s"; params.append(year)
    if div:  sql += " AND s.division=%s"; params.append(div)
    
    sql += " ORDER BY s.roll"
    rows = qry(sql, params)
    
    return render_template("admin/attendance_roll_call.html", 
                           rows=rows, view=view, dept=dept, year=year, div=div,
                           DEPARTMENTS=DEPARTMENTS, YEARS=YEARS, DIVISIONS=DIVISIONS)

@admin_intel_bp.route("/student/<int:student_id>/attendance")
@admin_required
def student_attendance_profile(student_id):
    from flask import abort
    student = qone("SELECT * FROM students WHERE id = %s", (student_id,))
    if not student:
        abort(404)
    
    # Subject breakdown with hardening
    subjects_data = qry("""
        SELECT subject, 
               COALESCE(COUNT(*), 0) as total,
               COALESCE(SUM(CASE WHEN status ILIKE 'Present' THEN 1 ELSE 0 END), 0) as present
        FROM attendance
        WHERE student_id = %s
        GROUP BY subject
    """, (student_id,))
    
    for s in subjects_data:
        s["pct"] = round(s["present"] * 100.0 / s["total"], 1) if s["total"] > 0 else 0

    # Global comparison
    global_avg = qone("SELECT AVG(CASE WHEN status ILIKE 'Present' THEN 100.0 ELSE 0 END) as avg FROM attendance")["avg"] or 0
    
    # Recent history timeline
    recent_history = qry("""
        SELECT a.date, a.subject, a.status, f.name as faculty_name
        FROM attendance a
        LEFT JOIN faculty f ON f.id::TEXT = a.faculty OR f.name = a.faculty
        WHERE a.student_id = %s
        ORDER BY a.date DESC LIMIT 20
    """, (student_id,))

    # Overall Metrics
    overall = qone("""
        SELECT COALESCE(COUNT(*), 0) as total, 
               COALESCE(SUM(CASE WHEN status ILIKE 'Present' THEN 1 ELSE 0 END), 0) as present 
        FROM attendance WHERE student_id=%s
    """, (student_id,))
    
    total = overall["total"]
    present = overall["present"]
    overall_pct = round(present * 100.0 / total, 1) if total > 0 else 0
    
    # Simple risk score: (100 - pct) + penalty for last absence
    last_status = recent_history[0]["status"] if recent_history else "Present"
    risk_score = (100 - overall_pct) + (15 if last_status.lower() == "absent" else 0)

    return render_template("admin/student_subject_breakdown.html", 
                           student=student, 
                           subjects_data=subjects_data, 
                           global_avg=global_avg,
                           recent_history=recent_history,
                           overall_pct=overall_pct,
                           risk_score=min(100, risk_score))
