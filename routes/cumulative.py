"""
cumulative_routes.py
────────────────────
Paste this entire file's content into app.py  OR  save as cumulative_routes.py
and add these two lines near the top of app.py:

    from routes.cumulative import cumulative_bp
    app.register_blueprint(cumulative_bp)

Also add this to init_db() → executescript (before the final triple-quote):

    CREATE TABLE IF NOT EXISTS cumulative_attendance (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        roll         TEXT NOT NULL,
        student_name TEXT NOT NULL,
        department   TEXT NOT NULL DEFAULT '',
        division     TEXT NOT NULL DEFAULT '',
        semester     TEXT NOT NULL DEFAULT '',
        acad_year    TEXT NOT NULL DEFAULT '',
        subject      TEXT NOT NULL,
        subject_code TEXT DEFAULT '',
        conducted    INTEGER NOT NULL DEFAULT 0,
        attended     INTEGER NOT NULL DEFAULT 0,
        percentage   REAL DEFAULT 0,
        updated_at   TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE(roll, subject_code, semester, acad_year)
    );

And add to migrate_db() migrations list:
    ("cumulative_attendance", "division",     "''"),
    ("cumulative_attendance", "subject_code", "''"),
"""

import os, json
from flask import Blueprint, request, redirect, session, jsonify, render_template, send_file
from io import BytesIO

# ── If pasting into app.py, remove this Blueprint wrapper and use app directly ──
try:
    from utils.cumulative_parser import parse_attendance_file
except ImportError:
    from cumulative_parser import parse_attendance_file  # fallback

cumulative_bp = Blueprint("cumulative", __name__)

_CLEAR_ALL_CUMULATIVE_PHRASE = "DELETE ALL CUMULATIVE DATA"

# ─── Helpers (already exist in app.py — skip if pasting) ─────────────────────
# qry, qone, exe, login_required, safe_int  ← use the ones from app.py


# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORT PAGE  /cumulative_import
# ═══════════════════════════════════════════════════════════════════════════════

@cumulative_bp.route("/cumulative_import", methods=["GET", "POST"])
def cumulative_import():
    """Upload one or more PDF/Excel attendance reports and preview before saving."""
    from utils.pg_wrapper import qry, qone, exe
    from blueprints.auth.decorators import login_required

    if session.get("role") != "admin":
        return redirect("/login")

    if request.method == "GET":
        return render_template("cumulative/cumulative_import.html")

    # ── POST: process uploaded files ──────────────────────────────────────────
    if parse_attendance_file is None:
        return jsonify({"error": "cumulative_parser.py not found next to app.py"}), 500

    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400

    all_parsed = []
    errors = []

    for f in files:
        if not f.filename:
            continue
        try:
            data = f.read()
            
            # Save a debug copy of the uploaded file
            try:
                os.makedirs("uploads", exist_ok=True)
                debug_path = os.path.join("uploads", f"debug_{f.filename}")
                with open(debug_path, "wb") as dbg_f:
                    dbg_f.write(data)
            except Exception:
                pass

            parsed = parse_attendance_file(data, f.filename)
            if isinstance(parsed, list):
                for p in parsed:
                    p["filename"] = f.filename
                    all_parsed.append(p)
            else:
                parsed["filename"] = f.filename
                all_parsed.append(parsed)
        except Exception as e:
            errors.append({"file": f.filename, "error": str(e)})

    return jsonify({"parsed": all_parsed, "errors": errors})


# ═══════════════════════════════════════════════════════════════════════════════
#  COMMIT IMPORT  /cumulative_commit   (JSON POST from preview UI)
# ═══════════════════════════════════════════════════════════════════════════════

@cumulative_bp.route("/cumulative_commit", methods=["POST"])
def cumulative_commit():
    from utils.pg_wrapper import exe
    from blueprints.auth.decorators import login_required

    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403

    payload = request.get_json(force=True) or {}
    rows    = payload.get("rows", [])
    inserted = updated = skipped = 0

    for r in rows:
        roll     = (r.get("roll") or "").strip()
        name     = (r.get("name") or "").strip()
        dept     = (r.get("department") or "").strip()
        div      = (r.get("division") or "").strip()
        sem      = (r.get("semester") or "").strip()
        year     = (r.get("acad_year") or "").strip()
        subj     = (r.get("subject") or "").strip()
        code     = (r.get("subject_code") or "").strip()
        conducted = int(r.get("conducted") or 0)
        attended  = int(r.get("attended") or 0)
        pct       = round(100.0 * attended / conducted, 2) if conducted else 0.0

        if not roll or not subj:
            skipped += 1
            continue
        try:
            exe("""
                INSERT INTO cumulative_attendance
                    (roll, student_name, department, division, semester, acad_year,
                     subject, subject_code, conducted, attended, percentage)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT(roll, subject_code, semester, acad_year)
                DO UPDATE SET
                    attended     = excluded.attended,
                    conducted    = excluded.conducted,
                    percentage   = excluded.percentage,
                    student_name = excluded.student_name,
                    updated_at   = CURRENT_TIMESTAMP
            """, (roll, name, dept, div, sem, year, subj, code,
                  conducted, attended, pct))
            inserted += 1
        except Exception as e:
            skipped += 1

    return jsonify({"inserted": inserted, "skipped": skipped})


@cumulative_bp.route("/cumulative_delete_row", methods=["POST"])
def cumulative_delete_row():
    """Delete one cumulative_attendance row (admin)."""
    from utils.pg_wrapper import exe, qone
    from utils.helpers import safe_redirect_target

    if session.get("role") != "admin":
        return redirect("/login")
    rid = request.form.get("row_id", "").strip()
    roll_chk = (request.form.get("roll", "") or "").strip()
    red = request.form.get("redirect", "/cumulative_report")
    if not rid.isdigit() or not roll_chk:
        return redirect(safe_redirect_target(red, "/cumulative_report"))
    row = qone("SELECT id, roll FROM cumulative_attendance WHERE id=%s", (int(rid),))
    if row and row["roll"] == roll_chk:
        exe("DELETE FROM cumulative_attendance WHERE id=%s", (int(rid),))
    return redirect(safe_redirect_target(red, "/cumulative_report"))


@cumulative_bp.route("/cumulative_clear_all", methods=["POST"])
def cumulative_clear_all():
    """Delete all rows in cumulative_attendance (admin, typed confirmation)."""
    from utils.pg_wrapper import exe
    from utils.helpers import safe_redirect_target

    if session.get("role") != "admin":
        return redirect("/login")
    phrase = (request.form.get("confirm_phrase") or "").strip()
    if phrase != _CLEAR_ALL_CUMULATIVE_PHRASE:
        return redirect("/cumulative_report?error=bad_cumulative_confirm")
    exe("DELETE FROM cumulative_attendance")
    return redirect("/cumulative_report?cumulative_cleared=1")


# ═══════════════════════════════════════════════════════════════════════════════
#  CUMULATIVE REPORT  /cumulative_report
# ═══════════════════════════════════════════════════════════════════════════════

@cumulative_bp.route("/cumulative_report")
def cumulative_report():
    from utils.pg_wrapper import qry
    from blueprints.auth.decorators import login_required
    from config import Config
    DEPARTMENTS = Config.DEPARTMENTS
    DIVISIONS = Config.DIVISIONS
    SEMESTERS = Config.SEMESTERS

    if session.get("role") not in ("admin", "faculty"):
        return redirect("/login")

    dept = request.args.get("dept", "")
    div  = request.args.get("div", "")
    sem  = request.args.get("sem", "")
    year = request.args.get("year", "")
    roll = request.args.get("roll", "")

    filters = "WHERE 1=1"
    params  = []
    if dept: filters += " AND department=%s";  params.append(dept)
    if div:  filters += " AND division=%s";    params.append(div)
    if sem:  filters += " AND semester=%s";    params.append(sem)
    if year: filters += " AND acad_year=%s";   params.append(year)
    if roll: filters += " AND roll LIKE %s";   params.append(f"%{roll}%")

    # Per-student cumulative summary
    students = qry(f"""
        SELECT roll, student_name, department, division,
               COUNT(DISTINCT semester||acad_year) AS sem_count,
               SUM(attended)  AS total_attended,
               SUM(conducted) AS total_conducted,
               ROUND(100.0 * SUM(attended) / NULLIF(SUM(conducted),0), 2) AS cumulative_pct
        FROM cumulative_attendance
        {filters}
        GROUP BY roll, student_name, department, division
        ORDER BY department, division, roll
    """, params)  # nosec B608 - filters is composed of safe hardcoded literals and parameters

    # Per-subject breakdown for selected student (exact roll match)
    student_detail = []
    detail_meta = None
    if roll:
        student_detail = qry("""
            SELECT id, student_name, semester, acad_year, subject, subject_code,
                   conducted, attended, percentage
            FROM cumulative_attendance
            WHERE roll=%s
            ORDER BY acad_year, semester, subject
        """, (roll,))
        if student_detail:
            tot_att = sum(int(r["attended"] or 0) for r in student_detail)
            tot_con = sum(int(r["conducted"] or 0) for r in student_detail)
            pct = round(100.0 * tot_att / tot_con, 2) if tot_con else 0.0
            detail_meta = {
                "student_name": (student_detail[0]["student_name"] or "").strip(),
                "total_attended": tot_att,
                "total_conducted": tot_con,
                "cumulative_pct": pct,
            }

    # Available years for filter dropdown
    years = qry("SELECT DISTINCT acad_year FROM cumulative_attendance ORDER BY acad_year DESC")

    # Shortage list (below 75%)
    shortage = [s for s in students if (s["cumulative_pct"] or 0) < 75]

    return render_template("cumulative/cumulative_report.html",
        students=students,
        student_detail=student_detail,
        detail_meta=detail_meta,
        shortage=shortage,
        dept=dept, div=div, sem=sem, year=year, roll=roll,
        DEPARTMENTS=DEPARTMENTS,
        DIVISIONS=DIVISIONS,
        SEMESTERS=SEMESTERS,
        years=years,
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  STUDENT PORTAL — cumulative summary  (call from student dashboard route)
# ═══════════════════════════════════════════════════════════════════════════════

def get_student_cumulative(roll: str) -> dict:
    """
    Call this from your student_dashboard route to get cumulative data.
    
    Usage in your route:
        from cumulative_routes import get_student_cumulative
        cum = get_student_cumulative(session["roll"])
        return render_template("student/student_dashboard.html", ..., cumulative=cum)
    """
    from utils.pg_wrapper import qry, qone
    overall = qone("""
        SELECT SUM(attended)  AS tot_att,
               SUM(conducted) AS tot_con,
               ROUND(100.0 * SUM(attended)/NULLIF(SUM(conducted),0), 2) AS cum_pct
        FROM cumulative_attendance WHERE roll=%s
    """, (roll,))

    by_sem = qry("""
        SELECT semester, acad_year,
               SUM(attended) AS att, SUM(conducted) AS con,
               ROUND(100.0*SUM(attended)/NULLIF(SUM(conducted),0), 2) AS pct
        FROM cumulative_attendance WHERE roll=%s
        GROUP BY acad_year, semester
        ORDER BY acad_year, semester
    """, (roll,))

    by_subject = qry("""
        SELECT subject, subject_code, semester, acad_year,
               conducted, attended, percentage
        FROM cumulative_attendance WHERE roll=%s
        ORDER BY acad_year, semester, subject
    """, (roll,))

    return {
        "overall":    dict(overall) if overall else {},
        "by_sem":     [dict(r) for r in by_sem],
        "by_subject": [dict(r) for r in by_subject],
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT  /cumulative_export
# ═══════════════════════════════════════════════════════════════════════════════

@cumulative_bp.route("/cumulative_export")
def cumulative_export():
    from utils.pg_wrapper import qry
    from blueprints.auth.decorators import login_required
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if session.get("role") not in ("admin", "faculty"):
        return redirect("/login")

    dept = request.args.get("dept", "")
    sem  = request.args.get("sem", "")
    year = request.args.get("year", "")

    filters = "WHERE 1=1"
    params  = []
    if dept: filters += " AND department=%s"; params.append(dept)
    if sem:  filters += " AND semester=%s";   params.append(sem)
    if year: filters += " AND acad_year=%s";  params.append(year)

    # 1. Fetch all distinct subjects in the current dataset
    subjects_in_data = qry(f"""
        SELECT DISTINCT subject
        FROM cumulative_attendance {filters}
        ORDER BY subject
    """, params)  # nosec B608
    subjects = [s["subject"] for s in subjects_in_data]

    # 2. Fetch max conducted lectures for each subject
    subject_conducted = qry(f"""
        SELECT subject, MAX(conducted) AS max_conducted
        FROM cumulative_attendance {filters}
        GROUP BY subject
        ORDER BY subject
    """, params)  # nosec B608
    conducted_map = {s["subject"]: s["max_conducted"] or 0 for s in subject_conducted}

    # 3. Fetch detailed student attendance records
    records = qry(f"""
        SELECT roll, student_name, department, division, semester, acad_year,
               subject, conducted, attended
        FROM cumulative_attendance {filters}
        ORDER BY department, division, roll, subject
    """, params)  # nosec B608

    from collections import defaultdict
    student_records = defaultdict(lambda: {
        "roll": "",
        "name": "",
        "dept": "",
        "div": "",
        "sem": "",
        "year": "",
        "attendance": {},
        "total_conducted": 0,
        "total_attended": 0
    })

    for r in records:
        key = (r["roll"], r["student_name"], r["department"], r["division"], r["semester"], r["acad_year"])
        s = student_records[key]
        s["roll"] = r["roll"]
        s["name"] = r["student_name"]
        s["dept"] = r["department"]
        s["div"] = r["division"]
        s["sem"] = r["semester"]
        s["year"] = r["acad_year"]
        
        subj = r["subject"]
        s["attendance"][subj] = r["attended"]
        s["total_conducted"] += r["conducted"] or 0
        s["total_attended"] += r["attended"] or 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Cumulative Attendance"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    cond_fill = PatternFill("solid", fgColor="F2F2F2")
    cond_font = Font(bold=True, italic=True, color="595959", size=10)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write Headers Row (Row 1)
    headers = ["Roll No", "Name", "Dept", "Div", "Semester", "Year"] + subjects + ["Total Attended", "Total Conducted", "Cumulative %", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Write Conducted Row (Row 2)
    conducted_values = ["", "", "", "", "", "LECTURES CONDUCTED"]
    total_conducted_sum = 0
    for s_name in subjects:
        cond = conducted_map.get(s_name, 0)
        conducted_values.append(cond)
        total_conducted_sum += cond
    conducted_values += ["", total_conducted_sum, "", ""]
    ws.append(conducted_values)
    for cell in ws[2]:
        cell.fill = cond_fill
        cell.font = cond_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    RED = PatternFill("solid", fgColor="FFCCCC")
    YEL = PatternFill("solid", fgColor="FFF2CC")
    GRN = PatternFill("solid", fgColor="CCFFCC")

    # Write Student Rows
    pct_col_idx = 6 + len(subjects) + 3  # Cumulative % column index (1-based)
    
    # Sort students by department, division, roll
    sorted_keys = sorted(student_records.keys(), key=lambda k: (k[2], k[3], k[0]))
    for key in sorted_keys:
        r = student_records[key]
        pct = round(100.0 * r["total_attended"] / r["total_conducted"], 2) if r["total_conducted"] else 0.0
        status = "✓ OK" if pct >= 75 else ("⚠ Low" if pct >= 60 else "✗ Shortage")

        row_values = [r["roll"], r["name"], r["dept"], r["div"], r["sem"], r["year"]]
        for s_name in subjects:
            row_values.append(r["attendance"].get(s_name, "-"))
        row_values += [r["total_attended"], r["total_conducted"], pct, status]

        ws.append(row_values)
        
        row_cells = ws[ws.max_row]
        fill = GRN if pct >= 75 else (YEL if pct >= 60 else RED)
        for cell in row_cells:
            cell.border = border
            if cell.column == pct_col_idx:
                cell.fill = fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 10)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"cumulative_{dept or 'all'}_{sem or 'all'}_{year or 'all'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@cumulative_bp.route("/api/cumulative/notify-parents", methods=["POST"])
def notify_cumulative_parents():
    from utils.pg_wrapper import qry, qone
    from services.parent_notification_service import ParentNotificationService
    
    if session.get("role") not in ("admin", "faculty"):
        return jsonify({"error": "Unauthorized"}), 403

    payload = request.get_json(force=True) or {}
    rolls = payload.get("rolls", [])
    if not isinstance(rolls, list):
        if rolls:
            rolls = [rolls]
        else:
            rolls = []

    if not rolls:
        return jsonify({"error": "No student rolls provided"}), 400

    results = []
    skipped = []
    
    for roll in rolls:
        roll = roll.strip()
        if not roll:
            continue
            
        # 1. Look up student in students table by roll
        student = qone("SELECT id, name FROM students WHERE roll=%s", (roll,))
        if not student:
            skipped.append({"roll": roll, "reason": "Student not found in main database."})
            continue

        # 2. Get student overall cumulative percentage
        cum = qone("""
            SELECT SUM(attended) AS total_attended, SUM(conducted) AS total_conducted
            FROM cumulative_attendance
            WHERE roll=%s
        """, (roll,))
        
        if not cum or not cum["total_conducted"]:
            skipped.append({"roll": roll, "reason": "No cumulative attendance records found."})
            continue
            
        pct = round(100.0 * cum["total_attended"] / cum["total_conducted"], 2)
        
        # 3. Notify student parents
        try:
            sms_res = ParentNotificationService.notify_student_parents(
                student_id=student["id"],
                category="attendance",
                template_slug="defaulter_alert",
                context={"percentage": f"{pct}%"}
            )
            if not sms_res:
                skipped.append({"roll": roll, "name": student["name"], "reason": "No primary parent contact mapped or notification disabled."})
            else:
                for r in sms_res:
                    results.append({
                        "roll": roll,
                        "name": student["name"],
                        "parent": r.get("parent", "Parent"),
                        "success": r.get("success", True),
                        "error": r.get("error")
                    })
        except Exception as e:
            skipped.append({"roll": roll, "name": student["name"], "reason": f"Error sending SMS: {str(e)}"})

    return jsonify({
        "status": "success",
        "notified": results,
        "skipped": skipped
    })
