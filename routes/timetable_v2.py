import logging
from flask import Blueprint, render_template, request, session, jsonify, flash, redirect, send_file
from services.timetable_service import TimetableService
from repositories.timetable_repository import TimetableRepository
from models.timetable_model import TimetableEntry
from config import DAYS, DAY_ORD, DEPARTMENTS, DIVISIONS
from blueprints.auth.decorators import login_required
from datetime import date
from utils.pg_wrapper import qry, qone, exe
from utils.helpers import safe_int
from routes.features import log_audit
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import io
import re

logger = logging.getLogger(__name__)

timetable_v2_bp = Blueprint('timetable_v2', __name__)

# Manual Dependency Injection (Simple for monolithic migration)
_repo = TimetableRepository()
_service = TimetableService(_repo)

SLOT_COLORS = {
    "Theory":    "#2563EB",
    "Lab":       "#7C3AED",
    "Elective":  "#059669",
    "Minor":     "#D97706",
    "Practical": "#0891B2",
    "Other":     "#6B7280",
}

def assign_color(subject, slot_type):
    if slot_type and slot_type in SLOT_COLORS:
        return SLOT_COLORS[slot_type]
    s = (subject or "").lower()
    if "lab" in s:       return SLOT_COLORS["Lab"]
    if "elective" in s:  return SLOT_COLORS["Elective"]
    if "minor" in s:     return SLOT_COLORS["Minor"]
    return SLOT_COLORS["Theory"]

@timetable_v2_bp.route("/timetable_v2")
@login_required(role=['admin', 'faculty', 'student'])
def timetable_view():
    """
    Renders the timetable grid and list views.
    Transferred from app.py to modular blueprint.
    """
    from extensions import db
    from models.timetable import Timetable
    from models.faculty import Faculty

    # 1. Capture Filters
    q = request.args.get("q", "").strip()
    f_day = request.args.get("day", "")
    f_subj = request.args.get("subject", "")
    f_teach = request.args.get("teacher", "")
    f_div = request.args.get("division", "")
    f_sem = request.args.get("semester", "")
    f_type = request.args.get("slot_type", "")
    view = request.args.get("view", "grid")
    
    # 2. Query all entries with coalesced teacher name
    all_entries = db.session.query(
        Timetable.id,
        Timetable.day,
        Timetable.time,
        Timetable.start_time,
        Timetable.end_time,
        Timetable.subject_id,
        Timetable.subject,
        db.func.coalesce(Faculty.name, Timetable.teacher).label('teacher'),
        Timetable.faculty_id,
        Timetable.room,
        Timetable.division,
        Timetable.branch,
        Timetable.year,
        Timetable.semester,
        Timetable.slot_type,
        Timetable.color,
        Timetable.published,
        Timetable.created_at
    ).outerjoin(Faculty, Timetable.faculty_id == Faculty.id).all()
    
    # Extract unique filter options from the grid entries
    subjects = sorted(list(set(e.subject for e in all_entries if e.subject)))
    teachers = sorted(list(set(e.teacher for e in all_entries if e.teacher)))
    divs = sorted(list(set(e.division for e in all_entries if e.division)))

    # Compute faculty workload stats
    workload = {}
    for e in all_entries:
        t = e.teacher
        if t:
            workload[t] = workload.get(t, 0) + 1
    workload = sorted(workload.items(), key=lambda x: -x[1])[:10]

    # 3. Query filtered entries for list view and grid rendering
    sql = "SELECT t.*, COALESCE(f.name, t.teacher) as teacher FROM timetable t LEFT JOIN faculty f ON t.faculty_id = f.id WHERE 1=1"
    params = []
    if q:
        sql += " AND (t.subject ILIKE %s OR f.name ILIKE %s OR t.teacher ILIKE %s OR t.room ILIKE %s OR f.department ILIKE %s)"
        params += [f"%{q}%"] * 5
    if f_day:
        sql += " AND t.day=%s"
        params.append(f_day)
    if f_subj:
        sql += " AND t.subject ILIKE %s"
        params.append(f"%{f_subj}%")
    if f_teach:
        sql += " AND (f.name ILIKE %s OR t.teacher ILIKE %s)"
        params += [f"%{f_teach}%"] * 2
    if f_div:
        sql += " AND t.division ILIKE %s"
        params.append(f"%{f_div}%")
    if f_type:
        sql += " AND t.slot_type=%s"
        params.append(f_type)
    if f_sem:
        # Robust semester filtering supporting both Arabic and Roman numerals
        roman_map = {"1": "I", "2": "II", "3": "III", "4": "IV", "5": "V", "6": "VI", "7": "VII", "8": "VIII"}
        roman_sem = roman_map.get(f_sem, f_sem)
        sql += " AND (t.semester=%s OR t.semester=%s)"
        params += [f_sem, roman_sem]
        
    sql += f" ORDER BY {DAY_ORD}, t.start_time"
    filtered_entries = qry(sql, params)
    
    any_filter = bool(q or f_day or f_subj or f_teach or f_div or f_type or f_sem)
    filtered_ids = set(e["id"] for e in filtered_entries) if any_filter else None
    
    # Chronological sorting for time slots
    def _sk(ts):
        m = re.match(r"(\d+):(\d+)", ts)
        if not m: return 999
        h = int(m.group(1))
        mn = int(m.group(2))
        if h < 7: h += 12
        return h * 60 + mn

    # Filter grid time slots to make it compact if a filter is applied
    if any_filter:
        time_slots = sorted(list(set(e["time"] for e in filtered_entries if e.get("time"))), key=_sk)
    else:
        time_slots = sorted(list(set(e.time for e in all_entries if e.time)), key=_sk)

    grid = {d: {ts: [] for ts in time_slots} for d in DAYS}
    for e in all_entries:
        if e.day in grid and e.time in grid[e.day]:
            grid[e.day][e.time].append(e)
            
    # Compute dynamic stats from filtered entries
    total = len(filtered_entries)
    theory = sum(1 for e in filtered_entries if (e.get("slot_type") or "Theory") == "Theory")
    lab = sum(1 for e in filtered_entries if e.get("slot_type") == "Lab")
    elec = sum(1 for e in filtered_entries if e.get("slot_type") == "Elective")
    filtered_teachers = len(set(e.get("teacher") for e in filtered_entries if e.get("teacher")))

    return render_template(
        "common/timetable.html",
        entries=filtered_entries, 
        grid=grid, 
        time_slots=time_slots,
        subjects=subjects,
        teachers=teachers,
        divs=divs,
        workload=workload,
        total=total,
        theory=theory,
        lab=lab,
        elec=elec,
        filtered_teachers=filtered_teachers,
        any_filter=any_filter,
        filtered_ids=filtered_ids,
        q=q, f_day=f_day, f_subj=f_subj, f_teach=f_teach,
        f_div=f_div, f_sem=f_sem, f_type=f_type, view=view,
        DAYS=DAYS,
        today_name=date.today().strftime("%A")
    )

@timetable_v2_bp.route("/api/add_time_slot", methods=["POST"])
@login_required(role=['admin'])
def api_add_slot():
    data = request.json
    if not data or 'slot' not in data:
        return jsonify({"ok": False, "error": "Missing slot data"}), 400
    
    # Construct Domain Model
    # Here we might be creating a placeholder slot
    entry = TimetableEntry(
        id=None,
        day=data.get('day', 'Monday'),
        time=data['slot'],
        subject=data.get('subject', 'New Slot'),
        department='General'
    )
    
    result = _service.add_or_update_slot(entry)
    return jsonify(result)

@timetable_v2_bp.route("/api/copy_day", methods=["POST"])
@login_required(role=['admin'])
def api_copy_day():
    data = request.json
    from_day = data.get('from_day')
    to_day = data.get('to_day')
    
    if not from_day or not to_day:
        return jsonify({"ok": False, "error": "Missing day parameters"}), 400
        
    result = _service.copy_day_schedule(from_day, to_day)
    return jsonify(result)

@timetable_v2_bp.route("/public/timetable/<branch>/<year>/<division>")
def public_timetable(branch, year, division):
    """Generates a weekly timetable view accessible to anyone without log in."""
    from utils.pg_wrapper import qry
    from config import DAYS, DAY_ORD
    import re

    # Fetch published timetable entries for this division
    all_entries = [dict(e) for e in qry(
        f"SELECT t.*, f.name as teacher FROM timetable t LEFT JOIN faculty f ON t.faculty_id = f.id "
        f"WHERE t.branch=%s AND t.year=%s AND t.division=%s AND t.published=TRUE "
        f"ORDER BY {DAY_ORD}, t.start_time",  # nosec B608 - DAY_ORD is safe static config
        (branch, year, division)
    )]

    # Extract all distinct time slots
    seen = set()
    raw = []
    for e in all_entries:
        t = e.get("time")
        if t and t not in seen:
            seen.add(t)
            raw.append(t)

    # Sort time slots
    def _sk(ts):
        m = re.match(r"(\d+):(\d+)", ts)
        if not m: return 999
        h = int(m.group(1))
        mn = int(m.group(2))
        if h < 7: h += 12
        return h * 60 + mn

    time_slots = sorted(raw, key=_sk)

    # Construct the grid
    grid = {d: {t: [] for t in time_slots} for d in DAYS}
    for e in all_entries:
        if e["day"] in grid and e["time"] in grid[e["day"]]:
            grid[e["day"]][e["time"]].append(e)

    # Render a beautiful standalone template
    return render_template(
        "timetable/public_timetable.html",
        branch=branch,
        year=year,
        division=division,
        grid=grid,
        time_slots=time_slots,
        DAYS=DAYS
    )


def _build_tt_body(teacher, slots):
    """Build formatted timetable message for a faculty member."""
    # Convert psycopg2.extras.DictCursor objects to dicts so .get() works
    slots = [dict(s) for s in slots]
    lines = ["Weekly Timetable - DY Patil University\n"]
    lines.append("Faculty: " + teacher)
    lines.append("=" * 42)
    by_day = {}
    for s in slots:
        by_day.setdefault(s["day"], []).append(s)
    for day in DAYS:
        if day not in by_day: continue
        lines.append("\n" + day + ":")
        for s in sorted(by_day[day], key=lambda x: x.get("time","") or ""):
            room = (" | Room: " + s["room"]) if s.get("room") else ""
            div  = (" | " + s["division"])  if s.get("division") else ""
            lines.append("  " + str(s["time"] or "").ljust(15) + " " + str(s["subject"]) + room + div)
    lines.append("\n\nSent from DY Patil ERP Admin.")
    return "\n".join(lines)


@timetable_v2_bp.route("/timetable_share", methods=["GET","POST"])
@login_required("admin")
def timetable_share():
    if request.method == "POST":
        target_role = request.form.get("target_role", "faculty")
        target_id   = safe_int(request.form.get("target_id","0"))
        
        if not target_id: return redirect(f"/timetable_share?error=no_{target_role}")

        if target_role == "faculty":
            row = qone("SELECT * FROM faculty WHERE id=%s", (target_id,))
            if not row: return redirect(f"/timetable_share?error=no_faculty")
            exact_name = row["name"]
            slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.faculty_id=%s ORDER BY t.day, t.start_time", (target_id,))
        else:
            row = qone("SELECT * FROM students WHERE id=%s", (target_id,))
            if not row: return redirect(f"/timetable_share?role=student&error=not_found")
            exact_name = row["name"]
            slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.branch=%s AND t.division=%s ORDER BY t.day, t.start_time", 
                        (row["department"], row["division"]))

        if not slots: return redirect(f"/timetable_share?role={target_role}&error=no_slots&id={target_id}")

        body = _build_tt_body(exact_name, slots)
        subj = f"Your Weekly Timetable — {len(slots)} slots"
        exe("""INSERT INTO messages(from_role,from_id,from_name,to_role,to_id,to_name,subject,body)
               VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
            ("admin", 1, "Administrator", target_role, target_id, exact_name, subj, body))
        
        log_audit("Share Timetable", f"Sent individual timetable to {target_role}: {exact_name} (ID: {target_id})")
        return redirect(f"/timetable_share?role={target_role}&sent=1&to={exact_name}")

    # GET
    role_sel = request.args.get("role", "faculty")
    id_sel   = safe_int(request.args.get("id","0"))
    selected_slots = []
    matched_entity = None
    
    if id_sel:
        if role_sel == "faculty":
            matched_entity = qone("SELECT * FROM faculty WHERE id=%s", (id_sel,))
            if matched_entity:
                matched_entity = dict(matched_entity)
                selected_slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.faculty_id=%s ORDER BY t.day, t.start_time", (id_sel,))
        else:
            matched_entity = qone("SELECT * FROM students WHERE id=%s", (id_sel,))
            if matched_entity:
                matched_entity = dict(matched_entity)
                selected_slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.branch=%s AND t.division=%s ORDER BY t.day, t.start_time", 
                                    (matched_entity["department"], matched_entity["division"]))

    return render_template("common/timetable_share.html",
        role_sel=role_sel,
        selected_slots=selected_slots,
        matched_entity=matched_entity,
        faculty_list=qry("SELECT id,name,department FROM faculty ORDER BY name"),
        student_list=qry("SELECT id,name,department,division,roll FROM students ORDER BY department, division, name"),
        DEPARTMENTS=DEPARTMENTS,
        DIVISIONS=DIVISIONS,
        DAYS=DAYS,
        today_name=date.today().strftime("%A")
    )


@timetable_v2_bp.route("/timetable_send_all", methods=["POST"])
@login_required("admin")
def timetable_send_all():
    target_role = request.form.get("target_role", "faculty")
    sent = skipped = 0
    
    if target_role == "faculty":
        distinct_faculties = qry("SELECT DISTINCT faculty_id FROM timetable WHERE faculty_id IS NOT NULL")
        for row in distinct_faculties:
            fac_id = row["faculty_id"]
            filter_val = (fac_id,)
            f = qone("SELECT name FROM faculty WHERE id=%s", filter_val)
            if not f: 
                skipped += 1
                continue
            fac_name = f["name"]
            slots = qry("SELECT t.*, f2.name as teacher FROM timetable t JOIN faculty f2 ON t.faculty_id=f2.id WHERE t.faculty_id=%s ORDER BY t.day, t.start_time", (fac_id,))
            if not slots: continue
            
            body = _build_tt_body(fac_name, slots)
            subj = f"Your Weekly Timetable — {len(slots)} slots"
            exe("""INSERT INTO messages(from_role,from_id,from_name,to_role,to_id,to_name,subject,body)
                   VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
                ("admin", 1, "Administrator", "faculty", fac_id, fac_name, subj, body))
            sent += 1
    else:
        students = qry("SELECT id, name, department, division FROM students")
        cache = {}
        for s in students:
            key = (s["department"], s["division"])
            if key not in cache:
                slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.branch=%s AND t.division=%s ORDER BY t.day, t.start_time", (key[0], key[1]))
                cache[key] = slots
            
            slots = cache[key]
            if not slots:
                skipped += 1
                continue
            
            body = _build_tt_body(s["name"], slots)
            subj = f"Your Class Timetable — {len(slots)} slots"
            exe("""INSERT INTO messages(from_role,from_id,from_name,to_role,to_id,to_name,subject,body)
                   VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
                ("admin", 1, "Administrator", "student", s["id"], s["name"], subj, body))
            sent += 1

    log_audit("Bulk Share Timetable", f"Initiated bulk share for {target_role}s. Successfully sent: {sent}, Skipped: {skipped}")
    return redirect(f"/timetable_share?role={target_role}&bulk_sent={sent}&bulk_skip={skipped}")


# Helper functions for timetable actions
def normalize_time(t):
    import re
    if not t: return t
    t = str(t).strip()
    parts = re.split(r'\s*-\s*', t)
    result = []
    for part in parts:
        m = re.match(r'^(\d{1,2}):(\d{2})$', part.strip())
        if m:
            result.append(f"{int(m.group(1)):02d}:{m.group(2)}")
        else:
            result.append(part.strip())
    return '-'.join(result)

def _check_conflicts(day, time_slot, teacher, room, division, exclude_id=None):
    conflicts = []
    sql = "SELECT * FROM timetable WHERE day=%s AND time=%s"
    params = [day, time_slot]
    if exclude_id:
        sql += " AND id != %s"
        params.append(int(exclude_id))
    rows = qry(sql, params)
    for r in rows:
        if teacher and r["teacher"] and r["teacher"].strip() == teacher.strip():
            conflicts.append(f"Faculty {teacher!r} already assigned at {day} {time_slot}")
        if room and r["room"] and r["room"].strip() == room.strip():
            conflicts.append(f"Room {room!r} already booked at {day} {time_slot}")
        if division and r["division"] and r["division"].strip() == division.strip():
            conflicts.append(f"Division {division!r} already has a class at {day} {time_slot}")
    return conflicts

@timetable_v2_bp.route("/save_timetable", methods=["POST"])
@login_required("admin")
def save_timetable():
    day      = request.form.get("day","")
    time_s   = normalize_time(request.form.get("time",""))
    subject  = request.form.get("subject","").strip()
    teacher  = request.form.get("teacher","").strip()
    room     = request.form.get("room","").strip()
    division = request.form.get("division","").strip()
    semester = request.form.get("semester","").strip()
    slot_type= request.form.get("slot_type","Theory")
    color    = assign_color(subject, slot_type)
    if not subject: return redirect("/timetable_v2?error=nosubject")

    # Add backend protection insertion logic:
    faculty_id = None
    if teacher:
        f_row = qone("SELECT id FROM faculty WHERE name=%s LIMIT 1", (teacher,))
        if f_row:
            faculty_id = f_row["id"]
    
    branch = ""
    year = ""
    if division:
        s_row = qone("SELECT department, year FROM students WHERE division=%s LIMIT 1", (division,))
        if s_row:
            branch = s_row["department"]
            year = s_row["year"]

    # Parse start and end time
    start_time = None
    end_time = None
    if time_s:
        m = re.match(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", time_s)
        if m:
            h1, m1, h2, m2 = map(int, m.groups())
            if h1 < 7: h1 += 12
            if h2 < 7: h2 += 12
            start_time = f"{h1:02d}:{m1:02d}:00"
            end_time = f"{h2:02d}:{m2:02d}:00"

    sub_row = qone("SELECT id FROM subjects WHERE name=%s LIMIT 1", (subject,))
    subject_id = sub_row["id"] if sub_row else None

    # --- CLASH DETECTION LOGIC ---
    clash_cond = "NOT (end_time <= %s OR start_time >= %s)"
    if faculty_id and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND faculty_id=%s AND {clash_cond}", (day, faculty_id, start_time, end_time)):  # nosec B608
            return "Teacher clash: Faculty is already assigned here.", 409
    if branch and year and division and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND branch=%s AND year=%s AND division=%s AND {clash_cond}", (day, branch, year, division, start_time, end_time)):  # nosec B608
            return redirect("/timetable_v2?error=" + f"Class clash: {division} is busy at {time_s}.")
    if room and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND room=%s AND {clash_cond}", (day, room, start_time, end_time)):  # nosec B608
            return redirect("/timetable_v2?error=" + f"Room clash: {room} is already booked.")

    exe("INSERT INTO timetable(day,time,start_time,end_time,subject_id,subject,teacher,room,division,semester,slot_type,color,faculty_id,branch,year) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (day,time_s,start_time,end_time,subject_id,subject,teacher,room,division,semester,slot_type,color,faculty_id,branch,year))
    
    # Auto-create faculty assignment
    if faculty_id and subject_id:
        try:
            exe("""
                INSERT INTO faculty_subject_assignments
                    (faculty_id, subject_id, subject_name, department, semester,
                     class_name, division, academic_year)
                VALUES (%s, %s, %s, %s, %s, %s, %s, '2025-26')
                ON CONFLICT (faculty_id, subject_id, division) DO NOTHING
            """, (faculty_id, subject_id, subject, branch or '',
                  semester or '', f"{year}-{branch}" if year and branch else (division or ''),
                  division or ''))
            log_audit("Auto Faculty Assignment", f"Auto-assigned {subject} → faculty_id {faculty_id} (div {division})")
        except Exception as e:
            logger.warning(f"Auto-assignment failed: {e}")

    return redirect("/timetable_v2?saved=1")

@timetable_v2_bp.route("/edit_timetable", methods=["POST"])
@login_required("admin")
def edit_timetable():
    tid      = request.form.get("tt_id","")
    day      = request.form.get("day","")
    time_s   = normalize_time(request.form.get("time",""))
    subject  = request.form.get("subject","").strip()
    teacher  = request.form.get("teacher","").strip()
    room     = request.form.get("room","").strip()
    division = request.form.get("division","").strip()
    semester = request.form.get("semester","").strip()
    slot_type= request.form.get("slot_type","Theory")
    color    = assign_color(subject, slot_type)

    faculty_id = None
    if teacher:
        f_row = qone("SELECT id FROM faculty WHERE name=%s LIMIT 1", (teacher,))
        if f_row: faculty_id = f_row["id"]
    
    branch = ""
    year = ""
    if division:
        s_row = qone("SELECT department, year FROM students WHERE division=%s LIMIT 1", (division,))
        if s_row:
            branch = s_row["department"]
            year = s_row["year"]

    start_time = None
    end_time = None
    if time_s:
        m = re.match(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", time_s)
        if m:
            h1, m1, h2, m2 = map(int, m.groups())
            if h1 < 7: h1 += 12
            if h2 < 7: h2 += 12
            start_time = f"{h1:02d}:{m1:02d}:00"
            end_time = f"{h2:02d}:{m2:02d}:00"

    sub_row = qone("SELECT id FROM subjects WHERE name=%s LIMIT 1", (subject,))
    subject_id = sub_row["id"] if sub_row else None

    # --- CLASH DETECTION LOGIC (EXCLUDING CURRENT ID) ---
    clash_cond = "NOT (end_time <= %s OR start_time >= %s) AND id != %s"
    if faculty_id and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND faculty_id=%s AND {clash_cond}", # nosec B608
                (day, faculty_id, start_time, end_time, tid)):
            return "Teacher clash detected during edit.", 409
    if branch and year and division and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND branch=%s AND year=%s AND division=%s AND {clash_cond}", # nosec B608
                (day, branch, year, division, start_time, end_time, tid)):
            return redirect(f"/timetable_v2?error=Class clash detected during edit: {division} busy at {time_s}.")
    if room and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND room=%s AND {clash_cond}", # nosec B608
                (day, room, start_time, end_time, tid)):
            return redirect(f"/timetable_v2?error=Room clash detected during edit: {room} already booked.")

    exe("""UPDATE timetable 
           SET day=%s, time=%s, start_time=%s, end_time=%s, subject_id=%s, 
               subject=%s, teacher=%s, room=%s, division=%s, semester=%s, 
               slot_type=%s, color=%s, faculty_id=%s, branch=%s, year=%s 
           WHERE id=%s""",
        (day, time_s, start_time, end_time, subject_id, subject, teacher, room, division, semester, slot_type, color, faculty_id, branch, year, tid))
    
    # Update faculty assignment if faculty changed
    if faculty_id and subject_id:
        try:
            exe("""
                INSERT INTO faculty_subject_assignments
                    (faculty_id, subject_id, subject_name, department, semester,
                     class_name, division, academic_year)
                VALUES (%s, %s, %s, %s, %s, %s, %s, '2025-26')
                ON CONFLICT (faculty_id, subject_id, division) DO UPDATE SET
                    subject_name = EXCLUDED.subject_name,
                    department = EXCLUDED.department
            """, (faculty_id, subject_id, subject, branch or '',
                  semester or '', f"{year}-{branch}" if year and branch else (division or ''),
                  division or ''))
        except Exception as e:
            logger.warning(f"Auto-assignment update failed: {e}")

    return redirect("/timetable_v2?saved=1")

@timetable_v2_bp.route("/export_timetable_excel")
@login_required(role=['admin', 'faculty', 'student'])
def export_timetable_excel():
    q = request.args.get("q", "").strip()
    f_day = request.args.get("day", "")
    f_subj = request.args.get("subject", "")
    f_teach = request.args.get("teacher", "")
    f_div = request.args.get("division", "")
    f_sem = request.args.get("semester", "")
    f_type = request.args.get("slot_type", "")

    sql = "SELECT t.*, COALESCE(f.name, t.teacher) as teacher FROM timetable t LEFT JOIN faculty f ON t.faculty_id = f.id WHERE 1=1"
    params = []
    if q:
        sql += " AND (t.subject ILIKE %s OR f.name ILIKE %s OR t.teacher ILIKE %s OR t.room ILIKE %s OR f.department ILIKE %s)"
        params += [f"%{q}%"] * 5
    if f_day:
        sql += " AND t.day=%s"
        params.append(f_day)
    if f_subj:
        sql += " AND t.subject ILIKE %s"
        params.append(f"%{f_subj}%")
    if f_teach:
        sql += " AND (f.name ILIKE %s OR t.teacher ILIKE %s)"
        params += [f"%{f_teach}%"] * 2
    if f_div:
        sql += " AND t.division ILIKE %s"
        params.append(f"%{f_div}%")
    if f_type:
        sql += " AND t.slot_type=%s"
        params.append(f_type)
    if f_sem:
        roman_map = {"1": "I", "2": "II", "3": "III", "4": "IV", "5": "V", "6": "VI", "7": "VII", "8": "VIII"}
        roman_sem = roman_map.get(f_sem, f_sem)
        sql += " AND (t.semester=%s OR t.semester=%s)"
        params += [f_sem, roman_sem]
        
    sql += " ORDER BY t.day, t.time"
    rows = qry(sql, params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"
    hdrs = ["Day","Time","Subject","Faculty","Room","Division","Semester","Type"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
    for ri, row in enumerate(rows, 2):
        for c, k in enumerate(["day","time","subject","teacher","room","division","semester","slot_type"], 1):
            ws.cell(ri, c, row[k] or "")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="timetable.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@timetable_v2_bp.route("/delete_timetable", methods=["POST"])
@login_required("admin")
def delete_timetable():
    tid = request.form.get("tt_id","")
    exe("DELETE FROM timetable WHERE id=%s", (tid,))
    return redirect("/timetable/?deleted=1")

@timetable_v2_bp.route("/clear_timetable", methods=["POST"])
@login_required("admin")
def clear_timetable():
    exe("DELETE FROM timetable")
    return redirect("/timetable/?cleared=1")

@timetable_v2_bp.route("/clean_timetable", methods=["POST"])
@login_required("admin")
def clean_timetable():
    valid = "'Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'"
    n = qone(f"SELECT COUNT(*) as c FROM timetable WHERE day NOT IN ({valid}) OR subject='' OR subject IS NULL")["c"]  # nosec B608
    exe(f"DELETE FROM timetable WHERE day NOT IN ({valid}) OR subject='' OR subject IS NULL")  # nosec B608
    return redirect(f"/timetable/?cleaned={n}")

@timetable_v2_bp.route("/duplicate_timetable", methods=["POST"])
@login_required("admin")
def duplicate_timetable():
    r = qone("SELECT * FROM timetable WHERE id=%s", (request.form.get("tt_id",""),))
    if r:
        exe("INSERT INTO timetable(day,time,start_time,end_time,subject_id,subject,teacher,room,division,semester,slot_type,color,faculty_id,branch,year) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (r["day"],r["time"],r.get("start_time"),r.get("end_time"),r.get("subject_id"),r["subject"],r["teacher"],r["room"] or "",
             r["division"] or "",r["semester"] or "",r["slot_type"] or "Theory",r["color"] or "", r.get("faculty_id"), r.get("branch"), r.get("year")))
    return redirect("/timetable/?added=1")

@timetable_v2_bp.route("/move_timetable", methods=["POST"])
@login_required("admin")
def move_timetable():
    data  = request.get_json() or {}
    tid   = data.get("id","")
    day   = data.get("day","")
    time_s= normalize_time(data.get("time",""))
    r = qone("SELECT * FROM timetable WHERE id=%s", (tid,))
    if not r: return jsonify({"ok":False}), 404
    start_time, end_time = None, None
    import re
    m = re.match(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", time_s)
    if m:
        h1, m1, h2, m2 = map(int, m.groups())
        if h1 < 7: h1 += 12
        if h2 < 7: h2 += 12
        start_time = f"{h1:02d}:{m1:02d}:00"
        end_time = f"{h2:02d}:{m2:02d}:00"
    exe("UPDATE timetable SET day=%s,time=%s,start_time=%s,end_time=%s WHERE id=%s", (day,time_s,start_time,end_time,tid))
    return jsonify({"ok":True})

@timetable_v2_bp.route("/check_conflicts_api", methods=["POST"])
@login_required("admin")
def check_conflicts_api():
    d = request.get_json() or {}
    conflicts = _check_conflicts(d.get("day",""), d.get("time",""),
                                  d.get("teacher",""), d.get("room",""),
                                  d.get("division",""), d.get("exclude_id"))
    return jsonify({"conflicts": conflicts})

@timetable_v2_bp.route("/api/send_timetable", methods=["POST"])
@login_required
def api_send_timetable():
    if session.get("role") not in ["admin", "faculty"]:
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    
    target_dept = request.form.get("department")
    target_year = request.form.get("year")
    target_div  = request.form.get("division")
    
    # Filter students
    sql = "SELECT id, name, department, division FROM students WHERE 1=1"
    params = []
    if target_dept and target_dept != "All": sql += " AND department=%s"; params.append(target_dept)
    if target_year and target_year != "All": sql += " AND year=%s"; params.append(target_year)
    if target_div and target_div != "All": sql += " AND division=%s"; params.append(target_div)
    
    students = qry(sql, params)
    sent = 0
    
    for s in students:
        fid = session.get("faculty_id")
        fname = session.get("name")
        body = f"Hello {s['name']}, \nYour timetable has been updated. You can view it here: {request.host_url}share/timetable/{fid if fid else 1}"
        subj = "Timetable Update Notification"
        
        exe("""INSERT INTO messages(from_role, from_id, from_name, to_role, to_id, to_name, subject, body) 
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
            (session.get("role"), session.get("user_id") or fid or 1, fname, "student", s["id"], s["name"], subj, body))
        sent += 1
        
    return jsonify({"status": "ok", "sent": sent})
