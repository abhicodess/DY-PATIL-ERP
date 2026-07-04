
from flask import Blueprint, render_template, request, redirect, session, jsonify, flash, send_file, url_for
from utils.db_helpers import safe_query, safe_fetch_one, safe_fetch_scalar, safe_execute, log_audit
from blueprints.auth.decorators import login_required

def normalize_branch(branch):
    if not branch: return ""
    return str(branch).strip()

def normalize_division(division):
    if not division: return ""
    return str(division).strip()

def parse_class_details(division_str, branch_str):
    """
    Parses class details (year, departments, division) from strings like
    division_str="SE CSE-A" or "SE AIDS" and branch_str="" or "CS".
    Returns (year, departments, division) where departments is a list of potential department names.
    """
    import re
    full_str = f"{division_str or ''} {branch_str or ''}".upper().strip()
    
    # 1. Parse Year
    year = None
    if re.search(r'\b(FE|FY|I|1ST)\b', full_str) or re.search(r'\bFIRST\b', full_str):
        year = 'I'
    elif re.search(r'\b(SE|SY|II|2ND)\b', full_str) or re.search(r'\bSECOND\b', full_str):
        year = 'II'
    elif re.search(r'\b(TE|TY|III|3RD)\b', full_str) or re.search(r'\bTHIRD\b', full_str):
        year = 'TY'
    elif re.search(r'\b(BE|LY|IV|4TH|FINAL)\b', full_str) or re.search(r'\bFOURTH\b', full_str):
        year = 'IV'
        
    # 2. Parse Division
    # Look for a standalone letter A, B, C, D or preceded by hyphen/space
    div_match = re.search(r'(?:[-/\s]|^)([A-D])\b', full_str)
    division = div_match.group(1) if div_match else None
    
    # 3. Parse Departments
    departments = []
    if 'AIDS' in full_str or 'AI' in full_str:
        departments = ['AIDS', 'AIML']
    elif 'CSE' in full_str or 'CS' in full_str or 'COMPUTER' in full_str:
        departments = ['CS', 'Computer']
    elif 'IT' in full_str or 'INFORMATION' in full_str:
        departments = ['IT']
        
    return year, departments, division

def resolve_assigned_class(fid, subject, slot_division):
    """
    Given a faculty ID, a subject name, and a timetable slot's division,
    checks if there is an active assignment for this faculty.
    Returns (resolved_division, resolved_branch/department) if found, otherwise (slot_division, None).
    """
    if not subject:
        return slot_division, None
        
    # Get all assignments for this faculty
    assignments = safe_query("""
        SELECT division, department, subject_name 
        FROM faculty_subject_assignments 
        WHERE faculty_id = %s
    """, (fid,))
    
    if not assignments:
        return slot_division, None
        
    def norm(s):
        import re
        if not s: return ""
        return re.sub(r'[^A-Z0-9]', '', s.upper())
        
    normalized_target = norm(subject)
    matching_assignments = [a for a in assignments if norm(a['subject_name']) == normalized_target]
    
    if not matching_assignments:
        return slot_division, None
        
    # If there is exactly one assignment, use it!
    if len(matching_assignments) == 1:
        return matching_assignments[0]['division'], matching_assignments[0]['department']
        
    # If there are multiple, look for one that shares the division letter (e.g. 'A' or 'B')
    if slot_division:
        # Extract division letter from slot_division (e.g. 'A' from 'SE CSE-A')
        import re
        m = re.search(r'\b([A-D])\b', slot_division.upper())
        slot_letter = m.group(1) if m else slot_division.upper()
        
        for a in matching_assignments:
            a_div = a['division'] or ''
            am = re.search(r'\b([A-D])\b', a_div.upper())
            a_letter = am.group(1) if am else a_div.upper()
            if slot_letter in a_letter or a_letter in slot_letter:
                return a['division'], a['department']
                
    return matching_assignments[0]['division'], matching_assignments[0]['department']

from datetime import datetime
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Configure logger
logger = logging.getLogger(__name__)

faculty_att_bp = Blueprint('faculty_att', __name__)

# ─── ADMIN: SUBJECT-FACULTY ASSIGNMENT ───────────────────

@faculty_att_bp.route("/admin/faculty_assignments", methods=["GET", "POST"])
@login_required("admin")
def faculty_assignments():
    if request.method == "POST":
        action = request.form.get("action")
        
        if action == "assign":
            faculty_id = request.form.get("faculty_id")
            subject_name = request.form.get("subject_name")
            class_name = request.form.get("class_name")
            department = request.form.get("department")
            semester = request.form.get("semester")
            division = request.form.get("division")
            
            if not all([faculty_id, subject_name, class_name, department, semester, division]):
                flash("All fields are required for assignment.", "error")
            else:
                safe_execute("""
                    INSERT INTO faculty_subject_assignments 
                    (faculty_id, subject_name, class_name, department, semester, division)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (faculty_id, subject_name, class_name) 
                    DO UPDATE SET 
                        department = EXCLUDED.department,
                        semester = EXCLUDED.semester,
                        division = EXCLUDED.division
                """, (faculty_id, subject_name, class_name, department, semester, division))
                flash(f"Successfully assigned {subject_name} to class {class_name}", "success")
        
        elif action == "delete":
            assignment_id = request.form.get("id")
            safe_execute("DELETE FROM faculty_subject_assignments WHERE id = %s", (assignment_id,))
            flash("Assignment removed successfully", "info")
            
        return redirect(url_for('faculty_att.faculty_assignments'))

    # GET Request: Fetch data for the page
    assignments = safe_query("""
        SELECT a.*, f.name as faculty_name 
        FROM faculty_subject_assignments a 
        JOIN faculty f ON a.faculty_id = f.id 
        ORDER BY f.name, a.class_name
    """)
    
    faculty_list = safe_query("SELECT id, name, department FROM faculty ORDER BY name")
    subject_list = safe_query("SELECT DISTINCT name FROM subjects ORDER BY name")
    
    from config import DEPARTMENTS, SEMESTERS, DIVISIONS
    
    return render_template("admin/faculty_assignments.html", 
                         assignments=assignments, 
                         faculty_list=faculty_list, 
                         subject_list=subject_list,
                         DEPARTMENTS=DEPARTMENTS,
                         SEMESTERS=SEMESTERS,
                         DIVISIONS=DIVISIONS)


@faculty_att_bp.route("/admin/auto_link_faculty", methods=["POST"])
@login_required("admin")
def auto_link_faculty():
    # 1. Exact match update
    safe_execute("""
        UPDATE timetable t
        SET faculty_id = f.id
        FROM faculty f
        WHERE t.faculty_id IS NULL
        AND TRIM(REPLACE(t.teacher, ' ', '')) ILIKE TRIM(REPLACE(f.name, ' ', ''))
    """)
    
    # 2. Resilient Python fuzzy match fallback
    import re
    import difflib

    def clean(s):
        if not s: return ""
        s = re.sub(r'\b(PROF\.|DR\.|PROF|DR|MR\.|MRS\.|MS\.)\b', '', s.upper())
        s = re.sub(r'[^A-Z0-9]', '', s)
        return s.strip()

    def clean_words(s):
        if not s: return []
        words = re.findall(r'[A-Z]+', s.upper())
        return [w for w in words if w not in ('PROF', 'DR', 'MR', 'MRS', 'MS', 'KUMAR', 'TRA', 'IP')]

    def fuzzy_match_names(t_name, f_name):
        t_clean = clean(t_name)
        f_clean = clean(f_name)
        if not t_clean or not f_clean:
            return False
        if t_clean == f_clean:
            return True
        ratio = difflib.SequenceMatcher(None, t_clean, f_clean).ratio()
        if ratio >= 0.75:
            return True
        t_words = clean_words(t_name)
        f_words = clean_words(f_name)
        if t_words and f_words:
            w_ratio = difflib.SequenceMatcher(None, t_words[0], f_words[0]).ratio()
            if w_ratio >= 0.75:
                if len(t_words) > 1 and len(f_words) > 1:
                    l_ratio = difflib.SequenceMatcher(None, t_words[-1], f_words[-1]).ratio()
                    if l_ratio >= 0.75:
                        return True
                else:
                    return True
            if len(t_words) > 1 and len(f_words) > 1:
                if t_words[-1] == f_words[-1] and w_ratio >= 0.5:
                    return True
        return False

    unmatched_rows = safe_query("SELECT DISTINCT teacher FROM timetable WHERE faculty_id IS NULL")
    if unmatched_rows:
        faculties = safe_query("SELECT id, name FROM faculty")
        for ut in unmatched_rows:
            teacher_name = ut['teacher']
            if not teacher_name:
                continue
            for f in faculties:
                if fuzzy_match_names(teacher_name, f['name']):
                    safe_execute("""
                        UPDATE timetable 
                        SET faculty_id = %s 
                        WHERE teacher = %s AND faculty_id IS NULL
                    """, (f['id'], teacher_name))
                    logger.info(f"Fuzzy-matched timetable teacher '{teacher_name}' to faculty '{f['name']}' (ID: {f['id']})")
                    break

    # Check fixed count
    fixed_count = safe_fetch_scalar("SELECT COUNT(*) FROM timetable WHERE faculty_id IS NOT NULL")
    
    # Run the timetable -> faculty_subject_assignments insert SQL
    pre_count = safe_fetch_scalar("SELECT COUNT(*) FROM faculty_subject_assignments")
    
    safe_execute("""
        INSERT INTO faculty_subject_assignments (faculty_id, subject_name, class_name, department, semester, division)
        SELECT DISTINCT t.faculty_id, t.subject, (f.department || '-' || t.division) as class_name, f.department, t.semester, t.division
        FROM timetable t
        JOIN faculty f ON t.faculty_id = f.id
        WHERE t.faculty_id IS NOT NULL
        ON CONFLICT (faculty_id, subject_name, class_name) DO NOTHING
    """)
    
    # Get count after
    post_count = safe_fetch_scalar("SELECT COUNT(*) FROM faculty_subject_assignments")
    inserted_assignments = post_count - pre_count
    
    # Get remaining unmatched teachers
    remaining_unmatched_rows = safe_query("SELECT DISTINCT teacher FROM timetable WHERE faculty_id IS NULL")
    remaining_unmatched = [r['teacher'] for r in remaining_unmatched_rows]
    
    if remaining_unmatched:
        flash(f"Timetable linking complete! Inserted {inserted_assignments} new subject assignments. Unmatched: {', '.join(remaining_unmatched)}", "info")
    else:
        flash(f"Timetable linking complete! Inserted {inserted_assignments} new subject assignments.", "success")
        
    return redirect(url_for('faculty_att.faculty_assignments'))

@faculty_att_bp.route("/admin/fix_student_divisions", methods=["POST"])
@login_required("admin")
def fix_student_divisions():
    # Update students SET division using pattern from roll number if division is empty
    safe_execute("""
        UPDATE students 
        SET division = SUBSTRING(TRIM(roll) FROM 1 FOR 1) 
        WHERE (division IS NULL OR division = '') 
        AND roll IS NOT NULL 
        AND char_length(TRIM(roll)) > 0
    """)
    flash("Students divisions updated based on their roll numbers.", "success")
    return redirect("/students/")



# ─── FACULTY: PORTAL & ATTENDANCE FLOW ───────────────────

@faculty_att_bp.route("/faculty/attendance_portal")
@login_required("faculty")
def attendance_portal():
    fid = session.get("faculty_id")
    # Assigned subjects
    my_assignments = safe_query("""
        SELECT * FROM faculty_subject_assignments 
        WHERE faculty_id = %s 
        ORDER BY class_name, subject_name
    """, (fid,))
    
    # Session stats per assignment
    enhanced_assignments = []
    for a in my_assignments:
        count = safe_fetch_scalar("""
            SELECT COUNT(*) FROM attendance_sessions 
            WHERE faculty_id = %s AND subject = %s AND division = %s
        """, (fid, a['subject_name'], a['division']))
        a_dict = dict(a)
        a_dict['session_count'] = count
    # Recent sessions
    recent_sessions = safe_query("""
        SELECT * FROM attendance_sessions 
        WHERE faculty_id = %s 
        ORDER BY lecture_date DESC, created_at DESC 
        LIMIT 10
    """, (fid,))
    
    # Today's schedule from Master Timetable
    from datetime import datetime
    today_name = datetime.now().strftime("%A")
    faculty_name = session.get("name")
    
    # ─── SELF-HEALING LOGIC ───
    # We use a robust query with TRIM and similarity handles for name mismatches
    # Fallback to 0.3 similarity if exact trim match fails
    today_schedule = safe_query("""
        SELECT t.*, s.status as attendance_status, s.id as session_id
        FROM timetable t
        LEFT JOIN attendance_sessions s ON t.id = s.timetable_id AND s.lecture_date = CURRENT_DATE
        WHERE (
            t.faculty_id = %s 
            OR TRIM(REPLACE(t.teacher, ' ', '')) ILIKE TRIM(REPLACE(%s, ' ', ''))
            OR (SELECT EXISTS(SELECT 1 FROM pg_extension WHERE extname='pg_trgm') AND similarity(t.teacher, %s::text) > 0.4)
        ) 
        AND t.day = %s
        ORDER BY t.start_time
    """, (fid, faculty_name, faculty_name, today_name))

    # Auto-insert missing assignments to bridge the data gap automatically
    for slot in today_schedule:
        # Check if assignment exists
        exists = safe_fetch_one("""
            SELECT id FROM faculty_subject_assignments 
            WHERE faculty_id = %s AND subject_name = %s AND division = %s
        """, (fid, slot['subject'], slot['division']))
        
        if not exists:
            branch_val = slot.get('branch')
            if not branch_val:
                fac_dept_row = safe_fetch_one("SELECT department FROM faculty WHERE id = %s", (fid,))
                branch_val = fac_dept_row['department'] if fac_dept_row else 'CORE'
            branch_val = branch_val or 'CORE'
            class_name_val = f"{branch_val}-{slot['division']}" if branch_val else slot['division']
            
            safe_execute("""
                INSERT INTO faculty_subject_assignments (faculty_id, subject_name, class_name, department, semester, division)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT DO NOTHING
            """, (fid, slot['subject'], class_name_val, branch_val, slot['semester'] or '', slot['division']))
            # Refresh assignments if we added a new one
            my_assignments = safe_query("SELECT * FROM faculty_subject_assignments WHERE faculty_id = %s ORDER BY class_name, subject_name", (fid,))

    # Session stats per assignment (Recalculate after auto-insert)
    enhanced_assignments = []
    for a in my_assignments:
        count = safe_fetch_scalar("""
            SELECT COUNT(*) FROM attendance_sessions 
            WHERE faculty_id = %s AND subject = %s AND division = %s
        """, (fid, a['subject_name'], a['division']))
        a_dict = dict(a)
        a_dict['session_count'] = count
        enhanced_assignments.append(a_dict)
    
    return render_template("faculty/attendance_portal.html", 
                         assignments=enhanced_assignments, 
                         recent_sessions=recent_sessions,
                         today_schedule=today_schedule,
                         today_date=datetime.now().strftime("%Y-%m-%d"))

@faculty_att_bp.route("/faculty/create_session_from_slot/<int:slot_id>")
@login_required("faculty")
def create_session_from_slot(slot_id):
    fid = session.get("faculty_id")
    faculty_name = session.get("name", "")

    # Fetch slot details from master timetable
    slot = safe_fetch_one("SELECT * FROM timetable WHERE id = %s", (slot_id,))
    if not slot:
        flash("Timetable slot not found", "error")
        return redirect(url_for('faculty_att.attendance_portal'))

    # --- Security guard: Verify this slot belongs to the requesting faculty ---
    # Allow if faculty_id matches, or if teacher name matches (for name-based assignments)
    slot_faculty_id = slot.get('faculty_id')
    slot_teacher = (slot.get('teacher') or '').strip()
    name_match = faculty_name and slot_teacher and (
        slot_teacher.upper().replace(' ', '') == faculty_name.upper().replace(' ', '')
        or faculty_name.upper() in slot_teacher.upper()
        or slot_teacher.upper() in faculty_name.upper()
    )

    if slot_faculty_id != fid and not name_match:
        logger.warning(
            f"Faculty {fid} ({faculty_name}) tried to mark slot {slot_id} "
            f"owned by faculty_id={slot_faculty_id} (teacher='{slot_teacher}'). Blocked."
        )
        flash("You are not assigned to this timetable slot.", "error")
        return redirect(url_for('faculty_att.attendance_portal'))

    # Check if a session already exists for today
    today = datetime.now().strftime("%Y-%m-%d")
    existing = safe_fetch_one("""
        SELECT id FROM attendance_sessions 
        WHERE timetable_id = %s AND lecture_date = %s
    """, (slot_id, today))
    
    if existing:
        return redirect(url_for('faculty_att.marking_session', session_id=existing['id']))
        
    # Resolve division and branch based on faculty's actual assignments
    resolved_div, resolved_branch = resolve_assigned_class(fid, slot['subject'], slot['division'])
    
    # Create new session
    session_id = safe_fetch_scalar("""
        INSERT INTO attendance_sessions 
        (faculty_id, subject, lecture_date, lecture_type, time_slot, timetable_id, division, branch, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'draft')
        RETURNING id
    """, (fid, slot['subject'] or '', today, slot['slot_type'] or 'Theory', slot['time'] or '', slot_id, resolved_div or slot['division'] or '', resolved_branch or slot['branch'] or '',))
    
    return redirect(url_for('faculty_att.marking_session', session_id=session_id))

@faculty_att_bp.route("/faculty/create_session", methods=["POST"])
@login_required("faculty")
def create_session():
    fid = session.get("faculty_id")
    subject = request.form.get("subject")
    lecture_date = request.form.get("lecture_date") or datetime.now().strftime("%Y-%m-%d")
    lecture_type = request.form.get("lecture_type", "Theory")
    time_slot = request.form.get("time_slot", "")
    timetable_id = request.form.get("timetable_id")
    
    div = request.form.get("division")
    branch = request.form.get("department")
    
    if timetable_id:
        # Use timetable slot as reference
        tt = safe_fetch_one("SELECT * FROM timetable WHERE id = %s", (timetable_id,))
        if tt:
            resolved_div, resolved_branch = resolve_assigned_class(fid, tt['subject'], tt['division'])
            div = resolved_div or tt['division']
            branch = resolved_branch or tt['branch']
            subject = tt['subject'] # Ensure subject matches timetable
    
    if not div:
        flash("Could not determine class or division for this session", "error")
        return redirect(url_for('faculty_att.attendance_portal'))

    # If manual creation (no timetable_id), validate against weekly timetable
    if not timetable_id:
        try:
            dt = datetime.strptime(lecture_date, "%Y-%m-%d")
            day_of_week = dt.strftime("%A")
        except Exception:
            flash("Invalid lecture date format.", "error")
            return redirect(url_for('faculty_att.attendance_portal'))

        valid_slot = safe_fetch_one("""
            SELECT id FROM timetable 
            WHERE day = %s 
            AND TRIM(REPLACE(subject, ' ', '')) ILIKE TRIM(REPLACE(%s, ' ', ''))
            AND division = %s 
            AND (branch IS NULL OR %s IS NULL OR branch = '' OR %s = '' OR branch = %s)
            AND faculty_id = %s
        """, (day_of_week, subject, div, branch, branch, branch, fid))
        
        if not valid_slot:
            flash("Error: No matching slot found in the weekly timetable for this subject, division, and day.", "error")
            return redirect(url_for('faculty_att.attendance_portal'))
        
        timetable_id = valid_slot['id']

    # Prevent duplicate sessions (same subject + division + date)
    existing = safe_fetch_one("""
        SELECT id FROM attendance_sessions 
        WHERE subject = %s AND division = %s AND lecture_date = %s
    """, (subject, div, lecture_date))
    
    if existing:
        flash("Session already exists for this subject, division, and date.", "info")
        return redirect(url_for('faculty_att.marking_session', session_id=existing['id']))
    
    # Create new session
    session_id = safe_fetch_scalar("""
        INSERT INTO attendance_sessions (faculty_id, subject, division, branch, lecture_date, lecture_type, time_slot, status, timetable_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s, 'draft', %s)
        RETURNING id
    """, (fid, subject or '', div or '', branch or '', lecture_date, lecture_type or 'Theory', time_slot or '', timetable_id))
    
    log_audit(fid, 'CREATED', f"Created attendance session ID {session_id} for {subject}")
    
    return redirect(url_for('faculty_att.marking_session', session_id=session_id))

@faculty_att_bp.route("/faculty/new_session", methods=["GET", "POST"])
@login_required("faculty")
def new_session():
    fid = session.get("faculty_id")
    subject = request.values.get("subject")
    division = request.values.get("division")
    
    if not subject or not division:
        flash("Subject and Division are required for manual marking.", "error")
        return redirect(url_for('faculty_att.attendance_portal'))
        
    # Verify faculty assignment
    assignment = safe_fetch_one("""
        SELECT * FROM faculty_subject_assignments 
        WHERE faculty_id = %s AND subject_name = %s AND division = %s
    """, (fid, subject, division))
    
    if not assignment:
        flash("Error: You are not assigned to this subject and division.", "error")
        return redirect(url_for('faculty_att.attendance_portal'))

    today = datetime.now().strftime("%Y-%m-%d")
    day_of_week = datetime.now().strftime("%A")
    branch = assignment.get('department')

    valid_slot = safe_fetch_one("""
        SELECT id FROM timetable 
        WHERE day = %s 
        AND TRIM(REPLACE(subject, ' ', '')) ILIKE TRIM(REPLACE(%s, ' ', ''))
        AND division = %s 
        AND (branch IS NULL OR %s IS NULL OR branch = '' OR %s = '' OR branch = %s)
        AND faculty_id = %s
    """, (day_of_week, subject, division, branch, branch, branch, fid))
    
    if not valid_slot:
        flash("Error: No matching slot found in the weekly timetable for today.", "error")
        return redirect(url_for('faculty_att.attendance_portal'))
        
    # Prevent duplicate sessions (same subject + division + date)
    existing = safe_fetch_one("""
        SELECT id FROM attendance_sessions 
        WHERE subject = %s AND division = %s AND lecture_date = %s
    """, (subject, division, today))
    
    if existing:
        return redirect(url_for('faculty_att.marking_session', session_id=existing['id']))
        
    # Create new manual session
    session_id = safe_fetch_scalar("""
        INSERT INTO attendance_sessions (faculty_id, subject, division, branch, lecture_date, status, timetable_id)
        VALUES (%s, %s, %s, %s, %s, 'draft', %s)
        RETURNING id
    """, (fid, subject or '', division or '', branch or '', today, valid_slot['id']))
    
    if session_id:
        log_audit(fid, 'MANUAL_SESSION_CREATED', f"Created manual session {session_id} for {subject} ({division})")
        return redirect(url_for('faculty_att.marking_session', session_id=session_id))
    else:
        flash("Critical Error: Failed to initialize attendance session.", "error")
        return redirect(url_for('faculty_att.attendance_portal'))


@faculty_att_bp.route("/faculty/marking_session/<int:session_id>")
@login_required("faculty")
def marking_session(session_id):
    fid = session.get("faculty_id")
    sess = safe_fetch_one("SELECT * FROM attendance_sessions WHERE id = %s", (session_id,))
    
    if not sess or sess['faculty_id'] != fid:
        flash("Unauthorized or session not found", "error")
        return redirect(url_for('faculty_att.attendance_portal'))
    
    # --- RESILIENT STUDENT FETCH ---
    division_str = sess.get('division', '')
    branch_str = sess.get('branch', '')
    
    year, depts, div = parse_class_details(division_str, branch_str)
    logger.info(f"Marking session {session_id}: Original=(Branch '{branch_str}', Div '{division_str}') -> Parsed=(Year '{year}', Depts '{depts}', Div '{div}')")
    
    students = []
    
    # Try matching with parsed details
    if depts:
        dept_placeholders = ", ".join(["%s"] * len(depts))
        dept_params = [d.upper() for d in depts]
        
        # 1. Match year, depts, div
        if not students and year and div:
            students = safe_query(f"""
                SELECT id, name, roll, prn 
                FROM students 
                WHERE TRIM(UPPER(year)) = %s 
                  AND TRIM(UPPER(department)) IN ({dept_placeholders})
                  AND TRIM(UPPER(division)) = %s
                ORDER BY roll, name
            """, [year.upper()] + dept_params + [div.upper()])
            
        # 2. Match depts, div
        if not students and div:
            students = safe_query(f"""
                SELECT id, name, roll, prn 
                FROM students 
                WHERE TRIM(UPPER(department)) IN ({dept_placeholders})
                  AND TRIM(UPPER(division)) = %s
                ORDER BY roll, name
            """, dept_params + [div.upper()])
            
        # 3. Match year, depts
        if not students and year:
            students = safe_query(f"""
                SELECT id, name, roll, prn 
                FROM students 
                WHERE TRIM(UPPER(year)) = %s 
                  AND TRIM(UPPER(department)) IN ({dept_placeholders})
                ORDER BY roll, name
            """, [year.upper()] + dept_params)
            
        # 4. Match depts only
        if not students:
            students = safe_query(f"""
                SELECT id, name, roll, prn 
                FROM students 
                WHERE TRIM(UPPER(department)) IN ({dept_placeholders})
                ORDER BY roll, name
            """, dept_params)
            
    else:
        # If no department is parsed, match by division and year
        # 1. Match year, div
        if not students and year and div:
            students = safe_query("""
                SELECT id, name, roll, prn 
                FROM students 
                WHERE TRIM(UPPER(year)) = %s 
                  AND TRIM(UPPER(division)) = %s
                ORDER BY roll, name
            """, (year.upper(), div.upper()))
            
        # 2. Match div only
        if not students and div:
            students = safe_query("""
                SELECT id, name, roll, prn 
                FROM students 
                WHERE TRIM(UPPER(division)) = %s
                ORDER BY roll, name
            """, (div.upper(),))
            
        # 3. Match year only
        if not students and year:
            students = safe_query("""
                SELECT id, name, roll, prn 
                FROM students 
                WHERE TRIM(UPPER(year)) = %s
                ORDER BY roll, name
            """, (year.upper(),))
            
    # Standard string normalization fallbacks for backward compatibility (only if branch specified)
    if not students:
        norm_branch = normalize_branch(branch_str)
        norm_div = normalize_division(division_str)
        if norm_branch:
            students = safe_query("""
                SELECT id, name, roll, prn 
                FROM students 
                WHERE TRIM(UPPER(division)) = TRIM(UPPER(%s)) 
                AND TRIM(UPPER(department)) = TRIM(UPPER(%s)) 
                ORDER BY roll, name
            """, (norm_div, norm_branch))
            
            if not students:
                students = safe_query("""
                    SELECT id, name, roll, prn 
                    FROM students 
                    WHERE TRIM(UPPER(division)) = TRIM(UPPER(%s))
                    AND (TRIM(UPPER(department)) ILIKE '%%' || TRIM(UPPER(%s)) || '%%' OR TRIM(UPPER(%s)) ILIKE '%%' || TRIM(UPPER(department)) || '%%')
                    ORDER BY roll, name
                """, (norm_div, norm_branch, norm_branch))
        else:
            students = safe_query("""
                SELECT id, name, roll, prn 
                FROM students 
                WHERE TRIM(UPPER(division)) = TRIM(UPPER(%s))
                ORDER BY roll, name
            """, (norm_div,))

    # Ultimate safety net: fetch first 100 students so page is not blank
    if not students:
        logger.warning("All student fetch fallbacks failed. Fetching first 100 students as safety net.")
        students = safe_query("""
            SELECT id, name, roll, prn 
            FROM students 
            ORDER BY name 
            LIMIT 100
        """)
    else:
        logger.info(f"Successfully loaded {len(students)} students.")
    
    # Fetch existing markings
    marks = safe_query("SELECT student_id, status FROM attendance WHERE lecture_id = %s", (session_id,))
    mark_map = {m['student_id']: m['status'] for m in marks}

    # Allow editing submitted sessions within 24 hours only
    from datetime import timedelta
    can_edit_submitted = False
    if sess['status'] == 'submitted':
        time_ref = sess.get('updated_at') or sess.get('created_at')
        if time_ref and (datetime.now() - time_ref) <= timedelta(hours=24):
            can_edit_submitted = True
    
    return render_template("faculty/marking_session.html", 
                         sess=sess, 
                         students=students, 
                         mark_map=mark_map,
                         can_edit_submitted=can_edit_submitted)

@faculty_att_bp.route("/faculty/api/save_attendance", methods=["POST"])
@login_required("faculty")
def api_save_attendance():
    data = request.get_json()
    session_id = data.get("session_id")
    markings = data.get("markings", {}) # student_id -> status
    fid = session.get("faculty_id")
    
    sess = safe_fetch_one("SELECT * FROM attendance_sessions WHERE id = %s", (session_id,))
    if not sess or sess['faculty_id'] != fid:
        return jsonify({"ok": False, "error": "Unauthorized"}), 403

    # Allow editing submitted session within 24 hours only
    from datetime import timedelta
    can_edit = False
    if sess['status'] == 'draft':
        can_edit = True
    elif sess['status'] == 'submitted':
        time_ref = sess.get('updated_at') or sess.get('created_at')
        if time_ref and (datetime.now() - time_ref) <= timedelta(hours=24):
            can_edit = True

    if not can_edit:
        return jsonify({"ok": False, "error": "This session is locked and cannot be edited anymore (24-hour window passed)."}), 403
    
    if not markings:
        return jsonify({"ok": True})

    # ── Fetch all existing records in ONE query ────────────────────────────
    student_ids = [int(k) for k in markings.keys()]
    placeholders = ",".join(["%s"] * len(student_ids))
    existing_rows = safe_query(
        f"SELECT student_id, status FROM attendance WHERE lecture_id = %s AND student_id IN ({placeholders})",
        (session_id, *student_ids)
    ) or []
    old_status_map = {str(r['student_id']): r['status'] for r in existing_rows}

    # ── Build the lists of changes ─────────────────────────────────────────
    upsert_rows  = []   # (sid, subject, date, status, fid, session_id, branch, division, sid)
    audit_rows   = []   # (fid, session_id, sid, old_status, new_status)

    faculty_row = safe_fetch_one("SELECT name FROM faculty WHERE id = %s", (fid,))
    faculty_name = faculty_row['name'] if faculty_row else "Unknown"
    time_slot = sess.get('time_slot')

    for sid_str, status in markings.items():
        sid = int(sid_str)
        old_status = old_status_map.get(sid_str)
        if old_status != status:
            upsert_rows.append((sid, sess['subject'], sess['lecture_date'], status,
                                fid, session_id, sess['branch'], sess['division'],
                                faculty_name, time_slot, sid))
            audit_rows.append((fid, session_id, sid, old_status, status))

    if upsert_rows:
        from utils.pg_wrapper import get_db
        conn = None
        try:
            conn = get_db()
            cur  = conn.cursor()
            if hasattr(cur, 'mogrify'):
                # Build multi-row INSERT … SELECT with a single round-trip
                args_str = " UNION ALL ".join(
                    cur.mogrify(
                        "(SELECT %s, name, %s, %s::date, %s, %s, %s, %s, %s, %s, %s FROM students WHERE id = %s)",
                        row
                    ).decode()
                    for row in upsert_rows
                )
                cur.execute(f"""
                    INSERT INTO attendance (student_id, student_name, subject, date, status, faculty_id, lecture_id, branch, division, faculty, time_slot)
                    {args_str}
                    ON CONFLICT (student_id, lecture_id) DO UPDATE
                        SET status = EXCLUDED.status, updated_at = CURRENT_TIMESTAMP,
                            faculty = EXCLUDED.faculty, time_slot = EXCLUDED.time_slot
                """)
            else:
                for row in upsert_rows:
                    cur.execute("""
                        INSERT INTO attendance (student_id, student_name, subject, date, status, faculty_id, lecture_id, branch, division, faculty, time_slot)
                        SELECT ?, name, ?, ?, ?, ?, ?, ?, ?, ?, ? FROM students WHERE id = ?
                        ON CONFLICT (student_id, lecture_id) DO UPDATE
                            SET status = EXCLUDED.status, updated_at = CURRENT_TIMESTAMP,
                                faculty = EXCLUDED.faculty, time_slot = EXCLUDED.time_slot
                    """, row)

            # ── Single bulk audit insert ───────────────────────────────────
            if audit_rows:
                if hasattr(cur, 'mogrify'):
                    from psycopg2.extras import execute_values
                    execute_values(cur, """
                        INSERT INTO attendance_audit (faculty_id, session_id, student_id, action, prev_status, new_status)
                        VALUES %s
                    """, [(r[0], r[1], r[2], 'STATUS_CHANGE', r[3], r[4]) for r in audit_rows])
                else:
                    cur.executemany("""
                        INSERT INTO attendance_audit (faculty_id, session_id, student_id, action, prev_status, new_status)
                        VALUES (?, ?, ?, 'STATUS_CHANGE', ?, ?)
                    """, [(r[0], r[1], r[2], r[3], r[4]) for r in audit_rows])

            conn.commit()
            cur.close()
        except Exception as e:
            if conn:
                conn.rollback()
            logger.error(f"Bulk save_attendance failed: {e}")
            return jsonify({"ok": False, "error": "DB error during save"}), 500
        finally:
            if conn:
                conn.close()

    # ── Update session timestamp ───────────────────────────────────────────
    safe_execute("UPDATE attendance_sessions SET updated_at = CURRENT_TIMESTAMP WHERE id = %s", (session_id,))

    try:
        from services.admin_notification_service import admin_notifier
        present = sum(1 for s in markings.values() if s == 'Present')
        total   = len(markings)
        admin_notifier.notify_admin(
            event_type     = 'attendance_submitted',
            faculty_id     = session.get('faculty_id'),
            faculty_name   = session.get('name', 'Unknown'),
            subject        = sess['subject'],
            division       = sess['division'],
            date           = sess['lecture_date'],
            total_students = total,
            present_count  = present,
        )
    except Exception:
        pass

    return jsonify({"ok": True})

@faculty_att_bp.route("/faculty/submit_session", methods=["POST"])
@login_required("faculty")
def submit_session():
    session_id = request.form.get("session_id")
    fid = session.get("faculty_id")
    
    sess = safe_fetch_one("SELECT * FROM attendance_sessions WHERE id = %s", (session_id,))
    if not sess or sess['faculty_id'] != fid:
        return "Unauthorized", 403
        
    safe_execute("UPDATE attendance_sessions SET status = 'submitted', updated_at = CURRENT_TIMESTAMP, is_locked = TRUE WHERE id = %s", (session_id,))
    log_audit(fid, 'SUBMITTED', f"Submitted session {session_id}")
    
    # Trigger background tasks
    from tasks.attendance_tasks import process_submitted_attendance
    process_submitted_attendance.delay(session_id)
    
    flash("Attendance submitted and locked successfully", "success")
    return redirect(url_for('faculty_att.attendance_portal'))

@faculty_att_bp.route("/faculty/session_history")
@login_required("faculty")
def session_history():
    fid = session.get("faculty_id")
    subject = request.args.get("subject", "").strip()
    class_name = request.args.get("class", "").strip()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    
    sql = "SELECT * FROM attendance_sessions WHERE faculty_id = %s"
    params = [fid]
    
    if subject: sql += " AND subject = %s"; params.append(subject)
    if class_name: sql += " AND division = %s"; params.append(class_name)
    if start_date: sql += " AND lecture_date >= %s"; params.append(start_date)
    if end_date: sql += " AND lecture_date <= %s"; params.append(end_date)
    
    sql += " ORDER BY lecture_date DESC"
    sessions = safe_query(sql, params)
    
    # Calculate attendance % for each session
    enhanced = []
    for s in sessions:
        total = safe_fetch_scalar("SELECT COUNT(*) FROM attendance WHERE lecture_id = %s", (s['id'],))
        present = safe_fetch_scalar("SELECT COUNT(*) FROM attendance WHERE lecture_id = %s AND status = 'Present'", (s['id'],))
        s_dict = dict(s)
        s_dict['attendance_pct'] = round(present/total*100, 1) if total > 0 else 0
        s_dict['total_students'] = total
        s_dict['present_students'] = present
        enhanced.append(s_dict)
        
    my_subjects = safe_query("SELECT DISTINCT subject_name FROM faculty_subject_assignments WHERE faculty_id = %s", (fid,))
    
    return render_template("faculty/session_history.html", sessions=enhanced, subjects=my_subjects)

@faculty_att_bp.route("/faculty/export_session/<int:session_id>")
@login_required("faculty")
def export_session(session_id):
    fid = session.get("faculty_id")
    sess = safe_fetch_one("SELECT * FROM attendance_sessions WHERE id = %s", (session_id,))
    if not sess or sess['faculty_id'] != fid:
        return "Unauthorized", 403
        
    records = safe_query("""
        SELECT s.name, s.roll, s.prn, a.status 
        FROM attendance a 
        JOIN students s ON a.student_id = s.id 
        WHERE a.lecture_id = %s 
        ORDER BY s.roll, s.name
    """, (session_id,))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    ws.append([f"Attendance Report: {sess['subject']} ({sess['lecture_type']})"])
    ws.append([f"Date: {sess['lecture_date']}", f"Class: {sess['branch']} {sess['division']}"])
    ws.append([]) # Blank
    
    headers = ["Roll No", "PRN", "Student Name", "Status"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
        
    for r in records:
        ws.append([r['roll'], r['prn'], r['name'], r['status']])
        
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"Attendance_{sess['subject']}_{sess['lecture_date']}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename, 
                   mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@faculty_att_bp.route("/faculty/my_students")
@login_required("faculty")
def my_students():
    fid = session.get("faculty_id")
    # Get all assignments for this faculty
    assignments = safe_query("SELECT division, department FROM faculty_subject_assignments WHERE faculty_id = %s", (fid,))
    
    students = []
    if assignments:
        # Collect conditions
        conditions = []
        params = []
        
        for a in assignments:
            division_str = a.get('division', '')
            branch_str = a.get('department', '')
            
            year, depts, div = parse_class_details(division_str, branch_str)
            
            # For each assignment, try to generate a specific matching clause
            clauses = []
            if year:
                clauses.append("TRIM(UPPER(year)) = %s")
                params.append(year.upper())
            if depts:
                placeholder = ", ".join(["%s"] * len(depts))
                clauses.append(f"TRIM(UPPER(department)) IN ({placeholder})")
                params.extend([d.upper() for d in depts])
            if div:
                clauses.append("TRIM(UPPER(division)) = %s")
                params.append(div.upper())
                
            if clauses:
                conditions.append(f"({' AND '.join(clauses)})")
                
        if conditions:
            # Query students matching any of the assignment clauses
            students = safe_query(f"""
                SELECT DISTINCT id, name, roll, prn, department, division, year
                FROM students
                WHERE {' OR '.join(conditions)}
                ORDER BY division, roll
            """, params)
            
        # Fallback 1: If no students found through parsing, try exact string joins
        if not students:
            students = safe_query("""
                SELECT DISTINCT s.id, s.name, s.roll, s.prn, s.department, s.division, s.year
                FROM students s
                JOIN faculty_subject_assignments a ON s.division = a.division AND s.department = a.department
                WHERE a.faculty_id = %s
                ORDER BY s.division, s.roll
            """, (fid,))
            
        # Fallback 2: Try loose division match
        if not students:
            # Fetch all assigned departments from faculty_subject_assignments to restrict fallback
            depts_assigned = set()
            for a in assignments:
                division_str = a.get('division', '')
                branch_str = a.get('department', '')
                _, parsed_depts, _ = parse_class_details(division_str, branch_str)
                for pd in parsed_depts:
                    depts_assigned.add(pd.upper())
            
            # Look for any student where division or department letters/words match roughly
            divs = [a.get('division') for a in assignments if a.get('division')]
            if divs:
                # Extract first parsed division letter from each assignment
                parsed_divs = []
                for d in divs:
                    _, _, div = parse_class_details(d, "")
                    if div:
                        parsed_divs.append(div.upper())
                if parsed_divs:
                    placeholder = ", ".join(["%s"] * len(parsed_divs))
                    if depts_assigned:
                        dept_placeholder = ", ".join(["%s"] * len(depts_assigned))
                        students = safe_query(f"""
                            SELECT DISTINCT id, name, roll, prn, department, division, year
                            FROM students
                            WHERE TRIM(UPPER(division)) IN ({placeholder})
                              AND TRIM(UPPER(department)) IN ({dept_placeholder})
                            ORDER BY division, roll
                        """, parsed_divs + list(depts_assigned))
                    else:
                        students = safe_query(f"""
                            SELECT DISTINCT id, name, roll, prn, department, division, year
                            FROM students
                            WHERE TRIM(UPPER(division)) IN ({placeholder})
                            ORDER BY division, roll
                        """, parsed_divs)
                    
    # Ultimate fallback if no students assigned/found
    if not students:
        students = safe_query("""
            SELECT id, name, roll, prn, department, division, year 
            FROM students 
            ORDER BY division, name 
            LIMIT 100
        """)
    
    # Calculate attendance % across all subjects taught by this faculty for each student
    enhanced = []
    for s in students:
        stats = safe_fetch_one("""
            SELECT COUNT(*) as total, 
                   SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present
            FROM attendance
            WHERE student_id = %s AND faculty_id = %s
        """, (s['id'], fid))
        s_dict = dict(s)
        s_dict['attendance_pct'] = round(stats['present']/stats['total']*100, 1) if stats['total'] > 0 else 0
        enhanced.append(s_dict)
        
    return render_template("faculty/my_students.html", students=enhanced)

@faculty_att_bp.route("/faculty/student_report/<int:student_id>")
@login_required("faculty")
def faculty_student_report(student_id):
    fid = session.get("faculty_id")
    student = safe_fetch_one("SELECT * FROM students WHERE id = %s", (student_id,))
    
    # Subject breakdown directly from attendance table
    subjects_data_raw = safe_query("""
        SELECT subject, 
             COUNT(*) as total,
             COUNT(*) FILTER (WHERE status='Present') as present,
             COUNT(*) FILTER (WHERE status='Absent') as absent,
             COUNT(*) FILTER (WHERE status='Leave') as leave,
             ROUND(COUNT(*) FILTER (WHERE status='Present') * 100.0 / NULLIF(COUNT(*),0), 1) as percentage
        FROM attendance 
        WHERE student_id = %s AND faculty_id = %s
        GROUP BY subject
    """, (student_id, fid))
    
    subjects_data = []
    total_sessions = 0
    total_present = 0
    total_absent = 0
    total_leave = 0
    
    for row in subjects_data_raw:
        total_sessions += row['total']
        total_present += row['present']
        total_absent += row['absent']
        total_leave += row['leave']
        row_dict = dict(row)
        row_dict['percentage'] = float(row['percentage']) if row['percentage'] is not None else 0.0
        subjects_data.append(row_dict)
        
    overall_pct = round((total_present / total_sessions * 100.0), 1) if total_sessions > 0 else 0.0
    
    if total_sessions == 0:
        grade = "no_data"
    elif overall_pct >= 75:
        grade = "good"
    elif overall_pct >= 60:
        grade = "warning"
    else:
        grade = "defaulter"
        
    recent_history = safe_query("""
        SELECT date, subject, status 
        FROM attendance 
        WHERE student_id = %s AND faculty_id = %s
        ORDER BY date DESC, id DESC LIMIT 10
    """, (student_id, fid))

    return render_template("faculty/student_report.html", 
        student=student, 
        subjects_data=subjects_data, 
        overall_pct=overall_pct, 
        total_sessions=total_sessions,
        total_present=total_present, 
        total_absent=total_absent,
        total_leave=total_leave,
        grade=grade,
        history=recent_history
    )
