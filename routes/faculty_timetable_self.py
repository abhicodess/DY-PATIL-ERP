import json
from datetime import datetime
from flask import Blueprint, request, jsonify, session, render_template, abort
from blueprints.auth.decorators import login_required
from utils.pg_wrapper import qry, qone, exe, get_tenant_db
from services.admin_notification_service import admin_notifier

faculty_timetable_self_bp = Blueprint('faculty_timetable_self', __name__)

from extensions import csrf, limiter, db
csrf.exempt(faculty_timetable_self_bp)

def safe_execute(cur, sql, params=None):
    from utils.pg_wrapper import _prepare_query_and_params, db
    is_postgres = db.engine.dialect.name == 'postgresql'
    sql, params = _prepare_query_and_params(sql, params, is_postgres)
    if params is not None:
        cur.execute(sql, params)
    else:
        cur.execute(sql)

def serialize_row(row):
    if not row:
        return None
    d = dict(row)
    for k, v in d.items():
        if hasattr(v, 'isoformat'):
            d[k] = v.isoformat()
    return d

def serialize_rows(rows):
    return [serialize_row(r) for r in rows]

# ── GET /api/faculty/my-timetable ──────────────────────────────
@faculty_timetable_self_bp.route('/api/faculty/my-timetable', methods=['GET'])
@login_required('faculty')
def get_my_timetable():
    faculty_id = session.get('faculty_id')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    rows = qry("""
        SELECT * FROM faculty_timetable
        WHERE faculty_id = %s
        ORDER BY
          CASE day
            WHEN 'Monday'    THEN 1
            WHEN 'Tuesday'   THEN 2
            WHEN 'Wednesday' THEN 3
            WHEN 'Thursday'  THEN 4
            WHEN 'Friday'    THEN 5
            WHEN 'Saturday'  THEN 6
          END, time_slot
    """, (faculty_id,))

    slots = serialize_rows(rows)
    
    slots_by_day = {
        'Monday': [],
        'Tuesday': [],
        'Wednesday': [],
        'Thursday': [],
        'Friday': [],
        'Saturday': []
    }
    
    total = len(slots)
    pending = 0
    approved = 0
    
    for s in slots:
        day = s.get('day')
        if day in slots_by_day:
            slots_by_day[day].append(s)
        status = s.get('status')
        if status == 'pending':
            pending += 1
        elif status == 'approved':
            approved += 1

    return jsonify({
        'slots_by_day': slots_by_day,
        'total': total,
        'pending': pending,
        'approved': approved
    }), 200

# ── POST /api/faculty/my-timetable ─────────────────────────────
@faculty_timetable_self_bp.route('/api/faculty/my-timetable', methods=['POST'])
@login_required('faculty')
@limiter.limit("30 per hour")
def add_my_timetable_slot():
    faculty_id = session.get('faculty_id')
    faculty_name = session.get('name', '')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    body = request.get_json() or {}
    day = body.get('day')
    time_slot = body.get('time_slot')
    subject = body.get('subject')
    division = body.get('division')
    slot_type = body.get('slot_type')
    room = body.get('room', '')
    semester = body.get('semester', '')
    academic_year = body.get('academic_year', '')

    errors = {}
    if not day or day not in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']:
        errors['day'] = "Day is required and must be Monday-Saturday"
    if not time_slot:
        errors['time_slot'] = "Time slot is required"
    if not subject or len(subject.strip()) == 0 or len(subject) > 100:
        errors['subject'] = "Subject is required and max 100 characters"
    if not division:
        errors['division'] = "Division is required"
    if not slot_type or slot_type not in ['Theory', 'Lab', 'Elective', 'Minor']:
        errors['slot_type'] = "Slot type is required and must be Theory/Lab/Elective/Minor"

    if errors:
        return jsonify({"errors": errors}), 400

    # Duplicate check
    dup = qone("""
        SELECT id FROM faculty_timetable
        WHERE faculty_id = %s AND day = %s AND time_slot = %s
        AND status != 'rejected'
    """, (faculty_id, day, time_slot))
    
    if dup:
        return jsonify({
            "error": f"You already have a slot at {day} {time_slot}. Edit or delete it first."
        }), 409

    # Insert slot
    try:
        row = qone("""
            INSERT INTO faculty_timetable
            (faculty_id, faculty_name, day, time_slot, subject, division, room, slot_type, semester, academic_year, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'draft')
            RETURNING id
        """, (faculty_id, faculty_name, day, time_slot, subject, division, room, slot_type, semester, academic_year))
    except Exception as e:
        err_name = type(e).__name__
        if "IntegrityError" in err_name or "UniqueViolation" in err_name:
            return jsonify({"error": "You already have a slot at that day and time"}), 409
        raise e

    slot_id = row[0]
    new_slot = qone("SELECT * FROM faculty_timetable WHERE id = %s", (slot_id,))
    slot_dict = serialize_row(new_slot)

    return jsonify(slot_dict), 201

# ── PUT /api/faculty/my-timetable/<int:slot_id> ─────────────────
@faculty_timetable_self_bp.route('/api/faculty/my-timetable/<int:slot_id>', methods=['PUT'])
@login_required('faculty')
def update_my_timetable_slot(slot_id):
    faculty_id = session.get('faculty_id')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    slot = qone("SELECT * FROM faculty_timetable WHERE id = %s", (slot_id,))
    if not slot:
        return jsonify({"error": "Not Found"}), 404
    if slot['faculty_id'] != faculty_id:
        return jsonify({"error": "Forbidden"}), 403
    if slot['status'] not in ['draft', 'rejected']:
        return jsonify({"error": "Only draft or rejected slots can be edited."}), 403

    body = request.get_json() or {}
    day = body.get('day')
    time_slot = body.get('time_slot')
    subject = body.get('subject')
    division = body.get('division')
    slot_type = body.get('slot_type')
    room = body.get('room', '')
    semester = body.get('semester', '')
    academic_year = body.get('academic_year', '')

    errors = {}
    if not day or day not in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']:
        errors['day'] = "Day is required and must be Monday-Saturday"
    if not time_slot:
        errors['time_slot'] = "Time slot is required"
    if not subject or len(subject.strip()) == 0 or len(subject) > 100:
        errors['subject'] = "Subject is required and max 100 characters"
    if not division:
        errors['division'] = "Division is required"
    if not slot_type or slot_type not in ['Theory', 'Lab', 'Elective', 'Minor']:
        errors['slot_type'] = "Slot type is required and must be Theory/Lab/Elective/Minor"

    if errors:
        return jsonify({"errors": errors}), 400

    # Duplicate check for other slots
    dup = qone("""
        SELECT id FROM faculty_timetable
        WHERE faculty_id = %s AND day = %s AND time_slot = %s
        AND status != 'rejected' AND id != %s
    """, (faculty_id, day, time_slot, slot_id))
    
    if dup:
        return jsonify({
            "error": f"You already have a slot at {day} {time_slot}. Edit or delete it first."
        }), 409

    if slot['status'] == 'rejected':
        last_rejected_note = slot['admin_note'] or ''
    else:
        last_rejected_note = slot['last_rejected_note'] or ''

    exe("""
        UPDATE faculty_timetable
        SET day = %s, time_slot = %s, subject = %s, division = %s, room = %s,
            slot_type = %s, semester = %s, academic_year = %s, status = 'draft',
            admin_note = '', last_rejected_note = %s,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = %s AND faculty_id = %s
    """, (day, time_slot, subject, division, room, slot_type, semester, academic_year, last_rejected_note, slot_id, faculty_id))

    updated_slot = qone("SELECT * FROM faculty_timetable WHERE id = %s", (slot_id,))
    slot_dict = serialize_row(updated_slot)

    return jsonify(slot_dict), 200

# ── DELETE /api/faculty/my-timetable/<int:slot_id> ──────────────
@faculty_timetable_self_bp.route('/api/faculty/my-timetable/<int:slot_id>', methods=['DELETE'])
@login_required('faculty')
def delete_my_timetable_slot(slot_id):
    faculty_id = session.get('faculty_id')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    slot = qone("SELECT * FROM faculty_timetable WHERE id = %s", (slot_id,))
    if not slot:
        return jsonify({"error": "Not Found"}), 404
    if slot['faculty_id'] != faculty_id:
        return jsonify({"error": "Forbidden"}), 403
    if slot['status'] not in ['draft', 'rejected']:
        return jsonify({"error": "Only draft or rejected slots can be deleted."}), 403

    exe("DELETE FROM faculty_timetable WHERE id = %s AND faculty_id = %s", (slot_id, faculty_id))
    return '', 204

# ── POST /api/faculty/my-timetable/submit ──────────────────────
@faculty_timetable_self_bp.route('/api/faculty/my-timetable/submit', methods=['POST'])
@login_required('faculty')
@limiter.limit("10 per hour")
def submit_my_timetable():
    faculty_id = session.get('faculty_id')
    faculty_name = session.get('name', '')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    drafts = qry("""
        SELECT id, subject, day, time_slot, division, room, slot_type, semester,
               last_rejected_note, resubmission_count
        FROM faculty_timetable
        WHERE faculty_id = %s AND status IN ('draft', 'rejected')
    """, (faculty_id,))
    if not drafts:
        return jsonify({"error": "No draft or rejected slots found to submit."}), 400

    for draft in drafts:
        resub_count = draft['resubmission_count'] or 0
        last_note = draft['last_rejected_note'] or ''
        
        if last_note.strip() != '':
            resub_count += 1
            
        exe("""
            UPDATE faculty_timetable
            SET status = 'pending', resubmission_count = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (resub_count, draft['id']))

        try:
            if resub_count > 0:
                admin_notifier.notify_admin(
                    event_type         = 'timetable_resubmitted',
                    faculty_id         = faculty_id,
                    faculty_name       = faculty_name,
                    subject            = draft['subject'],
                    day                = draft['day'],
                    time_slot          = draft['time_slot'],
                    division           = draft['division'],
                    last_rejected_note = last_note,
                    resubmission_count = resub_count,
                )
            else:
                admin_notifier.notify_admin(
                    event_type         = 'timetable_submitted',
                    faculty_id         = faculty_id,
                    faculty_name       = faculty_name,
                )
        except:
            pass

    return jsonify({"message": f"Successfully submitted {len(drafts)} slots for approval."}), 200

# ── GET /api/faculty/divisions ──────────────────────────────────
@faculty_timetable_self_bp.route('/api/faculty/divisions', methods=['GET'])
def get_divisions():
    # Fetch from timetable
    t_rows = qry("SELECT DISTINCT division FROM timetable ORDER BY division")
    t_divs = [r[0] for r in t_rows if r[0]]
    
    # Fetch from faculty_subject_assignments
    a_rows = qry("SELECT DISTINCT division FROM faculty_subject_assignments ORDER BY division")
    a_divs = [r[0] for r in a_rows if r[0]]
    
    # Default list
    default_divs = ['SE CSE-A', 'SE CSE-B', 'SE AIDS', 'SE IT']
    
    # Merge and preserve uniqueness
    all_divs = []
    for d in default_divs + t_divs + a_divs:
        d_clean = str(d).strip()
        if d_clean and d_clean not in all_divs:
            all_divs.append(d_clean)
            
    return jsonify({"divisions": all_divs}), 200

# ── GET /api/admin/timetable-requests ───────────────────────────
@faculty_timetable_self_bp.route('/api/admin/timetable-requests', methods=['GET'])
@login_required('admin')
def get_admin_timetable_requests():
    rows = qry("""
        SELECT ft.*, f.name as faculty_display_name
        FROM faculty_timetable ft
        LEFT JOIN faculty f ON ft.faculty_id = f.id
        WHERE ft.status = 'pending'
        ORDER BY ft.created_at DESC
    """)
    requests = serialize_rows(rows)
    return jsonify({"requests": requests, "count": len(requests)}), 200

# ── PATCH /api/admin/timetable-requests/<int:slot_id>/approve ───
@faculty_timetable_self_bp.route('/api/admin/timetable-requests/<int:slot_id>/approve', methods=['PATCH'])
@login_required('admin')
def approve_timetable_request(slot_id):
    slot = qone("SELECT * FROM faculty_timetable WHERE id = %s", (slot_id,))
    if not slot:
        return jsonify({"error": "Not Found"}), 404
    if slot['status'] != 'pending':
        return jsonify({"error": "Already processed"}), 400

    from flask import current_app
    is_testing = False
    try:
        if current_app and current_app.config.get("TESTING"):
            is_testing = True
    except RuntimeError:
        pass

    if is_testing:
        conn = db.session.connection().connection
        cur = conn.cursor()
        try:
            safe_execute(cur, "UPDATE faculty_timetable SET status='approved', updated_at=CURRENT_TIMESTAMP WHERE id=%s", (slot_id,))
            safe_execute(cur, """
                INSERT INTO timetable (day, time, subject, teacher, room, slot_type, division, semester, faculty_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (slot['day'], slot['time_slot'], slot['subject'], slot['faculty_name'], slot['room'], slot['slot_type'], slot['division'], slot['semester'], slot['faculty_id']))
        finally:
            cur.close()
    else:
        conn = db.engine.raw_connection()
        try:
            with conn:  # auto-commits or rolls back
                cur = conn.cursor()
                safe_execute(cur, "UPDATE faculty_timetable SET status='approved', updated_at=CURRENT_TIMESTAMP WHERE id=%s", (slot_id,))
                safe_execute(cur, """
                    INSERT INTO timetable (day, time, subject, teacher, room, slot_type, division, semester, faculty_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (slot['day'], slot['time_slot'], slot['subject'], slot['faculty_name'], slot['room'], slot['slot_type'], slot['division'], slot['semester'], slot['faculty_id']))
        except Exception:
            conn.rollback()
            return jsonify({"error": "DB error"}), 500
        finally:
            conn.close()

    try:
        from datetime import datetime
        msg_body = f"Hello {slot['faculty_name']},\n\nYour weekly timetable slot request for {slot['subject']} on {slot['day']} ({slot['time_slot']}) has been APPROVED by the admin and is now active."
        exe("""
            INSERT INTO messages(from_role, from_id, from_name, to_role, to_id, to_name, subject, body, sent_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, ("admin", 0, "Admin Panel", "faculty", slot['faculty_id'], slot['faculty_name'], "Timetable Slot Approved", msg_body, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    except Exception as e:
        pass

    try:
        admin_notifier.notify_admin(
            event_type   = 'timetable_slot_approved',
            faculty_id   = slot['faculty_id'],
            faculty_name = slot['faculty_name'],
            subject      = slot['subject'],
            day          = slot['day'],
            time_slot    = slot['time_slot'],
        )
    except:
        pass

    return jsonify({"message": "Approved and added to master timetable"}), 200

# ── PATCH /api/admin/timetable-requests/<int:slot_id>/reject ────
@faculty_timetable_self_bp.route('/api/admin/timetable-requests/<int:slot_id>/reject', methods=['PATCH'])
@login_required('admin')
def reject_timetable_request(slot_id):
    body = request.get_json() or {}
    note = body.get('note', '').strip()
    if not note:
        return jsonify({"error": "rejection note is required"}), 400

    slot = qone("SELECT * FROM faculty_timetable WHERE id = %s", (slot_id,))
    if not slot:
        return jsonify({"error": "Not Found"}), 404
    if slot['status'] != 'pending':
        return jsonify({"error": "Already processed"}), 400

    exe("UPDATE faculty_timetable SET status='rejected', admin_note=%s, updated_at=CURRENT_TIMESTAMP WHERE id=%s", (note, slot_id))

    try:
        from datetime import datetime
        msg_body = f"Hello {slot['faculty_name']},\n\nYour weekly timetable slot request for {slot['subject']} on {slot['day']} ({slot['time_slot']}) was NOT approved by the admin.\n\nReason: {note}"
        exe("""
            INSERT INTO messages(from_role, from_id, from_name, to_role, to_id, to_name, subject, body, sent_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, ("admin", 0, "Admin Panel", "faculty", slot['faculty_id'], slot['faculty_name'], "Timetable Slot Rejected", msg_body, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    except Exception as e:
        pass

    try:
        admin_notifier.notify_admin(
            event_type   = 'timetable_slot_rejected',
            faculty_id   = slot['faculty_id'],
            faculty_name = slot['faculty_name'],
            subject      = slot['subject'],
            day          = slot['day'],
            time_slot    = slot['time_slot'],
            admin_note   = note,
        )
    except:
        pass

    return jsonify({"message": "Rejected"}), 200

# ── GET /api/admin/notifications ────────────────────────────────
@faculty_timetable_self_bp.route('/api/admin/notifications', methods=['GET'])
@login_required('admin')
def get_admin_notifications():
    unread_only = request.args.get('unread', 'false').lower() == 'true'
    
    if unread_only:
        rows = qry("SELECT * FROM admin_notifications WHERE is_read = FALSE ORDER BY created_at DESC LIMIT 50")
    else:
        rows = qry("SELECT * FROM admin_notifications ORDER BY created_at DESC LIMIT 50")

    notifications = serialize_rows(rows)
    
    count_row = qone("SELECT COUNT(*) FROM admin_notifications WHERE is_read = FALSE")
    unread_count = count_row[0] if count_row else 0

    return jsonify({"notifications": notifications, "unread_count": unread_count}), 200

# ── PATCH /api/admin/notifications/<int:id>/read ────────────────
@faculty_timetable_self_bp.route('/api/admin/notifications/<int:notif_id>/read', methods=['PATCH'])
@login_required('admin')
def mark_notification_read(notif_id):
    exe("UPDATE admin_notifications SET is_read=TRUE WHERE id=%s", (notif_id,))
    return jsonify({"message": "Marked as read"}), 200

# ── PATCH /api/admin/notifications/read-all ─────────────────────
@faculty_timetable_self_bp.route('/api/admin/notifications/read-all', methods=['PATCH'])
@login_required('admin')
def mark_all_notifications_read():
    row = qone("SELECT COUNT(*) FROM admin_notifications WHERE is_read=FALSE")
    unread_count = row[0] if row else 0
    exe("UPDATE admin_notifications SET is_read=TRUE WHERE is_read=FALSE")
    return jsonify({"updated": unread_count}), 200

# ── GET /admin/timetable-requests (View requests page) ──────────
@faculty_timetable_self_bp.route('/admin/timetable-requests', methods=['GET'])
@login_required('admin')
def admin_timetable_requests_page():
    return render_template('admin/timetable_requests.html')


# ================================================================
# COMPONENT 4 — Subject self-assignment API
# ================================================================

# ── POST /api/faculty/subjects/request ──────────────────────────
@faculty_timetable_self_bp.route('/api/faculty/subjects/request', methods=['POST'])
@login_required('faculty')
@limiter.limit("20 per hour")
def request_subject():
    faculty_id = session.get('faculty_id')
    faculty_name = session.get('name', '')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    body = request.get_json() or {}
    subject_name = body.get('subject_name')
    division = body.get('division')
    class_name = body.get('class_name')
    department = body.get('department')
    semester = body.get('semester')

    errors = {}
    for fld in ['subject_name', 'division', 'class_name', 'department', 'semester']:
        if not body.get(fld):
            errors[fld] = f"{fld} is required"

    if errors:
        return jsonify({"errors": errors}), 400

    # Duplicate check
    dup = qone("""
        SELECT id FROM faculty_subject_assignments
        WHERE faculty_id=%s AND subject_name=%s AND division=%s
        AND status != 'rejected'
    """, (faculty_id, subject_name, division))

    if dup:
        return jsonify({
            "error": f"Request already exists for {subject_name} / {division}"
        }), 409

    exe("""
        INSERT INTO faculty_subject_assignments
        (faculty_id, subject_name, division, class_name, department, semester, status, requested_at)
        VALUES (%s, %s, %s, %s, %s, %s, 'pending', CURRENT_TIMESTAMP)
    """, (faculty_id, subject_name, division, class_name, department, semester))

    try:
        admin_notifier.notify_admin(
            event_type   = 'subject_assignment_requested',
            faculty_id   = faculty_id,
            faculty_name = faculty_name,
            subject_name = subject_name,
            division     = division,
        )
    except:
        pass

    return '', 201

# ── GET /api/faculty/subjects ───────────────────────────────────
@faculty_timetable_self_bp.route('/api/faculty/subjects', methods=['GET'])
@login_required('faculty')
def get_my_subjects():
    faculty_id = session.get('faculty_id')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    rows = qry("""
        SELECT * FROM faculty_subject_assignments
        WHERE faculty_id=%s
        ORDER BY status, subject_name
    """, (faculty_id,))

    assignments = serialize_rows(rows)
    approved = [a for a in assignments if a.get('status') == 'approved']
    pending = [a for a in assignments if a.get('status') == 'pending']

    return jsonify({
        "approved": approved,
        "pending": pending
    }), 200

# ── DELETE /api/faculty/subjects/<int:assignment_id> ────────────
@faculty_timetable_self_bp.route('/api/faculty/subjects/<int:assignment_id>', methods=['DELETE'])
@login_required('faculty')
def delete_subject_request(assignment_id):
    faculty_id = session.get('faculty_id')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    assignment = qone("SELECT * FROM faculty_subject_assignments WHERE id = %s", (assignment_id,))
    if not assignment:
        return jsonify({"error": "Not Found"}), 404
    if assignment['faculty_id'] != faculty_id:
        return jsonify({"error": "Forbidden"}), 403
    if assignment['status'] == 'approved':
        return jsonify({"error": "Approved assignments cannot be removed. Contact admin."}), 403

    exe("DELETE FROM faculty_subject_assignments WHERE id=%s AND faculty_id=%s AND status='pending'", (assignment_id, faculty_id))
    return '', 204

# ── GET /api/admin/subject-requests/pending ─────────────────────
@faculty_timetable_self_bp.route('/api/admin/subject-requests/pending', methods=['GET'])
@login_required('admin')
def get_admin_pending_subject_requests():
    rows = qry("""
        SELECT fsa.*, f.name as faculty_display_name
        FROM faculty_subject_assignments fsa
        LEFT JOIN faculty f ON fsa.faculty_id = f.id
        WHERE fsa.status = 'pending'
        ORDER BY fsa.requested_at DESC
    """)
    requests = serialize_rows(rows)
    return jsonify({"requests": requests}), 200

# ── PATCH /api/admin/subject-requests/<int:id>/approve ──────────
@faculty_timetable_self_bp.route('/api/admin/subject-requests/<int:req_id>/approve', methods=['PATCH'])
@login_required('admin')
def approve_subject_request(req_id):
    req = qone("SELECT * FROM faculty_subject_assignments WHERE id = %s", (req_id,))
    if not req:
        return jsonify({"error": "Not Found"}), 404

    exe("""
        UPDATE faculty_subject_assignments
        SET status='approved', updated_at=CURRENT_TIMESTAMP
        WHERE id=%s
    """, (req_id,))

    try:
        admin_notifier.notify_admin(
            event_type   = 'subject_assignment_approved',
            faculty_id   = req['faculty_id'],
            subject_name = req['subject_name'],
            division     = req['division'],
        )
    except:
        pass

    return jsonify({"message": "Approved"}), 200

# ── PATCH /api/admin/subject-requests/<int:id>/reject ───────────
@faculty_timetable_self_bp.route('/api/admin/subject-requests/<int:req_id>/reject', methods=['PATCH'])
@login_required('admin')
def reject_subject_request(req_id):
    body = request.get_json() or {}
    note = body.get('note', '').strip()
    if not note:
        return jsonify({"error": "Rejection note is required"}), 400

    req = qone("SELECT * FROM faculty_subject_assignments WHERE id = %s", (req_id,))
    if not req:
        return jsonify({"error": "Not Found"}), 404

    exe("""
        UPDATE faculty_subject_assignments
        SET status='rejected', admin_note=%s, updated_at=CURRENT_TIMESTAMP
        WHERE id=%s
    """, (note, req_id))

    try:
        admin_notifier.notify_admin(
            event_type   = 'subject_assignment_rejected',
            faculty_id   = req['faculty_id'],
            subject_name = req['subject_name'],
            division     = req['division'],
            admin_note   = note,
        )
    except:
        pass

    return jsonify({"message": "Rejected"}), 200


# ================================================================
# COMPONENT 2 & 3 — Exports & Admin Timetable Dashboard
# ================================================================

# ── GET /api/faculty/my-timetable/export/excel ──────────────────
@faculty_timetable_self_bp.route('/api/faculty/my-timetable/export/excel', methods=['GET'])
@login_required('faculty')
@limiter.limit("10 per hour")
def export_my_timetable_excel():
    faculty_id = session.get('faculty_id')
    faculty_name = session.get('name', 'Faculty')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    MAX_EXPORT_ROWS = 5000
    rows = qry("""
        SELECT * FROM faculty_timetable
        WHERE faculty_id = %s AND status = 'approved'
        ORDER BY
          CASE day
            WHEN 'Monday'    THEN 1
            WHEN 'Tuesday'   THEN 2
            WHEN 'Wednesday' THEN 3
            WHEN 'Thursday'  THEN 4
            WHEN 'Friday'    THEN 5
            WHEN 'Saturday'  THEN 6
          END, time_slot
        LIMIT %s
    """, (faculty_id, MAX_EXPORT_ROWS))
    
    slots = serialize_rows(rows)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Timetable"

    headers = ["Day", "Time", "Subject", "Division", "Room", "Type", "Semester"]
    ws.append(headers)

    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    for col_num in range(1, 8):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    widths = {"A": 12, "B": 14, "C": 30, "D": 15, "E": 12, "F": 10, "G": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_odd = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    theory_cnt = 0
    lab_cnt = 0
    elective_cnt = 0
    minor_cnt = 0
    active_days = set()

    for idx, s in enumerate(slots, start=2):
        row_data = [
            s.get('day', ''),
            s.get('time_slot', ''),
            s.get('subject', ''),
            s.get('division', ''),
            s.get('room', ''),
            s.get('slot_type', 'Theory'),
            s.get('semester', '')
        ]
        ws.append(row_data)
        
        if s.get('day'):
            active_days.add(s.get('day'))
        stype = s.get('slot_type', 'Theory')
        if stype == 'Theory': theory_cnt += 1
        elif stype == 'Lab': lab_cnt += 1
        elif stype == 'Elective': elective_cnt += 1
        elif stype == 'Minor': minor_cnt += 1

        fill = fill_odd if idx % 2 == 1 else fill_even
        for col_num in range(1, 8):
            cell = ws.cell(row=idx, column=col_num)
            cell.fill = fill
            cell.border = thin_border
            if col_num in [1, 2, 4, 5, 6, 7]:
                cell.alignment = center_align

    start_r = len(slots) + 4
    ws.cell(row=start_r, column=1, value="Total slots:").font = Font(bold=True)
    ws.cell(row=start_r, column=2, value=len(slots))

    ws.cell(row=start_r+1, column=1, value="Theory:").font = Font(bold=True)
    ws.cell(row=start_r+1, column=2, value=theory_cnt)
    ws.cell(row=start_r+1, column=3, value="Lab:").font = Font(bold=True)
    ws.cell(row=start_r+1, column=4, value=lab_cnt)
    ws.cell(row=start_r+1, column=5, value="Elective:").font = Font(bold=True)
    ws.cell(row=start_r+1, column=6, value=elective_cnt)
    ws.cell(row=start_r+1, column=7, value=f"Minor: {minor_cnt}").font = Font(bold=True)

    ws.cell(row=start_r+2, column=1, value="Active days:").font = Font(bold=True)
    ws.cell(row=start_r+2, column=2, value=len(active_days))

    ws.cell(row=start_r+3, column=1, value="Exported:").font = Font(bold=True)
    ws.cell(row=start_r+3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    ws.cell(row=start_r+4, column=1, value="Faculty:").font = Font(bold=True)
    ws.cell(row=start_r+4, column=2, value=faculty_name)

    if len(slots) == MAX_EXPORT_ROWS:
        ws.insert_rows(1)
        ws['A1'] = f"WARNING: Export limited to {MAX_EXPORT_ROWS} rows. Apply filters to get full data."
        ws['A1'].font = Font(color="FF0000", bold=True)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"{faculty_name.replace(' ', '_')}_timetable_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ── GET /api/faculty/my-timetable/export/pdf ───────────────────
@faculty_timetable_self_bp.route('/api/faculty/my-timetable/export/pdf', methods=['GET'])
@login_required('faculty')
@limiter.limit("10 per hour")
def export_my_timetable_pdf():
    faculty_id = session.get('faculty_id')
    faculty_name = session.get('name', 'Faculty')
    if not faculty_id:
        return jsonify({"error": "Unauthorized"}), 401

    MAX_EXPORT_ROWS = 5000
    rows = qry("""
        SELECT * FROM faculty_timetable
        WHERE faculty_id = %s AND status = 'approved'
        ORDER BY
          CASE day
            WHEN 'Monday'    THEN 1
            WHEN 'Tuesday'   THEN 2
            WHEN 'Wednesday' THEN 3
            WHEN 'Thursday'  THEN 4
            WHEN 'Friday'    THEN 5
            WHEN 'Saturday'  THEN 6
          END, time_slot
        LIMIT %s
    """, (faculty_id, MAX_EXPORT_ROWS))
    
    slots = serialize_rows(rows)

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from io import BytesIO
    from flask import send_file

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_page_number(num_pages)
                super().showPage()
            super().save()

        def draw_page_number(self, page_count):
            self.saveState()
            self.setFont("Helvetica", 9)
            self.setFillColor(colors.HexColor("#6B7280"))
            self.drawCentredString(420.9, 30, f"DY Patil University ERP | Page {self._pageNumber} of {page_count}")
            self.restoreState()

    out = BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=54
    )

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#1F2937"),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=15
    )
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor("#374151")
    )
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )
    summary_label_style = ParagraphStyle(
        'SumLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#1F2937")
    )
    summary_val_style = ParagraphStyle(
        'SumVal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#4B5563")
    )

    elements = []
    
    elements.append(Paragraph(f"My Weekly Timetable — {faculty_name}", title_style))
    elements.append(Paragraph(f"DY Patil University | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))

    if len(slots) == MAX_EXPORT_ROWS:
        warning_style = ParagraphStyle(
            'Warning',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.HexColor('#FF0000'),
            spaceAfter=10
        )
        elements.append(Paragraph(f"WARNING: Export limited to {MAX_EXPORT_ROWS} rows. Apply filters to get full data.", warning_style))

    headers = ["Day", "Time", "Subject", "Division", "Room", "Type", "Semester"]
    table_data = [[Paragraph(h, header_style) for h in headers]]

    theory_cnt = 0
    lab_cnt = 0
    elective_cnt = 0
    minor_cnt = 0
    active_days = set()

    for s in slots:
        if s.get('day'):
            active_days.add(s.get('day'))
        stype = s.get('slot_type', 'Theory')
        if stype == 'Theory': theory_cnt += 1
        elif stype == 'Lab': lab_cnt += 1
        elif stype == 'Elective': elective_cnt += 1
        elif stype == 'Minor': minor_cnt += 1

        table_data.append([
            Paragraph(s.get('day') or '', cell_style),
            Paragraph(s.get('time_slot') or '', cell_style),
            Paragraph(s.get('subject') or '', cell_style),
            Paragraph(s.get('division') or '', cell_style),
            Paragraph(s.get('room') or '', cell_style),
            Paragraph(s.get('slot_type') or 'Theory', cell_style),
            Paragraph(s.get('semester') or '', cell_style),
        ])

    col_widths = [80, 100, 220, 100, 80, 70, 70]
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7C3AED')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
    ])
    
    for idx in range(1, len(table_data)):
        bg = colors.HexColor('#F5F3FF') if idx % 2 == 1 else colors.white
        t_style.add('BACKGROUND', (0, idx), (-1, idx), bg)

    t = Table(table_data, colWidths=col_widths)
    t.setStyle(t_style)
    elements.append(t)
    elements.append(Spacer(1, 20))

    sum_data = [
        [
            Paragraph("Total slots:", summary_label_style), Paragraph(str(len(slots)), summary_val_style),
            Paragraph("Theory:", summary_label_style), Paragraph(str(theory_cnt), summary_val_style),
            Paragraph("Lab:", summary_label_style), Paragraph(str(lab_cnt), summary_val_style)
        ],
        [
            Paragraph("Active days:", summary_label_style), Paragraph(str(len(active_days)), summary_val_style),
            Paragraph("Elective:", summary_label_style), Paragraph(str(elective_cnt), summary_val_style),
            Paragraph("Minor:", summary_label_style), Paragraph(str(minor_cnt), summary_val_style)
        ],
        [
            Paragraph("Faculty:", summary_label_style), Paragraph(faculty_name, summary_val_style),
            Paragraph("", summary_label_style), Paragraph("", summary_val_style),
            Paragraph("", summary_label_style), Paragraph("", summary_val_style)
        ]
    ]
    sum_table = Table(sum_data, colWidths=[90, 100, 90, 100, 90, 100])
    sum_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(sum_table)

    doc.build(elements, canvasmaker=NumberedCanvas)
    out.seek(0)

    filename = f"{faculty_name.replace(' ', '_')}_timetable_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    return send_file(
        out,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )


# ── GET /api/admin/timetable/export/excel ──────────────────────
@faculty_timetable_self_bp.route('/api/admin/timetable/export/excel', methods=['GET'])
@login_required('admin')
@limiter.limit("10 per hour")
def export_admin_timetable_excel():
    export_type = request.args.get('type', 'master')
    if export_type not in ['master', 'faculty', 'pending']:
        return jsonify({"error": "Invalid export type. Must be 'master', 'faculty', or 'pending'."}), 400

    faculty_id = request.args.get('faculty_id')
    division = request.args.get('division')

    rows = []
    MAX_EXPORT_ROWS = 5000
    if export_type == 'master':
        where_clauses = []
        params = []
        if faculty_id:
            where_clauses.append("faculty_id = %s")
            params.append(int(faculty_id))
        if division:
            where_clauses.append("division = %s")
            params.append(division)
        
        sql = """
            SELECT day, time as time_slot, subject, teacher as faculty_name,
                   room, slot_type, division, semester
            FROM timetable
        """
        if where_clauses:
            sql += " WHERE " + " AND ".join(where_clauses)
        sql += """
            ORDER BY division,
                     CASE day
                       WHEN 'Monday'    THEN 1
                       WHEN 'Tuesday'   THEN 2
                       WHEN 'Wednesday' THEN 3
                       WHEN 'Thursday'  THEN 4
                       WHEN 'Friday'    THEN 5
                       WHEN 'Saturday'  THEN 6
                     END, time
            LIMIT %s
        """
        params.append(MAX_EXPORT_ROWS)
        rows = qry(sql, tuple(params))
    else:
        status_val = 'approved' if export_type == 'faculty' else 'pending'
        where_clauses = ["ft.status = %s"]
        params = [status_val]
        if faculty_id:
            where_clauses.append("ft.faculty_id = %s")
            params.append(int(faculty_id))
        if division:
            where_clauses.append("ft.division = %s")
            params.append(division)

        sql = """
            SELECT ft.day, ft.time_slot, ft.subject, f.name as faculty_name,
                   ft.room, ft.slot_type, ft.division, ft.semester
            FROM faculty_timetable ft
            LEFT JOIN faculty f ON ft.faculty_id = f.id
        """
        if where_clauses:
            sql += " WHERE " + " AND ".join(where_clauses)
        sql += """
            ORDER BY faculty_name,
                     CASE ft.day
                       WHEN 'Monday'    THEN 1
                       WHEN 'Tuesday'   THEN 2
                       WHEN 'Wednesday' THEN 3
                       WHEN 'Thursday'  THEN 4
                       WHEN 'Friday'    THEN 5
                       WHEN 'Saturday'  THEN 6
                     END, ft.time_slot
            LIMIT %s
        """
        params.append(MAX_EXPORT_ROWS)
        rows = qry(sql, tuple(params))

    slots = serialize_rows(rows)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from flask import send_file

    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Timetable"

    headers = ["Faculty Name", "Day", "Time", "Subject", "Division", "Room", "Type", "Semester"]
    ws1.append(headers)

    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, 9):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    widths = {"A": 25, "B": 12, "C": 14, "D": 30, "E": 15, "F": 12, "G": 10, "H": 10}
    for col, w in widths.items():
        ws1.column_dimensions[col].width = w

    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_odd = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    from collections import defaultdict
    div_stats = defaultdict(lambda: {"total": 0, "Theory": 0, "Lab": 0, "Elective": 0, "Minor": 0})

    for idx, s in enumerate(slots, start=2):
        row_data = [
            s.get('faculty_name') or 'N/A',
            s.get('day', ''),
            s.get('time_slot', ''),
            s.get('subject', ''),
            s.get('division', ''),
            s.get('room', ''),
            s.get('slot_type', 'Theory'),
            s.get('semester', '')
        ]
        ws1.append(row_data)

        div = s.get('division') or 'N/A'
        div_stats[div]["total"] += 1
        stype = s.get('slot_type', 'Theory')
        if stype in ["Theory", "Lab", "Elective", "Minor"]:
            div_stats[div][stype] += 1

        fill = fill_odd if idx % 2 == 1 else fill_even
        for col_num in range(1, 9):
            cell = ws1.cell(row=idx, column=col_num)
            cell.fill = fill
            cell.border = thin_border
            if col_num in [2, 3, 5, 6, 7, 8]:
                cell.alignment = center_align

    ws2 = wb.create_sheet(title="Summary")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 12

    ws2.append(["Export Type:", export_type])
    ws2.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws2.append(["Total rows:", len(slots)])
    ws2.append([])

    sum_headers = ["Division", "Total Slots", "Theory", "Lab", "Elective", "Minor"]
    ws2.append(sum_headers)
    
    for col_num in range(1, 7):
        cell = ws2.cell(row=5, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for s_idx, (div, counts) in enumerate(div_stats.items(), start=6):
        ws2.append([
            div,
            counts["total"],
            counts["Theory"],
            counts["Lab"],
            counts["Elective"],
            counts["Minor"]
        ])
        fill = fill_odd if s_idx % 2 == 1 else fill_even
        for col_num in range(1, 7):
            cell = ws2.cell(row=s_idx, column=col_num)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = center_align

    if len(slots) == MAX_EXPORT_ROWS:
        ws1.insert_rows(1)
        ws1['A1'] = f"WARNING: Export limited to {MAX_EXPORT_ROWS} rows. Apply filters to get full data."
        ws1['A1'].font = Font(color="FF0000", bold=True)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"DYPatil_timetable_{export_type}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ── GET /api/admin/timetable/export/pdf ────────────────────────
@faculty_timetable_self_bp.route('/api/admin/timetable/export/pdf', methods=['GET'])
@login_required('admin')
@limiter.limit("10 per hour")
def export_admin_timetable_pdf():
    export_type = request.args.get('type', 'master')
    if export_type not in ['master', 'faculty', 'pending']:
        return jsonify({"error": "Invalid export type. Must be 'master', 'faculty', or 'pending'."}), 400

    faculty_id = request.args.get('faculty_id')
    division = request.args.get('division')

    rows = []
    MAX_EXPORT_ROWS = 5000
    if export_type == 'master':
        where_clauses = []
        params = []
        if faculty_id:
            where_clauses.append("faculty_id = %s")
            params.append(int(faculty_id))
        if division:
            where_clauses.append("division = %s")
            params.append(division)
        
        sql = """
            SELECT day, time as time_slot, subject, teacher as faculty_name,
                   room, slot_type, division, semester
            FROM timetable
        """
        if where_clauses:
            sql += " WHERE " + " AND ".join(where_clauses)
        sql += """
            ORDER BY division,
                     CASE day
                       WHEN 'Monday'    THEN 1
                       WHEN 'Tuesday'   THEN 2
                       WHEN 'Wednesday' THEN 3
                       WHEN 'Thursday'  THEN 4
                       WHEN 'Friday'    THEN 5
                       WHEN 'Saturday'  THEN 6
                     END, time
            LIMIT %s
        """
        params.append(MAX_EXPORT_ROWS)
        rows = qry(sql, tuple(params))
    else:
        status_val = 'approved' if export_type == 'faculty' else 'pending'
        where_clauses = ["ft.status = %s"]
        params = [status_val]
        if faculty_id:
            where_clauses.append("ft.faculty_id = %s")
            params.append(int(faculty_id))
        if division:
            where_clauses.append("ft.division = %s")
            params.append(division)

        sql = """
            SELECT ft.day, ft.time_slot, ft.subject, f.name as faculty_name,
                   ft.room, ft.slot_type, ft.division, ft.semester
            FROM faculty_timetable ft
            LEFT JOIN faculty f ON ft.faculty_id = f.id
        """
        if where_clauses:
            sql += " WHERE " + " AND ".join(where_clauses)
        sql += """
            ORDER BY faculty_name,
                     CASE ft.day
                       WHEN 'Monday'    THEN 1
                       WHEN 'Tuesday'   THEN 2
                       WHEN 'Wednesday' THEN 3
                       WHEN 'Thursday'  THEN 4
                       WHEN 'Friday'    THEN 5
                       WHEN 'Saturday'  THEN 6
                     END, ft.time_slot
            LIMIT %s
        """
        params.append(MAX_EXPORT_ROWS)
        rows = qry(sql, tuple(params))

    slots = serialize_rows(rows)

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from io import BytesIO
    from flask import send_file

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.draw_page_number(num_pages)
                super().showPage()
            super().save()

        def draw_page_number(self, page_count):
            self.saveState()
            self.setFont("Helvetica", 9)
            self.setFillColor(colors.HexColor("#6B7280"))
            self.drawCentredString(420.9, 30, f"DY Patil University ERP | Page {self._pageNumber} of {page_count}")
            self.restoreState()

    out = BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=54
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=20, leading=24, spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14, spaceAfter=15
    )
    cell_style = ParagraphStyle(
        'CellText', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10
    )
    header_style = ParagraphStyle(
        'HeaderStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white
    )
    section_title_style = ParagraphStyle(
        'SecTitle', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=14, leading=18, spaceBefore=15, spaceAfter=8
    )

    elements = []
    
    elements.append(Paragraph("DY Patil University — Timetable Report", title_style))
    elements.append(Paragraph(f"Type: {export_type} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))

    if len(slots) == MAX_EXPORT_ROWS:
        warning_style = ParagraphStyle(
            'Warning',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.HexColor('#FF0000'),
            spaceAfter=10
        )
        elements.append(Paragraph(f"WARNING: Export limited to {MAX_EXPORT_ROWS} rows. Apply filters to get full data.", warning_style))

    headers = ["Faculty Name", "Day", "Time", "Subject", "Division", "Room", "Type", "Semester"]
    table_data = [[Paragraph(h, header_style) for h in headers]]

    from collections import defaultdict
    div_stats = defaultdict(lambda: {"total": 0, "Theory": 0, "Lab": 0, "Elective": 0, "Minor": 0})

    for s in slots:
        div = s.get('division') or 'N/A'
        div_stats[div]["total"] += 1
        stype = s.get('slot_type', 'Theory')
        if stype in ["Theory", "Lab", "Elective", "Minor"]:
            div_stats[div][stype] += 1

        table_data.append([
            Paragraph(s.get('faculty_name') or 'N/A', cell_style),
            Paragraph(s.get('day') or '', cell_style),
            Paragraph(s.get('time_slot') or '', cell_style),
            Paragraph(s.get('subject') or '', cell_style),
            Paragraph(s.get('division') or '', cell_style),
            Paragraph(s.get('room') or '', cell_style),
            Paragraph(s.get('slot_type') or 'Theory', cell_style),
            Paragraph(s.get('semester') or '', cell_style),
        ])

    col_widths = [130, 70, 90, 180, 90, 70, 70, 60]
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7C3AED')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
    ])
    
    if export_type == 'pending':
        row_bg = colors.HexColor('#FEF3C7')
    elif export_type == 'faculty':
        row_bg = colors.HexColor('#F0FDF4')
    else:
        row_bg = None

    for idx in range(1, len(table_data)):
        if row_bg:
            bg = row_bg
        else:
            bg = colors.HexColor('#F5F3FF') if idx % 2 == 1 else colors.white
        t_style.add('BACKGROUND', (0, idx), (-1, idx), bg)

    t = Table(table_data, colWidths=col_widths)
    t.setStyle(t_style)
    elements.append(t)

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Division Coverage Summary", section_title_style))

    sum_headers = ["Division", "Total Slots", "Theory", "Lab", "Elective", "Minor"]
    sum_table_data = [[Paragraph(h, header_style) for h in sum_headers]]
    
    for div, counts in div_stats.items():
        sum_table_data.append([
            Paragraph(div, cell_style),
            Paragraph(str(counts["total"]), cell_style),
            Paragraph(str(counts["Theory"]), cell_style),
            Paragraph(str(counts["Lab"]), cell_style),
            Paragraph(str(counts["Elective"]), cell_style),
            Paragraph(str(counts["Minor"]), cell_style)
        ])

    sum_col_widths = [130, 90, 90, 90, 90, 90]
    sum_t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7C3AED')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
    ])
    for idx in range(1, len(sum_table_data)):
        bg = colors.HexColor('#F5F3FF') if idx % 2 == 1 else colors.white
        sum_t_style.add('BACKGROUND', (0, idx), (-1, idx), bg)

    sum_table = Table(sum_table_data, colWidths=sum_col_widths)
    sum_table.setStyle(sum_t_style)
    elements.append(sum_table)

    doc.build(elements, canvasmaker=NumberedCanvas)
    out.seek(0)

    filename = f"DYPatil_timetable_{export_type}_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    return send_file(
        out,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )


# ── GET /api/admin/timetable/stats ─────────────────────────────
@faculty_timetable_self_bp.route('/api/admin/timetable/stats', methods=['GET'])
@login_required('admin')
def get_admin_timetable_stats():
    from extensions import db
    master_total = qone("SELECT COUNT(*) FROM timetable")[0] or 0
    faculty_approved = qone("SELECT COUNT(*) FROM faculty_timetable WHERE status = 'approved'")[0] or 0
    pending_review = qone("SELECT COUNT(*) FROM faculty_timetable WHERE status = 'pending'")[0] or 0
    
    is_postgres = db.engine.dialect.name == 'postgresql'
    if is_postgres:
        rejected_row = qone("SELECT COUNT(*) FROM faculty_timetable WHERE status = 'rejected' AND updated_at >= NOW() - INTERVAL '24 hours'")
    else:
        rejected_row = qone("SELECT COUNT(*) FROM faculty_timetable WHERE status = 'rejected' AND updated_at >= datetime('now', '-1 day')")
    rejected_today = rejected_row[0] if rejected_row else 0
    
    resubmissions = qone("SELECT COUNT(*) FROM faculty_timetable WHERE resubmission_count > 0")[0] or 0
    faculty_active = qone("SELECT COUNT(DISTINCT faculty_id) FROM faculty_timetable WHERE status = 'approved'")[0] or 0
    divisions_covered = qone("SELECT COUNT(DISTINCT division) FROM timetable")[0] or 0

    return jsonify({
        "master_total": master_total,
        "faculty_approved": faculty_approved,
        "pending_review": pending_review,
        "rejected_today": rejected_today,
        "resubmissions": resubmissions,
        "faculty_active": faculty_active,
        "divisions_covered": divisions_covered
    }), 200


# ── GET /api/admin/timetable/activity ──────────────────────────
@faculty_timetable_self_bp.route('/api/admin/timetable/activity', methods=['GET'])
@login_required('admin')
def get_admin_timetable_activity():
    limit = request.args.get('limit', 20, type=int)
    rows = qry("""
        SELECT
          an.created_at,
          an.event_type,
          an.faculty_name,
          an.message,
          ft.subject,
          ft.division,
          ft.day,
          ft.time_slot,
          ft.status
        FROM admin_notifications an
        LEFT JOIN faculty_timetable ft
          ON an.faculty_id = ft.faculty_id
          AND an.event_type IN (
            'timetable_submitted','timetable_resubmitted',
            'timetable_slot_approved','timetable_slot_rejected'
          )
        ORDER BY an.created_at DESC
        LIMIT %s
    """, (limit,))
    
    return jsonify(serialize_rows(rows)), 200


# ── GET /api/admin/timetable/coverage ──────────────────────────
@faculty_timetable_self_bp.route('/api/admin/timetable/coverage', methods=['GET'])
@login_required('admin')
def get_admin_timetable_coverage():
    rows = qry("SELECT division, day, slot_type FROM timetable")
    coverage = {}
    
    for r in rows:
        div = r['division'] or 'N/A'
        day = r['day'] or 'N/A'
        slot_type = r['slot_type'] or 'Theory'
        
        if div not in coverage:
            coverage[div] = {d: {"total": 0, "theory": 0, "lab": 0} for d in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]}
            
        if day in coverage[div]:
            coverage[div][day]["total"] += 1
            if slot_type.lower() == 'theory':
                coverage[div][day]["theory"] += 1
            elif slot_type.lower() == 'lab':
                coverage[div][day]["lab"] += 1
                
    return jsonify(coverage), 200


# ── GET /admin/timetable-dashboard ─────────────────────────────
@faculty_timetable_self_bp.route('/admin/timetable-dashboard', methods=['GET'])
@login_required('admin')
def admin_timetable_dashboard():
    return render_template('admin/timetable_dashboard.html')


# ── GET /api/admin/timetable-slots ─────────────────────────────
@faculty_timetable_self_bp.route('/api/admin/timetable-slots', methods=['GET'])
@login_required('admin')
def get_admin_timetable_slots():
    status_filter = request.args.get('status', 'all')
    if status_filter == 'approved':
        rows = qry("SELECT * FROM faculty_timetable WHERE status = 'approved' ORDER BY id DESC")
    elif status_filter == 'pending':
        rows = qry("SELECT * FROM faculty_timetable WHERE status = 'pending' ORDER BY id DESC")
    else:
        rows = qry("SELECT * FROM faculty_timetable ORDER BY id DESC")
    return jsonify(serialize_rows(rows)), 200


# ── PUT /api/admin/timetable-slots/<int:slot_id> ─────────────────
@faculty_timetable_self_bp.route('/api/admin/timetable-slots/<int:slot_id>', methods=['PUT'])
@login_required('admin')
def admin_override_slot(slot_id):
    slot = qone("SELECT * FROM faculty_timetable WHERE id = %s", (slot_id,))
    if not slot:
        return jsonify({"error": "Not Found"}), 404
        
    body = request.get_json() or {}
    day = body.get('day')
    time_slot = body.get('time_slot')
    subject = body.get('subject')
    division = body.get('division')
    slot_type = body.get('slot_type')
    room = body.get('room', '')
    semester = body.get('semester', '')
    status = body.get('status')

    errors = {}
    if not day or day not in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']:
        errors['day'] = "Day is required and must be Monday-Saturday"
    if not time_slot:
        errors['time_slot'] = "Time slot is required"
    if not subject or len(subject.strip()) == 0 or len(subject) > 100:
        errors['subject'] = "Subject is required and max 100 characters"
    if not division:
        errors['division'] = "Division is required"
    if not slot_type or slot_type not in ['Theory', 'Lab', 'Elective', 'Minor']:
        errors['slot_type'] = "Slot type is required and must be Theory/Lab/Elective/Minor"

    if errors:
        return jsonify({"errors": errors}), 400

    with get_tenant_db() as cur:
        # If it was already approved, delete old master timetable slot first
        if slot['status'] == 'approved':
            safe_execute(cur, """
                DELETE FROM timetable 
                WHERE day = %s AND time = %s AND teacher = %s AND division = %s AND subject = %s
            """, (slot['day'], slot['time_slot'], slot['faculty_name'], slot['division'], slot['subject']))
            
        new_status = status if status in ['approved', 'pending', 'rejected', 'draft'] else slot['status']
        safe_execute(cur, """
            UPDATE faculty_timetable
            SET day = %s, time_slot = %s, subject = %s, division = %s, room = %s, slot_type = %s, semester = %s, status = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (day, time_slot, subject, division, room, slot_type, semester, new_status, slot_id))
        
        # If the new status is approved, insert into master timetable
        if new_status == 'approved':
            safe_execute(cur, """
                INSERT INTO timetable (day, time, subject, teacher, room, slot_type, division, semester, faculty_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (day, time_slot, subject, slot['faculty_name'], room, slot_type, division, semester, slot['faculty_id']))

    try:
        from datetime import datetime
        msg_body = f"Hello {slot['faculty_name']},\n\nAn admin has directly modified your timetable slot for '{slot['subject']}' on {slot['day']} ({slot['time_slot']}). The updated slot is: {subject} on {day} ({time_slot}) and is currently {new_status}."
        exe("""
            INSERT INTO messages(from_role, from_id, from_name, to_role, to_id, to_name, subject, body, sent_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, ("admin", 0, "Admin Panel", "faculty", slot['faculty_id'], slot['faculty_name'], "Timetable Slot Modified by Admin", msg_body, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    except:
        pass

    return jsonify({"message": "Slot modified successfully by admin override"}), 200


# ── POST /api/admin/timetable-slots/<int:slot_id>/notify ─────────────────
@faculty_timetable_self_bp.route('/api/admin/timetable-slots/<int:slot_id>/notify', methods=['POST'])
@login_required('admin')
def notify_timetable_slot(slot_id):
    slot = qone("SELECT * FROM faculty_timetable WHERE id = %s", (slot_id,))
    if not slot:
        slot = qone("SELECT * FROM timetable WHERE id = %s", (slot_id,))
        if not slot:
            return jsonify({"error": "Slot Not Found"}), 404
        slot = {
            'subject': slot['subject'],
            'faculty_name': slot['teacher'],
            'slot_type': slot.get('slot_type', 'Theory'),
            'day': slot['day'],
            'time_slot': slot['time'],
            'room': slot.get('room', 'N/A'),
            'division': slot['division']
        }
        
    title = f"New Class Scheduled: {slot['subject']}"
    message = f"Faculty {slot['faculty_name']} has been scheduled for {slot['subject']} ({slot['slot_type']}) on {slot['day']} at {slot['time_slot']} in Room {slot.get('room') or 'N/A'} for division {slot['division']}."
    
    from routes.features import _ensure_notifications_table, log_audit
    _ensure_notifications_table()
    
    exe("INSERT INTO notifications (title, message, role_target) VALUES (%s, %s, %s)",
        (title, message, "all"))
        
    try:
        log_audit("Publish Notification", f"Title: {title} (Timetable Notification)")
    except Exception:
        pass
        
    return jsonify({"message": "Notification broadcast successfully"}), 200



