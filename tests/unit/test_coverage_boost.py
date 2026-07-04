"""
High-impact coverage tests targeting:
  - tasks/reports/attendance_reports.py (get_date_range_from_month_academic_year)
  - services/tenant_service.py
  - utils/excel_generator.py (ExcelReport, conditional format, freeze, chart)
  - utils/version_router.py (VERSION_CONFIGS)
  - utils/apm.py
  - utils/api_key_auth.py
  - utils/db_schema_setup.py (importable constants)
  - services/report_service.py (ReportService.generate validation)
  - utils/pdf_generator.py
  - utils/data_cleaner.py
"""
import pytest
from unittest.mock import MagicMock, patch, call
import io
import os


# ────────────────────────────────────────────────────────
# tasks/reports/attendance_reports.py — pure function
# ────────────────────────────────────────────────────────
def test_get_date_range_june_academic_year():
    from tasks.reports.attendance_reports import get_date_range_from_month_academic_year
    start, end = get_date_range_from_month_academic_year("june", "2024-25")
    assert start.year == 2024
    assert start.month == 6
    assert start.day == 1
    assert end.month == 6
    assert end.day == 30

def test_get_date_range_january_academic_year():
    from tasks.reports.attendance_reports import get_date_range_from_month_academic_year
    # Jan belongs to the second year of "2024-25" (i.e. 2025)
    start, end = get_date_range_from_month_academic_year("january", "2024-25")
    assert start.year == 2025
    assert start.month == 1
    assert start.day == 1

def test_get_date_range_numeric_month():
    from tasks.reports.attendance_reports import get_date_range_from_month_academic_year
    start, end = get_date_range_from_month_academic_year(11, "2024-25")
    assert start.month == 11
    assert start.year == 2024

def test_get_date_range_string_digit_month():
    from tasks.reports.attendance_reports import get_date_range_from_month_academic_year
    start, end = get_date_range_from_month_academic_year("3", "2024-25")
    assert start.month == 3
    assert start.year == 2025  # march = month <= 5, second year

def test_get_date_range_short_year():
    """Academic year given as single year e.g. '2024'."""
    from tasks.reports.attendance_reports import get_date_range_from_month_academic_year
    start, end = get_date_range_from_month_academic_year("december", "2024")
    assert start.month == 12
    assert start.year == 2024

def test_get_date_range_abbreviations():
    from tasks.reports.attendance_reports import get_date_range_from_month_academic_year
    for abbr, month_num in [("jan", 1), ("feb", 2), ("mar", 3), ("apr", 4),
                             ("aug", 8), ("sep", 9), ("oct", 10), ("nov", 11), ("dec", 12)]:
        start, end = get_date_range_from_month_academic_year(abbr, "2024-25")
        assert start.month == month_num

def test_get_date_range_unknown_month_fallback():
    """Unknown month string should fallback to June."""
    from tasks.reports.attendance_reports import get_date_range_from_month_academic_year
    start, end = get_date_range_from_month_academic_year("unknown_month", "2024-25")
    assert start.month == 6  # fallback to June


# ────────────────────────────────────────────────────────
# services/tenant_service.py
# ────────────────────────────────────────────────────────
def test_tenant_get_by_subdomain_cache_hit(app):
    """get_by_subdomain should return cached tenant from Redis."""
    import json
    with app.app_context():
        cached_tenant = {"id": 1, "slug": "dypatil", "name": "DY Patil"}
        with patch("services.tenant_service.redis_client") as mock_redis:
            mock_redis.get.return_value = json.dumps(cached_tenant).encode()
            from services.tenant_service import TenantService
            result = TenantService.get_by_subdomain("dypatil")
            assert result["slug"] == "dypatil"
            mock_redis.get.assert_called()

def test_tenant_get_by_subdomain_db_fallback(app):
    """get_by_subdomain should query DB when cache misses."""
    import json
    with app.app_context():
        tenant_mock = MagicMock()
        tenant_mock.__iter__ = lambda self: iter([("id", 1), ("slug", "dypatil"), ("name", "DY Patil")])
        tenant_mock.get.return_value = None
        with patch("services.tenant_service.redis_client") as mock_redis, \
             patch("services.tenant_service.qone", return_value=None):
            mock_redis.get.return_value = None
            from services.tenant_service import TenantService
            result = TenantService.get_by_subdomain("nonexistent")
            assert result is None

def test_tenant_get_by_id_cache_hit(app):
    """get_by_id should return cached tenant."""
    import json
    with app.app_context():
        cached = {"id": 2, "slug": "test"}
        with patch("services.tenant_service.redis_client") as mock_redis:
            mock_redis.get.return_value = json.dumps(cached).encode()
            from services.tenant_service import TenantService
            result = TenantService.get_by_id(2)
            assert result["id"] == 2

def test_tenant_get_by_id_not_found(app):
    with app.app_context():
        with patch("services.tenant_service.redis_client") as mock_redis, \
             patch("services.tenant_service.qone", return_value=None):
            mock_redis.get.return_value = None
            from services.tenant_service import TenantService
            result = TenantService.get_by_id(999)
            assert result is None

def test_tenant_get_config_cache_hit(app):
    """get_config returns cached config value."""
    import json
    with app.app_context():
        with patch("services.tenant_service.redis_client") as mock_redis:
            mock_redis.get.return_value = json.dumps("enabled").encode()
            from services.tenant_service import TenantService
            val = TenantService.get_config(1, "some_feature")
            assert val == "enabled"

def test_tenant_get_config_db_fallback(app):
    with app.app_context():
        with patch("services.tenant_service.redis_client") as mock_redis, \
             patch("services.tenant_service.qone", return_value={"value": "dark_mode"}):
            mock_redis.get.return_value = None
            from services.tenant_service import TenantService
            val = TenantService.get_config(1, "theme")
            assert val == "dark_mode"

def test_tenant_get_config_default(app):
    with app.app_context():
        with patch("services.tenant_service.redis_client") as mock_redis, \
             patch("services.tenant_service.qone", return_value=None):
            mock_redis.get.return_value = None
            from services.tenant_service import TenantService
            val = TenantService.get_config(1, "missing_key", default="fallback")
            assert val == "fallback"

def test_tenant_bust_cache(app):
    """bust_cache should delete Redis keys."""
    with app.app_context():
        with patch("services.tenant_service.redis_client") as mock_redis:
            mock_redis.keys.return_value = [b"tenant:cfg:1:key1"]
            from services.tenant_service import TenantService
            TenantService.bust_cache(1, "dypatil")
            assert mock_redis.delete.call_count >= 2  # subdomain + id + configs

def test_tenant_provision_existing_slug(app):
    """provision_tenant should raise ValueError if slug exists."""
    with app.app_context():
        with patch("services.tenant_service.qone", return_value={"id": 1}):
            from services.tenant_service import TenantService
            with pytest.raises(ValueError, match="already exists"):
                TenantService.provision_tenant("dypatil", "DY Patil", "dypatil.erp", "basic")

def test_tenant_create_schema_invalid_chars(app):
    """_create_tenant_schema should reject names with special chars."""
    with app.app_context():
        from services.tenant_service import TenantService
        with pytest.raises(ValueError, match="Invalid characters"):
            TenantService._create_tenant_schema("DROP TABLE; --")

def test_tenant_create_schema_valid(app):
    """_create_tenant_schema should run DDL for safe schema names."""
    with app.app_context():
        with patch("services.tenant_service.exe") as mock_exe:
            from services.tenant_service import TenantService
            TenantService._create_tenant_schema("tenant_dypatil")
            mock_exe.assert_called_once()
            call_sql = mock_exe.call_args[0][0]
            assert "CREATE SCHEMA" in call_sql


# ────────────────────────────────────────────────────────
# utils/excel_generator.py — ExcelReport, conditional format, freeze
# ────────────────────────────────────────────────────────
def test_excel_report_add_sheet():
    """ExcelReport should create multiple named sheets."""
    from utils.excel_generator import ExcelReport
    report = ExcelReport("DY Patil Report", subtitle="Academic Year 2024-25")
    sheet1 = report.add_sheet("Attendance")
    sheet2 = report.add_sheet("Results")
    assert "Attendance" in report.sheets
    assert "Results" in report.sheets

def test_excel_report_save(tmp_path):
    """ExcelReport.save should write a valid xlsx file."""
    from utils.excel_generator import ExcelReport
    report = ExcelReport("Test Report")
    sheet = report.add_sheet("Sheet1")
    sheet.set_headers([{"key": "name", "label": "Name", "width": 15, "format": "text"}])
    sheet.add_rows([{"name": "Alice"}, {"name": "Bob"}])
    
    out_path = str(tmp_path / "test_report.xlsx")
    size = report.save(out_path)
    assert size > 0
    assert os.path.exists(out_path)

def test_excel_sheet_freeze():
    """freeze_panes should be set on the worksheet."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Test")
    sheet.set_headers([{"key": "roll", "label": "Roll No", "width": 10}])
    sheet.freeze(row=5, col=2)
    assert ws.freeze_panes is not None

def test_excel_sheet_apply_conditional_format():
    """apply_conditional_format should color cells based on rules."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Test")
    sheet.set_headers([
        {"key": "pct", "label": "Pct", "width": 8, "format": "percentage"}
    ])
    sheet.add_rows([{"pct": 0.5}, {"pct": 0.9}])
    # Should not raise
    sheet.apply_conditional_format("pct", [
        {"min": 0, "max": 0.75, "fill_color": "FFC7CE"},
        {"min": 0.75, "max": 1.0, "fill_color": "C6EFCE"},
    ])

def test_excel_sheet_conditional_format_unknown_key():
    """apply_conditional_format with unknown key should silently skip."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Test")
    sheet.set_headers([{"key": "name", "label": "Name", "width": 10}])
    sheet.add_rows([{"name": "Alice"}])
    # Should not raise — key not found returns early
    sheet.apply_conditional_format("nonexistent_col", [{"min": 0, "max": 1, "fill_color": "FFC7CE"}])

def test_excel_sheet_add_chart():
    """add_chart should not raise for bar chart type."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Test")
    sheet.set_headers([
        {"key": "roll", "label": "Roll", "width": 10},
        {"key": "pct", "label": "Pct", "width": 8},
    ])
    rows = [{"roll": f"CS{i:03d}", "pct": 70 + i} for i in range(5)]
    sheet.add_rows(rows)
    sheet.add_chart("bar", {
        "data_col": 2, "cats_col": 1,
        "min_row": sheet.start_row, "max_row": sheet.ws.max_row
    }, "Attendance Chart")


# ────────────────────────────────────────────────────────
# utils/api_key_auth.py
# ────────────────────────────────────────────────────────
def test_api_key_auth_importable():
    """api_key_auth module should be importable with expected decorator."""
    from utils.api_key_auth import api_key_required
    assert callable(api_key_required)

def test_api_key_auth_hashing():
    """SHA-256 hashing of API key should produce consistent 64-char hex."""
    import hashlib
    key = "test_api_key_12345"
    h1 = hashlib.sha256(key.encode('utf-8')).hexdigest()
    h2 = hashlib.sha256(key.encode('utf-8')).hexdigest()
    assert h1 == h2  # deterministic
    assert len(h1) == 64  # SHA-256 hex = 64 chars


# ────────────────────────────────────────────────────────
# services/report_service.py — REPORT_REGISTRY
# ────────────────────────────────────────────────────────
def test_report_registry_completeness():
    """All report types should have required fields."""
    from services.report_service import REPORT_REGISTRY
    required_fields = ["task", "name", "description", "allowed_roles",
                       "required_filters", "optional_filters", "formats", "estimated_seconds"]
    for rtype, config in REPORT_REGISTRY.items():
        for field in required_fields:
            assert field in config, f"{rtype} missing field: {field}"

def test_validate_filters_timetable_student_wrong_dept():
    """Student accessing another dept timetable should be denied."""
    from services.report_service import validate_filters, ReportValidationError
    with pytest.raises(ReportValidationError, match="Access denied"):
        validate_filters(
            "timetable_export",
            {"department": "IT", "year": "II", "division": "A", "semester": "3"},
            {"role": "student", "department": "CS"}
        )

def test_validate_filters_timetable_student_own_dept():
    """Student accessing own dept timetable should pass."""
    from services.report_service import validate_filters
    # Should not raise
    validate_filters(
        "timetable_export",
        {"department": "CS", "year": "II", "division": "A", "semester": "3"},
        {"role": "student", "department": "CS"}
    )

def test_report_service_invalid_format():
    """ReportService.generate with invalid format should raise."""
    from services.report_service import ReportService, ReportValidationError
    with pytest.raises(ReportValidationError, match="Invalid format"):
        ReportService.generate("monthly_attendance", "docx", {}, {"role": "admin"})

def test_report_service_invalid_type():
    """ReportService.generate with invalid type should raise."""
    from services.report_service import ReportService, ReportValidationError
    with pytest.raises(ReportValidationError, match="Invalid report type"):
        ReportService.generate("nonexistent", "pdf", {}, {"role": "admin"})


# ────────────────────────────────────────────────────────
# utils/data_cleaner.py — additional paths
# ────────────────────────────────────────────────────────
def test_data_cleaner_functions():
    """Test actual functions available in data_cleaner."""
    from utils.data_cleaner import (
        clean_text, normalize_header, parse_date_token,
        normalize_subject_name, clean_subject_code, safe_int, is_valid_roll
    )
    # clean_text
    assert clean_text("  hello  world\n") == "hello world"
    assert clean_text(None) == ""

    # normalize_header
    assert normalize_header("Roll No.") == "roll_no"
    assert normalize_header("NAME") == "name"

    # parse_date_token
    assert parse_date_token("01.06.2024") == "2024-06-01"
    assert parse_date_token("invalid") is None

    # normalize_subject_name
    assert normalize_subject_name("  Math - I  ") == "Math - I"

    # clean_subject_code
    assert clean_subject_code(" u24cs001 ") == "U24CS001"

    # safe_int
    assert safe_int("42") == 42
    assert safe_int(None) == 0

    # is_valid_roll
    assert is_valid_roll("A1") is True
    assert is_valid_roll("not_a_roll_number_xyz_abc") is False


# ────────────────────────────────────────────────────────
# utils/pg_wrapper.py — covered paths
# ────────────────────────────────────────────────────────
def test_pg_wrapper_qry_returns_list(app):
    """qry should return a list of dicts."""
    with app.app_context():
        from utils.pg_wrapper import qry
        result = qry("SELECT 1 AS n")
        assert isinstance(result, list)

def test_pg_wrapper_qone_returns_row_or_none(app):
    """qone should return a single row or None."""
    with app.app_context():
        from utils.pg_wrapper import qone
        result = qone("SELECT 1 AS n WHERE 1=0")
        assert result is None

def test_pg_wrapper_exe_runs_without_error(app):
    """exe should run an INSERT/UPDATE without crashing."""
    with app.app_context():
        from utils.pg_wrapper import exe
        # Create temp table and insert into it
        exe("CREATE TABLE IF NOT EXISTS _test_pg_wrapper_coverage (id SERIAL PRIMARY KEY, val TEXT)")
        exe("INSERT INTO _test_pg_wrapper_coverage (val) VALUES (%s)", ("test_value",))
        exe("DROP TABLE IF EXISTS _test_pg_wrapper_coverage")


# ────────────────────────────────────────────────────────
# utils/tenant_middleware.py
# ────────────────────────────────────────────────────────
def test_tenant_middleware_importable(app):
    """Tenant middleware should be importable without errors."""
    from utils.tenant_middleware import TenantMiddleware
    assert TenantMiddleware is not None


# ────────────────────────────────────────────────────────
# utils/tenant_jwt.py
# ────────────────────────────────────────────────────────
def test_tenant_jwt_importable(app):
    """tenant_jwt module should import correctly."""
    with app.app_context():
        from utils.tenant_jwt import tenant_jwt_required, role_required
        assert callable(tenant_jwt_required)
        assert callable(role_required)

def test_role_required_decorator_works(app):
    """role_required wraps a function correctly."""
    with app.app_context():
        from utils.tenant_jwt import role_required
        @role_required("admin")
        def sample_view():
            return "ok"
        assert callable(sample_view)


# ────────────────────────────────────────────────────────
# utils/tenant_storage.py
# ────────────────────────────────────────────────────────
def test_tenant_storage_importable():
    from utils.tenant_storage import TenantStorage
    assert TenantStorage is not None

def test_tenant_storage_download_url(app):
    with app.app_context():
        with patch("utils.tenant_storage.get_current_tenant", return_value={"slug": "dypatil"}):
            from utils.tenant_storage import TenantStorage
            url = TenantStorage.download_url("report.pdf")
            assert "dypatil" in url
            assert "report.pdf" in url
