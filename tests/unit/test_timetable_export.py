import pytest
import openpyxl
from io import BytesIO
from utils.pg_wrapper import qone, exe, qry

@pytest.fixture
def logged_in_faculty(client):
    with client.session_transaction() as sess:
        sess['faculty_id'] = 999
        sess['name'] = 'Prof. Test Faculty'
        sess['role'] = 'faculty'
    return 999

@pytest.fixture
def logged_in_admin(client):
    with client.session_transaction() as sess:
        sess['role'] = 'admin'
        sess['name'] = 'Admin User'
    return 'admin'

def test_faculty_excel_export_returns_xlsx(client, logged_in_faculty):
    # Ensure there is at least one approved slot
    exe("DELETE FROM faculty_timetable")
    exe("""
        INSERT INTO faculty_timetable (faculty_id, faculty_name, day, time_slot, subject, division, slot_type, status)
        VALUES (999, 'Prof. Test Faculty', 'Monday', '8:30-9:30', 'Maths', 'SE CSE-A', 'Theory', 'approved')
    """)
    
    resp = client.get('/api/faculty/my-timetable/export/excel')
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.content_type
    assert '.xlsx' in resp.headers.get('Content-Disposition', '')

def test_faculty_pdf_export_returns_pdf(client, logged_in_faculty):
    resp = client.get('/api/faculty/my-timetable/export/pdf')
    assert resp.status_code == 200
    assert resp.content_type == 'application/pdf'
    assert '.pdf' in resp.headers.get('Content-Disposition', '')

def test_faculty_export_only_approved_slots(client, logged_in_faculty):
    exe("DELETE FROM faculty_timetable")
    # One approved slot
    exe("""
        INSERT INTO faculty_timetable (faculty_id, faculty_name, day, time_slot, subject, division, slot_type, status)
        VALUES (999, 'Prof. Test Faculty', 'Monday', '8:30-9:30', 'Approved Subject', 'SE CSE-A', 'Theory', 'approved')
    """)
    # One pending slot
    exe("""
        INSERT INTO faculty_timetable (faculty_id, faculty_name, day, time_slot, subject, division, slot_type, status)
        VALUES (999, 'Prof. Test Faculty', 'Tuesday', '8:30-9:30', 'Pending Subject', 'SE CSE-A', 'Theory', 'pending')
    """)

    resp = client.get('/api/faculty/my-timetable/export/excel')
    assert resp.status_code == 200
    
    wb = openpyxl.load_workbook(BytesIO(resp.data))
    ws = wb.active
    
    # Header is row 1. Data rows start at row 2.
    # Read row 2 subject
    row2_subject = ws.cell(row=2, column=3).value
    assert row2_subject == 'Approved Subject'
    
    # Row 3 should not be a slot data row (it should be empty or summary)
    row3_val = ws.cell(row=3, column=3).value
    assert row3_val is None

def test_admin_excel_export_master(client, logged_in_admin):
    resp = client.get('/api/admin/timetable/export/excel?type=master')
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.content_type

def test_admin_excel_export_pending(client, logged_in_admin):
    resp = client.get('/api/admin/timetable/export/excel?type=pending')
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.content_type

def test_admin_export_invalid_type_returns_400(client, logged_in_admin):
    resp = client.get('/api/admin/timetable/export/excel?type=garbage')
    assert resp.status_code == 400

def test_export_requires_auth(client):
    resp = client.get('/api/faculty/my-timetable/export/excel')
    assert resp.status_code in [302, 401]
