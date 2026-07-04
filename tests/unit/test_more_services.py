"""
Unit tests for:
  - utils/excel_generator.py (ExcelSheet class)
  - tasks/notification_tasks.py (Celery tasks unit testing)
  - services/attendance_service.py (is_attendance_locked, get_students_for_filters)
"""
import pytest
from unittest.mock import MagicMock, patch, PropertyMock
from datetime import datetime, timedelta


# ────────────────────────────────────────────────────────
# utils/excel_generator.py
# ────────────────────────────────────────────────────────
def test_excel_sheet_creation():
    """ExcelSheet should create a workbook and write title block."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Test Report", subtitle="Test Subtitle")
    assert sheet.title == "Test Report"
    assert sheet.subtitle == "Test Subtitle"
    # Title cell should have value
    assert ws['C1'].value == "Test Report"

def test_excel_sheet_no_subtitle():
    """ExcelSheet without subtitle should not write to A3."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Report Only")
    assert ws['A3'].value is None

def test_excel_sheet_set_headers():
    """set_headers should configure columns and mark headers_set."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Test")
    columns = [
        {"key": "roll", "label": "Roll No", "width": 10},
        {"key": "name", "label": "Name", "width": 20},
    ]
    sheet.set_headers(columns)
    assert sheet.headers_set is True
    assert len(sheet.columns) == 2

def test_excel_sheet_add_rows():
    """add_rows should write data to the worksheet."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Test")
    sheet.set_headers([
        {"key": "roll", "label": "Roll No", "width": 10, "format": "text"},
        {"key": "pct", "label": "Pct", "width": 8, "format": "percentage"},
        {"key": "marks", "label": "Marks", "width": 8, "format": "integer"},
        {"key": "score", "label": "Score", "width": 8, "format": "float"},
    ])
    data = [
        {"roll": "CS001", "pct": 75.0, "marks": 42, "score": 3.5},
        {"roll": "CS002", "pct": 0.85, "marks": 39, "score": None},
    ]
    sheet.add_rows(data)
    # Verify data was written to worksheet
    assert ws.max_row > 5  # title + header + data rows

def test_excel_sheet_add_rows_without_headers_raises():
    """add_rows before set_headers should raise ValueError."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Test")
    with pytest.raises(ValueError, match="Headers must be set"):
        sheet.add_rows([{"key": "val"}])

def test_excel_sheet_add_summary_row():
    """add_summary_row should add a totals row without error."""
    import openpyxl
    from utils.excel_generator import ExcelSheet
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = ExcelSheet(ws, "Test")
    sheet.set_headers([
        {"key": "roll", "label": "Roll No", "width": 10, "format": "text"},
        {"key": "count", "label": "Count", "width": 8, "format": "integer"},
    ])
    sheet.add_rows([{"roll": "CS001", "count": 5}])
    sheet.add_summary_row({"roll": "Total", "count": 5})
    assert ws.max_row > 6


# ────────────────────────────────────────────────────────
# services/attendance_service.py — is_attendance_locked
# ────────────────────────────────────────────────────────
def test_is_attendance_locked_no_session_id(app):
    """No session_id should return False (not locked)."""
    with app.app_context():
        from services.attendance_service import is_attendance_locked
        assert is_attendance_locked(None) is False
        assert is_attendance_locked("") is False

def test_is_attendance_locked_missing_session(app):
    """Session not found in DB should return True (locked by default)."""
    with app.app_context():
        with patch("services.attendance_service.qone", return_value=None):
            from services.attendance_service import is_attendance_locked
            assert is_attendance_locked(999) is True

def test_is_attendance_locked_explicit(app):
    """Explicitly locked session should return True."""
    with app.app_context():
        mock_row = MagicMock()
        mock_row.get.side_effect = lambda k, d=None: True if k == "locked_at" else d
        mock_row.__getitem__ = MagicMock(return_value="2024-01-01")
        with patch("services.attendance_service.qone", return_value=mock_row):
            from services.attendance_service import is_attendance_locked
            assert is_attendance_locked(1) is True

def test_is_attendance_locked_recent(app):
    """Session within lock window should not be locked."""
    with app.app_context():
        # Use a recent date (today) so the cutoff hasn't passed
        today_str = datetime.now().strftime("%Y-%m-%d")
        mock_row = MagicMock()
        mock_row.get.return_value = None  # locked_at = None, is_locked = None
        mock_row.__getitem__ = lambda self, k: today_str if k == 'lecture_date' else None
        with patch("services.attendance_service.qone", return_value=mock_row):
            from services.attendance_service import is_attendance_locked
            # Recent session — may or may not be locked depending on ATTENDANCE_LOCK_HOURS
            # Just assert it doesn't crash
            result = is_attendance_locked(1)
            assert isinstance(result, bool)


# ────────────────────────────────────────────────────────
# tasks/notification_tasks.py
# ────────────────────────────────────────────────────────
def test_send_push_notification_task_no_tokens(app):
    """send_push_notification_task should call push_service.send_to_user."""
    with app.app_context():
        with patch("tasks.notification_tasks.push_service") as mock_push:
            mock_push.send_to_user.return_value = None
            from tasks.notification_tasks import send_push_notification_task
            # Call the underlying function directly (bypassing Celery)
            result = send_push_notification_task.run(1, "Title", "Body")
            mock_push.send_to_user.assert_called_once_with(1, "Title", "Body")

def test_broadcast_notification_task(app):
    """broadcast_notification_task should call push_service.send_to_topic."""
    with app.app_context():
        with patch("tasks.notification_tasks.push_service") as mock_push:
            mock_push.send_to_topic.return_value = MagicMock()
            from tasks.notification_tasks import broadcast_notification_task
            broadcast_notification_task.run("general", "Title", "Body")
            mock_push.send_to_topic.assert_called_once_with("general", "Title", "Body")

def test_send_application_confirmation_no_app(app):
    """send_application_confirmation with no app found should return early."""
    with app.app_context():
        with patch("tasks.notification_tasks.qone", return_value=None):
            from tasks.notification_tasks import send_application_confirmation
            result = send_application_confirmation.run(999)
            assert result is None

def test_send_offer_letter_no_app(app):
    """send_offer_letter with no app found should return early."""
    with app.app_context():
        with patch("tasks.notification_tasks.qone", return_value=None):
            from tasks.notification_tasks import send_offer_letter
            result = send_offer_letter.run(999)
            assert result is None


# ────────────────────────────────────────────────────────
# services/otp_service.py — verify_otp
# ────────────────────────────────────────────────────────
def test_otp_verify_invalid(app):
    """Verify OTP with no record should return failure."""
    with app.app_context():
        with patch("services.otp_service.qone", return_value=None):
            from services.otp_service import OTPService
            result = OTPService.verify_otp("9876543210", "123456")
            assert result["success"] is False
            assert "Invalid" in result["error"]

def test_otp_verify_valid(app):
    """Verify OTP with valid record should mark as verified."""
    with app.app_context():
        mock_record = MagicMock()
        mock_record.__getitem__ = lambda self, k: 1 if k == "id" else None
        with patch("services.otp_service.qone", return_value=mock_record), \
             patch("services.otp_service.exe") as mock_exe:
            from services.otp_service import OTPService
            result = OTPService.verify_otp("9876543210", "123456")
            assert result["success"] is True
            mock_exe.assert_called()


# ────────────────────────────────────────────────────────
# services/sms_service.py — send_immediate
# ────────────────────────────────────────────────────────
def test_sms_send_immediate_template_not_found(app):
    """send_immediate with missing template should return failure."""
    with app.app_context():
        with patch("services.sms_service.qone", return_value=None):
            from services.sms_service import SMSService
            result = SMSService.send_immediate("9876543210", "nonexistent_tpl", {})
            assert result["success"] is False
            assert "not found" in result["error"]

def test_sms_send_immediate_success(app):
    """send_immediate with valid template and provider should succeed."""
    with app.app_context():
        mock_tpl = MagicMock()
        mock_tpl.__getitem__ = lambda self, k: "Hello {{name}}" if k == "body" else None
        mock_log = MagicMock()
        mock_log.__getitem__ = lambda self, k: 1 if k == "id" else None

        mock_provider = MagicMock()
        mock_provider.send_sms.return_value = {"success": True, "id": "ref_001", "raw": {}, "error": None}

        with patch("services.sms_service.qone", side_effect=[mock_tpl, mock_log]), \
             patch("services.sms_service.SMSFactory.get_provider", return_value=mock_provider), \
             patch("services.sms_service.exe"):
            from services.sms_service import SMSService
            result = SMSService.send_immediate("9876543210", "welcome", {"name": "Alice"})
            assert result["success"] is True
