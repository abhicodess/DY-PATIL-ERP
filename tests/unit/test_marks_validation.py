import pytest
import io
from openpyxl import Workbook, load_workbook
from services.results_service import calculate_result
from models.results import Mark
from models.student import Student
from utils.pg_wrapper import qry, qone, exe

# 1. test_total_auto_calculated
def test_total_auto_calculated(client, session, faculty, student):
    # Set faculty session
    with client.session_transaction() as sess:
        sess['role'] = 'faculty'
        sess['faculty_id'] = faculty.id
        sess['name'] = faculty.name
        
    # POST to faculty_save_marks
    resp = client.post("/faculty_save_marks", json={
        "student_name": student.name,
        "subject": "Statistics & Probability",
        "assignment_marks": 4.0,
        "attendance_marks": 4.0,
        "teaching_assessment": 8.0,
        "ut_marks": 16.0,
        "mse_marks": 16.0,
        "remarks": "Good",
        "semester": "SEM IV"
    })
    assert resp.status_code == 200
    
    # Query database to assert total = sum
    mark = qone("SELECT * FROM marks WHERE student_id = %s", (student.id,))
    assert mark is not None
    assert mark["marks"] == 48.0 # 4+4+8+16+16
    assert mark["total"] == 60.0

# 2. test_ut_max_enforced
def test_ut_max_enforced(client, session, faculty, student):
    with client.session_transaction() as sess:
        sess['role'] = 'faculty'
        sess['faculty_id'] = faculty.id
        sess['name'] = faculty.name
        
    resp = client.post("/faculty_save_marks", json={
        "student_name": student.name,
        "subject": "Statistics & Probability",
        "ut_marks": 25.0 # Max is 20
    })
    assert resp.status_code == 400
    assert "ut_marks" in resp.json["errors"]

# 3. test_assignment_max_enforced
def test_assignment_max_enforced(client, session, faculty, student):
    with client.session_transaction() as sess:
        sess['role'] = 'faculty'
        sess['faculty_id'] = faculty.id
        sess['name'] = faculty.name
        
    resp = client.post("/faculty_save_marks", json={
        "student_name": student.name,
        "subject": "Statistics & Probability",
        "assignment_marks": 7.0 # Max is 5
    })
    assert resp.status_code == 400
    assert "assignment_marks" in resp.json["errors"]

# 4. test_grade_calculation_pass
def test_grade_calculation_pass():
    # total=45/60 = 75% -> grade O
    total, grade, result, passed = calculate_result(4.0, 4.0, 7.0, 15.0, 15.0)
    assert total == 45.0
    assert grade == 'O'
    assert result == 'Pass'
    assert passed is True

# 5. test_grade_calculation_fail_ut
def test_grade_calculation_fail_ut():
    # ut=5 (below 8) even if total=35 -> grade F, result Fail
    total, grade, result, passed = calculate_result(4.0, 4.0, 7.0, 5.0, 15.0)
    assert not passed
    assert grade == 'F'
    assert result == 'Fail'

# 6. test_grade_calculation_fail_mse
def test_grade_calculation_fail_mse():
    # mse=5 (below 8) even if total=35 -> grade F, result Fail
    total, grade, result, passed = calculate_result(4.0, 4.0, 7.0, 15.0, 5.0)
    assert not passed
    assert grade == 'F'
    assert result == 'Fail'

# 7. test_ut_publish_makes_visible_to_student
def test_ut_publish_makes_visible_to_student(client, session, faculty, student):
    # Add marks row
    # Seeding subject U24AIMLPC401
    exe("INSERT INTO subjects_master (subject_code, subject_name, department, semester) VALUES (%s, %s, %s, %s) ON CONFLICT DO NOTHING",
        ("U24AIMLPC401", "Statistics & Probability", "AIML", "SEM IV"))
        
    with client.session_transaction() as sess:
        sess['role'] = 'faculty'
        sess['faculty_id'] = faculty.id
        sess['name'] = faculty.name
        
    # Save marks
    resp = client.post("/faculty_save_marks", json={
        "student_name": student.name,
        "subject": "Statistics & Probability",
        "ut_marks": 15.0,
        "semester": "SEM IV"
    })
    assert resp.status_code == 200
    
    # Student login
    with client.session_transaction() as sess:
        sess['role'] = 'student'
        sess['student_id'] = student.id
        sess['name'] = student.name
        
    # Initially UT is not published
    resp = client.get("/student/marks/ut")
    assert resp.status_code == 200
    assert len(resp.json) == 0
    
    # Faculty publish UT
    with client.session_transaction() as sess:
        sess['role'] = 'faculty'
        sess['faculty_id'] = faculty.id
        sess['name'] = faculty.name
        
    resp_pub = client.post("/faculty/marks/publish-ut", json={
        "subject_code": "U24AIMLPC401",
        "division": student.division,
        "semester": "SEM IV"
    })
    assert resp_pub.status_code == 200
    
    # Student logs back in and checks
    with client.session_transaction() as sess:
        sess['role'] = 'student'
        sess['student_id'] = student.id
        sess['name'] = student.name
        
    resp_stud = client.get("/student/marks/ut")
    assert resp_stud.status_code == 200
    assert len(resp_stud.json) == 1
    assert resp_stud.json[0]["ut_marks"] == 15.0

# 8. test_result_not_visible_before_publish
def test_result_not_visible_before_publish(client, session, faculty, student):
    exe("INSERT INTO subjects_master (subject_code, subject_name, department, semester) VALUES (%s, %s, %s, %s) ON CONFLICT DO NOTHING",
        ("U24AIMLPC401", "Statistics & Probability", "AIML", "SEM IV"))
        
    with client.session_transaction() as sess:
        sess['role'] = 'faculty'
        sess['faculty_id'] = faculty.id
        sess['name'] = faculty.name
        
    client.post("/faculty_save_marks", json={
        "student_name": student.name,
        "subject": "Statistics & Probability",
        "ut_marks": 15.0,
        "semester": "SEM IV"
    })
    
    with client.session_transaction() as sess:
        sess['role'] = 'student'
        sess['student_id'] = student.id
        
    resp = client.get("/student/marks/full")
    assert resp.status_code == 200
    assert len(resp.json) == 0

# 9. test_excel_import_reads_prn_correctly
def test_excel_import_reads_prn_correctly(client, session, student):
    # Set PRN for student
    student.prn = "PRN12345"
    session.commit()
    
    # Create excel in memory
    wb = Workbook()
    ws = wb.active
    ws.append([]) # Row 1 blank
    ws.append([None, None, None, "U24AIMLPC401"]) # Row 2 subject code
    ws.append(["Sr.No", "Student Name", "PRN Number", "Assignment", "Attendance", "Teaching", "UT", "MSE", "TOTAL"]) # Row 3
    ws.append([1, student.name, "PRN12345", 4.0, 4.0, 8.0, 16.0, 16.0, 48.0]) # Row 4
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    with client.session_transaction() as sess:
        sess['role'] = 'admin'
        sess['name'] = 'Admin User'
        
    # POST to import endpoint
    resp = client.post("/admin/marks/import-excel", data={
        "file": (output, "import_test.xlsx")
    })
    assert resp.status_code == 200
    assert resp.json["imported"] == 1
    
    # Assert DB marks updated
    mark = qone("SELECT * FROM marks WHERE student_id = %s", (student.id,))
    assert mark is not None
    assert mark["prn_number"] == "PRN12345"
    assert mark["marks"] == 48.0

# 10. test_excel_export_has_correct_structure
def test_excel_export_has_correct_structure(client, session, student):
    student.prn = "PRN55555"
    student.department = "AIML"
    student.division = "A"
    session.commit()
    
    # Seed subject
    exe("INSERT INTO subjects_master (subject_code, subject_name, department, semester) VALUES (%s, %s, %s, %s) ON CONFLICT DO NOTHING",
        ("U24AIMLPC401", "Statistics & Probability", "AIML", "SEM IV"))
        
    with client.session_transaction() as sess:
        sess['role'] = 'admin'
        sess['name'] = 'Admin User'
        
    # GET export
    resp = client.get("/admin/marks/export-excel?division=A&semester=SEM IV")
    assert resp.status_code == 200
    
    # Read excel response
    wb = load_workbook(io.BytesIO(resp.data))
    assert "AIML" in wb.sheetnames
    
    ws = wb["AIML"]
    assert ws.cell(row=3, column=4).value == "Assignment(05)"
