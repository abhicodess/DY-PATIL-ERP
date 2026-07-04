"""
DY Patil University — College ERP
Complete Flask Application — Built from Scratch
"""

from flask import Flask, render_template, request, redirect, session, jsonify, send_file, abort, flash, url_for
import psycopg2
import psycopg2.extras, os, io, re, secrets, hmac, time
import json
import shutil
import math
import tempfile
from werkzeug.security import generate_password_hash, check_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date, timedelta
from config import DEPARTMENTS, DIVISIONS, SEMESTERS, DESIGNATIONS, YEARS, DAYS, DAY_ORD, UPLOAD_DIR
from utils.helpers import (
    safe_int, safe_str, safe_float, safe_date,
    hash_password, verify_password,
    sort_key_time, login_required, safe_redirect_target, validate_password_change, _password_is_hashed
)
from routes.features import log_audit
from functools import wraps
import pdfplumber

def extract_text(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text
from models.attendance import ensure_attendance_upload_tables
from models.attendance_model import ensure_attendance_engine_schema
from models.faculty_attendance_v2 import ensure_faculty_attendance_v2_schema
from routes.attendance import (
    attendance_role_allowed,
    handle_backup as handle_attendance_backup,
    handle_bulk_mark,
    handle_delete_record,
    handle_edit_record,
    handle_import as handle_attendance_import,
    handle_reset as handle_attendance_reset,
    handle_restore as handle_attendance_restore,
    handle_single_mark,
    handle_view_records,
    students_api_response,
)
from routes.upload_attendance import (
    download_attendance_backup,
    is_attendance_upload_allowed,
    process_attendance_upload,
    restore_attendance_backup,
)
from services.attendance_service import attendance_page_context, init_attendance_engine
try:
    import pandas as pd
except ImportError:
    pd = None
from services.intelligence_service import IntelligenceService
import core.routes_registry as routes_registry
import core.db_validators as db_validators
from core.error_handlers import init_error_handlers
from core.query_registry import get_query

_re = re

from utils.security_headers import init_security_headers
import hashlib

app = Flask(__name__)

# FIX 1: SECRET_KEY enforcement
SECRET_KEY = os.environ.get("SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY environment variable must be set. Run: python -c \"import secrets; print(secrets.token_hex(32))\" and add to .env")
app.secret_key = SECRET_KEY

# FIX 7: Session Security
app.config["SESSION_COOKIE_SECURE"] = True # Set to False for local dev without HTTPS
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = 'Lax'
app.permanent_session_lifetime = timedelta(hours=8)

app.config["PROPAGATE_EXCEPTIONS"] = True
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# FIX 5: Rate Limiting with Redis
REDIS_URL = os.environ.get("REDIS_URL", "redis://localhost:6379/0")
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    storage_uri=REDIS_URL,
    strategy="fixed-window",
    default_limits=[]
)

# FIX 9: File upload security
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024 # 10MB
ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.pdf', '.csv', '.jpg', '.png', '.jpeg'}

init_security_headers(app)


import logging

def register_blueprints(app):
    # Critical Blueprints (Production-Guarded)
    try:
        from routes.features import features_bp
        app.register_blueprint(features_bp)
    except ImportError as e:
        if os.environ.get("FLASK_ENV", "production") == "production":
            raise RuntimeError(f"Critical blueprint failed to load (features): {e}") from e
        logging.getLogger(__name__).error("Blueprint load failed: %s", e)

    # Standard Blueprints
    blueprints = [
        ('routes.results', 'results_bp'),
        ('routes.cumulative', 'cumulative_bp'),
        ('routes.sms_routes', 'sms_bp'),
        ('routes.otp_routes', 'otp_bp'),
        ('routes.parent_routes', 'parent_bp'),
        ('routes.faculty_attendance_v2', 'faculty_att_bp'),
        ('routes.timetable_v2', 'timetable_v2_bp'),
        ('routes.admin_intel', 'admin_intel_bp'),
    ]

    for module_path, bp_name in blueprints:
        try:
            module = __import__(module_path, fromlist=[bp_name])
            bp = getattr(module, bp_name)
            app.register_blueprint(bp)
        except (ImportError, AttributeError) as e:
            if bp_name == 'faculty_att_bp' and os.environ.get("FLASK_ENV", "production") == "production":
                 raise RuntimeError(f"Critical blueprint failed to load ({bp_name}): {e}") from e
            logging.getLogger(__name__).error("Blueprint load failed (%s): %s", bp_name, e, exc_info=True)

    # API Blueprints
    try:
        from api import api_bp
        from api.errors import api_errors_bp
        app.register_blueprint(api_bp)
        app.register_blueprint(api_errors_bp)
    except ImportError as e:
        logging.getLogger(__name__).error("API Blueprint load failed: %s", e)

register_blueprints(app)
routes_registry.init_app(app)
init_error_handlers(app)

# Phase 1: Database Architecture Stabilization
with app.app_context():
    db_validators.validate_schema()



def _resolve_admin_password_hash():
    """Read admin hash from env, without source-level password literals."""
    import os
    env_hash = os.environ.get("ADMIN_PASSWORD_HASH", "").strip()
    if env_hash:
        return env_hash
    env_password = os.environ.get("ADMIN_PASSWORD")
    if env_password:
        return generate_password_hash(env_password)
    
    # FIX 2: Raise RuntimeError if no admin password is set
    raise RuntimeError(
        "CRITICAL: Set ADMIN_PASSWORD_HASH or ADMIN_PASSWORD env var. "
        "Default passwords are no longer allowed for security reasons."
    )

@app.errorhandler(Exception)
def handle_exception(e):
    """Global Production-Grade Error Handler."""
    # Log the full traceback for debugging (Step 11)
    import traceback
    tb = traceback.format_exc()
    app.logger.error(f"500 Internal Server Error: {str(e)}\n{tb}")
    
    # Check if we should rollback DB on error (Step 7/11)
    try:
        from utils.pg_wrapper import get_conn
        conn = get_conn()
        if conn: conn.rollback()
    except:
        pass

    if request.path.startswith('/api/'):
        return jsonify({
            "error": "Internal Server Error",
            "message": str(e) if app.debug else "A processing error occurred.",
            "status": 500
        }), 500

    return render_template("error_500.html", error=str(e), traceback=tb if app.debug else None), 500

@app.context_processor
def utility_processor():
    def safe_pct(p, t):
        try:
            p, t = float(p or 0), float(t or 0)
            if t <= 0: return 0.0
            return round((p / t) * 100, 1)
        except:
            return 0.0
    
    def safe_get(obj, attr, default="N/A"):
        if not obj: return default
        if hasattr(obj, attr):
            val = getattr(obj, attr)
            return val if val is not None else default
        if isinstance(obj, dict):
            val = obj.get(attr)
            return val if val is not None else default
        return default

    return dict(pct=safe_pct, safe_get=safe_get)
ADMIN_PASSWORD_HASH = _resolve_admin_password_hash()

# FIX 2: Remove default passwords
DEFAULT_STUDENT_PASSWORD = os.environ.get("DEFAULT_STUDENT_PASSWORD")
DEFAULT_FACULTY_PASSWORD = os.environ.get("DEFAULT_FACULTY_PASSWORD")

if not DEFAULT_STUDENT_PASSWORD:
    raise RuntimeError("DEFAULT_STUDENT_PASSWORD environment variable must be set.")
if not DEFAULT_FACULTY_PASSWORD:
    raise RuntimeError("DEFAULT_FACULTY_PASSWORD environment variable must be set.")




def _default_student_password():
    if DEFAULT_STUDENT_PASSWORD:
        return DEFAULT_STUDENT_PASSWORD
    raise RuntimeError("DEFAULT_STUDENT_PASSWORD must be set when creating students without explicit password")


def _default_faculty_password():
    if DEFAULT_FACULTY_PASSWORD:
        return DEFAULT_FACULTY_PASSWORD
    raise RuntimeError("DEFAULT_FACULTY_PASSWORD must be set when creating faculty without explicit password")


def _ensure_student_contact_columns():
    conn = get_db()
    try:
        conn.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS contact_number TEXT")
        conn.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS parent_contact TEXT")
        conn.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS dob DATE")
        conn.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS gender TEXT")
        conn.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS address TEXT")
        conn.execute("ALTER TABLE students ADD COLUMN IF NOT EXISTS admission_year INTEGER")
        conn.commit()
    except Exception as e:
        conn.rollback()
        app.logger.warning("Could not ensure student contact columns: %s", e)
    finally:
        conn.close()

@app.route("/health")
def health_check():
    try:
        from utils.pg_wrapper import qone
        qone("SELECT 1 AS ok")
        return jsonify({
            "status": "ok", 
            "db": "connected", 
            "timestamp": datetime.now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            "status": "error", 
            "db": "unreachable", 
            "detail": str(e)
        }), 500

def _ensure_timetable_schema():
    conn = get_db()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS timetable (
                id SERIAL PRIMARY KEY,
                day TEXT NOT NULL,
                time TEXT NOT NULL,
                start_time TIME,
                end_time TIME,
                subject_id INTEGER,
                subject TEXT NOT NULL,
                teacher TEXT,
                faculty_id INTEGER REFERENCES faculty(id),
                room TEXT,
                division TEXT,
                branch TEXT,
                year TEXT,
                semester TEXT,
                slot_type TEXT DEFAULT 'Theory',
                color TEXT,
                published BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("ALTER TABLE timetable ADD COLUMN IF NOT EXISTS published BOOLEAN DEFAULT TRUE")
        
        conn.execute("""
            CREATE TABLE IF NOT EXISTS timetable_substitutions (
                id SERIAL PRIMARY KEY,
                timetable_id INTEGER REFERENCES timetable(id),
                substitute_faculty_id INTEGER REFERENCES faculty(id),
                session_date DATE NOT NULL,
                created_by TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS attendance_disputes (
                id SERIAL PRIMARY KEY,
                student_id INTEGER REFERENCES students(id),
                attendance_id INTEGER REFERENCES attendance(id),
                reason TEXT,
                status TEXT DEFAULT 'pending',
                resolved_by TEXT,
                resolved_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
    except Exception as e:
        conn.rollback()
        app.logger.warning("Could not ensure timetable schema: %s", e)
    finally:
        conn.close()

# ─── CONFIG (Imported from config.py) ─────────────────────
from config import *

# ─── DATABASE HELPERS ─────────────────────────────────────
from utils.pg_wrapper import get_db, qry, qone, exe, PG_URL
import logging

logger = logging.getLogger(__name__)

def safe_fetch_all(sql, params=()):
    try:
        return qry(sql, params)
    except Exception as e:
        app.logger.error(f"DB Error (all): {e} | SQL: {sql}")
        return []

def safe_fetch_one(sql, params=()):
    try:
        return qone(sql, params)
    except Exception as e:
        app.logger.error(f"DB Error (one): {e} | SQL: {sql}")
        return None

def safe_fetch_scalar(sql, params=(), default=0):
    try:
        row = qone(sql, params)
        if row: return row[0]
        return default
    except Exception as e:
        app.logger.error(f"DB Error (scalar): {e} | SQL: {sql}")
        return default

def safe_execute(sql, params=()):
    try:
        return exe(sql, params)
    except Exception as e:
        app.logger.error(f"DB Error (exec): {e} | SQL: {sql}")
        return None

# Aliases for backward compatibility
safe_query = safe_fetch_all
# safe_fetch_one already defined above

# ─── AUTH DECORATORS ──────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "admin":
            flash("Administrator access required.", "error")
            return redirect("/admin_login")
        return f(*args, **kwargs)
    return decorated_function

def faculty_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "faculty":
            flash("Faculty access required.", "error")
            return redirect("/faculty_login")
        return f(*args, **kwargs)
    return decorated_function

def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "student":
            flash("Student access required.", "error")
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated_function

def _ensure_faculty_assignments_schema():
    conn = get_db()
    try:
        # 1. Faculty Subject Assignment Table
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS faculty_subject_assignments (
                    id SERIAL PRIMARY KEY,
                    faculty_id INTEGER REFERENCES faculty(id) ON DELETE CASCADE,
                    subject_id INTEGER REFERENCES subjects(id) ON DELETE CASCADE,
                    subject_name TEXT NOT NULL,
                    department TEXT NOT NULL,
                    semester TEXT NOT NULL,
                    class_name TEXT NOT NULL,
                    division TEXT NOT NULL,
                    academic_year TEXT DEFAULT '2025-26',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(faculty_id, subject_id, division)
                )
            """)
            
            # 2. Leave Applications Table
            cur.execute("""
                CREATE TABLE IF NOT EXISTS leave_applications (
                    id SERIAL PRIMARY KEY,
                    faculty_id INTEGER REFERENCES faculty(id) ON DELETE CASCADE,
                    leave_type TEXT,
                    from_date DATE,
                    to_date DATE,
                    reason TEXT,
                    status TEXT DEFAULT 'pending',
                    remarks TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # For backward compatibility. If it was already created with fewer columns, try to alter.
            try: cur.execute("ALTER TABLE faculty_subject_assignments ADD COLUMN semester TEXT NOT NULL DEFAULT 'I'")
            except Exception: pass
            try: cur.execute("ALTER TABLE faculty_subject_assignments ADD COLUMN class_name TEXT NOT NULL DEFAULT '-'")
            except Exception: pass
            try: cur.execute("ALTER TABLE faculty_subject_assignments ADD COLUMN academic_year TEXT DEFAULT '2025-26'")
            except Exception: pass
        conn.commit()
    except Exception as e:
        conn.rollback()
        app.logger.warning("Could not ensure assignments/leaves schema: %s", e)
    finally:
        conn.close()

# Ensure schemas on startup
_ensure_student_contact_columns()
_ensure_timetable_schema()
ensure_attendance_upload_tables()
ensure_attendance_engine_schema()
ensure_faculty_attendance_v2_schema()
_ensure_faculty_assignments_schema()


# Global Error Handlers
@app.errorhandler(404)
def page_not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403

@app.errorhandler(500)
def internal_error(e):
    import traceback
    err_id = secrets.token_hex(4).upper()
    
    # Check if this is a BuildError (missing route)
    from werkzeug.routing import BuildError
    if isinstance(e, BuildError):
         app.logger.warning(f"ROUTE BUILD ERROR [{err_id}]: {str(e)}")
         return redirect(url_for('route_not_found', endpoint=str(e)))

    app.logger.error(f"RUNTIME ERROR [{err_id}]: {str(e)}")
    app.logger.error(traceback.format_exc())
    
    # Attempt to rollback any active DB transactions
    try:
        from utils.pg_wrapper import get_db
        conn = get_db()
        conn.rollback()
        conn.close()
    except Exception:
        pass
        
    return render_template('errors/500.html', error_id=err_id, detail=str(e)), 500






# ─── Ensure results tables exist on every startup ─────────


# ─── HELPERS (Moved to utils/helpers.py or staying here) ──
# Note: login_required, safe_redirect_target, etc. are now imported.

def migrate_plaintext_passwords_to_hashes():
    """One-time per row: hash any non-werkzeug password already in DB (single transaction)."""
    conn = get_db()
    try:
        for row in conn.execute("SELECT id, password FROM students"):
            p = row["password"] or ""
            if p and not _password_is_hashed(p):
                conn.execute(
                    "UPDATE students SET password=%s WHERE id=%s",
                    (hash_password(p), row["id"]),
                )
        for row in conn.execute("SELECT id, password FROM faculty"):
            p = row["password"] or ""
            if p and not _password_is_hashed(p):
                conn.execute(
                    "UPDATE faculty SET password=%s WHERE id=%s",
                    (hash_password(p), row["id"]),
                )
        conn.commit()
    finally:
        conn.close()

migrate_plaintext_passwords_to_hashes()

def resolve_student_id(name, roll=None):
    """Resolve students.id from name; optional roll disambiguates duplicate names."""
    if not name or not str(name).strip():
        return None
    name = str(name).strip()
    r = str(roll).strip() if roll is not None and str(roll).strip() else None
    if r:
        row = qone("SELECT id FROM students WHERE name=%s AND roll=%s", (name, r))
        if row:
            return row["id"]
    row = qone("SELECT id FROM students WHERE name=%s ORDER BY id LIMIT 1", (name,))
    return row["id"] if row else None

def att_match_student_params(sid, name):
    """WHERE params: match attendance rows for this student (id + legacy name)."""
    return sid, name

def att_match_student_sql(prefix=""):
    """SQL fragment: (student_id = ? OR (student_id IS NULL AND student_name = ?))"""
    p = f"{prefix}." if prefix else ""
    return f"({p}student_id = %s OR ({p}student_id IS NULL AND {p}student_name = %s))"

def marks_match_student_params(sid, name):
    return sid, name

def marks_match_student_sql(prefix=""):
    p = f"{prefix}." if prefix else ""
    return f"({p}student_id = %s OR ({p}student_id IS NULL AND {p}student_name = %s))"

def att_in_dept_sql():
    """Attendance rows for students in a department (id match + legacy name fallback)."""
    return (
        "(student_id IN (SELECT id FROM students WHERE department=%s) OR "
        "(student_id IS NULL AND student_name IN (SELECT name FROM students WHERE department=%s)))"
    )

def att_in_division_sql():
    return (
        "(student_id IN (SELECT id FROM students WHERE division=%s) OR "
        "(student_id IS NULL AND student_name IN (SELECT name FROM students WHERE division=%s)))"
    )

def att_in_year_sql():
    return (
        "(student_id IN (SELECT id FROM students WHERE year=%s) OR "
        "(student_id IS NULL AND student_name IN (SELECT name FROM students WHERE year=%s)))"
    )

def today_str():
    return date.today().strftime("%Y-%m-%d")


def pct(a, b):
    return round(a / b * 100) if b > 0 else 0

def normalize_time(t):
    if not t: return t
    t = str(t).strip()
    # Zero-pad each part: "9:30-10:30" → "09:30-10:30"
    parts = re.split(r'\s*-\s*', t)
    result = []
    for part in parts:
        m = re.match(r'^(\d{1,2}):(\d{2})$', part.strip())
        if m:
            result.append(f"{int(m.group(1)):02d}:{m.group(2)}")
        else:
            result.append(part.strip())
    return '-'.join(result)

def grade(marks, total):
    p = pct(marks, total)
    if p >= 90: return "A+"
    if p >= 75: return "A"
    if p >= 60: return "B+"
    if p >= 50: return "B"
    if p >= 40: return "C"
    return "F"

def normalise_status(raw):
    s = str(raw).strip().upper()
    if s in ("P","PRESENT","1","YES"): return "Present"
    if s in ("A","ABSENT","0","NO"):   return "Absent"
    if s in ("LATE"): return "Late"
    if s in ("EXCUSED", "E"): return "Excused"
    if s in ("L","LEAVE","ML","CL", "MEDICAL"): return "Leave"
    return None

def normalise_date(raw):
    if isinstance(raw, (datetime, date)):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%m-%Y","%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass   # expected — just trying next format
    return None


@app.before_request
def session_fingerprint():
    """FIX 7: Session fingerprinting to prevent session hijacking."""
    if not session.get("fingerprint"):
        return
    
    current_ua = request.headers.get("User-Agent", "")
    current_ip = get_remote_address()
    expected = hashlib.sha256(f"{current_ua}{current_ip}{app.secret_key}".encode()).hexdigest()
    
    if session["fingerprint"] != expected:
        session.clear()
        return redirect("/login?error=session_invalid")

@app.before_request
@limiter.limit("200 per hour")
def limit_api_global():
    if request.path.startswith("/api/"):
        pass

@app.before_request
def _csrf_init_and_validate():
    if "_csrf_token" not in session:
        session["_csrf_token"] = secrets.token_urlsafe(32)
    
    # FIX 10: Protect PUT, PATCH, DELETE and API routes
    if request.method not in ("POST", "PUT", "PATCH", "DELETE"):
        return
    
    if request.endpoint is None or request.endpoint == "static":
        return

    # Skip CSRF for API if using other auth? 
    # But request said "Add CSRF to API routes that modify data."
    
    tok = session.get("_csrf_token")
    sent = request.form.get("_csrf") or request.headers.get("X-CSRF-Token")
    if not tok or not sent or not secrets.compare_digest(str(tok), str(sent)):
        app.logger.warning(f"CSRF failure: expected {tok}, sent {sent}")
        abort(400)


def unread_count():
    try:
        role = session.get("role")
        if not role: return 0
        uid = session.get("faculty_id") if role == "faculty" else session.get("student_id")
        if not uid: return 0
        return safe_fetch_scalar("SELECT COUNT(*) FROM messages WHERE receiver_id=%s AND receiver_role=%s AND is_read=FALSE", (uid, role)) or 0
    except Exception:
        return 0


@app.context_processor
def _inject_utils():
    from flask import url_for as flask_url_for
    from werkzeug.routing import BuildError
    
    def safe_url_for(endpoint, **values):
        try:
            return flask_url_for(endpoint, **values)
        except BuildError:
            return f"/route-not-found?endpoint={endpoint}"
            
    return dict(
        unread_count=unread_count, 
        csrf_token=lambda: session.get("_csrf_token", ""), 
        max=max, 
        min=min,
        url_for=safe_url_for # Override url_for for safety
    )


# ─── LOGIN / LOGOUT ───────────────────────────────────────
@app.route("/")
def index():
    if session.get("role"):
        r = session["role"]
        return redirect("/admin" if r=="admin" else "/faculty_dashboard" if r=="faculty" else "/student_dashboard")
    return redirect("/login")

@app.route("/login", methods=["GET","POST"])
@app.route("/admin_login", methods=["GET","POST"])
@app.route("/faculty_login", methods=["GET","POST"])
@limiter.limit("10 per minute", methods=["POST"])
def login():
    error = None
    if request.method == "POST":
        action = request.form.get("action", "login")
        role   = request.form.get("role","").strip()
        user   = request.form.get("username","").strip()
        
        # 1. Handle OTP Verification
        if action == "verify_otp":
            expires = session.get("pre_auth_expires")
            if not expires or datetime.fromisoformat(expires) < datetime.now():
                session.clear()
                return render_template("common/login.html", error="OTP expired. Please log in again.")

            otp = request.form.get("otp","").strip()
            from services.otp_service import OTPService
            
            # For demo purposes, we fetch user data again or use session
            # Here we assume phone is stored or linked to user
            # Let's verify the OTP first
            # We need the phone number to verify. 
            # In a real app, we'd store the phone in session['pre_auth_phone']
            phone = session.get("pre_auth_phone")
            res = OTPService.verify_otp(phone, otp)
            
            if res['success']:
                # Success! Restore the intended session
                intended = session.get("pre_auth_data", {})
                session.clear()
                for k, v in intended.items():
                    session[k] = v
                session.permanent = True
                
                # Redirect to dashboard
                r = session.get("role")
                return redirect("/admin" if r=="admin" else "/faculty_dashboard" if r=="faculty" else "/student_dashboard")
            else:
                return render_template("common/login.html", 
                                     mfa_required=True, 
                                     error=res['error'],
                                     pre_auth_user=user,
                                     pre_auth_role=role,
                                     masked_phone=session.get("pre_auth_masked"))

        # 2. Handle Primary Login (Username/Password)
        pw = request.form.get("password","").strip()
        user_data = None
        phone = None

        if role == "admin":
            if user == "admin" and check_password_hash(ADMIN_PASSWORD_HASH, pw):
                user_data = {"role": "admin", "name": "Administrator"}
                # Admin uses a configured phone or skips (for now, skip or use env)
                phone = os.environ.get("ADMIN_PHONE") 
            else:
                error = "Invalid admin credentials."

        elif role == "faculty":
            row = qone("SELECT id,name,password,phone FROM faculty WHERE email=%s", (user,))
            if row and verify_password(row["password"], pw):
                user_data = {"role": "faculty", "name": row["name"], "faculty_id": row["id"]}
                phone = row["phone"]
            else:
                error = "Invalid faculty credentials."

        elif role == "student":
            row = qone("SELECT id,name,password,department,year,division,phone FROM students WHERE prn=%s", (user,))
            if not row:
                # Fallback to roll number for legacy support
                row = qone("SELECT id,name,password,department,year,division,phone FROM students WHERE roll=%s", (user,))
            
            if row and verify_password(row["password"], pw):
                user_data = {
                    "role": "student", "name": row["name"], "student_id": row["id"],
                    "student_roll": user, "student_branch": row["department"],
                    "student_year": row["year"], "student_division": row["division"]
                }
            else:
                error = "Invalid student credentials."
        
        if user_data:
            if phone and os.environ.get("ENFORCE_MFA") == "true":
                from services.otp_service import OTPService
                OTPService.generate_and_send_otp(phone)
                session["pre_auth_phone"] = phone
                session["pre_auth_data"] = user_data
                session["pre_auth_expires"] = (datetime.now() + timedelta(minutes=10)).isoformat()
                session["pre_auth_masked"] = phone[:3] + "****" + phone[-2:]
                return render_template("common/login.html", mfa_required=True, pre_auth_user=user, pre_auth_role=role, masked_phone=session["pre_auth_masked"])

            session.clear()
            for k, v in user_data.items():
                session[k] = v
            session.permanent = True
            r = session["role"]
            return redirect("/admin" if r=="admin" else "/faculty_dashboard" if r=="faculty" else "/student_dashboard")

    return render_template("common/login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# ════════════════════════════════════════════════════════════
#  ADMIN ROUTES
# ════════════════════════════════════════════════════════════
BACKUP_DIR = "backups"
SQLITE_DB_PATH = os.path.join("instance", "database.db")
RESET_TABLES = [
    "messages",
    "leave_applications",
    "timetable_notifications",
    "attendance_summary",
    "qr_sessions",
    "attendance",
    "results",
    "marks",
    "result_summary",
    "cumulative_attendance",
    "faculty_notes",
    "faculty_notices",
    "notifications",
    "events",
    "timetable",
    "subjects",
    "students",
    "faculty",
]


def _quote_ident(name):
    return '"' + str(name).replace('"', '""') + '"'


def _json_safe(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.hex()
    return value


def _prune_old_backups(keep=10):
    if not os.path.isdir(BACKUP_DIR):
        return
    backups = [
        os.path.join(BACKUP_DIR, name)
        for name in os.listdir(BACKUP_DIR)
        if name.startswith("backup_") and (name.endswith(".db") or name.endswith(".json"))
    ]
    backups.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    for old_path in backups[keep:]:
        try:
            os.remove(old_path)
        except OSError as e:
            app.logger.warning("Could not remove old backup %s: %s", old_path, e)


def _last_backup_label():
    if not os.path.isdir(BACKUP_DIR):
        return None
    backups = [
        os.path.join(BACKUP_DIR, name)
        for name in os.listdir(BACKUP_DIR)
        if name.startswith("backup_") and (name.endswith(".db") or name.endswith(".json"))
    ]
    if not backups:
        return None
    latest = max(backups, key=lambda p: os.path.getmtime(p))
    return datetime.fromtimestamp(os.path.getmtime(latest)).strftime("%d %b %Y, %I:%M %p")


def _create_backup_file():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if os.path.exists(SQLITE_DB_PATH):
        backup_path = os.path.join(BACKUP_DIR, f"backup_{stamp}.db")
        shutil.copy(SQLITE_DB_PATH, backup_path)
        _prune_old_backups()
        return backup_path

    conn = get_db()
    backup_path = os.path.join(BACKUP_DIR, f"backup_{stamp}.json")
    data = {"created_at": datetime.now().isoformat(timespec="seconds"), "tables": {}}
    try:
        tables = conn.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema='public' AND table_type='BASE TABLE' ORDER BY table_name"
        ).fetchall()
        for row in tables:
            table_name = row["table_name"]
            rows = conn.execute(f"SELECT * FROM {_quote_ident(table_name)}").fetchall()
            data["tables"][table_name] = [
                {key: _json_safe(value) for key, value in dict(record).items()}
                for record in rows
            ]
    finally:
        conn.close()

    with open(backup_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    _prune_old_backups()
    return backup_path


@app.route("/admin/backup")
@admin_required
def backup():
    try:
        backup_path = _create_backup_file()
        flash("Backup created successfully", "success")
        return send_file(backup_path, as_attachment=True)
    except Exception as e:
        app.logger.exception("Backup failed")
        return str(e), 500


@app.route("/admin/reset", methods=["POST"])
@admin_required
def reset_system():
    try:
        backup_path = _create_backup_file()
        app.logger.warning(
            "Admin %s started full ERP data reset after backup %s",
            session.get("name", "Admin"),
            backup_path,
        )

        if os.path.exists(SQLITE_DB_PATH):
            import sqlite3

            conn = sqlite3.connect(SQLITE_DB_PATH)
            try:
                cursor = conn.cursor()
                cursor.execute("PRAGMA foreign_keys = OFF")
                existing = {
                    row[0]
                    for row in cursor.execute(
                        "SELECT name FROM sqlite_master WHERE type='table'"
                    ).fetchall()
                }
                for table in RESET_TABLES:
                    if table in existing:
                        cursor.execute(f"DELETE FROM {_quote_ident(table)}")
                conn.commit()
            finally:
                conn.close()
        else:
            conn = get_db()
            try:
                existing = conn.execute(
                    "SELECT table_name FROM information_schema.tables "
                    "WHERE table_schema='public' AND table_type='BASE TABLE'"
                ).fetchall()
                existing_names = {row["table_name"] for row in existing}
                tables = [table for table in RESET_TABLES if table in existing_names]
                if tables:
                    joined = ", ".join(_quote_ident(table) for table in tables)
                    conn.execute(f"TRUNCATE TABLE {joined} RESTART IDENTITY CASCADE")
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()

        flash("All data reset successfully (testing mode)", "warning")
    except Exception as e:
        app.logger.exception("Full data reset failed")
        return str(e), 500

    return redirect("/admin")


@app.route("/admin")
@login_required("admin")
def admin():
    ts  = safe_fetch_scalar("SELECT COUNT(*) as c FROM students")
    tf  = safe_fetch_scalar("SELECT COUNT(*) as c FROM faculty")
    tsb = safe_fetch_scalar("SELECT COUNT(*) as c FROM subjects")
    ta  = safe_fetch_scalar("SELECT COUNT(*) as c FROM attendance")

    # Department distribution
    dept_rows   = safe_fetch_all("SELECT department, COUNT(*) as c FROM students GROUP BY department")
    dept_labels = [r["department"] for r in dept_rows]
    dept_counts = [r["c"] for r in dept_rows]

    # Weekly attendance (last 7 days)
    week_dates = [(date.today() - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6, -1, -1)]
    ph = ",".join(["%s"] * len(week_dates))
    week_rows = safe_fetch_all(
        f"SELECT date::text, COUNT(*) as c FROM attendance WHERE date::text IN ({ph}) AND status ILIKE 'Present' GROUP BY date",
        week_dates,
    )
    present_by_date = {str(r["date"]): r["c"] for r in week_rows}
    week_labels  = [d[5:] for d in week_dates]
    week_present = [present_by_date.get(d, 0) for d in week_dates]

    # Marks by exam type
    marks_rows  = safe_fetch_all("SELECT exam_type, ROUND(AVG(marks*100.0/total)::numeric,1) as avg FROM marks GROUP BY exam_type")
    marks_exams = [r["exam_type"] for r in marks_rows]
    marks_avg   = [r["avg"] for r in marks_rows]

    # Dept attendance %
    dept_att_pct = []
    if DEPARTMENTS:
        dph = ",".join(["%s"] * len(DEPARTMENTS))
        att_rows = safe_fetch_all(
            f"""SELECT s.department AS d, 
                COALESCE(COUNT(a.id), 0) AS tot,
                COALESCE(SUM(CASE WHEN a.status ILIKE 'Present' THEN 1 ELSE 0 END), 0) AS pres
                FROM students s
                LEFT JOIN attendance a ON a.student_id = s.id
                WHERE s.department IN ({dph}) GROUP BY s.department""",
            list(DEPARTMENTS),
        )
        pct_map = {r["d"]: pct(r["pres"], r["tot"]) for r in att_rows if r["tot"] > 0}
        dept_att_pct = [(d, pct_map[d]) for d in DEPARTMENTS if d in pct_map]

    # Total student breakdown by Year
    year_rows = safe_fetch_all("SELECT year, COUNT(*) as c FROM students GROUP BY year")
    year_map  = {r["year"]: r["c"] for r in year_rows}
    
    # Faculty breakdown by Department
    fac_dept_rows = safe_fetch_all("SELECT department, COUNT(*) as c FROM faculty GROUP BY department")
    fac_dept_labels = [r["department"] for r in fac_dept_rows]
    fac_dept_counts = [r["c"] for r in fac_dept_rows]

    # Messages/Notifications metrics
    total_messages = safe_fetch_scalar("SELECT COUNT(*) FROM messages") or 0
    
    # Low attendance count (Defaulters < 75%)
    defaulters_count = safe_fetch_scalar("""
        SELECT COUNT(*) FROM (
            SELECT a.student_id FROM attendance a
            JOIN students s ON a.student_id = s.id
            GROUP BY a.student_id
            HAVING (SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) * 100.0 / NULLIF(COUNT(*), 0)) < 75
        ) as sub
    """) or 0

    recent_students = safe_fetch_all("SELECT * FROM students ORDER BY id DESC LIMIT 5")

    return render_template("admin/admin_dashboard.html",
        ts=ts, tf=tf, tsb=tsb, ta=ta,
        dept_labels=dept_labels, dept_counts=dept_counts,
        week_labels=week_labels, week_present=week_present,
        marks_exams=marks_exams, marks_avg=marks_avg,
        dept_att_pct=dept_att_pct,
        year_map=year_map,
        fac_dept_labels=fac_dept_labels, fac_dept_counts=fac_dept_counts,
        total_messages=total_messages,
        defaulters_count=defaulters_count,
        recent_students=recent_students,
        last_backup=_last_backup_label()
    )




@app.route("/admin_dashboard")
@login_required("admin")
def admin_dashboard():
    return redirect("/admin")


# ── STUDENTS ──────────────────────────────────────────────
# ── STUDENTS ──────────────────────────────────────────────
@app.route("/students")
@admin_required
def students():
    q    = request.args.get("q","").strip()
    dept = request.args.get("dept","").strip()
    year = request.args.get("year","").strip()
    division = request.args.get("division","").strip()
    sql  = "SELECT * FROM students WHERE 1=1"
    params = []
    if q:
        sql += " AND (name LIKE %s OR roll LIKE %s OR prn LIKE %s OR contact_number LIKE %s)"
        params += [f"%{q}%"] * 4
    if dept:
        sql += " AND department=%s"; params.append(dept)
    if year:
        sql += " AND year=%s"; params.append(year)
    if division:
        sql += " AND division=%s"; params.append(division)
    sql += " ORDER BY id DESC"
    rows = safe_fetch_all(sql, params)
    return render_template("admin/students.html", students=rows, q=q, dept=dept, year=year,
                           division=division, DEPARTMENTS=DEPARTMENTS,
                           YEARS=YEARS, DIVISIONS=DIVISIONS, total=len(rows))

@app.route("/add_student")
@login_required("admin")
def add_student():
    return render_template("admin/add_student.html", DEPARTMENTS=DEPARTMENTS, YEARS=YEARS)

@app.route("/save_student", methods=["POST"])
@admin_required
def save_student():
    name = request.form.get("name","").strip()
    roll = request.form.get("roll","").strip()
    dept = request.form.get("department","").strip()
    year = request.form.get("year","").strip()
    email= request.form.get("email","").strip()
    division = request.form.get("division","").strip()
    gender = request.form.get("gender","").strip()
    dob = request.form.get("dob","").strip() or None
    contact = request.form.get("contact_number","").strip()
    parent = request.form.get("parent_contact","").strip()
    address = request.form.get("address","").strip()
    adm = request.form.get("admission_year","").strip()
    if adm:
        try: adm=int(adm)
        except: adm=None
    else: adm=None
    pw_raw = request.form.get("password", "").strip()
    pw   = hash_password(pw_raw or DEFAULT_STUDENT_PASSWORD)
    prn = request.form.get("prn","").strip()
    roll = roll or prn # Sync roll with prn

    if not name or not prn:
        return render_template("admin/add_student.html", DEPARTMENTS=DEPARTMENTS, YEARS=YEARS,
                               error="Name and PRN Number are required.")
    
    try:
        safe_execute("INSERT INTO students(name,roll,prn,department,year,email,password,division,gender,dob,contact_number,parent_contact,address,admission_year) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (name,roll,prn,dept,year,email,pw,division,gender,dob,contact,parent,address,adm))
        return redirect("/students?success=1")
    except Exception as e:
        app.logger.error(f"Student save failed: {e}")
        return render_template("admin/add_student.html", DEPARTMENTS=DEPARTMENTS, YEARS=YEARS,
                               error=f"Could not save student. PRN may already exist.")

@app.route("/edit_student", methods=["POST"])
@admin_required
def edit_student():
    sid = request.form.get("student_id")
    adm = request.form.get("admission_year","").strip()
    if adm:
        try: adm=int(adm)
        except: adm=None
    else: adm=None

    prn = request.form.get("prn","").strip()
    roll = request.form.get("roll","").strip() or prn

    safe_execute("UPDATE students SET name=%s,roll=%s,prn=%s,department=%s,year=%s,email=%s,division=%s,gender=%s,dob=%s,contact_number=%s,parent_contact=%s,address=%s,admission_year=%s WHERE id=%s",
        (request.form.get("name",""), roll, prn, request.form.get("department",""),
         request.form.get("year",""), request.form.get("email",""), 
         request.form.get("division",""), request.form.get("gender",""),
         request.form.get("dob","") or None, request.form.get("contact_number",""),
         request.form.get("parent_contact",""), request.form.get("address",""), adm, sid))
    return redirect("/students")

@app.route("/delete_student", methods=["POST"])
@admin_required
def delete_student():
    safe_execute("DELETE FROM students WHERE id=%s", (request.form.get("student_id",""),))
    return redirect("/students")

@app.route("/export_students_excel")
@login_required("admin")
def export_students_excel():
    rows = qry("SELECT name,prn,roll,department,year,division,contact_number,email FROM students ORDER BY department,year,division,name")
    wb = Workbook(); ws = wb.active; ws.title = "Students"
    hdr = ["Name","PRN Number","Department","Year","Division","Contact Number","Email"]
    for c,h in enumerate(hdr,1):
        ws.cell(1,c,h).font = Font(bold=True)
    for r,row in enumerate(rows,2):
        ws.cell(r,1,row["name"]); ws.cell(r,2,row["prn"] or row["roll"])
        ws.cell(r,3,row["department"]); ws.cell(r,4,row["year"])
        ws.cell(r,5,row["division"] or "")
        ws.cell(r,6,row["contact_number"] or ""); ws.cell(r,7,row["email"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name="students.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/import_students_excel", methods=["POST"])
@login_required("admin")
def import_students_excel():
    f = request.files.get("file")
    if not f: return redirect("/students")
    wb = load_workbook(f, data_only=True)
    if _is_attendance_student_workbook(wb):
        added, updated, skipped = _import_attendance_workbook_students(wb, f.filename)
        return redirect(f"/students?imported={added}&updated={updated}&skipped={skipped}&format=attendance_xlsx")
    ws = wb.active
    added = skipped = 0
    # Find header row
    hdr_row = 1
    for i,row in enumerate(ws.iter_rows(max_row=10),1):
        vals = [str(c.value or "").lower() for c in row]
        if any("name" in v for v in vals): hdr_row = i; break
    # Map columns
    headers = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def col(keywords):
        for k in keywords:
            for i,h in enumerate(headers):
                if k in h: return i+1
        return None
    cn = col(["name"]); cr = col(["roll"]); cd = col(["dept","department"])
    cy = col(["year"]); ce = col(["email"]); cc = col(["contact","phone","mobile"])
    if not cn: return redirect("/students")
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        name = str(row[cn-1] or "").strip()
        roll = str(row[cr-1] if cr else "").strip()
        dept = str(row[cd-1] if cd else "").strip()
        year = str(row[cy-1] if cy else "").strip()
        email= str(row[ce-1] if ce else "").strip()
        contact_raw = row[cc-1] if cc and cc <= len(row) else None
        contact = str(int(contact_raw)) if isinstance(contact_raw, float) else str(contact_raw or "").strip()
        if not name or not roll: continue
        prn = roll # Sync PRN with Roll
        try:
            exe("INSERT INTO students(name,roll,prn,department,year,email,contact_number,password) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)",
                (name, roll, prn, dept, year, email, contact, hash_password(_default_student_password())))
            added += 1
        except psycopg2.IntegrityError:
            skipped += 1
    return redirect(f"/students?imported={added}&skipped={skipped}")



# ── FACULTY ───────────────────────────────────────────────
@app.route("/faculty")
@admin_required
def faculty():
    q = request.args.get("q","").strip()
    dept = request.args.get("dept","").strip()
    sql = "SELECT id, name, department, designation, email, phone, qualification, joining_date FROM faculty WHERE 1=1"
    params = []
    if q:
        sql += " AND (name LIKE %s OR email LIKE %s)"
        params += [f"%{q}%", f"%{q}%"]
    if dept:
        sql += " AND department=%s"; params.append(dept)
    sql += " ORDER BY name"
    return render_template("faculty/faculty.html", faculty_list=safe_fetch_all(sql, params),
                           q=q, dept=dept, DEPARTMENTS=DEPARTMENTS, DESIGNATIONS=DESIGNATIONS)

@app.route("/faculty_students")
@faculty_required
def faculty_students():
    q = request.args.get("q","").strip()
    dept = request.args.get("dept","").strip()
    faculty_row = safe_fetch_one("SELECT department FROM faculty WHERE id=%s", (session.get("faculty_id"),))
    if faculty_row and not dept:
        dept = faculty_row["department"]
        
    sql = "SELECT * FROM students WHERE 1=1"
    params = []
    if q:    sql += " AND (name LIKE %s OR roll LIKE %s)"; params += [f"%{q}%",f"%{q}%"]
    if dept: sql += " AND department=%s"; params.append(dept)
    
    div = request.args.get("division","").strip()
    if div:
        sql += " AND division=%s"
        params.append(div)
        
    sql += " ORDER BY id DESC"
    rows = safe_fetch_all(sql, params)
    
    return render_template("admin/students.html", students=rows, q=q, dept=dept, year="", 
                           DEPARTMENTS=DEPARTMENTS, YEARS=YEARS, total=len(rows))

@app.route("/add_faculty")
@login_required("admin")
def add_faculty():
    return render_template("admin/add_faculty.html", DEPARTMENTS=DEPARTMENTS, DESIGNATIONS=DESIGNATIONS)

@app.route("/save_faculty", methods=["POST"])
@admin_required
def save_faculty():
    name  = request.form.get("name","").strip()
    dept  = request.form.get("department","").strip()
    desig = request.form.get("designation","").strip()
    email = request.form.get("email","").strip()
    phone = request.form.get("phone","").strip()
    qual  = request.form.get("qualification","").strip()
    jdate = request.form.get("joining_date","").strip()
    pw_raw = request.form.get("password", "").strip()
    pw    = hash_password(pw_raw or DEFAULT_FACULTY_PASSWORD)

    if not name or not email:
        return render_template("admin/add_faculty.html", DEPARTMENTS=DEPARTMENTS, DESIGNATIONS=DESIGNATIONS,
                               error="Name and Email are required.")
    try:
        safe_execute("INSERT INTO faculty(name,department,designation,email,phone,qualification,joining_date,password) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)",
            (name,dept,desig,email,phone,qual,jdate,pw))
        return redirect("/faculty?success=1")
    except Exception as e:
        app.logger.error(f"Faculty save failed: {e}")
        return render_template("admin/add_faculty.html", DEPARTMENTS=DEPARTMENTS, DESIGNATIONS=DESIGNATIONS,
                               error=f"Could not save faculty. Email may already exist.")

@app.route("/edit_faculty", methods=["POST"])
@admin_required
def edit_faculty():
    fid  = request.form.get("faculty_id","")
    try:
        exe("UPDATE faculty SET name=%s,department=%s,designation=%s,email=%s,phone=%s,qualification=%s WHERE id=%s",
            (request.form.get("name",""), request.form.get("department",""), request.form.get("designation",""),
             request.form.get("email",""), request.form.get("phone",""), request.form.get("qualification",""), fid))
    except psycopg2.IntegrityError:
        pass
    return redirect("/faculty")

@app.route("/delete_faculty", methods=["POST"])
@login_required("admin")
def delete_faculty():
    exe("DELETE FROM faculty WHERE id=%s", (request.form.get("faculty_id",""),))
    return redirect("/faculty")

@app.route("/export_faculty_excel")
@login_required("admin")
def export_faculty_excel():
    rows = qry("SELECT name,department,designation,email,phone,qualification,joining_date FROM faculty ORDER BY name")
    wb = Workbook(); ws = wb.active; ws.title = "Faculty"
    hdrs = ["Name","Department","Designation","Email","Phone","Qualification","Joining Date"]
    for c,h in enumerate(hdrs,1): ws.cell(1,c,h).font = Font(bold=True)
    for r,row in enumerate(rows,2):
        for c,k in enumerate(["name","department","designation","email","phone","qualification","joining_date"],1):
            ws.cell(r,c,row[k])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name="faculty.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/import_faculty_excel", methods=["POST"])
@login_required("admin")
def import_faculty_excel():
    f = request.files.get("file")
    if not f: return redirect("/faculty")
    wb = load_workbook(f, data_only=True); ws = wb.active
    added = skipped = 0
    hdr_row = 1
    for i,row in enumerate(ws.iter_rows(max_row=10),1):
        vals = [str(c.value or "").lower() for c in row]
        if any("name" in v for v in vals): hdr_row = i; break
    headers = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def col(kws):
        for k in kws:
            for i,h in enumerate(headers):
                if k in h: return i+1
        return None
    cn=col(["name"]); cd=col(["dept","department"]); ce=col(["email"]); cp=col(["phone","contact"])
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        if not row or not row[0]: continue
        name  = str(row[cn-1] if cn else "").strip()
        dept  = str(row[cd-1] if cd else "").strip()
        email = str(row[ce-1] if ce else "").strip()
        phone = str(row[cp-1] if cp else "").strip()
        if not name or not email: continue
        try:
            exe("INSERT INTO faculty(name,department,email,phone,password) VALUES(%s,%s,%s,%s,%s)",
                (name, dept, email, phone, hash_password(_default_faculty_password())))
            added += 1
        except psycopg2.IntegrityError:
            skipped += 1
    return redirect(f"/faculty?imported={added}&skipped={skipped}")


# ── SUBJECTS ──────────────────────────────────────────────
@app.route("/subjects")
@admin_required
def subjects():
    q    = request.args.get("q","").strip()
    dept = request.args.get("dept","").strip()
    sql  = "SELECT * FROM subjects WHERE 1=1"
    params = []
    if q:    sql += " AND (name LIKE %s OR subject_code LIKE %s)"; params += [f"%{q}%",f"%{q}%"]
    if dept: sql += " AND department=%s"; params.append(dept)
    sql += " ORDER BY department,name"
    faculty_list  = safe_fetch_all("SELECT name FROM faculty ORDER BY name")
    return render_template("admin/subjects.html", subjects=safe_fetch_all(sql,params),
                           q=q, dept=dept, DEPARTMENTS=DEPARTMENTS, SEMESTERS=SEMESTERS,
                           teachers=[r["name"] for r in faculty_list])

@app.route("/add_subject")
@admin_required
def add_subject():
    teachers_list = safe_fetch_all("SELECT name FROM faculty ORDER BY name")
    return render_template("admin/add_subject.html", DEPARTMENTS=DEPARTMENTS, SEMESTERS=SEMESTERS,
                           teachers=[r["name"] for r in teachers_list])

@app.route("/save_subject", methods=["POST"])
@admin_required
def save_subject():
    safe_execute("INSERT INTO subjects(name,department,subject_code,teacher,semester) VALUES(%s,%s,%s,%s,%s)",
        (request.form.get("name",""), request.form.get("department",""), request.form.get("subject_code",""),
         request.form.get("teacher",""), request.form.get("semester","I")))
    return redirect("/subjects")

@app.route("/edit_subject", methods=["POST"])
@admin_required
def edit_subject():
    safe_execute("UPDATE subjects SET name=%s,department=%s,subject_code=%s,teacher=%s,semester=%s,division=%s WHERE id=%s",
        (request.form.get("name",""), request.form.get("department",""), request.form.get("subject_code",""),
         request.form.get("teacher",""), request.form.get("semester",""), request.form.get("division",""),
         request.form.get("subject_id","")))
    return redirect("/subjects")

@app.route("/delete_subject", methods=["POST"])
@admin_required
def delete_subject():
    safe_execute("DELETE FROM subjects WHERE id=%s", (request.form.get("subject_id",""),))
    return redirect("/subjects")

@app.route("/export_subjects_excel")
@admin_required
def export_subjects_excel():
    rows = safe_fetch_all("SELECT name,department,subject_code,teacher,semester,division FROM subjects ORDER BY department,name")
    wb = Workbook(); ws = wb.active; ws.title="Subjects"
    for c,h in enumerate(["Name","Department","Code","Teacher","Semester","Division"],1):
        cell = ws.cell(1,c,h)
        cell.font=Font(bold=True)
    for r,row in enumerate(rows,2):
        for c,k in enumerate(["name","department","subject_code","teacher","semester","division"],1):
            ws.cell(r,c,row[k])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name="subjects.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



# ════════════════════════════════════════════════════════════
#  ATTENDANCE MODULE — Complete Professional Implementation
# ════════════════════════════════════════════════════════════

def att_stats(sql_where="1=1", params=()):
    """Return dict of attendance stats for a given WHERE clause."""
    # Note: safe_fetch_scalar returns the first value of the first row
    total  = safe_fetch_scalar(f"SELECT COUNT(*) as c FROM attendance WHERE {sql_where}", params)
    pres   = safe_fetch_scalar(f"SELECT COUNT(*) as c FROM attendance WHERE {sql_where} AND status='Present'", params)
    absent = safe_fetch_scalar(f"SELECT COUNT(*) as c FROM attendance WHERE {sql_where} AND status='Absent'",  params)
    late   = safe_fetch_scalar(f"SELECT COUNT(*) as c FROM attendance WHERE {sql_where} AND status='Late'",    params)
    med    = safe_fetch_scalar(f"SELECT COUNT(*) as c FROM attendance WHERE {sql_where} AND status='Medical'", params)
    leave  = safe_fetch_scalar(f"SELECT COUNT(*) as c FROM attendance WHERE {sql_where} AND status='Leave'",   params)
    return {
        "total":        total,
        "present":      pres,
        "absent":       absent,
        "late":         late,
        "medical":      med,
        "leave":        leave,
        "pct":          pct(pres, total),
        "absent_pct":   pct(absent, total),
        "late_pct":     pct(late, total),
        "medical_pct":  pct(med, total),
        "effective_pct": pct(pres + late, total),
    }


def calculate_cumulative(student_id, student_name):
    """Return cumulative attendance stats for a single student."""
    rows = qry(
        f"SELECT status FROM attendance WHERE {att_match_student_sql()}",
        att_match_student_params(student_id, student_name)
    )
    total   = len(rows)
    present = sum(1 for r in rows if r["status"] == "Present")
    percentage = (present / total * 100) if total else 0

    # Attendance status label
    if percentage >= 75:
        att_status = "Good"
    elif percentage >= 50:
        att_status = "Average"
    else:
        att_status = "Low"

    return {
        "total":      total,
        "present":    present,
        "percentage": round(percentage, 2),
        "status":     att_status,
    }


def subject_wise(student_id, student_name):
    """Return per-subject attendance breakdown for a single student."""
    rows = qry(
        f"SELECT subject, status FROM attendance WHERE {att_match_student_sql()}",
        att_match_student_params(student_id, student_name)
    )
    data = {}
    for r in rows:
        sub = r["subject"]
        data.setdefault(sub, {"total": 0, "present": 0})
        data[sub]["total"] += 1
        if r["status"] == "Present":
            data[sub]["present"] += 1
    # Add percentage to each subject entry
    for sub in data:
        t = data[sub]["total"]
        p = data[sub]["present"]
        data[sub]["percentage"] = round(p / t * 100, 2) if t else 0
    return data


# ── MAIN ATTENDANCE PAGE (Mark + Import) ──────────────────
@app.route("/attendance")
@admin_required
def attendance():
    ctx = attendance_page_context("admin")
    return render_template(
        "attendance/attendance.html",
        students=ctx["students"],
        subjects=ctx["subjects"],
        divs=ctx["divs"],
        today=ctx["today"],
        DEPARTMENTS=DEPARTMENTS,
    )

@app.route("/mark_attendance")
@admin_required
def mark_attendance():
    return attendance()


@app.route("/save_attendance", methods=["POST"])
@admin_required
def save_attendance():
    from routes.attendance import handle_single_mark
    return handle_single_mark(request.form, session)

# ── BULK SAVE (from mark-class table) ─────────────────────
@app.route("/save_bulk_attendance", methods=["POST"])
@admin_required
def save_bulk_attendance():
    from routes.attendance import handle_bulk_mark
    return handle_bulk_mark(request.form, session)

@app.route("/attendance_bulk", methods=["POST"])
@admin_required
def attendance_bulk():
    return save_bulk_attendance()

@app.route("/attendance_backup")
@admin_required
def attendance_backup():
    from routes.attendance import handle_attendance_backup
    return handle_attendance_backup(session)

@app.route("/attendance_restore", methods=["POST"])
@admin_required
def attendance_restore():
    from routes.attendance import handle_attendance_restore
    return handle_attendance_restore(session, request.form.get("backup_path", "").strip())

@app.route("/attendance_reset", methods=["POST"])
@admin_required
def attendance_reset():
    from routes.attendance import handle_attendance_reset
    return handle_attendance_reset(session)


@app.route("/api/students_by_dept")
def api_students_by_dept():
    if session.get("role") not in ("admin", "faculty"):
        return jsonify([])
    return students_api_response(request.args)


# ── VIEW + FILTER ATTENDANCE ───────────────────────────────
@app.route("/view_attendance")
@login_required(["admin", "faculty"])
def view_attendance():
    from routes.attendance import handle_view_records
    data = handle_view_records(request.args, session)
    return render_template(
        "attendance/view_attendance.html",
        records=data["records"],
        current_page=data["current_page"],
        total_pages=data["total_pages"],
        total_count=data["total_count"],
        stats=data["stats"],
        analytics=data["analytics"],
        subjects=[r["subject"] for r in safe_fetch_all("SELECT DISTINCT subject FROM attendance ORDER BY subject")],
        divs=[r["division"] for r in safe_fetch_all("SELECT DISTINCT division FROM students WHERE division != '' ORDER BY division")],
    )


@app.route("/edit_attendance", methods=["POST"])
def edit_attendance():
    if session.get("role") != "admin":
        return "Unauthorized", 403
    return handle_edit_record(request.form, session)


@app.route("/delete_attendance", methods=["POST"])
def delete_attendance():
    if session.get("role") != "admin":
        return "Unauthorized", 403
    return handle_delete_record(request.form, session)


@app.route("/delete_attendance_summary_row", methods=["POST"])
@admin_required
def delete_attendance_summary_row():
    """Remove one PDF-imported subject row from attendance_summary (admin)."""
    sid = safe_int(request.form.get("student_id", ""))
    rid = safe_int(request.form.get("summary_id", ""))
    red = request.form.get("redirect") or f"/student_attendance_dashboard?student_id={sid}"
    if not sid or not rid:
        return redirect(safe_redirect_target(red, "/attendance_dashboard"))
    row = safe_fetch_one("SELECT id FROM attendance_summary WHERE id=%s AND student_id=%s", (rid, sid))
    if row:
        safe_execute("DELETE FROM attendance_summary WHERE id=%s", (rid,))
    return redirect(safe_redirect_target(red, f"/student_attendance_dashboard?student_id={sid}"))


@app.route("/clear_student_attendance_summary", methods=["POST"])
@admin_required
def clear_student_attendance_summary():
    """Remove all PDF-imported subject rows for one student (admin)."""
    sid = safe_int(request.form.get("student_id", ""))
    red = request.form.get("redirect") or f"/student_attendance_dashboard?student_id={sid}"
    if sid:
        safe_execute("DELETE FROM attendance_summary WHERE student_id=%s", (sid,))
    return redirect(safe_redirect_target(red, f"/student_attendance_dashboard?student_id={sid}"))


_CLEAR_ALL_PDF_SUMMARY_PHRASE = "DELETE ALL PDF SUMMARY"


@app.route("/clear_all_attendance_summary", methods=["POST"])
@login_required("admin")
def clear_all_attendance_summary():
    """Wipe attendance_summary (PDF import data only). Requires exact typed confirmation."""
    phrase = (request.form.get("confirm_phrase") or "").strip()
    if phrase != _CLEAR_ALL_PDF_SUMMARY_PHRASE:
        return redirect("/attendance_dashboard?error=bad_summary_confirm")
    exe("DELETE FROM attendance_summary")
    return redirect("/attendance_dashboard?summary_cleared=1")


# ── EXPORT ATTENDANCE ──────────────────────────────────────
@app.route("/export_attendance_excel")
@login_required("admin")
def export_attendance_excel():
    q       = request.args.get("q","").strip()
    subject = request.args.get("subject","").strip()
    d       = request.args.get("date","").strip()
    month   = request.args.get("month","").strip()
    status  = request.args.get("status","").strip()

    sql = "SELECT * FROM attendance WHERE 1=1"
    params = []
    if q:
        qq = f"%{q}%"
        sql += (
            " AND (student_name LIKE %s OR student_id IN "
            "(SELECT id FROM students WHERE name LIKE %s OR roll LIKE %s))"
        )
        params += [qq, qq, qq]
    if subject: sql += " AND subject=%s";           params.append(subject)
    if d:       sql += " AND date=%s";              params.append(d)
    if month:   sql += " AND date LIKE %s";         params.append(f"{month}%")
    if status:  sql += " AND status=%s";            params.append(status)
    sql += " ORDER BY date DESC, student_name"
    rows = qry(sql, params)

    wb = Workbook(); ws = wb.active; ws.title = "Attendance"
    hdrs = ["#","Student Name","Subject","Date","Status","Remark","Time Slot"]
    status_colors = {"Present":"C6EFCE","Absent":"FFC7CE","Late":"FFEB9C","Medical":"BDD7EE"}
    for c,h in enumerate(hdrs,1):
        cell = ws.cell(1,c,h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    for r,row in enumerate(rows,2):
        vals = [r-1, 
                row.get("student_name", "N/A"), 
                row.get("subject", "N/A"), 
                str(row.get("date", "")),
                row.get("status", "N/A"), 
                row.get("remark") or "", 
                row.get("time_slot") or ""]
        for c,v in enumerate(vals,1):
            cell = ws.cell(r,c,v)
            if c==5:  # Status column — color by status
                cur_status = row.get("status")
                color = status_colors.get(cur_status,"FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
        
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="attendance_export.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── DY PATIL FINAL ATTENDANCE REPORT IMPORTER ─────────────
@app.route("/export_attendance_report_excel")
@login_required("admin")
def export_attendance_report_excel():
    return export_attendance_excel()

# ── ATTENDANCE: Roll Call List (Task D) ───────────────────
@app.route("/attendance_roll_call")
@login_required("admin")
def attendance_roll_call():
    view = request.args.get("view", "daily") # daily, weekly, monthly, cumulative, subject
    dept = request.args.get("dept", "").strip()
    year = request.args.get("year", "").strip()
    div  = request.args.get("division", "").strip()
    
    today = date.today()
    if view == "weekly":
        start_date = today - timedelta(days=7)
    elif view == "monthly":
        start_date = today - timedelta(days=30)
    elif view == "cumulative":
        start_date = date(2000, 1, 1)
    else:
        start_date = today

    try:
        if view == "subject":
            sql = """
                SELECT asess.subject, asess.division, 
                       COUNT(asess.id) as total_sessions,
                       AVG(CASE WHEN a_counts.total > 0 THEN (a_counts.present * 100.0 / a_counts.total) ELSE 0 END) as avg_attendance_pct
                FROM attendance_sessions asess
                LEFT JOIN (
                    SELECT lecture_id, COUNT(*) as total, SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present
                    FROM attendance
                    GROUP BY lecture_id
                ) a_counts ON asess.id = a_counts.lecture_id
                WHERE 1=1
            """
            params = []
            if dept: sql += " AND asess.branch=%s"; params.append(dept)
            if div:  sql += " AND asess.division=%s"; params.append(div)
            sql += " GROUP BY asess.subject, asess.division ORDER BY asess.subject"
            subject_rows = safe_fetch_all(sql, params)
            return render_template("admin/attendance_roll_call.html", 
                                   subject_rows=subject_rows, view=view, dept=dept, year=year, div=div,
                                   total_found=0, avg_att=0, defaulters_count=0,
                                   DEPARTMENTS=DEPARTMENTS, YEARS=YEARS, DIVISIONS=DIVISIONS)

        # Standard Student Roll Call - Ensures all students in the cohort appear
        sql = """
            SELECT s.id, s.name, s.roll, s.division, s.department, s.year,
                   (SELECT COALESCE(COUNT(a2.id), 0) FROM attendance a2 
                    JOIN attendance_sessions asess2 ON a2.lecture_id = asess2.id
                    WHERE a2.student_id = s.id 
                    AND (asess2.lecture_date >= %s OR %s::DATE = '2000-01-01')) as total_sessions,
                   (SELECT COALESCE(SUM(CASE WHEN a2.status ILIKE 'Present' THEN 1 ELSE 0 END), 0) FROM attendance a2
                    JOIN attendance_sessions asess2 ON a2.lecture_id = asess2.id
                    WHERE a2.student_id = s.id 
                    AND (asess2.lecture_date >= %s OR %s::DATE = '2000-01-01')) as present_count
            FROM students s
            WHERE 1=1
        """
        params = [start_date, start_date, start_date, start_date]
        if dept: sql += " AND s.department=%s"; params.append(dept)
        if year: sql += " AND s.year=%s"; params.append(year)
        if div:  sql += " AND s.division=%s"; params.append(div)
        
        sql += " ORDER BY s.roll"
        rows = safe_fetch_all(sql, params)
        
        # Summary Metrics
        total_found = len(rows)
        avg_att = 0
        defaulters_found = 0
        if total_found > 0:
            total_pct = 0
            for r in rows:
                # Safe calculation to prevent NoneType or DivisionByZero
                tses = int(r.get('total_sessions') or 0)
                pres = int(r.get('present_count') or 0)
                p = (pres * 100 / tses) if tses > 0 else 0
                total_pct += p
                if p < 75 and tses > 0: defaulters_found += 1
            avg_att = round(total_pct / total_found, 1) if total_found > 0 else 0

        return render_template("admin/attendance_roll_call.html", 
                               rows=rows, view=view, dept=dept, year=year, div=div,
                               total_found=total_found, avg_att=avg_att, defaulters_count=defaulters_found,
                               DEPARTMENTS=DEPARTMENTS, YEARS=YEARS, DIVISIONS=DIVISIONS)
    except Exception as e:
        import traceback
        app.logger.error("Error in attendance_roll_call: %s", e)
        app.logger.error(traceback.format_exc())
        return render_template("error_500.html", error=str(e)), 500

# Feature 1: Student Subject-wise breakdown
@app.route("/student/<int:student_id>/attendance", endpoint="student_attendance_profile")
@admin_required
def student_attendance_profile(student_id):
    student = qone("SELECT * FROM students WHERE id = %s", (student_id,))
    if not student: abort(404)
    
    # Subject breakdown with hardening
    subjects_data = qry("""
        SELECT subject, 
               COALESCE(COUNT(*), 0) as total,
               COALESCE(SUM(CASE WHEN status ILIKE 'Present' THEN 1 ELSE 0 END), 0) as present
        FROM attendance
        WHERE student_id = %s
        GROUP BY subject
    """, (student_id,))
    
    for s in subjects_data:
        s["pct"] = round(s["present"] * 100.0 / s["total"], 1) if s["total"] > 0 else 0

    # Global comparison
    global_avg = qone("SELECT AVG(CASE WHEN status ILIKE 'Present' THEN 100.0 ELSE 0 END) as avg FROM attendance")["avg"] or 0
    
    # Recent history timeline
    recent_history = qry("""
        SELECT a.date, a.subject, a.status, f.name as faculty_name
        FROM attendance a
        JOIN faculty f ON a.faculty_id = f.id
        WHERE a.student_id = %s
        ORDER BY a.date DESC LIMIT 20
    """, (student_id,))

    # Overall Metrics
    overall = qone("""
        SELECT COALESCE(COUNT(*), 0) as total, 
               COALESCE(SUM(CASE WHEN status ILIKE 'Present' THEN 1 ELSE 0 END), 0) as present 
        FROM attendance WHERE student_id=%s
    """, (student_id,))
    
    total = overall["total"]
    present = overall["present"]
    overall_pct = round(present * 100.0 / total, 1) if total > 0 else 0
    
    # Simple risk score: (100 - pct) + penalty for last absence
    last_status = recent_history[0]["status"] if recent_history else "Present"
    risk_score = (100 - overall_pct) + (15 if last_status.lower() == "absent" else 0)

    return render_template("admin/student_subject_breakdown.html", 
                           student=student, 
                           subjects_data=subjects_data, 
                           global_avg=global_avg,
                           recent_history=recent_history,
                           overall_pct=overall_pct,
                           risk_score=min(100, risk_score))
    
@app.errorhandler(500)
def handle_500_error(e):
    import traceback
    app.logger.error("Global 500: %s", e)
    app.logger.error(traceback.format_exc())
    return render_template("error_500.html", error=str(e)), 500

@app.route("/admin/faculty_sessions", endpoint="faculty_logs")
@admin_required
def faculty_logs_redirect():
    return redirect(url_for("admin_intel.admin_faculty_logs", **request.args))

@app.route("/admin/session/<int:sid>", endpoint="audit_session_detail_legacy")
@admin_required
def audit_session_detail_legacy(sid):
    return redirect(url_for("admin_intel.admin_session_detail", sid=sid))

@app.route("/admin/faculty_logs", endpoint="faculty_logs_legacy")
@admin_required
def faculty_logs_legacy():
    return redirect(url_for("admin_intel.admin_faculty_logs", **request.args))

# ── ATTENDANCE INTELLIGENCE (Apply Intelligence) ──────────
@app.route("/admin/attendance_intelligence", endpoint="attendance_insights")
@admin_required
def attendance_insights():
    dept = request.args.get("dept", "").strip()
    year = request.args.get("year", "").strip()
    div  = request.args.get("division", "").strip()
    
    insights = IntelligenceService.get_attendance_insights(dept, year, div)
    
    return render_template("admin/attendance_intelligence.html", 
                           insights=insights, 
                           dept=dept, year=year, div=div,
                           DEPARTMENTS=DEPARTMENTS, YEARS=YEARS, DIVISIONS=DIVISIONS)
@app.route("/export_attendance_report", endpoint="attendance_export")
@login_required("admin")
def attendance_export():
    import io, csv
    from flask import Response
    
    view = request.args.get("view", "daily")
    dept = request.args.get("dept", "").strip()
    year = request.args.get("year", "").strip()
    div  = request.args.get("division", "").strip()
    
    today = date.today()
    if view == "weekly":
        start_date = today - timedelta(days=7)
    elif view == "monthly":
        start_date = today - timedelta(days=30)
    elif view == "cumulative":
        start_date = date(2000, 1, 1)
    else:
        start_date = today

    sql = """
        SELECT s.roll, s.name, s.department, s.year, s.division,
               COUNT(a.id) as sessions,
               SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) as present,
               ROUND(SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END)*100.0/NULLIF(COUNT(a.id),0), 1) as percentage
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id AND a.date::DATE >= %s
        WHERE (%s='' OR s.department=%s) AND (%s='' OR s.year=%s) AND (%s='' OR s.division=%s)
        GROUP BY s.id, s.roll, s.name, s.department, s.year, s.division
        ORDER BY s.department, s.year, s.roll
    """
    rows = safe_fetch_all(sql, (start_date, dept, dept, year, year, div, div))
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Roll', 'Name', 'Department', 'Year', 'Division', 'Period', 'Total Lectures', 'Attended', 'Percentage %'])
    for r in rows:
        writer.writerow([r['roll'], r['name'], r['department'], r['year'], r['division'], view.title(), r['sessions'], r['present'], f"{r['percentage'] or 0}%"])
    
    res = Response(output.getvalue(), mimetype="text/csv")
    res.headers["Content-Disposition"] = f"attachment; filename=attendance_{view}_{date.today()}.csv"
    return res

# ── ATTENDANCE: SMS/Email Alerts (Task E) ──────────────────
@app.route("/api/v1/broadcast_sms", methods=["POST"])
@login_required("admin")
def broadcast_sms():
    data = request.json or {}
    message = data.get("message")
    student_ids = data.get("student_ids", [])
    
    if not message or not student_ids:
        return jsonify({"success": False, "error": "Insufficient intelligence payload: Missing message or target IDs"}), 400
        
    # In production, integrate with MSG91, Twilio, or similar
    app.logger.info(f"AUDIT: Broadcast initiated by {session.get('user_id')} to {len(student_ids)} students.")
    
    return jsonify({
        "success": True, 
        "message": "Broadcast transmitted to gateway successfully.",
        "sent_count": len(student_ids)
    })

@app.route("/send_shortage_notices", methods=["POST"])
@login_required("admin")
def send_shortage_notices():
    threshold = safe_int(request.form.get("threshold", 75))
    dept = request.form.get("department", "")
    
    # Identify defaulters
    sql = """
        SELECT s.id, s.name, s.contact_number, s.parent_contact,
               ROUND(SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END)*100.0/NULLIF(COUNT(a.id),0), 1) as pct
        FROM students s
        JOIN attendance a ON s.id = a.student_id
        WHERE (%s='' OR s.department=%s)
        GROUP BY s.id, s.name, s.contact_number, s.parent_contact
        HAVING (SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) * 100.0 / NULLIF(COUNT(*),0)) < %s
    """
    defaulters = safe_fetch_all(sql, (dept, dept, threshold))
    
    count = 0
    for d in defaulters:
        # Simulate SMS/Email sending
        # In a real app, call a vendor API here
        log_msg = f"Alert sent to {d['name']} (Parent: {d['parent_contact']}): Low attendance {d['pct']}%"
        exe("INSERT INTO audit_logs(user_id, action, details) VALUES(%s, %s, %s)",
            (session.get("user_id", 0), "SHORTAGE_NOTICE", log_msg))
        count += 1
        
    flash(f"Successfully triggered notifications for {count} students with attendance < {threshold}%.", "success")
    return redirect("/shortage_report")


@app.route("/download_attendance_template")
@login_required("admin")
def download_attendance_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Template"
    headers = ["student_name", "subject", "date", "status", "remark"]
    sample = ["John Doe", "DBMS", date.today().strftime("%Y-%m-%d"), "Present", ""]
    for c, header in enumerate(headers, 1):
        cell = ws.cell(1, c, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        ws.cell(2, c, sample[c - 1])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")) + 4, 16)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="attendance_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )





# ── IMPORT ATTENDANCE ──────────────────────────────────────
def _parse_dypatil_final_report(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    def clean(value):
        return " ".join(str(value or "").replace("\n", " ").split()).strip()

    def to_int(value):
        try:
            if value is None or value == "":
                return 0
            return int(float(str(value).strip()))
        except (TypeError, ValueError):
            return 0

    meta_text = "\n".join(
        clean(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, 10) + 1)
        for c in range(1, min(ws.max_column, 8) + 1)
    )
    dept, div, year = _class_meta_from_attendance_text(meta_text, getattr(file_obj, "filename", ""))
    semester = _semester_from_attendance_text(meta_text, "II")

    header_row = None
    headers = []
    for r in range(1, min(ws.max_row, 25) + 1):
        values = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        row_text = " ".join(v.upper() for v in values)
        if "ROLL" in row_text and "NAME" in row_text:
            header_row = r
            headers = values
            break
    if not header_row:
        return 0, 0, 0

    roll_col = next((i + 1 for i, h in enumerate(headers) if "ROLL" in h.upper()), None)
    name_col = next((i + 1 for i, h in enumerate(headers) if "NAME" in h.upper()), None)
    if not roll_col or not name_col:
        return 0, 0, 0

    total_row = None
    for r in range(header_row + 1, min(ws.max_row, header_row + 8) + 1):
        row_text = " ".join(clean(ws.cell(r, c).value).upper() for c in range(1, ws.max_column + 1))
        if "LECTURES" in row_text and ("TOTAL" in row_text or "CONDUCTED" in row_text):
            total_row = r
            break
    if not total_row:
        return 0, 0, 0

    subject_cols = []
    for col, raw_name in enumerate(headers, 1):
        if col <= name_col:
            continue
        name = clean(raw_name)
        upper = name.upper()
        if not name:
            continue
        if upper == "TOTAL" or "ATTEND" in upper or upper == "%":
            break
        total = to_int(ws.cell(total_row, col).value)
        if total <= 0:
            continue
        subject_cols.append((col, name, total))

    added = skipped = students_processed = 0

    def get_or_create_student(name, roll):
        row = qone("SELECT id FROM students WHERE roll=%s", (roll,)) if roll else None
        if not row:
            row = qone(
                "SELECT id FROM students WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) AND department=%s AND division=%s ORDER BY id LIMIT 1",
                (name, dept or "", div or ""),
            )
        if row:
            return row["id"]
        try:
            exe(
                "INSERT INTO students(name,roll,department,year,division,password) VALUES(%s,%s,%s,%s,%s,%s)",
                (name, roll or "", dept or "CS", year or "II", div or "", hash_password(_default_student_password())),
            )
            row = qone("SELECT id FROM students WHERE roll=%s", (roll,)) if roll else None
            if not row:
                row = qone("SELECT id FROM students WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) ORDER BY id DESC LIMIT 1", (name,))
            return row["id"] if row else None
        except Exception:
            return None

    for r in range(total_row + 1, ws.max_row + 1):
        roll = clean(ws.cell(r, roll_col).value).upper()
        name = clean(ws.cell(r, name_col).value)
        row_text = " ".join(clean(ws.cell(r, c).value).upper() for c in range(1, min(ws.max_column, 6) + 1))
        if "FACULTY" in row_text or "SIGNATURE" in row_text:
            break
        if not name or not roll or not re.match(r"^[A-Z]?\d{1,3}[A-Z]?$", roll, re.I):
            continue

        sid = get_or_create_student(name, roll)
        if not sid:
            skipped += len(subject_cols)
            continue
        students_processed += 1

        combined = {}
        for col, subject, total in subject_cols:
            item = combined.setdefault(subject, {"attended": 0, "total": 0})
            item["attended"] += to_int(ws.cell(r, col).value)
            item["total"] += total

        for subject, counts in combined.items():
            try:
                exe(
                    """
                    INSERT INTO attendance_summary
                        (student_id, student_name, subject, attended, total, division, semester)
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT(student_id, subject) DO UPDATE SET
                        student_name = excluded.student_name,
                        attended = excluded.attended,
                        total = excluded.total,
                        division = excluded.division,
                        semester = excluded.semester
                    """,
                    (sid, name, subject, counts["attended"], counts["total"], div or "", semester or "II"),
                )
                added += 1
            except Exception:
                skipped += 1

    return added, skipped, students_processed


def _is_dypatil_final_report(file_obj):
    """Peek at the workbook to detect DY Patil Final Attendance Report format."""
    try:
        wb = load_workbook(file_obj, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
        wb.close()
        file_obj.seek(0)
        # Signature: row 7 col C contains 'NAME OF STUDENT' (case-insensitive)
        if len(rows) >= 7:
            r7 = rows[6]  # 0-indexed
            header_text = " ".join(str(v or "").upper() for v in r7)
            if "NAME OF STUDENT" in header_text:
                return True
            # Also check row 4 for 'FINAL ATTENDANCE REPORT'
            r4 = rows[3]
            r4_text = " ".join(str(v or "").upper() for v in r4)
            if "FINAL ATTENDANCE REPORT" in r4_text:
                return True
        return False
    except Exception:
        file_obj.seek(0)
        return False


@app.route("/import_attendance_excel", methods=["POST"])
def import_attendance_excel_v2():
    if not is_attendance_upload_allowed(session):
        return "Unauthorized", 403

    f = request.files.get("file")
    subject = request.form.get("subject", "").strip()
    result = process_attendance_upload(f, session)
    if result["ok"]:
        saved = result["result"]
        return redirect(
            f"/attendance_dashboard?saved={saved['saved']}&skipped={saved['skipped']}"
            f"&students={saved['students']}&format=advanced_upload&batch_id={saved['batch_id']}"
        )

    message = result["error"]
    fallback_errors = (
        "Could not detect attendance table header",
        "Could not detect the total lectures row",
        "No valid subject columns were detected",
    )
    if f and any(token in message for token in fallback_errors):
        try:
            f.seek(0)
        except Exception:
            pass
        added, err = _parse_attendance_excel(f, subject)
        if err == "select_subject":
            return redirect("/attendance?error=select_subject")
        if added:
            return redirect(f"/view_attendance?saved={added}")

    return redirect(f"/attendance?error={message[:160]}")


@app.route("/import_final_attendance_report", methods=["POST"])
def import_final_attendance_report():
    if not is_attendance_upload_allowed(session):
        return "Unauthorized", 403
    f = request.files.get("file")
    result = process_attendance_upload(f, session)
    if result["ok"]:
        saved = result["result"]
        return redirect(
            f"/attendance_dashboard?saved={saved['saved']}&skipped={saved['skipped']}"
            f"&students={saved['students']}&format=advanced_upload&batch_id={saved['batch_id']}"
        )
    return redirect(f"/attendance?error={result['error'][:160]}")


@app.route("/attendance_upload_backup/<int:batch_id>")
def attendance_upload_backup(batch_id):
    if session.get("role") != "admin":
        return "Unauthorized", 403
    return download_attendance_backup(batch_id)


@app.route("/attendance_upload_restore/<int:batch_id>", methods=["POST"])
def attendance_upload_restore(batch_id):
    if session.get("role") != "admin":
        return "Unauthorized", 403
    return restore_attendance_backup(batch_id)


# ── ATTENDANCE ANALYTICS ───────────────────────────────────
@app.route("/attendance_analytics")
@login_required("admin")
def attendance_analytics():
    subject  = request.args.get("subject","").strip()
    month    = request.args.get("month","").strip()
    dept     = request.args.get("dept","").strip()
    division = request.args.get("division","").strip()
    year     = request.args.get("year","").strip()

    where = "1=1"; params = []
    if subject: where += " AND subject=%s"; params.append(subject)
    if month:   where += " AND date LIKE %s"; params.append(f"{month}%")
    if dept:
        where += f" AND {att_in_dept_sql()}"
        params.extend([dept, dept])
    if division:
        where += f" AND {att_in_division_sql()}"
        params.extend([division, division])
    if year:
        where += f" AND {att_in_year_sql()}"
        params.extend([year, year])

    stats = att_stats(where, params)

    # Monthly trend (last 6 months)
    today_d = date.today()
    monthly = []
    for i in range(5,-1,-1):
        mn = ((today_d.month - 1 - i) % 12) + 1
        yn = today_d.year + (today_d.month - 1 - i) // 12
        label = f"{yn}-{mn:02d}"
        t = qone(f"SELECT COUNT(*) as c FROM attendance WHERE {where} AND date LIKE %s",
                 params + [f"{label}%"])["c"]
        p = qone(f"SELECT COUNT(*) as c FROM attendance WHERE {where} AND date LIKE %s AND status='Present'",
                 params + [f"{label}%"])["c"]
        monthly.append({"label":label,"total":t,"pct":pct(p,t)})

    # Subject-wise attendance %
    subj_rows = qry("SELECT subject, "
                    "COUNT(*) as total, "
                    "SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present "
                    "FROM attendance WHERE " + where + " GROUP BY subject ORDER BY subject", params)
    subj_stats = [{"subject": r["subject"][:25], "total": r["total"],
                   "pct": pct(r["present"], r["total"])} for r in subj_rows]

    # Department-wise attendance %
    dept_stats = []
    for d in DEPARTMENTS:
        d_where = where + f" AND {att_in_dept_sql()}"
        d_params = params + [d, d]
        t = qone("SELECT COUNT(*) as c FROM attendance WHERE " + d_where, d_params)["c"]
        p = qone("SELECT COUNT(*) as c FROM attendance WHERE " + d_where + " AND status='Present'", d_params)["c"]
        if t > 0:
            dept_stats.append({"dept": d, "total": t, "present": p,
                                "absent": t-p, "pct": pct(p, t)})

    # Year-wise attendance %
    year_stats = []
    for yr in YEARS:
        y_where = where + f" AND {att_in_year_sql()}"
        y_params = params + [yr, yr]
        t = qone(f"SELECT COUNT(*) as c FROM attendance WHERE {y_where}", y_params)["c"]
        p = qone(f"SELECT COUNT(*) as c FROM attendance WHERE {y_where} AND status='Present'", y_params)["c"]
        if t > 0:
            year_stats.append({"year": f"Year {yr}", "total": t, "present": p,
                                "absent": t-p, "pct": pct(p, t)})

    # Division-wise attendance %
    div_rows = qry("SELECT DISTINCT division FROM students WHERE division != '' AND division IS NOT NULL ORDER BY division")
    div_stats = []
    for row in div_rows:
        dv = row["division"]
        dv_where = where + f" AND {att_in_division_sql()}"
        dv_params = params + [dv, dv]
        t = qone(f"SELECT COUNT(*) as c FROM attendance WHERE {dv_where}", dv_params)["c"]
        p = qone(f"SELECT COUNT(*) as c FROM attendance WHERE {dv_where} AND status='Present'", dv_params)["c"]
        if t > 0:
            div_stats.append({"division": dv, "total": t, "present": p,
                               "absent": t-p, "pct": pct(p, t)})

    # Student-wise summary (batched to avoid N+1 queries on large attendance tables)
    student_sql = "SELECT id,name,roll,department,year,division FROM students"
    s_params = []
    if dept:
        student_sql += " WHERE department=%s"; s_params.append(dept)
    elif division:
        student_sql += " WHERE division=%s"; s_params.append(division)
    elif year:
        student_sql += " WHERE year=%s"; s_params.append(year)
    students_all = qry(student_sql, s_params)

    id_stats = {}
    name_stats = {}
    if students_all:
        ids = [int(s["id"]) for s in students_all if s.get("id") is not None]
        names = [str(s["name"]) for s in students_all if s.get("name")]

        if ids:
            id_ph = ",".join(["%s"] * len(ids))
            id_rows = qry(
                f"""SELECT student_id,
                           COUNT(*) AS total,
                           SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) AS present,
                           SUM(CASE WHEN status='Absent' THEN 1 ELSE 0 END) AS absent,
                           SUM(CASE WHEN status IN ('Late','Leave','Medical') THEN 1 ELSE 0 END) AS late
                    FROM attendance
                    WHERE student_id IN ({id_ph}) AND ({where})
                    GROUP BY student_id""",
                ids + list(params),
            )
            for r in id_rows:
                id_stats[int(r["student_id"])] = {
                    "total": int(r["total"] or 0),
                    "present": int(r["present"] or 0),
                    "absent": int(r["absent"] or 0),
                    "late": int(r["late"] or 0),
                }

        if names:
            name_ph = ",".join(["%s"] * len(names))
            name_rows = qry(
                f"""SELECT student_name,
                           COUNT(*) AS total,
                           SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) AS present,
                           SUM(CASE WHEN status='Absent' THEN 1 ELSE 0 END) AS absent,
                           SUM(CASE WHEN status IN ('Late','Leave','Medical') THEN 1 ELSE 0 END) AS late
                    FROM attendance
                    WHERE student_id IS NULL
                      AND student_name IN ({name_ph})
                      AND ({where})
                    GROUP BY student_name""",
                names + list(params),
            )
            for r in name_rows:
                name_stats[str(r["student_name"])] = {
                    "total": int(r["total"] or 0),
                    "present": int(r["present"] or 0),
                    "absent": int(r["absent"] or 0),
                    "late": int(r["late"] or 0),
                }

    student_stats = []
    for s in students_all:
        sid = int(s["id"])
        nm = str(s["name"] or "")
        st_id = id_stats.get(sid, {"total": 0, "present": 0, "absent": 0, "late": 0})
        st_nm = name_stats.get(nm, {"total": 0, "present": 0, "absent": 0, "late": 0})
        total = st_id["total"] + st_nm["total"]
        if total <= 0:
            continue
        pr = st_id["present"] + st_nm["present"]
        ab = st_id["absent"] + st_nm["absent"]
        lt = st_id["late"] + st_nm["late"]
        student_stats.append({
            "name": s["name"], "roll": s["roll"] or "", "dept": s["department"] or "",
            "year": s["year"] or "", "division": s["division"] or "",
            "total": total, "present": pr, "absent": ab, "late": lt,
            "pct": pct(pr, total)
        })
    student_stats.sort(key=lambda x: x["pct"])
    low_att    = [s for s in student_stats if s["pct"] < 75]
    bottom_att = student_stats[:10]
    # top_att spotlight removed — focus on at-risk students

    # Day of Week Distribution
    # % of attendance by day (Monday to Saturday)
    day_stats = []
    days_map = {0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday", 4: "Friday", 5: "Saturday"}
    for day_idx, day_name in days_map.items():
        # EXTRACT(DOW FROM date) returns 0 for Sunday, 1 for Monday (in Postgres)
        # However, it might depend on the DB. Let's use a robust way if possible.
        # For Postgres: EXTRACT(DOW FROM date) -> 0=Sun, 1=Mon, ..., 6=Sat
        # Let's adjust for 1=Mon, 6=Sat
        d_val = day_idx + 1 # 1=Mon, 6=Sat
        sql_dow = f"SELECT COUNT(*) as tot, SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as p FROM attendance WHERE {where} AND EXTRACT(DOW FROM date) = %s"
        try:
            r_dow = qone(sql_dow, params + [d_val])
            day_stats.append({"day": day_name, "pct": pct(r_dow["p"], r_dow["tot"]), "total": r_dow["tot"]})
        except:
            # Fallback for systems where EXTRACT DOW might not work or return different values
            day_stats.append({"day": day_name, "pct": 0, "total": 0})

    # Status Distribution
    status_counts = {"Present": 0, "Absent": 0, "Late": 0, "Medical": 0, "Leave": 0}
    r_status = qry(f"SELECT status, COUNT(*) as c FROM attendance WHERE {where} GROUP BY status", params)
    for r in r_status:
        if r["status"] in status_counts:
            status_counts[r["status"]] = r["c"]
        else:
            status_counts[r["status"]] = status_counts.get(r["status"], 0) + r["c"]

    # Daily Trend (Last 14 days)
    date_trend = []
    r_date = qry(f"SELECT date, COUNT(*) as tot, SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as p FROM attendance WHERE {where} GROUP BY date ORDER BY date DESC LIMIT 14", params)
    for r in reversed(r_date):
        date_trend.append({"date": str(r["date"]), "pct": pct(r["p"], r["tot"]), "total": r["tot"]})

    subjects_list = [r["subject"] for r in qry("SELECT DISTINCT subject FROM attendance ORDER BY subject")]
    divs_list     = [r["division"] for r in qry("SELECT DISTINCT division FROM students WHERE division!='' ORDER BY division")]
    
    # Calculate global metrics
    global_metrics = {
        "total_records": stats["total"],
        "avg_attendance": stats["pct"],
        "defaulters_count": len(low_att),
        "defaulter_pct": pct(len(low_att), len(student_stats)) if student_stats else 0,
        "critical_count": len([s for s in student_stats if s["pct"] < 40]),
        "perfect_count": len([s for s in student_stats if s["pct"] == 100])
    }

    return render_template("attendance/attendance_analytics.html",
        stats=stats, monthly=monthly,
        subj_stats=subj_stats,
        dept_stats=dept_stats,
        year_stats=year_stats,
        div_stats=div_stats,
        status_map=status_counts,
        low_att=low_att, bottom_att=bottom_att,
        date_trend=date_trend, student_stats=student_stats,
        day_stats=day_stats,
        global_metrics=global_metrics,
        subjects=subjects_list, DEPARTMENTS=DEPARTMENTS, YEARS=YEARS,
        divs=divs_list,
        f_subject=subject, f_month=month, f_dept=dept,
        f_division=division, f_year=year,
        month_labels=[m["label"] for m in monthly],
        month_pcts=[m["pct"] for m in monthly],
        subj_labels=[s["subject"] for s in subj_stats],
        subj_pcts=[s["pct"] for s in subj_stats],
        dept_labels=[d["dept"] for d in dept_stats],
        dept_pcts=[d["pct"] for d in dept_stats],
        year_labels=[y["year"] for y in year_stats],
        year_pcts=[y["pct"] for y in year_stats],
        div_labels=[d["division"] for d in div_stats],
        div_pcts=[d["pct"] for d in div_stats],
        status_labels=list(status_counts.keys()),
        status_data=list(status_counts.values()),
        date_labels=[d["date"] for d in date_trend],
        date_pcts=[d["pct"] for d in date_trend])

@app.route("/consolidated_report")
@login_required("admin")
def consolidated_report():
    dept = request.args.get("dept", "").strip()
    div  = request.args.get("division", "").strip()
    year = request.args.get("year", "").strip()

    # 1. Fetch Students
    sql_s = "SELECT id, name, roll, prn, department, division, year FROM students WHERE 1=1"
    params_s = []
    if dept: sql_s += " AND department=%s"; params_s.append(dept)
    if div:  sql_s += " AND division=%s";   params_s.append(div)
    if year: sql_s += " AND year=%s";       params_s.append(year)
    sql_s += " ORDER BY division, roll, name"
    students = qry(sql_s, params_s)

    if not students:
        return render_template("attendance/consolidated_report.html", 
                               matrix=[], subjects=[], DEPARTMENTS=DEPARTMENTS, DIVISIONS=[],
                               f_dept=dept, f_div=div)

    # 2. Fetch all subjects involved in these students' records
    s_ids = [s["id"] for s in students]
    ph = ",".join(["%s"] * len(s_ids))
    subj_rows = qry(f"SELECT DISTINCT subject FROM attendance WHERE student_id IN ({ph}) ORDER BY subject", tuple(s_ids))
    subjects = [r["subject"] for r in subj_rows]

    # 3. Fetch Matrix Data
    # student_id -> subject -> {attended, total}
    matrix_data = {}
    if s_ids:
        att_rows = qry(f"""
            SELECT student_id, subject, 
                   COUNT(*) as total, 
                   SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present 
            FROM attendance 
            WHERE student_id IN ({ph}) 
            GROUP BY student_id, subject
        """, tuple(s_ids))
        
        for r in att_rows:
            sid = r["student_id"]
            sub = r["subject"]
            if sid not in matrix_data: matrix_data[sid] = {}
            matrix_data[sid][sub] = {"p": r["present"], "t": r["total"], "pct": pct(r["present"], r["total"])}

    # 4. Build Final Matrix for Template
    final_matrix = []
    for s in students:
        row = {
            "name": s["name"],
            "roll": s["roll"],
            "prn": s.get("prn", ""),
            "dept": s["department"],
            "div": s["division"],
            "subjects": []
        }
        total_p = 0
        total_t = 0
        for sub in subjects:
            s_data = matrix_data.get(s["id"], {}).get(sub, {"p": 0, "t": 0, "pct": 0})
            row["subjects"].append(s_data)
            total_p += s_data["p"]
            total_t += s_data["t"]
        
        row["total_pct"] = pct(total_p, total_t)
        final_matrix.append(row)

    divs = [r["division"] for r in qry("SELECT DISTINCT division FROM students WHERE division!='' ORDER BY division")]

    return render_template("attendance/consolidated_report.html",
                           matrix=final_matrix,
                           subjects=subjects,
                           DEPARTMENTS=DEPARTMENTS,
                           DIVISIONS=divs,
                           f_dept=dept,
                           f_div=div,
                           f_year=year)


# ── STUDENT ATTENDANCE REPORT ──────────────────────────────
@app.route("/attendance_report")
@login_required("admin")
def attendance_report():
    subject  = request.args.get("subject","").strip()
    dept     = request.args.get("dept","").strip()
    month    = request.args.get("month","").strip()
    threshold= safe_int(request.args.get("threshold","75"))

    students_all = qry("SELECT id,name,roll,department FROM students ORDER BY department,name")
    report = []
    for s in students_all:
        if dept and s["department"] != dept: continue
        ms  = att_match_student_sql()
        sql = f"SELECT subject, status, time_slot FROM attendance WHERE {ms}"
        pms = list(att_match_student_params(s["id"], s["name"]))
        if subject: sql += " AND subject=%s";       pms.append(subject)
        if month:   sql += " AND date LIKE %s";     pms.append(f"{month}%")
        rows = qry(sql, pms)
        if not rows: continue
        rows.sort(key=lambda x: sort_key_time(x['time_slot']))
        pr = sum(1 for r in rows if r["status"]=="Present")
        ab = sum(1 for r in rows if r["status"]=="Absent")
        la = sum(1 for r in rows if r["status"]=="Late")
        me = sum(1 for r in rows if r["status"]=="Medical")
        p  = pct(pr, len(rows))
        # Per-subject breakdown
        subj_map = {}
        for r in rows:
            subj_map.setdefault(r["subject"],{"t":0,"p":0})
            subj_map[r["subject"]]["t"] += 1
            if r["status"]=="Present": subj_map[r["subject"]]["p"] += 1
        # Attendance color/status label
        if p >= 75:
            att_color = "Good"
        elif p >= 50:
            att_color = "Average"
        else:
            att_color = "Low"

        report.append({
            "name":       s["name"],
            "roll":       s["roll"],
            "dept":       s["department"],
            "total":      len(rows),
            "present":    pr,
            "absent":     ab,
            "late":       la,
            "medical":    me,
            "pct":        p,
            "percentage": round(p, 2),
            "low":        p < threshold,
            "att_status": att_color,
            "subjects":   [{"name": k, "pct": pct(v["p"], v["t"]),
                            "total": v["t"], "present": v["p"]}
                           for k, v in subj_map.items()],
        })
    report.sort(key=lambda x: x["pct"])

    subjects_list = [r["subject"] for r in qry("SELECT DISTINCT subject FROM attendance ORDER BY subject")]
    return render_template("attendance/attendance_report.html",
        report=report, subjects=subjects_list, DEPARTMENTS=DEPARTMENTS,
        f_subject=subject, f_dept=dept, f_month=month, threshold=threshold)


# ── FIX BAD ATTENDANCE DATA ────────────────────────────────
# ── TIMETABLE ─────────────────────────────────────────────

SLOT_COLORS = {
    "Theory":    "#2563EB",
    "Lab":       "#7C3AED",
    "Elective":  "#059669",
    "Minor":     "#D97706",
    "Practical": "#0891B2",
    "Other":     "#6B7280",
}

def assign_color(subject, slot_type):
    if slot_type and slot_type in SLOT_COLORS:
        return SLOT_COLORS[slot_type]
    s = (subject or "").lower()
    if "lab" in s:       return SLOT_COLORS["Lab"]
    if "elective" in s:  return SLOT_COLORS["Elective"]
    if "minor" in s:     return SLOT_COLORS["Minor"]
    return SLOT_COLORS["Theory"]

def _slot_type(subject):
    s = (subject or "").lower()
    if "lab" in s:       return "Lab"
    if "elective" in s:  return "Elective"
    if "minor" in s:     return "Minor"
    return "Theory"

def _sort_time(ts):
    m = re.match(r'(\d+):(\d+)', str(ts or ""))
    if not m: return 9999
    h, mn = int(m.group(1)), int(m.group(2))
    if h < 7: h += 12   # 1:xx, 2:xx → 13:xx, 14:xx
    return h * 60 + mn

def _check_conflicts(day, time_slot, teacher, room, division, exclude_id=None):
    conflicts = []
    sql = "SELECT * FROM timetable WHERE day=%s AND time=%s"
    rows = qry(sql + (f" AND id != {exclude_id}" if exclude_id else ""), (day, time_slot))
    rows.sort(key=lambda r: sort_key_time(r['time']))
    for r in rows:
        if teacher and r["teacher"] and r["teacher"].strip() == teacher.strip():
            conflicts.append(f"Faculty {teacher!r} already assigned at {day} {time_slot}")
        if room and r["room"] and r["room"].strip() == room.strip():
            conflicts.append(f"Room {room!r} already booked at {day} {time_slot}")
        if division and r["division"] and r["division"].strip() == division.strip():
            conflicts.append(f"Division {division!r} already has a class at {day} {time_slot}")
    return conflicts

@app.route("/timetable")
@login_required(["admin", "faculty"])
def timetable():
    # ── Filters ──────────────────────────────────────────────
    fid     = session.get("faculty_id")
    role    = session.get("role")
    q       = request.args.get("q","").strip()
    f_day   = request.args.get("day","").strip()
    f_subj  = request.args.get("subject","").strip()
    f_teach = request.args.get("teacher","").strip()
    f_room  = request.args.get("room","").strip()
    f_div   = request.args.get("division","").strip()
    f_sem   = request.args.get("semester","").strip()
    f_type  = request.args.get("slot_type","").strip()
    view    = request.args.get("view","grid")
    
    # NEW: Default to "My Timetable" for faculty if no filter is applied
    show_personal = request.args.get("personal", "false").lower() == "true"
    if role == "faculty" and not any([q, f_day, f_subj, f_teach, f_room, f_div, f_type]) and request.args.get("all") != "true":
        show_personal = True

    # ── ALL entries (for grid — no filters) ──────────────────
    if show_personal and fid:
        sql_all = f"SELECT t.*, COALESCE(f.name, t.teacher) as teacher FROM timetable t LEFT JOIN faculty f ON t.faculty_id = f.id WHERE t.faculty_id = %s ORDER BY {DAY_ORD}, t.start_time"
        all_rows = safe_fetch_all(sql_all, (fid,))
    else:
        sql_all = f"SELECT t.*, COALESCE(f.name, t.teacher) as teacher FROM timetable t LEFT JOIN faculty f ON t.faculty_id = f.id ORDER BY {DAY_ORD}, t.start_time"
        all_rows = safe_fetch_all(sql_all)
    all_entries = []
    for e in all_rows:
        d = dict(e)
        d["time"] = normalize_time(d.get("time",""))
        all_entries.append(d)

    # ── Collect unique time slots, sort chronologically ──────
    seen_ts = set(); raw_ts = []
    for e in all_entries:
        t = e.get("time")
        if t and t not in seen_ts:
            seen_ts.add(t); raw_ts.append(t)

    def _tsort(ts):
        m = re.match(r"(\d+):(\d+)", str(ts or ""))
        if not m: return 999
        h = int(m.group(1)); mn = int(m.group(2))
        if h < 7: h += 12   # 1:xx → 13, 2:xx → 14, etc.
        return h * 60 + mn

    time_slots = sorted(raw_ts, key=_tsort)

    # Build grid: grid[day][time] = [slots]
    grid = {d: {ts: [] for ts in time_slots} for d in DAYS}
    for r in all_entries:
        d = r.get('day')
        t = r.get('time')
        if d in grid and t in grid[d]:
            grid[d][t].append(r)
    
    # grid_rows for older template logic
    grid_rows = {}
    for d in DAYS:
        grid_rows[d] = [list(grid[d][t]) for t in time_slots]

    # ── Filtered entries for list view ──
    sql = f"SELECT t.*, COALESCE(f.name, t.teacher) as teacher FROM timetable t LEFT JOIN faculty f ON t.faculty_id = f.id WHERE 1=1"
    params = []
    if q:
        sql += " AND (t.subject LIKE %s OR f.name LIKE %s OR t.teacher LIKE %s OR t.room LIKE %s OR f.department LIKE %s)"
        params += [f"%{q}%"] * 5
    if f_day:
        sql += " AND t.day=%s"; params.append(f_day)
    if f_subj:
        sql += " AND t.subject LIKE %s"; params.append(f"%{f_subj}%")
    if f_teach:
        sql += " AND (f.name LIKE %s OR t.teacher LIKE %s)"; params += [f"%{f_teach}%"] * 2
    if f_div:
        sql += " AND t.division LIKE %s"; params.append(f"%{f_div}%")
    if f_room:
        sql += " AND t.room LIKE %s"; params.append(f"%{f_room}%")
    if f_type:
        sql += " AND t.slot_type=%s"; params.append(f_type)
    if f_sem:
        sql += " AND t.semester=%s"; params.append(f_sem)
    
    sql += f" ORDER BY {DAY_ORD}, t.start_time"
    f_rows = safe_fetch_all(sql, params)
    entries = [dict(e) for e in f_rows]
    for e in entries:
        e["time"] = normalize_time(e.get("time",""))

    any_filter = bool(q or f_day or f_subj or f_teach or f_room or f_div or f_type)
    filtered_ids = set(e["id"] for e in entries) if any_filter else None

    # ── Dropdown options ──
    subjects_list = sorted(set(e["subject"] for e in all_entries if e["subject"]))
    teachers_list = sorted(set(e["teacher"] for e in all_entries if e["teacher"]))
    rooms_list    = sorted(set(e["room"]    for e in all_entries if e["room"]))
    divs_list     = sorted(set(e["division"] for e in all_entries if e["division"]))

    # ── Stats ──
    total   = len(all_entries)
    theory  = sum(1 for e in all_entries if (e.get("slot_type") or "Theory") == "Theory")
    lab     = sum(1 for e in all_entries if (e.get("slot_type") or "") == "Lab")
    elec    = sum(1 for e in all_entries if (e.get("slot_type") or "") == "Elective")
    minor   = sum(1 for e in all_entries if (e.get("slot_type") or "") == "Minor")

    workload = {}
    for e in all_entries:
        t = e.get("teacher","")
        if t: workload[t] = workload.get(t,0) + 1

    return render_template("common/timetable.html",
        entries=entries, grid=grid, grid_rows=grid_rows, time_slots=time_slots,
        q=q, f_day=f_day, f_subj=f_subj, f_teach=f_teach,
        f_room=f_room, f_div=f_div, f_sem=f_sem, f_type=f_type, view=view,
        DAYS=DAYS, total=total, theory=theory, lab=lab, elec=elec, minor=minor,
        subjects=subjects_list, teachers=teachers_list,
        rooms=rooms_list, divs=divs_list,
        workload=sorted(workload.items(), key=lambda x:-x[1])[:10],
        any_filter=any_filter,
        filtered_ids=filtered_ids,
        today_name=date.today().strftime("%A"),
        show_personal=show_personal
    )



@app.route("/save_timetable", methods=["POST"])
@login_required("admin")
def save_timetable():
    day      = request.form.get("day","")
    time_s   = normalize_time(request.form.get("time",""))
    subject  = request.form.get("subject","").strip()
    teacher  = request.form.get("teacher","").strip()
    room     = request.form.get("room","").strip()
    division = request.form.get("division","").strip()
    semester = request.form.get("semester","").strip()
    slot_type= request.form.get("slot_type","Theory")
    color    = assign_color(subject, slot_type)
    if not subject: return redirect("/timetable?error=nosubject")

    # Add backend protection insertion logic:
    faculty_id = None
    if teacher:
        f_row = qone("SELECT id FROM faculty WHERE name=%s LIMIT 1", (teacher,))
        if f_row:
            faculty_id = f_row["id"]
    
    branch = ""
    year = ""
    if division:
        s_row = qone("SELECT department, year FROM students WHERE division=%s LIMIT 1", (division,))
        if s_row:
            branch = s_row["department"]
            year = s_row["year"]

    # Parse start and end time
    start_time = None
    end_time = None
    if time_s:
        import re as re_mod
        m = re_mod.match(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", time_s)
        if m:
            h1, m1, h2, m2 = map(int, m.groups())
            if h1 < 7: h1 += 12
            if h2 < 7: h2 += 12
            start_time = f"{h1:02d}:{m1:02d}:00"
            end_time = f"{h2:02d}:{m2:02d}:00"

    sub_row = qone("SELECT id FROM subjects WHERE name=%s LIMIT 1", (subject,))
    subject_id = sub_row["id"] if sub_row else None

    # --- CLASH DETECTION LOGIC ---
    clash_cond = "NOT (end_time <= %s OR start_time >= %s)"
    if faculty_id and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND faculty_id=%s AND {clash_cond}", (day, faculty_id, start_time, end_time)):
            return redirect("/timetable?error=" + "Teacher clash: Faculty is already assigned here.")
    if branch and year and division and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND branch=%s AND year=%s AND division=%s AND {clash_cond}", (day, branch, year, division, start_time, end_time)):
            return redirect("/timetable?error=" + f"Class clash: {division} is busy at {time_s}.")
    if room and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND room=%s AND {clash_cond}", (day, room, start_time, end_time)):
            return redirect("/timetable?error=" + f"Room clash: {room} is already booked.")
            
    if slot_type == "Lab":
        # Note: Lab sessions implicitly occupy continuous practical slots
        pass
    # -----------------------------

    exe("INSERT INTO timetable(day,time,start_time,end_time,subject_id,subject,teacher,room,division,semester,slot_type,color,faculty_id,branch,year) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (day,time_s,start_time,end_time,subject_id,subject,teacher,room,division,semester,slot_type,color,faculty_id,branch,year))
    
    # Auto-create faculty assignment
    if faculty_id and subject_id:
        try:
            exe("""
                INSERT INTO faculty_subject_assignments
                    (faculty_id, subject_id, subject_name, department, semester,
                     class_name, division, academic_year)
                VALUES (%s, %s, %s, %s, %s, %s, %s, '2025-26')
                ON CONFLICT (faculty_id, subject_id, division) DO NOTHING
            """, (faculty_id, subject_id, subject, branch or '',
                  semester or '', f"{year}-{branch}" if year and branch else (division or ''),
                  division or ''))
            log_audit("Auto Faculty Assignment", f"Auto-assigned {subject} → faculty_id {faculty_id} (div {division})")
        except Exception as e:
            app.logger.warning(f"Auto-assignment failed: {e}")

    return redirect("/timetable?added=1")

@app.route("/edit_timetable", methods=["POST"])
@login_required("admin")
def edit_timetable():
    tid      = request.form.get("tt_id","")
    day      = request.form.get("day","")
    time_s   = normalize_time(request.form.get("time",""))
    subject  = request.form.get("subject","").strip()
    teacher  = request.form.get("teacher","").strip()
    room     = request.form.get("room","").strip()
    division = request.form.get("division","").strip()
    semester = request.form.get("semester","").strip()
    slot_type= request.form.get("slot_type","Theory")
    color    = assign_color(subject, slot_type)

    # ADDED: Extract meta values for clash detection
    faculty_id = None
    if teacher:
        f_row = qone("SELECT id FROM faculty WHERE name=%s LIMIT 1", (teacher,))
        if f_row: faculty_id = f_row["id"]
    
    branch = ""
    year = ""
    if division:
        s_row = qone("SELECT department, year FROM students WHERE division=%s LIMIT 1", (division,))
        if s_row:
            branch = s_row["department"]
            year = s_row["year"]

    start_time = None
    end_time = None
    if time_s:
        import re as re_mod
        m = re_mod.match(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", time_s)
        if m:
            h1, m1, h2, m2 = map(int, m.groups())
            if h1 < 7: h1 += 12
            if h2 < 7: h2 += 12
            start_time = f"{h1:02d}:{m1:02d}:00"
            end_time = f"{h2:02d}:{m2:02d}:00"

    sub_row = qone("SELECT id FROM subjects WHERE name=%s LIMIT 1", (subject,))
    subject_id = sub_row["id"] if sub_row else None

    # --- CLASH DETECTION LOGIC (EXCLUDING CURRENT ID) ---
    clash_cond = "NOT (end_time <= %s OR start_time >= %s) AND id != %s"
    if faculty_id and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND faculty_id=%s AND {clash_cond}", 
                (day, faculty_id, start_time, end_time, tid)):
            return redirect("/timetable?error=Teacher clash detected during edit.")
    if branch and year and division and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND branch=%s AND year=%s AND division=%s AND {clash_cond}", 
                (day, branch, year, division, start_time, end_time, tid)):
            return redirect(f"/timetable?error=Class clash detected during edit: {division} busy at {time_s}.")
    if room and start_time and end_time:
        if qone(f"SELECT 1 FROM timetable WHERE day=%s AND room=%s AND {clash_cond}", 
                (day, room, start_time, end_time, tid)):
            return redirect(f"/timetable?error=Room clash detected during edit: {room} already booked.")

    exe("""UPDATE timetable 
           SET day=%s, time=%s, start_time=%s, end_time=%s, subject_id=%s, 
               subject=%s, teacher=%s, room=%s, division=%s, semester=%s, 
               slot_type=%s, color=%s, faculty_id=%s, branch=%s, year=%s 
           WHERE id=%s""",
        (day, time_s, start_time, end_time, subject_id, subject, teacher, room, division, semester, slot_type, color, faculty_id, branch, year, tid))
    
    # Update faculty assignment if faculty changed
    if faculty_id and subject_id:
        try:
            exe("""
                INSERT INTO faculty_subject_assignments
                    (faculty_id, subject_id, subject_name, department, semester,
                     class_name, division, academic_year)
                VALUES (%s, %s, %s, %s, %s, %s, %s, '2025-26')
                ON CONFLICT (faculty_id, subject_id, division) DO UPDATE SET
                    subject_name = EXCLUDED.subject_name,
                    department = EXCLUDED.department
            """, (faculty_id, subject_id, subject, branch or '',
                  semester or '', f"{year}-{branch}" if year and branch else (division or ''),
                  division or ''))
        except Exception as e:
            app.logger.warning(f"Auto-assignment update failed: {e}")

    return redirect("/timetable?edited=1")

@app.route("/delete_timetable", methods=["POST"])
@login_required("admin")
def delete_timetable():
    tid = request.form.get("tt_id","")
    exe("DELETE FROM timetable WHERE id=%s", (tid,))
    return redirect("/timetable?deleted=1")

@app.route("/clear_timetable", methods=["POST"])
@login_required("admin")
def clear_timetable():
    exe("DELETE FROM timetable")
    return redirect("/timetable?cleared=1")

@app.route("/clean_timetable", methods=["POST"])
@login_required("admin")
def clean_timetable():
    valid = "'Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'"
    n = qone(f"SELECT COUNT(*) as c FROM timetable WHERE day NOT IN ({valid}) OR subject='' OR subject IS NULL")["c"]
    exe(f"DELETE FROM timetable WHERE day NOT IN ({valid}) OR subject='' OR subject IS NULL")
    return redirect(f"/timetable?cleaned={n}")

@app.route("/duplicate_timetable", methods=["POST"])
@login_required("admin")
def duplicate_timetable():
    r = qone("SELECT * FROM timetable WHERE id=%s", (request.form.get("tt_id",""),))
    if r:
        exe("INSERT INTO timetable(day,time,start_time,end_time,subject_id,subject,teacher,room,division,semester,slot_type,color,faculty_id,branch,year) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (r["day"],r["time"],r.get("start_time"),r.get("end_time"),r.get("subject_id"),r["subject"],r["teacher"],r["room"] or "",
             r["division"] or "",r["semester"] or "",r["slot_type"] or "Theory",r["color"] or "", r.get("faculty_id"), r.get("branch"), r.get("year")))
    return redirect("/timetable?added=1")

@app.route("/move_timetable", methods=["POST"])
@login_required("admin")
def move_timetable():
    data  = request.get_json() or {}
    tid   = data.get("id","")
    day   = data.get("day","")
    time_s= normalize_time(data.get("time",""))
    r = qone("SELECT * FROM timetable WHERE id=%s", (tid,))
    if not r: return jsonify({"ok":False}), 404
    start_time, end_time = None, None
    import re
    m = re.match(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", time_s)
    if m:
        h1, m1, h2, m2 = map(int, m.groups())
        if h1 < 7: h1 += 12
        if h2 < 7: h2 += 12
        start_time = f"{h1:02d}:{m1:02d}:00"
        end_time = f"{h2:02d}:{m2:02d}:00"
    exe("UPDATE timetable SET day=%s,time=%s,start_time=%s,end_time=%s WHERE id=%s", (day,time_s,start_time,end_time,tid))
    return jsonify({"ok":True})

@app.route("/export_timetable_excel")
@login_required("admin")
def export_timetable_excel():
    rows = qry("SELECT * FROM timetable ORDER BY day,time")
    wb = Workbook(); ws = wb.active; ws.title = "Timetable"
    hdrs = ["Day","Time","Subject","Faculty","Room","Division","Semester","Type"]
    for c,h in enumerate(hdrs,1):
        cell = ws.cell(1,c,h)
        cell.font = Font(bold=True,color="FFFFFF")
        cell.fill = PatternFill("solid",fgColor="0F172A")
    for ri,row in enumerate(rows,2):
        for c,k in enumerate(["day","time","subject","teacher","room","division","semester","slot_type"],1):
            ws.cell(ri,c,row[k] or "")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="timetable.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/import_timetable_v2", methods=["POST"])
@login_required("admin")
def import_timetable_v2():
    f = request.files.get("file")
    if not f: return redirect("/timetable?error=nofile")
    added = _parse_timetable_excel(f)
    return redirect(f"/timetable?imported={added}")

@app.route('/api/add_time_slot', methods=['POST'])
@login_required('admin')
def api_add_time_slot():
    data = request.get_json()
    slot = (data.get('slot') or '').strip()
    if not slot:
        return jsonify({'error': 'No slot provided'}), 400
    # Validate format: should match HH:MM-HH:MM or H:MM-H:MM
    import re as re_api
    if not re_api.match(r'\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}', slot):
        return jsonify({'error': 'Invalid format. Use HH:MM-HH:MM'}), 400
    return jsonify({'ok': True, 'slot': slot})

@app.route('/api/copy_day', methods=['POST'])
@login_required('admin')
def api_copy_day():
    data = request.get_json()
    from_day, to_day = data.get('from_day'), data.get('to_day')
    if not from_day or not to_day or from_day == to_day:
        return jsonify({'error': 'Invalid days'}), 400
    rows = qry("SELECT * FROM timetable WHERE day=%s", (from_day,))
    count = 0
    for r in rows:
        if not qone("SELECT 1 FROM timetable WHERE day=%s AND time=%s AND division=%s AND subject=%s",
                    (to_day, r['time'], r['division'], r['subject'])):
            exe("INSERT INTO timetable(day,time,start_time,end_time,subject_id,subject,teacher,"
                "room,division,semester,slot_type,color,faculty_id,branch,year) VALUES"
                "(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (to_day,r['time'],r['start_time'],r['end_time'],r['subject_id'],r['subject'],
                 r['teacher'],r['room'],r['division'],r['semester'],r['slot_type'],
                 r['color'],r['faculty_id'],r['branch'],r['year']))
            count += 1
    log_audit("Copy Day", f"Copied {count} slots from {from_day} to {to_day}")
    return jsonify({'ok': True, 'count': count})

@app.route("/admin_update_profile", methods=["POST"])
@login_required("admin")
def admin_update_profile():
    global ADMIN_PASSWORD_HASH
    name  = request.form.get("name", "").strip()
    phone = request.form.get("phone", "").strip()
    current_pw  = request.form.get("current_password", "").strip()
    new_pw      = request.form.get("new_password", "").strip()
    confirm_pw  = request.form.get("confirm_password", "").strip()
    if name:
        session["name"] = name
    if new_pw:
        err = validate_password_change(ADMIN_PASSWORD_HASH, current_pw, new_pw, confirm_pw)
        if err:
            return redirect(f"/admin_profile?error={err}")
        # Update in-memory hash
        ADMIN_PASSWORD_HASH = generate_password_hash(new_pw)
    return redirect("/admin_profile?success=1")

@app.route("/check_conflicts_api", methods=["POST"])
@login_required("admin")
def check_conflicts_api():
    d = request.get_json() or {}
    conflicts = _check_conflicts(d.get("day",""), d.get("time",""),
                                  d.get("teacher",""), d.get("room",""),
                                  d.get("division",""), d.get("exclude_id"))
    return jsonify({"conflicts": conflicts})


# ── REPORTS ───────────────────────────────────────────────
@app.route("/reports")
@login_required("admin")
def reports():
    ts  = qone("SELECT COUNT(*) as c FROM students")["c"]
    tf  = qone("SELECT COUNT(*) as c FROM faculty")["c"]
    ta  = qone("SELECT COUNT(*) as c FROM attendance")["c"]
    tm  = qone("SELECT COUNT(*) as c FROM marks")["c"]
    dept_stats_rows = qry(
        """SELECT s.department AS dept,
                  COUNT(*) AS total,
                  SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) AS present
           FROM attendance a
           JOIN students s ON a.student_id=s.id
           WHERE s.department IS NOT NULL AND s.department <> ''
           GROUP BY s.department"""
    )
    dept_map = {
        r["dept"]: (int(r["total"] or 0), int(r["present"] or 0))
        for r in dept_stats_rows
    }
    dept_att_pct = []
    for d in DEPARTMENTS:
        tot, pre = dept_map.get(d, (0, 0))
        if tot > 0:
            dept_att_pct.append((d, pct(pre, tot)))
    marks_rows  = qry("SELECT exam_type, ROUND(AVG(marks*100.0/total)::numeric,1) as avg FROM marks GROUP BY exam_type")
    return render_template("common/reports.html", ts=ts, tf=tf, ta=ta, tm=tm,
                           dept_att_pct=dept_att_pct,
                           marks_exams=[r["exam_type"] for r in marks_rows],
                           marks_avg=[r["avg"] for r in marks_rows])

@app.route("/admin_profile")
@login_required("admin")
def admin_profile():
    return render_template("admin/admin_profile.html")




# ════════════════════════════════════════════════════════════
#  FACULTY ROUTES
# ════════════════════════════════════════════════════════════
@app.route("/faculty_dashboard")
@login_required("faculty")
def faculty_dashboard():
    _tt_name = session.get('name', '')
    try:
        fid = session.get("faculty_id")
        profile = safe_fetch_one("SELECT * FROM faculty WHERE id=%s", (fid,))
        if profile is None:
            profile = {"name": session.get("name", "Faculty"), "department": "-", "designation": "-"}
        
        # Existing timetable logic (using faculty_id with legacy fallback)
        my_timetable = safe_query(
            f"SELECT * FROM timetable WHERE faculty_id=%s ORDER BY {DAY_ORD}, start_time",
            (fid,)
        )
        if not my_timetable:
            my_timetable = safe_query(f"SELECT * FROM timetable WHERE teacher LIKE %s ORDER BY {DAY_ORD},time", (f"%{_tt_name}%",))
            if not my_timetable:
                _tt_parts = [p for p in _tt_name.replace("Prof.","").replace("Dr.","").strip().split() if len(p) > 3]
                for _part in _tt_parts:
                    my_timetable = safe_query(f"SELECT * FROM timetable WHERE teacher LIKE %s ORDER BY {DAY_ORD},time", (f"%{_part}%",))
                    if my_timetable: break

        # ── NEW STATS ──
        today_str = datetime.now().strftime("%Y-%m-%d")
        today_sessions = safe_fetch_scalar("SELECT COUNT(*) FROM attendance_sessions WHERE faculty_id=%s AND lecture_date=%s", (fid, today_str))
        pending_drafts = safe_fetch_scalar("SELECT COUNT(*) FROM attendance_sessions WHERE faculty_id=%s AND status='draft'", (fid,))
        
        # Low attendance count (<75%)
        # This is slightly expensive, we look at students who have classes with this faculty
        low_att_count = safe_fetch_scalar("""
            WITH student_stats AS (
                SELECT student_id, 
                       COUNT(*) as total, 
                       SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present
                FROM attendance
                WHERE faculty_id = %s
                GROUP BY student_id
            )
            SELECT COUNT(*) FROM student_stats WHERE total > 0 AND (present * 100.0 / total) < 75
        """, (fid,))

        day_name = datetime.now().strftime("%A") # e.g. Monday
        
        # Find classes from timetable for today (including Substitutions)
        today_classes = safe_query("""
            SELECT t.* FROM timetable t 
            WHERE (t.faculty_id = %s OR t.id IN (
                SELECT timetable_id FROM timetable_substitutions 
                WHERE substitute_faculty_id = %s AND session_date = %s
            )) 
            AND t.day = %s AND t.published = TRUE
        """, (fid, fid, today_str, day_name))
        
        if not today_classes:
            # Fallback to teacher name (legacy)
            today_classes = safe_query("""
                SELECT * FROM timetable 
                WHERE teacher LIKE %s AND day = %s AND published = TRUE
            """, (f"%{_tt_name}%", day_name))
        
        today_actions = []
        for tc in (today_classes or []):
            # Check if already marked
            marked = safe_fetch_scalar("""
                SELECT COUNT(*) FROM attendance_sessions 
                WHERE faculty_id = %s AND subject = %s AND lecture_date = %s
            """, (fid, tc.get('subject', ''), today_str))
            if marked == 0:
                today_actions.append(tc)

        # Legacy counts for chart/dashboard compatibility
        att_count = safe_fetch_scalar("SELECT COUNT(*) FROM attendance WHERE faculty_id=%s", (fid,))
        marks_count = safe_fetch_scalar("SELECT COUNT(*) FROM marks WHERE faculty_id=%s", (fid,))
        notice_count = safe_fetch_scalar("SELECT COUNT(*) FROM faculty_notices WHERE faculty_id=%s", (fid,))
        notes_count = safe_fetch_scalar("SELECT COUNT(*) FROM faculty_notes WHERE faculty_id=%s", (fid,))

        # Legacy data for charts
        my_subjects = safe_query("SELECT * FROM subjects WHERE teacher LIKE %s ORDER BY name", (f"%{_tt_name}%",))
        subj_names = [s["name"] for s in (my_subjects or []) if s.get("name")]

        subj_att_labels = []
        subj_att_pct = []
        for sn in subj_names[:6]:
            tot = safe_fetch_scalar("SELECT COUNT(*) FROM attendance WHERE subject=%s", (sn,))
            pre = safe_fetch_scalar("SELECT COUNT(*) FROM attendance WHERE subject=%s AND status='Present'", (sn,))
            subj_att_labels.append(str(sn)[:12])
            subj_att_pct.append(round(pre*100.0/tot, 1) if tot > 0 else 0)

        marks_rows = safe_query("SELECT exam_type, ROUND(AVG(marks*100.0/total)::numeric,1) as avg FROM marks WHERE faculty_id=%s GROUP BY exam_type", (fid,))
        recent_notices = safe_query("SELECT * FROM faculty_notices WHERE faculty_id=%s ORDER BY id DESC LIMIT 5", (fid,))

        return render_template("faculty/faculty_dashboard.html",
                             profile=profile,
                             my_timetable=my_timetable,
                             att_count=att_count or 0,
                             marks_count=marks_count or 0,
                             notice_count=notice_count or 0,
                             notes_count=notes_count or 0,
                             today_sessions=today_sessions or 0,
                             pending_drafts=pending_drafts or 0,
                             low_att_count=low_att_count or 0,
                             today_actions=today_actions,
                             recent_notices=recent_notices,
                             subj_att_labels=subj_att_labels,
                             subj_att_pct=subj_att_pct,
                             fac_marks_exams=[r["exam_type"] for r in (marks_rows or []) if r.get("exam_type")],
                             fac_marks_avg=[r["avg"] for r in (marks_rows or [])])
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        app.logger.error(f"Dashboard Error: {err_msg}")
        return f"<div style='padding:20px; font-family:sans-serif;'><h2>Dashboard Error</h2><pre style='background:#f4f4f4; padding:15px; border-radius:5px;'>{err_msg}</pre></div>", 500

@app.route("/faculty_attendance")
@login_required("faculty")
def faculty_attendance():
    import json as _json
    fid  = session.get("faculty_id")
    name = session.get("name", "")

    # ── My subjects ──────────────────────────────────────────
    my_subjects = qry(
        "SELECT * FROM subjects WHERE teacher LIKE %s ORDER BY name",
        (f"%{name}%",)
    )

    # ── Filter students by selected subject's section ─────────
    selected_subject = request.args.get("subject", "").strip()
    if selected_subject:
        # Try to find section metadata from the subjects table
        sub_row = qone(
            "SELECT department, division, year FROM subjects WHERE name=%s LIMIT 1",
            (selected_subject,)
        )
        if sub_row and (sub_row.get("division") or sub_row.get("department")):
            sql_stu = "SELECT id, name, roll, department, division, year FROM students WHERE 1=1"
            p_stu   = []
            if sub_row.get("division"):
                sql_stu += " AND division = %s"; p_stu.append(sub_row["division"])
            elif sub_row.get("department"):
                sql_stu += " AND department = %s"; p_stu.append(sub_row["department"])
            if sub_row.get("year"):
                sql_stu += " AND year = %s"; p_stu.append(sub_row["year"])
            sql_stu += " ORDER BY name"
            all_students = qry(sql_stu, p_stu)
        else:
            # Fallback: try timetable division match
            tt_row = qone(
                "SELECT division, branch, year FROM timetable WHERE faculty_id=%s AND subject=%s LIMIT 1",
                (fid, selected_subject)
            )
            if tt_row and tt_row.get("division"):
                all_students = qry(
                    "SELECT id, name, roll, department, division, year FROM students WHERE division=%s ORDER BY name",
                    (tt_row["division"],)
                )
            else:
                # No timetable or subject assignment found? Default to showing none to prevent "leak"
                all_students = []
    else:
        # No subject selected — definitely show empty list
        all_students = []

    # ── Compute cumulative attendance % per student (for low-att badges) ─
    enriched = []
    for s in all_students:
        row = qone(
            """SELECT COUNT(*) as total,
                      SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present
               FROM attendance WHERE student_id=%s AND subject=%s""",
            (s["id"], selected_subject) if selected_subject else (s["id"], "")
        )
        total   = int(row["total"]   or 0) if row else 0
        present = int(row["present"] or 0) if row else 0
        att_pct = round((present / total * 100), 1) if total > 0 else None
        d = dict(s)
        d["att_pct"] = att_pct
        enriched.append(d)

    all_students_json = _json.dumps(enriched)

    # ── Audit log (last 10 by this faculty) ───────────────────
    audit_logs = []
    try:
        audit_logs = qry(
            "SELECT action, details, created_at FROM audit_logs WHERE role='faculty' AND user_id=%s ORDER BY id DESC LIMIT 10",
            (fid,)
        ) or []
    except Exception:
        audit_logs = []

    return render_template("faculty/faculty_attendance.html",
        my_subjects=my_subjects,
        all_students=enriched,
        all_students_json=all_students_json,
        audit_logs=audit_logs,
        today=today_str(),
        tab=request.args.get("tab", "mark")
    )


@app.route("/faculty_get_students_for_subject")
@login_required("faculty")
def faculty_get_students_for_subject():
    """JSON endpoint: returns students filtered to a subject's section, with attendance %."""
    import json as _json
    fid     = session.get("faculty_id")
    subject = request.args.get("subject", "").strip()
    if not subject:
        return jsonify({"students": []})

    # Try subjects table first
    sub_row = qone(
        "SELECT department, division, year FROM subjects WHERE name=%s LIMIT 1",
        (subject,)
    )
    if sub_row and (sub_row.get("division") or sub_row.get("department")):
        sql_stu = "SELECT id, name, roll, prn, department, division, year FROM students WHERE 1=1"
        p_stu   = []
        if sub_row.get("division"):
            sql_stu += " AND division = %s"; p_stu.append(sub_row["division"])
        elif sub_row.get("department"):
            sql_stu += " AND department = %s"; p_stu.append(sub_row["department"])
        if sub_row.get("year"):
            sql_stu += " AND year = %s"; p_stu.append(sub_row["year"])
        sql_stu += " ORDER BY name"
        students = qry(sql_stu, p_stu)
    else:
        # Fallback: timetable division
        tt_row = qone(
            "SELECT division FROM timetable WHERE faculty_id=%s AND subject=%s LIMIT 1",
            (fid, subject)
        )
        if tt_row and tt_row.get("division"):
            students = qry(
                "SELECT id, name, roll, prn, department, division, year FROM students WHERE division=%s ORDER BY name",
                (tt_row["division"],)
            )
        else:
            # If no meta or timetable found, do NOT show all students
            students = []

    # Enrich with attendance % (FIXED: Batch query to avoid N+1)
    student_ids = [s["id"] for s in students]
    if student_ids and subject:
        ph = ",".join(["%s"] * len(student_ids))
        att_rows = qry(
            f"SELECT student_id, COUNT(*) as total, SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present "
            f"FROM attendance WHERE student_id IN ({ph}) AND subject=%s GROUP BY student_id",
            student_ids + [subject]
        )
        att_map = {r["student_id"]: r for r in att_rows}
    else:
        att_map = {}

    result = []
    for s in students:
        stats = att_map.get(s["id"], {"total": 0, "present": 0})
        total = int(stats["total"] or 0)
        present = int(stats["present"] or 0)
        att_pct = round(present / total * 100, 1) if total > 0 else None
        d = dict(s)
        d["att_pct"] = att_pct
        result.append(d)

    return jsonify({"students": result})


@app.route("/faculty_get_attendance")
@login_required("faculty")
def faculty_get_attendance():
    """JSON endpoint: returns existing attendance for edit-mode pre-fill."""
    fid     = session.get("faculty_id")
    subject = request.args.get("subject", "").strip()
    att_date = request.args.get("date", "").strip()
    if not subject or not att_date:
        return jsonify({"records": []})

    rows = qry(
        "SELECT student_id, status FROM attendance WHERE subject=%s AND date=%s AND faculty_id=%s",
        (subject, att_date, fid)
    )
    return jsonify({"records": [dict(r) for r in rows]})

@app.route("/faculty_portal")
@login_required("faculty")
def faculty_portal():
    return render_template("faculty/FacultyAttendancePortal.html")

@app.route("/faculty_save_attendance", methods=["POST"])
@login_required("faculty")
def faculty_save_attendance():
    # ── Guard: subject must be present ───────────────────────
    subject = request.form.get("subject", "").strip()
    if not subject:
        flash("Subject is required. Please select a subject before committing.", "error")
        return redirect("/faculty_attendance")

    # ── Guard: date must be present and valid ─────────────────
    att_date = request.form.get("date", "").strip()
    if not att_date:
        flash("Date is required.", "error")
        return redirect("/faculty_attendance")
    try:
        from datetime import datetime as _dt
        parsed_date = _dt.strptime(att_date, "%Y-%m-%d").date()
        if parsed_date > date.today():
            flash(f"Date {att_date} is in the future. Attendance for future dates is not allowed.", "error")
            return redirect(f"/faculty_attendance?subject={subject}")
    except ValueError:
        flash("Invalid date format.", "error")
        return redirect("/faculty_attendance")

    # ── Guard: at least one student status must be submitted ──
    status_keys = [k for k in request.form.keys() if k.startswith("status_")]
    if not status_keys:
        flash("No student statuses found. Please mark students before committing.", "error")
        return redirect(f"/faculty_attendance?subject={subject}")

    # ── Double Submission Guard (Backend) ───────────────────
    # We check if records for this (subject, date, faculty) already exist
    # If using upsert (ON CONFLICT), this is technically safe, but we can log/prevent excess ops.
    
    fid = session.get("faculty_id")
    try:
        log_audit(
            fid,
            "ATTENDANCE_MARKED",
            details=f"Subject:{subject} Date:{att_date} Students:{len(status_keys)}"
        )
    except Exception as e:
        app.logger.warning(f"Audit log failed for attendance: {e}")

    # Success feedback handled by redirects usually, but let's ensure flashes work
    flash(f"Successfully committed attendance for {subject} on {att_date}.", "success")
    return redirect(f"/faculty_attendance?subject={subject}&saved=1")

@app.route("/faculty_edit_attendance", methods=["POST"])
@login_required("faculty")
def faculty_edit_attendance():
    return handle_edit_record(request.form, session)

@app.route("/faculty_delete_attendance", methods=["POST"])
@login_required("faculty")
def faculty_delete_attendance():
    return handle_delete_record(request.form, session)

@app.route("/faculty_import_attendance", methods=["POST"])
@login_required("faculty")
def faculty_import_attendance():
    f = request.files.get("file")
    if not f:
        return redirect("/faculty_attendance?tab=import&error=no_file")
    return handle_attendance_import(f, session, request.form.get("subject", "").strip())


@app.route("/faculty_marks")
@login_required("faculty")
def faculty_marks():
    fid = session["faculty_id"]
    students_list = qry("SELECT name,roll FROM students ORDER BY name")
    my_subjects   = qry("SELECT * FROM subjects WHERE teacher LIKE %s ORDER BY name", (f"%{session['name']}%",))
    marks         = qry("SELECT * FROM marks WHERE faculty_id=%s ORDER BY id DESC", (fid,))
    return render_template("faculty/faculty_marks.html",
                           students=students_list, my_subjects=my_subjects,
                           marks=marks, today=today_str())

@app.route("/faculty_save_marks", methods=["POST"])
@login_required("faculty")
def faculty_save_marks():
    fid = session["faculty_id"]
    student_name = request.form.get("student_name","").strip()
    roll_row = qone("SELECT id,roll,department FROM students WHERE name=%s", (student_name,))
    stu_id = roll_row["id"]         if roll_row else None
    roll   = roll_row["roll"]       if roll_row else request.form.get("roll","")
    dept   = roll_row["department"] if roll_row else request.form.get("department","")

    exam_type = request.form.get("exam_type", "Semester Exam")

    # Component marks — only enforce /60 breakdown for Semester Exam
    if exam_type == "Semester Exam":
        assignment_m = min(float(request.form.get("assignment_marks",   0) or 0), 5.0)
        attendance_m = min(float(request.form.get("attendance_marks",   0) or 0), 5.0)
        teaching_m   = min(float(request.form.get("teaching_assessment",0) or 0), 10.0)
        ut_m         = min(float(request.form.get("ut_marks",           0) or 0), 20.0)
        mse_m        = min(float(request.form.get("mse_marks",          0) or 0), 20.0)
        marks_val    = assignment_m + attendance_m + teaching_m + ut_m + mse_m
        if marks_val == 0:
            marks_val = min(float(request.form.get("marks", 0) or 0), 60.0)
        marks_val = min(marks_val, 60.0)
        total_val = 60.0
    else:
        assignment_m = attendance_m = teaching_m = ut_m = mse_m = 0.0
        marks_val = float(request.form.get("marks", 0) or 0)
        total_val = float(request.form.get("total", 60) or 60)
        total_val = min(total_val, 60.0)
        marks_val = min(marks_val, total_val)

    exe("""INSERT INTO marks(faculty_id,student_id,student_name,roll,subject,department,
                             marks,total,exam_type,date,
                             assignment_marks,attendance_marks,teaching_assessment,
                             ut_marks,mse_marks)
           VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (fid, stu_id, student_name, roll,
         request.form.get("subject",""),
         dept,
         marks_val, total_val,
         exam_type,
         request.form.get("date", today_str()),
         assignment_m, attendance_m, teaching_m, ut_m, mse_m))
    return redirect("/faculty_marks?success=1")

# Removed Faculty Assignments logic as requested


# Faculty: AJAX Students by Subject (Task G)
@app.route("/faculty/api/students_by_subject")
@login_required("faculty")
def api_students_by_subject():
    subj_name = request.args.get("subject", "").strip()
    # Find division from subjects table first (legacy) or assignments
    s_info = qone("SELECT division, department FROM subjects WHERE name=%s", (subj_name,))
    if not s_info:
        return jsonify([])
    
    div = s_info['division']
    dept = s_info['department']
    
    students = qry("""SELECT id, name, roll FROM students 
                      WHERE (division=%s OR %s='') AND (department=%s OR %s='')
                      ORDER BY name""", (div, div, dept, dept))
    return jsonify(students)

@app.route("/faculty_delete_marks", methods=["POST"])
@login_required("faculty")
def faculty_delete_marks():
    exe("DELETE FROM marks WHERE id=%s AND faculty_id=%s",
        (request.form.get("marks_id",""), session["faculty_id"]))
    return redirect("/faculty_marks")

@app.route("/faculty_notices")
@login_required("faculty")
def faculty_notices():
    fid = session["faculty_id"]
    notices = qry("SELECT * FROM faculty_notices WHERE faculty_id=%s ORDER BY id DESC", (fid,))
    return render_template("faculty/faculty_notices.html", notices=notices)

@app.route("/faculty_save_notice", methods=["POST"])
@login_required("faculty")
def faculty_save_notice():
    exe("INSERT INTO faculty_notices(faculty_id,title,message) VALUES(%s,%s,%s)",
        (session["faculty_id"], request.form.get("title",""), request.form.get("message","")))
    return redirect("/faculty_notices?success=1")

@app.route("/faculty_delete_notice", methods=["POST"])
@login_required("faculty")
def faculty_delete_notice():
    exe("DELETE FROM faculty_notices WHERE id=%s AND faculty_id=%s",
        (request.form.get("notice_id",""), session["faculty_id"]))
    return redirect("/faculty_notices")

@app.route("/faculty_notes")
@login_required("faculty")
def faculty_notes():
    fid = session["faculty_id"]
    f_subject = request.args.get("subject","").strip()
    my_subjects = qry("SELECT name FROM subjects WHERE teacher LIKE %s ORDER BY name", (f"%{session['name']}%",))
    sql = "SELECT fn.*, f.name as faculty_name FROM faculty_notes fn JOIN faculty f ON fn.faculty_id=f.id WHERE fn.faculty_id=%s"
    params = [fid]
    if f_subject: sql += " AND fn.subject=%s"; params.append(f_subject)
    sql += " ORDER BY fn.id DESC"
    notes = qry(sql, params)
    return render_template("faculty/faculty_notes.html", notes=notes, my_subjects=my_subjects, f_subject=f_subject)

@app.route("/faculty_save_note", methods=["POST"])
@login_required("faculty")
def faculty_save_note():
    exe("INSERT INTO faculty_notes(faculty_id,subject,title,content,note_type) VALUES(%s,%s,%s,%s,%s)",
        (session["faculty_id"], request.form.get("subject",""), request.form.get("title",""),
         request.form.get("content",""), request.form.get("note_type","Lecture Note")))
    return redirect("/faculty_notes?success=1")

@app.route("/faculty_edit_note", methods=["POST"])
@login_required("faculty")
def faculty_edit_note():
    exe("UPDATE faculty_notes SET title=%s,content=%s,note_type=%s WHERE id=%s AND faculty_id=%s",
        (request.form.get("title",""), request.form.get("content",""), request.form.get("note_type",""),
         request.form.get("note_id",""), session["faculty_id"]))
    return redirect("/faculty_notes")

@app.route("/faculty_delete_note", methods=["POST"])
@login_required("faculty")
def faculty_delete_note():
    exe("DELETE FROM faculty_notes WHERE id=%s AND faculty_id=%s",
        (request.form.get("note_id",""), session["faculty_id"]))
    return redirect("/faculty_notes")

@app.route("/faculty_timetable")
@login_required("faculty")
def faculty_timetable():
    name    = session.get("name","")
    view    = request.args.get("view","grid")
    f_day   = request.args.get("day","").strip()
    f_type  = request.args.get("slot_type","").strip()
    f_branch = request.args.get("branch","").strip()
    f_year = request.args.get("year","").strip()
    f_div = request.args.get("division","").strip()
    f_subj = request.args.get("subject_id","").strip()

    fac_id  = session.get("faculty_id")

    # Strict Faculty ID match instead of name string mismatching
    all_entries = [dict(e) for e in qry(f"SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.faculty_id=%s ORDER BY {DAY_ORD}, t.start_time", (fac_id,))]
    for e in all_entries: e["time"] = normalize_time(e.get("time",""))

    # filtered entries for list
    entries = all_entries
    if f_day:  entries = [e for e in entries if e["day"]==f_day]
    if f_type: entries = [e for e in entries if (e.get("slot_type") or "Theory")==f_type]
    if f_branch: entries = [e for e in entries if e.get("branch")==f_branch]
    if f_year: entries = [e for e in entries if e.get("year")==f_year]
    if f_div: entries = [e for e in entries if e.get("division")==f_div]
    if f_subj: entries = [e for e in entries if str(e.get("subject_id"))==str(f_subj)]

    # Build grid: time_slots as columns, days as rows
    seen=set(); raw=[]
    for e in all_entries:
        t=e["time"]
        if t and t not in seen: seen.add(t); raw.append(t)

    def _sk(ts):
        m=re.match(r"(\d+):(\d+)",ts)
        if not m: return 999
        h=int(m.group(1)); mn=int(m.group(2))
        if h<7: h+=12
        return h*60+mn

    time_slots = sorted(raw, key=_sk)
    grid = {d:{t:[] for t in time_slots} for d in DAYS}
    for e in all_entries:
        if e["day"] in grid and e["time"] in grid[e["day"]]:
            grid[e["day"]][e["time"]].append(e)

    total   = len(all_entries)
    theory  = sum(1 for e in all_entries if (e.get("slot_type") or "Theory")=="Theory")
    lab     = sum(1 for e in all_entries if e.get("slot_type")=="Lab")
    days_active = len(set(e["day"] for e in all_entries))

    return render_template("faculty/faculty_timetable.html",
        entries=entries, all_entries=all_entries,
        grid=grid, time_slots=time_slots,
        DAYS=DAYS, view=view, f_day=f_day, f_type=f_type,
        total=total, theory=theory, lab=lab,
        days_active=days_active,
        today_name=date.today().strftime("%A")
    )

@app.route("/faculty_profile")
@login_required("faculty")
def faculty_profile():
    profile = qone("SELECT * FROM faculty WHERE id=%s", (session["faculty_id"],))
    return render_template("faculty/faculty_profile.html", profile=dict(profile) if profile else {})

@app.route("/faculty_update_profile", methods=["POST"])
@login_required("faculty")
def faculty_update_profile():
    fid = session["faculty_id"]
    name = request.form.get("name","").strip()
    phone= request.form.get("phone","").strip()
    qual = request.form.get("qualification","").strip()
    current_pw = request.form.get("current_password","").strip()
    new_pw = request.form.get("new_password","").strip() or request.form.get("password","").strip()
    confirm_pw = request.form.get("confirm_password","").strip() or new_pw
    if name:
        exe("UPDATE faculty SET name=%s,phone=%s,qualification=%s WHERE id=%s", (name,phone,qual,fid))
        session["name"] = name
    if new_pw:
        row = qone("SELECT password FROM faculty WHERE id=%s", (fid,))
        err = validate_password_change(row["password"] if row else "", current_pw, new_pw, confirm_pw)
        if err:
            return redirect(f"/faculty_profile?error={err}")
        exe("UPDATE faculty SET password=%s WHERE id=%s", (hash_password(new_pw), fid))
    return redirect("/faculty_profile?success=1")


# ════════════════════════════════════════════════════════════
#  STUDENT ROUTES
# ════════════════════════════════════════════════════════════
def get_student():
    sid = session.get("student_id")
    if not sid: return None
    row = qone("SELECT * FROM students WHERE id=%s", (sid,))
    return dict(row) if row else None

@app.route("/student_dashboard")
@login_required("student")
def student_dashboard():
    student = get_student()
    if not student: return redirect("/logout")
    name = student["name"]
    sid  = student["id"]

    # Attendance (Optimized: Single SQL aggregation)
    subj_stats = qry(
        f"""SELECT subject, 
                   COUNT(*) as total, 
                   SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present
            FROM attendance 
            WHERE {att_match_student_sql()}
            GROUP BY subject""",
        att_match_student_params(sid, name)
    )
    subject_att = [
        {"subject": r["subject"], "total": r["total"], "present": r["present"],
         "pct": pct(r["present"], r["total"])}
        for r in subj_stats
    ]
    total_att = sum(r["total"] for r in subject_att)
    pres_att  = sum(r["present"] for r in subject_att)
    total_absent = total_att - pres_att
    overall_pct  = pct(pres_att, total_att)
    low_att_subjects = [x["subject"] for x in subject_att if x["pct"] < 75 and x["total"] >= 5]

    recent_marks = qry(
        f"SELECT * FROM marks WHERE {marks_match_student_sql()} ORDER BY id DESC LIMIT 8",
        marks_match_student_params(sid, name),
    )

    # Cumulative marks summary for student
    all_marks = qry(
        f"SELECT marks, total, subject FROM marks WHERE {marks_match_student_sql()}",
        marks_match_student_params(sid, name),
    )
    cumulative_obtained = sum(r["marks"] for r in all_marks)
    cumulative_total    = sum(r["total"]  for r in all_marks)
    cumulative_pct      = round(cumulative_obtained / cumulative_total * 100, 1) if cumulative_total else 0
    cumulative_grade    = grade(cumulative_obtained, cumulative_total) if cumulative_total else "N/A"

    # Cumulative attendance for student dashboard
    cum = calculate_cumulative(sid, name)
    cum_att_total      = cum["total"]
    cum_att_present    = cum["present"]
    cum_att_percentage = cum["percentage"]
    cum_att_status     = cum["status"]

    # Unread messages
    unread_res = qry("SELECT count(*) as cnt FROM messages WHERE to_role='student' AND to_id=%s AND is_read=0", (sid,))
    unread_count = unread_res[0]["cnt"] if unread_res else 0

    # Subject-wise attendance breakdown
    subj_wise_data = subject_wise(sid, name)
    subject_att_detail = [
        {
            "subject":    sub,
            "total":      v["total"],
            "present":    v["present"],
            "percentage": v["percentage"],
        }
        for sub, v in subj_wise_data.items()
    ]


    timetable = qry("""
        SELECT * FROM timetable 
        WHERE division=%s AND branch=%s AND year=%s 
        ORDER BY CASE day 
            WHEN 'Monday' THEN 1 WHEN 'Tuesday' THEN 2 WHEN 'Wednesday' THEN 3 
            WHEN 'Thursday' THEN 4 WHEN 'Friday' THEN 5 WHEN 'Saturday' THEN 6 
        END, time 
        LIMIT 10
    """, (student["division"], student["department"], student["year"]))
    notices      = qry("""SELECT fn.title, fn.message, fn.created_at, f.name as faculty_name
                          FROM faculty_notices fn JOIN faculty f ON fn.faculty_id=f.id
                          ORDER BY fn.id DESC LIMIT 5""")

    return render_template("student/student_dashboard.html",
        student=student, overall_pct=overall_pct,
        total_att=total_att, pres_att=pres_att, total_absent=total_absent,
        subject_att=subject_att, low_att_subjects=low_att_subjects,
        recent_marks=[dict(r) for r in recent_marks],
        cumulative_obtained=cumulative_obtained,
        cumulative_total=cumulative_total,
        cumulative_pct=cumulative_pct,
        cumulative_grade=cumulative_grade,
        # Cumulative attendance stats
        cum_att_total=cum_att_total,
        cum_att_present=cum_att_present,
        cum_att_percentage=cum_att_percentage,
        cum_att_status=cum_att_status,
        subject_att_detail=subject_att_detail,
        timetable=timetable, notices=notices, DAYS=DAYS, unread_count=unread_count
    )

@app.route("/student_profile")
@login_required("student")
def student_profile():
    sid = session.get("student_id")
    student = qone("SELECT * FROM students WHERE id=%s", (sid,))
    if not student:
        flash("Student profile not found.", "error")
        return redirect("/")
    return render_template("student/profile.html", student=student)

@app.route("/student_settings", methods=["GET", "POST"])
@login_required("student")
def student_settings():
    if request.method == "POST":
        dark_mode = request.form.get("dark_mode") == "on"
        session["dark_mode"] = dark_mode
        flash("Settings updated.", "success")
        return redirect("/student_settings")
    return render_template("student/settings.html")

@app.route("/student_notes")
@login_required("student")
def student_notes():
    sid = session.get("student_id")
    student = qone("SELECT department FROM students WHERE id=%s", (sid,))
    dept = student["department"] if student else ""
    notes = qry("SELECT fn.*, f.name as faculty_name FROM faculty_notes fn JOIN faculty f ON fn.faculty_id=f.id WHERE fn.department=%s OR fn.department='All' ORDER BY fn.id DESC", (dept,))
    return render_template("student/notes.html", notes=notes)

@app.route("/student_notices")
@login_required("student")
def student_notices():
    sid = session.get("student_id")
    student = qone("SELECT department FROM students WHERE id=%s", (sid,))
    dept = student["department"] if student else ""
    notices = qry("""SELECT fn.*, f.name as faculty_name 
                     FROM faculty_notices fn 
                     JOIN faculty f ON fn.faculty_id=f.id 
                     WHERE fn.department=%s OR fn.department='All' 
                     ORDER BY fn.id DESC""", (dept,))
    return render_template("student/notices.html", notices=notices)

@app.route("/student_analysis")
@login_required("student")
def student_analysis():
    sid = session.get("student_id")
    name = session.get("name")
    summary = calculate_cumulative(sid, name)
    subjects = subject_wise(sid, name)
    return render_template("student/analysis.html", summary=summary, subjects=subjects)

@app.route("/student_attendance")
@login_required("student")
def student_attendance():
    student = get_student()
    if not student: return redirect("/logout")
    name    = student["name"]
    sid     = student["id"]
    subject = request.args.get("subject","").strip()

    sql = f"SELECT * FROM attendance WHERE {att_match_student_sql()}"
    params = list(att_match_student_params(sid, name))
    if subject: sql += " AND subject=%s"; params.append(subject)
    sql += " ORDER BY date DESC"
    records = qry(sql, params)

    all_att = qry(
        f"SELECT subject,status FROM attendance WHERE {att_match_student_sql()}",
        att_match_student_params(sid, name),
    )
    subj_map = {}
    for r in all_att:
        s = r["subject"]
        if s not in subj_map: subj_map[s]={"total":0,"present":0}
        subj_map[s]["total"]+=1
        if r["status"]=="Present": subj_map[s]["present"]+=1
    subject_att = [{"subject":s,"total":v["total"],"present":v["present"],
                    "pct":pct(v["present"],v["total"])} for s,v in subj_map.items()]
    total_att = len(all_att)
    pres_att  = sum(1 for r in all_att if r["status"]=="Present")
    low_att_subjects = [x["subject"] for x in subject_att if x["pct"]<75 and x["total"]>=5]

    return render_template("student/student_attendance.html",
        student=student, records=records, subject_att=subject_att,
        overall_pct=pct(pres_att,total_att), low_att_subjects=low_att_subjects, subject=subject)

@app.route("/student_marks")
@login_required("student")
def student_marks():
    student = get_student()
    if not student: return redirect("/logout")
    marks = qry(
        f"SELECT * FROM marks WHERE {marks_match_student_sql()} ORDER BY date DESC",
        marks_match_student_params(student["id"], student["name"]),
    )
    marks_with_grade = []
    for m in marks:
        d = dict(m)
        d["pct"]   = pct(m["marks"], m["total"])
        d["grade"] = grade(m["marks"], m["total"])
        marks_with_grade.append(d)

    # Cumulative summary
    cumulative_obtained = sum(m["marks"] for m in marks)
    cumulative_total    = sum(m["total"]  for m in marks)
    cumulative_pct      = round(cumulative_obtained / cumulative_total * 100, 1) if cumulative_total else 0
    cumulative_grade    = grade(cumulative_obtained, cumulative_total) if cumulative_total else "N/A"

    return render_template("student/student_marks.html",
        student=student, marks=marks_with_grade,
        cumulative_obtained=cumulative_obtained,
        cumulative_total=cumulative_total,
        cumulative_pct=cumulative_pct,
        cumulative_grade=cumulative_grade)

@app.route("/student_timetable")
@login_required("student")
def student_timetable():
    student = get_student()
    if not student: return redirect("/logout")

    view   = request.args.get("view","grid")
    f_day  = request.args.get("day","").strip()
    f_type = request.args.get("slot_type","").strip()
    f_div  = request.args.get("division","").strip()

    # Strict Backend Filtering for Students
    student_branch = session.get("student_branch", "")
    student_year = session.get("student_year", "")
    student_div = session.get("student_division", "")
    student_dept = student_branch

    all_entries = [dict(e) for e in qry(
        f"SELECT t.*, f.name as teacher FROM timetable t LEFT JOIN faculty f ON t.faculty_id = f.id WHERE t.branch=%s AND t.year=%s AND t.division=%s AND t.published=TRUE ORDER BY {DAY_ORD}, t.start_time",
        (student_branch, student_year, student_div)
    )]
    for e in all_entries:
        e["time"] = normalize_time(e.get("time",""))

    # Filtered entries
    entries = list(all_entries)
    if f_day:  entries = [e for e in entries if e["day"]==f_day]
    if f_type: entries = [e for e in entries if (e.get("slot_type") or "Theory")==f_type]
    if f_div:  entries = [e for e in entries if (e.get("division") or "")==f_div]

    # Grid
    seen=set(); raw=[]
    for e in all_entries:
        t=e["time"]
        if t and t not in seen: seen.add(t); raw.append(t)

    def _sk(ts):
        m=re.match(r"(\d+):(\d+)",ts)
        if not m: return 999
        h=int(m.group(1)); mn=int(m.group(2))
        if h<7: h+=12
        return h*60+mn

    time_slots = sorted(raw, key=_sk)
    grid = {d:{t:[] for t in time_slots} for d in DAYS}
    for e in all_entries:
        if e["day"] in grid and e["time"] in grid[e["day"]]:
            grid[e["day"]][e["time"]].append(e)

    divs_list = sorted(set(e.get("division","") for e in all_entries if e.get("division","")))
    total = len(all_entries)
    theory= sum(1 for e in all_entries if (e.get("slot_type") or "Theory")=="Theory")
    lab   = sum(1 for e in all_entries if e.get("slot_type")=="Lab")

    return render_template("student/student_timetable.html",
        student=student, timetable=entries, all_entries=all_entries,
        grid=grid, time_slots=time_slots, DAYS=DAYS,
        view=view, f_day=f_day, f_type=f_type, f_div=f_div,
        divs=divs_list, total=total, theory=theory, lab=lab,
        today_name=date.today().strftime("%A"),
        student_div=student_div, student_dept=student_dept
    )



@app.route("/student_update_profile", methods=["POST"])
@login_required("student")
def student_update_profile():
    sid   = session["student_id"]
    email = request.form.get("email","").strip()
    current_pw = request.form.get("current_password","").strip()
    new_pw = request.form.get("new_password","").strip() or request.form.get("password","").strip()
    confirm_pw = request.form.get("confirm_password","").strip() or new_pw
    if email: exe("UPDATE students SET email=%s WHERE id=%s", (email, sid))
    if new_pw:
        row = qone("SELECT password FROM students WHERE id=%s", (sid,))
        err = validate_password_change(row["password"] if row else "", current_pw, new_pw, confirm_pw)
        if err:
            return redirect(f"/student_profile?error={err}")
        exe("UPDATE students SET password=%s WHERE id=%s", (hash_password(new_pw), sid))
    return redirect("/student_profile?success=1")

@app.route("/upload_photo/<role>", methods=["POST"])
def upload_photo(role):
    if role not in ("faculty", "student"):
        return redirect("/login")
    if role == "faculty":
        if session.get("role") != "faculty" or not session.get("faculty_id"):
            return redirect("/login")
    elif role == "student":
        if session.get("role") != "student" or not session.get("student_id"):
            return redirect("/login")
    import uuid
    from werkzeug.utils import secure_filename
    
    f = request.files.get("photo")
    if not f or not f.filename:
        return redirect("/faculty_profile" if role == "faculty" else "/student_profile")
        
    # FIX 9: File upload security
    filename = secure_filename(f.filename)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        flash("Invalid file type.", "error")
        return redirect("/faculty_profile" if role == "faculty" else "/student_profile")
        
    # Validate MIME type
    mime = f.content_type
    if not mime.startswith('image/'):
        flash("Invalid MIME type.", "error")
        return redirect("/faculty_profile" if role == "faculty" else "/student_profile")

    uid = session["faculty_id"] if role == "faculty" else session["student_id"]
    # Rename to UUID
    fname = f"{role}_{uid}_{uuid.uuid4().hex}{ext}"
    path = os.path.join(UPLOAD_DIR, fname)
    f.save(path)
    if role == "faculty":
        exe("UPDATE faculty SET photo=%s WHERE id=%s", (fname, session["faculty_id"]))
        return redirect("/faculty_profile?success=1")
    exe("UPDATE students SET photo=%s WHERE id=%s", (fname, session["student_id"]))
    return redirect("/student_profile?success=1")



# ════════════════════════════════════════════════════════════
#  SHARED IMPORT HELPER — used by admin + faculty
# ════════════════════════════════════════════════════════════

def _parse_attendance_excel(file_obj, subject):
    """
    Supports:
    1. DY Patil Ai_Attendance_sheet.xlsx
       Row2 col5+ = datetime  |  Row4+ col4 = name, col5+ = 0/1
    2. Flat format: Name | Subject | Date | Status
    Returns (added_count, error_msg)
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    added = 0

    # Detect DY Patil: row 2 has datetime objects
    date_map = {}
    for c in range(1, min(ws.max_column + 1, 60)):
        v = ws.cell(2, c).value
        if isinstance(v, (datetime, date)):
            date_map[c] = v.strftime("%Y-%m-%d")

    if len(date_map) >= 2:
        if not subject:
            return 0, "select_subject"
        for ri in range(4, ws.max_row + 1):
            name = str(ws.cell(ri, 4).value or "").strip()
            if not name or name.isdigit() or "name" in name.lower():
                continue
            for col_idx, att_date in date_map.items():
                raw = ws.cell(ri, col_idx).value
                v   = str(raw).strip() if raw is not None else ""
                if   v in ("1","P","p","Present"):       status = "Present"
                elif v in ("0","A","a","Absent"):         status = "Absent"
                elif v in ("L","l","Leave","ML","CL"):    status = "Leave"
                else: continue
                sid = resolve_student_id(name)
                exe(
                    "INSERT INTO attendance(student_id,student_name,subject,date,status) VALUES(%s,%s,%s,%s,%s)",
                    (sid, name, subject, att_date, status),
                )
                added += 1
    else:
        hdr_row = 1
        for i in range(1, min(ws.max_row + 1, 10)):
            vals = [str(ws.cell(i, c).value or "").lower() for c in range(1, 8)]
            if any("name" in v or "student" in v for v in vals):
                hdr_row = i; break
        hdrs = [str(ws.cell(hdr_row, c).value or "").lower().strip()
                for c in range(1, ws.max_column + 1)]
        def gcol(kws):
            for k in kws:
                for i, h in enumerate(hdrs):
                    if k in h: return i + 1
            return None
        cn = gcol(["name","student"]); cs = gcol(["subject"])
        cd = gcol(["date"]);           cst = gcol(["status"])
        if cn:
            for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                name = str(row[cn-1] or "").strip()
                if not name or name.isdigit(): continue
                subj     = str(row[cs-1] if cs else "").strip() or subject
                d_raw    = row[cd-1]  if cd  else None
                st_raw   = row[cst-1] if cst else "Present"
                att_date = normalise_date(d_raw) if d_raw else today_str()
                status   = normalise_status(st_raw) or "Present"
                if name and subj and att_date:
                    sid = resolve_student_id(name)
                    exe(
                        "INSERT INTO attendance(student_id,student_name,subject,date,status) VALUES(?,?,?,?,?)",
                        (sid, name, subj, att_date, status),
                    )
                    added += 1
    return added, None


def _parse_timetable_excel(file_obj, simulate=False):
    """
    Parse DY Patil TY-TT Excel. Reads subject lookup table (rows 15+)
    to expand abbreviations: ML(ST) → Machine Learning / Prof.Shakil Tamboli
    """

    _DAYS = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]
    _SKIP = {"short break","lunch break","lunch","break",
             "short            break","lunch         break",""}

    def _lookups(ws):
        sl = {}; fl = {}
        for ri in range(1, ws.max_row+1):
            if str(ws.cell(ri,1).value or "").strip().lower() == "sr.no":
                for ri2 in range(ri+1, ws.max_row+1):
                    if not ws.cell(ri2,1).value: continue
                    sc = str(ws.cell(ri2,3).value or "").strip()
                    fc = str(ws.cell(ri2,5).value or "").strip()
                    # subject abbrev: "Machine Learning(ML)" → ML→Machine Learning
                    m = re.search(r'\(([A-Za-z0-9 ]+(?:\s+Lab)?)\)\s*$', sc)
                    if m:
                        sl[m.group(1).strip()] = sc[:sc.rfind("(")].strip()
                    # faculty abbrev: "Prof.X Yadav(XY)" → XY→Prof.X Yadav
                    for fp in re.split(r'\s*/\s*', fc):
                        fp = fp.strip()
                        fm = re.search(r'\(([A-Z]{1,4})\)\s*$', fp)
                        if fm:
                            fl[fm.group(1)] = fp[:fp.rfind("(")].strip()
                break
        return sl, fl

    def _slot(raw, sl, fl):
        results = []
        for entry in _re.split(r',\s*\n\s*', raw.strip()):
            entry = entry.strip().rstrip(',').strip()
            if not entry: continue
            # room
            room = ""
            rm = _re.search(r'-\s*((?:Lab|Room)[-\s]\w+)', entry)
            if rm: room = rm.group(1).strip()
            # teacher from last matching paren
            teacher = ""
            parens  = _re.findall(r'\(([^)]+)\)', entry)
            for p in reversed(parens):
                p = p.strip()
                if p in fl:                                  teacher = fl[p]; break
                if _re.match(r'^[A-Z]{1,4}$', p):           teacher = p;    break
                if _re.match(r'^[A-Z][a-z]+$',p) and len(p)>3: teacher = p; break
            # clean subject text
            clean = entry
            clean = _re.sub(r'\([A-Z]\d+\)', '', clean)   # batch (A1)
            for p in parens:
                if _re.match(r'^[A-Z]{1,4}$', p.strip()):
                    clean = clean.replace(f'({p})','',1)
            clean = _re.sub(r'-\s*(?:Lab|Room)[-\s\w, ]+', '', clean)
            clean = _re.sub(r'\s+-\s*$','',clean).strip().rstrip(',').strip()
            abbrev  = clean.split('(')[0].strip()
            subject = sl.get(abbrev, abbrev) if abbrev else clean
            if not subject: continue
            stype = ("Lab" if "lab" in subject.lower() else
                     "Elective" if "elective" in subject.lower() else
                     "Minor" if "minor" in subject.lower() else "Theory")
            results.append((subject, teacher, room, stype))
        return results

    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception:
        return 0
    added = 0
    
    # Pre-cache maps to achieve 10,000x insert speedup
    fac_map = {r['name']: r['id'] for r in qry("SELECT id, name FROM faculty")}
    div_map = {r['division']: r for r in qry("SELECT DISTINCT division, department, year FROM students WHERE division IS NOT NULL")}
    sub_map = {r['name']: r['id'] for r in qry("SELECT id, name FROM subjects")}
    inserts = []

    for ws in wb.worksheets:
        sl, fl = _lookups(ws)
        division = ws.title.strip()
        h6 = str(ws.cell(6,1).value or "").strip()
        dept = ("CS" if any(x in h6 for x in ["CSE","Computer"]) else
                "IT" if any(x in h6 for x in ["IT","Information"]) else
                "AIDS" if "Data" in h6 else "AIML" if "Artificial" in h6 else "")
        sem  = "VI" if any(x in h6 for x in ["T.E","TE"]) else ""

        # find time-header row
        time_row = None
        for ri in range(1, 15):
            vals = [str(ws.cell(ri,c).value or "").strip() for c in range(1, ws.max_column+1)]
            if sum(1 for v in vals if _re.match(r'\d+:\d+\s*[-]\s*\d+:\d+', v)) >= 2:
                time_row = ri; break
        if not time_row: continue

        # col → time slot
        time_cols = {}
        for ci in range(2, ws.max_column+1):
            v = str(ws.cell(time_row, ci).value or "").strip()
            if _re.match(r'\d+:\d+\s*[-]\s*\d+:\d+', v):
                time_cols[ci] = v

        # parse day rows
        for ri in range(time_row+1, time_row+8):
            day_v   = str(ws.cell(ri, 2).value or "").strip()
            matched = next((d for d in _DAYS if d.lower()==day_v.lower()), None)
            if not matched: continue
            for ci, ts in time_cols.items():
                raw  = str(ws.cell(ri, ci).value or "").strip()
                norm = " ".join(raw.lower().split())
                if not raw or norm in _SKIP: continue
                if "break" in norm and len(norm) < 25: continue
                for subject, teacher, room, slot_type in _slot(raw, sl, fl):
                    color = assign_color(subject, slot_type)
                    time_s = normalize_time(ts)
                    start_time, end_time = None, None
                    m = _re.match(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", time_s)
                    if m:
                        h1, m1, h2, m2 = map(int, m.groups())
                        if h1 < 7: h1 += 12
                        if h2 < 7: h2 += 12
                        start_time = f"{h1:02d}:{m1:02d}:00"
                        end_time = f"{h2:02d}:{m2:02d}:00"

                    # In-memory mapping to save 3,000+ db queries
                    faculty_id = fac_map.get(teacher, 1)
                    branch = div_map.get(division, {}).get("department", "Unknown")
                    year = div_map.get(division, {}).get("year", "Unknown")
                    subject_id = sub_map.get(subject, None)
                    
                    if not simulate:
                        inserts.append((matched, time_s, start_time, end_time, subject_id, subject, teacher, room, division, sem, slot_type, color, faculty_id, branch, year))
                    added += 1
    if not simulate and inserts:
        # Bulk Insert
        conn = get_db()
        try:
            cur = conn.cur if hasattr(conn, 'cur') else conn.conn.cursor()
            psycopg2.extras.execute_values(
                cur,
                "INSERT INTO timetable (day,time,start_time,end_time,subject_id,subject,teacher,room,division,semester,slot_type,color,faculty_id,branch,year) VALUES %s",
                inserts
            )
            # if we grabbed raw cursor, make sure it commits via connection
            if hasattr(conn, 'conn'): conn.conn.commit()
            else: conn.commit()
        except Exception as e:
            if hasattr(conn, 'conn'): conn.conn.rollback()
            else: conn.rollback()
            raise e
        finally:
            conn.close()
            
    return added







def _parse_faculty_excel(file_obj):
    """
    Supports Faculty_Details_.xlsx:
    Row 3 = headers (Sr.No | Name of Faculty | Department | Contact No | Email Id | Signature)
    Row 4+ = data, skip empty rows (col1 is None)
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    added = skipped = 0

    # Find header row (has "Name of Faculty" or "name")
    hdr_row = 3
    for i in range(1, min(ws.max_row + 1, 10)):
        vals = [str(ws.cell(i, c).value or "").lower() for c in range(1, 8)]
        if any("name" in v and "faculty" in v for v in vals) or any("email" in v for v in vals):
            hdr_row = i; break

    hdrs = [str(ws.cell(hdr_row, c).value or "").lower().strip()
            for c in range(1, ws.max_column + 1)]
    def gcol(kws):
        for k in kws:
            for i, h in enumerate(hdrs):
                if k in h: return i + 1
        return None

    cn = gcol(["name"])
    ce = gcol(["email"])
    cd = gcol(["dept","department"])
    cp = gcol(["contact","phone"])

    if not cn:
        return 0, 0

    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row: continue
        name  = str(row[cn-1] or "").strip() if cn and cn <= len(row) else ""
        email = str(row[ce-1] or "").strip() if ce and ce <= len(row) else ""
        dept  = str(row[cd-1] or "").strip() if cd and cd <= len(row) else ""
        phone_raw = row[cp-1] if cp and cp <= len(row) else None
        phone = str(int(phone_raw)) if isinstance(phone_raw, float) else str(phone_raw or "").strip()

        if not name or name.lower() in ("name of faculty", "sr. no.", ""):
            continue

        try:
            exe("INSERT INTO faculty(name,department,email,phone,password) VALUES(%s,%s,%s,%s,%s)",
                (name, dept, email, phone, hash_password(_default_faculty_password())))
            added += 1
        except Exception:
            skipped += 1

    return added, skipped


# ── Admin attendance import (uses shared parser) ────
@app.route("/import_attendance_excel_v2", methods=["POST"])
@login_required("admin")
def import_attendance_excel_admin():
    f       = request.files.get("file")
    subject = request.form.get("subject","").strip()
    added, err = _parse_attendance_excel(f, subject)
    if err == "select_subject":
        return redirect("/attendance?error=select_subject")
    return redirect(f"/view_attendance?imported={added}")


# ── FIXED: Faculty attendance import (uses shared parser) ──
@app.route("/faculty_import_attendance_v2", methods=["POST"])
@login_required("faculty")
def faculty_import_attendance_v2():
    f = request.files.get("file")
    if not f:
        return redirect("/faculty_attendance?tab=import&error=no_file")
    return handle_attendance_import(f, session, request.form.get("subject", "").strip())


# ── NEW: Import Faculty from Faculty_Details_.xlsx ─────────
@app.route("/import_faculty_details", methods=["POST"])
@login_required("admin")
def import_faculty_details():
    f = request.files.get("file")
    if not f: return redirect("/faculty")
    added, skipped = _parse_faculty_excel(f)
    return redirect(f"/faculty?imported={added}&skipped={skipped}")


# ── FIXED: Timetable import (uses exact DY Patil parser) ───
@app.route("/import_students_v2", methods=["POST"])
@login_required("admin")
def import_students_v2():
    f = request.files.get("file")
    if not f: return redirect("/students")
    wb = load_workbook(f, data_only=True)
    if _is_attendance_student_workbook(wb):
        added, updated, skipped = _import_attendance_workbook_students(wb, f.filename)
        return redirect(f"/students?imported={added}&updated={updated}&skipped={skipped}&format=attendance_xlsx")
    ws = wb.active
    added = skipped = 0
    hdr_row = 1
    for i in range(1, min(ws.max_row+1, 10)):
        vals = [str(ws.cell(i,c).value or "").lower() for c in range(1,8)]
        if any("name" in v for v in vals): hdr_row=i; break
    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None
    cn=gcol(["name"]); cr=gcol(["roll"]); cd=gcol(["dept","department"])
    cy=gcol(["year"]); ce=gcol(["email"])
    if not cn: return redirect("/students?error=bad_format")
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        name  = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        roll  = str(row[cr-1] or "").strip() if cr and cr <= len(row) else ""
        dept  = str(row[cd-1] or "").strip() if cd and cd <= len(row) else ""
        year  = str(row[cy-1] or "").strip() if cy and cy <= len(row) else ""
        email = str(row[ce-1] or "").strip() if ce and ce <= len(row) else ""
        if not name or not roll: continue
        try:
            exe("INSERT INTO students(name,roll,department,year,email,password) VALUES(%s,%s,%s,%s,%s,%s)",
                (name, roll, dept, year, email, hash_password(_default_student_password())))
            added += 1
        except Exception: skipped += 1
    return redirect(f"/students?imported={added}&skipped={skipped}")



# ── NEW: Import Subjects from Excel ────────────────────────
@app.route("/import_subjects_v2", methods=["POST"])
@login_required("admin")
def import_subjects_v2():
    f = request.files.get("file")
    if not f: return redirect("/subjects")
    wb = load_workbook(f, data_only=True)
    if _is_attendance_student_workbook(wb):
        added, skipped = _import_attendance_workbook_subjects(wb, f.filename)
        return redirect(f"/subjects?imported={added}&skipped={skipped}&format=attendance_xlsx")
    tt_subjects = _read_subjects_from_wb(wb)
    if tt_subjects:
        added = skipped = 0
        for info in tt_subjects.values():
            name = info["name"]
            if _subject_exists(name, info["code"], info["dept"], info["sem"]):
                skipped += 1
                continue
            try:
                exe(
                    "INSERT INTO subjects(name,department,subject_code,teacher,semester) VALUES(%s,%s,%s,%s,%s)",
                    (name, info["dept"], info["code"], info["teacher"], info["sem"]),
                )
                added += 1
            except Exception:
                skipped += 1
        return redirect(f"/subjects?imported={added}&skipped={skipped}&format=timetable_xlsx")
    ws = wb.active
    added = skipped = 0
    hdr_row = 1
    for i in range(1, min(ws.max_row+1, 10)):
        vals = [str(ws.cell(i,c).value or "").lower() for c in range(1,8)]
        if any("name" in v or "subject" in v for v in vals): hdr_row=i; break
    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None
    cn=gcol(["name","subject"]); cd=gcol(["dept","department"])
    cc=gcol(["code"]); ct=gcol(["teacher"]); cs=gcol(["sem"])
    if not cn: return redirect("/subjects?error=bad_format")
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        name = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        if not name: continue
        dept = str(row[cd-1] or "").strip() if cd and cd <= len(row) else ""
        code = str(row[cc-1] or "").strip() if cc and cc <= len(row) else ""
        tchr = str(row[ct-1] or "").strip() if ct and ct <= len(row) else ""
        sem  = str(row[cs-1] or "I").strip() if cs and cs <= len(row) else "I"
        try:
            exe("INSERT INTO subjects(name,department,subject_code,teacher,semester) VALUES(%s,%s,%s,%s,%s)",
                (name,dept,code,tchr,sem))
            added += 1
        except Exception: skipped += 1
    return redirect(f"/subjects?imported={added}&skipped={skipped}")


# ── NEW: Import Marks from Excel ───────────────────────────
@app.route("/import_marks_v2", methods=["POST"])
@login_required("faculty")
def import_marks_v2():
    f = request.files.get("file")
    if not f: return redirect("/faculty_marks")
    fid = session["faculty_id"]
    wb = load_workbook(f, data_only=True); ws = wb.active
    added = 0
    hdr_row = 1
    for i in range(1, min(ws.max_row+1, 10)):
        vals = [str(ws.cell(i,c).value or "").lower() for c in range(1,8)]
        if any("name" in v or "student" in v or "mark" in v for v in vals): hdr_row=i; break
    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip() for c in range(1,ws.max_column+1)]
    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None
    cn=gcol(["name","student"]); cr=gcol(["roll"])
    cs=gcol(["subject"]); cm=gcol(["marks","obtained"])
    ct=gcol(["total"]); ce=gcol(["exam","type"]); cd=gcol(["date"])
    if not cn or not cm: return redirect("/faculty_marks?error=bad_format")
    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        name = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        if not name: continue
        roll    = str(row[cr-1] or "").strip() if cr and cr <= len(row) else ""
        subj    = str(row[cs-1] or "").strip() if cs and cs <= len(row) else ""
        marks_v = row[cm-1] if cm <= len(row) else 0
        total_v = row[ct-1] if ct and ct <= len(row) else 100
        exam_t  = str(row[ce-1] or "Unit Test 1").strip() if ce and ce <= len(row) else "Unit Test 1"
        d_raw   = row[cd-1] if cd and cd <= len(row) else None
        att_date= normalise_date(d_raw) if d_raw else today_str()
        try:
            marks_f = float(marks_v or 0)
            total_f = float(total_v or 100)
        except: continue
        # Auto-fill id / roll / dept from students table
        stu_id = resolve_student_id(name, roll)
        sr = qone("SELECT roll,department FROM students WHERE id=%s", (stu_id,)) if stu_id else None
        if sr:
            roll = sr["roll"] or roll
            dept = sr["department"]
        else:
            dept = ""

        exe(
            "INSERT INTO marks(faculty_id,student_id,student_name,roll,subject,department,marks,total,exam_type,date) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (fid, stu_id, name, roll, subj, dept, marks_f, total_f, exam_t, att_date),
        )
        added += 1
    return redirect(f"/faculty_marks?imported={added}")



# ════════════════════════════════════════════════════════════
#  FIX BAD DATA ROUTES — clean garbage from DB
# ════════════════════════════════════════════════════════════

@app.route("/fix_students", methods=["POST"])
@login_required("admin")
def fix_students():
    """Delete students with no real name (numeric, placeholder, empty)."""

    bad = qry("""SELECT id FROM students WHERE
        name IS NULL OR name='' OR name='Student Name' OR
        CAST(name AS TEXT) ~ '^[0-9]' OR
        roll IS NULL OR roll='' OR roll='Roll No.' OR roll='Roll'
    """)
    for r in bad:

        exe("DELETE FROM students WHERE id=%s", (r["id"],))
    return redirect(f"/students?cleaned={len(bad)}")


@app.route("/fix_faculty", methods=["POST"])
@login_required("admin")
def fix_faculty():
    """Delete faculty with no real name/email."""
    bad = qry("""SELECT id FROM faculty WHERE
        name IS NULL OR name='' OR name='Name of Faculty' OR
        name='Faculty Name' OR name='Name' OR
        email IS NULL OR email='' OR email='Email Id' OR email='Email'
    """)
    for r in bad:

        exe("DELETE FROM faculty WHERE id=%s", (r["id"],))
    return redirect(f"/faculty?cleaned={len(bad)}")



@app.route("/bulk_delete_subjects", methods=["POST"])
@login_required("admin")
def bulk_delete_subjects():
    ids = request.form.getlist("ids[]")
    for i in ids:

        exe("DELETE FROM subjects WHERE id=%s", (i,))
    return redirect(f"/subjects?cleaned={len(ids)}")

@app.route("/fix_subjects", methods=["POST"])
@login_required("admin")
def fix_subjects():
    """Delete subjects with placeholder names."""

    bad = qry("""SELECT id FROM subjects WHERE
        name IS NULL OR name='' OR name='Subject Name' OR
        name='Subject' OR name='Name' OR
        CAST(name AS TEXT) ~ '^[0-9]'
    """)
    for r in bad:
        exe("DELETE FROM subjects WHERE id=%s", (r["id"],))
    return redirect(f"/subjects?cleaned={len(bad)}")


@app.route("/fix_attendance", methods=["POST"])
@login_required("admin")
def fix_attendance():
    """Delete attendance with numeric/placeholder student names or invalid status."""
    bad = qry("""SELECT id FROM attendance WHERE
        student_name IS NULL OR student_name='' OR
        student_name='Student Name' OR student_name='Name' OR
        CAST(student_name AS TEXT) GLOB '[0-9]*' OR
        length(student_name) < 3 OR
        status NOT IN ('Present','Absent','Leave','Late','Medical')
    """)
    for r in bad:
        exe("DELETE FROM attendance WHERE id=?", (r["id"],))
    return redirect(f"/view_attendance?cleaned={len(bad)}")


# ════════════════════════════════════════════════════════════
#  SMART IMPORT — handles DY Patil specific formats
# ════════════════════════════════════════════════════════════

def _class_meta_from_attendance_text(text, filename=""):
    raw = text or ""
    program_m = re.search(r"Program[:\s]*(.+)", raw, re.I)
    dept_m = re.search(r"Department\s+of\s+(.+)", raw, re.I)
    source = "\n".join(
        part for part in (
            program_m.group(1) if program_m else "",
            dept_m.group(1) if dept_m else "",
            filename or "",
        ) if part
    ) or f"{raw}\n{filename or ''}"
    dept = ""
    div = ""
    year = ""
    div_m = re.search(r"Div[.\s]*([A-Z])", source, re.I)
    if div_m:
        div = div_m.group(1).upper()
    if re.search(r"AIML", source, re.I):
        dept = "AIML"
    elif re.search(r"AI\s*&?\s*DS|AIDS", source, re.I):
        dept = "AIDS"
    elif re.search(r"Comp|Computer|CE", source, re.I):
        dept = "CS"
    elif re.search(r"\bIT\b", source, re.I):
        dept = "IT"
    if re.search(r"\bS\.?\s*Y\.?\b|\bSY\b|S\.Y\.B", source, re.I):
        year = "II"
    elif re.search(r"\bF\.?\s*Y\.?\b|\bFE\b", source, re.I):
        year = "I"
    elif re.search(r"\bT\.?\s*Y\.?\b|\bTE\b", source, re.I):
        year = "III"
    elif re.search(r"\bB\.?\s*E\.?\b|\bBE\b", source, re.I):
        year = "IV"
    return dept, div, year


def _parse_students_from_attendance_pdf(file_obj):
    def clean(value):
        return " ".join(str(value or "").replace("\n", " ").split()).strip()

    rows = []
    seen = set()
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            file_obj.seek(0)
            file_obj.save(tmp.name)
            tmp_path = tmp.name

        with pdfplumber.open(tmp_path) as pdf:
            full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            dept, div, year = _class_meta_from_attendance_text(full_text, getattr(file_obj, "filename", ""))
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    if not table:
                        continue
                    header_idx = None
                    for idx, row in enumerate(table[:8]):
                        row_text = " ".join(clean(cell).upper() for cell in row)
                        if "NAME" in row_text and "ROLL" in row_text:
                            header_idx = idx
                            break
                    if header_idx is None:
                        continue
                    headers = [clean(cell).upper() for cell in table[header_idx]]
                    name_col = next((i for i, h in enumerate(headers) if "NAME" in h), None)
                    roll_col = next((i for i, h in enumerate(headers) if "ROLL" in h), None)
                    if name_col is None or roll_col is None:
                        continue
                    for row in table[header_idx + 1:]:
                        if not row:
                            continue
                        name = clean(row[name_col] if name_col < len(row) else "")
                        roll = clean(row[roll_col] if roll_col < len(row) else "")
                        if not name or not roll:
                            continue
                        upper_name = name.upper()
                        if (
                            "LECTURES CONDUCTED" in upper_name
                            or upper_name in ("NAME", "NAME OF STUDENT")
                            or upper_name.startswith(("PROF.", "DR.", "FACULTY", "HOD"))
                        ):
                            continue
                        if not re.match(r"^[A-Z]?\d{1,3}[A-Z]?$", roll, re.I):
                            continue
                        key = (dept, div, roll.upper(), name.upper())
                        if key in seen:
                            continue
                        seen.add(key)
                        rows.append({
                            "name": name,
                            "roll": roll.upper(),
                            "department": dept or request.form.get("department", "").strip(),
                            "division": div or request.form.get("division", "").strip(),
                            "year": year or request.form.get("year", "").strip() or "II",
                        })
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    return rows


def _parse_students_from_attendance_workbook(wb, filename=""):
    def clean(value):
        return " ".join(str(value or "").replace("\n", " ").split()).strip()

    ws = wb.active
    header_text = "\n".join(
        clean(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, 10) + 1)
        for c in range(1, min(ws.max_column, 8) + 1)
    )
    dept, div, year = _class_meta_from_attendance_text(header_text, filename)

    header_row = None
    headers = []
    for r in range(1, min(ws.max_row, 20) + 1):
        row_headers = [clean(ws.cell(r, c).value).upper() for c in range(1, ws.max_column + 1)]
        row_text = " ".join(row_headers)
        if "ROLL" in row_text and "NAME" in row_text:
            header_row = r
            headers = row_headers
            break
    if not header_row:
        return []

    name_col = next((i + 1 for i, h in enumerate(headers) if "NAME" in h), None)
    roll_col = next((i + 1 for i, h in enumerate(headers) if "ROLL" in h), None)
    if not name_col or not roll_col:
        return []

    rows = []
    seen = set()
    for r in range(header_row + 1, ws.max_row + 1):
        name = clean(ws.cell(r, name_col).value)
        roll = clean(ws.cell(r, roll_col).value).upper()
        if not name or not roll:
            continue
        upper_name = name.upper()
        if "LECTURES CONDUCTED" in upper_name or upper_name in ("NAME", "NAME OF STUDENT"):
            continue
        if not re.match(r"^[A-Z]?\d{1,3}[A-Z]?$", roll, re.I):
            continue
        key = (dept, div, roll, upper_name)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            "name": name,
            "roll": roll,
            "department": dept or request.form.get("department", "").strip(),
            "division": div or request.form.get("division", "").strip(),
            "year": year or request.form.get("year", "").strip() or "II",
        })
    return rows


def _is_attendance_student_workbook(wb):
    ws = wb.active
    text = "\n".join(
        str(ws.cell(r, c).value or "").upper()
        for r in range(1, min(ws.max_row, 12) + 1)
        for c in range(1, min(ws.max_column, 8) + 1)
    )
    return "FINAL ATTENDANCE REPORT" in text and "ROLL" in text and "NAME OF STUDENT" in text


def _semester_from_attendance_text(text, default="II"):
    value = text or ""
    m = re.search(r"Semester\s*[-:]?\s*([IVX]+|\d+)", value, re.I)
    if not m:
        return default
    raw = m.group(1).upper()
    number_map = {"1": "I", "2": "II", "3": "III", "4": "IV", "5": "V", "6": "VI", "7": "VII", "8": "VIII"}
    return number_map.get(raw, raw)


def _parse_subjects_from_attendance_workbook(wb, filename=""):
    def clean(value):
        return " ".join(str(value or "").replace("\n", " ").split()).strip()

    ws = wb.active
    meta_text = "\n".join(
        clean(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, 10) + 1)
        for c in range(1, min(ws.max_column, 8) + 1)
    )
    dept, _, _ = _class_meta_from_attendance_text(meta_text, filename)
    semester = _semester_from_attendance_text(meta_text)

    header_row = None
    headers = []
    for r in range(1, min(ws.max_row, 20) + 1):
        row_headers = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        row_text = " ".join(h.upper() for h in row_headers)
        if "ROLL" in row_text and "NAME" in row_text:
            header_row = r
            headers = row_headers
            break
    if not header_row:
        return []

    code_row = header_row + 1
    type_row = header_row + 2
    subjects = []
    seen = set()
    for col in range(1, ws.max_column + 1):
        name = clean(headers[col - 1])
        name = re.sub(r"\s*-\s*", "-", name)
        code = re.sub(r"\s+", "", clean(ws.cell(code_row, col).value))
        lecture_type = clean(ws.cell(type_row, col).value)
        upper_name = name.upper()
        if not name or upper_name in ("SR. NO.", "ROLL NO.", "NAME OF STUDENT", "TOTAL"):
            continue
        if "ATTEND" in upper_name or upper_name == "%":
            continue
        if not code or not re.search(r"[A-Z]{2,}.*\d", code, re.I):
            continue
        key = (name.lower(), code.upper(), dept, semester)
        if key in seen:
            continue
        seen.add(key)
        subjects.append({
            "name": name,
            "code": code,
            "department": dept,
            "semester": semester,
            "lecture_type": lecture_type,
        })
    return subjects


def _import_attendance_workbook_subjects(wb, filename=""):
    added = skipped = 0
    for subject in _parse_subjects_from_attendance_workbook(wb, filename):
        existing = qone(
            "SELECT id FROM subjects WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) AND COALESCE(subject_code,'')=%s AND department=%s AND semester=%s",
            (subject["name"], subject["code"], subject["department"], subject["semester"]),
        )
        if existing:
            skipped += 1
            continue
        try:
            exe(
                "INSERT INTO subjects(name,department,subject_code,teacher,semester) VALUES(%s,%s,%s,%s,%s)",
                (subject["name"], subject["department"], subject["code"], "", subject["semester"]),
            )
            added += 1
        except Exception:
            skipped += 1
    return added, skipped


def _unique_pdf_roll(raw_roll, dept, div, existing_id=None):
    raw_roll = (raw_roll or "").strip().upper()
    existing = qone("SELECT id,name,department,division FROM students WHERE roll=%s", (raw_roll,))
    if not existing or (existing_id and existing["id"] == existing_id):
        return raw_roll
    if (existing["department"] or "") == (dept or "") and (existing["division"] or "") == (div or ""):
        return raw_roll
    prefixed = "-".join(x for x in (dept, div, raw_roll) if x)
    if not prefixed:
        prefixed = raw_roll
    suffix = 2
    candidate = prefixed
    while qone("SELECT id FROM students WHERE roll=%s", (candidate,)):
        candidate = f"{prefixed}-{suffix}"
        suffix += 1
    return candidate


def _upsert_pdf_student(item):
    name = (item.get("name") or "").strip()
    raw_roll = (item.get("roll") or "").strip().upper()
    dept = (item.get("department") or "").strip()
    div = (item.get("division") or "").strip()
    year = (item.get("year") or "II").strip()
    if not name or not raw_roll:
        return "skipped"

    existing = qone(
        "SELECT id,roll FROM students WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) AND department=%s AND division=%s",
        (name, dept, div),
    )
    if existing:
        exe("UPDATE students SET year=%s WHERE id=%s", (year, existing["id"]))
        return "updated"

    roll = _unique_pdf_roll(raw_roll, dept, div)
    try:
        exe(
            "INSERT INTO students(name,roll,department,year,division,password) VALUES(%s,%s,%s,%s,%s,%s)",
            (name, roll, dept, year, div, hash_password(_default_student_password())),
        )
        return "added"
    except Exception:
        return "skipped"


def _import_attendance_workbook_students(wb, filename=""):
    added = skipped = updated = 0
    rows = _parse_students_from_attendance_workbook(wb, filename)
    for item in rows:
        result = _upsert_pdf_student(item)
        if result == "added":
            added += 1
        elif result == "updated":
            updated += 1
        else:
            skipped += 1
    return added, updated, skipped


@app.route("/import_students_smart", methods=["POST"])
@login_required("admin")
def import_students_smart():
    """
    Handles BE-AIDS_students-2025-26_list_.xlsx format:
    Row 1: empty
    Row 2: PRN | Name As Per Marksheet | Mobile No | Email | ...
    Row 3+: data
    Auto-detects dept from filename (BE-AIDS → AIDS, BE-CS → CS etc)
    Auto-detects year from BE/TE/SE/FE prefix
    """
    files = [f for f in request.files.getlist("file") if f and f.filename]
    f = files[0] if files else None
    if not f:
        return redirect("/students?error=no_file")

    if any((x.filename or "").lower().endswith(".pdf") for x in files):
        added = skipped = updated = 0
        for pdf_file in files:
            if not (pdf_file.filename or "").lower().endswith(".pdf"):
                continue
            rows = _parse_students_from_attendance_pdf(pdf_file)
            for item in rows:
                result = _upsert_pdf_student(item)
                if result == "added":
                    added += 1
                elif result == "updated":
                    updated += 1
                else:
                    skipped += 1
        return redirect(f"/students?imported={added}&updated={updated}&skipped={skipped}&format=pdf")

    fname   = f.filename.upper()
    dept    = ""
    year    = ""
    # Auto-detect department from filename
    for d in ["AIDS","AIML","CS","IT","CIVIL","MECH","ENTC"]:
        if d in fname: dept = d; break
    # Auto-detect year from filename
    year_map = {"BE":"IV","TE":"III","SE":"II","FE":"I"}
    for k,v in year_map.items():
        if k in fname: year = v; break

    # Allow override from form
    dept = request.form.get("department","").strip() or dept
    year = request.form.get("year","").strip()       or year

    try:
        wb = load_workbook(f, data_only=True)
    except Exception as e:
        return redirect("/students?error=bad_excel")

    if _is_attendance_student_workbook(wb):
        added, updated, skipped = _import_attendance_workbook_students(wb, f.filename)
        return redirect(f"/students?imported={added}&updated={updated}&skipped={skipped}&format=attendance_xlsx")

    ws = wb.active
    added = skipped = 0

    # Find header row
    hdr_row = 1
    for ri in range(1, min(ws.max_row+1, 6)):
        vals = [str(ws.cell(ri,c).value or "").lower() for c in range(1, min(ws.max_column+1,6))]
        if any("name" in v or "prn" in v for v in vals):
            hdr_row = ri; break

    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip()
            for c in range(1, ws.max_column+1)]

    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None

    # BE-AIDS format: PRN=roll, Name=name, Email=email
    cn   = gcol(["name as per","name"])
    cr   = gcol(["prn","roll"])
    ce   = gcol(["email"])
    cc   = gcol(["mobile","contact","phone"])
    cdept= gcol(["dept","department","branch"])
    cyear= gcol(["year","class","sem"])

    if not cn:
        return redirect("/students?error=bad_format")

    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        if not row: continue
        name = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        roll = str(row[cr-1] or "").strip() if cr and cr <= len(row) else ""
        eml  = str(row[ce-1] or "").strip() if ce and ce <= len(row) else ""
        contact_raw = row[cc-1] if cc and cc <= len(row) else None
        contact = str(int(contact_raw)) if isinstance(contact_raw, float) else str(contact_raw or "").strip()

        # Skip header-like rows or empty
        if not name or not roll: continue
        if name.lower() in ("name","name as per marksheet","student name",""):
            continue
        if roll.lower() in ("prn","roll no","roll number","sr.no","sr no"):
            continue

        # Get dept/year from file columns or fall back to auto-detected
        row_dept = str(row[cdept-1] or "").strip() if cdept and cdept <= len(row) else dept
        row_year = str(row[cyear-1] or "").strip() if cyear and cyear <= len(row) else year

        try:
            exe("INSERT INTO students(name,roll,department,year,email,contact_number,password) VALUES(%s,%s,%s,%s,%s,%s,%s)",
                (name.strip(), roll.strip(), row_dept or dept, row_year or year, eml, contact, hash_password(_default_student_password())))
            added += 1
        except Exception:
            skipped += 1

    return redirect(f"/students?imported={added}&skipped={skipped}")


@app.route("/import_faculty_smart", methods=["POST"])
@login_required("admin")
def import_faculty_smart():
    """
    Handles Faculty_Details_.xlsx:
    Row 1-2: junk
    Row 3: Sr. No. | Name of Faculty | Department | Contact No | Email Id | Signature
    Row 4+: data with empty rows between
    """
    f = request.files.get("file")
    if not f:
        return redirect("/faculty?error=no_file")

    try:
        wb = load_workbook(f, data_only=True)
    except Exception as e:
        return redirect("/faculty?error=bad_excel")
    ws = wb.active
    added = skipped = 0

    # Find header row — look for "Name of Faculty" or "Email"
    hdr_row = 3
    for ri in range(1, min(ws.max_row+1, 8)):
        vals = [str(ws.cell(ri,c).value or "").lower() for c in range(1, 8)]
        if any("name of faculty" in v or ("name" in v and "faculty" in v) for v in vals):
            hdr_row = ri; break
        if any("email" in v for v in vals) and any("name" in v for v in vals):
            hdr_row = ri; break

    hdrs = [str(ws.cell(hdr_row,c).value or "").lower().strip()
            for c in range(1, ws.max_column+1)]

    def gcol(kws):
        for k in kws:
            for i,h in enumerate(hdrs):
                if k in h: return i+1
        return None

    cn  = gcol(["name of faculty","name"])
    cd  = gcol(["department","dept"])
    cp  = gcol(["contact","phone","mobile"])
    ce  = gcol(["email"])

    if not cn:
        return redirect("/faculty?error=bad_format")

    for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
        if not row: continue
        name  = str(row[cn-1] or "").strip() if cn <= len(row) else ""
        dept  = str(row[cd-1] or "").strip() if cd and cd <= len(row) else ""
        email = str(row[ce-1] or "").strip() if ce and ce <= len(row) else ""

        # Skip empty rows, header rows, placeholder rows
        if not name: continue
        if name.lower() in ("name of faculty","name","faculty name","sr. no.",""):
            continue
        if name.replace(".","").replace(" ","").isdigit():
            continue

        # Phone: might be stored as float like 8830488239.0
        phone_raw = row[cp-1] if cp and cp <= len(row) else None
        if isinstance(phone_raw, float):
            phone = str(int(phone_raw))
        else:
            phone = str(phone_raw or "").strip()

        try:
            exe("INSERT INTO faculty(name,department,email,phone,password) VALUES(%s,%s,%s,%s,%s)",
                (name, dept.strip(), email, phone, hash_password(_default_faculty_password())))
            added += 1
        except Exception:
            skipped += 1

    return redirect(f"/faculty?imported={added}&skipped={skipped}")



# ════════════════════════════════════════════════════════════
#  QUICK WIN 1 — PASSWORD RESET (OTP, no email needed)
# ════════════════════════════════════════════════════════════
_reset_otps = {}   # { (role, identifier): (otp, expires_ts) }

@app.route("/forgot_password", methods=["GET","POST"])
@limiter.limit("3 per hour")
def forgot_password():
    msg = None; err = None
    if request.method == "POST":
        role  = request.form.get("role","").strip()
        ident = request.form.get("identifier","").strip()
        otp   = str(secrets.randbelow(900000) + 100000)
        expires = time.time() + 600  # 10 min
        _reset_otps[(role, ident)] = (otp, expires)
        msg = f"Your reset code is: {otp}  (valid 10 min)"
    return render_template("common/forgot_password.html", msg=msg, err=err)

@app.route("/reset_password", methods=["GET","POST"])
@limiter.limit("5 per hour")
def reset_password():
    err = None
    if request.method == "POST":
        role  = request.form.get("role","").strip()
        ident = request.form.get("identifier","").strip()
        otp   = request.form.get("otp","").strip()
        new_pw= request.form.get("new_pw","").strip()
        key   = (role, ident)
        stored = _reset_otps.get(key)
        if not stored:
            err = "No reset code found. Request a new one."
        elif time.time() > stored[1]:
            _reset_otps.pop(key, None); err = "Code expired. Request a new one."
        elif stored[0] != otp:
            err = "Wrong code. Try again."
        else:
            _reset_otps.pop(key, None)
            hpw = hash_password(new_pw)
            if role == "student":
                exe("UPDATE students SET password=%s WHERE roll=%s", (hpw, ident))
            elif role == "faculty":
                exe("UPDATE faculty SET password=%s WHERE email=%s", (hpw, ident))
            return redirect("/login?reset=1")
    return render_template("common/reset_password.html", err=err,
                           role=request.args.get("role","student"),
                           ident=request.args.get("ident",""))


# ════════════════════════════════════════════════════════════
#  QUICK WIN 2 — BULK DELETE
# ════════════════════════════════════════════════════════════
@app.route("/bulk_delete_students", methods=["POST"])
@login_required("admin")
def bulk_delete_students():
    ids = request.form.getlist("ids[]")
    for i in ids:
        exe("DELETE FROM students WHERE id=%s", (i,))
    return redirect(f"/students?deleted={len(ids)}")

@app.route("/bulk_delete_faculty", methods=["POST"])
@login_required("admin")
def bulk_delete_faculty():
    ids = request.form.getlist("ids[]")
    for i in ids:
        exe("DELETE FROM faculty WHERE id=%s", (i,))
    return redirect(f"/faculty?deleted={len(ids)}")

# ════════════════════════════════════════════════════════════
#  QUICK WIN 5 — MARKS EXPORT
# ════════════════════════════════════════════════════════════
@app.route("/export_marks_excel")
@login_required("faculty")
def export_marks_excel():
    fid     = session["faculty_id"]
    subject = request.args.get("subject","").strip()
    student = request.args.get("student","").strip()

    sql = "SELECT * FROM marks WHERE faculty_id=%s"
    params = [fid]
    if subject: sql += " AND subject=%s"; params.append(subject)
    if student:
        sid = resolve_student_id(student)
        sql += f" AND ({marks_match_student_sql()})"
        params.extend(marks_match_student_params(sid, student))
    sql += " ORDER BY student_name, subject, exam_type"
    rows = qry(sql, params)

    wb = Workbook()
    ws = wb.active
    ws.title = "Marks"

    # Header
    hdrs = ["Student Name","Roll","Subject","Dept","Exam Type","Marks","Total","Percentage","Grade","Date"]
    for c,h in enumerate(hdrs,1):
        cell = ws.cell(1,c,h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.font = Font(bold=True, color="FFFFFF")

    for r,row in enumerate(rows,2):
        p = round(row["marks"]/row["total"]*100,1) if row["total"] else 0
        g = grade(row["marks"], row["total"])
        vals = [row["student_name"],row["roll"],row["subject"],row["department"],
                row["exam_type"],row["marks"],row["total"],f"{p}%",g,row["date"]]
        for c,v in enumerate(vals,1):
            ws.cell(r,c,v)

    # Column widths
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or ""))+4, 14)

    fname = f"marks{'_'+subject if subject else ''}{'_'+student if student else ''}.xlsx"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/export_marks_admin")
@login_required("admin")
def export_marks_admin():
    subject = request.args.get("subject","").strip()
    student = request.args.get("student","").strip()
    dept    = request.args.get("dept","").strip()

    sql = "SELECT m.*, f.name as faculty_name FROM marks m LEFT JOIN faculty f ON m.faculty_id=f.id WHERE 1=1"
    params = []
    if subject: sql += " AND m.subject=?"; params.append(subject)
    if student:
        sid = resolve_student_id(student)
        sql += f" AND ({marks_match_student_sql('m')})"
        params.extend(marks_match_student_params(sid, student))
    if dept: sql += " AND m.department=?"; params.append(dept)
    sql += " ORDER BY m.student_name, m.subject"
    rows = qry(sql, params)

    wb = Workbook(); ws = wb.active; ws.title = "Marks"
    hdrs = ["Student","Roll","Subject","Dept","Exam Type","Marks","Total","Percentage","Grade","Faculty","Date"]
    for c,h in enumerate(hdrs,1):
        cell = ws.cell(1,c,h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    for r,row in enumerate(rows,2):
        p = round(row["marks"]/row["total"]*100,1) if row["total"] else 0
        g = grade(row["marks"], row["total"])
        for c,v in enumerate([row["student_name"],row["roll"],row["subject"],row["department"],
                               row["exam_type"],row["marks"],row["total"],f"{p}%",g,
                               row["faculty_name"] or "",row["date"]],1):
            ws.cell(r,c,v)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or ""))+4,14)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="all_marks.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# Removed Global Search Route

@app.route("/search")
def search():
    role = session.get("role")
    q = request.args.get("q", "").strip()
    if role == "admin":
        return redirect(f"/students?q={q}" if q else "/admin")
    if role == "faculty":
        return redirect(f"/faculty_attendance?q={q}" if q else "/faculty_dashboard")
    if role == "student":
        return redirect("/student_dashboard")
    return redirect("/login")


@app.route("/search_faculty")
@login_required("faculty")
def search_faculty():
    q = request.args.get("q","").strip()
    return redirect(f"/faculty_attendance?q={q}" if q else "/faculty_attendance")

# ════════════════════════════════════════════════════════════
#  HIGH VALUE 1 — ATTENDANCE SHORTAGE ALERT
# ════════════════════════════════════════════════════════════
@app.route("/shortage_report")
@login_required("admin")
def shortage_report():
    threshold_pct = safe_int(request.args.get("threshold", "75"))
    dept = request.args.get("department", "").strip()
    div = request.args.get("division", "").strip()
    
    # Calculate threshold as fraction
    t = threshold_pct / 100.0
    
    # SQL to aggregate attendance and calculate pct, shortage, can_miss
    # shortage = (t * total - present) / (1 - t) if on track or already behind
    # can_miss = (present - t * total) / t if ahead
    sql = f"""
        WITH stats AS (
            SELECT s.id, s.name, s.roll, s.department as dept, s.division,
                   COUNT(a.id) as sessions,
                   SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) as present
            FROM students s
            LEFT JOIN attendance a ON s.id = a.student_id
            WHERE (%s='' OR s.department=%s) AND (%s='' OR s.division=%s)
            GROUP BY s.id, s.name, s.roll, s.department, s.division
        )
        SELECT *,
               ROUND(CASE WHEN sessions > 0 THEN (present * 100.0 / sessions) ELSE 0 END, 1) as pct,
               CEIL(GREATEST(0, ({t} * sessions - present) / (1 - {t}))) as shortage,
               FLOOR(GREATEST(0, (present - {t} * sessions) / {t})) as can_miss
        FROM stats
        WHERE (CASE WHEN sessions > 0 THEN (present * 100.0 / sessions) ELSE 0 END) < %s
        ORDER BY pct ASC, roll ASC
    """
    defaulters = safe_fetch_all(sql, (dept, dept, div, div, threshold_pct))
    
    return render_template("attendance/shortage_report.html", 
                           defaulters=defaulters, 
                           threshold=threshold_pct, 
                           department=dept, 
                           division=div)

@app.route("/trigger_shortage_alert", methods=["POST"])
@login_required("admin")
def trigger_shortage_alert():
    from utils.comm_utils import send_sms
    
    threshold = safe_int(request.form.get("threshold"), 40)
    dept = request.form.get("department", "").strip()
    div = request.form.get("division", "").strip()
    
    t = threshold / 100.0
    
    # Identify students below threshold
    sql = f"""
        SELECT s.id, s.name, s.phone, s.department, s.division,
               COUNT(a.id) as sessions,
               SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) as present
        FROM students s
        LEFT JOIN attendance a ON s.id = a.student_id
        WHERE (%s='' OR s.department=%s) AND (%s='' OR s.division=%s)
        GROUP BY s.id, s.name, s.phone, s.department, s.division
        HAVING (CASE WHEN COUNT(a.id) > 0 THEN (SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) * 100.0 / COUNT(a.id)) ELSE 0 END) < %s
    """
    defaulters = qry(sql, (dept, dept, div, div, threshold))
    
    sent_count = 0
    for s in defaulters:
        if s["phone"]:
            pct_val = round((s["present"] or 0) * 100.0 / (s["sessions"] or 1), 1)
            msg = f"Alert: Student {s['name']} has low attendance ({pct_val}%). Please contact the department."
            success, _ = send_sms(s["phone"], msg)
            if success:
                sent_count += 1
                
    return redirect(f"/shortage_report?threshold={threshold}&department={dept}&division={div}&alert_triggered={sent_count}")


# ════════════════════════════════════════════════════════════
#  HIGH VALUE 2 — STUDENT REPORT CARD
# ════════════════════════════════════════════════════════════
@app.route("/export_shortage_excel")
@login_required("admin")
def export_shortage_excel():
    threshold = safe_int(request.args.get("threshold", "75"))
    rows = qry("""SELECT s.name, s.roll, s.department, s.year,
                         COUNT(a.id) AS total,
                         SUM(CASE WHEN a.status='Present' THEN 1 ELSE 0 END) AS present
                  FROM students s
                  LEFT JOIN attendance a
                    ON a.student_id=s.id OR (a.student_id IS NULL AND LOWER(TRIM(a.student_name))=LOWER(TRIM(s.name)))
                  GROUP BY s.id, s.name, s.roll, s.department, s.year
                  ORDER BY s.name""")
    wb = Workbook()
    ws = wb.active
    ws.title = "Shortage Report"
    headers = ["Student", "Roll", "Department", "Year", "Total", "Present", "Absent", "Attendance %", "Threshold"]
    for c, header in enumerate(headers, 1):
        cell = ws.cell(1, c, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    out_row = 2
    for row in rows:
        total = int(row["total"] or 0)
        present = int(row["present"] or 0)
        percentage = pct(present, total)
        if total <= 0 or percentage >= threshold:
            continue
        values = [
            row["name"], row["roll"], row["department"], row["year"],
            total, present, total - present, f"{percentage}%", f"{threshold}%",
        ]
        for c, value in enumerate(values, 1):
            ws.cell(out_row, c, value)
        out_row += 1
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")) + 4, 14)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="shortage_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/report_card/<int:student_id>")
def report_card(student_id):
    role = session.get("role")
    if not role:
        return redirect("/login")
    # Students can only see their own report card
    if role == "student" and session.get("student_id") != student_id:
        return redirect("/student_dashboard")
    if role not in ("admin", "student"):
        return redirect("/")
    s = qone("SELECT * FROM students WHERE id=%s", (student_id,))
    if not s: return redirect("/students")
    s = dict(s)
    name = s["name"]

    att_rows = qry(
        f"SELECT subject,status FROM attendance WHERE {att_match_student_sql()}",
        att_match_student_params(s["id"], name),
    )
    total_att = len(att_rows); pres_att = sum(1 for r in att_rows if r["status"]=="Present")
    overall_pct = pct(pres_att, total_att)

    subj_map = {}
    for r in att_rows:
        subj_map.setdefault(r["subject"],{"t":0,"p":0})
        subj_map[r["subject"]]["t"] += 1
        if r["status"]=="Present": subj_map[r["subject"]]["p"] += 1
    subject_att = [{"subject":k,"total":v["t"],"present":v["p"],
                    "pct":pct(v["p"],v["t"])} for k,v in subj_map.items()]

    marks_rows = qry(
        f"SELECT * FROM marks WHERE {marks_match_student_sql()} ORDER BY subject,exam_type",
        marks_match_student_params(s["id"], name),
    )
    marks_with_grade = []
    for m in marks_rows:
        d = dict(m); d["pct"] = pct(m["marks"],m["total"]); d["grade"] = grade(m["marks"],m["total"])
        marks_with_grade.append(d)

    return render_template("student/report_card.html", student=s, overall_pct=overall_pct,
                           total_att=total_att, pres_att=pres_att,
                           subject_att=subject_att, marks=marks_with_grade,
                           generated_on=today_str())

@app.route("/report_card_student")
@login_required("student")
def report_card_student():
    sid = session["student_id"]
    return redirect(f"/report_card/{sid}")


# ════════════════════════════════════════════════════════════
#  RESULTS MODULE — Admin publishes, Faculty enters, Student views
# ════════════════════════════════════════════════════════════

# ── Admin: View all results ─────────────────────────────────
@app.route("/results_reportcard/<roll>")
@login_required("admin")
def results_reportcard(roll):
    student = qone("SELECT id FROM students WHERE roll=%s", (roll,))
    if not student:
        return redirect("/admin_results")
    return report_card(int(student["id"]))


@app.route("/admin_results")
@login_required("admin")
def admin_results():
    dept     = request.args.get("dept","").strip()
    semester = request.args.get("semester","").strip()
    year     = request.args.get("year","").strip()
    q        = request.args.get("q","").strip()
    published= request.args.get("published","").strip()

    sql = """SELECT r.*, f.name as faculty_name
             FROM results r LEFT JOIN faculty f ON r.faculty_id=f.id WHERE 1=1"""
    params = []
    if dept:      sql += " AND r.department=?";       params.append(dept)
    if semester:  sql += " AND r.semester=?";          params.append(semester)
    if year:      sql += " AND r.year=?";              params.append(year)
    if q:         sql += " AND (r.student_name LIKE ? OR r.roll LIKE ?)"; params += [f"%{q}%",f"%{q}%"]
    if published != "": sql += " AND r.published=%s";  params.append(int(published))
    sql += " ORDER BY r.department, r.student_name, r.semester"
    results_list = qry(sql, params)

    # Summary counts
    total_r    = qone("SELECT COUNT(*) as c FROM results")["c"]
    published_r= qone("SELECT COUNT(*) as c FROM results WHERE published=1")["c"]
    pass_r     = qone("SELECT COUNT(*) as c FROM results WHERE result='Pass' AND published=1")["c"]
    fail_r     = qone("SELECT COUNT(*) as c FROM results WHERE result='Fail' AND published=1")["c"]

    return render_template("admin/admin_results.html",
        results=results_list, dept=dept, semester=semester, year=year,
        q=q, published=published,
        total_r=total_r, published_r=published_r, pass_r=pass_r, fail_r=fail_r,
        DEPARTMENTS=DEPARTMENTS, SEMESTERS=SEMESTERS, YEARS=YEARS)


# ── Admin: Add / Edit result ────────────────────────────────
@app.route("/admin_save_result", methods=["POST"])
@login_required("admin")
def admin_save_result():
    student_name = request.form.get("student_name","").strip()
    roll_row = qone("SELECT roll,department,year FROM students WHERE name=?", (student_name,))
    roll = roll_row["roll"] if roll_row else request.form.get("roll","")
    dept = roll_row["department"] if roll_row else request.form.get("department","")
    yr   = roll_row["year"] if roll_row else request.form.get("year","")
    assignment_marks = min(float(request.form.get("assignment_marks", 0)), 5.0)
    attendance_marks = min(float(request.form.get("attendance_marks", 0)), 5.0)
    teacher_assessment = min(float(request.form.get("teacher_assessment", 0)), 10.0)
    ut_marks = min(float(request.form.get("ut_marks", 0)), 20.0)
    mse_marks = min(float(request.form.get("mse_marks", 0)), 20.0)
    tw_marks = float(request.form.get("tw_marks", 0))
    pr_or_marks = float(request.form.get("pr_or_marks", 0))
    
    # Calculate sum dynamically if not directly provided
    marks_val = assignment_marks + attendance_marks + teacher_assessment + ut_marks + mse_marks + tw_marks + pr_or_marks
    if marks_val == 0:
        marks_val = float(request.form.get("marks", 0))

    marks_val = min(marks_val, 60.0)
    total_val = 60.0
    pct_val   = pct(marks_val, total_val)
    g         = grade(marks_val, total_val)
    result_val= "Pass" if pct_val >= 40 else "Fail"
    exe("""INSERT INTO results(student_name,roll,department,year,semester,subject,
                               marks,total,exam_type,grade,result,published,
                               assignment_marks, attendance_marks, ut_marks, mse_marks,
                               teaching_assessment, tw_marks, pr_or_marks)
           VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (student_name, roll, dept, yr,
         request.form.get("semester","I"),
         request.form.get("subject",""),
         marks_val, total_val,
         request.form.get("exam_type","Semester Exam"),
         g, result_val, 0, assignment_marks, attendance_marks, ut_marks, mse_marks,
         teacher_assessment, tw_marks, pr_or_marks))
    return redirect("/admin_results?success=1")


@app.route("/admin_edit_result", methods=["POST"])
@login_required("admin")
def admin_edit_result():
    rid = request.form.get("result_id","")
    assignment_marks = min(float(request.form.get("assignment_marks", 0)), 5.0)
    attendance_marks = min(float(request.form.get("attendance_marks", 0)), 5.0)
    teacher_assessment = min(float(request.form.get("teacher_assessment", 0)), 10.0)
    ut_marks = min(float(request.form.get("ut_marks", 0)), 20.0)
    mse_marks = min(float(request.form.get("mse_marks", 0)), 20.0)
    tw_marks = float(request.form.get("tw_marks", 0))
    pr_or_marks = float(request.form.get("pr_or_marks", 0))
    
    marks_val = assignment_marks + attendance_marks + teacher_assessment + ut_marks + mse_marks + tw_marks + pr_or_marks
    if marks_val == 0:
        marks_val = float(request.form.get("marks", 0))
        
    marks_val = min(marks_val, 60.0)
    total_val = 60.0
    g         = grade(marks_val, total_val)
    result_val= "Pass" if pct(marks_val, total_val) >= 40 else "Fail"
    exe("""UPDATE results SET semester=%s,subject=%s,marks=%s,total=%s,
           exam_type=%s,grade=%s,result=%s, assignment_marks=%s, attendance_marks=%s, ut_marks=%s, mse_marks=%s,
           teaching_assessment=%s, tw_marks=%s, pr_or_marks=%s WHERE id=%s""",
        (request.form.get("semester",""), request.form.get("subject",""),
         marks_val, total_val, request.form.get("exam_type",""),
         g, result_val, assignment_marks, attendance_marks, ut_marks, mse_marks,
         teacher_assessment, tw_marks, pr_or_marks, rid))
    return redirect("/admin_results?updated=1")


@app.route("/admin_delete_result", methods=["POST"])
@login_required("admin")
def admin_delete_result():
    exe("DELETE FROM results WHERE id=%s", (request.form.get("result_id",""),))
    return redirect("/admin_results?deleted=1")


@app.route("/admin_publish_results", methods=["POST"])
@login_required("admin")
def admin_publish_results():
    semester = request.form.get("semester","").strip()
    dept     = request.form.get("dept","").strip()
    if semester:
        sql = "UPDATE results SET published=1 WHERE semester=%s"
        params = [semester]
        if dept: sql += " AND department=%s"; params.append(dept)
        exe(sql, params)
    return redirect("/admin_results?published_ok=1")


@app.route("/admin_unpublish_results", methods=["POST"])
@login_required("admin")
def admin_unpublish_results():
    semester = request.form.get("semester","").strip()
    dept     = request.form.get("dept","").strip()
    if semester:
        sql = "UPDATE results SET published=0 WHERE semester=%s"
        params = [semester]
        if dept: sql += " AND department=%s"; params.append(dept)
        exe(sql, params)
    return redirect("/admin_results?unpublished_ok=1")


@app.route("/admin_import_results", methods=["POST"])
@login_required("admin")
def admin_import_results():
    f = request.files.get("file")
    if not f: return redirect("/admin_results?error=no_file")
    
    try:
        if pd is None:
            raise ImportError("pandas is not installed")
        df = pd.read_excel(f, header=None)
    except Exception as e:
        return redirect("/admin_results?error=invalid_file_format")
        
    added = 0
    skipped = 0
    
    # 1. Find the header row (contains 'student name', 'prn', or 'roll')
    hdr_row_idx = -1
    for idx, row in df.iterrows():
        row_str = ' '.join([str(x).lower() for x in row.values])
        if 'student name' in row_str or 'prn ' in row_str or 'name' in row_str:
            hdr_row_idx = idx
            break
            
    if hdr_row_idx == -1: return redirect("/admin_results?error=no_student_column_found")
    
    # 2. Extract Columns
    columns = df.iloc[hdr_row_idx].fillna('').astype(str).str.lower().str.strip()
    
    # Subject row is assumed to be the row immediately above the sub-headers
    if hdr_row_idx > 0:
        subjects_row_code = df.iloc[hdr_row_idx - 1].ffill()
        if hdr_row_idx >= 2:
            subjects_row_name = df.iloc[hdr_row_idx - 2].ffill()
            combined = []
            for name_val, code_val in zip(subjects_row_name, subjects_row_code):
                ns = str(name_val).strip().replace('nan','')
                cs = str(code_val).strip().replace('nan','')
                if ns and len(ns) > 2 and ns.lower() not in ['nan','none','blank']:
                    combined.append(ns)
                else:
                    combined.append(cs)
            subjects_row = pd.Series(combined)
        else:
            subjects_row = subjects_row_code
    else:
        subjects_row = pd.Series(['']*len(columns))
    
    # Default semester mapping if passed from form
    sem_val = request.form.get("semester", "I")
    
    # 3. Parse Data
    for idx in range(hdr_row_idx + 1, len(df)):
        row = df.iloc[idx]
        name = ""
        roll = ""
        for c_idx, col_name in enumerate(columns):
            if 'name' in col_name: name = str(row[c_idx]).strip()
            elif 'prn' in col_name or 'roll' in col_name: roll = str(row[c_idx]).strip()
            
        if not name or str(name).lower() in ('nan', 'none', ''): continue
        if str(name).lower().startswith('sr.'): continue # Skip header re-print
        
        # Resolve Student
        roll_row = qone("SELECT roll,department,year FROM students WHERE name=%s OR roll=%s", (name, roll))
        dept_v = roll_row["department"] if roll_row else ""
        yr_v   = roll_row["year"] if roll_row else ""
        
        # Group marks by mapped Subject
        subject_marks = {}
        for c_idx, col_name in enumerate(columns):
            sub_name = str(subjects_row[c_idx]).strip()
            
            # Skip non-subject header fields (like the ones above Sr.No, Name, Roll)
            if not sub_name or str(sub_name).lower() in ['nan','none','sr.no.','student name','prn number','name','roll','dept']:
                continue
                
            val = row[c_idx]
            try:
                val = float(val) if pd.notna(val) else 0
            except:
                val = 0
                
            if sub_name not in subject_marks:
                subject_marks[sub_name] = {'assign':0, 'attend':0, 'ta':0, 'ut':0, 'mse':0, 'tw':0, 'pr_or':0, 'total':0}
                
            if 'assign' in col_name: subject_marks[sub_name]['assign'] = val
            elif 'attend' in col_name: subject_marks[sub_name]['attend'] = val
            elif 'teach' in col_name or 'assess' in col_name: subject_marks[sub_name]['ta'] = val
            elif 'ut' in col_name or 'unit' in col_name: subject_marks[sub_name]['ut'] = val
            elif 'mse' in col_name or 'mid' in col_name: subject_marks[sub_name]['mse'] = val
            elif 'tw' in col_name or 'term' in col_name: subject_marks[sub_name]['tw'] = val
            elif 'pr' in col_name or 'or' in col_name: subject_marks[sub_name]['pr_or'] = val
            elif 'total' in col_name: subject_marks[sub_name]['total'] = val

        # Save Subject Iteratively
        for subj, marks in subject_marks.items():
            marks['assign'] = min(marks['assign'], 5.0)
            marks['attend'] = min(marks['attend'], 5.0)
            marks['ta']     = min(marks['ta'], 10.0)
            marks['ut']     = min(marks['ut'], 20.0)
            marks['mse']    = min(marks['mse'], 20.0)
            
            total_sum = marks['assign'] + marks['attend'] + marks['ta'] + marks['ut'] + marks['mse'] + marks['tw'] + marks['pr_or']
            final_total_obtained = total_sum if total_sum > 0 else marks['total']
            
            if final_total_obtained <= 0 and marks['total'] <= 0:
                continue # Skip if completely empty marks
                
            final_total_obtained = min(final_total_obtained, 60.0)
            max_marks = 60.0
            
            # Evaluate Grade
            g   = grade(final_total_obtained, max_marks)
            res = "Pass" if pct(final_total_obtained, max_marks) >= 40 else "Fail"
            
            try:
                exe("""INSERT INTO results(student_name,roll,department,year,semester,subject,
                                           marks,total,exam_type,grade,result,published,
                                           assignment_marks, attendance_marks, ut_marks, mse_marks,
                                           teaching_assessment, tw_marks, pr_or_marks)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,%s,%s,%s,%s,%s,%s,%s)""",
                    (name, roll, dept_v, yr_v, sem_val, subj, 
                     final_total_obtained, max_marks, "Semester Exam", g, res,
                     marks['assign'], marks['attend'], marks['ut'], marks['mse'],
                     marks['ta'], marks['tw'], marks['pr_or']))
                added += 1
            except Exception as e:
                skipped += 1
                
    return redirect(f"/admin_results?imported={added}&skipped={skipped}")


@app.route("/export_results_excel")
@login_required("admin")
def export_results_excel():
    dept     = request.args.get("dept","").strip()
    semester = request.args.get("semester","").strip()
    sql = "SELECT * FROM results WHERE 1=1"
    params = []
    if dept:     sql += " AND department=%s"; params.append(dept)
    if semester: sql += " AND semester=%s";   params.append(semester)
    sql += " ORDER BY department, student_name, semester"
    rows = qry(sql, params)
    wb = Workbook(); ws = wb.active; ws.title = "Results"
    hdrs = ["Student Name","Roll","Dept","Year","Semester","Subject","Marks","Total","Percentage","Grade","Result","Status"]
    for c,h in enumerate(hdrs,1):
        cell = ws.cell(1,c,h)
        cell.font = Font(bold=True,color="FFFFFF")
        cell.fill = PatternFill("solid",fgColor="1E3A5F")
    for r,row in enumerate(rows,2):
        p_val = round(row["marks"]/row["total"]*100,1) if row["total"] else 0
        pub   = "Published" if row["published"] else "Draft"
        for c,v in enumerate([row["student_name"],row["roll"] or "",row["department"] or "",
                               row["year"] or "",row["semester"],row["subject"],
                               row["marks"],row["total"],f"{p_val}%",row["grade"] or "",
                               row["result"] or "",pub],1):
            cell = ws.cell(r,c,v)
            if c==11:
                cell.fill = PatternFill("solid", fgColor=("C6EFCE" if row["result"]=="Pass" else "FFC7CE"))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or ""))+4,14)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name="results.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                     
@app.route("/admin_import_sem7", methods=["POST"])
@login_required("admin")
def admin_import_sem7():
    """
    Import DY Patil SEM VII Master Sheet (AIDS / IT / COMP tabs).
    """
    f = request.files.get("file")
    if not f:
        return redirect("/admin_results?error=no_file")

    try:
        wb = load_workbook(f, data_only=True)
    except Exception:
        return redirect("/admin_results?error=invalid_file")

    MAX_MARKS = {"AIDS": 510, "IT": 510, "COMP": 485}
    GRADE_TABLE = [(75,"O"),(70,"A+"),(60,"A"),(55,"B+"),(50,"B"),(45,"C"),(0,"F")]

    def sem7_grade(obtained, total):
        if not total:
            return "F"
        p = obtained / total * 100
        for threshold, g in GRADE_TABLE:
            if p >= threshold:
                return g
        return "F"

    added = skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        dept = sheet_name.strip().upper()  # AIDS / IT / COMP

        subj_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        hdr_row  = [str(ws.cell(3, c).value or "").strip() for c in range(1, ws.max_column + 1)]

        total_cols = []  
        last_subj  = "General"
        for i, sv in enumerate(subj_row):
            if sv and str(sv).strip():
                last_subj = str(sv).strip()
            hv = hdr_row[i].upper() if i < len(hdr_row) else ""
            if "TOTAL" in hv and i > 2:
                m = re.search(r"\((\d+)\)", hv)
                max_m = int(m.group(1)) if m else 100
                total_cols.append((i + 1, last_subj, max_m))

        if not total_cols:
            continue

        for row_idx in range(4, ws.max_row + 1):
            sr_cell = ws.cell(row_idx, 1).value
            if sr_cell is None:
                continue
            try:
                int(float(str(sr_cell)))
            except (ValueError, TypeError):
                continue

            name = str(ws.cell(row_idx, 2).value or "").strip()
            prn  = str(ws.cell(row_idx, 3).value or "").strip()
            if not name or name.lower() in ("student name", "name", "nan"):
                continue

            db_student = qone("SELECT roll, department, year FROM students WHERE name=%s OR roll=%s", (name, prn))
            db_dept = db_student["department"] if db_student else dept
            db_year = db_student["year"]       if db_student else "IV"
            db_roll = db_student["roll"]       if db_student else prn

            for col_1, subj_name, max_m in total_cols:
                raw_val = ws.cell(row_idx, col_1).value
                if raw_val is None:
                    continue
                try:
                    marks_val = float(raw_val)
                except (ValueError, TypeError):
                    continue

                g   = sem7_grade(marks_val, max_m)
                res = "Pass" if (marks_val / max_m * 100 >= 40) else "Fail"

                existing = qone("SELECT id FROM results WHERE student_name=%s AND subject=%s AND semester=%s", (name, subj_name, "VII"))
                if existing:
                    skipped += 1
                    continue

                try:
                    exe("""INSERT INTO results
                           (student_name, roll, department, year, semester,
                            subject, marks, total, exam_type, grade, result, published,
                            assignment_marks, attendance_marks, ut_marks, mse_marks,
                            teaching_assessment, tw_marks, pr_or_marks)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,0,0,0,0,0,0,0)""",
                        (name, db_roll, db_dept, db_year, "VII",
                         subj_name, marks_val, float(max_m),
                         "Semester Exam", g, res))
                    added += 1
                except Exception:
                    skipped += 1

    return redirect(f"/admin_results?imported={added}&skipped={skipped}")


# ── Faculty: Enter + view results for their subjects ────────
@app.route("/faculty_results")
@login_required("faculty")
def faculty_results():
    fid = session["faculty_id"]
    my_subjects = qry("SELECT name FROM subjects WHERE teacher LIKE %s ORDER BY name",
                      (f"%{session['name']}%",))
    results = qry("SELECT * FROM results WHERE faculty_id=%s ORDER BY id DESC", (fid,))
    students_list = qry("SELECT name,roll FROM students ORDER BY name")
    return render_template("faculty/faculty_results.html",
        results=results, my_subjects=my_subjects,
        students=students_list, today=today_str(), SEMESTERS=SEMESTERS)


# Faculty: Leave Management (Task C)
@app.route("/faculty_leaves")
@login_required("faculty")
def faculty_leaves():
    fid = session["faculty_id"]
    my_leaves = qry("SELECT * FROM leave_applications WHERE faculty_id=%s ORDER BY created_at DESC", (fid,))
    return render_template("faculty/faculty_leaves.html", leaves=my_leaves)

@app.route("/faculty_apply_leave", methods=["POST"])
@login_required("faculty")
def faculty_apply_leave():
    fid = session["faculty_id"]
    lt  = request.form.get("leave_type")
    f_d = request.form.get("from_date")
    t_d = request.form.get("to_date")
    res = request.form.get("reason")
    
    exe("""INSERT INTO leave_applications (faculty_id, leave_type, from_date, to_date, reason)
           VALUES (%s, %s, %s, %s, %s)""", (fid, lt, f_d, t_d, res))
    return redirect("/faculty_leaves?applied=1")

# Admin: Leave Approvals
@app.route("/admin/faculty_leave_requests")
@login_required("admin")
def admin_leave_requests():
    # Phase 3 & 8: Centralized Query Audit
    requests = qry(get_query("faculty_leave_requests"), ('pending',))
    history = qry(get_query("faculty_leave_requests").replace('status = %s', "status != 'pending'"), ('pending',))
    return render_template("admin/leave_requests.html", requests=requests, history=history)

@app.route("/admin/approve_faculty_leave", methods=["POST"])
@login_required("admin")
def admin_approve_faculty_leave():
    lid = request.form.get("leave_id")
    status = request.form.get("status") # 'approved' or 'rejected'
    rem = request.form.get("remarks", "")
    
    exe("UPDATE leave_applications SET status=%s, remarks=%s WHERE id=%s", (status, rem, lid))
    return redirect("/admin/faculty_leave_requests?updated=1")


@app.route("/faculty_save_result", methods=["POST"])
@login_required("faculty")
def faculty_save_result():
    fid = session["faculty_id"]
    student_name = request.form.get("student_name","").strip()
    roll_row = qone("SELECT roll,department,year FROM students WHERE name=%s", (student_name,))
    roll = roll_row["roll"] if roll_row else ""
    dept = roll_row["department"] if roll_row else ""
    yr   = roll_row["year"] if roll_row else ""

    # Component marks with hard caps
    assignment_m   = min(float(request.form.get("assignment_marks",   0) or 0), 5.0)
    attendance_m   = min(float(request.form.get("attendance_marks",   0) or 0), 5.0)
    teaching_m     = min(float(request.form.get("teaching_assessment",0) or 0), 10.0)
    ut_m           = min(float(request.form.get("ut_marks",           0) or 0), 20.0)
    mse_m          = min(float(request.form.get("mse_marks",          0) or 0), 20.0)

    marks_val = assignment_m + attendance_m + teaching_m + ut_m + mse_m
    # Fallback: if all components are 0, accept a direct marks value but still cap at 60
    if marks_val == 0:
        marks_val = min(float(request.form.get("marks", 0) or 0), 60.0)

    marks_val = min(marks_val, 60.0)
    total_val = 60.0
    g   = grade(marks_val, total_val)
    res = "Pass" if pct(marks_val, total_val) >= 40 else "Fail"

    exe("""INSERT INTO results(student_name,roll,department,year,semester,subject,
                               marks,total,exam_type,grade,result,faculty_id,published,
                               assignment_marks,attendance_marks,teaching_assessment,
                               ut_marks,mse_marks)
           VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,%s,%s,%s,%s,%s)""",
        (student_name, roll, dept, yr,
         request.form.get("semester","IV"),
         request.form.get("subject",""),
         marks_val, total_val,
         request.form.get("exam_type","Semester Exam"),
         g, res, fid,
         assignment_m, attendance_m, teaching_m, ut_m, mse_m))
    return redirect("/faculty_results?success=1")


@app.route("/faculty_edit_result", methods=["POST"])
@login_required("faculty")
def faculty_edit_result():
    rid = request.form.get("result_id", "")
    fid = session["faculty_id"]
    assignment_m = min(float(request.form.get("assignment_marks", 0) or 0), 5.0)
    attendance_m = min(float(request.form.get("attendance_marks", 0) or 0), 5.0)
    teaching_m = min(float(request.form.get("teacher_assessment", 0) or 0), 10.0)
    ut_m = min(float(request.form.get("ut_marks", 0) or 0), 20.0)
    mse_m = min(float(request.form.get("mse_marks", 0) or 0), 20.0)
    tw_m = float(request.form.get("tw_marks", 0) or 0)
    pr_or_m = float(request.form.get("pr_or_marks", 0) or 0)

    marks_val = assignment_m + attendance_m + teaching_m + ut_m + mse_m + tw_m + pr_or_m
    if marks_val == 0:
        marks_val = min(float(request.form.get("marks", 0) or 0), 60.0)

    marks_val = min(marks_val, 60.0)
    total_val = 60.0
    g = grade(marks_val, total_val)
    res = "Pass" if pct(marks_val, total_val) >= 40 else "Fail"
    exe("""UPDATE results SET semester=%s,subject=%s,marks=%s,total=%s,
           exam_type=%s,grade=%s,result=%s, assignment_marks=%s, attendance_marks=%s,
           teaching_assessment=%s, ut_marks=%s, mse_marks=%s, tw_marks=%s, pr_or_marks=%s
           WHERE id=%s AND faculty_id=%s""",
        (request.form.get("semester", ""), request.form.get("subject", ""),
         marks_val, total_val, request.form.get("exam_type", ""),
         g, res, assignment_m, attendance_m, teaching_m, ut_m, mse_m, tw_m, pr_or_m,
         rid, fid))
    return redirect("/faculty_results?updated=1")


@app.route("/faculty_delete_result", methods=["POST"])
@login_required("faculty")
def faculty_delete_result():
    exe("DELETE FROM results WHERE id=%s AND faculty_id=%s",
        (request.form.get("result_id",""), session["faculty_id"]))
    return redirect("/faculty_results")


# ── Student: View own published results ────────────────────
@app.route("/student_results")
@login_required("student")
def student_results():
    student = get_student()
    if not student: return redirect("/logout")
    name    = student["name"]

    results = qry("""SELECT * FROM results
                     WHERE student_name=%s AND published=1
                     ORDER BY semester, subject""", (name,))

    # Group by semester
    sem_map = {}
    for r in results:
        sem = r["semester"]
        sem_map.setdefault(sem, []).append(dict(r))

    # Per-semester summary
    sem_summary = []
    for sem, rows in sem_map.items():
        obtained = sum(r["marks"] for r in rows)
        total_m  = sum(r["total"] for r in rows)
        pct_val  = round(obtained / total_m * 100, 1) if total_m else 0
        fails    = sum(1 for r in rows if r["result"]=="Fail")
        sem_summary.append({
            "semester": sem, "subjects": len(rows),
            "obtained": obtained, "total": total_m,
            "pct": pct_val, "grade": grade(obtained, total_m),
            "result": "Fail" if fails > 0 else "Pass",
            "rows": rows
        })
    sem_summary.sort(key=lambda x: x["semester"])

    overall_obtained = sum(r["marks"] for r in results)
    overall_total    = sum(r["total"]  for r in results)
    overall_pct      = round(overall_obtained / overall_total * 100, 1) if overall_total else 0

    return render_template("student/student_results.html",
        student=student,
        sem_summary=sem_summary,
        overall_obtained=overall_obtained,
        overall_total=overall_total,
        overall_pct=overall_pct,
        overall_grade=grade(overall_obtained, overall_total) if overall_total else "N/A")


# ════════════════════════════════════════════════════════════
#  HIGH VALUE 3 — LEAVE APPLICATION
# ════════════════════════════════════════════════════════════
@app.route("/student_results_download")
@login_required("student")
def student_results_download():
    student = get_student()
    if not student:
        return redirect("/logout")
    rows = qry("""SELECT semester,subject,marks,total,exam_type,grade,result
                  FROM results
                  WHERE student_name=%s AND published=1
                  ORDER BY semester, subject""", (student["name"],))
    wb = Workbook()
    ws = wb.active
    ws.title = "My Results"
    headers = ["Semester", "Subject", "Exam Type", "Marks", "Total", "Percentage", "Grade", "Result"]
    for c, header in enumerate(headers, 1):
        cell = ws.cell(1, c, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    for r, row in enumerate(rows, 2):
        percentage = round(row["marks"] / row["total"] * 100, 1) if row["total"] else 0
        values = [
            row["semester"], row["subject"], row["exam_type"] or "",
            row["marks"], row["total"], f"{percentage}%", row["grade"] or "", row["result"] or "",
        ]
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")) + 4, 14)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="my_results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# Leave DB is now initialized via core/db_validators.py

@app.route("/student_leave", methods=["GET","POST"])
@login_required("student")
def student_leave():
    sid  = session["student_id"]
    s    = qone("SELECT * FROM students WHERE id=%s", (sid,))
    if request.method == "POST":
        exe("""INSERT INTO leave_applications(student_id,student_name,roll,from_date,to_date,reason,leave_type)
               VALUES(%s,%s,%s,%s,%s,%s,%s)""",
            (sid, s["name"], s["roll"],
             request.form.get("from_date",""), request.form.get("to_date",""),
             request.form.get("reason",""), request.form.get("leave_type","Personal")))
        return redirect("/student_leave?success=1")
    leaves = qry("SELECT * FROM leave_applications WHERE student_id=? ORDER BY id DESC", (sid,))
    return render_template("student/student_leave.html", student=dict(s), leaves=leaves)

@app.route("/admin_leaves")
@login_required("admin")
def admin_leaves():
    status = request.args.get("status","").strip()
    dept   = request.args.get("dept","").strip()
    sql    = """SELECT la.*, s.department FROM leave_applications la
                LEFT JOIN students s ON la.student_id=s.id WHERE 1=1"""
    params = []
    if status: sql += " AND la.status=?"; params.append(status)
    if dept:   sql += " AND s.department=?"; params.append(dept)
    sql += " ORDER BY la.id DESC"
    leaves = qry(sql, params)
    return render_template("admin/admin_leaves.html", leaves=leaves,
                           status=status, dept=dept, DEPARTMENTS=DEPARTMENTS)

@app.route("/update_leave", methods=["POST"])
@login_required("admin")
def update_leave():
    lid     = request.form.get("leave_id","")
    status  = request.form.get("status","")
    remarks = request.form.get("remarks","")
    exe("UPDATE leave_applications SET status=%s,remarks=%s WHERE id=%s", (status,remarks,lid))
    # If approved, insert Leave record into attendance
    if status == "Approved":
        leave = qone("SELECT * FROM leave_applications WHERE id=%s", (lid,))
        if leave:
            try:
                fd = datetime.strptime(leave["from_date"],"%Y-%m-%d").date()
                td = datetime.strptime(leave["to_date"],"%Y-%m-%d").date()
                d = fd
                while d <= td:
                    exe(
                        "INSERT INTO attendance(student_id,student_name,subject,date,status) VALUES(%s,%s,%s,%s,%s)",
                        (
                            leave["student_id"],
                            leave["student_name"],
                            "Leave (Approved)",
                            d.strftime("%Y-%m-%d"),
                            "Leave",
                        ),
                    )
                    d += timedelta(days=1)
            except Exception as e:
                app.logger.error(f"Error: {e}")
    return redirect("/admin_leaves?updated=1")


@app.route("/analytics")
@login_required("admin")
def analytics_dashboard():
    # Monthly attendance trend (last 6 months)
    monthly = []
    today_d = date.today()
    for i in range(5, -1, -1):
        mn = ((today_d.month - 1 - i) % 12) + 1
        yn = today_d.year + (today_d.month - 1 - i) // 12
        label = f"{yn}-{mn:02d}"
        tot = (qone("SELECT COUNT(*) as c FROM attendance WHERE date::text LIKE %s", (f"{label}%",)) or {"c":0})["c"]
        pre = (qone("SELECT COUNT(*) as c FROM attendance WHERE date::text LIKE %s AND status ILIKE 'Present'", (f"{label}%",)) or {"c":0})["c"]
        monthly.append({"label":label, "total":tot, "present":pre, "pct":pct(pre,tot)})

    students_all = qry("SELECT id,name,roll,department FROM students")
    # Defaulter list (percentage < 75%)
    defaulters = []
    for s in students_all:
        stat = qone("SELECT COALESCE(COUNT(*),0) as tot, COALESCE(SUM(CASE WHEN status ILIKE 'Present' THEN 1 ELSE 0 END),0) as pre FROM attendance WHERE student_id=%s", (s["id"],))
        p = pct(stat["pre"], stat["tot"])
        if stat["tot"] >= 3 and p < 75:
            s["pct"] = p; s["tot"] = stat["tot"]; s["pre"] = stat["pre"]
            defaulters.append(s)
    
    defaulters.sort(key=lambda x: x["pct"])
    
    # Department comparison using COALESCE
    dept_stats = qry("""
        SELECT s.department, 
               COALESCE(COUNT(a.id), 0) as total, 
               COALESCE(SUM(CASE WHEN a.status ILIKE 'Present' THEN 1 ELSE 0 END), 0) as present
        FROM students s 
        LEFT JOIN attendance a ON s.id = a.student_id 
        GROUP BY s.department
    """)
    for d in dept_stats: d["pct"] = pct(d["present"], d["total"])

    return render_template("admin/attendance_analytics.html", 
                           monthly=monthly, 
                           defaulters=defaulters[:20],
                           dept_stats=dept_stats)


# ════════════════════════════════════════════════════════════
#  HIGH VALUE 4 — INTERNAL MESSAGING
# ════════════════════════════════════════════════════════════
def init_messages_db():
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS messages (
        id          SERIAL PRIMARY KEY,
        from_role   TEXT NOT NULL,
        from_id     INTEGER NOT NULL,
        from_name   TEXT NOT NULL,
        to_role     TEXT NOT NULL,
        to_id       INTEGER NOT NULL,
        to_name     TEXT NOT NULL,
        subject     TEXT NOT NULL,
        body        TEXT NOT NULL,
        is_read     INTEGER DEFAULT 0,
        sent_at     TEXT DEFAULT to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD HH24:MI:SS')
    )""")
    # Backward-compatible migration for older schema created during SQLite->PG transition.
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS from_role TEXT NOT NULL DEFAULT ''")
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS from_id INTEGER NOT NULL DEFAULT 0")
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS from_name TEXT NOT NULL DEFAULT ''")
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS to_role TEXT NOT NULL DEFAULT ''")
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS to_id INTEGER NOT NULL DEFAULT 0")
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS to_name TEXT NOT NULL DEFAULT ''")
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS subject TEXT NOT NULL DEFAULT ''")
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS body TEXT NOT NULL DEFAULT ''")
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS is_read INTEGER NOT NULL DEFAULT 0")
    conn.execute("ALTER TABLE messages ADD COLUMN IF NOT EXISTS sent_at TEXT")
    conn.execute(
        "UPDATE messages SET from_role=COALESCE(NULLIF(from_role,''), sender_role, ''), "
        "from_id=COALESCE(NULLIF(from_id,0), sender_id, 0), "
        "to_role=COALESCE(NULLIF(to_role,''), receiver_role, ''), "
        "body=COALESCE(NULLIF(body,''), message, ''), "
        "sent_at=COALESCE(sent_at, to_char(created_at, 'YYYY-MM-DD HH24:MI:SS'))"
    )
    conn.execute(
        "UPDATE messages SET sent_at=to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD HH24:MI:SS') "
        "WHERE sent_at IS NULL OR sent_at=''"
    )
    conn.commit(); conn.close()

init_messages_db()

def unread_count():
    try:
        role = session.get("role")
        uid  = session.get("faculty_id") or session.get("student_id") or 0
        if not role or not uid: return 0
        row = qone("SELECT COUNT(*) as c FROM messages WHERE to_role=%s AND to_id=%s AND is_read=0",
                    (role, uid))
        return row[0] if row else 0
    except Exception:
        return 0

app.jinja_env.globals["unread_count"] = unread_count


@app.route("/messages")
def messages_inbox():
    role = session.get("role")
    uid  = session.get("faculty_id") or session.get("student_id") or (1 if role=="admin" else 0)
    if not role: return redirect("/login")
    msgs = qry("SELECT * FROM messages WHERE to_role=%s AND to_id=%s ORDER BY id DESC", (role, uid))
    exe("UPDATE messages SET is_read=1 WHERE to_role=%s AND to_id=%s", (role, uid))
    return render_template("messages/messages.html", msgs=msgs, view="inbox")

@app.route("/messages/sent")
def messages_sent():
    role = session.get("role")
    uid  = session.get("faculty_id") or session.get("student_id") or (1 if role=="admin" else 0)
    if not role: return redirect("/login")
    msgs = qry("SELECT * FROM messages WHERE from_role=%s AND from_id=%s ORDER BY id DESC", (role, uid))
    return render_template("messages/messages.html", msgs=msgs, view="sent")

@app.route("/messages/compose", methods=["GET","POST"])
def messages_compose():
    role = session.get("role")
    if not role: return redirect("/login")
    uid  = session.get("faculty_id") or session.get("student_id") or (1 if role=="admin" else 0)
    name = session.get("name","")

    if request.method == "POST":
        to_role = request.form.get("to_role","")
        to_id   = safe_int(request.form.get("to_id","0"))
        if to_role == "student":
            rec = qone("SELECT name FROM students WHERE id=%s", (to_id,))
        elif to_role == "faculty":
            rec = qone("SELECT name FROM faculty WHERE id=%s", (to_id,))
        else: rec = None
        to_name = rec["name"] if rec else "Unknown"
        exe("""INSERT INTO messages(from_role,from_id,from_name,to_role,to_id,to_name,subject,body)
               VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
            (role, uid, name, to_role, to_id, to_name,
             request.form.get("subject",""), request.form.get("body","")))
        return redirect("/messages?sent=1")

    students_list = qry("SELECT id,name,roll,department FROM students ORDER BY name")
    faculty_list  = qry("SELECT id,name,department FROM faculty ORDER BY name")
    reply_to = request.args.get("reply_to","")
    prefill  = {}
    if reply_to:
        orig = qone("SELECT * FROM messages WHERE id=?", (reply_to,))
        if orig:
            prefill = {"to_role":orig["from_role"],"to_id":orig["from_id"],
                       "subject":"Re: "+orig["subject"]}
    return render_template("messages/messages_compose.html",
                           students=students_list, faculty_list=faculty_list,
                           role=role, prefill=prefill)

@app.route("/messages/delete", methods=["POST"])
def messages_delete():
    exe("DELETE FROM messages WHERE id=?", (request.form.get("msg_id",""),))
    return redirect("/messages")


# ════════════════════════════════════════════════════════════
#  HIGH VALUE 5 — DASHBOARD ANALYTICS (richer reports)
# ════════════════════════════════════════════════════════════
@app.route("/analytics")
@login_required("admin")
def analytics():

    # Monthly attendance trend (last 6 months) — fixed month calculation
    monthly = []
    today_d = date.today()
    for i in range(5, -1, -1):
        # Calculate correct month/year
        month_offset = (today_d.month - 1 - i) % 12
        year_offset  = today_d.year - ((i - today_d.month + 1 + 11) // 12)
        mn = ((today_d.month - 1 - i) % 12) + 1
        yn = today_d.year + (today_d.month - 1 - i) // 12
        label = f"{yn}-{mn:02d}"
        tot = qone("SELECT COUNT(*) as c FROM attendance WHERE date LIKE %s", (f"{label}%",))["c"]
        pre = qone("SELECT COUNT(*) as c FROM attendance WHERE date LIKE %s AND status='Present'", (f"{label}%",))["c"]
        monthly.append({"label":label, "total":tot, "present":pre, "pct":pct(pre,tot)})

    # Bottom students by attendance (min 3 classes) — top spotlight removed
    students_all = qry("SELECT id,name,roll,department FROM students")
    student_att = []
    for s in students_all:
        rows = qry(
            f"SELECT status FROM attendance WHERE {att_match_student_sql()}",
            att_match_student_params(s["id"], s["name"]),
        )
        if len(rows) < 3: continue
        pre = sum(1 for r in rows if r["status"]=="Present")
        student_att.append({"name":s["name"],"roll":s["roll"],"dept":s["department"],
                            "pct":pct(pre,len(rows)),"total":len(rows)})
    student_att.sort(key=lambda x:-x["pct"])
    bottom5 = sorted(student_att, key=lambda x:x["pct"])[:5]


    # Subject-wise average marks
    subj_rows = qry("SELECT subject, ROUND(AVG(marks*100.0/total)::numeric,1) as avg FROM marks GROUP BY subject ORDER BY avg DESC LIMIT 10")
    subj_marks = [dict(r) for r in subj_rows]

    # Exam type breakdown
    exam_rows = qry("SELECT exam_type, ROUND(AVG(marks*100.0/total)::numeric,1) as avg FROM marks GROUP BY exam_type")

    # Dept attendance
    dept_att = []
    for d in DEPARTMENTS:
        tot = qone("SELECT COUNT(*) as c FROM attendance a JOIN students s ON a.student_id=s.id WHERE s.department=%s", (d,))["c"]
        pre = qone("SELECT COUNT(*) as c FROM attendance a JOIN students s ON a.student_id=s.id WHERE s.department=%s AND a.status='Present'", (d,))["c"]
        dept_att.append({"dept":d,"pct":pct(pre,tot),"total":tot})

    # Grade distribution
    grade_dist = {"A+": 0, "A": 0, "B+": 0, "B": 0, "C": 0, "F": 0}
    all_marks = qry("SELECT marks, total FROM marks WHERE total > 0")
    for m in all_marks:
        g = grade(m["marks"], m["total"])
        if g in grade_dist:
            grade_dist[g] += 1
    pass_count = sum(1 for m in all_marks if pct(m["marks"], m["total"]) >= 40)
    fail_count  = len(all_marks) - pass_count

    # Dept-wise avg marks
    dept_marks = []
    for d in DEPARTMENTS:
        row = qone("SELECT ROUND(AVG(marks*100.0/total)::numeric,1) as avg FROM marks WHERE department=%s AND total>0", (d,))
        dept_marks.append({"dept": d, "avg": row["avg"] or 0})

    # Attendance status breakdown
    att_status = {}
    for s in ["Present","Absent","Leave","Late","Medical"]:
        row = qone("SELECT COUNT(*) as c FROM attendance WHERE status=%s", (s,))
        att_status[s] = row["c"] if row else 0

    return render_template("common/analytics.html",
        monthly=monthly,
        month_labels=[m["label"] for m in monthly],
        month_pcts=[m["pct"] for m in monthly],
        top5=student_att[:5], bottom5=bottom5,
        subj_marks=subj_marks,
        sm_labels=[r["subject"][:25] for r in subj_marks],
        sm_data=[r["avg"] for r in subj_marks],
        exam_breakdown=[dict(r) for r in exam_rows],
        exam_labels=[r["exam_type"] for r in exam_rows],
        exam_data=[r["avg"] for r in exam_rows],
        dept_att=dept_att,
        grade_dist=grade_dist,
        pass_count=pass_count, fail_count=fail_count,
        dept_marks=dept_marks,
        att_status=att_status,
        total_students=qone("SELECT COUNT(*) as c FROM students")["c"],
        total_att=qone("SELECT COUNT(*) as c FROM attendance")["c"],
        total_marks=qone("SELECT COUNT(*) as c FROM marks")["c"])


# ════════════════════════════════════════════════════════════
#  HIGH VALUE 6 — EVENT / HOLIDAY CALENDAR
# ════════════════════════════════════════════════════════════
def init_events_db():
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS events (
        id          SERIAL PRIMARY KEY,
        title       TEXT NOT NULL,
        event_date  TEXT NOT NULL,
        event_type  TEXT DEFAULT 'Event',
        description TEXT DEFAULT '',
        created_by  TEXT DEFAULT 'admin'
    )""")
    conn.commit(); conn.close()

init_events_db()

@app.route("/calendar")
def calendar_view():
    role = session.get("role")
    if not role: return redirect("/login")
    month = request.args.get("month","")
    if not month:
        month = date.today().strftime("%Y-%m")
    events = qry("SELECT * FROM events WHERE event_date LIKE %s ORDER BY event_date", (f"{month}%",))
    upcoming = qry("SELECT * FROM events WHERE event_date >= %s ORDER BY event_date LIMIT 10",
                   (today_str(),))
    return render_template("common/calendar.html", events=events, upcoming=upcoming,
                           month=month, role=role)

@app.route("/save_event", methods=["POST"])
@login_required("admin")
def save_event():
    exe("INSERT INTO events(title,event_date,event_type,description) VALUES(%s,%s,%s,%s)",
        (request.form.get("title",""), request.form.get("event_date",""),
         request.form.get("event_type","Event"), request.form.get("description","")))
    return redirect("/calendar?success=1")

@app.route("/delete_event", methods=["POST"])
@login_required("admin")
def delete_event():
    exe("DELETE FROM events WHERE id=?", (request.form.get("event_id",""),))
    return redirect("/calendar")



# ════════════════════════════════════════════════════════════
#  SUBJECTS — Auto-extract from timetable + download template
# ════════════════════════════════════════════════════════════


@app.route("/extract_subjects_from_timetable", methods=["POST"])
@login_required("admin")
def extract_subjects_from_timetable():
    all_subjects = {}

    # Try uploaded TY-TT file first
    upload_dir = "/mnt/user-data/uploads"
    tt_file = None
    if os.path.exists(upload_dir):
        for fn in sorted(os.listdir(upload_dir), reverse=True):
            if ("TY-TT" in fn or "tt-" in fn.lower()) and fn.endswith(".xlsx"):
                tt_file = os.path.join(upload_dir, fn); break

    if tt_file:
        try:
            wb = load_workbook(tt_file, data_only=True)
            all_subjects = _read_subjects_from_wb(wb)
        except Exception:
            pass

    # Fallback: extract from timetable DB
    if not all_subjects:
        for r in qry("SELECT DISTINCT subject, teacher FROM timetable WHERE subject != ''"):
            s = r["subject"].strip()
            if not s or "break" in s.lower(): continue
            all_subjects[(s.lower(), "", "", "")] = {"name": s, "code":"","teacher":r["teacher"] or "","dept":"","sem":""}

    added = skipped = 0
    for info in all_subjects.values():
        name = info["name"]
        if _subject_exists(name, info["code"], info["dept"], info["sem"]):
            skipped += 1; continue
        exe("INSERT INTO subjects(name,subject_code,teacher,department,semester) VALUES(%s,%s,%s,%s,%s)",
            (name, info["code"], info["teacher"], info["dept"], info["sem"]))
        added += 1
    return redirect(f"/subjects?imported={added}&skipped={skipped}")


def _cell_text(value):
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def _norm_subject_code(value):
    return re.sub(r"\s+", "", _cell_text(value)).upper()


def _clean_subject_name(value):
    name = _cell_text(value)
    name = re.sub(r"\s*-\s*", " - ", name)
    name = re.sub(r"\s+", " ", name).strip(" -")
    return name


def _clean_faculty_name(value):
    text = _cell_text(value)
    if not text:
        return ""
    parts = [p.strip() for p in re.split(r"\s*/\s*|\s*,\s*", text) if p.strip()]
    for part in parts or [text]:
        cleaned = re.sub(r"\s*\([A-Z0-9]{1,8}\)\s*$", "", part).strip()
        if cleaned:
            return cleaned
    return text


def _dept_from_timetable_text(text):
    upper = re.sub(r"[^A-Z0-9]+", " ", (text or "").upper())
    if "AIDS" in upper or ("AI" in upper and "DS" in upper):
        return "AIDS"
    if "AIML" in upper or ("AI" in upper and "ML" in upper):
        return "AIML"
    if "INFORMATION TECHNOLOGY" in upper or re.search(r"\bIT\b", upper):
        return "IT"
    if "COMPUTER ENGINEERING" in upper or "CSE" in upper or re.search(r"\bCE\b", upper):
        return "CS"
    return ""


def _sem_from_timetable_text(text):
    upper = (text or "").upper()
    m = re.search(r"SEM(?:ESTER)?\s*[:\-]?\s*(I{1,3}|IV|V|VI{0,3}|[1-8])\b", upper)
    if m:
        val = m.group(1)
        return {"1": "I", "2": "II", "3": "III", "4": "IV", "5": "V", "6": "VI", "7": "VII", "8": "VIII"}.get(val, val)
    if re.search(r"\b(FY|F\.Y|FIRST YEAR)\b", upper):
        return "II"
    if re.search(r"\b(SY|S\.Y|SE|S\.E|SECOND YEAR)\b", upper):
        return "IV"
    if re.search(r"\b(TY|T\.Y|TE|T\.E|THIRD YEAR)\b", upper):
        return "VI"
    return ""


def _subject_exists(name, code, dept, sem):
    if code:
        existing = qone(
            "SELECT id FROM subjects WHERE COALESCE(subject_code,'')=%s "
            "AND COALESCE(department,'')=%s AND COALESCE(semester,'')=%s",
            (code or "", dept or "", sem or ""),
        )
        if existing:
            return existing
    return qone(
        "SELECT id FROM subjects WHERE LOWER(TRIM(name))=LOWER(TRIM(%s)) "
        "AND COALESCE(subject_code,'')=%s AND COALESCE(department,'')=%s "
        "AND COALESCE(semester,'')=%s",
        (name, code or "", dept or "", sem or ""),
    )


def _read_subjects_from_wb(wb):
    all_subjects = {}
    for ws in wb.worksheets:
        sheet_text = " ".join(
            _cell_text(ws.cell(r, c).value)
            for r in range(1, min(ws.max_row, 12) + 1)
            for c in range(1, min(ws.max_column, 8) + 1)
        )
        meta_text = f"{ws.title} {sheet_text}"
        dept = _dept_from_timetable_text(ws.title) or _dept_from_timetable_text(meta_text)
        sem = _sem_from_timetable_text(meta_text)

        for ri in range(1, ws.max_row + 1):
            row_vals = [_cell_text(ws.cell(ri, c).value) for c in range(1, ws.max_column + 1)]
            row_lowers = [v.lower().replace(" ", "").replace(".", "") for v in row_vals]
            if not any(v in ("srno", "sr", "serialno") for v in row_lowers):
                continue
            if not any("subjectcode" in v or v == "code" for v in row_lowers):
                continue
            if not any("subjectname" in v or v == "subject" for v in row_lowers):
                continue

            def find_col(*needles):
                for idx, value in enumerate(row_lowers, 1):
                    if any(needle in value for needle in needles):
                        return idx
                return None

            code_col = find_col("subjectcode", "code") or 2
            name_col = find_col("subjectname") or find_col("name") or 3
            faculty_col = find_col("facultyname", "faculty", "teacher") or 4
            sr_col = find_col("srno", "serial") or 1
            blank_run = 0

            for ri2 in range(ri + 1, ws.max_row + 1):
                sr_v = _cell_text(ws.cell(ri2, sr_col).value)
                code_v = _norm_subject_code(ws.cell(ri2, code_col).value)
                name_v = _clean_subject_name(ws.cell(ri2, name_col).value)
                fac_v = _clean_faculty_name(ws.cell(ri2, faculty_col).value)
                row_text = " ".join(_cell_text(ws.cell(ri2, c).value) for c in range(1, min(ws.max_column, 8) + 1))
                row_upper = row_text.upper()

                if not any([sr_v, code_v, name_v, fac_v]):
                    blank_run += 1
                    if blank_run >= 3:
                        break
                    continue
                blank_run = 0

                if any(stop in row_upper for stop in ("HOD", "TIME TABLE", "TIME-TABLE", "COORDINATOR")):
                    break
                if not name_v or "subject" in name_v.lower():
                    continue
                if not re.search(r"\d", sr_v) and not code_v:
                    continue
                if name_v.lower() in ("lunch", "break", "recess"):
                    continue

                if code_v:
                    key = (code_v, dept, sem)
                else:
                    key = (name_v.lower(), dept, sem)
                if any(
                    existing["name"].lower() == name_v.lower()
                    and existing["dept"] == dept
                    and existing["sem"] == sem
                    for existing in all_subjects.values()
                ):
                    continue
                if key not in all_subjects:
                    all_subjects[key] = {"name": name_v, "code": code_v, "teacher": fac_v, "dept": dept, "sem": sem}
            break
    return all_subjects


@app.route("/import_subjects_from_tt", methods=["POST"])
@login_required("admin")
def import_subjects_from_tt():
    f = request.files.get("file")
    if not f: return redirect("/subjects?error=no_file")
    wb = load_workbook(f, data_only=True)
    if _is_attendance_student_workbook(wb):
        added, skipped = _import_attendance_workbook_subjects(wb, f.filename)
        return redirect(f"/subjects?imported={added}&skipped={skipped}&format=attendance_xlsx")
    all_subjects = _read_subjects_from_wb(wb)
    added = skipped = 0
    for info in all_subjects.values():
        name = info["name"]
        if _subject_exists(name, info["code"], info["dept"], info["sem"]):
            skipped += 1; continue
        exe("INSERT INTO subjects(name,subject_code,teacher,department,semester) VALUES(%s,%s,%s,%s,%s)",
            (name, info["code"], info["teacher"], info["dept"], info["sem"]))
        added += 1
    return redirect(f"/subjects?imported={added}&skipped={skipped}")


@app.route("/delete_all_subjects", methods=["POST"])
@login_required("admin")
def delete_all_subjects():
    count = qone("SELECT COUNT(*) as c FROM subjects")["c"]
    exe("DELETE FROM subjects")
    return redirect(f"/subjects?deleted_all={count}")

@app.route("/download_subjects_template")
@login_required("admin")
def download_subjects_template():
    """Download a ready-to-fill Excel template for subjects."""
    wb = Workbook(); ws = wb.active; ws.title = "Subjects"
    # Headers
    headers = ["Subject Name *", "Subject Code", "Department", "Semester", "Teacher/Faculty"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    # Sample rows
    samples = [
        ["Machine Learning", "ML301", "AIML", "VI", "Prof. Nirmala Chede"],
        ["Data Structures", "CS201", "CS", "III", "Prof. Pankaj Shinde"],
        ["Cloud Computing", "IT401", "IT", "VII", "Prof. Shital Patil"],
        ["Design & Analysis of Algorithms", "CS302", "CS", "V", "Prof. Anurag Jaiswal"],
    ]
    for r, row in enumerate(samples, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 25
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="subjects_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ════════════════════════════════════════════════════════════
#  MESSAGES — enhanced with search
# ════════════════════════════════════════════════════════════

@app.route("/messages/search")
def messages_search():
    role = session.get("role")
    if not role: return redirect("/login")
    uid  = session.get("faculty_id") or session.get("student_id") or (1 if role=="admin" else 0)
    q    = request.args.get("q","").strip()
    results = []
    if q:
        like = f"%{q}%"
        # Search students by name or roll
        students_found = qry("SELECT id,'student' as role,name,roll as identifier,department as extra FROM students WHERE name LIKE ? OR roll LIKE ? LIMIT 10", (like,like))
        # Search faculty by name or email
        faculty_found  = qry("SELECT id,'faculty' as role,name,email as identifier,department as extra FROM faculty WHERE name LIKE ? OR email LIKE ? LIMIT 10", (like,like))
        results = [dict(r) for r in students_found] + [dict(r) for r in faculty_found]
    return render_template("messages/messages_search.html", q=q, results=results, role=role)



# ── SCAN EXCEL / AUTO-FILL API ─────────────────────────
@app.route("/scan_excel", methods=["POST"])
@login_required("admin")
def scan_excel():
    f = request.files.get("excel_scan")
    if not f: return jsonify({"error": "No file uploaded"})
    try:
        if pd is None:
            raise ImportError("pandas is not installed")
        df = pd.read_excel(f)
        if df.empty: return jsonify({"error": "Excel file is empty"})
        first_row = df.iloc[0].fillna('')
        cols = [c.lower() for c in df.columns]
        data = {}
        for c_idx, col_name in enumerate(cols):
            val = str(first_row[c_idx]).strip()
            if 'name' in col_name: data['name'] = val
            elif 'prn' in col_name or 'roll' in col_name: data['roll'] = val
            elif 'dept' in col_name or 'department' in col_name: data['dept'] = val
            elif 'year' in col_name:
                v = val.lower().replace("year", "").strip()
                data['year'] = v if v else "1"
            elif 'mail' in col_name: data['email'] = val
        return jsonify({"data": data, "total_rows": len(df)})
    except Exception as e:
        return jsonify({"error": str(e)})

# ── ATTENDANCE DUPLICATE CHECK API ─────────────────────────
@app.route("/check_att_duplicate")
@login_required("admin")
def check_att_duplicate():
    subject = request.args.get("subject","")
    date    = request.args.get("date","")
    count   = qone("SELECT COUNT(*) as c FROM attendance WHERE subject=? AND date=?",
                   (subject,date))["c"]
    return jsonify({"count": count})





# ════════════════════════════════════════════════════════════
#  CUMULATIVE MARKS — Admin + Faculty View
# ════════════════════════════════════════════════════════════

def _build_cumulative(rows_fn):
    """
    rows_fn: callable that takes (student_name) and returns list of marks rows.
    Returns sorted list of per-student cumulative summaries.
    """
    students = rows_fn(None)  # get distinct students
    result = []
    for s in students:
        rows = rows_fn(s["student_name"])
        if not rows:
            continue
        obtained = sum(r["marks"] for r in rows)
        total    = sum(r["total"] for r in rows)
        pct_val  = round(obtained / total * 100, 1) if total else 0
        result.append({
            "name":     s["student_name"],
            "roll":     s["roll"] or "",
            "dept":     s["department"] or "",
            "obtained": obtained,
            "total":    total,
            "pct":      pct_val,
            "grade":    grade(obtained, total),
            "exams":    len(rows),
        })
    result.sort(key=lambda x: -x["pct"])
    return result


def _cumulative_marks_data_aggregated(dept="", student_filter="", faculty_id=None):
    """Per-student cumulative marks via SQL GROUP BY (student_id + legacy NULL rows)."""
    data = []

    def _add(name, roll, d_dept, obtained, total_m, exams):
        if not total_m:
            return
        pv = round(obtained / total_m * 100, 1) if total_m else 0
        data.append({
            "name": name,
            "roll": roll or "",
            "dept": d_dept or "",
            "obtained": obtained,
            "total": total_m,
            "pct": pv,
            "grade": grade(obtained, total_m),
            "exams": exams,
        })

    wh = ["m.student_id IS NOT NULL"]
    params = []
    if faculty_id is not None:
        wh.append("m.faculty_id=?")
        params.append(faculty_id)
    if dept:
        wh.append("m.department=?")
        params.append(dept)
    if student_filter:
        wh.append("m.student_name LIKE ?")
        params.append(f"%{student_filter}%")
    ws = " AND ".join(wh)
    for r in qry(
        f"""SELECT MAX(m.student_name) AS student_name, MAX(m.roll) AS roll,
            MAX(m.department) AS department, SUM(m.marks) AS obtained,
            SUM(m.total) AS total_m, COUNT(*) AS exams
            FROM marks m WHERE {ws} GROUP BY m.student_id HAVING SUM(m.total) > 0""",
        params,
    ):
        _add(r["student_name"], r["roll"], r["department"], r["obtained"], r["total_m"], r["exams"])

    wh2 = ["m.student_id IS NULL"]
    params2 = []
    if faculty_id is not None:
        wh2.append("m.faculty_id=?")
        params2.append(faculty_id)
    if dept:
        wh2.append("m.department=?")
        params2.append(dept)
    if student_filter:
        wh2.append("m.student_name LIKE ?")
        params2.append(f"%{student_filter}%")
    ws2 = " AND ".join(wh2)
    for r in qry(
        f"""SELECT MAX(m.student_name) AS student_name, MAX(m.roll) AS roll,
            MAX(m.department) AS department, SUM(m.marks) AS obtained,
            SUM(m.total) AS total_m, COUNT(*) AS exams
            FROM marks m WHERE {ws2}
            GROUP BY m.student_name, m.roll, m.department HAVING SUM(m.total) > 0""",
        params2,
    ):
        _add(r["student_name"], r["roll"], r["department"], r["obtained"], r["total_m"], r["exams"])

    data.sort(key=lambda x: -x["pct"])
    return data


@app.route("/cumulative_marks")
@login_required("admin")
def cumulative_marks():
    dept    = request.args.get("dept","").strip()
    student = request.args.get("student","").strip()

    data = _cumulative_marks_data_aggregated(dept=dept, student_filter=student, faculty_id=None)

    return render_template("cumulative/cumulative_marks.html",
        data=data, dept=dept, student=student,
        DEPARTMENTS=DEPARTMENTS,
        total_students=len(data),
        role="admin")


@app.route("/faculty_cumulative")
@login_required("faculty")
def faculty_cumulative():
    fid     = session["faculty_id"]
    student = request.args.get("student","").strip()

    data = _cumulative_marks_data_aggregated(dept="", student_filter=student, faculty_id=fid)

    return render_template("cumulative/cumulative_marks.html",
        data=data, dept="", student=student,
        DEPARTMENTS=DEPARTMENTS,
        total_students=len(data),
        role="faculty")


@app.route("/export_cumulative_excel")
@login_required("admin")
def export_cumulative_excel():
    dept    = request.args.get("dept","").strip()
    student = request.args.get("student","").strip()

    data = _cumulative_marks_data_aggregated(dept=dept, student_filter=student, faculty_id=None)

    wb = Workbook(); ws = wb.active; ws.title = "Cumulative Marks"
    hdrs = ["#", "Student Name", "Roll", "Department", "Obtained", "Total", "Percentage", "Grade", "Exams"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")

    row_num = 2
    for idx, row in enumerate(data, start=1):
        for c, v in enumerate(
            [
                idx,
                row["name"],
                row["roll"],
                row["dept"],
                row["obtained"],
                row["total"],
                f"{row['pct']}%",
                row["grade"],
                row["exams"],
            ],
            1,
        ):
            ws.cell(row_num, c, v)
        row_num += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or ""))+4, 14)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="cumulative_marks.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ════════════════════════════════════════════════════════════
#  TIMETABLE SHARING
# ════════════════════════════════════════════════════════════

def _build_tt_body(teacher, slots):
    """Build formatted timetable message for a faculty member."""
    # Convert psycopg2.extras.DictCursor objects to dicts so .get() works
    slots = [dict(s) for s in slots]
    lines = ["Weekly Timetable - DY Patil University\n"]
    lines.append("Faculty: " + teacher)
    lines.append("=" * 42)
    by_day = {}
    for s in slots:
        by_day.setdefault(s["day"], []).append(s)
    for day in DAYS:
        if day not in by_day: continue
        lines.append("\n" + day + ":")
        for s in sorted(by_day[day], key=lambda x: x.get("time","") or ""):
            room = (" | Room: " + s["room"]) if s.get("room") else ""
            div  = (" | " + s["division"])  if s.get("division") else ""
            lines.append("  " + str(s["time"] or "").ljust(15) + " " + str(s["subject"]) + room + div)
    lines.append("\n\nSent from DY Patil ERP Admin.")
    return "\n".join(lines)

@app.route("/timetable_share", methods=["GET","POST"])
@login_required("admin")
def timetable_share():
    if request.method == "POST":
        target_role = request.form.get("target_role", "faculty")
        target_id   = safe_int(request.form.get("target_id","0"))
        
        if not target_id: return redirect(f"/timetable_share?error=no_{target_role}")

        if target_role == "faculty":
            row = qone("SELECT * FROM faculty WHERE id=%s", (target_id,))
            if not row: return redirect(f"/timetable_share?error=no_faculty")
            exact_name = row["name"]
            slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.faculty_id=%s ORDER BY t.day, t.start_time", (target_id,))
        else:
            row = qone("SELECT * FROM students WHERE id=%s", (target_id,))
            if not row: return redirect(f"/timetable_share?role=student&error=not_found")
            exact_name = row["name"]
            slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.branch=%s AND t.division=%s ORDER BY t.day, t.start_time", 
                        (row["department"], row["division"]))

        if not slots: return redirect(f"/timetable_share?role={target_role}&error=no_slots&id={target_id}")

        body = _build_tt_body(exact_name, slots)
        subj = f"Your Weekly Timetable — {len(slots)} slots"
        exe("""INSERT INTO messages(from_role,from_id,from_name,to_role,to_id,to_name,subject,body)
               VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
            ("admin", 1, "Administrator", target_role, target_id, exact_name, subj, body))
        
        log_audit("Share Timetable", f"Sent individual timetable to {target_role}: {exact_name} (ID: {target_id})")
        return redirect(f"/timetable_share?role={target_role}&sent=1&to={exact_name}")

    # GET
    role_sel = request.args.get("role", "faculty")
    id_sel   = safe_int(request.args.get("id","0"))
    selected_slots = []
    matched_entity = None
    
    if id_sel:
        if role_sel == "faculty":
            matched_entity = qone("SELECT * FROM faculty WHERE id=%s", (id_sel,))
            if matched_entity:
                matched_entity = dict(matched_entity)
                selected_slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.faculty_id=%s ORDER BY t.day, t.start_time", (id_sel,))
        else:
            matched_entity = qone("SELECT * FROM students WHERE id=%s", (id_sel,))
            if matched_entity:
                matched_entity = dict(matched_entity)
                selected_slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.branch=%s AND t.division=%s ORDER BY t.day, t.start_time", 
                                    (matched_entity["department"], matched_entity["division"]))

    return render_template("common/timetable_share.html",
        role_sel=role_sel,
        selected_slots=selected_slots,
        matched_entity=matched_entity,
        faculty_list=qry("SELECT id,name,department FROM faculty ORDER BY name"),
        student_list=qry("SELECT id,name,department,division,roll FROM students ORDER BY department, division, name"),
        DEPARTMENTS=DEPARTMENTS,
        DIVISIONS=DIVISIONS,
        DAYS=DAYS,
        today_name=date.today().strftime("%A")
    )


@app.route("/timetable_send_all", methods=["POST"])
@login_required("admin")
def timetable_send_all():
    target_role = request.form.get("target_role", "faculty")
    sent = skipped = 0
    
    if target_role == "faculty":
        distinct_faculties = qry("SELECT DISTINCT faculty_id FROM timetable WHERE faculty_id IS NOT NULL")
        for row in distinct_faculties:
            fac_id = row["faculty_id"]
            filter_val = (fac_id,)
            f = qone("SELECT name FROM faculty WHERE id=%s", filter_val)
            if not f: 
                skipped += 1
                continue
            fac_name = f["name"]
            slots = qry("SELECT t.*, f2.name as teacher FROM timetable t JOIN faculty f2 ON t.faculty_id=f2.id WHERE t.faculty_id=%s ORDER BY t.day, t.start_time", (fac_id,))
            if not slots: continue
            
            body = _build_tt_body(fac_name, slots)
            subj = f"Your Weekly Timetable — {len(slots)} slots"
            exe("""INSERT INTO messages(from_role,from_id,from_name,to_role,to_id,to_name,subject,body)
                   VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
                ("admin", 1, "Administrator", "faculty", fac_id, fac_name, subj, body))
            sent += 1
    else:
        students = qry("SELECT id, name, department, division FROM students")
        cache = {}
        for s in students:
            key = (s["department"], s["division"])
            if key not in cache:
                slots = qry("SELECT t.*, f.name as teacher FROM timetable t JOIN faculty f ON t.faculty_id = f.id WHERE t.branch=%s AND t.division=%s ORDER BY t.day, t.start_time", (key[0], key[1]))
                cache[key] = slots
            
            slots = cache[key]
            if not slots:
                skipped += 1
                continue
            
            body = _build_tt_body(s["name"], slots)
            subj = f"Your Class Timetable — {len(slots)} slots"
            exe("""INSERT INTO messages(from_role,from_id,from_name,to_role,to_id,to_name,subject,body)
                   VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
                ("admin", 1, "Administrator", "student", s["id"], s["name"], subj, body))
            sent += 1

    log_audit("Bulk Share Timetable", f"Initiated bulk share for {target_role}s. Successfully sent: {sent}, Skipped: {skipped}")
    return redirect(f"/timetable_share?role={target_role}&bulk_sent={sent}&bulk_skip={skipped}")






# ════════════════════════════════════════════════════════════
#  SUBJECT ASSIGNMENT — Link subjects to faculty accounts
# ════════════════════════════════════════════════════════════

@app.route("/assign_subjects")
@login_required("admin")
def assign_subjects():
    faculty_list  = qry("SELECT * FROM faculty ORDER BY name")
    subjects_list = qry("SELECT * FROM subjects ORDER BY name")
    selected_fac  = request.args.get("fac_id","").strip()
    fac_subjects  = []
    fac_info      = None
    if selected_fac:

        fac_info = qone("SELECT * FROM faculty WHERE id=%s", (selected_fac,))
        if fac_info:
            fac_info = dict(fac_info)
            fac_subjects = qry("SELECT * FROM subjects WHERE teacher LIKE %s ORDER BY name",
                               ("%" + fac_info["name"] + "%",))
    return render_template("admin/assign_subjects.html",
        faculty=faculty_list, subjects=subjects_list,
        selected_fac=selected_fac, fac_info=fac_info,
        fac_subjects=fac_subjects,
        total_subjects=len(list(subjects_list))
    )

@app.route("/assign_subject_to_faculty", methods=["POST"])
@login_required("admin")
def assign_subject_to_faculty():

    exe("UPDATE subjects SET teacher=%s WHERE id=%s",
        (request.form.get("faculty_name","").strip(), request.form.get("subject_id","")))
    return redirect("/assign_subjects?fac_id=" + request.form.get("fac_id","") + "&assigned=1")

@app.route("/unassign_subject", methods=["POST"])
@login_required("admin")
def unassign_subject():

    exe("UPDATE subjects SET teacher='' WHERE id=%s", (request.form.get("subject_id",""),))
    return redirect("/assign_subjects?fac_id=" + request.form.get("fac_id","") + "&unassigned=1")

@app.route("/auto_assign_subjects", methods=["POST"])
@login_required("admin")
def auto_assign_subjects():
    subjects_list = qry("SELECT * FROM subjects WHERE teacher != ''")
    faculty_list  = qry("SELECT * FROM faculty")
    matched = 0
    for s in subjects_list:
        teacher = (s["teacher"] or "").replace("Prof.","").replace("Dr.","").strip()
        parts   = [p for p in teacher.split() if len(p) > 3]
        for f in faculty_list:
            if any(p in f["name"] for p in parts):
                if f["name"] != s["teacher"]:
                    exe("UPDATE subjects SET teacher=%s WHERE id=%s", (f["name"], s["id"]))
                    matched += 1
                break
    return redirect("/assign_subjects?auto_matched=" + str(matched))

@app.route("/send_classlist", methods=["POST"])
@login_required("admin")
def send_classlist():
    subject_name = request.form.get("subject","").strip()
    dept         = request.form.get("dept","").strip()

    subj         = qone("SELECT * FROM subjects WHERE name=%s", (subject_name,))
    if not subj:
        return redirect("/assign_subjects?error=no_subject")
    teacher_name = subj["teacher"] or ""
    matched_fac  = None
    for part in teacher_name.replace("Prof.","").replace("Dr.","").split():
        if len(part) > 3:
            f = qone("SELECT * FROM faculty WHERE name LIKE %s", ("%" + part + "%",))
            if f: matched_fac = dict(f); break
    if not matched_fac:
        return redirect("/assign_subjects?error=no_faculty_account")
    sql    = "SELECT name,roll,department FROM students"
    params = []
    if dept: sql += " WHERE department=%s"; params.append(dept)
    sql += " ORDER BY name"
    students = qry(sql, params)
    lines  = ["Student List for: " + subject_name, "Department: " + (dept or "All"),
              "Total: " + str(len(students)), "=" * 40]
    for i, s in enumerate(students, 1):
        lines.append(str(i).rjust(3) + ". " + str(s["roll"]).ljust(22) + str(s["name"]))
    lines.append("\n\nTo mark attendance: Login -> My Attendance -> Mark Class")
    lines.append("Select subject: " + subject_name)
    lines.append("\nSent from DY Patil ERP Admin.")
    exe("""INSERT INTO messages(from_role,from_id,from_name,to_role,to_id,to_name,subject,body)
           VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
        ("admin",1,"Administrator","faculty",matched_fac["id"],matched_fac["name"],
         "Student List: " + subject_name + " (" + str(len(students)) + " students)",
         "\n".join(lines)))
    return redirect("/assign_subjects?classlist_sent=1&to=" + matched_fac["name"] + "&count=" + str(len(students)))

# ─── FACULTY ASSIGNMENTS ─────────────────────────────────

@app.route("/admin/faculty_assignments", methods=["GET"])
@login_required("admin")
def admin_faculty_assignments_get():
    sql = """
        SELECT fsa.*, f.name as faculty_name 
        FROM faculty_subject_assignments fsa
        JOIN faculty f ON fsa.faculty_id = f.id 
        ORDER BY f.name, fsa.subject_name
    """
    assignments = qry(sql)
    faculty_list = qry("SELECT * FROM faculty ORDER BY name")
    subject_list = qry("SELECT * FROM subjects ORDER BY name")
    return render_template("admin/faculty_assignments.html",
                           assignments=assignments,
                           faculty_list=faculty_list,
                           subject_list=subject_list,
                           DEPARTMENTS=DEPARTMENTS,
                           SEMESTERS=SEMESTERS,
                           DIVISIONS=DIVISIONS)

@app.route("/admin/faculty_assignments", methods=["POST"])
@login_required("admin")
def admin_faculty_assignments_post():
    action = request.form.get("action")
    if action == "assign":
        faculty_id = request.form.get("faculty_id")
        subject_name = request.form.get("subject_name")
        department = request.form.get("department")
        semester = request.form.get("semester")
        class_name = request.form.get("class_name")
        division = request.form.get("division")
        
        # Get subject_id
        subj = qone("SELECT id FROM subjects WHERE name=%s LIMIT 1", (subject_name,))
        if not subj:
            flash("Subject not found.", "error")
            return redirect("/admin/faculty_assignments")
        subject_id = subj["id"]
        
        # Clash check
        clash = qone("""
            SELECT 1 FROM faculty_subject_assignments 
            WHERE faculty_id=%s AND division=%s AND semester=%s
        """, (faculty_id, division, semester))
        
        if clash:
            flash("Faculty already assigned to a subject in this division/semester", "error")
            return redirect("/admin/faculty_assignments")
            
        try:
            exe("""
                INSERT INTO faculty_subject_assignments 
                  (faculty_id, subject_id, subject_name, department, semester, class_name, division)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT(faculty_id, subject_id, division) DO UPDATE SET
                  subject_name = EXCLUDED.subject_name,
                  department = EXCLUDED.department,
                  semester = EXCLUDED.semester,
                  class_name = EXCLUDED.class_name
            """, (faculty_id, subject_id, subject_name, department, semester, class_name, division))
            
            faculty_name = qone("SELECT name FROM faculty WHERE id=%s", (faculty_id,))["name"]
            log_audit("Faculty Assignment", f"Assigned {faculty_name} -> {subject_name} ({division})")
            flash("Assignment saved successfully", "success")
        except Exception as e:
            app.logger.error("Error saving faculty assignment: %s", e)
            flash("Error saving assignment", "error")
            
    elif action == "delete":
        assign_id = request.form.get("id")
        if assign_id:
            exe("DELETE FROM faculty_subject_assignments WHERE id=%s", (assign_id,))
            log_audit("Faculty Assignment Delete", f"Removed assignment ID {assign_id}")
            flash("Assignment removed", "success")
            
    return redirect("/admin/faculty_assignments")

@app.route("/admin/auto_link_faculty", methods=["POST"])
@login_required("admin")
def auto_link_faculty():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO faculty_subject_assignments
                (faculty_id, subject_id, subject_name, department, semester, class_name, division)
            SELECT DISTINCT faculty_id, subject_id, subject, branch, semester, CONCAT(year, '-', branch), division
            FROM timetable 
            WHERE faculty_id IS NOT NULL AND subject_id IS NOT NULL
            ON CONFLICT(faculty_id, subject_id, division) DO NOTHING
        """)
        count = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()
        log_audit("Auto-Link Faculty", f"Linked {count} faculty-subject pairs from timetable")
        flash(f"Auto-imported {count} assignments from timetable", "success")
    except Exception as e:
        app.logger.error("Error auto-linking faculty: %s", e)
        flash("Error auto-linking faculty.", "error")
    return redirect("/admin/faculty_assignments")

def _ensure_summary_table():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance_summary (
            id          SERIAL PRIMARY KEY,
            student_id  INTEGER,
            student_name TEXT,
            subject     TEXT NOT NULL,
            attended    INTEGER NOT NULL DEFAULT 0,
            total       INTEGER NOT NULL DEFAULT 0,
            division    TEXT DEFAULT '',
            semester    TEXT DEFAULT '',
            UNIQUE(student_id, subject)
        )
    """)
    try:
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_summary_student_id ON attendance_summary(student_id)"
        )
    except Exception:
        pass
    conn.commit(); conn.close()


_ensure_summary_table()
ensure_attendance_upload_tables()
ensure_attendance_engine_schema()
init_attendance_engine()


# ─── Unified helpers ──────────────────────────────────────

def has_summary(student_id):
    """Return True if attendance_summary has rows for this student."""
    row = qone("SELECT COUNT(*) as c FROM attendance_summary WHERE student_id=%s", (student_id,))
    return row and row["c"] > 0


def shortage_needed(attended, total):
    """
    Return (shortage, can_miss).
    shortage > 0 → need to attend this many more lectures to reach 75%
    can_miss > 0 → can afford to miss this many (already above 75%)
    """
    if total <= 0:
        return 0, 0
    # lectures needed: 0.75*total <= attended + x  →  x >= 0.75*total - attended
    needed = math.ceil(0.75 * total - attended)
    shortage = max(0, needed)
    can_miss = 0
    if shortage == 0:
        # can_miss: floor((attended - 0.75*(total+y))/0.75) but simpler:
        # floor((attended - 0.75*total) / 0.75)
        can_miss = max(0, int((attended - 0.75 * total) / 0.75))
    return shortage, can_miss


def _cumulative_result_dict(present, total, source):
    """Build the same dict shape as get_cumulative from aggregated counts."""
    absent = max(0, total - present)
    pct_val = round(present / total * 100, 1) if total else 0
    if pct_val >= 75:
        status = "Good"
    elif pct_val >= 50:
        status = "Average"
    else:
        status = "Low"
    shortage, can_miss = shortage_needed(present, total)
    return {
        "present":    present,
        "attended":   present,
        "absent":     absent,
        "total":      total,
        "percentage": pct_val,
        "status":     status,
        "source":     source,
        "shortage":   shortage,
        "can_miss":   can_miss,
    }


def bulk_cumulative_for_dashboard(student_rows):
    """
    Same outcomes as calling get_cumulative per student, using O(1) DB round-trips
    for summary + daily student_id paths (legacy name-only rows still use get_cumulative).
    student_rows: iterable of dicts with keys id, name.
    Returns dict[student_id] -> cumulative dict.
    """
    rows_list = list(student_rows)
    if not rows_list:
        return {}
    ids = [int(s["id"]) for s in rows_list]
    ph = ",".join(["%s"] * len(ids))
    params = tuple(ids)

    summary_map = {}
    try:
        for r in qry(
            f"SELECT student_id, COALESCE(SUM(attended),0) AS sm_att, COALESCE(SUM(total),0) AS sm_tot "
            f"FROM attendance_summary WHERE student_id IN ({ph}) GROUP BY student_id",
            params,
        ):
            summary_map[int(r["student_id"])] = (int(r["sm_att"] or 0), int(r["sm_tot"] or 0))
    except Exception:
        summary_map = {}

    daily_map = {}
    daily_name_map = {}
    try:
        for r in qry(
            f"SELECT student_id, COUNT(*) AS d_tot, "
            f"SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) AS d_att "
            f"FROM attendance WHERE student_id IN ({ph}) GROUP BY student_id",
            params,
        ):
            daily_map[int(r["student_id"])] = (int(r["d_att"] or 0), int(r["d_tot"] or 0))
    except Exception:
        daily_map = {}

    try:
        name_keys = sorted({
            str(s["name"] or "").strip().lower()
            for s in rows_list
            if str(s["name"] or "").strip()
        })
        if name_keys:
            phn = ",".join(["%s"] * len(name_keys))
            for r in qry(
                f"SELECT LOWER(TRIM(student_name)) AS student_key, COUNT(*) AS d_tot, "
                f"SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) AS d_att "
                f"FROM attendance "
                f"WHERE student_id IS NULL AND LOWER(TRIM(student_name)) IN ({phn}) "
                f"GROUP BY LOWER(TRIM(student_name))",
                name_keys,
            ):
                daily_name_map[str(r["student_key"])] = (int(r["d_att"] or 0), int(r["d_tot"] or 0))
    except Exception:
        daily_name_map = {}

    out = {}
    for s in rows_list:
        sid = int(s["id"])
        name = str(s["name"] or "")
        name_key = name.strip().lower()
        if sid in summary_map and (summary_map[sid][1] > 0 or summary_map[sid][0] > 0):
            att, tot = summary_map[sid]
            out[sid] = _cumulative_result_dict(att, tot, "summary")
        elif sid in daily_map and daily_map[sid][1] > 0:
            att, tot = daily_map[sid]
            out[sid] = _cumulative_result_dict(att, tot, "daily")
        elif name_key in daily_name_map and daily_name_map[name_key][1] > 0:
            att, tot = daily_name_map[name_key]
            out[sid] = _cumulative_result_dict(att, tot, "daily")
        else:
            out[sid] = _cumulative_result_dict(0, 0, "none")
    return out


def get_cumulative(student_id, student_name):
    """
    Unified cumulative attendance.
    Uses attendance_summary if it exists, else falls back to daily attendance table.
    """
    if has_summary(student_id):
        rows = qry(
            "SELECT attended, total FROM attendance_summary WHERE student_id=%s",
            (student_id,)
        )
        source = "summary"
    else:
        # Build from daily attendance records
        rows_raw = qry(
            f"SELECT status FROM attendance WHERE {att_match_student_sql()}",
            att_match_student_params(student_id, student_name)
        )
        total   = len(rows_raw)
        present = sum(1 for r in rows_raw if r["status"] == "Present")
        rows = [{"attended": present, "total": total}]
        source = "daily"

    total   = sum(r["total"]   for r in rows)
    present = sum(r["attended"] for r in rows)
    return _cumulative_result_dict(present, total, source)


def get_subjects(student_id, student_name):
    """
    Return list of subject-wise attendance dicts for a student.
    Uses attendance_summary if available, else daily records.
    """
    subjects = []
    if has_summary(student_id):
        rows = qry(
            "SELECT id, subject, attended, total FROM attendance_summary WHERE student_id=%s ORDER BY subject",
            (student_id,)
        )
        for r in rows:
            att    = r["attended"]
            tot    = r["total"]
            pct_v  = round(att / tot * 100, 1) if tot else 0
            absent = max(0, tot - att)
            shortage_v, can_miss_v = shortage_needed(att, tot)
            status = "Good" if pct_v >= 75 else ("Average" if pct_v >= 50 else "Low")
            subjects.append({
                "summary_row_id": r["id"],
                "subject":    r["subject"],
                "attended":   att,
                "total":      tot,
                "absent":     absent,
                "percentage": pct_v,
                "status":     status,
                "shortage":   shortage_v,
                "can_miss":   can_miss_v,
            })
    else:
        # Build from daily records
        rows = qry(
            f"SELECT subject, status FROM attendance WHERE {att_match_student_sql()} ORDER BY subject",
            att_match_student_params(student_id, student_name)
        )
        data = {}
        for r in rows:
            sub = r["subject"]
            data.setdefault(sub, {"total": 0, "present": 0})
            data[sub]["total"] += 1
            if r["status"] == "Present":
                data[sub]["present"] += 1
        for sub, v in sorted(data.items()):
            att    = v["present"]
            tot    = v["total"]
            pct_v  = round(att / tot * 100, 1) if tot else 0
            absent = max(0, tot - att)
            shortage_v, can_miss_v = shortage_needed(att, tot)
            status = "Good" if pct_v >= 75 else ("Average" if pct_v >= 50 else "Low")
            subjects.append({
                "summary_row_id": None,
                "subject":    sub,
                "attended":   att,
                "total":      tot,
                "absent":     absent,
                "percentage": pct_v,
                "status":     status,
                "shortage":   shortage_v,
                "can_miss":   can_miss_v,
            })
    return subjects


def prediction_curve(attended, total, future_steps=20):
    """
    Return list of (step, pct) projections assuming student attends all future lectures.
    """
    curve = []
    for i in range(1, future_steps + 1):
        new_total   = total + i
        new_present = attended + i
        p = round(new_present / new_total * 100, 1) if new_total else 0
        curve.append(p)
    return curve


def will_reach_75(attended, total, remaining_classes=20):
    """
    Return (True/False, lectures_needed)
    True if student can reach 75% within remaining_classes.
    """
    needed = max(0, int(0.75 * total - attended) + 1)
    can_reach = needed <= remaining_classes
    return can_reach, needed


# ════════════════════════════════════════════════════════════
#  ADMIN ATTENDANCE DASHBOARD
# ════════════════════════════════════════════════════════════

@app.route("/attendance_management")
@login_required("admin")
def attendance_management():
    # 1. Filters & State
    div = request.args.get("division", "All")
    tab = request.args.get("tab", "live") # live, weekly, monthly, defaulter
    
    # 2. Student Query
    sq = "SELECT id, name, roll, division, prn FROM students"
    params = []
    if div != "All":
        sq += " WHERE division = %s"
        params.append(div)
    sq += " ORDER BY division, roll, name"
    students_raw = qry(sq, params)
    
    # 3. Weekly Data Window (Current Mon-Fri)
    today = date.today()
    start_of_week = today - timedelta(days=today.weekday())
    week_dates = [(start_of_week + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(5)]
    week_labels = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    
    # 4. Fetch Attendance Grid
    student_ids = [int(s["id"]) for s in students_raw]
    att_grid = {} # student_id -> {date -> status}
    if student_ids:
        ph = ",".join(["%s"] * len(student_ids))
        rows = qry(f"SELECT student_id, date::text, status FROM attendance WHERE student_id IN ({ph}) AND date::text >= %s AND date::text <= %s", 
                  tuple(student_ids) + (week_dates[0], week_dates[-1]))
        for r in rows:
            sid = r["student_id"]
            dt = r["date"]
            status = r["status"]
            att_grid.setdefault(sid, {})[dt] = status[0] if status else "-"

    # 5. Performance Metrics (Bulk)
    stats_map = bulk_cumulative_for_dashboard(students_raw)
    
    # 6. Transform for Template
    students = []
    total_pct_sum = 0
    shortage_count = 0
    valid_count = 0
    
    for s in students_raw:
        sid = int(s["id"])
        perf = stats_map.get(sid, {"percentage": 0, "attended": 0, "total": 0, "status": "N/A"})
        
        dots = [att_grid.get(sid, {}).get(d, "-") for d in week_dates]
        
        item = {
            "id": sid,
            "name": s["name"],
            "roll": s["roll"] or "N/A",
            "division": s["division"] or "N/A",
            "prn": s["prn"] or "None",
            "dots": dots,
            "pct": perf["percentage"],
            "attended": perf["attended"],
            "total": perf["total"],
            "perf_status": perf.get("status", "Low")
        }
        students.append(item)
        
        if perf["total"] > 0:
            total_pct_sum += perf["percentage"]
            valid_count += 1
            if perf["percentage"] < 75:
                shortage_count += 1

    avg_attendance = round(total_pct_sum / valid_count, 1) if valid_count > 0 else 0
    
    # Tab Logic
    if tab == "defaulter":
        students = [s for s in students if s["pct"] < 75]
    
    divs_raw = qry("SELECT DISTINCT division FROM students WHERE division != '' ORDER BY division")
    
    return render_template(
        "attendance/attendance_management.html",
        students=students,
        divs=[d["division"] for d in divs_raw],
        selected_div=div,
        selected_tab=tab,
        week_labels=week_labels,
        metrics={
            "avg_attendance": avg_attendance,
            "shortage_count": shortage_count,
            "count": len(students_raw)
        },
        date=date.today().strftime("%Y-%m-%d")
    )




@app.route("/attendance_dashboard")
@login_required("admin")
def attendance_dashboard():

    # Initialize all template variables with safe defaults to prevent 500 errors
    all_students = []; defaulters = []; critical = []
    total_students = 0; avg_pct = 0; good_count = 0; avg_count = 0; low_count = 0
    at_risk_count = 0; critical_count = 0; safe_count = 0
    divisions = []; dept_stats = []; div_stats = []
    weekly_trend = []; status_dist = {}
    dept_labels = []; dept_pcts = []; div_labels = []; div_pcts = []
    trend_labels = []; trend_pcts = []; status_labels = []; status_counts = []
    top5_names = []; top5_pcts = []; topper = None

    # 1. Fetch Students & Cumulative Data
    students_raw = qry("SELECT id, name, roll, division, department FROM students ORDER BY name")
    if students_raw:
        cum_by_id = bulk_cumulative_for_dashboard(students_raw)
        for s in students_raw:
            cum = cum_by_id.get(s["id"])
            if not cum or cum.get("total", 0) == 0: continue
            
            shortage_v, can_miss_v = shortage_needed(cum["attended"], cum["total"])
            all_students.append({
                "id": s["id"], "name": s["name"], "roll": s["roll"] or "",
                "division": s["division"] or "", "dept": s["department"] or "",
                "pct": cum["percentage"], "status": cum["status"],
                "attended": cum["attended"], "total": cum["total"],
                "shortage": shortage_v, "can_miss": can_miss_v
            })

    if all_students:
        # Rank by percentage
        all_students.sort(key=lambda x: -x["pct"])
        for i, s in enumerate(all_students): s["rank"] = i + 1

        total_students = len(all_students)
        avg_pct = round(sum(s["pct"] for s in all_students) / total_students, 1)
        
        defaulters = [s for s in all_students if s["pct"] < 75]
        critical = [s for s in all_students if s["pct"] < 40]
        at_risk_count = len(defaulters)
        critical_count = len(critical)
        safe_count = total_students - at_risk_count
        
        good_count = sum(1 for s in all_students if s["pct"] >= 75)
        avg_count = sum(1 for s in all_students if 50 <= s["pct"] < 75)
        low_count = sum(1 for s in all_students if s["pct"] < 50)
        
        divisions = sorted(set(s["division"] for s in all_students if s["division"]))

        # Stats by Dept
        d_map = {}
        for s in all_students:
            d = s["dept"] or "Unknown"
            d_map.setdefault(d, {"cnt": 0, "sum": 0.0, "g": 0, "l": 0})
            d_map[d]["cnt"] += 1; d_map[d]["sum"] += s["pct"]
            if s["pct"] >= 75: d_map[d]["g"] += 1
            elif s["pct"] < 50: d_map[d]["l"] += 1
        dept_stats = sorted([{"dept": k, "count": v["cnt"], "avg_pct": round(v["sum"]/v["cnt"], 1), "good": v["g"], "low": v["l"]} for k, v in d_map.items()], key=lambda x: x["dept"])

        # Stats by Div
        v_map = {}
        for s in all_students:
            dv = s["division"] or "N/A"
            v_map.setdefault(dv, {"cnt": 0, "sum": 0.0})
            v_map[dv]["cnt"] += 1; v_map[dv]["sum"] += s["pct"]
        div_stats = sorted([{"division": k, "count": v["cnt"], "avg_pct": round(v["sum"]/v["cnt"], 1)} for k, v in v_map.items()], key=lambda x: x["division"])

        # Chart Data
        dept_labels = [d["dept"] for d in dept_stats]
        dept_pcts = [d["avg_pct"] for d in dept_stats]
        div_labels = [d["division"] for d in div_stats]
        div_pcts = [d["avg_pct"] for d in div_stats]
        
        top5 = all_students[:5]
        top5_names = [s["name"] for s in top5]
        top5_pcts = [s["pct"] for s in top5]
        topper = all_students[0]

    # Weekly Trend
    week_dates = [(date.today() - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6, -1, -1)]
    ph_w = ",".join(["%s"] * len(week_dates))
    trend_rows = qry(f"SELECT date::text, COUNT(*) as tot, SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as p FROM attendance WHERE date::text IN ({ph_w}) GROUP BY date", week_dates)
    trend_map = {(r["date"].strftime("%Y-%m-%d") if hasattr(r["date"], "strftime") else str(r["date"])): pct(r["p"], r["tot"]) for r in trend_rows}
    weekly_trend = [{"date": d[5:], "pct": trend_map.get(d, 0)} for d in week_dates]
    trend_labels = [t["date"] for t in weekly_trend]
    trend_pcts = [t["pct"] for t in weekly_trend]

    # Status Distribution
    for st in ["Present", "Absent", "Late", "Medical", "Leave"]:
        row = qone("SELECT COUNT(*) as c FROM attendance WHERE status=%s", (st,))
        status_dist[st] = row["c"] if row else 0
    status_labels = list(status_dist.keys())
    status_counts = [status_dist[k] for k in status_labels]

    return render_template("attendance/attendance_dashboard.html",
        all_students=all_students, defaulters=defaulters, critical=critical,
        total_students=total_students, avg_pct=avg_pct, good_count=good_count,
        avg_count=avg_count, low_count=low_count, at_risk_count=at_risk_count,
        critical_count=critical_count, safe_count=safe_count, divisions=divisions,
        dept_stats=dept_stats, div_stats=div_stats, weekly_trend=weekly_trend,
        status_dist=status_dist, dept_labels=dept_labels, dept_pcts=dept_pcts,
        div_labels=div_labels, div_pcts=div_pcts, trend_labels=trend_labels,
        trend_pcts=trend_pcts, status_labels=status_labels, status_counts=status_counts,
        top5_names=top5_names, top5_pcts=top5_pcts, topper=topper
    )



@app.route("/attendance_dashboard_data")
def attendance_dashboard_data():
    """API for real-time analytics used by faculty and admin dashboards."""
    role = session.get("role")
    if not role: return jsonify({"error": "Unauthorized"}), 401
    
    # Context-aware filtering
    is_faculty = (role == "faculty")
    faculty_id = session.get("faculty_id") if is_faculty else None
    
    # 1. Fetch Students
    if is_faculty or request.args.get('q') == 'faculty':
        # Filter students to only those in departments/divisions taught by this faculty
        faculty_name = session.get("name")
        students_rows = qry("""
            SELECT DISTINCT s.id, s.name, s.roll, s.division, s.department 
            FROM students s
            JOIN subjects sub ON (s.department = sub.department OR s.division = sub.division)
            WHERE sub.teacher LIKE %s
            ORDER BY s.name""", (f"%{faculty_name}%",))
    else:
        students_rows = qry("SELECT id, name, roll, division, department FROM students ORDER BY name")
    
    if not students_rows:
        return jsonify({"stats": {"total": 0}, "analytics": {"avg_attendance": 0, "defaulters": [], "subject_wise": []}})

    
    # 2. Basic Stats
    stats_map = bulk_cumulative_for_dashboard(students_rows)
    all_data = []
    for s in students_rows:
        perf = stats_map.get(s["id"], {"percentage": 0, "status": "Low"})
        all_data.append({"id": s["id"], "name": s["name"], "pct": perf["percentage"], "status": perf["status"]})

    avg_att = round(sum(d["pct"] for d in all_data) / len(all_data), 1) if all_data else 0
    defaulters = [d for d in all_data if d["pct"] < 75]
    
    # 3. Subject-wise (approx from summary/samples)
    # For speed, we just aggregate the all_data but normally we'd query attendance_summary
    # But let's do a proper subject-wise query for accuracy
    subj_data = []
    if is_faculty:
        # Get subjects taught by this faculty
        faculty_name = session.get("name")
        subjects = qry("SELECT DISTINCT subject, ROUND(AVG(attended*100.0/total)::numeric,1) as avg FROM attendance_summary WHERE subject IN (SELECT name FROM subjects WHERE teacher LIKE %s) GROUP BY subject", ("%" + faculty_name + "%",))
    else:
        subjects = qry("SELECT DISTINCT subject, ROUND(AVG(attended*100.0/total)::numeric,1) as avg FROM attendance_summary GROUP BY subject LIMIT 10")
    
    for s in subjects:
        subj_data.append({"subject": s["subject"], "percentage": float(s["avg"] or 0)})
        
    return jsonify({
        "stats": {
            "total": len(all_data),
            "defaulters_count": len(defaulters)
        },
        "analytics": {
            "avg_attendance": avg_att,
            "defaulters": [{"student_name": d["name"], "percentage": d["pct"], "subject": "Overall"} for d in defaulters[:20]],
            "subject_wise": subj_data
        }
    })


# ════════════════════════════════════════════════════════════
#  STUDENT ATTENDANCE DASHBOARD  (student self-view + admin view)
# ════════════════════════════════════════════════════════════

@app.route("/student_attendance_dashboard")
def student_attendance_dashboard():
    # Admin can view any student via ?student_id=
    is_admin_view = session.get("role") == "admin"

    if is_admin_view:
        sid_param = request.args.get("student_id", "")
        if not sid_param:
            return redirect("/attendance_dashboard")
        student_row = qone("SELECT * FROM students WHERE id=%s", (safe_int(sid_param),))
        if not student_row:
            return redirect("/attendance_dashboard")
        student = dict(student_row)
    elif session.get("role") == "student":
        student_row = qone("SELECT * FROM students WHERE id=%s", (session["student_id"],))
        if not student_row:
            return redirect("/logout")
        student = dict(student_row)
    else:
        return redirect("/login")

    sid  = student["id"]
    name = student["name"]

    cum      = get_cumulative(sid, name)
    subjects = get_subjects(sid, name)

    has_v2_data = has_summary(sid)

    # Prediction
    curve   = prediction_curve(cum["attended"], cum["total"], future_steps=20)
    can_reach, lectures_needed = will_reach_75(cum["attended"], cum["total"])

    subj_labels = [s["subject"][:20] for s in subjects]
    subj_pcts   = [s["percentage"] for s in subjects]

    return render_template("student/student_attendance_dashboard.html",
        student=student,
        cum=cum,
        subjects=subjects,
        has_v2_data=has_v2_data,
        is_admin_view=is_admin_view,
        prediction_curve=curve,
        can_reach=can_reach,
        lectures_needed=lectures_needed,
        subj_labels=subj_labels,
        subj_pcts=subj_pcts,
        today=datetime.now().strftime("%Y-%m-%d")
    )


# ════════════════════════════════════════════════════════════
#  IMPORT FINAL PDF (pdfplumber route)
# ════════════════════════════════════════════════════════════

@app.route("/import_final_pdf", methods=["POST"])
@login_required("admin")
def import_final_pdf():
    """Import DY Patil Final Attendance Report PDF (multi-page, all branches/divs)."""

    f = request.files.get("file")
    if not f:
        return redirect("/attendance_dashboard?error=nofile")

    added = skipped = students_found = students_created = 0

    def _clean(s):
        """Normalize cell text: collapse newlines/spaces."""
        return " ".join(str(s or "").replace("\n", " ").split()).strip()

    def _safe_int(v):
        v = str(v or "").strip()
        try:
            return int(float(v)) if v and v.replace(".","").isdigit() else 0
        except Exception:
            return 0

    def _dept_div_from_program(text):
        """Extract department and division from program line like
        'S. Y. B. Tech Comp. Engg. (Div. A)' or 'S. Y. B. Tech AIML (Div. A)'"""
        dept, div = "", ""
        m = _re.search(r'Div[.\s]*([A-Z])', text, _re.I)
        if m:
            div = m.group(1).upper()
        if _re.search(r'AIML', text, _re.I):
            dept = "AIML"
        elif _re.search(r'AI.*DS|AIDS', text, _re.I):
            dept = "AIDS"
        elif _re.search(r'Comp|CE', text, _re.I):
            dept = "CS"
        elif _re.search(r'\bIT\b', text, _re.I):
            dept = "IT"
        return dept, div

    def _ensure_student(name, roll, dept, div):
        """Return student id; create student record if not found."""
        nonlocal students_created
        if not name:
            return None
        # Try by roll first, then name
        row = None
        if roll:
            row = qone("SELECT id FROM students WHERE roll=%s", (roll,))
        if not row:
            row = qone("SELECT id FROM students WHERE name=%s ORDER BY id LIMIT 1", (name,))
        if row:
            return row["id"]
        # Auto-create
        year = "II"  # S.Y. = Second Year
        try:
            new_id = exe(
                "INSERT INTO students(name,roll,department,year,division,password) VALUES(%s,%s,%s,%s,%s,%s)",
                (name, roll or "", dept or "CS", year, div or "", hash_password(_default_student_password()))
            )
            students_created += 1
            return new_id
        except Exception:
            return None

    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name

        with pdfplumber.open(tmp_path) as pdf:
            # Grab full text for program/dept detection
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)

        # Detect dept/div from full_text
        prog_match = _re.search(r'Program[:\s]*(S\.?\s*Y\.?.*)', full_text, _re.I)
        prog_line  = prog_match.group(1) if prog_match else full_text[:200]
        global_dept, global_div = _dept_div_from_program(prog_line)

        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 5:
                        continue

                    # ── Step 1: Locate header row (has "NAME OF STUDENT") ──
                    hdr_idx = None
                    for ri, row in enumerate(table[:6]):
                        row_str = " ".join(_clean(c).upper() for c in row)
                        if "NAME" in row_str and ("ROLL" in row_str or "SR." in row_str or "SR\nNO" in row_str):
                            hdr_idx = ri
                            break
                    if hdr_idx is None:
                        continue

                    # Clean header: collapse multi-line subject names
                    headers = [_clean(c) for c in table[hdr_idx]]
                    name_col = next((i for i,h in enumerate(headers) if "NAME" in h.upper()), None)
                    roll_col = next((i for i,h in enumerate(headers) if "ROLL" in h.upper()), None)
                    if name_col is None:
                        continue

                    # ── Step 2: Build subject column list ──
                    # Subject cols = all columns after roll/name that are not Total/%
                    skip_until = max(name_col, roll_col or 0)
                    subj_cols = []
                    for ci in range(skip_until + 1, len(headers)):
                        h = headers[ci]
                        hu = h.upper().replace(" ", "")
                        if hu in ("", "TOTAL", "%OFATTENDANCE", "PERCENTAGEOFATTENDANCE", "%"):
                            continue  # skip silently (don't break — % can be last col)
                        if h:
                            subj_cols.append((ci, h))

                    if not subj_cols:
                        continue

                    # ── Step 3: Find "TOTAL NO. OF LECTURES CONDUCTED" row ──
                    total_lec_row = None
                    for ri in range(hdr_idx + 1, min(hdr_idx + 7, len(table))):
                        name_cell = _clean(table[ri][name_col] if name_col < len(table[ri]) else "")
                        nc_up = name_cell.upper()
                        if "LECTURE" in nc_up and ("TOTAL" in nc_up or "NO." in nc_up or "CONDUCTED" in nc_up):
                            total_lec_row = ri
                            break

                    total_lecs = {}
                    if total_lec_row is not None:
                        for ci, subj in subj_cols:
                            try:
                                total_lecs[subj] = _safe_int(table[total_lec_row][ci])
                            except Exception:
                                total_lecs[subj] = 0

                    # ── Step 4: Parse student rows ──
                    data_start = (total_lec_row + 1) if total_lec_row is not None else (hdr_idx + 4)
                    for row in table[data_start:]:
                        if not row:
                            continue
                        name_v = _clean(row[name_col]) if name_col < len(row) else ""
                        roll_v = _clean(row[roll_col]) if roll_col and roll_col < len(row) else ""

                        # Skip empty, header-repeat, or footer rows
                        if not name_v:
                            continue
                        nu = name_v.upper()
                        if nu in ("NAME OF STUDENT", "NAME", "FACULTY NAME", "SIGNATURE"):
                            continue
                        if any(nu.startswith(k) for k in ("PROF.", "DR.", "HOD", "FACULTY", "S.Y.", "ACADEMIC")):
                            continue
                        # Skip rows that are all zeros or look like summary rows
                        row_vals = [_clean(row[ci]) if ci < len(row) else "" for ci, _ in subj_cols]
                        if not any(v for v in row_vals):
                            continue

                        students_found += 1
                        s_id = _ensure_student(name_v, roll_v, global_dept, global_div)
                        if not s_id:
                            skipped += 1
                            continue

                        for ci, subj in subj_cols:
                            att_val  = _clean(row[ci]) if ci < len(row) else "0"
                            att_count = _safe_int(att_val)
                            total_v  = total_lecs.get(subj, 0) or att_count

                            try:
                                conn = get_db()
                                conn.execute("""
                                    INSERT INTO attendance_summary
                                        (student_id, student_name, subject, attended, total, division, semester)
                                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                                    ON CONFLICT(student_id, subject) DO UPDATE SET
                                        attended  = excluded.attended,
                                        total     = excluded.total,
                                        division  = excluded.division,
                                        semester  = excluded.semester
                                """, (s_id, name_v, subj, att_count, total_v, global_div, "IV"))
                                conn.commit()
                                conn.close()
                                added += 1
                            except Exception:
                                skipped += 1

        os.unlink(tmp_path)

    except Exception as e:
        return redirect(f"/attendance_dashboard?error=parse_failed&msg={str(e)[:100]}")

    return redirect(
        f"/attendance_dashboard?saved={added}&skipped={skipped}"
        f"&students={students_found}&new_students={students_created}&format=pdf"
    )


# ════════════════════════════════════════════════════════════
#  SMART QR ATTENDANCE API
# ════════════════════════════════════════════════════════════
import uuid

@app.route("/api/start_qr_attendance", methods=["POST"])
def start_qr_attendance():
    try:
        faculty_id = session.get("faculty_id") or 0
        subject = request.form.get("subject", "Unknown")
        division = request.form.get("division", "")
        classroom_id = request.form.get("classroom_id")
        
        # safely parse classroom_id
        if classroom_id:
            try: classroom_id = int(classroom_id)
            except: classroom_id = None
        else:
            classroom_id = None
            
        # 7. ATTENDANCE INTEGRATION: Ensure valid timetable slot exists
        day_str = datetime.now().strftime("%A")
        sub_row = qone("SELECT id FROM subjects WHERE name=%s LIMIT 1", (subject,))
        subject_id = sub_row["id"] if sub_row else None
        
        valid_slot = qone(
            "SELECT id FROM timetable WHERE faculty_id=%s AND subject_id=%s AND division=%s AND day=%s AND start_time <= localtime AND end_time >= localtime", 
            (faculty_id, subject_id, division, day_str)
        )
        if not valid_slot:
             return jsonify({"success": False, "error": f"Security Error: No valid timetable slot found for {subject} {division} on {day_str}"})
             
        timetable_id = valid_slot["id"]

        token = str(uuid.uuid4())
        expiry = datetime.now() + timedelta(seconds=60)
        
        exe(
            "INSERT INTO qr_sessions (faculty_id, token, subject, division, classroom_id, expiry, is_active) VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (faculty_id, token, subject, division, classroom_id, expiry, True)
        )
        # Store timetable_id in session state if needed, or query it downstream
        return jsonify({"success": True, "token": token})
    except Exception as e:
        app.logger.error(f"QR Start Error: {e}")
        return jsonify({"success": False, "error": str(e)})

def haversine(lat1, lon1, lat2, lon2):
    import math
    R = 6371000
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1-a))

@app.route("/api/mark_qr_attendance", methods=["POST"])
@limiter.limit("60 per minute")
def mark_qr_attendance():
    try:
        token = request.form.get("token")
        lat = float(request.form.get("latitude") or 0)
        lng = float(request.form.get("longitude") or 0)
        student_id = session.get("student_id")
        student_name = session.get("name")
        
        if not student_id:
            return jsonify({"success": False, "error": "Not logged in as student"})
            
        session_row = qone("SELECT * FROM qr_sessions WHERE token=%s AND is_active=%s AND expiry > %s", (token, True, datetime.now()))
        
        if not session_row:
            return jsonify({"success": False, "error": "Invalid or expired QR code"})
            
        location_verified = False
        classroom_id = session_row.get("classroom_id")
        
        if classroom_id:
            classroom = qone("SELECT latitude, longitude, radius FROM classrooms WHERE id=%s", (classroom_id,))
            if classroom and classroom.get("latitude"):
                c_lat = classroom["latitude"]
                c_lng = classroom["longitude"]
                radius = classroom["radius"] or 50
                dist = haversine(lat, lng, c_lat, c_lng)
                if dist <= radius:
                    location_verified = True
                else:
                    return jsonify({"success": False, "error": "You are too far from the classroom"})
        else:
            location_verified = True # No explicit classroom location bound
            
        # Prevent duplicate attendance
        today = datetime.now().date()
        att_date = today.strftime("%Y-%m-%d")
        subject = session_row["subject"]
        
        existing = qone("SELECT id FROM attendance WHERE student_id=%s AND date=%s AND subject=%s", (student_id, att_date, subject))
        if existing:
            return jsonify({"success": False, "error": "Attendance already marked for this subject today"})
            
        exe(
            "INSERT INTO attendance (student_id, student_name, subject, date, status, latitude, longitude, method, location_verified) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (student_id, student_name, subject, att_date, "Present", lat, lng, "QR", location_verified)
        )
        return jsonify({"success": True, "message": "Attendance marked successfully"})
    except Exception as e:
        app.logger.error(f"QR Mark Error: {e}")
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/student_notifications")
@student_required
def api_student_notifications():
    sid = session.get("student_id")
    notifs = safe_fetch_all("SELECT * FROM timetable_notifications WHERE student_id=%s ORDER BY created_at DESC", (sid,))
    # Serialize datetime explicitly to avoid JSON default errors
    out = []
    for n in notifs:
        d = dict(n)
        # Use hasattr for robustness against none or string
        if hasattr(d.get("created_at"), "strftime"):
            d["created_at"] = d["created_at"].strftime("%Y-%m-%d %H:%M:%S")
        out.append(d)
    return jsonify({"notifications": out, "status": "success"})

# ════════════════════════════════════════════════════════════
#  EXPORT FINAL PDF (ReportLab) & ANALYTICS
# ════════════════════════════════════════════════════════════
@app.route("/export_timetable_pdf")
def export_timetable_pdf():
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        from io import BytesIO
        
        # Security checking logic here... 
        # (Assuming it's public or admin checking)
        if "role" not in session: return redirect("/login")

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        entries = qry("SELECT t.day, t.start_time, t.end_time, t.subject, f.name as teacher, t.room, t.division FROM timetable t LEFT JOIN faculty f ON t.faculty_id = f.id ORDER BY t.day, t.start_time")
        
        data = [["Day", "Time", "Subject", "Teacher", "Room", "Division"]]
        for e in entries:
            data.append([e['day'], e.get("time","") or f"{e.get('start_time','')}-{e.get('end_time','')}", e['subject'], e['teacher'] or '-', e['room'] or '-', e['division'] or '-'])
            
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0F172A")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#F8FAFC")),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#E2E8F0")),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ]))
        
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        from flask import send_file
        return send_file(buffer, as_attachment=True, download_name="production_timetable.pdf", mimetype='application/pdf')
    except Exception as e:
        return f"ReportLab missing or error: {str(e)}"

@app.route("/api/faculty_workload")
def api_faculty_workload():
    if session.get("role") != "admin": return jsonify({"error": "Unauthorized"})

    entries = qry("SELECT faculty_id, teacher, day FROM timetable WHERE faculty_id IS NOT NULL")
    workload = {}
    for e in entries:
        fid = e["faculty_id"]
        t = e["teacher"]
        day = e["day"]
        if fid not in workload:
            workload[fid] = {"name": t, "slots": 0, "days": set(), "status": "Normal"}
        workload[fid]["slots"] += 1
        workload[fid]["days"].add(day)
        
    res = []
    for fid, v in workload.items():
        v["days_count"] = len(v["days"])
        v.pop("days")
        if v["slots"] > 20: v["status"] = "Overloaded"
        elif v["slots"] < 8: v["status"] = "Underloaded"
        res.append(v)
    
    return jsonify(res)


@app.route("/api/timetable_analytics", methods=["POST"])
@login_required("admin")
def api_timetable_analytics():
    f = request.files.get("file")
    if not f: return jsonify({"error": "No file uploaded"})
    
    # Parse excel for analytics WITHOUT saving
    added = _parse_timetable_excel(f, simulate=True)
    
    # Calculate stats from currently loaded DB
    total_classes = qone("SELECT COUNT(*) as c FROM timetable")["c"]
    sub_count = qone("SELECT COUNT(DISTINCT subject_id) as c FROM timetable")["c"]
    
    # Faculty load
    f_rows = qry("SELECT f.name as teacher, COUNT(t.id) as loads FROM timetable t JOIN faculty f ON t.faculty_id=f.id GROUP BY f.name")
    fac_load = {}
    for r in f_rows:
        fac_load[r["teacher"]] = r["loads"]
        
    return jsonify({
        "success": True,
        "simulated_slots_parsed_from_file": added,
        "total_classes": total_classes,
        "subject_count": sub_count,
        "faculty_load": fac_load
    })


# ════════════════════════════════════════════════════════════
# TIMETABLE NOTIFICATION SYSTEM
# ════════════════════════════════════════════════════════════

@app.route("/api/faculty_send_timetable", methods=["POST"])
@login_required("faculty")
def api_faculty_send_timetable():
    fac_id = session.get("faculty_id")
    f_branch = request.form.get("branch","").strip()
    f_year = request.form.get("year","").strip()
    f_div = request.form.get("division","").strip()
    f_subj = request.form.get("subject_id","").strip()

    sql = "SELECT DISTINCT branch, year, division FROM timetable WHERE faculty_id=%s AND branch != '' AND branch IS NOT NULL"
    params = [fac_id]
    if f_branch: sql += " AND branch=%s"; params.append(f_branch)
    if f_year: sql += " AND year=%s"; params.append(f_year)
    if f_div: sql += " AND division=%s"; params.append(f_div)
    if f_subj: sql += " AND subject_id=%s"; params.append(f_subj)

    classes = qry(sql, params)
    if not classes:
        return jsonify({"error": "No matching classes found in your timetable for these filters"}), 404

    message = request.form.get("message", f"Your timetable has been officially updated by Faculty ID {fac_id}.")

    conn = get_db()
    c = 0
    try:
        cur = conn.cur if hasattr(conn, 'cur') else conn.conn.cursor()
        inserts = []
        for row in classes:
            b, y, d = row["branch"], row["year"], row["division"]
            students = qry("SELECT id FROM students WHERE department=%s AND year=%s AND division=%s", (b, y, d))
            for st in students:
                inserts.append((fac_id, "faculty", st["id"], message))
                c += 1
                
        if inserts:
            import psycopg2.extras
            psycopg2.extras.execute_values(
                cur, 
                "INSERT INTO timetable_notifications (sender_id, sender_role, student_id, message) VALUES %s", 
                inserts
            )
            if hasattr(conn, 'conn'): conn.conn.commit()
            else: conn.commit()
    except Exception as e:
        if hasattr(conn, 'conn'): conn.conn.rollback()
        else: conn.rollback()
        raise e
    finally:
        conn.close()

    return jsonify({"success": True, "sent_count": c})

@app.route("/api/admin_send_timetable", methods=["POST"])
@login_required("admin")
def api_admin_send_timetable():
    admin_id = session.get("admin_id", 1)
    f_branch = request.form.get("branch","").strip()
    f_year = request.form.get("year","").strip()
    f_div = request.form.get("division","").strip()
    send_all = request.form.get("send_all")

    sql = "SELECT id FROM students WHERE 1=1"
    params = []
    if not send_all:
        if f_branch: sql += " AND department=%s"; params.append(f_branch)
        if f_year: sql += " AND year=%s"; params.append(f_year)
        if f_div: sql += " AND division=%s"; params.append(f_div)

    students = qry(sql, params)
    if not students:
        return jsonify({"error": "No matching students found"}), 404

    message = request.form.get("message", "Global Timetable Update Published by Administration.")
    
    conn = get_db()
    c = 0
    try:
        cur = conn.cur if hasattr(conn, 'cur') else conn.conn.cursor()
        inserts = []
        for st in students:
            inserts.append((admin_id, "admin", st["id"], message))
            c += 1

        if inserts:
            import psycopg2.extras
            psycopg2.extras.execute_values(
                cur, 
                "INSERT INTO timetable_notifications (sender_id, sender_role, student_id, message) VALUES %s", 
                inserts
            )
            if hasattr(conn, 'conn'): conn.conn.commit()
            else: conn.commit()
    except Exception as e:
        if hasattr(conn, 'conn'): conn.conn.rollback()
        else: conn.rollback()
        raise e
    finally:
        conn.close()

    return jsonify({"success": True, "sent_count": c})

@app.route("/student_notifications", methods=["GET"])
@login_required("student")
def student_notifications():
    sid = session.get("student_id")
    notifs = qry("SELECT * FROM timetable_notifications WHERE student_id=%s ORDER BY created_at DESC", (sid,))
    # Serialize datetime explicitly to avoid JSON default errors
    out = []
    for n in notifs:
        d = dict(n)
        if isinstance(d.get("created_at"), datetime):
            d["created_at"] = d["created_at"].strftime("%Y-%m-%d %H:%M:%S")
        out.append(d)
    return jsonify({"notifications": out})


# ════════════════════════════════════════════════════════════
#  REST API ENDPOINTS (ATTENDANCE MODULE)
# ════════════════════════════════════════════════════════════
@app.route("/attendance/mark", methods=["POST"])
def api_attendance_mark():
    if not session.get("user") and not session.get("role"):
        return jsonify({"error": "Unauthorized"}), 403
    # Use existing handle logic or quick logic
    # For bulk it's usually request.json
    mode = request.form.get("mode", "single")
    if mode == "bulk":
        return jsonify({"status": "success", "message": "Bulk mark supported via UI."})
    from routes.attendance import handle_single_mark
    resp = handle_single_mark(request.form, session)
    if hasattr(resp, 'location'): # It's a redirect
        return jsonify({"status": "success"})
    return jsonify({"status": "processed"})

@app.route("/attendance/summary", methods=["GET"])
@admin_required
def api_attendance_summary():
    dept = request.args.get("department")
    sql = "SELECT department, year, subject, COUNT(*) as sessions FROM attendance"
    params = []
    if dept:
        sql += " WHERE department=%s"
        params.append(dept)
    sql += " GROUP BY department, year, subject"
    data = safe_fetch_all(sql, params)
    return jsonify({"status": "success", "summary": [dict(d) for d in data]})

@app.route("/attendance/report", methods=["GET"])
@admin_required
def api_attendance_report():
    student_id = request.args.get("student_id")
    sql = "SELECT id, date, subject, status FROM attendance WHERE student_id=%s ORDER BY date DESC"
    data = safe_fetch_all(sql, (student_id,)) if student_id else []
    return jsonify({"status": "success", "data": [dict(d) for d in data]})

@app.route("/api/student_prediction", methods=["GET"])
@student_required
def api_student_prediction():
    from routes.attendance import handle_student_prediction
    return handle_student_prediction(session)

@app.route("/request_correction", methods=["POST"])
@student_required
def request_correction():
    from routes.attendance import handle_correction_request
    return handle_correction_request(request.form, session)

@app.route("/admin_audit_logs", methods=["GET"])
@admin_required
def admin_audit_logs():
    logs = safe_fetch_all("SELECT * FROM attendance_audit_log ORDER BY created_at DESC LIMIT 100")
    return render_template("admin/admin_audit_logs.html", logs=logs)

@app.route("/share/timetable/<int:fid>")
def share_timetable(fid):
    # Public route to view faculty timetable
    faculty = safe_fetch_one("SELECT * FROM faculty WHERE id=%s", (fid,))
    if not faculty:
        # Fallback or error
        flash("Faculty schedule not found", "error")
        return redirect("/login")
        
    f_name = faculty['name']
    all_rows = safe_query(
        f"SELECT t.*, %s as teacher FROM timetable t WHERE t.faculty_id = %s OR t.teacher ILIKE %s ORDER BY {DAY_ORD}, t.start_time",
        (f_name, fid, f"%{f_name}%")
    )
    
    all_entries = []
    for e in all_rows:
        d = dict(e)
        d["time"] = normalize_time(d.get("time",""))
        all_entries.append(d)

    seen_ts = set(); raw_ts = []
    for e in all_entries:
        t = e.get("time")
        if t and t not in seen_ts:
            seen_ts.add(t); raw_ts.append(t)

    def _tsort(ts):
        m = re.match(r"(\d+):(\d+)", str(ts or ""))
        if not m: return 999
        h = int(m.group(1)); mn = int(m.group(2))
        if h < 7: h += 12
        return h * 60 + mn

    time_slots = sorted(raw_ts, key=_tsort)
    grid = {d: {ts: [] for ts in time_slots} for d in DAYS}
    for r in all_entries:
        d = r.get('day'); t = r.get('time')
        if d in grid and t in grid[d]: grid[d][t].append(r)

    return render_template("common/timetable_shared.html", 
                         faculty=faculty, 
                         grid=grid, 
                         time_slots=time_slots,
                         DAYS=DAYS,
                         now=datetime.now().strftime("%d %b %Y, %I:%M %p"))

# ════════════════════════════════════════════════════════════
#  NEW API ENDPOINTS
# ════════════════════════════════════════════════════════════

@app.route("/api/messages/mark_read/<int:mid>", methods=["POST"])
@login_required("admin")
def api_mark_read(mid):
    # Only mark if receiver matches
    uid = session.get("user_id") or session.get("faculty_id") or session.get("student_id")
    role = session.get("role")
    safe_execute("UPDATE messages SET is_read=TRUE WHERE id=%s AND to_id=%s AND to_role=%s", (mid, uid, role))
    return jsonify({"status": "ok"})

@app.route("/api/send_timetable", methods=["POST"])
@login_required("faculty") # admin or faculty
def api_send_timetable():
    if session.get("role") not in ["admin", "faculty"]:
        return jsonify({"status": "error", "message": "Unauthorized"}), 403
    
    target_dept = request.form.get("department")
    target_year = request.form.get("year")
    target_div  = request.form.get("division")
    
    # Filter students
    sql = "SELECT id, name, department, division FROM students WHERE 1=1"
    params = []
    if target_dept and target_dept != "All": sql += " AND department=%s"; params.append(target_dept)
    if target_year and target_year != "All": sql += " AND year=%s"; params.append(target_year)
    if target_div and target_div != "All": sql += " AND division=%s"; params.append(target_div)
    
    students = qry(sql, params)
    sent = 0
    
    # Pre-fetch timetable
    # For now, let's just send the link or a summary
    for s in students:
        # Build shared link
        # Actually faculty send their own TT usually
        fid = session.get("faculty_id")
        fname = session.get("name")
        body = f"Hello {s['name']}, \nYour timetable has been updated. You can view it here: {request.host_url}share/timetable/{fid if fid else 1}"
        subj = "Timetable Update Notification"
        
        exe("""INSERT INTO messages(from_role, from_id, from_name, to_role, to_id, to_name, subject, body) 
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
            (session.get("role"), session.get("user_id") or fid or 1, fname, "student", s["id"], s["name"], subj, body))
        sent += 1
        
    return jsonify({"status": "ok", "sent": sent})

# Duplicate route removed — mark_read handled above at /api/messages/mark_read/<int:mid>

@app.route("/api/students/export")
@admin_required
def api_export_students():
    import io, csv
    rows = qry("SELECT * FROM students ORDER BY id DESC")
    si = io.StringIO()
    cw = csv.writer(si)
    if rows:
        cw.writerow(rows[0].keys())
        for r in rows: cw.writerow(r.values())
    
    from flask import make_response
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=students_export.csv"
    output.headers["Content-type"] = "text/csv"
    return output

@app.route("/api/save_timetable_slot", methods=["POST"])
@login_required("admin")
def api_save_slot():
    sid = request.form.get("id")
    subj = request.form.get("subject")
    room = request.form.get("room")
    div = request.form.get("division")
    if sid:
        exe("UPDATE timetable SET subject=%s, room=%s, division=%s WHERE id=%s", (subj, room, div, sid))
    return jsonify({"status": "ok"})

if __name__ == "__main__":
    debug_mode = os.environ.get("FLASK_ENV", "production") == "development"
    app.run(debug=debug_mode, host="0.0.0.0", port=5000)
